import sqlite3
import pandas as pd
import winsound
import openpyxl
from openpyxl.chart import (LineChart, Reference)
import openpyxl.styles

# подключаемся к базе данных
conn = sqlite3.connect('C:/SQLite/firstDB/stsDB.db')
# создаем курсор для выполнения запросов
cursor = conn.cursor()
start_date = '2023-02-10'
end_date = '2023-02-16' # надо брать на день позже

# используйте операторы сравнения для выборки строк в заданном диапазоне
query = f"SELECT * FROM LTEsts WHERE `Start Time` >= '{start_date}' AND `Start Time` <= '{end_date}'"

# выполняем запрос и получаем данные
cursor.execute(query)
data = cursor.fetchall()
data = [[None if col == 'NIL' else col for col in row] for row in data]
sts_df = pd.DataFrame(data, columns=[i[0] for i in cursor.description])

active_cell_number = 398  # ввести количество активных сот !!!!  19    398

directory = 'C:/test2/' # ввести директорию где лежит файл
csv_name = 'testf'  # ввести имя файла
output_comment = '_output_fromSQLdirectly'  # что добавится в конце к названию файла

sts_df['date'] = sts_df['Start Time'].str.split(' ').str[0]
sts_df['hour'] = sts_df['Start Time'].str.split(' ').str[1]
sts_df['date'] = pd.to_datetime(sts_df['date'])
sts_df['week'] = sts_df['date'].dt.isocalendar().week

list_1 = [ 'L.Cell.Avail.Dur (s)', 'L.ChMeas.PRB.DL.Avail (None)', 'L.ChMeas.PRB.DL.Used.Avg (None)', 'L.ChMeas.PRB.UL.Avail (None)', 'L.ChMeas.PRB.UL.Used.Avg (None)',\
           'L.CSFB.E2G (None)', 'L.CSFB.E2W (None)', 'L.E-RAB.AbnormRel (None)', 'L.E-RAB.AttEst (None)', 'L.E-RAB.FailEst.X2AP (None)', 'L.E-RAB.NormRel (None)', \
           'L.E-RAB.NormRel.IRatHOOut (None)', 'L.E-RAB.SuccEst (None)', 'L.HHO.IntereNB.InterFreq.ExecAttOut (None)', 'L.HHO.IntereNB.InterFreq.ExecSuccOut (None)', \
           'L.HHO.IntereNB.IntraFreq.ExecAttOut (None)', 'L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)', 'L.HHO.IntraeNB.InterFreq.ExecAttOut (None)', \
           'L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)', 'L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)', 'L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)', 'L.Thrp.bits.DL (bit)', \
           'L.Thrp.bits.DL.LastTTI (bit)', 'L.Thrp.bits.DL.QCI.1 (bit)', 'L.Thrp.bits.DL.QCI.2 (bit)', 'L.Thrp.bits.DL.QCI.3 (bit)', 'L.Thrp.bits.DL.QCI.4 (bit)', \
           'L.Thrp.bits.DL.QCI.5 (bit)', 'L.Thrp.bits.DL.QCI.6 (bit)', 'L.Thrp.bits.DL.QCI.7 (bit)', 'L.Thrp.bits.DL.QCI.8 (bit)', 'L.Thrp.bits.DL.QCI.9 (bit)', \
           'L.Thrp.bits.DL.QCI.65 (bit)', 'L.Thrp.bits.DL.QCI.66 (bit)', 'L.Thrp.bits.DL.QCI.69 (bit)', 'L.Thrp.bits.DL.QCI.70 (bit)', 'L.Thrp.bits.UE.UL.LastTTI (bit)', \
           'L.Thrp.bits.UL (bit)', 'L.Thrp.bits.UL.QCI.1 (bit)', 'L.Thrp.bits.UL.QCI.2 (bit)', 'L.Thrp.bits.UL.QCI.3 (bit)', 'L.Thrp.bits.UL.QCI.4 (bit)', 'L.Thrp.bits.UL.QCI.5 (bit)', \
           'L.Thrp.bits.UL.QCI.6 (bit)', 'L.Thrp.bits.UL.QCI.7 (bit)', 'L.Thrp.bits.UL.QCI.8 (bit)', 'L.Thrp.bits.UL.QCI.9 (bit)', 'L.Thrp.bits.UL.QCI.65 (bit)', \
           'L.Thrp.bits.UL.QCI.66 (bit)', 'L.Thrp.bits.UL.QCI.69 (bit)', 'L.Thrp.bits.UL.QCI.70 (bit)', 'L.Thrp.Time.DL.RmvLastTTI (ms)', 'L.Thrp.Time.UE.UL.RmvLastTTI (ms)', \
           'L.RRC.ConnReq.Att (None)', 'L.RRC.ConnReq.Succ (None)']  # список счётчиков
list_2 = [ 'PS traffic 4G, GB', 'Cell Availability 4G, %', 'Total LTE Cells Number', 'Downlink PRB UR,%', 'Uplink PRB UR,%', 'UE Downlink Av Thrp', \
           'UE Uplink Av Thrp', 'E-RAB Setup SR, %', 'E-RAB Drop Rate', 'Inter-Freq HO Out SR,%', 'Intra-Freq HO Out SR,%', \
           'CSFB to WCDMA', 'CSFB to GERAN', 'RRS setup SR,%', 'CCSR, %']  # список KPI

list_cluster = ['LABEL=UH29813, CellIndex=883, CGI=25094C3507475', 'LABEL=UH29812, CellIndex=882, CGI=25094C3507474', \
                'LABEL=UH29811, CellIndex=881, CGI=25094C3507473', 'LABEL=UH19473, CellIndex=880, CGI=25094C3504C11', \
                'LABEL=UH19472, CellIndex=879, CGI=25094C3504C10', 'LABEL=UH19471, CellIndex=878, CGI=25094C3504C0F', \
                'LABEL=UH08216, CellIndex=259, CGI=25094C3502018', 'LABEL=UH08215, CellIndex=258, CGI=25094C3502017', \
                'LABEL=UH08214, CellIndex=257, CGI=25094C3502016', 'LABEL=UH09703, CellIndex=306, CGI=25094C35025E7', \
                'LABEL=UH09702, CellIndex=305, CGI=25094C35025E6', \
                'LABEL=UH08815, CellIndex=270, CGI=25094C350226F', \
                'LABEL=UH08812, CellIndex=267, CGI=25094C350226C', \
                'LABEL=UH09701, CellIndex=304, CGI=25094C35025E5', \
                'LABEL=UH08811, CellIndex=266, CGI=25094C350226B', \
                'LABEL=UH39252, CellIndex=311, CGI=25094C3509954', \
                'LABEL=UH39251, CellIndex=310, CGI=25094C3509953', \
                'LABEL=UH08816, CellIndex=271, CGI=25094C3502270', \
                'LABEL=UH08814, CellIndex=269, CGI=25094C350226E', \
                'LABEL=UH08213, CellIndex=256, CGI=25094C3502015', \
                'LABEL=UH08813, CellIndex=268, CGI=25094C350226D', \
                'LABEL=UH08212, CellIndex=255, CGI=25094C3502014', \
                'LABEL=UH08211, CellIndex=254, CGI=25094C3502013', \
                'LABEL=UH39253, CellIndex=312, CGI=25094C3509955', ]  # кластер число сот 19
list_2600 = [
'eNodeB Function Name=UH0734, Local Cell ID=6, Cell Name=UH0734L6, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0734, Local Cell ID=5, Cell Name=UH0734L5, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=6, Cell Name=UH0763L6, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=5, Cell Name=UH0763L5, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=6, Cell Name=UH3922L6, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=5, Cell Name=UH3922L5, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=4, Cell Name=UH3922L4, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=6, Cell Name=UH0831L6, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=5, Cell Name=UH0831L5, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=4, Cell Name=UH0831L4, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=6, Cell Name=UH0960L6, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=5, Cell Name=UH0960L5, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=4, Cell Name=UH0960L4, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=6, Cell Name=UH0961L6, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=5, Cell Name=UH0961L5, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=4, Cell Name=UH0961L4, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=6, Cell Name=UH0965L6, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=5, Cell Name=UH0965L5, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=4, Cell Name=UH0965L4, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1905, Local Cell ID=5, Cell Name=UH1905L5, eNodeB ID=11905, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1905, Local Cell ID=4, Cell Name=UH1905L4, eNodeB ID=11905, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=6, Cell Name=UH1960L6, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=5, Cell Name=UH1960L5, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=4, Cell Name=UH1960L4, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=6, Cell Name=UH2763L6, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=5, Cell Name=UH2763L5, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=4, Cell Name=UH2763L4, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=6, Cell Name=UH1702L6, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=5, Cell Name=UH1702L5, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=4, Cell Name=UH1702L4, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=6, Cell Name=UH1707L6, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=5, Cell Name=UH1707L5, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=4, Cell Name=UH1707L4, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=7, Cell Name=UH0830L7, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=6, Cell Name=UH0830L6, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=5, Cell Name=UH0830L5, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=4, Cell Name=UH0830L4, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=7, Cell Name=UH0832L7, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=6, Cell Name=UH0832L6, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=5, Cell Name=UH0832L5, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=4, Cell Name=UH0832L4, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=5, Cell Name=UH0963L5, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=4, Cell Name=UH0963L4, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=6, Cell Name=UH0952L6, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=5, Cell Name=UH0952L5, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=4, Cell Name=UH0952L4, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0841, Local Cell ID=5, Cell Name=UH0841L5, eNodeB ID=10841, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0840, Local Cell ID=6, Cell Name=UH0840L6, eNodeB ID=10840, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=6, Cell Name=UH0962L6, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=5, Cell Name=UH0962L5, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=4, Cell Name=UH0962L4, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=6, Cell Name=UH0966L6, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=5, Cell Name=UH0966L5, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=4, Cell Name=UH0966L4, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=5, Cell Name=UH1703L5, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=4, Cell Name=UH1703L4, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1936, Local Cell ID=6, Cell Name=UH1936L6, eNodeB ID=1936, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1936, Local Cell ID=5, Cell Name=UH1936L5, eNodeB ID=1936, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1936, Local Cell ID=4, Cell Name=UH1936L4, eNodeB ID=1936, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=6, Cell Name=UH1939L6, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=5, Cell Name=UH1939L5, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=4, Cell Name=UH1939L4, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0821, Local Cell ID=6, Cell Name=UH0821L6, eNodeB ID=10821, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=6, Cell Name=UH0973L6, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=5, Cell Name=UH0973L5, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=4, Cell Name=UH0973L4, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1854, Local Cell ID=5, Cell Name=UH1854L5, eNodeB ID=11854, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0602, Local Cell ID=4, Cell Name=UH0602L4, eNodeB ID=10602, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0617, Local Cell ID=4, Cell Name=UH0617L4, eNodeB ID=10617, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0758, Local Cell ID=6, Cell Name=UH0758L6, eNodeB ID=10758, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0758, Local Cell ID=5, Cell Name=UH0758L5, eNodeB ID=10758, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=6, Cell Name=UH1908L6, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=5, Cell Name=UH1908L5, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1958, Local Cell ID=6, Cell Name=UH1958L6, eNodeB ID=11958, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1958, Local Cell ID=5, Cell Name=UH1958L5, eNodeB ID=11958, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=6, Cell Name=UH1976L6, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=5, Cell Name=UH1976L5, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=4, Cell Name=UH1976L4, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=6, Cell Name=UH3989L6, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=5, Cell Name=UH3989L5, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=4, Cell Name=UH3989L4, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=6, Cell Name=UH2938L6, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=5, Cell Name=UH2938L5, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=4, Cell Name=UH2938L4, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0742, Local Cell ID=4, Cell Name=UH0742L4, eNodeB ID=10742, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0742, Local Cell ID=5, Cell Name=UH0742L5, eNodeB ID=10742, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0879, Local Cell ID=6, Cell Name=UH0879L6, eNodeB ID=10879, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0879, Local Cell ID=5, Cell Name=UH0879L5, eNodeB ID=10879, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0879, Local Cell ID=4, Cell Name=UH0879L4, eNodeB ID=10879, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=61, Cell Name=UH1939L61, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=41, Cell Name=UH1939L41, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD'
]  # кластер число сот 91
list_1800 = [
'eNodeB Function Name=UH0748, Local Cell ID=1, Cell Name=UH0748L1, eNodeB ID=10748, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2981, Local Cell ID=3, Cell Name=UH02981L3, eNodeB ID=12981, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2981, Local Cell ID=1, Cell Name=UH02981L1, eNodeB ID=12981, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2981, Local Cell ID=2, Cell Name=UH02981L2, eNodeB ID=12981, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0737, Local Cell ID=3, Cell Name=UH0737L3, eNodeB ID=10737, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0737, Local Cell ID=2, Cell Name=UH0737L2, eNodeB ID=10737, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0737, Local Cell ID=1, Cell Name=UH0737L1, eNodeB ID=10737, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0752, Local Cell ID=2, Cell Name=UH00752L2, eNodeB ID=10752, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0752, Local Cell ID=3, Cell Name=UH00752L3, eNodeB ID=10752, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0843, Local Cell ID=3, Cell Name=UH0843L3, eNodeB ID=10843, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0843, Local Cell ID=1, Cell Name=UH0843L1, eNodeB ID=10843, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0991, Local Cell ID=2, Cell Name=UH0991L2, eNodeB ID=10991, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0991, Local Cell ID=1, Cell Name=UH0991L1, eNodeB ID=10991, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1919, Local Cell ID=3, Cell Name=UH1919L3, eNodeB ID=11919, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1919, Local Cell ID=2, Cell Name=UH1919L2, eNodeB ID=11919, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0702, Local Cell ID=2, Cell Name=UH0702L2, eNodeB ID=10702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0702, Local Cell ID=1, Cell Name=UH0702L1, eNodeB ID=10702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0715, Local Cell ID=2, Cell Name=UH0715L2, eNodeB ID=10715, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0715, Local Cell ID=1, Cell Name=UH0715L1, eNodeB ID=10715, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0734, Local Cell ID=3, Cell Name=UH0734L3, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0734, Local Cell ID=2, Cell Name=UH0734L2, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0734, Local Cell ID=1, Cell Name=UH0734L1, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=3, Cell Name=UH0763L3, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=2, Cell Name=UH0763L2, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=1, Cell Name=UH0763L1, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0770, Local Cell ID=3, Cell Name=UH0770L3, eNodeB ID=10770, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0770, Local Cell ID=2, Cell Name=UH0770L2, eNodeB ID=10770, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0770, Local Cell ID=1, Cell Name=UH0770L1, eNodeB ID=10770, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1704, Local Cell ID=3, Cell Name=UH1704L3, eNodeB ID=11704, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1704, Local Cell ID=2, Cell Name=UH1704L2, eNodeB ID=11704, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1704, Local Cell ID=1, Cell Name=UH1704L1, eNodeB ID=11704, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=3, Cell Name=UH3922L3, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=2, Cell Name=UH3922L2, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=1, Cell Name=UH3922L1, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0738, Local Cell ID=3, Cell Name=UH0738L3, eNodeB ID=10738, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0738, Local Cell ID=2, Cell Name=UH0738L2, eNodeB ID=10738, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0738, Local Cell ID=1, Cell Name=UH0738L1, eNodeB ID=10738, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=3, Cell Name=UH0831L3, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=2, Cell Name=UH0831L2, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=1, Cell Name=UH0831L1, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=3, Cell Name=UH0960L3, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=2, Cell Name=UH0960L2, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=1, Cell Name=UH0960L1, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=3, Cell Name=UH0961L3, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=2, Cell Name=UH0961L2, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=1, Cell Name=UH0961L1, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=3, Cell Name=UH0965L3, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=2, Cell Name=UH0965L2, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=1, Cell Name=UH0965L1, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0971, Local Cell ID=2, Cell Name=UH0971L2, eNodeB ID=10971, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0971, Local Cell ID=1, Cell Name=UH0971L1, eNodeB ID=10971, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0975, Local Cell ID=3, Cell Name=UH0975L3, eNodeB ID=10975, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0975, Local Cell ID=2, Cell Name=UH0975L2, eNodeB ID=10975, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0975, Local Cell ID=1, Cell Name=UH0975L1, eNodeB ID=10975, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1905, Local Cell ID=2, Cell Name=UH1905L2, eNodeB ID=11905, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1905, Local Cell ID=1, Cell Name=UH1905L1, eNodeB ID=11905, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1917, Local Cell ID=3, Cell Name=UH1917L3, eNodeB ID=11917, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1917, Local Cell ID=2, Cell Name=UH1917L2, eNodeB ID=11917, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1917, Local Cell ID=1, Cell Name=UH1917L1, eNodeB ID=11917, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=3, Cell Name=UH1960L3, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=2, Cell Name=UH1960L2, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=1, Cell Name=UH1960L1, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=3, Cell Name=UH2763L3, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=2, Cell Name=UH2763L2, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=1, Cell Name=UH2763L1, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=3, Cell Name=UH1702L3, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=2, Cell Name=UH1702L2, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=1, Cell Name=UH1702L1, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3921, Local Cell ID=3, Cell Name=UH3921L3, eNodeB ID=13921, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3921, Local Cell ID=2, Cell Name=UH3921L2, eNodeB ID=13921, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3921, Local Cell ID=1, Cell Name=UH3921L1, eNodeB ID=13921, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=3, Cell Name=UH1707L3, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=2, Cell Name=UH1707L2, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=1, Cell Name=UH1707L1, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0508, Local Cell ID=2, Cell Name=UH0508L2, eNodeB ID=10508, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0508, Local Cell ID=1, Cell Name=UH0508L1, eNodeB ID=10508, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0614, Local Cell ID=3, Cell Name=UH0614L3, eNodeB ID=10614, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0614, Local Cell ID=2, Cell Name=UH0614L2, eNodeB ID=10614, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0614, Local Cell ID=1, Cell Name=UH0614L1, eNodeB ID=10614, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0611, Local Cell ID=3, Cell Name=UH0611L3, eNodeB ID=10611, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0611, Local Cell ID=1, Cell Name=UH0611L1, eNodeB ID=10611, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0736, Local Cell ID=3, Cell Name=UH0736L3, eNodeB ID=10736, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0736, Local Cell ID=2, Cell Name=UH0736L2, eNodeB ID=10736, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0736, Local Cell ID=1, Cell Name=UH0736L1, eNodeB ID=10736, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0747, Local Cell ID=2, Cell Name=UH0747L2, eNodeB ID=10747, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=3, Cell Name=UH0830L3, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=2, Cell Name=UH0830L2, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=1, Cell Name=UH0830L1, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=8, Cell Name=UH0832L8, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=3, Cell Name=UH0832L3, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=2, Cell Name=UH0832L2, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=1, Cell Name=UH0832L1, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0849, Local Cell ID=2, Cell Name=UH0849L2, eNodeB ID=10849, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0849, Local Cell ID=1, Cell Name=UH0849L1, eNodeB ID=10849, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=8, Cell Name=UH0963L8, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=3, Cell Name=UH0963L3, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=2, Cell Name=UH0963L2, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=1, Cell Name=UH0963L1, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1805, Local Cell ID=3, Cell Name=UH1805L3, eNodeB ID=11805, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1805, Local Cell ID=2, Cell Name=UH1805L2, eNodeB ID=11805, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1805, Local Cell ID=1, Cell Name=UH1805L1, eNodeB ID=11805, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1906, Local Cell ID=3, Cell Name=UH1906L3, eNodeB ID=11906, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1906, Local Cell ID=2, Cell Name=UH1906L2, eNodeB ID=11906, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1906, Local Cell ID=1, Cell Name=UH1906L1, eNodeB ID=11906, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1910, Local Cell ID=3, Cell Name=UH1910L3, eNodeB ID=11910, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1910, Local Cell ID=2, Cell Name=UH1910L2, eNodeB ID=11910, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1910, Local Cell ID=1, Cell Name=UH1910L1, eNodeB ID=11910, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1925, Local Cell ID=3, Cell Name=UH1925L3, eNodeB ID=11925, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1925, Local Cell ID=2, Cell Name=UH1925L2, eNodeB ID=11925, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1925, Local Cell ID=1, Cell Name=UH1925L1, eNodeB ID=11925, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2932, Local Cell ID=3, Cell Name=UH2932L3, eNodeB ID=12932, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2932, Local Cell ID=2, Cell Name=UH2932L2, eNodeB ID=12932, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2932, Local Cell ID=1, Cell Name=UH2932L1, eNodeB ID=12932, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3928, Local Cell ID=3, Cell Name=UH3928L3, eNodeB ID=13928, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3928, Local Cell ID=2, Cell Name=UH3928L2, eNodeB ID=13928, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3928, Local Cell ID=1, Cell Name=UH3928L1, eNodeB ID=13928, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3996, Local Cell ID=3, Cell Name=UH3996L3, eNodeB ID=13996, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3996, Local Cell ID=2, Cell Name=UH3996L2, eNodeB ID=13996, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3996, Local Cell ID=1, Cell Name=UH3996L1, eNodeB ID=13996, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0711, Local Cell ID=3, Cell Name=UH0711L3, eNodeB ID=10711, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0711, Local Cell ID=2, Cell Name=UH0711L2, eNodeB ID=10711, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0711, Local Cell ID=1, Cell Name=UH0711L1, eNodeB ID=10711, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0731, Local Cell ID=3, Cell Name=UH0731L3, eNodeB ID=10731, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0731, Local Cell ID=2, Cell Name=UH0731L2, eNodeB ID=10731, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0731, Local Cell ID=1, Cell Name=UH0731L1, eNodeB ID=10731, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0735, Local Cell ID=2, Cell Name=UH0735L2, eNodeB ID=10735, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0735, Local Cell ID=1, Cell Name=UH0735L1, eNodeB ID=10735, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0842, Local Cell ID=3, Cell Name=UH0842L3, eNodeB ID=10842, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0842, Local Cell ID=2, Cell Name=UH0842L2, eNodeB ID=10842, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0842, Local Cell ID=1, Cell Name=UH0842L1, eNodeB ID=10842, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0708, Local Cell ID=3, Cell Name=UH0708L3, eNodeB ID=10708, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0708, Local Cell ID=2, Cell Name=UH0708L2, eNodeB ID=10708, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0708, Local Cell ID=1, Cell Name=UH0708L1, eNodeB ID=10708, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=3, Cell Name=UH0952L3, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=2, Cell Name=UH0952L2, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=1, Cell Name=UH0952L1, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0807, Local Cell ID=3, Cell Name=UH0807L3, eNodeB ID=10807, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0807, Local Cell ID=2, Cell Name=UH0807L2, eNodeB ID=10807, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0807, Local Cell ID=1, Cell Name=UH0807L1, eNodeB ID=10807, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0838, Local Cell ID=3, Cell Name=UH0838L3, eNodeB ID=10838, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0838, Local Cell ID=1, Cell Name=UH0838L1, eNodeB ID=10838, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0839, Local Cell ID=1, Cell Name=UH0839L1, eNodeB ID=10839, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0841, Local Cell ID=2, Cell Name=UH0841L2, eNodeB ID=10841, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0841, Local Cell ID=1, Cell Name=UH0841L1, eNodeB ID=10841, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1933, Local Cell ID=3, Cell Name=UH1933L3, eNodeB ID=11933, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1933, Local Cell ID=2, Cell Name=UH1933L2, eNodeB ID=11933, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1933, Local Cell ID=1, Cell Name=UH1933L1, eNodeB ID=11933, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0622, Local Cell ID=3, Cell Name=UH0622L3, eNodeB ID=10622, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0622, Local Cell ID=2, Cell Name=UH0622L2, eNodeB ID=10622, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0622, Local Cell ID=1, Cell Name=UH0622L1, eNodeB ID=10622, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0869, Local Cell ID=3, Cell Name=UH0869L3, eNodeB ID=10869, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0869, Local Cell ID=2, Cell Name=UH0869L2, eNodeB ID=10869, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0869, Local Cell ID=1, Cell Name=UH0869L1, eNodeB ID=10869, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0646, Local Cell ID=3, Cell Name=UH0646L3, eNodeB ID=10646, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0646, Local Cell ID=1, Cell Name=UH0646L1, eNodeB ID=10646, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0742, Local Cell ID=2, Cell Name=UH0742L2, eNodeB ID=10742, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0742, Local Cell ID=1, Cell Name=UH0742L1, eNodeB ID=10742, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0751, Local Cell ID=3, Cell Name=UH0751L3, eNodeB ID=10751, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0751, Local Cell ID=1, Cell Name=UH0751L1, eNodeB ID=10751, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0976, Local Cell ID=3, Cell Name=UH0976L3, eNodeB ID=10976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0976, Local Cell ID=2, Cell Name=UH0976L2, eNodeB ID=10976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0976, Local Cell ID=1, Cell Name=UH0976L1, eNodeB ID=10976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0950, Local Cell ID=3, Cell Name=UH0950L3, eNodeB ID=10950, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0950, Local Cell ID=2, Cell Name=UH0950L2, eNodeB ID=10950, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0950, Local Cell ID=1, Cell Name=UH0950L1, eNodeB ID=10950, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0981, Local Cell ID=3, Cell Name=UH0981L3, eNodeB ID=10981, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0981, Local Cell ID=2, Cell Name=UH0981L2, eNodeB ID=10981, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0981, Local Cell ID=1, Cell Name=UH0981L1, eNodeB ID=10981, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1901, Local Cell ID=2, Cell Name=UH1901L2, eNodeB ID=11901, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1901, Local Cell ID=1, Cell Name=UH1901L1, eNodeB ID=11901, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1902, Local Cell ID=3, Cell Name=UH1902L3, eNodeB ID=11902, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1902, Local Cell ID=2, Cell Name=UH1902L2, eNodeB ID=11902, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1902, Local Cell ID=1, Cell Name=UH1902L1, eNodeB ID=11902, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0623, Local Cell ID=3, Cell Name=UH0623L3, eNodeB ID=10623, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0623, Local Cell ID=2, Cell Name=UH0623L2, eNodeB ID=10623, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0623, Local Cell ID=1, Cell Name=UH0623L1, eNodeB ID=10623, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1920, Local Cell ID=3, Cell Name=UH1920L3, eNodeB ID=11920, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1920, Local Cell ID=2, Cell Name=UH1920L2, eNodeB ID=11920, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1920, Local Cell ID=1, Cell Name=UH1920L1, eNodeB ID=11920, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0713, Local Cell ID=3, Cell Name=UH0713L3, eNodeB ID=10713, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0713, Local Cell ID=2, Cell Name=UH0713L2, eNodeB ID=10713, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0713, Local Cell ID=1, Cell Name=UH0713L1, eNodeB ID=10713, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1945, Local Cell ID=2, Cell Name=UH1945L2, eNodeB ID=11945, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1945, Local Cell ID=1, Cell Name=UH1945L1, eNodeB ID=11945, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0717, Local Cell ID=3, Cell Name=UH0717L3, eNodeB ID=10717, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0717, Local Cell ID=2, Cell Name=UH0717L2, eNodeB ID=10717, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0717, Local Cell ID=1, Cell Name=UH0717L1, eNodeB ID=10717, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=3, Cell Name=UH2938L3, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=2, Cell Name=UH2938L2, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=1, Cell Name=UH2938L1, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0840, Local Cell ID=3, Cell Name=UH0840L3, eNodeB ID=10840, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0825, Local Cell ID=3, Cell Name=UH0825L3, eNodeB ID=10825, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0825, Local Cell ID=2, Cell Name=UH0825L2, eNodeB ID=10825, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0825, Local Cell ID=1, Cell Name=UH0825L1, eNodeB ID=10825, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=3, Cell Name=UH0962L3, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=2, Cell Name=UH0962L2, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=1, Cell Name=UH0962L1, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=3, Cell Name=UH0966L3, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=2, Cell Name=UH0966L2, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=1, Cell Name=UH0966L1, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0972, Local Cell ID=3, Cell Name=UH0972L3, eNodeB ID=10972, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0972, Local Cell ID=2, Cell Name=UH0972L2, eNodeB ID=10972, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0972, Local Cell ID=1, Cell Name=UH0972L1, eNodeB ID=10972, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=3, Cell Name=UH1703L3, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=2, Cell Name=UH1703L2, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=1, Cell Name=UH1703L1, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1801, Local Cell ID=3, Cell Name=UH1801L3, eNodeB ID=11801, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1801, Local Cell ID=1, Cell Name=UH1801L1, eNodeB ID=11801, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1900, Local Cell ID=3, Cell Name=UH1900L3, eNodeB ID=11900, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1900, Local Cell ID=2, Cell Name=UH1900L2, eNodeB ID=11900, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1900, Local Cell ID=1, Cell Name=UH1900L1, eNodeB ID=11900, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1936, Local Cell ID=3, Cell Name=UH1936L3, eNodeB ID=1936, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1936, Local Cell ID=2, Cell Name=UH1936L2, eNodeB ID=1936, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1936, Local Cell ID=1, Cell Name=UH1936L1, eNodeB ID=1936, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=3, Cell Name=UH1939L3, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=2, Cell Name=UH1939L2, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=1, Cell Name=UH1939L1, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0765, Local Cell ID=3, Cell Name=UH0765L3, eNodeB ID=10765, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0765, Local Cell ID=2, Cell Name=UH0765L2, eNodeB ID=10765, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0765, Local Cell ID=1, Cell Name=UH0765L1, eNodeB ID=10765, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0821, Local Cell ID=3, Cell Name=UH0821L3, eNodeB ID=10821, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0821, Local Cell ID=2, Cell Name=UH0821L2, eNodeB ID=10821, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0821, Local Cell ID=1, Cell Name=UH0821L1, eNodeB ID=10821, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0822, Local Cell ID=3, Cell Name=UH0822L3, eNodeB ID=10822, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0822, Local Cell ID=2, Cell Name=UH0822L2, eNodeB ID=10822, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0822, Local Cell ID=1, Cell Name=UH0822L1, eNodeB ID=10822, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0881, Local Cell ID=3, Cell Name=UH0881L3, eNodeB ID=10881, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0881, Local Cell ID=2, Cell Name=UH0881L2, eNodeB ID=10881, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0881, Local Cell ID=1, Cell Name=UH0881L1, eNodeB ID=10881, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=3, Cell Name=UH0973L3, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=2, Cell Name=UH0973L2, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=1, Cell Name=UH0973L1, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0994, Local Cell ID=3, Cell Name=UH0994L3, eNodeB ID=10994, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0994, Local Cell ID=2, Cell Name=UH0994L2, eNodeB ID=10994, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0994, Local Cell ID=1, Cell Name=UH0994L1, eNodeB ID=10994, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1854, Local Cell ID=3, Cell Name=UH1854L3, eNodeB ID=11854, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1854, Local Cell ID=2, Cell Name=UH1854L2, eNodeB ID=11854, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1854, Local Cell ID=1, Cell Name=UH1854L1, eNodeB ID=11854, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1705, Local Cell ID=3, Cell Name=UH1705L3, eNodeB ID=11705, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1705, Local Cell ID=2, Cell Name=UH1705L2, eNodeB ID=11705, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1705, Local Cell ID=1, Cell Name=UH1705L1, eNodeB ID=11705, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1916, Local Cell ID=3, Cell Name=UH1916L3, eNodeB ID=11916, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1916, Local Cell ID=2, Cell Name=UH1916L2, eNodeB ID=11916, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1916, Local Cell ID=1, Cell Name=UH1916L1, eNodeB ID=11916, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3990, Local Cell ID=3, Cell Name=UH3990L3, eNodeB ID=13990, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3990, Local Cell ID=2, Cell Name=UH3990L2, eNodeB ID=13990, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3990, Local Cell ID=1, Cell Name=UH3990L1, eNodeB ID=13990, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0601, Local Cell ID=3, Cell Name=UH0601L3, eNodeB ID=10601, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0601, Local Cell ID=2, Cell Name=UH0601L2, eNodeB ID=10601, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0601, Local Cell ID=1, Cell Name=UH0601L1, eNodeB ID=10601, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0602, Local Cell ID=2, Cell Name=UH0602L2, eNodeB ID=10602, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0602, Local Cell ID=1, Cell Name=UH0602L1, eNodeB ID=10602, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0617, Local Cell ID=3, Cell Name=UH0617L3, eNodeB ID=10617, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0617, Local Cell ID=2, Cell Name=UH0617L2, eNodeB ID=10617, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0617, Local Cell ID=1, Cell Name=UH0617L1, eNodeB ID=10617, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0618, Local Cell ID=3, Cell Name=UH0618L3, eNodeB ID=10618, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0618, Local Cell ID=2, Cell Name=UH0618L2, eNodeB ID=10618, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0618, Local Cell ID=1, Cell Name=UH0618L1, eNodeB ID=10618, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0714, Local Cell ID=3, Cell Name=UH0714L3, eNodeB ID=10714, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0714, Local Cell ID=2, Cell Name=UH0714L2, eNodeB ID=10714, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0714, Local Cell ID=1, Cell Name=UH0714L1, eNodeB ID=10714, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0754, Local Cell ID=2, Cell Name=UH0754L2, eNodeB ID=10754, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0754, Local Cell ID=1, Cell Name=UH0754L1, eNodeB ID=10754, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0758, Local Cell ID=2, Cell Name=UH0758L2, eNodeB ID=10758, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0758, Local Cell ID=1, Cell Name=UH0758L1, eNodeB ID=10758, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0758, Local Cell ID=3, Cell Name=UH0758L3, eNodeB ID=10758, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=3, Cell Name=UH1908L3, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=2, Cell Name=UH1908L2, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=1, Cell Name=UH1908L1, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1913, Local Cell ID=3, Cell Name=UH1913L3, eNodeB ID=11913, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1913, Local Cell ID=2, Cell Name=UH1913L2, eNodeB ID=11913, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1913, Local Cell ID=1, Cell Name=UH1913L1, eNodeB ID=11913, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1951, Local Cell ID=3, Cell Name=UH1951L3, eNodeB ID=11951, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1951, Local Cell ID=2, Cell Name=UH1951L2, eNodeB ID=11951, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1951, Local Cell ID=1, Cell Name=UH1951L1, eNodeB ID=11951, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1958, Local Cell ID=3, Cell Name=UH1958L3, eNodeB ID=11958, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1958, Local Cell ID=2, Cell Name=UH1958L2, eNodeB ID=11958, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=3, Cell Name=UH1976L3, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=2, Cell Name=UH1976L2, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=1, Cell Name=UH1976L1, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=3, Cell Name=UH3989L3, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=2, Cell Name=UH3989L2, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=1, Cell Name=UH3989L1, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH4600, Local Cell ID=3, Cell Name=UH4600L3, eNodeB ID=14600, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH4600, Local Cell ID=2, Cell Name=UH4600L2, eNodeB ID=14600, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH4600, Local Cell ID=1, Cell Name=UH4600L1, eNodeB ID=14600, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1909, Local Cell ID=3, Cell Name=UH1909L3, eNodeB ID=11909, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1909, Local Cell ID=2, Cell Name=UH1909L2, eNodeB ID=11909, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0646, Local Cell ID=2, Cell Name=UH0646L2, eNodeB ID=10646, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0845, Local Cell ID=3, Cell Name=UH0845L3, eNodeB ID=10845, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0845, Local Cell ID=2, Cell Name=UH0845L2, eNodeB ID=10845, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0845, Local Cell ID=1, Cell Name=UH0845L1, eNodeB ID=10845, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3925, Local Cell ID=3, Cell Name=UH3925L3, eNodeB ID=13925, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3925, Local Cell ID=2, Cell Name=UH3925L2, eNodeB ID=13925, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3925, Local Cell ID=1, Cell Name=UH3925L1, eNodeB ID=13925, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0970, Local Cell ID=3, Cell Name=UH0970L3, eNodeB ID=10970, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0970, Local Cell ID=2, Cell Name=UH0970L2, eNodeB ID=10970, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0970, Local Cell ID=1, Cell Name=UH0970L1, eNodeB ID=10970, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1947, Local Cell ID=3, Cell Name=UH1947L3, eNodeB ID=11947, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1947, Local Cell ID=2, Cell Name=UH1947L2, eNodeB ID=11947, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1947, Local Cell ID=1, Cell Name=UH1947L1, eNodeB ID=11947, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0880, Local Cell ID=3, Cell Name=UH0880L3, eNodeB ID=10880, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0880, Local Cell ID=2, Cell Name=UH0880L2, eNodeB ID=10880, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0880, Local Cell ID=1, Cell Name=UH0880L1, eNodeB ID=10880, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0951, Local Cell ID=3, Cell Name=UH0951L3, eNodeB ID=10951, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0951, Local Cell ID=1, Cell Name=UH0951L1, eNodeB ID=10951, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0636, Local Cell ID=3, Cell Name=UH0636L3, eNodeB ID=10636, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0636, Local Cell ID=2, Cell Name=UH0636L2, eNodeB ID=10636, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0636, Local Cell ID=1, Cell Name=UH0636L1, eNodeB ID=10636, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2950, Local Cell ID=3, Cell Name=UH2950L3, eNodeB ID=12950, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2950, Local Cell ID=2, Cell Name=UH2950L2, eNodeB ID=12950, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2950, Local Cell ID=1, Cell Name=UH2950L1, eNodeB ID=12950, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0743, Local Cell ID=3, Cell Name=UH0743L3, eNodeB ID=10743, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0743, Local Cell ID=2, Cell Name=UH0743L2, eNodeB ID=10743, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0743, Local Cell ID=1, Cell Name=UH0743L1, eNodeB ID=10743, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2984, Local Cell ID=3, Cell Name=UH2984L3, eNodeB ID=12984, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2984, Local Cell ID=2, Cell Name=UH2984L2, eNodeB ID=12984, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2984, Local Cell ID=1, Cell Name=UH2984L1, eNodeB ID=12984, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0988, Local Cell ID=3, Cell Name=UH0988L3, eNodeB ID=10988, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0988, Local Cell ID=2, Cell Name=UH0988L2, eNodeB ID=10988, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0988, Local Cell ID=1, Cell Name=UH0988L1, eNodeB ID=10988, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0820, Local Cell ID=3, Cell Name=UH0820L3, eNodeB ID=10820, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0820, Local Cell ID=2, Cell Name=UH0820L2, eNodeB ID=10820, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0820, Local Cell ID=1, Cell Name=UH0820L1, eNodeB ID=10820, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1907, Local Cell ID=3, Cell Name=UH1907L3, eNodeB ID=11907, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1907, Local Cell ID=2, Cell Name=UH1907L2, eNodeB ID=11907, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1907, Local Cell ID=1, Cell Name=UH1907L1, eNodeB ID=11907, Cell FDD TDD indication=CELL_FDD'
]  # кластер число сот 327

list_dualband_1800 = [
'eNodeB Function Name=UH0734, Local Cell ID=3, Cell Name=UH0734L3, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0734, Local Cell ID=2, Cell Name=UH0734L2, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0734, Local Cell ID=1, Cell Name=UH0734L1, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=3, Cell Name=UH0763L3, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=2, Cell Name=UH0763L2, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=1, Cell Name=UH0763L1, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=3, Cell Name=UH3922L3, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=2, Cell Name=UH3922L2, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=1, Cell Name=UH3922L1, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=3, Cell Name=UH0831L3, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=2, Cell Name=UH0831L2, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=1, Cell Name=UH0831L1, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=3, Cell Name=UH0960L3, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=2, Cell Name=UH0960L2, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=1, Cell Name=UH0960L1, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=3, Cell Name=UH0961L3, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=2, Cell Name=UH0961L2, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=1, Cell Name=UH0961L1, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=3, Cell Name=UH0965L3, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=2, Cell Name=UH0965L2, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=1, Cell Name=UH0965L1, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1905, Local Cell ID=2, Cell Name=UH1905L2, eNodeB ID=11905, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1905, Local Cell ID=1, Cell Name=UH1905L1, eNodeB ID=11905, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=3, Cell Name=UH1960L3, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=2, Cell Name=UH1960L2, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=1, Cell Name=UH1960L1, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=3, Cell Name=UH2763L3, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=2, Cell Name=UH2763L2, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=1, Cell Name=UH2763L1, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=3, Cell Name=UH1702L3, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=2, Cell Name=UH1702L2, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=1, Cell Name=UH1702L1, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=3, Cell Name=UH1707L3, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=2, Cell Name=UH1707L2, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=1, Cell Name=UH1707L1, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=3, Cell Name=UH0830L3, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=2, Cell Name=UH0830L2, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=1, Cell Name=UH0830L1, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=7, Cell Name=UH0830L7, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=8, Cell Name=UH0832L8, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=3, Cell Name=UH0832L3, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=2, Cell Name=UH0832L2, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=1, Cell Name=UH0832L1, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=7, Cell Name=UH0832L7, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=8, Cell Name=UH0963L8, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=3, Cell Name=UH0963L3, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=2, Cell Name=UH0963L2, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=1, Cell Name=UH0963L1, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=3, Cell Name=UH0952L3, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=2, Cell Name=UH0952L2, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=1, Cell Name=UH0952L1, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0841, Local Cell ID=2, Cell Name=UH0841L2, eNodeB ID=10841, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0841, Local Cell ID=1, Cell Name=UH0841L1, eNodeB ID=10841, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0742, Local Cell ID=2, Cell Name=UH0742L2, eNodeB ID=10742, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0742, Local Cell ID=1, Cell Name=UH0742L1, eNodeB ID=10742, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=3, Cell Name=UH2938L3, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=2, Cell Name=UH2938L2, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=1, Cell Name=UH2938L1, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0840, Local Cell ID=3, Cell Name=UH0840L3, eNodeB ID=10840, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=3, Cell Name=UH0962L3, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=2, Cell Name=UH0962L2, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=1, Cell Name=UH0962L1, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=3, Cell Name=UH0966L3, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=2, Cell Name=UH0966L2, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=1, Cell Name=UH0966L1, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=3, Cell Name=UH1703L3, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=2, Cell Name=UH1703L2, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=1, Cell Name=UH1703L1, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=3, Cell Name=UH1939L3, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=2, Cell Name=UH1939L2, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=1, Cell Name=UH1939L1, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0821, Local Cell ID=3, Cell Name=UH0821L3, eNodeB ID=10821, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0821, Local Cell ID=2, Cell Name=UH0821L2, eNodeB ID=10821, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0821, Local Cell ID=1, Cell Name=UH0821L1, eNodeB ID=10821, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=3, Cell Name=UH0973L3, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=2, Cell Name=UH0973L2, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=1, Cell Name=UH0973L1, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1854, Local Cell ID=3, Cell Name=UH1854L3, eNodeB ID=11854, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1854, Local Cell ID=2, Cell Name=UH1854L2, eNodeB ID=11854, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1854, Local Cell ID=1, Cell Name=UH1854L1, eNodeB ID=11854, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0602, Local Cell ID=2, Cell Name=UH0602L2, eNodeB ID=10602, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0602, Local Cell ID=1, Cell Name=UH0602L1, eNodeB ID=10602, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0617, Local Cell ID=3, Cell Name=UH0617L3, eNodeB ID=10617, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0617, Local Cell ID=2, Cell Name=UH0617L2, eNodeB ID=10617, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0617, Local Cell ID=1, Cell Name=UH0617L1, eNodeB ID=10617, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=3, Cell Name=UH1908L3, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=2, Cell Name=UH1908L2, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=1, Cell Name=UH1908L1, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1958, Local Cell ID=3, Cell Name=UH1958L3, eNodeB ID=11958, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1958, Local Cell ID=2, Cell Name=UH1958L2, eNodeB ID=11958, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=3, Cell Name=UH1976L3, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=2, Cell Name=UH1976L2, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=1, Cell Name=UH1976L1, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=3, Cell Name=UH3989L3, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=2, Cell Name=UH3989L2, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=1, Cell Name=UH3989L1, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD'
] #сот 96
list_dualband_2600 = [
'eNodeB Function Name=UH0734, Local Cell ID=6, Cell Name=UH0734L6, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0734, Local Cell ID=5, Cell Name=UH0734L5, eNodeB ID=10734, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=6, Cell Name=UH0763L6, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0763, Local Cell ID=5, Cell Name=UH0763L5, eNodeB ID=10763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=6, Cell Name=UH3922L6, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=5, Cell Name=UH3922L5, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3922, Local Cell ID=4, Cell Name=UH3922L4, eNodeB ID=13922, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=6, Cell Name=UH0831L6, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=5, Cell Name=UH0831L5, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0831, Local Cell ID=4, Cell Name=UH0831L4, eNodeB ID=10831, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=6, Cell Name=UH0960L6, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=5, Cell Name=UH0960L5, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0960, Local Cell ID=4, Cell Name=UH0960L4, eNodeB ID=10960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=6, Cell Name=UH0961L6, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=5, Cell Name=UH0961L5, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0961, Local Cell ID=4, Cell Name=UH0961L4, eNodeB ID=10961, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=6, Cell Name=UH0965L6, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=5, Cell Name=UH0965L5, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0965, Local Cell ID=4, Cell Name=UH0965L4, eNodeB ID=10965, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1905, Local Cell ID=5, Cell Name=UH1905L5, eNodeB ID=11905, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1905, Local Cell ID=4, Cell Name=UH1905L4, eNodeB ID=11905, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=6, Cell Name=UH1960L6, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=5, Cell Name=UH1960L5, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1960, Local Cell ID=4, Cell Name=UH1960L4, eNodeB ID=11960, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=6, Cell Name=UH2763L6, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=5, Cell Name=UH2763L5, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2763, Local Cell ID=4, Cell Name=UH2763L4, eNodeB ID=12763, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=6, Cell Name=UH1702L6, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=5, Cell Name=UH1702L5, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1702, Local Cell ID=4, Cell Name=UH1702L4, eNodeB ID=11702, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=6, Cell Name=UH1707L6, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=5, Cell Name=UH1707L5, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1707, Local Cell ID=4, Cell Name=UH1707L4, eNodeB ID=11707, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=6, Cell Name=UH0830L6, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=5, Cell Name=UH0830L5, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0830, Local Cell ID=4, Cell Name=UH0830L4, eNodeB ID=10830, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=6, Cell Name=UH0832L6, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=5, Cell Name=UH0832L5, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0832, Local Cell ID=4, Cell Name=UH0832L4, eNodeB ID=10832, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=5, Cell Name=UH0963L5, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0963, Local Cell ID=4, Cell Name=UH0963L4, eNodeB ID=10963, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=6, Cell Name=UH0952L6, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=5, Cell Name=UH0952L5, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0952, Local Cell ID=4, Cell Name=UH0952L4, eNodeB ID=10952, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0841, Local Cell ID=5, Cell Name=UH0841L5, eNodeB ID=10841, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0840, Local Cell ID=6, Cell Name=UH0840L6, eNodeB ID=10840, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=6, Cell Name=UH0962L6, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=5, Cell Name=UH0962L5, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0962, Local Cell ID=4, Cell Name=UH0962L4, eNodeB ID=10962, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=6, Cell Name=UH0966L6, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=5, Cell Name=UH0966L5, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0966, Local Cell ID=4, Cell Name=UH0966L4, eNodeB ID=10966, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=5, Cell Name=UH1703L5, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1703, Local Cell ID=4, Cell Name=UH1703L4, eNodeB ID=11703, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=6, Cell Name=UH1939L6, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=5, Cell Name=UH1939L5, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1939, Local Cell ID=4, Cell Name=UH1939L4, eNodeB ID=11939, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0821, Local Cell ID=6, Cell Name=UH0821L6, eNodeB ID=10821, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=6, Cell Name=UH0973L6, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=5, Cell Name=UH0973L5, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0973, Local Cell ID=4, Cell Name=UH0973L4, eNodeB ID=10973, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1854, Local Cell ID=5, Cell Name=UH1854L5, eNodeB ID=11854, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0602, Local Cell ID=4, Cell Name=UH0602L4, eNodeB ID=10602, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0617, Local Cell ID=4, Cell Name=UH0617L4, eNodeB ID=10617, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=6, Cell Name=UH1908L6, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1908, Local Cell ID=5, Cell Name=UH1908L5, eNodeB ID=11908, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1958, Local Cell ID=6, Cell Name=UH1958L6, eNodeB ID=11958, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1958, Local Cell ID=5, Cell Name=UH1958L5, eNodeB ID=11958, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=6, Cell Name=UH1976L6, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=5, Cell Name=UH1976L5, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH1976, Local Cell ID=4, Cell Name=UH1976L4, eNodeB ID=11976, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=6, Cell Name=UH3989L6, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=5, Cell Name=UH3989L5, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH3989, Local Cell ID=4, Cell Name=UH3989L4, eNodeB ID=13989, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=6, Cell Name=UH2938L6, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=5, Cell Name=UH2938L5, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH2938, Local Cell ID=4, Cell Name=UH2938L4, eNodeB ID=12938, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0742, Local Cell ID=4, Cell Name=UH0742L4, eNodeB ID=10742, Cell FDD TDD indication=CELL_FDD', \
'eNodeB Function Name=UH0742, Local Cell ID=5, Cell Name=UH0742L5, eNodeB ID=10742, Cell FDD TDD indication=CELL_FDD'
] # сот 79

# обработка weekly:
weekly_df = sts_df.groupby(['week'])[list_1]. sum().reset_index()
weekly_df['PS traffic 4G, GB'] =(weekly_df['L.Thrp.bits.DL (bit)'] + weekly_df['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
weekly_df['Cell Availability 4G,%'] = 100 * weekly_df['L.Cell.Avail.Dur (s)'] / active_cell_number / 24 / 3600 / 7  # количество сот 398!!!
weekly_df['Total LTE Cells Number'] = active_cell_number # количество сот
weekly_df['Downlink PRB UR,%'] = weekly_df['L.ChMeas.PRB.DL.Used.Avg (None)'] / weekly_df['L.ChMeas.PRB.DL.Avail (None)'] * 100
weekly_df['Uplink PRB UR,%'] = weekly_df['L.ChMeas.PRB.UL.Used.Avg (None)'] /  weekly_df['L.ChMeas.PRB.UL.Avail (None)'] * 100
weekly_df['UE Downlink Av Thrp'] = (weekly_df['L.Thrp.bits.DL (bit)'] - weekly_df['L.Thrp.bits.DL.LastTTI (bit)']) / weekly_df['L.Thrp.Time.DL.RmvLastTTI (ms)']
weekly_df['UE Uplink Av Thrp'] = (weekly_df['L.Thrp.bits.UL (bit)'] - weekly_df['L.Thrp.bits.UE.UL.LastTTI (bit)']) / weekly_df['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
weekly_df['E-RAB Setup SR, %'] = weekly_df['L.E-RAB.SuccEst (None)'] / (weekly_df['L.E-RAB.AttEst (None)'] - weekly_df['L.E-RAB.FailEst.X2AP (None)']) * 100
weekly_df['E-RAB Drop Rate, %'] = weekly_df['L.E-RAB.AbnormRel (None)'] / (weekly_df['L.E-RAB.AbnormRel (None)'] + weekly_df['L.E-RAB.NormRel (None)'] + weekly_df['L.E-RAB.NormRel.IRatHOOut (None)'])*100
weekly_df['Inter-Freq HO Out SR,%'] = (weekly_df['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + weekly_df['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (weekly_df['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + weekly_df['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
weekly_df['Intra-Freq HO Out SR,%'] = (weekly_df['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + weekly_df['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (weekly_df['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + weekly_df['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
weekly_df['CSFB to WCDMA'] = weekly_df['L.CSFB.E2W (None)']
weekly_df['CSFB to GERAN'] = weekly_df['L.CSFB.E2G (None)']
weekly_df['RRS setup SR,%'] = weekly_df['L.RRC.ConnReq.Succ (None)'] / weekly_df['L.RRC.ConnReq.Att (None)'] * 100
weekly_df['DCSR4G, %'] = weekly_df['RRS setup SR,%'] * weekly_df['E-RAB Setup SR, %'] * (100 - weekly_df['E-RAB Drop Rate, %']) / 10000
weekly_df = weekly_df.drop(list_1, axis=1)
weekly_df_trans = weekly_df.transpose()

# ===обработка daily===
daily_df = sts_df.groupby(['date'])[list_1]. sum().reset_index()
daily_df['PS traffic 4G, GB'] =(daily_df['L.Thrp.bits.DL (bit)'] + daily_df['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
daily_df['Cell Availability 4G,%'] = 100 * daily_df['L.Cell.Avail.Dur (s)'] / active_cell_number / 24 / 3600  # количество сот 398!!!
daily_df['Total LTE Cells Number'] = active_cell_number # количество сот
daily_df['Downlink PRB UR,%'] = daily_df['L.ChMeas.PRB.DL.Used.Avg (None)'] / daily_df['L.ChMeas.PRB.DL.Avail (None)'] * 100
daily_df['Uplink PRB UR,%'] = daily_df['L.ChMeas.PRB.UL.Used.Avg (None)'] /  daily_df['L.ChMeas.PRB.UL.Avail (None)'] * 100
daily_df['UE Downlink Av Thrp'] = (daily_df['L.Thrp.bits.DL (bit)'] - daily_df['L.Thrp.bits.DL.LastTTI (bit)']) / daily_df['L.Thrp.Time.DL.RmvLastTTI (ms)']
daily_df['UE Uplink Av Thrp'] = (daily_df['L.Thrp.bits.UL (bit)'] - daily_df['L.Thrp.bits.UE.UL.LastTTI (bit)']) / daily_df['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
daily_df['E-RAB Setup SR, %'] = daily_df['L.E-RAB.SuccEst (None)'] / (daily_df['L.E-RAB.AttEst (None)'] - daily_df['L.E-RAB.FailEst.X2AP (None)']) * 100
daily_df['E-RAB Drop Rate, %'] = daily_df['L.E-RAB.AbnormRel (None)'] / (daily_df['L.E-RAB.AbnormRel (None)'] + daily_df['L.E-RAB.NormRel (None)'] + daily_df['L.E-RAB.NormRel.IRatHOOut (None)'])*100
daily_df['Inter-Freq HO Out SR,%'] = (daily_df['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + daily_df['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (daily_df['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + daily_df['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
daily_df['Intra-Freq HO Out SR,%'] = (daily_df['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + daily_df['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (daily_df['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + daily_df['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
daily_df['CSFB to WCDMA'] = daily_df['L.CSFB.E2W (None)']
daily_df['CSFB to GERAN'] = daily_df['L.CSFB.E2G (None)']
daily_df['RRS setup SR,%'] = daily_df['L.RRC.ConnReq.Succ (None)'] / daily_df['L.RRC.ConnReq.Att (None)'] * 100
daily_df['DCSR4G, %'] = daily_df['RRS setup SR,%'] * daily_df['E-RAB Setup SR, %'] * (100 - daily_df['E-RAB Drop Rate, %']) / 10000
daily_df = daily_df.drop(list_1, axis=1)

# фильтрация L2600
daily_dfL2600 = sts_df[sts_df['Cell'].isin(list_2600)]
daily_dfL2600 = daily_dfL2600.groupby(['date'])[list_1]. sum().reset_index()
daily_dfL2600['PS traffic 4G, GB_L2600'] =(daily_dfL2600['L.Thrp.bits.DL (bit)'] + daily_dfL2600['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
daily_dfL2600['Cell Availability 4G,%_L2600'] = 100 * daily_dfL2600['L.Cell.Avail.Dur (s)'] / 91 / 24 / 3600 # количество сот 91!!!
daily_dfL2600['Total LTE Cells Number_L2600'] = 91 # количество сот 91!!!
daily_dfL2600['Downlink PRB UR,%_L2600'] = daily_dfL2600['L.ChMeas.PRB.DL.Used.Avg (None)'] / daily_dfL2600['L.ChMeas.PRB.DL.Avail (None)'] * 100
daily_dfL2600['Uplink PRB UR,%_L2600'] = daily_dfL2600['L.ChMeas.PRB.UL.Used.Avg (None)'] /  daily_dfL2600['L.ChMeas.PRB.UL.Avail (None)'] * 100
daily_dfL2600['UE Downlink Av Thrp_L2600'] = (daily_dfL2600['L.Thrp.bits.DL (bit)'] - daily_dfL2600['L.Thrp.bits.DL.LastTTI (bit)']) / daily_dfL2600['L.Thrp.Time.DL.RmvLastTTI (ms)']
daily_dfL2600['UE Uplink Av Thrp_L2600'] = (daily_dfL2600['L.Thrp.bits.UL (bit)'] - daily_dfL2600['L.Thrp.bits.UE.UL.LastTTI (bit)']) / daily_dfL2600['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
daily_dfL2600['E-RAB Setup SR, %_L2600'] = daily_dfL2600['L.E-RAB.SuccEst (None)'] / (daily_dfL2600['L.E-RAB.AttEst (None)'] - daily_dfL2600['L.E-RAB.FailEst.X2AP (None)']) * 100
daily_dfL2600['E-RAB Drop Rate, %_L2600'] = daily_dfL2600['L.E-RAB.AbnormRel (None)'] / (daily_dfL2600['L.E-RAB.AbnormRel (None)'] + daily_dfL2600['L.E-RAB.NormRel (None)'] + daily_dfL2600['L.E-RAB.NormRel.IRatHOOut (None)'])*100
daily_dfL2600['Inter-Freq HO Out SR,%_L2600'] = (daily_dfL2600['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + daily_dfL2600['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (daily_dfL2600['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + daily_dfL2600['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
daily_dfL2600['Intra-Freq HO Out SR,%_L2600'] = (daily_dfL2600['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + daily_dfL2600['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (daily_dfL2600['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + daily_dfL2600['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
daily_dfL2600['CSFB to WCDMA_L2600'] = daily_dfL2600['L.CSFB.E2W (None)']
daily_dfL2600['CSFB to GERAN_L2600'] = daily_dfL2600['L.CSFB.E2G (None)']
daily_dfL2600['RRS setup SR,%_L2600'] = daily_dfL2600['L.RRC.ConnReq.Succ (None)'] / daily_dfL2600['L.RRC.ConnReq.Att (None)'] * 100
daily_dfL2600['DCSR4G, %_L2600'] = daily_dfL2600['RRS setup SR,%_L2600'] * daily_dfL2600['E-RAB Setup SR, %_L2600'] * (100 - daily_dfL2600['E-RAB Drop Rate, %_L2600']) / 10000
daily_dfL2600 = daily_dfL2600.drop(list_1, axis=1)

# фильтрация L1800
daily_dfL1800 = sts_df[sts_df['Cell'].isin(list_1800)]
daily_dfL1800 = daily_dfL1800.groupby(['date'])[list_1]. sum().reset_index()
daily_dfL1800['PS traffic 4G, GB_L1800'] =(daily_dfL1800['L.Thrp.bits.DL (bit)'] + daily_dfL1800['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
daily_dfL1800['Cell Availability 4G,%_L1800'] = 100 * daily_dfL1800['L.Cell.Avail.Dur (s)'] / 327 / 24 / 3600  # количество сот 327 !!!
daily_dfL1800['Total LTE Cells Number_L1800'] = 327 # количество сот 327 !!!
daily_dfL1800['Downlink PRB UR,%_L1800'] = daily_dfL1800['L.ChMeas.PRB.DL.Used.Avg (None)'] / daily_dfL1800['L.ChMeas.PRB.DL.Avail (None)'] * 100
daily_dfL1800['Uplink PRB UR,%_L1800'] = daily_dfL1800['L.ChMeas.PRB.UL.Used.Avg (None)'] /  daily_dfL1800['L.ChMeas.PRB.UL.Avail (None)'] * 100
daily_dfL1800['UE Downlink Av Thrp_L1800'] = (daily_dfL1800['L.Thrp.bits.DL (bit)'] - daily_dfL1800['L.Thrp.bits.DL.LastTTI (bit)']) / daily_dfL1800['L.Thrp.Time.DL.RmvLastTTI (ms)']
daily_dfL1800['UE Uplink Av Thrp_L1800'] = (daily_dfL1800['L.Thrp.bits.UL (bit)'] - daily_dfL1800['L.Thrp.bits.UE.UL.LastTTI (bit)']) / daily_dfL1800['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
daily_dfL1800['E-RAB Setup SR, %_L1800'] = daily_dfL1800['L.E-RAB.SuccEst (None)'] / (daily_dfL1800['L.E-RAB.AttEst (None)'] - daily_dfL1800['L.E-RAB.FailEst.X2AP (None)']) * 100
daily_dfL1800['E-RAB Drop Rate, %_L1800'] = daily_dfL1800['L.E-RAB.AbnormRel (None)'] / (daily_dfL1800['L.E-RAB.AbnormRel (None)'] + daily_dfL1800['L.E-RAB.NormRel (None)'] + daily_dfL1800['L.E-RAB.NormRel.IRatHOOut (None)'])*100
daily_dfL1800['Inter-Freq HO Out SR,%_L1800'] = (daily_dfL1800['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + daily_dfL1800['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (daily_dfL1800['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + daily_dfL1800['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
daily_dfL1800['Intra-Freq HO Out SR,%_L1800'] = (daily_dfL1800['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + daily_dfL1800['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (daily_dfL1800['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + daily_dfL1800['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
daily_dfL1800['CSFB to WCDMA_L1800'] = daily_dfL1800['L.CSFB.E2W (None)']
daily_dfL1800['CSFB to GERAN_L1800'] = daily_dfL1800['L.CSFB.E2G (None)']
daily_dfL1800['RRS setup SR,%_L1800'] = daily_dfL1800['L.RRC.ConnReq.Succ (None)'] / daily_dfL1800['L.RRC.ConnReq.Att (None)'] * 100
daily_dfL1800['DCSR4G, %_L1800'] = daily_dfL1800['RRS setup SR,%_L1800'] * daily_dfL1800['E-RAB Setup SR, %_L1800'] * (100 - daily_dfL1800['E-RAB Drop Rate, %_L1800']) / 10000
daily_dfL1800 = daily_dfL1800.drop(list_1, axis=1)

#daily_dfall = pd.merge(daily_df, daily_dfL2600, how="left")
#daily_dfall = pd.merge(daily_dfall, daily_dfL1800, how="left")
#daily_dfall = daily_dfall.transpose()

# фильтрация dual band L1800
daily_df_db_L18 = sts_df[sts_df['Cell'].isin(list_dualband_1800)]
daily_df_db_L18 = daily_df_db_L18.groupby(['date'])[list_1]. sum().reset_index()
daily_df_db_L18['PS traffic 4G, GB_db_L1800'] =(daily_df_db_L18['L.Thrp.bits.DL (bit)'] + daily_df_db_L18['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
daily_df_db_L18['Cell Availability 4G,%_db_L1800'] = 100 * daily_df_db_L18['L.Cell.Avail.Dur (s)'] / 96 / 24 / 3600  # количество сот 327 !!!
daily_df_db_L18['Total LTE Cells Number_db_L1800'] = 327 # количество сот 327 !!!
daily_df_db_L18['Downlink PRB UR,%_db_L1800'] = daily_df_db_L18['L.ChMeas.PRB.DL.Used.Avg (None)'] / daily_df_db_L18['L.ChMeas.PRB.DL.Avail (None)'] * 100
daily_df_db_L18['Uplink PRB UR,%_db_L1800'] = daily_df_db_L18['L.ChMeas.PRB.UL.Used.Avg (None)'] /  daily_df_db_L18['L.ChMeas.PRB.UL.Avail (None)'] * 100
daily_df_db_L18['UE Downlink Av Thrp_db_L1800'] = (daily_df_db_L18['L.Thrp.bits.DL (bit)'] - daily_df_db_L18['L.Thrp.bits.DL.LastTTI (bit)']) / daily_df_db_L18['L.Thrp.Time.DL.RmvLastTTI (ms)']
daily_df_db_L18['UE Uplink Av Thrp_db_L1800'] = (daily_df_db_L18['L.Thrp.bits.UL (bit)'] - daily_df_db_L18['L.Thrp.bits.UE.UL.LastTTI (bit)']) / daily_df_db_L18['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
daily_df_db_L18['E-RAB Setup SR, %_db_L1800'] = daily_df_db_L18['L.E-RAB.SuccEst (None)'] / (daily_df_db_L18['L.E-RAB.AttEst (None)'] - daily_df_db_L18['L.E-RAB.FailEst.X2AP (None)']) * 100
daily_df_db_L18['E-RAB Drop Rate, %_db_L1800'] = daily_df_db_L18['L.E-RAB.AbnormRel (None)'] / (daily_df_db_L18['L.E-RAB.AbnormRel (None)'] + daily_df_db_L18['L.E-RAB.NormRel (None)'] + daily_df_db_L18['L.E-RAB.NormRel.IRatHOOut (None)'])*100
daily_df_db_L18['Inter-Freq HO Out SR,%_db_L1800'] = (daily_df_db_L18['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + daily_df_db_L18['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (daily_df_db_L18['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + daily_df_db_L18['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
daily_df_db_L18['Intra-Freq HO Out SR,%_db_L1800'] = (daily_df_db_L18['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + daily_df_db_L18['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (daily_df_db_L18['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + daily_df_db_L18['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
daily_df_db_L18['CSFB to WCDMA_db_L1800'] = daily_df_db_L18['L.CSFB.E2W (None)']
daily_df_db_L18['CSFB to GERAN_db_L1800'] = daily_df_db_L18['L.CSFB.E2G (None)']
daily_df_db_L18['RRS setup SR,%_db_L1800'] = daily_df_db_L18['L.RRC.ConnReq.Succ (None)'] / daily_df_db_L18['L.RRC.ConnReq.Att (None)'] * 100
daily_df_db_L18['DCSR4G, %_db_L1800'] = daily_df_db_L18['RRS setup SR,%_db_L1800'] * daily_df_db_L18['E-RAB Setup SR, %_db_L1800'] * (100 - daily_df_db_L18['E-RAB Drop Rate, %_db_L1800']) / 10000
daily_df_db_L18 = daily_df_db_L18.drop(list_1, axis=1)

# фильтрация dual band L2600
daily_df_db_L26 = sts_df[sts_df['Cell'].isin(list_dualband_2600)]
daily_df_db_L26 = daily_df_db_L26.groupby(['date'])[list_1]. sum().reset_index()
daily_df_db_L26['PS traffic 4G, GB_db_L2600'] =(daily_df_db_L26['L.Thrp.bits.DL (bit)'] + daily_df_db_L26['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
daily_df_db_L26['Cell Availability 4G,%_db_L2600'] = 100 * daily_df_db_L26['L.Cell.Avail.Dur (s)'] / 79 / 24 / 3600  # количество сот 327 !!!
daily_df_db_L26['Total LTE Cells Number_db_L2600'] = 327 # количество сот 327 !!!
daily_df_db_L26['Downlink PRB UR,%_db_L2600'] = daily_df_db_L26['L.ChMeas.PRB.DL.Used.Avg (None)'] / daily_df_db_L26['L.ChMeas.PRB.DL.Avail (None)'] * 100
daily_df_db_L26['Uplink PRB UR,%_db_L2600'] = daily_df_db_L26['L.ChMeas.PRB.UL.Used.Avg (None)'] /  daily_df_db_L26['L.ChMeas.PRB.UL.Avail (None)'] * 100
daily_df_db_L26['UE Downlink Av Thrp_db_L2600'] = (daily_df_db_L26['L.Thrp.bits.DL (bit)'] - daily_df_db_L26['L.Thrp.bits.DL.LastTTI (bit)']) / daily_df_db_L26['L.Thrp.Time.DL.RmvLastTTI (ms)']
daily_df_db_L26['UE Uplink Av Thrp_db_L2600'] = (daily_df_db_L26['L.Thrp.bits.UL (bit)'] - daily_df_db_L26['L.Thrp.bits.UE.UL.LastTTI (bit)']) / daily_df_db_L26['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
daily_df_db_L26['E-RAB Setup SR, %_db_L2600'] = daily_df_db_L26['L.E-RAB.SuccEst (None)'] / (daily_df_db_L26['L.E-RAB.AttEst (None)'] - daily_df_db_L26['L.E-RAB.FailEst.X2AP (None)']) * 100
daily_df_db_L26['E-RAB Drop Rate, %_db_L2600'] = daily_df_db_L26['L.E-RAB.AbnormRel (None)'] / (daily_df_db_L26['L.E-RAB.AbnormRel (None)'] + daily_df_db_L26['L.E-RAB.NormRel (None)'] + daily_df_db_L26['L.E-RAB.NormRel.IRatHOOut (None)'])*100
daily_df_db_L26['Inter-Freq HO Out SR,%_db_L2600'] = (daily_df_db_L26['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + daily_df_db_L26['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (daily_df_db_L26['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + daily_df_db_L26['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
daily_df_db_L26['Intra-Freq HO Out SR,%_db_L2600'] = (daily_df_db_L26['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + daily_df_db_L26['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (daily_df_db_L26['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + daily_df_db_L26['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
daily_df_db_L26['CSFB to WCDMA_db_L2600'] = daily_df_db_L26['L.CSFB.E2W (None)']
daily_df_db_L26['CSFB to GERAN_db_L2600'] = daily_df_db_L26['L.CSFB.E2G (None)']
daily_df_db_L26['RRS setup SR,%_db_L2600'] = daily_df_db_L26['L.RRC.ConnReq.Succ (None)'] / daily_df_db_L26['L.RRC.ConnReq.Att (None)'] * 100
daily_df_db_L26['DCSR4G, %_db_L2600'] = daily_df_db_L26['RRS setup SR,%_db_L2600'] * daily_df_db_L26['E-RAB Setup SR, %_db_L2600'] * (100 - daily_df_db_L26['E-RAB Drop Rate, %_db_L2600']) / 10000
daily_df_db_L26 = daily_df_db_L26.drop(list_1, axis=1)

daily_dfall = pd.merge(daily_df, daily_dfL2600, how="left")
daily_dfall = pd.merge(daily_dfall, daily_dfL1800, how="left")
daily_dfall = pd.merge(daily_dfall, daily_df_db_L18, how="left")
daily_dfall = pd.merge(daily_dfall, daily_df_db_L26, how="left")
daily_dfall_trans = daily_dfall.transpose()

# обработка часовая
hourly_df = sts_df.groupby(['date', 'hour'])[list_1].sum().reset_index()
hourly_df['PS traffic 4G, GB'] =(hourly_df['L.Thrp.bits.DL (bit)'] + hourly_df['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
hourly_df['Cell Availability 4G, %'] = 100 * hourly_df['L.Cell.Avail.Dur (s)'] / active_cell_number / 3600  # количество сот  398!!!
hourly_df['Total LTE Cells Number'] = active_cell_number # количество сот
hourly_df['Downlink PRB UR,%'] = hourly_df['L.ChMeas.PRB.DL.Used.Avg (None)'] / hourly_df['L.ChMeas.PRB.DL.Avail (None)'] * 100
hourly_df['Uplink PRB UR,%'] = hourly_df['L.ChMeas.PRB.UL.Used.Avg (None)'] /  hourly_df['L.ChMeas.PRB.UL.Avail (None)'] * 100
hourly_df['UE Downlink Av Thrp'] = (hourly_df['L.Thrp.bits.DL (bit)'] - hourly_df['L.Thrp.bits.DL.LastTTI (bit)']) / hourly_df['L.Thrp.Time.DL.RmvLastTTI (ms)']
hourly_df['UE Uplink Av Thrp'] = (hourly_df['L.Thrp.bits.UL (bit)'] - hourly_df['L.Thrp.bits.UE.UL.LastTTI (bit)']) / hourly_df['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
hourly_df['E-RAB Setup SR,%'] = hourly_df['L.E-RAB.SuccEst (None)'] / (hourly_df['L.E-RAB.AttEst (None)'] - hourly_df['L.E-RAB.FailEst.X2AP (None)']) * 100
hourly_df['E-RAB Drop Rate'] = hourly_df['L.E-RAB.AbnormRel (None)'] / (hourly_df['L.E-RAB.AbnormRel (None)'] + hourly_df['L.E-RAB.NormRel (None)'] + hourly_df['L.E-RAB.NormRel.IRatHOOut (None)'])*100
hourly_df['Inter-Freq HO Out SR,%'] = (hourly_df['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + hourly_df['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (hourly_df['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + hourly_df['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
hourly_df['Intra-Freq HO Out SR,%'] = (hourly_df['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + hourly_df['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (hourly_df['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + hourly_df['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
hourly_df['CSFB to WCDMA'] = hourly_df['L.CSFB.E2W (None)']
hourly_df['CSFB to GERAN'] = hourly_df['L.CSFB.E2G (None)']
hourly_df['RRS setup SR,%'] = hourly_df['L.RRC.ConnReq.Succ (None)'] / hourly_df['L.RRC.ConnReq.Att (None)'] * 100
hourly_df['DCSR4G, %'] = hourly_df['RRS setup SR,%'] * hourly_df['E-RAB Setup SR,%'] * (100 - hourly_df['E-RAB Drop Rate']) / 10000
hourly_df = hourly_df.drop(list_1, axis=1)

# фильтрация L2600
hourly_dfL2600 = sts_df[sts_df['Cell'].isin(list_2600)]
hourly_dfL2600 = hourly_dfL2600.groupby(['date', 'hour'])[list_1]. sum().reset_index()
hourly_dfL2600['PS traffic 4G, GB_L2600'] =(hourly_dfL2600['L.Thrp.bits.DL (bit)'] + hourly_dfL2600['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
hourly_dfL2600['Cell Availability 4G,%_L2600'] = 100 * hourly_dfL2600['L.Cell.Avail.Dur (s)'] / 91 / 24 / 3600  # количество сот 91!!!
hourly_dfL2600['Total LTE Cells Number_L2600'] = 91 # количество сот 91!!!
hourly_dfL2600['Downlink PRB UR,%_L2600'] = hourly_dfL2600['L.ChMeas.PRB.DL.Used.Avg (None)'] / hourly_dfL2600['L.ChMeas.PRB.DL.Avail (None)'] * 100
hourly_dfL2600['Uplink PRB UR,%_L2600'] = hourly_dfL2600['L.ChMeas.PRB.UL.Used.Avg (None)'] /  hourly_dfL2600['L.ChMeas.PRB.UL.Avail (None)'] * 100
hourly_dfL2600['UE Downlink Av Thrp_L2600'] = (hourly_dfL2600['L.Thrp.bits.DL (bit)'] - hourly_dfL2600['L.Thrp.bits.DL.LastTTI (bit)']) / hourly_dfL2600['L.Thrp.Time.DL.RmvLastTTI (ms)']
hourly_dfL2600['UE Uplink Av Thrp_L2600'] = (hourly_dfL2600['L.Thrp.bits.UL (bit)'] - hourly_dfL2600['L.Thrp.bits.UE.UL.LastTTI (bit)']) / hourly_dfL2600['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
hourly_dfL2600['E-RAB Setup SR, %_L2600'] = hourly_dfL2600['L.E-RAB.SuccEst (None)'] / (hourly_dfL2600['L.E-RAB.AttEst (None)'] - hourly_dfL2600['L.E-RAB.FailEst.X2AP (None)']) * 100
hourly_dfL2600['E-RAB Drop Rate, %_L2600'] = hourly_dfL2600['L.E-RAB.AbnormRel (None)'] / (hourly_dfL2600['L.E-RAB.AbnormRel (None)'] + hourly_dfL2600['L.E-RAB.NormRel (None)'] + hourly_dfL2600['L.E-RAB.NormRel.IRatHOOut (None)'])*100
hourly_dfL2600['Inter-Freq HO Out SR,%_L2600'] = (hourly_dfL2600['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + hourly_dfL2600['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (hourly_dfL2600['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + hourly_dfL2600['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
hourly_dfL2600['Intra-Freq HO Out SR,%_L2600'] = (hourly_dfL2600['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + hourly_dfL2600['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (hourly_dfL2600['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + hourly_dfL2600['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
hourly_dfL2600['CSFB to WCDMA_L2600'] = hourly_dfL2600['L.CSFB.E2W (None)']
hourly_dfL2600['CSFB to GERAN_L2600'] = hourly_dfL2600['L.CSFB.E2G (None)']
hourly_dfL2600['RRS setup SR,%_L2600'] = hourly_dfL2600['L.RRC.ConnReq.Succ (None)'] / hourly_dfL2600['L.RRC.ConnReq.Att (None)'] * 100
hourly_dfL2600['DCSR4G, %_L2600'] = hourly_dfL2600['RRS setup SR,%_L2600'] * hourly_dfL2600['E-RAB Setup SR, %_L2600'] * (100 - hourly_dfL2600['E-RAB Drop Rate, %_L2600']) / 10000
hourly_dfL2600 = hourly_dfL2600.drop(list_1, axis=1)

# фильтрация L1800
hourly_dfL1800 = sts_df[sts_df['Cell'].isin(list_1800)]
hourly_dfL1800 = hourly_dfL1800.groupby(['date', 'hour'])[list_1]. sum().reset_index()
hourly_dfL1800['PS traffic 4G, GB_L1800'] =(hourly_dfL1800['L.Thrp.bits.DL (bit)'] + hourly_dfL1800['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
hourly_dfL1800['Cell Availability 4G,%_L1800'] = 100 * hourly_dfL1800['L.Cell.Avail.Dur (s)'] / 327 / 24 / 3600  # количество сот 327!!!
hourly_dfL1800['Total LTE Cells Number_L1800'] = 327 # количество сот 327 !!!
hourly_dfL1800['Downlink PRB UR,%_L1800'] = hourly_dfL1800['L.ChMeas.PRB.DL.Used.Avg (None)'] / hourly_dfL1800['L.ChMeas.PRB.DL.Avail (None)'] * 100
hourly_dfL1800['Uplink PRB UR,%_L1800'] = hourly_dfL1800['L.ChMeas.PRB.UL.Used.Avg (None)'] /  hourly_dfL1800['L.ChMeas.PRB.UL.Avail (None)'] * 100
hourly_dfL1800['UE Downlink Av Thrp_L1800'] = (hourly_dfL1800['L.Thrp.bits.DL (bit)'] - hourly_dfL1800['L.Thrp.bits.DL.LastTTI (bit)']) / hourly_dfL1800['L.Thrp.Time.DL.RmvLastTTI (ms)']
hourly_dfL1800['UE Uplink Av Thrp_L1800'] = (hourly_dfL1800['L.Thrp.bits.UL (bit)'] - hourly_dfL1800['L.Thrp.bits.UE.UL.LastTTI (bit)']) / hourly_dfL1800['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
hourly_dfL1800['E-RAB Setup SR, %_L1800'] = hourly_dfL1800['L.E-RAB.SuccEst (None)'] / (hourly_dfL1800['L.E-RAB.AttEst (None)'] - hourly_dfL1800['L.E-RAB.FailEst.X2AP (None)']) * 100
hourly_dfL1800['E-RAB Drop Rate, %_L1800'] = hourly_dfL1800['L.E-RAB.AbnormRel (None)'] / (hourly_dfL1800['L.E-RAB.AbnormRel (None)'] + hourly_dfL1800['L.E-RAB.NormRel (None)'] + hourly_dfL1800['L.E-RAB.NormRel.IRatHOOut (None)'])*100
hourly_dfL1800['Inter-Freq HO Out SR,%_L1800'] = (hourly_dfL1800['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + hourly_dfL1800['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (hourly_dfL1800['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + hourly_dfL1800['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
hourly_dfL1800['Intra-Freq HO Out SR,%_L1800'] = (hourly_dfL1800['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + hourly_dfL1800['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (hourly_dfL1800['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + hourly_dfL1800['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
hourly_dfL1800['CSFB to WCDMA_L1800'] = hourly_dfL1800['L.CSFB.E2W (None)']
hourly_dfL1800['CSFB to GERAN_L1800'] = hourly_dfL1800['L.CSFB.E2G (None)']
hourly_dfL1800['RRS setup SR,%_L1800'] = hourly_dfL1800['L.RRC.ConnReq.Succ (None)'] / hourly_dfL1800['L.RRC.ConnReq.Att (None)'] * 100
hourly_dfL1800['DCSR4G, %_L1800'] = hourly_dfL1800['RRS setup SR,%_L1800'] * hourly_dfL1800['E-RAB Setup SR, %_L1800'] * (100 - hourly_dfL1800['E-RAB Drop Rate, %_L1800']) / 10000
hourly_dfL1800 = hourly_dfL1800.drop(list_1, axis=1)

hourly_dfall = pd.merge(hourly_df, hourly_dfL2600, how="left")
hourly_dfall = pd.merge(hourly_dfall, hourly_dfL1800, how="left")
hourly_dfall_trans = hourly_dfall.transpose()

######
# обработка busy hour
hourlyBH_df = sts_df.groupby(['date', 'hour'])[list_1].sum().reset_index()
max_index = hourlyBH_df.groupby('date')['L.Thrp.bits.DL (bit)'].idxmax()
hourlyBH_df = hourlyBH_df.loc[max_index]
hourlyBH_df['PS traffic 4G, GB'] =(hourlyBH_df['L.Thrp.bits.DL (bit)'] + hourlyBH_df['L.Thrp.bits.UL (bit)'])/8/1024/1024/1024
hourlyBH_df['Cell Availability 4G, %'] = 100 * hourlyBH_df['L.Cell.Avail.Dur (s)'] / active_cell_number / 3600  # количество сот
hourlyBH_df['Total LTE Cells Number'] = active_cell_number # количество сот
hourlyBH_df['Downlink PRB UR,%'] = hourlyBH_df['L.ChMeas.PRB.DL.Used.Avg (None)'] / hourlyBH_df['L.ChMeas.PRB.DL.Avail (None)'] * 100
hourlyBH_df['Uplink PRB UR,%'] = hourlyBH_df['L.ChMeas.PRB.UL.Used.Avg (None)'] /  hourlyBH_df['L.ChMeas.PRB.UL.Avail (None)'] * 100
hourlyBH_df['UE Downlink Av Thrp'] = (hourlyBH_df['L.Thrp.bits.DL (bit)'] - hourlyBH_df['L.Thrp.bits.DL.LastTTI (bit)']) / hourlyBH_df['L.Thrp.Time.DL.RmvLastTTI (ms)']
hourlyBH_df['UE Uplink Av Thrp'] = (hourlyBH_df['L.Thrp.bits.UL (bit)'] - hourlyBH_df['L.Thrp.bits.UE.UL.LastTTI (bit)']) / hourlyBH_df['L.Thrp.Time.UE.UL.RmvLastTTI (ms)']
hourlyBH_df['E-RAB Setup SR,%'] = hourlyBH_df['L.E-RAB.SuccEst (None)'] / (hourlyBH_df['L.E-RAB.AttEst (None)'] - hourlyBH_df['L.E-RAB.FailEst.X2AP (None)']) * 100
hourlyBH_df['E-RAB Drop Rate'] = hourlyBH_df['L.E-RAB.AbnormRel (None)'] / (hourlyBH_df['L.E-RAB.AbnormRel (None)'] + hourlyBH_df['L.E-RAB.NormRel (None)'] + hourlyBH_df['L.E-RAB.NormRel.IRatHOOut (None)'])*100
hourlyBH_df['Inter-Freq HO Out SR,%'] = (hourlyBH_df['L.HHO.IntraeNB.InterFreq.ExecSuccOut (None)'] + hourlyBH_df['L.HHO.IntereNB.InterFreq.ExecSuccOut (None)']) / \
                                                        (hourlyBH_df['L.HHO.IntraeNB.InterFreq.ExecAttOut (None)'] + hourlyBH_df['L.HHO.IntereNB.InterFreq.ExecAttOut (None)']) * 100
hourlyBH_df['Intra-Freq HO Out SR,%'] = (hourlyBH_df['L.HHO.IntraeNB.IntraFreq.ExecSuccOut (None)'] + hourlyBH_df['L.HHO.IntereNB.IntraFreq.ExecSuccOut (None)']) / \
                                                        (hourlyBH_df['L.HHO.IntraeNB.IntraFreq.ExecAttOut (None)'] + hourlyBH_df['L.HHO.IntereNB.IntraFreq.ExecAttOut (None)']) * 100
hourlyBH_df['CSFB to WCDMA'] = hourlyBH_df['L.CSFB.E2W (None)']
hourlyBH_df['CSFB to GERAN'] = hourlyBH_df['L.CSFB.E2G (None)']
hourlyBH_df['RRS setup SR,%'] = hourlyBH_df['L.RRC.ConnReq.Succ (None)'] / hourlyBH_df['L.RRC.ConnReq.Att (None)'] * 100
hourlyBH_df['DCSR4G, %'] = hourlyBH_df['RRS setup SR,%'] * hourlyBH_df['E-RAB Setup SR,%'] * (100 - hourlyBH_df['E-RAB Drop Rate']) / 10000
hourlyBH_df = hourlyBH_df.drop(list_1, axis=1)
hourlyBH_df_trans = hourlyBH_df.transpose()

with pd.ExcelWriter(f"{directory}{csv_name}{output_comment}.xlsx", engine='openpyxl') as writer:
    weekly_df.to_excel(writer, sheet_name='weekly')
    daily_dfall.to_excel(writer, sheet_name='daily')
    hourly_dfall.to_excel(writer, sheet_name='hourly')
    hourlyBH_df.to_excel(writer, sheet_name='busy_hour')
    weekly_df_trans.to_excel(writer, sheet_name='weekly_trans')
    daily_dfall_trans.to_excel(writer, sheet_name='daily_trans')
    hourly_dfall_trans.to_excel(writer, sheet_name='hourly_trans')
    hourlyBH_df_trans.to_excel(writer, sheet_name='busy_hour_trans')

''' переходим к работе с эксель файлом - форматирование строк и добавление графиков
    используем модуль openpyxl'''

my_file = openpyxl.load_workbook(f"{directory}{csv_name}{output_comment}.xlsx")

weekly_sheet = my_file["weekly"]
daily_sheet = my_file["daily"]
hourly_sheet = my_file["hourly"]
busyhour_sheet = my_file["busy_hour"]

weekly_sheet_trans = my_file["weekly_trans"]
daily_sheet_trans = my_file["daily_trans"]
hourly_sheet_trans = my_file["hourly_trans"]
busyhour_sheet_trans = my_file["busy_hour_trans"]

weekly_sheet.column_dimensions["A"].width = 2
daily_sheet.column_dimensions["B"].width = 11
hourly_sheet.column_dimensions["A"].width = 2
busyhour_sheet.column_dimensions["A"].width = 11

weekly_sheet_trans.column_dimensions["A"].width = 35
daily_sheet_trans.column_dimensions["A"].width = 35
hourly_sheet_trans.column_dimensions["A"].width = 35
busyhour_sheet_trans.column_dimensions["A"].width = 35

hourly_sheet.delete_cols(1) # удаляем первые столбцы чтобы номера столбцов для всех KPI были одинаковыми как в дневной статистике
busyhour_sheet.delete_cols(1) # удаляем первые столбцы чтобы номера столбцов для всех KPI были одинаковыми как в дневной статистике

# определение количества строк в таблицах
last_row_weekly = weekly_sheet.max_row
last_row_daily = daily_sheet.max_row
last_row_hourly = hourly_sheet.max_row
last_row_BH = busyhour_sheet.max_row
# выставление правильного формата для столбцов с датами
for r in range(2,(last_row_daily+1)):
    daily_sheet[f'B{r}'].number_format ='DD.MM.YYYY'
for r in range(2, (last_row_hourly+1)):
    hourly_sheet[f'A{r}'].number_format ='DD'
for r in range(2,(last_row_BH+1)):
    busyhour_sheet[f'A{r}'].number_format ='DD.MM.YYYY'
for cell in daily_sheet_trans[2]:
    cell.number_format ='DD.MM.YYYY'

# выставление переноса строк для названий KPI
for cell in weekly_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in daily_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in hourly_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in busyhour_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

#  графики в недельной таблице weekly_sheet
x_values = Reference(weekly_sheet, range_string=(f"weekly!$B$2:$B${last_row_weekly}"))

PStraffic4GGB= Reference(weekly_sheet, min_col=3, min_row=1, max_row=last_row_weekly)
CellAvailability4G= Reference(weekly_sheet, min_col=4, min_row=1, max_row=last_row_weekly)
TotalLTECellsNumber= Reference(weekly_sheet, min_col=5, min_row=1, max_row=last_row_weekly)
DownlinkPRBUR= Reference(weekly_sheet, min_col=6, min_row=1, max_row=last_row_weekly)
UplinkPRBUR= Reference(weekly_sheet, min_col=7, min_row=1, max_row=last_row_weekly)
UEDownlinkAvThrp= Reference(weekly_sheet, min_col=8, min_row=1, max_row=last_row_weekly)
UEUplinkAvThrp= Reference(weekly_sheet, min_col=9, min_row=1, max_row=last_row_weekly)
ERABSetupSR= Reference(weekly_sheet, min_col=10, min_row=1, max_row=last_row_weekly)
ERABDropRate= Reference(weekly_sheet, min_col=11, min_row=1, max_row=last_row_weekly)
InterFreqHOOutSR= Reference(weekly_sheet, min_col=12, min_row=1, max_row=last_row_weekly)
IntraFreqHOOutSR= Reference(weekly_sheet, min_col=13, min_row=1, max_row=last_row_weekly)
CSFBtoWCDMA= Reference(weekly_sheet, min_col=14, min_row=1, max_row=last_row_weekly)
CSFBtoGERAN= Reference(weekly_sheet, min_col=15, min_row=1, max_row=last_row_weekly)
RRSsetupSR= Reference(weekly_sheet, min_col=16, min_row=1, max_row=last_row_weekly)
DCSR4G= Reference(weekly_sheet, min_col=17, min_row=1, max_row=last_row_weekly)

PStraffic_chart = LineChart()
PStraffic_chart.width = 40
PStraffic_chart.height = 10
PStraffic_chart.add_data(PStraffic4GGB, titles_from_data = True)  #
PStraffic_chart.set_categories(x_values)
PStraffic_chart.legend.position = 'b'
weekly_sheet.add_chart(PStraffic_chart, "A18")

CellAvailability4G_chart = LineChart()
CellAvailability4G_chart.width = 40
CellAvailability4G_chart.height = 10
CellAvailability4G_chart.add_data(CellAvailability4G, titles_from_data = True)  #
CellAvailability4G_chart.set_categories(x_values)
CellAvailability4G_chart.legend.position = 'b'
weekly_sheet.add_chart(CellAvailability4G_chart, "A38")

DownlinkPRBUR_chart = LineChart()
DownlinkPRBUR_chart.width = 40
DownlinkPRBUR_chart.height = 10
DownlinkPRBUR_chart.add_data(DownlinkPRBUR, titles_from_data = True)  #
DownlinkPRBUR_chart.set_categories(x_values)
DownlinkPRBUR_chart.legend.position = 'b'
weekly_sheet.add_chart(DownlinkPRBUR_chart, "A58")

UplinkPRBUR_chart = LineChart()
UplinkPRBUR_chart.width = 40
UplinkPRBUR_chart.height = 10
UplinkPRBUR_chart.add_data(UplinkPRBUR, titles_from_data = True)  #
UplinkPRBUR_chart.set_categories(x_values)
UplinkPRBUR_chart.legend.position = 'b'
weekly_sheet.add_chart(UplinkPRBUR_chart, "A78")

UEDownlinkAvThrp_chart = LineChart()
UEDownlinkAvThrp_chart.width = 40
UEDownlinkAvThrp_chart.height = 10
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrp, titles_from_data = True)  #
UEDownlinkAvThrp_chart.set_categories(x_values)
UEDownlinkAvThrp_chart.legend.position = 'b'
weekly_sheet.add_chart(UEDownlinkAvThrp_chart, "A98")

UEUplinkAvThrp_chart = LineChart()
UEUplinkAvThrp_chart.width = 40
UEUplinkAvThrp_chart.height = 10
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrp, titles_from_data = True)  #
UEUplinkAvThrp_chart.set_categories(x_values)
UEUplinkAvThrp_chart.legend.position = 'b'
weekly_sheet.add_chart(UEUplinkAvThrp_chart, "A118")

ERABSetupSR_chart = LineChart()
ERABSetupSR_chart.width = 40
ERABSetupSR_chart.height = 10
ERABSetupSR_chart.add_data(ERABSetupSR, titles_from_data = True)  #
ERABSetupSR_chart.set_categories(x_values)
ERABSetupSR_chart.legend.position = 'b'
weekly_sheet.add_chart(ERABSetupSR_chart, "A138")

ERABDropRate_chart = LineChart()
ERABDropRate_chart.width = 40
ERABDropRate_chart.height = 10
ERABDropRate_chart.add_data(ERABDropRate, titles_from_data = True)  #
ERABDropRate_chart.set_categories(x_values)
ERABDropRate_chart.legend.position = 'b'
weekly_sheet.add_chart(ERABDropRate_chart, "A158")

InterFreqHOOutSR_chart = LineChart()
InterFreqHOOutSR_chart.width = 40
InterFreqHOOutSR_chart.height = 10
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSR, titles_from_data = True)  #
InterFreqHOOutSR_chart.set_categories(x_values)
InterFreqHOOutSR_chart.legend.position = 'b'
weekly_sheet.add_chart(InterFreqHOOutSR_chart, "A178")

IntraFreqHOOutSR_chart = LineChart()
IntraFreqHOOutSR_chart.width = 40
IntraFreqHOOutSR_chart.height = 10
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSR, titles_from_data = True)  #
IntraFreqHOOutSR_chart.set_categories(x_values)
IntraFreqHOOutSR_chart.legend.position = 'b'
weekly_sheet.add_chart(IntraFreqHOOutSR_chart, "A198")

CSFBtoWCDMA_chart = LineChart()
CSFBtoWCDMA_chart.width = 40
CSFBtoWCDMA_chart.height = 10
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMA, titles_from_data = True)  #
CSFBtoWCDMA_chart.set_categories(x_values)
CSFBtoWCDMA_chart.legend.position = 'b'
weekly_sheet.add_chart(CSFBtoWCDMA_chart, "A218")

CSFBtoGERAN_chart = LineChart()
CSFBtoGERAN_chart.width = 40
CSFBtoGERAN_chart.height = 10
CSFBtoGERAN_chart.add_data(CSFBtoGERAN, titles_from_data = True)  #
CSFBtoGERAN_chart.set_categories(x_values)
CSFBtoGERAN_chart.legend.position = 'b'
weekly_sheet.add_chart(CSFBtoGERAN_chart, "A238")

RRSsetupSR_chart = LineChart()
RRSsetupSR_chart.width = 40
RRSsetupSR_chart.height = 10
RRSsetupSR_chart.add_data(RRSsetupSR, titles_from_data = True)  #
RRSsetupSR_chart.set_categories(x_values)
RRSsetupSR_chart.legend.position = 'b'
weekly_sheet.add_chart(RRSsetupSR_chart, "A258")

DCSR4G_chart = LineChart()
DCSR4G_chart.width = 40
DCSR4G_chart.height = 10
DCSR4G_chart.add_data(DCSR4G, titles_from_data = True)  #
DCSR4G_chart.set_categories(x_values)
DCSR4G_chart.legend.position = 'b'
weekly_sheet.add_chart(DCSR4G_chart, "A258")

#  графики в суточной таблице daily_sheet   last_row_daily
x_values = Reference(daily_sheet, range_string=(f"daily!$B$2:$B${last_row_daily}"))

PStraffic4GGB= Reference(daily_sheet, min_col=3, min_row=1, max_row=last_row_daily)
CellAvailability4G= Reference(daily_sheet, min_col=4, min_row=1, max_row=last_row_daily)
TotalLTECellsNumber= Reference(daily_sheet, min_col=5, min_row=1, max_row=last_row_daily)
DownlinkPRBUR= Reference(daily_sheet, min_col=6, min_row=1, max_row=last_row_daily)
UplinkPRBUR= Reference(daily_sheet, min_col=7, min_row=1, max_row=last_row_daily)
UEDownlinkAvThrp= Reference(daily_sheet, min_col=8, min_row=1, max_row=last_row_daily)
UEUplinkAvThrp= Reference(daily_sheet, min_col=9, min_row=1, max_row=last_row_daily)
ERABSetupSR= Reference(daily_sheet, min_col=10, min_row=1, max_row=last_row_daily)
ERABDropRate= Reference(daily_sheet, min_col=11, min_row=1, max_row=last_row_daily)
InterFreqHOOutSR= Reference(daily_sheet, min_col=12, min_row=1, max_row=last_row_daily)
IntraFreqHOOutSR= Reference(daily_sheet, min_col=13, min_row=1, max_row=last_row_daily)
CSFBtoWCDMA= Reference(daily_sheet, min_col=14, min_row=1, max_row=last_row_daily)
CSFBtoGERAN= Reference(daily_sheet, min_col=15, min_row=1, max_row=last_row_daily)
RRSsetupSR= Reference(daily_sheet, min_col=16, min_row=1, max_row=last_row_daily)
DCSR4G= Reference(daily_sheet, min_col=17, min_row=1, max_row=last_row_daily)
PStraffic4GGBL2600= Reference(daily_sheet, min_col=18, min_row=1, max_row=last_row_daily)
CellAvailability4GL2600= Reference(daily_sheet, min_col=19, min_row=1, max_row=last_row_daily)
TotalLTECellsNumberL2600= Reference(daily_sheet, min_col=20, min_row=1, max_row=last_row_daily)
DownlinkPRBURL2600= Reference(daily_sheet, min_col=21, min_row=1, max_row=last_row_daily)
UplinkPRBURL2600= Reference(daily_sheet, min_col=22, min_row=1, max_row=last_row_daily)
UEDownlinkAvThrpL2600= Reference(daily_sheet, min_col=23, min_row=1, max_row=last_row_daily)
UEUplinkAvThrpL2600= Reference(daily_sheet, min_col=24, min_row=1, max_row=last_row_daily)
ERABSetupSRL2600= Reference(daily_sheet, min_col=25, min_row=1, max_row=last_row_daily)
ERABDropRateL2600= Reference(daily_sheet, min_col=26, min_row=1, max_row=last_row_daily)
InterFreqHOOutSRL2600= Reference(daily_sheet, min_col=27, min_row=1, max_row=last_row_daily)
IntraFreqHOOutSRL2600= Reference(daily_sheet, min_col=28, min_row=1, max_row=last_row_daily)
CSFBtoWCDMAL2600= Reference(daily_sheet, min_col=29, min_row=1, max_row=last_row_daily)
CSFBtoGERANL2600= Reference(daily_sheet, min_col=30, min_row=1, max_row=last_row_daily)
RRSsetupSRL2600= Reference(daily_sheet, min_col=31, min_row=1, max_row=last_row_daily)
DCSR4GL2600= Reference(daily_sheet, min_col=32, min_row=1, max_row=last_row_daily)
PStraffic4GGBL1800= Reference(daily_sheet, min_col=33, min_row=1, max_row=last_row_daily)
CellAvailability4GL1800= Reference(daily_sheet, min_col=34, min_row=1, max_row=last_row_daily)
TotalLTECellsNumberL1800= Reference(daily_sheet, min_col=35, min_row=1, max_row=last_row_daily)
DownlinkPRBURL1800= Reference(daily_sheet, min_col=36, min_row=1, max_row=last_row_daily)
UplinkPRBURL1800= Reference(daily_sheet, min_col=37, min_row=1, max_row=last_row_daily)
UEDownlinkAvThrpL1800= Reference(daily_sheet, min_col=38, min_row=1, max_row=last_row_daily)
UEUplinkAvThrpL1800= Reference(daily_sheet, min_col=39, min_row=1, max_row=last_row_daily)
ERABSetupSRL1800= Reference(daily_sheet, min_col=40, min_row=1, max_row=last_row_daily)
ERABDropRateL1800= Reference(daily_sheet, min_col=41, min_row=1, max_row=last_row_daily)
InterFreqHOOutSRL1800= Reference(daily_sheet, min_col=42, min_row=1, max_row=last_row_daily)
IntraFreqHOOutSRL1800= Reference(daily_sheet, min_col=43, min_row=1, max_row=last_row_daily)
CSFBtoWCDMAL1800= Reference(daily_sheet, min_col=44, min_row=1, max_row=last_row_daily)
CSFBtoGERANL1800= Reference(daily_sheet, min_col=45, min_row=1, max_row=last_row_daily)
RRSsetupSRL1800= Reference(daily_sheet, min_col=46, min_row=1, max_row=last_row_daily)
DCSR4GL1800= Reference(daily_sheet, min_col=47, min_row=1, max_row=last_row_daily)
PStraffic4GGBdbL1800= Reference(daily_sheet, min_col=48, min_row=1, max_row=last_row_daily)
CellAvailability4GdbL1800= Reference(daily_sheet, min_col=49, min_row=1, max_row=last_row_daily)
TotalLTECellsNumberdbL1800= Reference(daily_sheet, min_col=50, min_row=1, max_row=last_row_daily)
DownlinkPRBURdbL1800= Reference(daily_sheet, min_col=51, min_row=1, max_row=last_row_daily)
UplinkPRBURdbL1800= Reference(daily_sheet, min_col=52, min_row=1, max_row=last_row_daily)
UEDownlinkAvThrpdbL1800= Reference(daily_sheet, min_col=53, min_row=1, max_row=last_row_daily)
UEUplinkAvThrpdbL1800= Reference(daily_sheet, min_col=54, min_row=1, max_row=last_row_daily)
ERABSetupSRdbL1800= Reference(daily_sheet, min_col=55, min_row=1, max_row=last_row_daily)
ERABDropRatedbL1800= Reference(daily_sheet, min_col=56, min_row=1, max_row=last_row_daily)
InterFreqHOOutSRdbL1800= Reference(daily_sheet, min_col=57, min_row=1, max_row=last_row_daily)
IntraFreqHOOutSRdbL1800= Reference(daily_sheet, min_col=58, min_row=1, max_row=last_row_daily)
CSFBtoWCDMAdbL1800= Reference(daily_sheet, min_col=59, min_row=1, max_row=last_row_daily)
CSFBtoGERANdbL1800= Reference(daily_sheet, min_col=60, min_row=1, max_row=last_row_daily)
RRSsetupSRdbL1800= Reference(daily_sheet, min_col=61, min_row=1, max_row=last_row_daily)
DCSR4GdbL1800= Reference(daily_sheet, min_col=62, min_row=1, max_row=last_row_daily)
PStraffic4GGBdbL2600= Reference(daily_sheet, min_col=63, min_row=1, max_row=last_row_daily)
CellAvailability4GdbL2600= Reference(daily_sheet, min_col=64, min_row=1, max_row=last_row_daily)
TotalLTECellsNumberdbL2600= Reference(daily_sheet, min_col=65, min_row=1, max_row=last_row_daily)
DownlinkPRBURdbL2600= Reference(daily_sheet, min_col=66, min_row=1, max_row=last_row_daily)
UplinkPRBURdbL2600= Reference(daily_sheet, min_col=67, min_row=1, max_row=last_row_daily)
UEDownlinkAvThrpdbL2600= Reference(daily_sheet, min_col=68, min_row=1, max_row=last_row_daily)
UEUplinkAvThrpdbL2600= Reference(daily_sheet, min_col=69, min_row=1, max_row=last_row_daily)
ERABSetupSRdbL2600= Reference(daily_sheet, min_col=70, min_row=1, max_row=last_row_daily)
ERABDropRatedbL2600= Reference(daily_sheet, min_col=71, min_row=1, max_row=last_row_daily)
InterFreqHOOutSRdbL2600= Reference(daily_sheet, min_col=72, min_row=1, max_row=last_row_daily)
IntraFreqHOOutSRdbL2600= Reference(daily_sheet, min_col=73, min_row=1, max_row=last_row_daily)
CSFBtoWCDMAdbL2600= Reference(daily_sheet, min_col=74, min_row=1, max_row=last_row_daily)
CSFBtoGERANdbL2600= Reference(daily_sheet, min_col=75, min_row=1, max_row=last_row_daily)
RRSsetupSRdbL2600= Reference(daily_sheet, min_col=76, min_row=1, max_row=last_row_daily)
DCSR4GdbL2600= Reference(daily_sheet, min_col=77, min_row=1, max_row=last_row_daily)

PStraffic_chart = LineChart()
PStraffic_chart.width = 40
PStraffic_chart.height = 10
PStraffic_chart.add_data(PStraffic4GGB, titles_from_data = True)  #
PStraffic_chart.add_data(PStraffic4GGBL2600, titles_from_data = True)
PStraffic_chart.add_data(PStraffic4GGBL1800, titles_from_data = True)
PStraffic_chart.set_categories(x_values)
PStraffic_chart.legend.position = 'b'
daily_sheet.add_chart(PStraffic_chart, "A18")

CellAvailability4G_chart = LineChart()
CellAvailability4G_chart.width = 40
CellAvailability4G_chart.height = 10
CellAvailability4G_chart.add_data(CellAvailability4G, titles_from_data = True)  #
CellAvailability4G_chart.add_data(CellAvailability4GL2600, titles_from_data = True)
CellAvailability4G_chart.add_data(CellAvailability4GL1800, titles_from_data = True)
CellAvailability4G_chart.set_categories(x_values)
CellAvailability4G_chart.legend.position = 'b'
daily_sheet.add_chart(CellAvailability4G_chart, "A38")

DownlinkPRBUR_chart = LineChart()
DownlinkPRBUR_chart.width = 40
DownlinkPRBUR_chart.height = 10
DownlinkPRBUR_chart.add_data(DownlinkPRBUR, titles_from_data = True)  #
DownlinkPRBUR_chart.add_data(DownlinkPRBURL2600, titles_from_data = True)
DownlinkPRBUR_chart.add_data(DownlinkPRBURL1800, titles_from_data = True)
DownlinkPRBUR_chart.set_categories(x_values)
DownlinkPRBUR_chart.legend.position = 'b'
daily_sheet.add_chart(DownlinkPRBUR_chart, "A58")

UplinkPRBUR_chart = LineChart()
UplinkPRBUR_chart.width = 40
UplinkPRBUR_chart.height = 10
UplinkPRBUR_chart.add_data(UplinkPRBUR, titles_from_data = True)  #
UplinkPRBUR_chart.add_data(UplinkPRBURL2600, titles_from_data = True)
UplinkPRBUR_chart.add_data(UplinkPRBURL1800, titles_from_data = True)
UplinkPRBUR_chart.set_categories(x_values)
UplinkPRBUR_chart.legend.position = 'b'
daily_sheet.add_chart(UplinkPRBUR_chart, "A78")

UEDownlinkAvThrp_chart = LineChart()
UEDownlinkAvThrp_chart.width = 40
UEDownlinkAvThrp_chart.height = 10
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrp, titles_from_data = True)  #
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrpL2600, titles_from_data = True)
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrpL1800, titles_from_data = True)
UEDownlinkAvThrp_chart.set_categories(x_values)
UEDownlinkAvThrp_chart.legend.position = 'b'
daily_sheet.add_chart(UEDownlinkAvThrp_chart, "A98")

UEUplinkAvThrp_chart = LineChart()
UEUplinkAvThrp_chart.width = 40
UEUplinkAvThrp_chart.height = 10
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrp, titles_from_data = True)  #
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrpL2600, titles_from_data = True)
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrpL1800, titles_from_data = True)
UEUplinkAvThrp_chart.set_categories(x_values)
UEUplinkAvThrp_chart.legend.position = 'b'
daily_sheet.add_chart(UEUplinkAvThrp_chart, "A118")

ERABSetupSR_chart = LineChart()
ERABSetupSR_chart.width = 40
ERABSetupSR_chart.height = 10
ERABSetupSR_chart.add_data(ERABSetupSR, titles_from_data = True)  #
ERABSetupSR_chart.add_data(ERABSetupSRL2600, titles_from_data = True)
ERABSetupSR_chart.add_data(ERABSetupSRL1800, titles_from_data = True)
ERABSetupSR_chart.set_categories(x_values)
ERABSetupSR_chart.legend.position = 'b'
daily_sheet.add_chart(ERABSetupSR_chart, "A138")

ERABDropRate_chart = LineChart()
ERABDropRate_chart.width = 40
ERABDropRate_chart.height = 10
ERABDropRate_chart.add_data(ERABDropRate, titles_from_data = True)  #
ERABDropRate_chart.add_data(ERABDropRateL2600, titles_from_data = True)
ERABDropRate_chart.add_data(ERABDropRateL1800, titles_from_data = True)
ERABDropRate_chart.set_categories(x_values)
ERABDropRate_chart.legend.position = 'b'
daily_sheet.add_chart(ERABDropRate_chart, "A158")

InterFreqHOOutSR_chart = LineChart()
InterFreqHOOutSR_chart.width = 40
InterFreqHOOutSR_chart.height = 10
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSR, titles_from_data = True)  #
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSRL2600, titles_from_data = True)
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSRL1800, titles_from_data = True)
InterFreqHOOutSR_chart.set_categories(x_values)
InterFreqHOOutSR_chart.legend.position = 'b'
daily_sheet.add_chart(InterFreqHOOutSR_chart, "A178")

IntraFreqHOOutSR_chart = LineChart()
IntraFreqHOOutSR_chart.width = 40
IntraFreqHOOutSR_chart.height = 10
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSR, titles_from_data = True)  #
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSRL2600, titles_from_data = True)
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSRL1800, titles_from_data = True)
IntraFreqHOOutSR_chart.set_categories(x_values)
IntraFreqHOOutSR_chart.legend.position = 'b'
daily_sheet.add_chart(IntraFreqHOOutSR_chart, "A198")

CSFBtoWCDMA_chart = LineChart()
CSFBtoWCDMA_chart.width = 40
CSFBtoWCDMA_chart.height = 10
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMA, titles_from_data = True)  #
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMAL2600, titles_from_data = True)
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMAL1800, titles_from_data = True)
CSFBtoWCDMA_chart.set_categories(x_values)
CSFBtoWCDMA_chart.legend.position = 'b'
daily_sheet.add_chart(CSFBtoWCDMA_chart, "A218")

CSFBtoGERAN_chart = LineChart()
CSFBtoGERAN_chart.width = 40
CSFBtoGERAN_chart.height = 10
CSFBtoGERAN_chart.add_data(CSFBtoGERAN, titles_from_data = True)  #
CSFBtoGERAN_chart.add_data(CSFBtoGERANL2600, titles_from_data = True)
CSFBtoGERAN_chart.add_data(CSFBtoGERANL1800, titles_from_data = True)
CSFBtoGERAN_chart.set_categories(x_values)
CSFBtoGERAN_chart.legend.position = 'b'
daily_sheet.add_chart(CSFBtoGERAN_chart, "A238")

RRSsetupSR_chart = LineChart()
RRSsetupSR_chart.width = 40
RRSsetupSR_chart.height = 10
RRSsetupSR_chart.add_data(RRSsetupSR, titles_from_data = True)  #
RRSsetupSR_chart.add_data(RRSsetupSRL2600, titles_from_data = True)
RRSsetupSR_chart.add_data(RRSsetupSRL1800, titles_from_data = True)
RRSsetupSR_chart.set_categories(x_values)
RRSsetupSR_chart.legend.position = 'b'
daily_sheet.add_chart(RRSsetupSR_chart, "A258")

DCSR4G_chart = LineChart()
DCSR4G_chart.width = 40
DCSR4G_chart.height = 10
DCSR4G_chart.add_data(DCSR4G, titles_from_data = True)  #
DCSR4G_chart.add_data(DCSR4GL2600, titles_from_data = True)
DCSR4G_chart.add_data(DCSR4GL1800, titles_from_data = True)
DCSR4G_chart.set_categories(x_values)
DCSR4G_chart.legend.position = 'b'
daily_sheet.add_chart(DCSR4G_chart, "A258")

# графики почасовые hourly_sheet  last_row_hourly
x_values = Reference(hourly_sheet, range_string=(f"hourly!$A$2:$B${last_row_hourly}"))

PStraffic4GGB= Reference(hourly_sheet, min_col=3, min_row=1, max_row=last_row_hourly)
CellAvailability4G= Reference(hourly_sheet, min_col=4, min_row=1, max_row=last_row_hourly)
TotalLTECellsNumber= Reference(hourly_sheet, min_col=5, min_row=1, max_row=last_row_hourly)
DownlinkPRBUR= Reference(hourly_sheet, min_col=6, min_row=1, max_row=last_row_hourly)
UplinkPRBUR= Reference(hourly_sheet, min_col=7, min_row=1, max_row=last_row_hourly)
UEDownlinkAvThrp= Reference(hourly_sheet, min_col=8, min_row=1, max_row=last_row_hourly)
UEUplinkAvThrp= Reference(hourly_sheet, min_col=9, min_row=1, max_row=last_row_hourly)
ERABSetupSR= Reference(hourly_sheet, min_col=10, min_row=1, max_row=last_row_hourly)
ERABDropRate= Reference(hourly_sheet, min_col=11, min_row=1, max_row=last_row_hourly)
InterFreqHOOutSR= Reference(hourly_sheet, min_col=12, min_row=1, max_row=last_row_hourly)
IntraFreqHOOutSR= Reference(hourly_sheet, min_col=13, min_row=1, max_row=last_row_hourly)
CSFBtoWCDMA= Reference(hourly_sheet, min_col=14, min_row=1, max_row=last_row_hourly)
CSFBtoGERAN= Reference(hourly_sheet, min_col=15, min_row=1, max_row=last_row_hourly)
RRSsetupSR= Reference(hourly_sheet, min_col=16, min_row=1, max_row=last_row_hourly)
DCSR4G= Reference(hourly_sheet, min_col=17, min_row=1, max_row=last_row_hourly)
PStraffic4GGBL2600= Reference(hourly_sheet, min_col=18, min_row=1, max_row=last_row_hourly)
CellAvailability4GL2600= Reference(hourly_sheet, min_col=19, min_row=1, max_row=last_row_hourly)
TotalLTECellsNumberL2600= Reference(hourly_sheet, min_col=20, min_row=1, max_row=last_row_hourly)
DownlinkPRBURL2600= Reference(hourly_sheet, min_col=21, min_row=1, max_row=last_row_hourly)
UplinkPRBURL2600= Reference(hourly_sheet, min_col=22, min_row=1, max_row=last_row_hourly)
UEDownlinkAvThrpL2600= Reference(hourly_sheet, min_col=23, min_row=1, max_row=last_row_hourly)
UEUplinkAvThrpL2600= Reference(hourly_sheet, min_col=24, min_row=1, max_row=last_row_hourly)
ERABSetupSRL2600= Reference(hourly_sheet, min_col=25, min_row=1, max_row=last_row_hourly)
ERABDropRateL2600= Reference(hourly_sheet, min_col=26, min_row=1, max_row=last_row_hourly)
InterFreqHOOutSRL2600= Reference(hourly_sheet, min_col=27, min_row=1, max_row=last_row_hourly)
IntraFreqHOOutSRL2600= Reference(hourly_sheet, min_col=28, min_row=1, max_row=last_row_hourly)
CSFBtoWCDMAL2600= Reference(hourly_sheet, min_col=29, min_row=1, max_row=last_row_hourly)
CSFBtoGERANL2600= Reference(hourly_sheet, min_col=30, min_row=1, max_row=last_row_hourly)
RRSsetupSRL2600= Reference(hourly_sheet, min_col=31, min_row=1, max_row=last_row_hourly)
DCSR4GL2600= Reference(hourly_sheet, min_col=32, min_row=1, max_row=last_row_hourly)
PStraffic4GGBL1800= Reference(hourly_sheet, min_col=33, min_row=1, max_row=last_row_hourly)
CellAvailability4GL1800= Reference(hourly_sheet, min_col=34, min_row=1, max_row=last_row_hourly)
TotalLTECellsNumberL1800= Reference(hourly_sheet, min_col=35, min_row=1, max_row=last_row_hourly)
DownlinkPRBURL1800= Reference(hourly_sheet, min_col=36, min_row=1, max_row=last_row_hourly)
UplinkPRBURL1800= Reference(hourly_sheet, min_col=37, min_row=1, max_row=last_row_hourly)
UEDownlinkAvThrpL1800= Reference(hourly_sheet, min_col=38, min_row=1, max_row=last_row_hourly)
UEUplinkAvThrpL1800= Reference(hourly_sheet, min_col=39, min_row=1, max_row=last_row_hourly)
ERABSetupSRL1800= Reference(hourly_sheet, min_col=40, min_row=1, max_row=last_row_hourly)
ERABDropRateL1800= Reference(hourly_sheet, min_col=41, min_row=1, max_row=last_row_hourly)
InterFreqHOOutSRL1800= Reference(hourly_sheet, min_col=42, min_row=1, max_row=last_row_hourly)
IntraFreqHOOutSRL1800= Reference(hourly_sheet, min_col=43, min_row=1, max_row=last_row_hourly)
CSFBtoWCDMAL1800= Reference(hourly_sheet, min_col=44, min_row=1, max_row=last_row_hourly)
CSFBtoGERANL1800= Reference(hourly_sheet, min_col=45, min_row=1, max_row=last_row_hourly)
RRSsetupSRL1800= Reference(hourly_sheet, min_col=46, min_row=1, max_row=last_row_hourly)
DCSR4GL1800= Reference(hourly_sheet, min_col=47, min_row=1, max_row=last_row_hourly)
PStraffic4GGBdbL1800= Reference(hourly_sheet, min_col=48, min_row=1, max_row=last_row_hourly)
CellAvailability4GdbL1800= Reference(hourly_sheet, min_col=49, min_row=1, max_row=last_row_hourly)
TotalLTECellsNumberdbL1800= Reference(hourly_sheet, min_col=50, min_row=1, max_row=last_row_hourly)
DownlinkPRBURdbL1800= Reference(hourly_sheet, min_col=51, min_row=1, max_row=last_row_hourly)
UplinkPRBURdbL1800= Reference(hourly_sheet, min_col=52, min_row=1, max_row=last_row_hourly)
UEDownlinkAvThrpdbL1800= Reference(hourly_sheet, min_col=53, min_row=1, max_row=last_row_hourly)
UEUplinkAvThrpdbL1800= Reference(hourly_sheet, min_col=54, min_row=1, max_row=last_row_hourly)
ERABSetupSRdbL1800= Reference(hourly_sheet, min_col=55, min_row=1, max_row=last_row_hourly)
ERABDropRatedbL1800= Reference(hourly_sheet, min_col=56, min_row=1, max_row=last_row_hourly)
InterFreqHOOutSRdbL1800= Reference(hourly_sheet, min_col=57, min_row=1, max_row=last_row_hourly)
IntraFreqHOOutSRdbL1800= Reference(hourly_sheet, min_col=58, min_row=1, max_row=last_row_hourly)
CSFBtoWCDMAdbL1800= Reference(hourly_sheet, min_col=59, min_row=1, max_row=last_row_hourly)
CSFBtoGERANdbL1800= Reference(hourly_sheet, min_col=60, min_row=1, max_row=last_row_hourly)
RRSsetupSRdbL1800= Reference(hourly_sheet, min_col=61, min_row=1, max_row=last_row_hourly)
DCSR4GdbL1800= Reference(hourly_sheet, min_col=62, min_row=1, max_row=last_row_hourly)
PStraffic4GGBdbL2600= Reference(hourly_sheet, min_col=63, min_row=1, max_row=last_row_hourly)
CellAvailability4GdbL2600= Reference(hourly_sheet, min_col=64, min_row=1, max_row=last_row_hourly)
TotalLTECellsNumberdbL2600= Reference(hourly_sheet, min_col=65, min_row=1, max_row=last_row_hourly)
DownlinkPRBURdbL2600= Reference(hourly_sheet, min_col=66, min_row=1, max_row=last_row_hourly)
UplinkPRBURdbL2600= Reference(hourly_sheet, min_col=67, min_row=1, max_row=last_row_hourly)
UEDownlinkAvThrpdbL2600= Reference(hourly_sheet, min_col=68, min_row=1, max_row=last_row_hourly)
UEUplinkAvThrpdbL2600= Reference(hourly_sheet, min_col=69, min_row=1, max_row=last_row_hourly)
ERABSetupSRdbL2600= Reference(hourly_sheet, min_col=70, min_row=1, max_row=last_row_hourly)
ERABDropRatedbL2600= Reference(hourly_sheet, min_col=71, min_row=1, max_row=last_row_hourly)
InterFreqHOOutSRdbL2600= Reference(hourly_sheet, min_col=72, min_row=1, max_row=last_row_hourly)
IntraFreqHOOutSRdbL2600= Reference(hourly_sheet, min_col=73, min_row=1, max_row=last_row_hourly)
CSFBtoWCDMAdbL2600= Reference(hourly_sheet, min_col=74, min_row=1, max_row=last_row_hourly)
CSFBtoGERANdbL2600= Reference(hourly_sheet, min_col=75, min_row=1, max_row=last_row_hourly)
RRSsetupSRdbL2600= Reference(hourly_sheet, min_col=76, min_row=1, max_row=last_row_hourly)
DCSR4GdbL2600= Reference(hourly_sheet, min_col=77, min_row=1, max_row=last_row_hourly)

PStraffic_chart = LineChart()
PStraffic_chart.width = 40
PStraffic_chart.height = 10
PStraffic_chart.add_data(PStraffic4GGB, titles_from_data = True)  #
PStraffic_chart.add_data(PStraffic4GGBL2600, titles_from_data = True)
PStraffic_chart.add_data(PStraffic4GGBL1800, titles_from_data = True)
PStraffic_chart.set_categories(x_values)
PStraffic_chart.legend.position = 'b'
hourly_sheet.add_chart(PStraffic_chart, "A18")

CellAvailability4G_chart = LineChart()
CellAvailability4G_chart.width = 40
CellAvailability4G_chart.height = 10
CellAvailability4G_chart.add_data(CellAvailability4G, titles_from_data = True)  #
CellAvailability4G_chart.add_data(CellAvailability4GL2600, titles_from_data = True)
CellAvailability4G_chart.add_data(CellAvailability4GL1800, titles_from_data = True)
CellAvailability4G_chart.set_categories(x_values)
CellAvailability4G_chart.legend.position = 'b'
hourly_sheet.add_chart(CellAvailability4G_chart, "A38")

DownlinkPRBUR_chart = LineChart()
DownlinkPRBUR_chart.width = 40
DownlinkPRBUR_chart.height = 10
DownlinkPRBUR_chart.add_data(DownlinkPRBUR, titles_from_data = True)  #
DownlinkPRBUR_chart.add_data(DownlinkPRBURL2600, titles_from_data = True)
DownlinkPRBUR_chart.add_data(DownlinkPRBURL1800, titles_from_data = True)
DownlinkPRBUR_chart.set_categories(x_values)
DownlinkPRBUR_chart.legend.position = 'b'
hourly_sheet.add_chart(DownlinkPRBUR_chart, "A58")

UplinkPRBUR_chart = LineChart()
UplinkPRBUR_chart.width = 40
UplinkPRBUR_chart.height = 10
UplinkPRBUR_chart.add_data(UplinkPRBUR, titles_from_data = True)  #
UplinkPRBUR_chart.add_data(UplinkPRBURL2600, titles_from_data = True)
UplinkPRBUR_chart.add_data(UplinkPRBURL1800, titles_from_data = True)
UplinkPRBUR_chart.set_categories(x_values)
UplinkPRBUR_chart.legend.position = 'b'
hourly_sheet.add_chart(UplinkPRBUR_chart, "A78")

UEDownlinkAvThrp_chart = LineChart()
UEDownlinkAvThrp_chart.width = 40
UEDownlinkAvThrp_chart.height = 10
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrp, titles_from_data = True)  #
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrpL2600, titles_from_data = True)
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrpL1800, titles_from_data = True)
UEDownlinkAvThrp_chart.set_categories(x_values)
UEDownlinkAvThrp_chart.legend.position = 'b'
hourly_sheet.add_chart(UEDownlinkAvThrp_chart, "A98")

UEUplinkAvThrp_chart = LineChart()
UEUplinkAvThrp_chart.width = 40
UEUplinkAvThrp_chart.height = 10
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrp, titles_from_data = True)  #
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrpL2600, titles_from_data = True)
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrpL1800, titles_from_data = True)
UEUplinkAvThrp_chart.set_categories(x_values)
UEUplinkAvThrp_chart.legend.position = 'b'
hourly_sheet.add_chart(UEUplinkAvThrp_chart, "A118")

ERABSetupSR_chart = LineChart()
ERABSetupSR_chart.width = 40
ERABSetupSR_chart.height = 10
ERABSetupSR_chart.add_data(ERABSetupSR, titles_from_data = True)  #
ERABSetupSR_chart.add_data(ERABSetupSRL2600, titles_from_data = True)
ERABSetupSR_chart.add_data(ERABSetupSRL1800, titles_from_data = True)
ERABSetupSR_chart.set_categories(x_values)
ERABSetupSR_chart.legend.position = 'b'
hourly_sheet.add_chart(ERABSetupSR_chart, "A138")

ERABDropRate_chart = LineChart()
ERABDropRate_chart.width = 40
ERABDropRate_chart.height = 10
ERABDropRate_chart.add_data(ERABDropRate, titles_from_data = True)  #
ERABDropRate_chart.add_data(ERABDropRateL2600, titles_from_data = True)
ERABDropRate_chart.add_data(ERABDropRateL1800, titles_from_data = True)
ERABDropRate_chart.set_categories(x_values)
ERABDropRate_chart.legend.position = 'b'
hourly_sheet.add_chart(ERABDropRate_chart, "A158")

InterFreqHOOutSR_chart = LineChart()
InterFreqHOOutSR_chart.width = 40
InterFreqHOOutSR_chart.height = 10
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSR, titles_from_data = True)  #
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSRL2600, titles_from_data = True)
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSRL1800, titles_from_data = True)
InterFreqHOOutSR_chart.set_categories(x_values)
InterFreqHOOutSR_chart.legend.position = 'b'
hourly_sheet.add_chart(InterFreqHOOutSR_chart, "A178")

IntraFreqHOOutSR_chart = LineChart()
IntraFreqHOOutSR_chart.width = 40
IntraFreqHOOutSR_chart.height = 10
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSR, titles_from_data = True)  #
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSRL2600, titles_from_data = True)
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSRL1800, titles_from_data = True)
IntraFreqHOOutSR_chart.set_categories(x_values)
IntraFreqHOOutSR_chart.legend.position = 'b'
hourly_sheet.add_chart(IntraFreqHOOutSR_chart, "A198")

CSFBtoWCDMA_chart = LineChart()
CSFBtoWCDMA_chart.width = 40
CSFBtoWCDMA_chart.height = 10
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMA, titles_from_data = True)  #
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMAL2600, titles_from_data = True)
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMAL1800, titles_from_data = True)
CSFBtoWCDMA_chart.set_categories(x_values)
CSFBtoWCDMA_chart.legend.position = 'b'
hourly_sheet.add_chart(CSFBtoWCDMA_chart, "A218")

CSFBtoGERAN_chart = LineChart()
CSFBtoGERAN_chart.width = 40
CSFBtoGERAN_chart.height = 10
CSFBtoGERAN_chart.add_data(CSFBtoGERAN, titles_from_data = True)  #
CSFBtoGERAN_chart.add_data(CSFBtoGERANL2600, titles_from_data = True)
CSFBtoGERAN_chart.add_data(CSFBtoGERANL1800, titles_from_data = True)
CSFBtoGERAN_chart.set_categories(x_values)
CSFBtoGERAN_chart.legend.position = 'b'
hourly_sheet.add_chart(CSFBtoGERAN_chart, "A238")

RRSsetupSR_chart = LineChart()
RRSsetupSR_chart.width = 40
RRSsetupSR_chart.height = 10
RRSsetupSR_chart.add_data(RRSsetupSR, titles_from_data = True)  #
RRSsetupSR_chart.add_data(RRSsetupSRL2600, titles_from_data = True)
RRSsetupSR_chart.add_data(RRSsetupSRL1800, titles_from_data = True)
RRSsetupSR_chart.set_categories(x_values)
RRSsetupSR_chart.legend.position = 'b'
hourly_sheet.add_chart(RRSsetupSR_chart, "A258")

DCSR4G_chart = LineChart()
DCSR4G_chart.width = 40
DCSR4G_chart.height = 10
DCSR4G_chart.add_data(DCSR4G, titles_from_data = True)  #
DCSR4G_chart.add_data(DCSR4GL2600, titles_from_data = True)
DCSR4G_chart.add_data(DCSR4GL1800, titles_from_data = True)
DCSR4G_chart.set_categories(x_values)
DCSR4G_chart.legend.position = 'b'
hourly_sheet.add_chart(DCSR4G_chart, "A258")

# графики в чнн last_row_BH, busyhour_sheet.max_row
x_values = Reference(hourly_sheet, range_string=(f"busy_hour!$A$2:$B${last_row_BH}"))

PStraffic4GGB= Reference(busyhour_sheet, min_col=3, min_row=1, max_row=last_row_BH)
CellAvailability4G= Reference(busyhour_sheet, min_col=4, min_row=1, max_row=last_row_BH)
TotalLTECellsNumber= Reference(busyhour_sheet, min_col=5, min_row=1, max_row=last_row_BH)
DownlinkPRBUR= Reference(busyhour_sheet, min_col=6, min_row=1, max_row=last_row_BH)
UplinkPRBUR= Reference(busyhour_sheet, min_col=7, min_row=1, max_row=last_row_BH)
UEDownlinkAvThrp= Reference(busyhour_sheet, min_col=8, min_row=1, max_row=last_row_BH)
UEUplinkAvThrp= Reference(busyhour_sheet, min_col=9, min_row=1, max_row=last_row_BH)
ERABSetupSR= Reference(busyhour_sheet, min_col=10, min_row=1, max_row=last_row_BH)
ERABDropRate= Reference(busyhour_sheet, min_col=11, min_row=1, max_row=last_row_BH)
InterFreqHOOutSR= Reference(busyhour_sheet, min_col=12, min_row=1, max_row=last_row_BH)
IntraFreqHOOutSR= Reference(busyhour_sheet, min_col=13, min_row=1, max_row=last_row_BH)
CSFBtoWCDMA= Reference(busyhour_sheet, min_col=14, min_row=1, max_row=last_row_BH)
CSFBtoGERAN= Reference(busyhour_sheet, min_col=15, min_row=1, max_row=last_row_BH)
RRSsetupSR= Reference(busyhour_sheet, min_col=16, min_row=1, max_row=last_row_BH)
DCSR4G= Reference(busyhour_sheet, min_col=17, min_row=1, max_row=last_row_BH)
PStraffic4GGBL2600= Reference(busyhour_sheet, min_col=18, min_row=1, max_row=last_row_BH)
CellAvailability4GL2600= Reference(busyhour_sheet, min_col=19, min_row=1, max_row=last_row_BH)
TotalLTECellsNumberL2600= Reference(busyhour_sheet, min_col=20, min_row=1, max_row=last_row_BH)
DownlinkPRBURL2600= Reference(busyhour_sheet, min_col=21, min_row=1, max_row=last_row_BH)
UplinkPRBURL2600= Reference(busyhour_sheet, min_col=22, min_row=1, max_row=last_row_BH)
UEDownlinkAvThrpL2600= Reference(busyhour_sheet, min_col=23, min_row=1, max_row=last_row_BH)
UEUplinkAvThrpL2600= Reference(busyhour_sheet, min_col=24, min_row=1, max_row=last_row_BH)
ERABSetupSRL2600= Reference(busyhour_sheet, min_col=25, min_row=1, max_row=last_row_BH)
ERABDropRateL2600= Reference(busyhour_sheet, min_col=26, min_row=1, max_row=last_row_BH)
InterFreqHOOutSRL2600= Reference(busyhour_sheet, min_col=27, min_row=1, max_row=last_row_BH)
IntraFreqHOOutSRL2600= Reference(busyhour_sheet, min_col=28, min_row=1, max_row=last_row_BH)
CSFBtoWCDMAL2600= Reference(busyhour_sheet, min_col=29, min_row=1, max_row=last_row_BH)
CSFBtoGERANL2600= Reference(busyhour_sheet, min_col=30, min_row=1, max_row=last_row_BH)
RRSsetupSRL2600= Reference(busyhour_sheet, min_col=31, min_row=1, max_row=last_row_BH)
DCSR4GL2600= Reference(busyhour_sheet, min_col=32, min_row=1, max_row=last_row_BH)
PStraffic4GGBL1800= Reference(busyhour_sheet, min_col=33, min_row=1, max_row=last_row_BH)
CellAvailability4GL1800= Reference(busyhour_sheet, min_col=34, min_row=1, max_row=last_row_BH)
TotalLTECellsNumberL1800= Reference(busyhour_sheet, min_col=35, min_row=1, max_row=last_row_BH)
DownlinkPRBURL1800= Reference(busyhour_sheet, min_col=36, min_row=1, max_row=last_row_BH)
UplinkPRBURL1800= Reference(busyhour_sheet, min_col=37, min_row=1, max_row=last_row_BH)
UEDownlinkAvThrpL1800= Reference(busyhour_sheet, min_col=38, min_row=1, max_row=last_row_BH)
UEUplinkAvThrpL1800= Reference(busyhour_sheet, min_col=39, min_row=1, max_row=last_row_BH)
ERABSetupSRL1800= Reference(busyhour_sheet, min_col=40, min_row=1, max_row=last_row_BH)
ERABDropRateL1800= Reference(busyhour_sheet, min_col=41, min_row=1, max_row=last_row_BH)
InterFreqHOOutSRL1800= Reference(busyhour_sheet, min_col=42, min_row=1, max_row=last_row_BH)
IntraFreqHOOutSRL1800= Reference(busyhour_sheet, min_col=43, min_row=1, max_row=last_row_BH)
CSFBtoWCDMAL1800= Reference(busyhour_sheet, min_col=44, min_row=1, max_row=last_row_BH)
CSFBtoGERANL1800= Reference(busyhour_sheet, min_col=45, min_row=1, max_row=last_row_BH)
RRSsetupSRL1800= Reference(busyhour_sheet, min_col=46, min_row=1, max_row=last_row_BH)
DCSR4GL1800= Reference(busyhour_sheet, min_col=47, min_row=1, max_row=last_row_BH)
PStraffic4GGBdbL1800= Reference(busyhour_sheet, min_col=48, min_row=1, max_row=last_row_BH)
CellAvailability4GdbL1800= Reference(busyhour_sheet, min_col=49, min_row=1, max_row=last_row_BH)
TotalLTECellsNumberdbL1800= Reference(busyhour_sheet, min_col=50, min_row=1, max_row=last_row_BH)
DownlinkPRBURdbL1800= Reference(busyhour_sheet, min_col=51, min_row=1, max_row=last_row_BH)
UplinkPRBURdbL1800= Reference(busyhour_sheet, min_col=52, min_row=1, max_row=last_row_BH)
UEDownlinkAvThrpdbL1800= Reference(busyhour_sheet, min_col=53, min_row=1, max_row=last_row_BH)
UEUplinkAvThrpdbL1800= Reference(busyhour_sheet, min_col=54, min_row=1, max_row=last_row_BH)
ERABSetupSRdbL1800= Reference(busyhour_sheet, min_col=55, min_row=1, max_row=last_row_BH)
ERABDropRatedbL1800= Reference(busyhour_sheet, min_col=56, min_row=1, max_row=last_row_BH)
InterFreqHOOutSRdbL1800= Reference(busyhour_sheet, min_col=57, min_row=1, max_row=last_row_BH)
IntraFreqHOOutSRdbL1800= Reference(busyhour_sheet, min_col=58, min_row=1, max_row=last_row_BH)
CSFBtoWCDMAdbL1800= Reference(busyhour_sheet, min_col=59, min_row=1, max_row=last_row_BH)
CSFBtoGERANdbL1800= Reference(busyhour_sheet, min_col=60, min_row=1, max_row=last_row_BH)
RRSsetupSRdbL1800= Reference(busyhour_sheet, min_col=61, min_row=1, max_row=last_row_BH)
DCSR4GdbL1800= Reference(busyhour_sheet, min_col=62, min_row=1, max_row=last_row_BH)
PStraffic4GGBdbL2600= Reference(busyhour_sheet, min_col=63, min_row=1, max_row=last_row_BH)
CellAvailability4GdbL2600= Reference(busyhour_sheet, min_col=64, min_row=1, max_row=last_row_BH)
TotalLTECellsNumberdbL2600= Reference(busyhour_sheet, min_col=65, min_row=1, max_row=last_row_BH)
DownlinkPRBURdbL2600= Reference(busyhour_sheet, min_col=66, min_row=1, max_row=last_row_BH)
UplinkPRBURdbL2600= Reference(busyhour_sheet, min_col=67, min_row=1, max_row=last_row_BH)
UEDownlinkAvThrpdbL2600= Reference(busyhour_sheet, min_col=68, min_row=1, max_row=last_row_BH)
UEUplinkAvThrpdbL2600= Reference(busyhour_sheet, min_col=69, min_row=1, max_row=last_row_BH)
ERABSetupSRdbL2600= Reference(busyhour_sheet, min_col=70, min_row=1, max_row=last_row_BH)
ERABDropRatedbL2600= Reference(busyhour_sheet, min_col=71, min_row=1, max_row=last_row_BH)
InterFreqHOOutSRdbL2600= Reference(busyhour_sheet, min_col=72, min_row=1, max_row=last_row_BH)
IntraFreqHOOutSRdbL2600= Reference(busyhour_sheet, min_col=73, min_row=1, max_row=last_row_BH)
CSFBtoWCDMAdbL2600= Reference(busyhour_sheet, min_col=74, min_row=1, max_row=last_row_BH)
CSFBtoGERANdbL2600= Reference(busyhour_sheet, min_col=75, min_row=1, max_row=last_row_BH)
RRSsetupSRdbL2600= Reference(busyhour_sheet, min_col=76, min_row=1, max_row=last_row_BH)
DCSR4GdbL2600= Reference(busyhour_sheet, min_col=77, min_row=1, max_row=last_row_BH)

PStraffic_chart = LineChart()
PStraffic_chart.width = 40
PStraffic_chart.height = 10
PStraffic_chart.add_data(PStraffic4GGB, titles_from_data = True)  #
PStraffic_chart.add_data(PStraffic4GGBL2600, titles_from_data = True)
PStraffic_chart.add_data(PStraffic4GGBL1800, titles_from_data = True)
PStraffic_chart.set_categories(x_values)
PStraffic_chart.legend.position = 'b'
busyhour_sheet.add_chart(PStraffic_chart, "A18")

CellAvailability4G_chart = LineChart()
CellAvailability4G_chart.width = 40
CellAvailability4G_chart.height = 10
CellAvailability4G_chart.add_data(CellAvailability4G, titles_from_data = True)  #
CellAvailability4G_chart.add_data(CellAvailability4GL2600, titles_from_data = True)
CellAvailability4G_chart.add_data(CellAvailability4GL1800, titles_from_data = True)
CellAvailability4G_chart.set_categories(x_values)
CellAvailability4G_chart.legend.position = 'b'
busyhour_sheet.add_chart(CellAvailability4G_chart, "A38")

DownlinkPRBUR_chart = LineChart()
DownlinkPRBUR_chart.width = 40
DownlinkPRBUR_chart.height = 10
DownlinkPRBUR_chart.add_data(DownlinkPRBUR, titles_from_data = True)  #
DownlinkPRBUR_chart.add_data(DownlinkPRBURL2600, titles_from_data = True)
DownlinkPRBUR_chart.add_data(DownlinkPRBURL1800, titles_from_data = True)
DownlinkPRBUR_chart.set_categories(x_values)
DownlinkPRBUR_chart.legend.position = 'b'
busyhour_sheet.add_chart(DownlinkPRBUR_chart, "A58")

UplinkPRBUR_chart = LineChart()
UplinkPRBUR_chart.width = 40
UplinkPRBUR_chart.height = 10
UplinkPRBUR_chart.add_data(UplinkPRBUR, titles_from_data = True)  #
UplinkPRBUR_chart.add_data(UplinkPRBURL2600, titles_from_data = True)
UplinkPRBUR_chart.add_data(UplinkPRBURL1800, titles_from_data = True)
UplinkPRBUR_chart.set_categories(x_values)
UplinkPRBUR_chart.legend.position = 'b'
busyhour_sheet.add_chart(UplinkPRBUR_chart, "A78")

UEDownlinkAvThrp_chart = LineChart()
UEDownlinkAvThrp_chart.width = 40
UEDownlinkAvThrp_chart.height = 10
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrp, titles_from_data = True)  #
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrpL2600, titles_from_data = True)
UEDownlinkAvThrp_chart.add_data(UEDownlinkAvThrpL1800, titles_from_data = True)
UEDownlinkAvThrp_chart.set_categories(x_values)
UEDownlinkAvThrp_chart.legend.position = 'b'
busyhour_sheet.add_chart(UEDownlinkAvThrp_chart, "A98")

UEUplinkAvThrp_chart = LineChart()
UEUplinkAvThrp_chart.width = 40
UEUplinkAvThrp_chart.height = 10
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrp, titles_from_data = True)  #
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrpL2600, titles_from_data = True)
UEUplinkAvThrp_chart.add_data(UEUplinkAvThrpL1800, titles_from_data = True)
UEUplinkAvThrp_chart.set_categories(x_values)
UEUplinkAvThrp_chart.legend.position = 'b'
busyhour_sheet.add_chart(UEUplinkAvThrp_chart, "A118")

ERABSetupSR_chart = LineChart()
ERABSetupSR_chart.width = 40
ERABSetupSR_chart.height = 10
ERABSetupSR_chart.add_data(ERABSetupSR, titles_from_data = True)  #
ERABSetupSR_chart.add_data(ERABSetupSRL2600, titles_from_data = True)
ERABSetupSR_chart.add_data(ERABSetupSRL1800, titles_from_data = True)
ERABSetupSR_chart.set_categories(x_values)
ERABSetupSR_chart.legend.position = 'b'
busyhour_sheet.add_chart(ERABSetupSR_chart, "A138")

ERABDropRate_chart = LineChart()
ERABDropRate_chart.width = 40
ERABDropRate_chart.height = 10
ERABDropRate_chart.add_data(ERABDropRate, titles_from_data = True)  #
ERABDropRate_chart.add_data(ERABDropRateL2600, titles_from_data = True)
ERABDropRate_chart.add_data(ERABDropRateL1800, titles_from_data = True)
ERABDropRate_chart.set_categories(x_values)
ERABDropRate_chart.legend.position = 'b'
busyhour_sheet.add_chart(ERABDropRate_chart, "A158")

InterFreqHOOutSR_chart = LineChart()
InterFreqHOOutSR_chart.width = 40
InterFreqHOOutSR_chart.height = 10
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSR, titles_from_data = True)  #
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSRL2600, titles_from_data = True)
InterFreqHOOutSR_chart.add_data(InterFreqHOOutSRL1800, titles_from_data = True)
InterFreqHOOutSR_chart.set_categories(x_values)
InterFreqHOOutSR_chart.legend.position = 'b'
busyhour_sheet.add_chart(InterFreqHOOutSR_chart, "A178")

IntraFreqHOOutSR_chart = LineChart()
IntraFreqHOOutSR_chart.width = 40
IntraFreqHOOutSR_chart.height = 10
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSR, titles_from_data = True)  #
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSRL2600, titles_from_data = True)
IntraFreqHOOutSR_chart.add_data(IntraFreqHOOutSRL1800, titles_from_data = True)
IntraFreqHOOutSR_chart.set_categories(x_values)
IntraFreqHOOutSR_chart.legend.position = 'b'
busyhour_sheet.add_chart(IntraFreqHOOutSR_chart, "A198")

CSFBtoWCDMA_chart = LineChart()
CSFBtoWCDMA_chart.width = 40
CSFBtoWCDMA_chart.height = 10
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMA, titles_from_data = True)  #
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMAL2600, titles_from_data = True)
CSFBtoWCDMA_chart.add_data(CSFBtoWCDMAL1800, titles_from_data = True)
CSFBtoWCDMA_chart.set_categories(x_values)
CSFBtoWCDMA_chart.legend.position = 'b'
busyhour_sheet.add_chart(CSFBtoWCDMA_chart, "A218")

CSFBtoGERAN_chart = LineChart()
CSFBtoGERAN_chart.width = 40
CSFBtoGERAN_chart.height = 10
CSFBtoGERAN_chart.add_data(CSFBtoGERAN, titles_from_data = True)  #
CSFBtoGERAN_chart.add_data(CSFBtoGERANL2600, titles_from_data = True)
CSFBtoGERAN_chart.add_data(CSFBtoGERANL1800, titles_from_data = True)
CSFBtoGERAN_chart.set_categories(x_values)
CSFBtoGERAN_chart.legend.position = 'b'
busyhour_sheet.add_chart(CSFBtoGERAN_chart, "A238")

RRSsetupSR_chart = LineChart()
RRSsetupSR_chart.width = 40
RRSsetupSR_chart.height = 10
RRSsetupSR_chart.add_data(RRSsetupSR, titles_from_data = True)  #
RRSsetupSR_chart.add_data(RRSsetupSRL2600, titles_from_data = True)
RRSsetupSR_chart.add_data(RRSsetupSRL1800, titles_from_data = True)
RRSsetupSR_chart.set_categories(x_values)
RRSsetupSR_chart.legend.position = 'b'
busyhour_sheet.add_chart(RRSsetupSR_chart, "A258")

DCSR4G_chart = LineChart()
DCSR4G_chart.width = 40
DCSR4G_chart.height = 10
DCSR4G_chart.add_data(DCSR4G, titles_from_data = True)  #
DCSR4G_chart.add_data(DCSR4GL2600, titles_from_data = True)
DCSR4G_chart.add_data(DCSR4GL1800, titles_from_data = True)
DCSR4G_chart.set_categories(x_values)
DCSR4G_chart.legend.position = 'b'
busyhour_sheet.add_chart(DCSR4G_chart, "A258")

my_file.save(f"{directory}{csv_name}{output_comment}.xlsx")
#daily_df.to_excel("C:/test/sts4G/daily.xls", engine='openpyxl', sheet_name='Book1')
print('готово')
frequency = 2500  # Set Frequency To 2500 Hertz
duration = 1000  # Set Duration To 1000 ms == 1 second
winsound.Beep(frequency, duration)


# закрываем соединение с базой данных
conn.close()