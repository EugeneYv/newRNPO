import sqlite3
import pandas as pd
import winsound
import openpyxl
from openpyxl.chart import (LineChart, Reference)
import openpyxl.styles
'''не забыть -  фильтрация для ucell и NodeB!!! не забыть output comment!!!
код для создания экселевского файла с KPI из статистики из БД stsDB '''

# подключаемся к базе данных
conn = sqlite3.connect('C:/SQLite/firstDB/stsDB.db')
# создаем курсор для выполнения запросов
cursor = conn.cursor()
start_date = '2023-05-08' # ввести вручную начальную дату !!!!
end_date = '2023-05-25' # ввести вручную конечную дату !!!!надо брать на день позже

# используйте операторы сравнения для выборки строк в заданном диапазоне
query1 = f"SELECT * FROM UMTS_1v2 WHERE `Start Time` >= '{start_date}' AND `Start Time` <= '{end_date}'"
query2 = f"SELECT * FROM UMTS_2_v2 WHERE `Start Time` >= '{start_date}' AND `Start Time` <= '{end_date}'"
query3 = f"SELECT * FROM NodeBsts WHERE `Start Time` >= '{start_date}' AND `Start Time` <= '{end_date}'"
# выполняем запрос и получаем данные
cursor.execute(query1)
data1 = cursor.fetchall()
data1 = [[None if col == 'NIL' else col for col in row] for row in data1]
sts1_df = pd.DataFrame(data1, columns=[i[0] for i in cursor.description])

cursor.execute(query2)
data2 = cursor.fetchall()
data2 = [[None if col == 'NIL' else col for col in row] for row in data2]
sts2_df = pd.DataFrame(data2, columns=[i[0] for i in cursor.description])
list_drop2 = ['Start Time', 'Period (min)', 'NE Name', 'BSC6910UCell']  # test
sts2_df = sts2_df.drop(list_drop2, axis=1) # test

cursor.execute(query3)
data3 = cursor.fetchall()
data3 = [[None if col == 'NIL' else col for col in row] for row in data3]
data3 = [[None if col == '' else col for col in row] for row in data3] # пробую решить проблему 'VS.DataOutput.AllHSDPA.Traffic (bit)'
stsN_df = pd.DataFrame(data3, columns=[i[0] for i in cursor.description])
conn.close()


directory = 'C:/wHesron/sts/3G/'
csv_name1 = (f'3G{end_date}')
output_comment = 'cluster_UH_noUH0845_top'  # что добавится в конце к названию файла

sts1_df['date'] = sts1_df['Start Time'].str.split(' ').str[0]
sts1_df['hour'] = sts1_df['Start Time'].str.split(' ').str[1]
sts1_df['date'] = pd.to_datetime(sts1_df['date'])
sts1_df['week'] = sts1_df['date'].dt.isocalendar().week

#sts_df = pd.merge(sts1_df, sts2_df, how="left")
sts_df = pd.concat([sts1_df, sts2_df], ignore_index=False, sort=False, axis=1)

stsN_df['date'] = stsN_df['Start Time'].str.split(' ').str[0]
stsN_df['hour'] = stsN_df['Start Time'].str.split(' ').str[1]
stsN_df['date'] = pd.to_datetime(stsN_df['date'])
stsN_df['week'] = stsN_df['date'].dt.isocalendar().week
print('sts сформированы')
list_1 = ['RRC.AttConnEstab.EmgCall (None)','RRC.AttConnEstab.OrgConvCall (None)','RRC.AttConnEstab.TmConvCall (None)',\
          'RRC.SuccConnEstab.EmgCall (None)','RRC.SuccConnEstab.OrgConvCall (None)','RRC.SuccConnEstab.TmConvCall (None)',\
          'VS.AttCellUpdt.EmgCall.PCH (None)','VS.AttCellUpdt.OrgConvCall.PCH (None)',\
          'VS.AttCellUpdt.TmConvCall.PCH (None)','VS.DCCC.D2P.Succ (None)','VS.DCCC.Succ.D2U (None)','VS.DCCC.Succ.F2P (None)',\
          'VS.DCCC.Succ.F2U (None)','VS.HSDPA.H2D.Succ (None)','VS.HSDPA.H2F.Succ (None)','VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)',\
          'VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)','VS.HSDPA.MeanChThroughput (kbit/s)','VS.HSDPA.MeanChThroughput.TotalBytes (byte)',\
          'VS.HSDPA.RAB.AbnormRel (None)','VS.HSDPA.RAB.NormRel (None)','VS.HSUPA.MeanChThroughput (kbit/s)','VS.HSUPA.MeanChThroughput.TotalBytes (byte)',\
          'VS.PS.Bkg.DL.8.Traffic (bit)','VS.PS.Bkg.DL.16.Traffic (bit)','VS.PS.Bkg.DL.32.Traffic (bit)','VS.PS.Bkg.DL.64.Traffic (bit)',\
          'VS.PS.Bkg.DL.128.Traffic (bit)','VS.PS.Bkg.DL.144.Traffic (bit)','VS.PS.Bkg.DL.256.Traffic (bit)','VS.PS.Bkg.DL.384.Traffic (bit)',\
          'VS.PS.Bkg.UL.8.Traffic (bit)','VS.PS.Bkg.UL.16.Traffic (bit)','VS.PS.Bkg.UL.32.Traffic (bit)','VS.PS.Bkg.UL.64.Traffic (bit)',\
          'VS.PS.Bkg.UL.128.Traffic (bit)','VS.PS.Bkg.UL.144.Traffic (bit)','VS.PS.Bkg.UL.256.Traffic (bit)','VS.PS.Bkg.UL.384.Traffic (bit)',\
          'VS.PS.Int.DL.8.Traffic (bit)','VS.PS.Int.DL.16.Traffic (bit)','VS.PS.Int.DL.32.Traffic (bit)','VS.PS.Int.DL.64.Traffic (bit)',\
          'VS.PS.Int.DL.128.Traffic (bit)','VS.PS.Int.DL.144.Traffic (bit)','VS.PS.Int.DL.256.Traffic (bit)','VS.PS.Int.DL.384.Traffic (bit)',\
          'VS.PS.Int.UL.8.Traffic (bit)','VS.PS.Int.UL.16.Traffic (bit)','VS.PS.Int.UL.32.Traffic (bit)','VS.PS.Int.UL.64.Traffic (bit)',\
          'VS.PS.Int.UL.128.Traffic (bit)','VS.PS.Int.UL.144.Traffic (bit)','VS.PS.Int.UL.256.Traffic (bit)','VS.PS.Int.UL.384.Traffic (bit)',\
          'VS.PS.Str.DL.32.Traffic (bit)','VS.PS.Str.DL.64.Traffic (bit)','VS.PS.Str.DL.128.Traffic (bit)','VS.PS.Str.DL.144.Traffic (bit)',\
          'VS.PS.Str.UL.16.Traffic (bit)','VS.PS.Str.UL.32.Traffic (bit)','VS.PS.Str.UL.64.Traffic (bit)','VS.RAB.AbnormRel.CS (None)',\
          'VS.RAB.AbnormRel.PS (None)','VS.RAB.AbnormRel.PS.D2P (None)','VS.RAB.AbnormRel.PS.F2P (None)','VS.RAB.AbnormRel.PS.PCH (None)',\
          'VS.RAB.NormRel.CS (None)','VS.RAB.NormRel.PS (None)','VS.RAB.NormRel.PS.PCH (None)','VS.SuccCellUpdt.EmgCall.PCH (None)',\
          'VS.SuccCellUpdt.OrgConvCall.PCH (None)','VS.SuccCellUpdt.TmConvCall.PCH (None)','CS Voice Traffic Volume (Erl)',\
          'VS.RAB.AttEstabPS.Bkg (None)','VS.RAB.AttEstabPS.Int (None)','VS.RAB.AttEstabPS.Str (None)','VS.RAB.FailEstabPS.Code.Cong (None)',\
          'VS.RAB.FailEstabPS.DLCE.Cong (None)','VS.RAB.FailEstabPS.DLIUBBand.Cong (None)','VS.RAB.FailEstabPS.DLPower.Cong (None)',\
          'VS.RAB.FailEstabPS.HSDPAUser.Cong (None)','VS.RAB.FailEstabPS.HSUPAUser.Cong (None)','VS.RAB.FailEstabPS.ULCE.Cong (None)',\
          'VS.RAB.FailEstabPS.ULIUBBand.Cong (None)','VS.RAB.FailEstabPS.ULPower.Cong (None)','VS.SHO.AttRLAdd (None)','VS.SHO.AttRLDel (None)',\
          'VS.SHO.SuccRLAdd (None)','VS.SHO.SuccRLDel (None)','VS.HHO.AttInterFreqOut (None)','VS.HHO.SuccInterFreqOut (None)','VS.IRATHOCS.Cancel.ReEstab (None)',\
          'VS.IRATHOPS.Cancel.ReEstab (None)','IRATHO.SuccOutCS (None)','IRATHO.AttOutCS (None)','IRATHO.SuccOutPSUTRAN (None)','IRATHO.AttOutPSUTRAN (None)',\
          'VS.RAB.AttEstab.AMR (None)','VS.RAB.SuccEstabCS.AMR (None)','VS.RAB.AttEstabPS.Conv (None)','VS.RAB.SuccEstabPS.Conv (None)','VS.RAB.SuccEstabPS.Bkg (None)',\
          'VS.RAB.SuccEstabPS.Int (None)','VS.RAB.SuccEstabPS.Str (None)','RRC.SuccConnEstab.sum (None)','VS.RRC.AttConnEstab.Sum (None)','RRC.AttConnRelCCCH.Cong (None)',\
          'RRC.AttConnRelCCCH.Preempt (None)','RRC.AttConnRelCCCH.ReEstRej (None)','RRC.AttConnRelCCCH.Unspec (None)','RRC.AttConnRelDCCH.Cong (None)',\
          'RRC.AttConnRelDCCH.Preempt (None)','RRC.AttConnRelDCCH.ReEstRej (None)','RRC.AttConnRelDCCH.Unspec (None)','VS.RRC.ConnRel.CellUpd (None)',\
          'RRC.AttConnRelDCCH.DSCR (None)','RRC.AttConnRelDCCH.UsrInact (None)','RRC.AttConnRelCCCH.DSCR (None)','RRC.AttConnRelDCCH.Norm (None)',\
          'RRC.AttConnRelCCCH.Norm (None)','RRC.AttConnRelCCCH.UsrInact (None)', 'VS.Cell.UnavailTime (s)', 'VS.Cell.UnavailTime.Sys (s)']  # список счётчиков
list_2 = ['CS traffic 3G, Erl','PS traffic 3G UL+DL, GB','CS RAB Drop Rate (%)','PS Blocking Rate (%)','PS RAB Drop Rate (%)','PS HS- Drop Rate (%)',\
          'HSDPA Throughput, kbps','HSUPA Throughput, kbps','Soft Handover Success rate, %','Hard Handover Success rate, %','CS W2G Inter-RAT Handover Out SR',\
          'RRC Assignment SucessRate (CS BH), %','RRC Assignment SucessRate (PS BH), %','RRC Drop Rate (CS BH), %','RRC Drop Rate (PS BH), %',\
          'RAB Assignment Success Rate (CS), %','RAB Assignment Success Rate (PS), %','CCSR CS,%','CCSR PS,%', 'UcellUnavail']  # список KPI
list_1N = [ 'VS.HSDPA.DataOutput.Traffic (bit)', 'VS.HSDPA.DataTtiNum.User (None)', 'VS.DataOutput.AllHSDPA.Traffic (bit)', 'VS.AllHSDPA.DataTtiNum.User (None)', \
            'VS.HSUPA.2msTTI.Traffic (kbit)', 'VS.HSUPA.10msTTI.Traffic (kbit)', 'VS.HSUPA.2msPDU.TTI.Num (None)', 'VS.HSUPA.10msPDU.TTI.Num (None)']
list_2N =[
'MeanThrHSDPA,kbps', \
'MeanThrHSDPA DC,kbps', \
'MeanThrHSUPA,kbps'
]
list_cluster_K = ['Label=UH0881_U94, CellID=48814, LogicRNCID=501', \
                        'Label=UH0881_U96, CellID=48816, LogicRNCID=501', \
                        'Label=UH0881_U95, CellID=48815, LogicRNCID=501', \
                        'Label=UH0821_U96, CellID=48216, LogicRNCID=501', \
                        'Label=UH0821_U95, CellID=48215, LogicRNCID=501', \
                        'Label=UH0821_U94, CellID=48214, LogicRNCID=501', \
                        'Label=UH2981_U4, CellID=29814, LogicRNCID=501', \
                        'Label=UH2981_U3, CellID=29813, LogicRNCID=501', \
                        'Label=UH2981_U2, CellID=29812, LogicRNCID=501', \
                        'Label=UH2981_U6, CellID=29816, LogicRNCID=501', \
                        'Label=UH2981_U5, CellID=29815, LogicRNCID=501', \
                        'Label=UH1947_U3, CellID=19473, LogicRNCID=501', \
                        'Label=UH1947_U2, CellID=19472, LogicRNCID=501', \
                        'Label=UH2981_U1, CellID=29811, LogicRNCID=501', \
                        'Label=UH1947_U6, CellID=19476, LogicRNCID=501', \
                        'Label=UH1947_U5, CellID=19475, LogicRNCID=501', \
                        'Label=UH1947_U4, CellID=19474, LogicRNCID=501', \
                        'Label=UH1947_U1, CellID=19471, LogicRNCID=501', \
                        'Label=UH3925_U1, CellID=39251, LogicRNCID=501', \
                        'Label=UH3925_U6, CellID=39256, LogicRNCID=501', \
                        'Label=UH3925_U5, CellID=39255, LogicRNCID=501', \
                        'Label=UH3925_U4, CellID=39254, LogicRNCID=501', \
                        'Label=UH3925_U3, CellID=39253, LogicRNCID=501', \
                        'Label=UH3925_U2, CellID=39252, LogicRNCID=501', \
                        'Label=UH0970_U3, CellID=9703, LogicRNCID=501', \
                        'Label=UH0970_U6, CellID=9706, LogicRNCID=501', \
                        'Label=UH0970_U5, CellID=9705, LogicRNCID=501', \
                        'Label=UH0970_U4, CellID=9704, LogicRNCID=501', \
                        'Label=UH0881_U3, CellID=8813, LogicRNCID=501', \
                        'Label=UH0970_U2, CellID=9702, LogicRNCID=501', \
                        'Label=UH0970_U1, CellID=9701, LogicRNCID=501', \
                        'Label=UH0881_U6, CellID=8816, LogicRNCID=501', \
                        'Label=UH0881_U5, CellID=8815, LogicRNCID=501', \
                        'Label=UH0821_U4, CellID=8214, LogicRNCID=501', \
                        'Label=UH0821_U2, CellID=8212, LogicRNCID=501', \
                        'Label=UH0821_U1, CellID=8211, LogicRNCID=501', \
                        'Label=UH0881_U4, CellID=8814, LogicRNCID=501', \
                        'Label=UH0821_U3, CellID=8213, LogicRNCID=501', \
                        'Label=UH0881_U2, CellID=8812, LogicRNCID=501', \
                        'Label=UH0881_U1, CellID=8811, LogicRNCID=501', \
                        'Label=UH0821_U6, CellID=8216, LogicRNCID=501', \
                        'Label=UH0821_U5, CellID=8215, LogicRNCID=501', \
                        'Label=UH0881_U97, CellID=48817, LogicRNCID=501', \
                        'Label=UH0821_U98, CellID=48218, LogicRNCID=501', \
                        'Label=UH0821_U97, CellID=48217, LogicRNCID=501', \
                        'Label=UH0821_U99, CellID=48219, LogicRNCID=501']  # кластер число сот 43
list_cluster_G = ['Label=UH0881_U94, CellID=48814, LogicRNCID=501', \
                'Label=UH0881_U96, CellID=48816, LogicRNCID=501', \
                'Label=UH0881_U95, CellID=48815, LogicRNCID=501', \
                'Label=UH0821_U96, CellID=48216, LogicRNCID=501', \
                'Label=UH0821_U95, CellID=48215, LogicRNCID=501', \
                'Label=UH0821_U94, CellID=48214, LogicRNCID=501', \
                'Label=UH2981_U4, CellID=29814, LogicRNCID=501', \
                'Label=UH2981_U3, CellID=29813, LogicRNCID=501', \
                'Label=UH2981_U2, CellID=29812, LogicRNCID=501', \
                'Label=UH2981_U6, CellID=29816, LogicRNCID=501', \
                'Label=UH2981_U5, CellID=29815, LogicRNCID=501', \
                'Label=UH1947_U3, CellID=19473, LogicRNCID=501', \
                'Label=UH1947_U2, CellID=19472, LogicRNCID=501', \
                'Label=UH2981_U1, CellID=29811, LogicRNCID=501', \
                'Label=UH1947_U6, CellID=19476, LogicRNCID=501', \
                'Label=UH1947_U5, CellID=19475, LogicRNCID=501', \
                'Label=UH1947_U4, CellID=19474, LogicRNCID=501', \
                'Label=UH1947_U1, CellID=19471, LogicRNCID=501', \
                'Label=UH3925_U1, CellID=39251, LogicRNCID=501', \
                'Label=UH3925_U6, CellID=39256, LogicRNCID=501', \
                'Label=UH3925_U5, CellID=39255, LogicRNCID=501', \
                'Label=UH3925_U4, CellID=39254, LogicRNCID=501', \
                'Label=UH3925_U3, CellID=39253, LogicRNCID=501', \
                'Label=UH3925_U2, CellID=39252, LogicRNCID=501', \
                'Label=UH0970_U3, CellID=9703, LogicRNCID=501', \
                'Label=UH0970_U6, CellID=9706, LogicRNCID=501', \
                'Label=UH0970_U5, CellID=9705, LogicRNCID=501', \
                'Label=UH0970_U4, CellID=9704, LogicRNCID=501', \
                'Label=UH0881_U3, CellID=8813, LogicRNCID=501', \
                'Label=UH0970_U2, CellID=9702, LogicRNCID=501', \
                'Label=UH0970_U1, CellID=9701, LogicRNCID=501', \
                'Label=UH0881_U6, CellID=8816, LogicRNCID=501', \
                'Label=UH0881_U5, CellID=8815, LogicRNCID=501', \
                'Label=UH0821_U4, CellID=8214, LogicRNCID=501', \
                'Label=UH0821_U2, CellID=8212, LogicRNCID=501', \
                'Label=UH0821_U1, CellID=8211, LogicRNCID=501', \
                'Label=UH0881_U4, CellID=8814, LogicRNCID=501', \
                'Label=UH0821_U3, CellID=8213, LogicRNCID=501', \
                'Label=UH0881_U2, CellID=8812, LogicRNCID=501', \
                'Label=UH0881_U1, CellID=8811, LogicRNCID=501', \
                'Label=UH0821_U6, CellID=8216, LogicRNCID=501', \
                'Label=UH0821_U5, CellID=8215, LogicRNCID=501', \
                'Label=UH0881_U97, CellID=48817, LogicRNCID=501', \
                'Label=UH0821_U98, CellID=48218, LogicRNCID=501', \
                'Label=UH0821_U97, CellID=48217, LogicRNCID=501', \
                'Label=UH0821_U99, CellID=48219, LogicRNCID=501']  # кластер число сот
list_U2100 = [
'Label=UH0742_U2, CellID=7422, LogicRNCID=501', \
'Label=UH1945_U3, CellID=19453, LogicRNCID=501', \
'Label=UH0742_U1, CellID=7421, LogicRNCID=501', \
'Label=UH1945_U2, CellID=19452, LogicRNCID=501', \
'Label=UH1945_U1, CellID=19451, LogicRNCID=501', \
'Label=UH0972_U3, CellID=9723, LogicRNCID=501', \
'Label=UH0844_U3, CellID=8443, LogicRNCID=501', \
'Label=UH0972_U2, CellID=9722, LogicRNCID=501', \
'Label=UH0844_U2, CellID=8442, LogicRNCID=501', \
'Label=UH0972_U1, CellID=9721, LogicRNCID=501', \
'Label=UH0844_U1, CellID=8441, LogicRNCID=501', \
'Label=UH2763_U3, CellID=27633, LogicRNCID=501', \
'Label=UH0880_U11, CellID=65520, LogicRNCID=501', \
'Label=UH2763_U2, CellID=27632, LogicRNCID=501', \
'Label=UH2763_U1, CellID=27631, LogicRNCID=501', \
'Label=UH1918_U3, CellID=19183, LogicRNCID=501', \
'Label=UH1918_U2, CellID=19182, LogicRNCID=501', \
'Label=UH1918_U1, CellID=19181, LogicRNCID=501', \
'Label=UH1892_U3, CellID=18923, LogicRNCID=501', \
'Label=UH1892_U2, CellID=18922, LogicRNCID=501', \
'Label=UH0970_U3, CellID=9703, LogicRNCID=501', \
'Label=UH0842_U3, CellID=8423, LogicRNCID=501', \
'Label=UH0714_U3, CellID=7143, LogicRNCID=501', \
'Label=UH0970_U2, CellID=9702, LogicRNCID=501', \
'Label=UH0842_U2, CellID=8422, LogicRNCID=501', \
'Label=UH0714_U2, CellID=7142, LogicRNCID=501', \
'Label=UH0970_U1, CellID=9701, LogicRNCID=501', \
'Label=UH0842_U1, CellID=8421, LogicRNCID=501', \
'Label=UH0765_U3, CellID=7653, LogicRNCID=501', \
'Label=UH0714_U1, CellID=7141, LogicRNCID=501', \
'Label=UH0765_U2, CellID=7652, LogicRNCID=501', \
'Label=UH0765_U1, CellID=7651, LogicRNCID=501', \
'Label=UH0995_U3, CellID=9953, LogicRNCID=501', \
'Label=UH0611_U3, CellID=6113, LogicRNCID=501', \
'Label=UH0995_U2, CellID=9952, LogicRNCID=501', \
'Label=UH3990_U3, CellID=39903, LogicRNCID=501', \
'Label=UH0995_U1, CellID=9951, LogicRNCID=501', \
'Label=UH0611_U1, CellID=6111, LogicRNCID=501', \
'Label=UH3990_U2, CellID=39902, LogicRNCID=501', \
'Label=UH3990_U1, CellID=39901, LogicRNCID=501', \
'Label=UH0841_U3, CellID=8413, LogicRNCID=501', \
'Label=UH0841_U2, CellID=8412, LogicRNCID=501', \
'Label=UH1916_U3, CellID=19163, LogicRNCID=501', \
'Label=UH0841_U1, CellID=8411, LogicRNCID=501', \
'Label=UH0636_U3, CellID=6363, LogicRNCID=501', \
'Label=UH0508_U3, CellID=5083, LogicRNCID=501', \
'Label=UH1916_U2, CellID=19162, LogicRNCID=501', \
'Label=UH0636_U2, CellID=6362, LogicRNCID=501', \
'Label=UH0508_U2, CellID=5082, LogicRNCID=501', \
'Label=UH1907_U66, CellID=65497, LogicRNCID=501', \
'Label=UH1916_U1, CellID=19161, LogicRNCID=501', \
'Label=UH1907_U55, CellID=65496, LogicRNCID=501', \
'Label=UH1907_U44, CellID=65495, LogicRNCID=501', \
'Label=UH0994_U3, CellID=9943, LogicRNCID=501', \
'Label=UH0738_U3, CellID=7383, LogicRNCID=501', \
'Label=UH0610_U3, CellID=6103, LogicRNCID=501', \
'Label=UH0994_U2, CellID=9942, LogicRNCID=501', \
'Label=UH0738_U2, CellID=7382, LogicRNCID=501', \
'Label=UH0610_U2, CellID=6102, LogicRNCID=501', \
'Label=UH3989_U3, CellID=39893, LogicRNCID=501', \
'Label=UH2709_U3, CellID=27093, LogicRNCID=501', \
'Label=UH0994_U1, CellID=9941, LogicRNCID=501', \
'Label=UH0738_U1, CellID=7381, LogicRNCID=501', \
'Label=UH3989_U2, CellID=39892, LogicRNCID=501', \
'Label=UH2709_U2, CellID=27092, LogicRNCID=501', \
'Label=UH3989_U1, CellID=39891, LogicRNCID=501', \
'Label=UH2709_U1, CellID=27091, LogicRNCID=501', \
'Label=UH0840_U3, CellID=8403, LogicRNCID=501', \
'Label=UH0840_U2, CellID=8402, LogicRNCID=501', \
'Label=UH0840_U1, CellID=8401, LogicRNCID=501', \
'Label=UH0763_U3, CellID=7633, LogicRNCID=501', \
'Label=UH0763_U2, CellID=7632, LogicRNCID=501', \
'Label=UH0763_U1, CellID=7631, LogicRNCID=501', \
'Label=UH2913_U3, CellID=29133, LogicRNCID=501', \
'Label=UH2913_U2, CellID=29132, LogicRNCID=501', \
'Label=UH2913_U1, CellID=29131, LogicRNCID=501', \
'Label=UH0839_U3, CellID=8393, LogicRNCID=501', \
'Label=UH0711_U3, CellID=7113, LogicRNCID=501', \
'Label=UH0839_U2, CellID=8392, LogicRNCID=501', \
'Label=UH0711_U2, CellID=7112, LogicRNCID=501', \
'Label=UH2938_U3, CellID=29383, LogicRNCID=501', \
'Label=UH0839_U1, CellID=8391, LogicRNCID=501', \
'Label=UH0711_U1, CellID=7111, LogicRNCID=501', \
'Label=UH2938_U2, CellID=29382, LogicRNCID=501', \
'Label=UH2938_U1, CellID=29381, LogicRNCID=501', \
'Label=UH1939_U3, CellID=19393, LogicRNCID=501', \
'Label=UH1939_U2, CellID=19392, LogicRNCID=501', \
'Label=UH1907_U3, CellID=19073, LogicRNCID=501', \
'Label=UH0960_U1, CellID=9601, LogicRNCID=501', \
'Label=UH0832_U1, CellID=8321, LogicRNCID=501', \
'Label=UH1907_U2, CellID=19072, LogicRNCID=501', \
'Label=UH1705_U2, CellID=17052, LogicRNCID=501', \
'Label=UH1833_U1, CellID=18331, LogicRNCID=501', \
'Label=UH1705_U1, CellID=17051, LogicRNCID=501', \
'Label=UH0988_U3, CellID=9883, LogicRNCID=501', \
'Label=UH2983_U3, CellID=29833, LogicRNCID=501', \
'Label=UH2932_U1, CellID=29321, LogicRNCID=501', \
'Label=UH1908_U1, CellID=19081, LogicRNCID=501', \
'Label=UH0963_U2, CellID=9632, LogicRNCID=501', \
'Label=UH0835_U2, CellID=8352, LogicRNCID=501', \
'Label=UH0966_U1, CellID=9661, LogicRNCID=501', \
'Label=UH0838_U1, CellID=8381, LogicRNCID=501', \
'Label=UH1935_U2, CellID=19352, LogicRNCID=501', \
'Label=UH1935_U1, CellID=19351, LogicRNCID=501', \
'Label=UH0962_U3, CellID=9623, LogicRNCID=501', \
'Label=UH1707_U1, CellID=17071, LogicRNCID=501', \
'Label=UH0734_U3, CellID=7343, LogicRNCID=501', \
'Label=UH0734_U2, CellID=7342, LogicRNCID=501', \
'Label=UH0734_U1, CellID=7341, LogicRNCID=501', \
'Label=UH0961_U1, CellID=9611, LogicRNCID=501', \
'Label=UH2932_U2, CellID=29322, LogicRNCID=501', \
'Label=UH1908_U2, CellID=19082, LogicRNCID=501', \
'Label=UH1960_U3, CellID=19603, LogicRNCID=501', \
'Label=UH1704_U3, CellID=17043, LogicRNCID=501', \
'Label=UH2984_U2, CellID=29842, LogicRNCID=501', \
'Label=UH1960_U2, CellID=19602, LogicRNCID=501', \
'Label=UH1939_U1, CellID=19391, LogicRNCID=501', \
'Label=UH0966_U3, CellID=9663, LogicRNCID=501', \
'Label=UH0838_U3, CellID=8383, LogicRNCID=501', \
'Label=UH0966_U2, CellID=9662, LogicRNCID=501', \
'Label=UH0838_U2, CellID=8382, LogicRNCID=501', \
'Label=UH0961_U3, CellID=9613, LogicRNCID=501', \
'Label=UH0602_U1, CellID=6021, LogicRNCID=501', \
'Label=UH1805_U2, CellID=18052, LogicRNCID=501', \
'Label=UH1704_U2, CellID=17042, LogicRNCID=501', \
'Label=UH2984_U1, CellID=29841, LogicRNCID=501', \
'Label=UH1960_U1, CellID=19601, LogicRNCID=501', \
'Label=UH1704_U1, CellID=17041, LogicRNCID=501', \
'Label=UH4600_U1, CellID=46001, LogicRNCID=501', \
'Label=UH1707_U3, CellID=17073, LogicRNCID=501', \
'Label=UH1707_U2, CellID=17072, LogicRNCID=501', \
'Label=UH1936_U2, CellID=19362, LogicRNCID=501', \
'Label=UH1936_U1, CellID=19361, LogicRNCID=501', \
'Label=UH0963_U3, CellID=9633, LogicRNCID=501', \
'Label=UH0835_U3, CellID=8353, LogicRNCID=501', \
'Label=UH1833_U3, CellID=18333, LogicRNCID=501', \
'Label=UH1705_U3, CellID=17053, LogicRNCID=501', \
'Label=UH0758_U1, CellID=7581, LogicRNCID=501', \
'Label=UH1833_U2, CellID=18332, LogicRNCID=501', \
'Label=UH1984_U2, CellID=19842, LogicRNCID=501', \
'Label=UH0770_U2, CellID=7702, LogicRNCID=501', \
'Label=UH0770_U1, CellID=7701, LogicRNCID=501', \
'Label=UH0821_U3, CellID=8213, LogicRNCID=501', \
'Label=UH0950_U1, CellID=9501, LogicRNCID=501', \
'Label=UH0617_U3, CellID=6173, LogicRNCID=501', \
'Label=UH1976_U2, CellID=19762, LogicRNCID=501', \
'Label=UH0952_U1, CellID=9521, LogicRNCID=501', \
'Label=UH3995_U3, CellID=39953, LogicRNCID=501', \
'Label=UH0618_U2, CellID=6182, LogicRNCID=501', \
'Label=UH0618_U1, CellID=6181, LogicRNCID=501', \
'Label=UH0820_U2, CellID=8202, LogicRNCID=501', \
'Label=UH1976_U3, CellID=19763, LogicRNCID=501', \
'Label=UH0845_U3, CellID=8453, LogicRNCID=501', \
'Label=UH0845_U2, CellID=8452, LogicRNCID=501', \
'Label=UH0973_U1, CellID=9731, LogicRNCID=501', \
'Label=UH0845_U1, CellID=8451, LogicRNCID=501', \
'Label=UH0717_U2, CellID=7172, LogicRNCID=501', \
'Label=UH0973_U2, CellID=9732, LogicRNCID=501', \
'Label=UH0717_U3, CellID=7173, LogicRNCID=501', \
'Label=UH0820_U1, CellID=8201, LogicRNCID=501', \
'Label=UH0743_U2, CellID=7432, LogicRNCID=501', \
'Label=UH0823_U2, CellID=8232, LogicRNCID=501', \
'Label=UH0951_U1, CellID=9511, LogicRNCID=501', \
'Label=UH0618_U3, CellID=6183, LogicRNCID=501', \
'Label=UH0822_U2, CellID=8222, LogicRNCID=501', \
'Label=UH0770_U3, CellID=7703, LogicRNCID=501', \
'Label=UH0743_U1, CellID=7431, LogicRNCID=501', \
'Label=UH0973_U3, CellID=9733, LogicRNCID=501', \
'Label=UH3995_U2, CellID=39952, LogicRNCID=501', \
'Label=UH1947_U2, CellID=19472, LogicRNCID=501', \
'Label=UH3995_U1, CellID=39951, LogicRNCID=501', \
'Label=UH0823_U3, CellID=8233, LogicRNCID=501', \
'Label=UH0821_U2, CellID=8212, LogicRNCID=501', \
'Label=UH0821_U1, CellID=8211, LogicRNCID=501', \
'Label=UH0950_U3, CellID=9503, LogicRNCID=501', \
'Label=UH0822_U3, CellID=8223, LogicRNCID=501', \
'Label=UH0951_U3, CellID=9513, LogicRNCID=501', \
'Label=UH0849_U3, CellID=8493, LogicRNCID=501', \
'Label=UH3921_U2, CellID=39212, LogicRNCID=501', \
'Label=UH0849_U2, CellID=8492, LogicRNCID=501', \
'Label=UH3921_U1, CellID=39211, LogicRNCID=501', \
'Label=UH3996_U3, CellID=39963, LogicRNCID=501', \
'Label=UH3996_U1, CellID=39961, LogicRNCID=501', \
'Label=UH3996_U2, CellID=39962, LogicRNCID=501', \
'Label=UH0617_U1, CellID=6171, LogicRNCID=501', \
'Label=UH0717_U1, CellID=7171, LogicRNCID=501', \
'Label=UH1947_U1, CellID=19471, LogicRNCID=501', \
'Label=UH0974_U1, CellID=9741, LogicRNCID=501', \
'Label=UH3920_U1, CellID=39201, LogicRNCID=501', \
'Label=UH0747_U2, CellID=7472, LogicRNCID=501', \
'Label=UH3921_U3, CellID=39213, LogicRNCID=501', \
'Label=UH3920_U2, CellID=39202, LogicRNCID=501', \
'Label=UH3918_U1, CellID=39181, LogicRNCID=501', \
'Label=UH0974_U2, CellID=9742, LogicRNCID=501', \
'Label=UH0974_U3, CellID=9743, LogicRNCID=501', \
'Label=UH0849_U1, CellID=8491, LogicRNCID=501', \
'Label=UH3918_U3, CellID=39183, LogicRNCID=501', \
'Label=UH0950_U2, CellID=9502, LogicRNCID=501', \
'Label=UH0823_U1, CellID=8231, LogicRNCID=501', \
'Label=UH0743_U3, CellID=7433, LogicRNCID=501', \
'Label=UH0952_U3, CellID=9523, LogicRNCID=501', \
'Label=UH1947_U3, CellID=19473, LogicRNCID=501', \
'Label=UH1976_U1, CellID=19761, LogicRNCID=501', \
'Label=UH0952_U2, CellID=9522, LogicRNCID=501', \
'Label=UH0617_U2, CellID=6172, LogicRNCID=501', \
'Label=UH0822_U1, CellID=8221, LogicRNCID=501', \
'Label=UH0960_U2, CellID=9602, LogicRNCID=501', \
'Label=UH0832_U2, CellID=8322, LogicRNCID=501', \
'Label=UH1984_U1, CellID=19841, LogicRNCID=501', \
'Label=UH0962_U2, CellID=9622, LogicRNCID=501', \
'Label=UH0962_U1, CellID=9621, LogicRNCID=501', \
'Label=UH0965_U1, CellID=9651, LogicRNCID=501', \
'Label=UH4600_U2, CellID=46002, LogicRNCID=501', \
'Label=UH0988_U2, CellID=9882, LogicRNCID=501', \
'Label=UH1935_U3, CellID=19353, LogicRNCID=501', \
'Label=UH0988_U1, CellID=9881, LogicRNCID=501', \
'Label=UH1703_U3, CellID=17033, LogicRNCID=501', \
'Label=UH0807_U3, CellID=8073, LogicRNCID=501', \
'Label=UH2983_U2, CellID=29832, LogicRNCID=501', \
'Label=UH1703_U2, CellID=17032, LogicRNCID=501', \
'Label=UH1936_U3, CellID=19363, LogicRNCID=501', \
'Label=UH1805_U1, CellID=18051, LogicRNCID=501', \
'Label=UH0960_U3, CellID=9603, LogicRNCID=501', \
'Label=UH0832_U3, CellID=8323, LogicRNCID=501', \
'Label=UH1882_U2, CellID=18822, LogicRNCID=501', \
'Label=UH0602_U2, CellID=6022, LogicRNCID=501', \
'Label=UH1882_U1, CellID=18821, LogicRNCID=501', \
'Label=UH1805_U3, CellID=18053, LogicRNCID=501', \
'Label=UH0965_U3, CellID=9653, LogicRNCID=501', \
'Label=UH0965_U2, CellID=9652, LogicRNCID=501', \
'Label=UH4600_U3, CellID=46003, LogicRNCID=501', \
'Label=UH0963_U1, CellID=9631, LogicRNCID=501', \
'Label=UH0835_U1, CellID=8351, LogicRNCID=501', \
'Label=UH0758_U3, CellID=7583, LogicRNCID=501', \
'Label=UH0758_U2, CellID=7582, LogicRNCID=501', \
'Label=UH0807_U2, CellID=8072, LogicRNCID=501', \
'Label=UH2983_U1, CellID=29831, LogicRNCID=501', \
'Label=UH1882_U3, CellID=18823, LogicRNCID=501', \
'Label=UH1703_U1, CellID=17031, LogicRNCID=501', \
'Label=UH0807_U1, CellID=8071, LogicRNCID=501', \
'Label=UH0602_U3, CellID=6023, LogicRNCID=501', \
'Label=UH0961_U2, CellID=9612, LogicRNCID=501', \
'Label=UH2932_U3, CellID=29323, LogicRNCID=501', \
'Label=UH1908_U3, CellID=19083, LogicRNCID=501', \
'Label=UH0706_U1, CellID=7061, LogicRNCID=501', \
'Label=UH2984_U3, CellID=29843, LogicRNCID=501', \
'Label=UH0708_U3, CellID=7083, LogicRNCID=501', \
'Label=UH0708_U2, CellID=7082, LogicRNCID=501', \
'Label=UH0708_U1, CellID=7081, LogicRNCID=501', \
'Label=UH1964_U2, CellID=19642, LogicRNCID=501', \
'Label=UH1964_U1, CellID=19641, LogicRNCID=501', \
'Label=UH3922_U1, CellID=39221, LogicRNCID=501', \
'Label=UH0880_U3, CellID=8803, LogicRNCID=501', \
'Label=UH0880_U1, CellID=8801, LogicRNCID=501', \
'Label=UH0980_U3, CellID=9803, LogicRNCID=501', \
'Label=UH0980_U2, CellID=9802, LogicRNCID=501', \
'Label=UH0980_U1, CellID=9801, LogicRNCID=501', \
'Label=UH1905_U2, CellID=19052, LogicRNCID=501', \
'Label=UH0881_U2, CellID=8812, LogicRNCID=501', \
'Label=UH1905_U1, CellID=19051, LogicRNCID=501', \
'Label=UH0881_U1, CellID=8811, LogicRNCID=501', \
'Label=UH1702_U2, CellID=17022, LogicRNCID=501', \
'Label=UH1702_U1, CellID=17021, LogicRNCID=501', \
'Label=UH0601_U3, CellID=6013, LogicRNCID=501', \
'Label=UH0601_U2, CellID=6012, LogicRNCID=501', \
'Label=UH0601_U1, CellID=6011, LogicRNCID=501', \
'Label=UH1801_U3, CellID=18013, LogicRNCID=501', \
'Label=UH0854_U1, CellID=8541, LogicRNCID=501', \
'Label=UH3928_U2, CellID=39282, LogicRNCID=501', \
'Label=UH0600_U2, CellID=6002, LogicRNCID=501', \
'Label=UH3928_U1, CellID=39281, LogicRNCID=501', \
'Label=UH0622_U2, CellID=6222, LogicRNCID=501', \
'Label=UH0622_U1, CellID=6221, LogicRNCID=501', \
'Label=UH0879_U2, CellID=8792, LogicRNCID=501', \
'Label=UH0623_U2, CellID=6232, LogicRNCID=501', \
'Label=UH0623_U1, CellID=6231, LogicRNCID=501', \
'Label=UH1958_U3, CellID=19583, LogicRNCID=501', \
'Label=UH1907_U1, CellID=19071, LogicRNCID=501', \
'Label=UH1702_U3, CellID=17023, LogicRNCID=501', \
'Label=UH1958_U2, CellID=19582, LogicRNCID=501', \
'Label=UH1928_U2, CellID=19282, LogicRNCID=501', \
'Label=UH1928_U1, CellID=19281, LogicRNCID=501', \
'Label=UH0955_U3, CellID=9553, LogicRNCID=501', \
'Label=UH2950_U1, CellID=29501, LogicRNCID=501', \
'Label=UH0825_U3, CellID=8253, LogicRNCID=501', \
'Label=UH0825_U2, CellID=8252, LogicRNCID=501', \
'Label=UH3925_U3, CellID=39253, LogicRNCID=501', \
'Label=UH0981_U3, CellID=9813, LogicRNCID=501', \
'Label=UH3928_U3, CellID=39283, LogicRNCID=501', \
'Label=UH2981_U1, CellID=29811, LogicRNCID=501', \
'Label=UH1854_U2, CellID=18542, LogicRNCID=501', \
'Label=UH0830_U2, CellID=8302, LogicRNCID=501', \
'Label=UH1854_U1, CellID=18541, LogicRNCID=501', \
'Label=UH0881_U3, CellID=8813, LogicRNCID=501', \
'Label=UH0830_U1, CellID=8301, LogicRNCID=501', \
'Label=UH3925_U2, CellID=39252, LogicRNCID=501', \
'Label=UH0981_U2, CellID=9812, LogicRNCID=501', \
'Label=UH3925_U1, CellID=39251, LogicRNCID=501', \
'Label=UH1928_U3, CellID=19283, LogicRNCID=501', \
'Label=UH0981_U1, CellID=9811, LogicRNCID=501', \
'Label=UH1951_U2, CellID=19512, LogicRNCID=501', \
'Label=UH1951_U1, CellID=19511, LogicRNCID=501', \
'Label=UH3922_U2, CellID=39222, LogicRNCID=501', \
'Label=UH1801_U1, CellID=18011, LogicRNCID=501', \
'Label=UH2981_U3, CellID=29813, LogicRNCID=501', \
'Label=UH1906_U1, CellID=19061, LogicRNCID=501', \
'Label=UH2981_U2, CellID=29812, LogicRNCID=501', \
'Label=UH0854_U3, CellID=8543, LogicRNCID=501', \
'Label=UH0854_U2, CellID=8542, LogicRNCID=501', \
'Label=UH1854_U3, CellID=18543, LogicRNCID=501', \
'Label=UH1803_U1, CellID=18031, LogicRNCID=501', \
'Label=UH0830_U3, CellID=8303, LogicRNCID=501', \
'Label=UH0825_U1, CellID=8251, LogicRNCID=501', \
'Label=UH1951_U3, CellID=19513, LogicRNCID=501', \
'Label=UH2950_U3, CellID=29503, LogicRNCID=501', \
'Label=UH2950_U2, CellID=29502, LogicRNCID=501', \
'Label=UH1906_U3, CellID=19063, LogicRNCID=501', \
'Label=UH0831_U1, CellID=8311, LogicRNCID=501', \
'Label=UH1906_U2, CellID=19062, LogicRNCID=501', \
'Label=UH0955_U2, CellID=9552, LogicRNCID=501', \
'Label=UH0955_U1, CellID=9551, LogicRNCID=501', \
'Label=UH0622_U3, CellID=6223, LogicRNCID=501', \
'Label=UH0879_U3, CellID=8793, LogicRNCID=501', \
'Label=UH0623_U3, CellID=6233, LogicRNCID=501', \
'Label=UH1803_U3, CellID=18033, LogicRNCID=501', \
'Label=UH0600_U1, CellID=6001, LogicRNCID=501', \
'Label=UH1803_U2, CellID=18032, LogicRNCID=501', \
'Label=UH0831_U3, CellID=8313, LogicRNCID=501', \
'Label=UH0831_U2, CellID=8312, LogicRNCID=501', \
'Label=UH0748_U3, CellID=7483, LogicRNCID=501', \
'Label=UH0748_U2, CellID=7482, LogicRNCID=501', \
'Label=UH0748_U1, CellID=7481, LogicRNCID=501', \
'Label=UH1945_U5, CellID=19455, LogicRNCID=501', \
'Label=UH1945_U4, CellID=19454, LogicRNCID=501', \
'Label=UH0972_U6, CellID=9726, LogicRNCID=501', \
'Label=UH0844_U6, CellID=8446, LogicRNCID=501', \
'Label=UH0972_U5, CellID=9725, LogicRNCID=501', \
'Label=UH0844_U5, CellID=8445, LogicRNCID=501', \
'Label=UH0972_U4, CellID=9724, LogicRNCID=501', \
'Label=UH0844_U4, CellID=8444, LogicRNCID=501', \
'Label=UH2763_U6, CellID=27636, LogicRNCID=501', \
'Label=UH2763_U5, CellID=27635, LogicRNCID=501', \
'Label=UH2763_U4, CellID=27634, LogicRNCID=501', \
'Label=UH1918_U6, CellID=19186, LogicRNCID=501', \
'Label=UH0880_U14, CellID=65521, LogicRNCID=501', \
'Label=UH1918_U5, CellID=19185, LogicRNCID=501', \
'Label=UH1918_U4, CellID=19184, LogicRNCID=501', \
'Label=UH1892_U6, CellID=18926, LogicRNCID=501', \
'Label=UH1892_U5, CellID=18925, LogicRNCID=501', \
'Label=UH0970_U6, CellID=9706, LogicRNCID=501', \
'Label=UH0842_U6, CellID=8426, LogicRNCID=501', \
'Label=UH0714_U6, CellID=7146, LogicRNCID=501', \
'Label=UH0970_U5, CellID=9705, LogicRNCID=501', \
'Label=UH0842_U5, CellID=8425, LogicRNCID=501', \
'Label=UH0714_U5, CellID=7145, LogicRNCID=501', \
'Label=UH0970_U4, CellID=9704, LogicRNCID=501', \
'Label=UH0842_U4, CellID=8424, LogicRNCID=501', \
'Label=UH0765_U6, CellID=7656, LogicRNCID=501', \
'Label=UH0714_U4, CellID=7144, LogicRNCID=501', \
'Label=UH0765_U5, CellID=7655, LogicRNCID=501', \
'Label=UH0765_U4, CellID=7654, LogicRNCID=501', \
'Label=UH0995_U6, CellID=9956, LogicRNCID=501', \
'Label=UH0611_U6, CellID=6116, LogicRNCID=501', \
'Label=UH0995_U5, CellID=9955, LogicRNCID=501', \
'Label=UH3990_U6, CellID=39906, LogicRNCID=501', \
'Label=UH0995_U4, CellID=9954, LogicRNCID=501', \
'Label=UH0611_U4, CellID=6114, LogicRNCID=501', \
'Label=UH3990_U5, CellID=39905, LogicRNCID=501', \
'Label=UH3990_U4, CellID=39904, LogicRNCID=501', \
'Label=UH0841_U6, CellID=8416, LogicRNCID=501', \
'Label=UH0841_U5, CellID=8415, LogicRNCID=501', \
'Label=UH1916_U6, CellID=19166, LogicRNCID=501', \
'Label=UH0841_U4, CellID=8414, LogicRNCID=501', \
'Label=UH0636_U6, CellID=6366, LogicRNCID=501', \
'Label=UH0508_U6, CellID=5086, LogicRNCID=501', \
'Label=UH1916_U5, CellID=19165, LogicRNCID=501', \
'Label=UH0636_U5, CellID=6365, LogicRNCID=501', \
'Label=UH0508_U5, CellID=5085, LogicRNCID=501', \
'Label=UH1916_U4, CellID=19164, LogicRNCID=501', \
'Label=UH0994_U6, CellID=9946, LogicRNCID=501', \
'Label=UH0738_U6, CellID=7386, LogicRNCID=501', \
'Label=UH0610_U6, CellID=6106, LogicRNCID=501', \
'Label=UH0994_U5, CellID=9945, LogicRNCID=501', \
'Label=UH0738_U5, CellID=7385, LogicRNCID=501', \
'Label=UH0610_U5, CellID=6105, LogicRNCID=501', \
'Label=UH3989_U6, CellID=39896, LogicRNCID=501', \
'Label=UH2709_U6, CellID=27096, LogicRNCID=501', \
'Label=UH0994_U4, CellID=9944, LogicRNCID=501', \
'Label=UH0738_U4, CellID=7384, LogicRNCID=501', \
'Label=UH3989_U5, CellID=39895, LogicRNCID=501', \
'Label=UH2709_U5, CellID=27095, LogicRNCID=501', \
'Label=UH1907_U33, CellID=65494, LogicRNCID=501', \
'Label=UH3989_U4, CellID=39894, LogicRNCID=501', \
'Label=UH2709_U4, CellID=27094, LogicRNCID=501', \
'Label=UH0840_U6, CellID=8406, LogicRNCID=501', \
'Label=UH1907_U22, CellID=65493, LogicRNCID=501', \
'Label=UH0840_U5, CellID=8405, LogicRNCID=501', \
'Label=UH1907_U11, CellID=65492, LogicRNCID=501', \
'Label=UH0840_U4, CellID=8404, LogicRNCID=501', \
'Label=UH0763_U6, CellID=7636, LogicRNCID=501', \
'Label=UH0763_U5, CellID=7635, LogicRNCID=501', \
'Label=UH0763_U4, CellID=7634, LogicRNCID=501', \
'Label=UH2913_U6, CellID=29136, LogicRNCID=501', \
'Label=UH2913_U5, CellID=29135, LogicRNCID=501', \
'Label=UH2913_U4, CellID=29134, LogicRNCID=501', \
'Label=UH0839_U6, CellID=8396, LogicRNCID=501', \
'Label=UH0711_U6, CellID=7116, LogicRNCID=501', \
'Label=UH0839_U5, CellID=8395, LogicRNCID=501', \
'Label=UH0711_U5, CellID=7115, LogicRNCID=501', \
'Label=UH2938_U6, CellID=29386, LogicRNCID=501', \
'Label=UH0839_U4, CellID=8394, LogicRNCID=501', \
'Label=UH0711_U4, CellID=7114, LogicRNCID=501', \
'Label=UH2938_U5, CellID=29385, LogicRNCID=501', \
'Label=UH2938_U4, CellID=29384, LogicRNCID=501', \
'Label=UH1939_U6, CellID=19396, LogicRNCID=501', \
'Label=UH1939_U5, CellID=19395, LogicRNCID=501', \
'Label=UH1939_U4, CellID=19394, LogicRNCID=501', \
'Label=UH0966_U6, CellID=9666, LogicRNCID=501', \
'Label=UH0838_U6, CellID=8386, LogicRNCID=501', \
'Label=UH0966_U5, CellID=9665, LogicRNCID=501', \
'Label=UH0838_U5, CellID=8385, LogicRNCID=501', \
'Label=UH0966_U4, CellID=9664, LogicRNCID=501', \
'Label=UH0838_U4, CellID=8384, LogicRNCID=501', \
'Label=UH1958_U5, CellID=19585, LogicRNCID=501', \
'Label=UH1702_U5, CellID=17025, LogicRNCID=501', \
'Label=UH1702_U4, CellID=17024, LogicRNCID=501', \
'Label=UH0988_U4, CellID=9884, LogicRNCID=501', \
'Label=UH1935_U5, CellID=19355, LogicRNCID=501', \
'Label=UH1703_U4, CellID=17034, LogicRNCID=501', \
'Label=UH0807_U4, CellID=8074, LogicRNCID=501', \
'Label=UH0602_U6, CellID=6026, LogicRNCID=501', \
'Label=UH1882_U5, CellID=18825, LogicRNCID=501', \
'Label=UH1833_U6, CellID=18336, LogicRNCID=501', \
'Label=UH1705_U6, CellID=17056, LogicRNCID=501', \
'Label=UH0758_U4, CellID=7584, LogicRNCID=501', \
'Label=UH1833_U5, CellID=18335, LogicRNCID=501', \
'Label=UH1705_U5, CellID=17055, LogicRNCID=501', \
'Label=UH1964_U4, CellID=19644, LogicRNCID=501', \
'Label=UH0962_U5, CellID=9625, LogicRNCID=501', \
'Label=UH0962_U4, CellID=9624, LogicRNCID=501', \
'Label=UH0706_U4, CellID=7064, LogicRNCID=501', \
'Label=UH0734_U4, CellID=7344, LogicRNCID=501', \
'Label=UH0708_U6, CellID=7086, LogicRNCID=501', \
'Label=UH1703_U5, CellID=17035, LogicRNCID=501', \
'Label=UH0807_U5, CellID=8075, LogicRNCID=501', \
'Label=UH2983_U4, CellID=29834, LogicRNCID=501', \
'Label=UH1882_U6, CellID=18826, LogicRNCID=501', \
'Label=UH1964_U5, CellID=19645, LogicRNCID=501', \
'Label=UH0961_U4, CellID=9614, LogicRNCID=501', \
'Label=UH2932_U5, CellID=29325, LogicRNCID=501', \
'Label=UH1908_U5, CellID=19085, LogicRNCID=501', \
'Label=UH2983_U6, CellID=29836, LogicRNCID=501', \
'Label=UH2932_U4, CellID=29324, LogicRNCID=501', \
'Label=UH0960_U5, CellID=9605, LogicRNCID=501', \
'Label=UH0832_U5, CellID=8325, LogicRNCID=501', \
'Label=UH1984_U4, CellID=19844, LogicRNCID=501', \
'Label=UH1907_U6, CellID=19076, LogicRNCID=501', \
'Label=UH0960_U4, CellID=9604, LogicRNCID=501', \
'Label=UH0734_U5, CellID=7345, LogicRNCID=501', \
'Label=UH0963_U4, CellID=9634, LogicRNCID=501', \
'Label=UH0835_U4, CellID=8354, LogicRNCID=501', \
'Label=UH0758_U6, CellID=7586, LogicRNCID=501', \
'Label=UH0758_U5, CellID=7585, LogicRNCID=501', \
'Label=UH0988_U5, CellID=9885, LogicRNCID=501', \
'Label=UH1935_U6, CellID=19356, LogicRNCID=501', \
'Label=UH0961_U6, CellID=9616, LogicRNCID=501', \
'Label=UH0961_U5, CellID=9615, LogicRNCID=501', \
'Label=UH2932_U6, CellID=29326, LogicRNCID=501', \
'Label=UH1908_U6, CellID=19086, LogicRNCID=501', \
'Label=UH1958_U6, CellID=19586, LogicRNCID=501', \
'Label=UH1907_U4, CellID=19074, LogicRNCID=501', \
'Label=UH1702_U6, CellID=17026, LogicRNCID=501', \
'Label=UH3996_U4, CellID=39964, LogicRNCID=501', \
'Label=UH3921_U6, CellID=39216, LogicRNCID=501', \
'Label=UH3995_U4, CellID=39954, LogicRNCID=501', \
'Label=UH1947_U4, CellID=19474, LogicRNCID=501', \
'Label=UH0974_U5, CellID=9745, LogicRNCID=501', \
'Label=UH0742_U4, CellID=7424, LogicRNCID=501', \
'Label=UH3920_U4, CellID=39204, LogicRNCID=501', \
'Label=UH0743_U6, CellID=7436, LogicRNCID=501', \
'Label=UH0743_U5, CellID=7435, LogicRNCID=501', \
'Label=UH0747_U5, CellID=7475, LogicRNCID=501', \
'Label=UH0973_U6, CellID=9736, LogicRNCID=501', \
'Label=UH0973_U5, CellID=9735, LogicRNCID=501', \
'Label=UH0617_U5, CellID=6175, LogicRNCID=501', \
'Label=UH3996_U6, CellID=39966, LogicRNCID=501', \
'Label=UH0770_U4, CellID=7704, LogicRNCID=501', \
'Label=UH0845_U4, CellID=8454, LogicRNCID=501', \
'Label=UH0717_U4, CellID=7174, LogicRNCID=501', \
'Label=UH0973_U4, CellID=9734, LogicRNCID=501', \
'Label=UH0717_U5, CellID=7175, LogicRNCID=501', \
'Label=UH0823_U5, CellID=8235, LogicRNCID=501', \
'Label=UH0823_U4, CellID=8234, LogicRNCID=501', \
'Label=UH3995_U6, CellID=39956, LogicRNCID=501', \
'Label=UH3995_U5, CellID=39955, LogicRNCID=501', \
'Label=UH1947_U5, CellID=19475, LogicRNCID=501', \
'Label=UH1947_U6, CellID=19476, LogicRNCID=501', \
'Label=UH0822_U4, CellID=8224, LogicRNCID=501', \
'Label=UH0617_U6, CellID=6176, LogicRNCID=501', \
'Label=UH0950_U4, CellID=9504, LogicRNCID=501', \
'Label=UH0618_U5, CellID=6185, LogicRNCID=501', \
'Label=UH0618_U6, CellID=6186, LogicRNCID=501', \
'Label=UH0951_U4, CellID=9514, LogicRNCID=501', \
'Label=UH0770_U6, CellID=7706, LogicRNCID=501', \
'Label=UH0770_U5, CellID=7705, LogicRNCID=501', \
'Label=UH0742_U5, CellID=7425, LogicRNCID=501', \
'Label=UH0950_U6, CellID=9506, LogicRNCID=501', \
'Label=UH0822_U5, CellID=8225, LogicRNCID=501', \
'Label=UH0849_U5, CellID=8495, LogicRNCID=501', \
'Label=UH3921_U4, CellID=39214, LogicRNCID=501', \
'Label=UH0849_U4, CellID=8494, LogicRNCID=501', \
'Label=UH3921_U5, CellID=39215, LogicRNCID=501', \
'Label=UH0950_U5, CellID=9505, LogicRNCID=501', \
'Label=UH0822_U6, CellID=8226, LogicRNCID=501', \
'Label=UH0820_U5, CellID=8205, LogicRNCID=501', \
'Label=UH1945_U6, CellID=19456, LogicRNCID=501', \
'Label=UH0823_U6, CellID=8236, LogicRNCID=501', \
'Label=UH0951_U6, CellID=9516, LogicRNCID=501', \
'Label=UH0974_U4, CellID=9744, LogicRNCID=501', \
'Label=UH3918_U4, CellID=39184, LogicRNCID=501', \
'Label=UH0821_U4, CellID=8214, LogicRNCID=501', \
'Label=UH0821_U5, CellID=8215, LogicRNCID=501', \
'Label=UH0821_U6, CellID=8216, LogicRNCID=501', \
'Label=UH0617_U4, CellID=6174, LogicRNCID=501', \
'Label=UH0618_U4, CellID=6184, LogicRNCID=501', \
'Label=UH0845_U5, CellID=8455, LogicRNCID=501', \
'Label=UH0717_U6, CellID=7176, LogicRNCID=501', \
'Label=UH0845_U6, CellID=8456, LogicRNCID=501', \
'Label=UH0743_U4, CellID=7434, LogicRNCID=501', \
'Label=UH0820_U4, CellID=8204, LogicRNCID=501', \
'Label=UH3920_U5, CellID=39205, LogicRNCID=501', \
'Label=UH0974_U6, CellID=9746, LogicRNCID=501', \
'Label=UH3918_U6, CellID=39186, LogicRNCID=501', \
'Label=UH0849_U6, CellID=8496, LogicRNCID=501', \
'Label=UH3996_U5, CellID=39965, LogicRNCID=501', \
'Label=UH2984_U6, CellID=29846, LogicRNCID=501', \
'Label=UH1960_U6, CellID=19606, LogicRNCID=501', \
'Label=UH1704_U6, CellID=17046, LogicRNCID=501', \
'Label=UH2984_U5, CellID=29845, LogicRNCID=501', \
'Label=UH1960_U5, CellID=19605, LogicRNCID=501', \
'Label=UH1704_U5, CellID=17045, LogicRNCID=501', \
'Label=UH1707_U4, CellID=17074, LogicRNCID=501', \
'Label=UH0734_U6, CellID=7346, LogicRNCID=501', \
'Label=UH1935_U4, CellID=19354, LogicRNCID=501', \
'Label=UH0962_U6, CellID=9626, LogicRNCID=501', \
'Label=UH0602_U5, CellID=6025, LogicRNCID=501', \
'Label=UH1882_U4, CellID=18824, LogicRNCID=501', \
'Label=UH1805_U6, CellID=18056, LogicRNCID=501', \
'Label=UH1936_U5, CellID=19365, LogicRNCID=501', \
'Label=UH1936_U4, CellID=19364, LogicRNCID=501', \
'Label=UH0963_U6, CellID=9636, LogicRNCID=501', \
'Label=UH0835_U6, CellID=8356, LogicRNCID=501', \
'Label=UH0963_U5, CellID=9635, LogicRNCID=501', \
'Label=UH0835_U5, CellID=8355, LogicRNCID=501', \
'Label=UH0832_U4, CellID=8324, LogicRNCID=501', \
'Label=UH1907_U5, CellID=19075, LogicRNCID=501', \
'Label=UH1805_U4, CellID=18054, LogicRNCID=501', \
'Label=UH0960_U6, CellID=9606, LogicRNCID=501', \
'Label=UH0832_U6, CellID=8326, LogicRNCID=501', \
'Label=UH1984_U5, CellID=19845, LogicRNCID=501', \
'Label=UH0965_U4, CellID=9654, LogicRNCID=501', \
'Label=UH4600_U5, CellID=46005, LogicRNCID=501', \
'Label=UH4600_U4, CellID=46004, LogicRNCID=501', \
'Label=UH1707_U6, CellID=17076, LogicRNCID=501', \
'Label=UH1707_U5, CellID=17075, LogicRNCID=501', \
'Label=UH1833_U4, CellID=18334, LogicRNCID=501', \
'Label=UH1705_U4, CellID=17054, LogicRNCID=501', \
'Label=UH0988_U6, CellID=9886, LogicRNCID=501', \
'Label=UH0602_U4, CellID=6024, LogicRNCID=501', \
'Label=UH1805_U5, CellID=18055, LogicRNCID=501', \
'Label=UH1908_U4, CellID=19084, LogicRNCID=501', \
'Label=UH1703_U6, CellID=17036, LogicRNCID=501', \
'Label=UH0807_U6, CellID=8076, LogicRNCID=501', \
'Label=UH2983_U5, CellID=29835, LogicRNCID=501', \
'Label=UH2984_U4, CellID=29844, LogicRNCID=501', \
'Label=UH1960_U4, CellID=19604, LogicRNCID=501', \
'Label=UH1704_U4, CellID=17044, LogicRNCID=501', \
'Label=UH0708_U5, CellID=7085, LogicRNCID=501', \
'Label=UH0708_U4, CellID=7084, LogicRNCID=501', \
'Label=UH1936_U6, CellID=19366, LogicRNCID=501', \
'Label=UH0965_U6, CellID=9656, LogicRNCID=501', \
'Label=UH0965_U5, CellID=9655, LogicRNCID=501', \
'Label=UH4600_U6, CellID=46006, LogicRNCID=501', \
'Label=UH0952_U6, CellID=9526, LogicRNCID=501', \
'Label=UH1976_U5, CellID=19765, LogicRNCID=501', \
'Label=UH0952_U5, CellID=9525, LogicRNCID=501', \
'Label=UH1976_U4, CellID=19764, LogicRNCID=501', \
'Label=UH0952_U4, CellID=9524, LogicRNCID=501', \
'Label=UH0854_U6, CellID=8546, LogicRNCID=501', \
'Label=UH0854_U5, CellID=8545, LogicRNCID=501', \
'Label=UH1801_U6, CellID=18016, LogicRNCID=501', \
'Label=UH0854_U4, CellID=8544, LogicRNCID=501', \
'Label=UH0601_U4, CellID=6014, LogicRNCID=501', \
'Label=UH0831_U6, CellID=8316, LogicRNCID=501', \
'Label=UH0831_U5, CellID=8315, LogicRNCID=501', \
'Label=UH0879_U6, CellID=8796, LogicRNCID=501', \
'Label=UH0623_U6, CellID=6236, LogicRNCID=501', \
'Label=UH1854_U6, CellID=18546, LogicRNCID=501', \
'Label=UH1803_U4, CellID=18034, LogicRNCID=501', \
'Label=UH0830_U6, CellID=8306, LogicRNCID=501', \
'Label=UH1854_U5, CellID=18545, LogicRNCID=501', \
'Label=UH0980_U5, CellID=9805, LogicRNCID=501', \
'Label=UH0980_U4, CellID=9804, LogicRNCID=501', \
'Label=UH0981_U6, CellID=9816, LogicRNCID=501', \
'Label=UH3925_U5, CellID=39255, LogicRNCID=501', \
'Label=UH0981_U5, CellID=9815, LogicRNCID=501', \
'Label=UH0601_U6, CellID=6016, LogicRNCID=501', \
'Label=UH0601_U5, CellID=6015, LogicRNCID=501', \
'Label=UH0955_U5, CellID=9555, LogicRNCID=501', \
'Label=UH0955_U4, CellID=9554, LogicRNCID=501', \
'Label=UH0622_U6, CellID=6226, LogicRNCID=501', \
'Label=UH0622_U5, CellID=6225, LogicRNCID=501', \
'Label=UH0825_U4, CellID=8254, LogicRNCID=501', \
'Label=UH1951_U6, CellID=19516, LogicRNCID=501', \
'Label=UH3925_U4, CellID=39254, LogicRNCID=501', \
'Label=UH1928_U6, CellID=19286, LogicRNCID=501', \
'Label=UH0981_U4, CellID=9814, LogicRNCID=501', \
'Label=UH1928_U5, CellID=19285, LogicRNCID=501', \
'Label=UH0600_U4, CellID=6004, LogicRNCID=501', \
'Label=UH1803_U5, CellID=18035, LogicRNCID=501', \
'Label=UH1905_U4, CellID=19054, LogicRNCID=501', \
'Label=UH0881_U4, CellID=8814, LogicRNCID=501', \
'Label=UH0880_U6, CellID=8806, LogicRNCID=501', \
'Label=UH0880_U4, CellID=8804, LogicRNCID=501', \
'Label=UH1928_U4, CellID=19284, LogicRNCID=501', \
'Label=UH0955_U6, CellID=9556, LogicRNCID=501', \
'Label=UH3922_U4, CellID=39224, LogicRNCID=501', \
'Label=UH1976_U6, CellID=19766, LogicRNCID=501', \
'Label=UH0879_U5, CellID=8795, LogicRNCID=501', \
'Label=UH0623_U5, CellID=6235, LogicRNCID=501', \
'Label=UH3928_U5, CellID=39285, LogicRNCID=501', \
'Label=UH0600_U5, CellID=6005, LogicRNCID=501', \
'Label=UH3928_U4, CellID=39284, LogicRNCID=501', \
'Label=UH1803_U6, CellID=18036, LogicRNCID=501', \
'Label=UH1801_U4, CellID=18014, LogicRNCID=501', \
'Label=UH0830_U4, CellID=8304, LogicRNCID=501', \
'Label=UH1905_U5, CellID=19055, LogicRNCID=501', \
'Label=UH0881_U5, CellID=8815, LogicRNCID=501', \
'Label=UH1951_U5, CellID=19515, LogicRNCID=501', \
'Label=UH1951_U4, CellID=19514, LogicRNCID=501', \
'Label=UH3922_U5, CellID=39225, LogicRNCID=501', \
'Label=UH2950_U4, CellID=29504, LogicRNCID=501', \
'Label=UH0825_U6, CellID=8256, LogicRNCID=501', \
'Label=UH0825_U5, CellID=8255, LogicRNCID=501', \
'Label=UH2981_U5, CellID=29815, LogicRNCID=501', \
'Label=UH3928_U6, CellID=39286, LogicRNCID=501', \
'Label=UH2981_U4, CellID=29814, LogicRNCID=501', \
'Label=UH2950_U6, CellID=29506, LogicRNCID=501', \
'Label=UH2950_U5, CellID=29505, LogicRNCID=501', \
'Label=UH0622_U4, CellID=6224, LogicRNCID=501', \
'Label=UH0980_U6, CellID=9806, LogicRNCID=501', \
'Label=UH0623_U4, CellID=6234, LogicRNCID=501', \
'Label=UH3925_U6, CellID=39256, LogicRNCID=501', \
'Label=UH0830_U5, CellID=8305, LogicRNCID=501', \
'Label=UH1854_U4, CellID=18544, LogicRNCID=501', \
'Label=UH0881_U6, CellID=8816, LogicRNCID=501', \
'Label=UH1906_U6, CellID=19066, LogicRNCID=501', \
'Label=UH0831_U4, CellID=8314, LogicRNCID=501', \
'Label=UH1906_U5, CellID=19065, LogicRNCID=501', \
'Label=UH2981_U6, CellID=29816, LogicRNCID=501', \
'Label=UH1906_U4, CellID=19064, LogicRNCID=501', \
'Label=UH0748_U5, CellID=7485, LogicRNCID=501', \
'Label=UH0748_U6, CellID=7486, LogicRNCID=501', \
'Label=UH0748_U4, CellID=7484, LogicRNCID=501'
]  # кластер число активных сот 471
list_U900 = [
'Label=UH0761_U94, CellID=47614, LogicRNCID=501', \
'Label=UH0735_U96, CellID=47356, LogicRNCID=501', \
'Label=UH0735_U95, CellID=47355, LogicRNCID=501', \
'Label=UH0735_U94, CellID=47354, LogicRNCID=501', \
'Label=UH0965_U97, CellID=49657, LogicRNCID=501', \
'Label=UH0965_U96, CellID=49656, LogicRNCID=501', \
'Label=UH0965_U95, CellID=49655, LogicRNCID=501', \
'Label=UH0965_U94, CellID=49654, LogicRNCID=501', \
'Label=UH0734_U96, CellID=47346, LogicRNCID=501', \
'Label=UH0734_U95, CellID=47345, LogicRNCID=501', \
'Label=UH0734_U94, CellID=47344, LogicRNCID=501', \
'Label=UH0708_U96, CellID=47086, LogicRNCID=501', \
'Label=UH0708_U95, CellID=47085, LogicRNCID=501', \
'Label=UH0759_U96, CellID=47596, LogicRNCID=501', \
'Label=UH0708_U94, CellID=47084, LogicRNCID=501', \
'Label=UH0759_U95, CellID=47595, LogicRNCID=501', \
'Label=UH1936_U96, CellID=59366, LogicRNCID=501', \
'Label=UH1936_U95, CellID=59365, LogicRNCID=501', \
'Label=UH0963_U96, CellID=49636, LogicRNCID=501', \
'Label=UH0835_U96, CellID=48356, LogicRNCID=501', \
'Label=UH0963_U95, CellID=49635, LogicRNCID=501', \
'Label=UH0835_U95, CellID=48355, LogicRNCID=501', \
'Label=UH0963_U94, CellID=49634, LogicRNCID=501', \
'Label=UH0835_U94, CellID=48354, LogicRNCID=501', \
'Label=UH0758_U96, CellID=47586, LogicRNCID=501', \
'Label=UH0758_U95, CellID=47585, LogicRNCID=501', \
'Label=UH0758_U94, CellID=47584, LogicRNCID=501', \
'Label=UH0732_U96, CellID=47326, LogicRNCID=501', \
'Label=UH0732_U95, CellID=47325, LogicRNCID=501', \
'Label=UH0732_U94, CellID=47324, LogicRNCID=501', \
'Label=UH0962_U96, CellID=49626, LogicRNCID=501', \
'Label=UH0706_U96, CellID=47066, LogicRNCID=501', \
'Label=UH0962_U95, CellID=49625, LogicRNCID=501', \
'Label=UH0706_U95, CellID=47065, LogicRNCID=501', \
'Label=UH0962_U94, CellID=49624, LogicRNCID=501', \
'Label=UH0706_U94, CellID=47064, LogicRNCID=501', \
'Label=UH0731_U96, CellID=47316, LogicRNCID=501', \
'Label=UH0731_U95, CellID=47315, LogicRNCID=501', \
'Label=UH0731_U94, CellID=47314, LogicRNCID=501', \
'Label=UH0833_U96, CellID=48336, LogicRNCID=501', \
'Label=UH0705_U96, CellID=47056, LogicRNCID=501', \
'Label=UH0833_U95, CellID=48335, LogicRNCID=501', \
'Label=UH0705_U95, CellID=47055, LogicRNCID=501', \
'Label=UH0833_U94, CellID=48334, LogicRNCID=501', \
'Label=UH0756_U96, CellID=47566, LogicRNCID=501', \
'Label=UH0705_U94, CellID=47054, LogicRNCID=501', \
'Label=UH0807_U96, CellID=48076, LogicRNCID=501', \
'Label=UH0756_U94, CellID=47564, LogicRNCID=501', \
'Label=UH0807_U95, CellID=48075, LogicRNCID=501', \
'Label=UH0807_U94, CellID=48074, LogicRNCID=501', \
'Label=UH0602_U96, CellID=46026, LogicRNCID=501', \
'Label=UH0602_U95, CellID=46025, LogicRNCID=501', \
'Label=UH0602_U94, CellID=46024, LogicRNCID=501', \
'Label=UH0960_U96, CellID=49606, LogicRNCID=501', \
'Label=UH0832_U96, CellID=48326, LogicRNCID=501', \
'Label=UH0704_U96, CellID=47046, LogicRNCID=501', \
'Label=UH0960_U95, CellID=49605, LogicRNCID=501', \
'Label=UH0832_U95, CellID=48325, LogicRNCID=501', \
'Label=UH0704_U95, CellID=47045, LogicRNCID=501', \
'Label=UH0960_U94, CellID=49604, LogicRNCID=501', \
'Label=UH0832_U94, CellID=48324, LogicRNCID=501', \
'Label=UH0601_U96, CellID=46016, LogicRNCID=501', \
'Label=UH0825_U96, CellID=48256, LogicRNCID=501', \
'Label=UH0623_U95, CellID=46235, LogicRNCID=501', \
'Label=UH0760_U97, CellID=50074, LogicRNCID=501', \
'Label=UH1945_U95, CellID=50057, LogicRNCID=501', \
'Label=UH1801_U96, CellID=58016, LogicRNCID=501', \
'Label=UH0831_U96, CellID=48316, LogicRNCID=501', \
'Label=UH0703_U96, CellID=47036, LogicRNCID=501', \
'Label=UH0831_U95, CellID=48315, LogicRNCID=501', \
'Label=UH0703_U95, CellID=47035, LogicRNCID=501', \
'Label=UH0831_U94, CellID=48314, LogicRNCID=501', \
'Label=UH0760_U95, CellID=50072, LogicRNCID=501', \
'Label=UH0760_U94, CellID=50071, LogicRNCID=501', \
'Label=UH0881_U95, CellID=48815, LogicRNCID=501', \
'Label=UH0881_U94, CellID=48814, LogicRNCID=501', \
'Label=UH1945_U96, CellID=50058, LogicRNCID=501', \
'Label=UH2709_U95, CellID=50066, LogicRNCID=501', \
'Label=UH0955_U94, CellID=49554, LogicRNCID=501', \
'Label=UH0750_U96, CellID=47506, LogicRNCID=501', \
'Label=UH0622_U96, CellID=46226, LogicRNCID=501', \
'Label=UH0601_U95, CellID=46015, LogicRNCID=501', \
'Label=UH0601_U94, CellID=46014, LogicRNCID=501', \
'Label=UH1945_U99, CellID=50061, LogicRNCID=501', \
'Label=UH1945_U98, CellID=50060, LogicRNCID=501', \
'Label=UH0723_U96, CellID=47236, LogicRNCID=501', \
'Label=UH2709_U94, CellID=50065, LogicRNCID=501', \
'Label=UH0750_U95, CellID=47505, LogicRNCID=501', \
'Label=UH0622_U95, CellID=46225, LogicRNCID=501', \
'Label=UH1951_U96, CellID=50064, LogicRNCID=501', \
'Label=UH0702_U95, CellID=47025, LogicRNCID=501', \
'Label=UH0881_U96, CellID=48816, LogicRNCID=501', \
'Label=UH0830_U94, CellID=48304, LogicRNCID=501', \
'Label=UH0702_U94, CellID=47024, LogicRNCID=501', \
'Label=UH0751_U96, CellID=47516, LogicRNCID=501', \
'Label=UH0623_U96, CellID=46236, LogicRNCID=501', \
'Label=UH0750_U94, CellID=47504, LogicRNCID=501', \
'Label=UH0622_U94, CellID=46224, LogicRNCID=501', \
'Label=UH1951_U95, CellID=50063, LogicRNCID=501', \
'Label=UH1951_U94, CellID=50062, LogicRNCID=501', \
'Label=UH0840_U95, CellID=48405, LogicRNCID=501', \
'Label=UH0840_U94, CellID=48404, LogicRNCID=501', \
'Label=UH0763_U96, CellID=47636, LogicRNCID=501', \
'Label=UH0879_U96, CellID=55060, LogicRNCID=501', \
'Label=UH0879_U95, CellID=55061, LogicRNCID=501', \
'Label=UH0636_U95, CellID=46365, LogicRNCID=501', \
'Label=UH0766_U95, CellID=47665, LogicRNCID=501', \
'Label=UH0611_U96, CellID=46116, LogicRNCID=501', \
'Label=UH0739_U95, CellID=47395, LogicRNCID=501', \
'Label=UH0611_U95, CellID=46115, LogicRNCID=501', \
'Label=UH0739_U96, CellID=47396, LogicRNCID=501', \
'Label=UH0839_U96, CellID=48396, LogicRNCID=501', \
'Label=UH0843_U95, CellID=48435, LogicRNCID=501', \
'Label=UH0766_U94, CellID=47666, LogicRNCID=501', \
'Label=UH0736_U94, CellID=47364, LogicRNCID=501', \
'Label=UH0765_U96, CellID=47656, LogicRNCID=501', \
'Label=UH0713_U94, CellID=47134, LogicRNCID=501', \
'Label=UH0738_U94, CellID=47384, LogicRNCID=501', \
'Label=UH0840_U96, CellID=48406, LogicRNCID=501', \
'Label=UH0736_U96, CellID=47366, LogicRNCID=501', \
'Label=UH0736_U95, CellID=47365, LogicRNCID=501', \
'Label=UH0839_U99, CellID=48399, LogicRNCID=501', \
'Label=UH0714_U95, CellID=47145, LogicRNCID=501', \
'Label=UH0763_U94, CellID=47634, LogicRNCID=501', \
'Label=UH0763_U95, CellID=47635, LogicRNCID=501', \
'Label=UH0713_U96, CellID=47136, LogicRNCID=501', \
'Label=UH0841_U95, CellID=48415, LogicRNCID=501', \
'Label=UH0713_U95, CellID=47135, LogicRNCID=501', \
'Label=UH0738_U96, CellID=47386, LogicRNCID=501', \
'Label=UH0738_U95, CellID=47385, LogicRNCID=501', \
'Label=UH0838_U95, CellID=48385, LogicRNCID=501', \
'Label=UH0761_U96, CellID=47616, LogicRNCID=501', \
'Label=UH0737_U94, CellID=47374, LogicRNCID=501', \
'Label=UH0839_U97, CellID=48397, LogicRNCID=501', \
'Label=UH0611_U94, CellID=46114, LogicRNCID=501', \
'Label=UH0740_U95, CellID=47405, LogicRNCID=501', \
'Label=UH0740_U96, CellID=47406, LogicRNCID=501', \
'Label=UH0841_U96, CellID=48416, LogicRNCID=501', \
'Label=UH0838_U94, CellID=48384, LogicRNCID=501', \
'Label=UH0761_U97, CellID=47617, LogicRNCID=501', \
'Label=UH0838_U96, CellID=48386, LogicRNCID=501', \
'Label=UH0740_U94, CellID=47404, LogicRNCID=501', \
'Label=UH0737_U95, CellID=47375, LogicRNCID=501', \
'Label=UH0879_U94, CellID=55062, LogicRNCID=501', \
'Label=UH0636_U96, CellID=46366, LogicRNCID=501', \
'Label=UH0841_U94, CellID=48414, LogicRNCID=501', \
'Label=UH0765_U94, CellID=47654, LogicRNCID=501', \
'Label=UH0714_U94, CellID=47144, LogicRNCID=501', \
'Label=UH0715_U94, CellID=47154, LogicRNCID=501', \
'Label=UH0843_U94, CellID=48434, LogicRNCID=501', \
'Label=UH0715_U95, CellID=47155, LogicRNCID=501', \
'Label=UH0839_U94, CellID=48394, LogicRNCID=501', \
'Label=UH0839_U95, CellID=48395, LogicRNCID=501', \
'Label=UH0737_U96, CellID=47376, LogicRNCID=501', \
'Label=UH0766_U96, CellID=47664, LogicRNCID=501', \
'Label=UH0636_U94, CellID=46364, LogicRNCID=501', \
'Label=UH1928_U95, CellID=50049, LogicRNCID=501', \
'Label=UH0830_U96, CellID=48306, LogicRNCID=501', \
'Label=UH0702_U96, CellID=47026, LogicRNCID=501', \
'Label=UH0881_U97, CellID=48817, LogicRNCID=501', \
'Label=UH0830_U95, CellID=48305, LogicRNCID=501', \
'Label=UH0751_U94, CellID=47514, LogicRNCID=501', \
'Label=UH0623_U94, CellID=46234, LogicRNCID=501', \
'Label=UH0760_U96, CellID=50073, LogicRNCID=501', \
'Label=UH1945_U94, CellID=50056, LogicRNCID=501', \
'Label=UH0752_U94, CellID=47524, LogicRNCID=501', \
'Label=UH0723_U95, CellID=47235, LogicRNCID=501', \
'Label=UH1928_U96, CellID=50050, LogicRNCID=501', \
'Label=UH0723_U94, CellID=47234, LogicRNCID=501', \
'Label=UH1801_U94, CellID=58014, LogicRNCID=501', \
'Label=UH1945_U97, CellID=50059, LogicRNCID=501', \
'Label=UH0955_U96, CellID=49556, LogicRNCID=501', \
'Label=UH2709_U96, CellID=50067, LogicRNCID=501', \
'Label=UH0955_U95, CellID=49555, LogicRNCID=501', \
'Label=UH0752_U96, CellID=47526, LogicRNCID=501', \
'Label=UH0752_U95, CellID=47525, LogicRNCID=501', \
'Label=UH0703_U94, CellID=47034, LogicRNCID=501', \
'Label=UH0754_U95, CellID=47545, LogicRNCID=501', \
'Label=UH0754_U94, CellID=47544, LogicRNCID=501', \
'Label=UH0843_U96, CellID=48436, LogicRNCID=501', \
'Label=UH0715_U96, CellID=47156, LogicRNCID=501', \
'Label=UH1803_U94, CellID=50019, LogicRNCID=501', \
'Label=UH0745_U96, CellID=47456, LogicRNCID=501', \
'Label=UH0973_U96, CellID=49736, LogicRNCID=501', \
'Label=UH0845_U96, CellID=48456, LogicRNCID=501', \
'Label=UH0717_U96, CellID=47176, LogicRNCID=501', \
'Label=UH0973_U95, CellID=49735, LogicRNCID=501', \
'Label=UH0845_U95, CellID=48455, LogicRNCID=501', \
'Label=UH0951_U96, CellID=49516, LogicRNCID=501', \
'Label=UH0823_U96, CellID=48236, LogicRNCID=501', \
'Label=UH1805_U96, CellID=50027, LogicRNCID=501', \
'Label=UH0823_U95, CellID=48235, LogicRNCID=501', \
'Label=UH1604_U94, CellID=50013, LogicRNCID=501', \
'Label=UH1964_U95, CellID=55388, LogicRNCID=501', \
'Label=UH1603_U96, CellID=50012, LogicRNCID=501', \
'Label=UH1964_U94, CellID=55387, LogicRNCID=501', \
'Label=UH1882_U96, CellID=50033, LogicRNCID=501', \
'Label=UH0820_U95, CellID=48205, LogicRNCID=501', \
'Label=UH0820_U94, CellID=48204, LogicRNCID=501', \
'Label=UH0743_U96, CellID=47436, LogicRNCID=501', \
'Label=UH0743_U95, CellID=47435, LogicRNCID=501', \
'Label=UH0821_U95, CellID=48215, LogicRNCID=501', \
'Label=UH0821_U94, CellID=48214, LogicRNCID=501', \
'Label=UH0825_U95, CellID=48255, LogicRNCID=501', \
'Label=UH0825_U94, CellID=48254, LogicRNCID=501', \
'Label=UH0820_U99, CellID=48209, LogicRNCID=501', \
'Label=UH0844_U95, CellID=48445, LogicRNCID=501', \
'Label=UH0844_U94, CellID=48444, LogicRNCID=501', \
'Label=UH0767_U96, CellID=47676, LogicRNCID=501', \
'Label=UH0744_U96, CellID=47446, LogicRNCID=501', \
'Label=UH0744_U95, CellID=47445, LogicRNCID=501', \
'Label=UH1902_U95, CellID=50035, LogicRNCID=501', \
'Label=UH0747_U95, CellID=47475, LogicRNCID=501', \
'Label=UH1902_U94, CellID=50034, LogicRNCID=501', \
'Label=UH0747_U94, CellID=47474, LogicRNCID=501', \
'Label=UH0849_U94, CellID=48494, LogicRNCID=501', \
'Label=UH1804_U96, CellID=50024, LogicRNCID=501', \
'Label=UH0618_U94, CellID=46184, LogicRNCID=501', \
'Label=UH1804_U95, CellID=50023, LogicRNCID=501', \
'Label=UH1804_U94, CellID=50022, LogicRNCID=501', \
'Label=UH1803_U96, CellID=50021, LogicRNCID=501', \
'Label=UH1803_U95, CellID=50020, LogicRNCID=501', \
'Label=UH0744_U94, CellID=47444, LogicRNCID=501', \
'Label=UH0869_U96, CellID=48696, LogicRNCID=501', \
'Label=UH0869_U95, CellID=48695, LogicRNCID=501', \
'Label=UH0869_U94, CellID=48694, LogicRNCID=501', \
'Label=UH1603_U95, CellID=50011, LogicRNCID=501', \
'Label=UH0821_U99, CellID=48219, LogicRNCID=501', \
'Label=UH1603_U94, CellID=50010, LogicRNCID=501', \
'Label=UH0821_U98, CellID=48218, LogicRNCID=501', \
'Label=UH0770_U96, CellID=47706, LogicRNCID=501', \
'Label=UH1902_U96, CellID=50036, LogicRNCID=501', \
'Label=UH1604_U96, CellID=50015, LogicRNCID=501', \
'Label=UH0745_U95, CellID=47455, LogicRNCID=501', \
'Label=UH1604_U95, CellID=50014, LogicRNCID=501', \
'Label=UH0745_U94, CellID=47454, LogicRNCID=501', \
'Label=UH1964_U96, CellID=55389, LogicRNCID=501', \
'Label=UH0717_U95, CellID=47175, LogicRNCID=501', \
'Label=UH0973_U94, CellID=49734, LogicRNCID=501', \
'Label=UH0845_U94, CellID=48454, LogicRNCID=501', \
'Label=UH0717_U94, CellID=47174, LogicRNCID=501', \
'Label=UH0646_U96, CellID=51012, LogicRNCID=501', \
'Label=UH0646_U95, CellID=51011, LogicRNCID=501', \
'Label=UH0646_U94, CellID=51010, LogicRNCID=501', \
'Label=UH0742_U96, CellID=47426, LogicRNCID=501', \
'Label=UH1882_U94, CellID=50031, LogicRNCID=501', \
'Label=UH0849_U95, CellID=48495, LogicRNCID=501', \
'Label=UH0767_U95, CellID=47675, LogicRNCID=501', \
'Label=UH0767_U94, CellID=47674, LogicRNCID=501', \
'Label=UH0869_U97, CellID=48697, LogicRNCID=501', \
'Label=UH0614_U94, CellID=46144, LogicRNCID=501', \
'Label=UH1602_U94, CellID=51007, LogicRNCID=501', \
'Label=UH0844_U96, CellID=48446, LogicRNCID=501', \
'Label=UH1907_U95, CellID=50038, LogicRNCID=501', \
'Label=UH1907_U94, CellID=50037, LogicRNCID=501', \
'Label=UH1805_U95, CellID=50026, LogicRNCID=501', \
'Label=UH0951_U94, CellID=49514, LogicRNCID=501', \
'Label=UH0823_U94, CellID=48234, LogicRNCID=501', \
'Label=UH0746_U96, CellID=47466, LogicRNCID=501', \
'Label=UH0618_U96, CellID=46186, LogicRNCID=501', \
'Label=UH1805_U94, CellID=50025, LogicRNCID=501', \
'Label=UH0746_U95, CellID=47465, LogicRNCID=501', \
'Label=UH0618_U95, CellID=46185, LogicRNCID=501', \
'Label=UH0614_U96, CellID=46146, LogicRNCID=501', \
'Label=UH1602_U96, CellID=51009, LogicRNCID=501', \
'Label=UH0742_U95, CellID=47425, LogicRNCID=501', \
'Label=UH0614_U95, CellID=46145, LogicRNCID=501', \
'Label=UH1602_U95, CellID=51008, LogicRNCID=501', \
'Label=UH0742_U94, CellID=47424, LogicRNCID=501', \
'Label=UH0820_U98, CellID=48208, LogicRNCID=501', \
'Label=UH0820_U97, CellID=48207, LogicRNCID=501', \
'Label=UH0820_U96, CellID=48206, LogicRNCID=501', \
'Label=UH0821_U97, CellID=48217, LogicRNCID=501', \
'Label=UH0770_U95, CellID=47705, LogicRNCID=501', \
'Label=UH0821_U96, CellID=48216, LogicRNCID=501', \
'Label=UH0770_U94, CellID=47704, LogicRNCID=501', \
'Label=UH1882_U95, CellID=50032, LogicRNCID=501', \
'Label=UH0849_U96, CellID=48496, LogicRNCID=501', \
'Label=UH1907_U96, CellID=50039, LogicRNCID=501', \
'Label=UH0711_U96, CellID=47116, LogicRNCID=501', \
'Label=UH0711_U95, CellID=47115, LogicRNCID=501', \
'Label=UH0711_U94, CellID=47114, LogicRNCID=501', \
'Label=UH0748_U94, CellID=47484, LogicRNCID=501', \
'Label=UH0748_U95, CellID=47485, LogicRNCID=501', \
'Label=UH0748_U96, CellID=47486, LogicRNCID=501'
]  # кластер число активных сот 216
list_F1_10612 = [
'Label=UH0742_U2, CellID=7422, LogicRNCID=501', \
'Label=UH1945_U3, CellID=19453, LogicRNCID=501', \
'Label=UH0742_U1, CellID=7421, LogicRNCID=501', \
'Label=UH1945_U2, CellID=19452, LogicRNCID=501', \
'Label=UH1945_U1, CellID=19451, LogicRNCID=501', \
'Label=UH0972_U3, CellID=9723, LogicRNCID=501', \
'Label=UH0844_U3, CellID=8443, LogicRNCID=501', \
'Label=UH0972_U2, CellID=9722, LogicRNCID=501', \
'Label=UH0844_U2, CellID=8442, LogicRNCID=501', \
'Label=UH0972_U1, CellID=9721, LogicRNCID=501', \
'Label=UH0844_U1, CellID=8441, LogicRNCID=501', \
'Label=UH2763_U3, CellID=27633, LogicRNCID=501', \
'Label=UH0880_U11, CellID=65520, LogicRNCID=501', \
'Label=UH2763_U2, CellID=27632, LogicRNCID=501', \
'Label=UH2763_U1, CellID=27631, LogicRNCID=501', \
'Label=UH1918_U3, CellID=19183, LogicRNCID=501', \
'Label=UH1918_U2, CellID=19182, LogicRNCID=501', \
'Label=UH1918_U1, CellID=19181, LogicRNCID=501', \
'Label=UH1892_U3, CellID=18923, LogicRNCID=501', \
'Label=UH1892_U2, CellID=18922, LogicRNCID=501', \
'Label=UH0970_U3, CellID=9703, LogicRNCID=501', \
'Label=UH0842_U3, CellID=8423, LogicRNCID=501', \
'Label=UH0714_U3, CellID=7143, LogicRNCID=501', \
'Label=UH0970_U2, CellID=9702, LogicRNCID=501', \
'Label=UH0842_U2, CellID=8422, LogicRNCID=501', \
'Label=UH0714_U2, CellID=7142, LogicRNCID=501', \
'Label=UH0970_U1, CellID=9701, LogicRNCID=501', \
'Label=UH0842_U1, CellID=8421, LogicRNCID=501', \
'Label=UH0765_U3, CellID=7653, LogicRNCID=501', \
'Label=UH0714_U1, CellID=7141, LogicRNCID=501', \
'Label=UH0765_U2, CellID=7652, LogicRNCID=501', \
'Label=UH0765_U1, CellID=7651, LogicRNCID=501', \
'Label=UH0995_U3, CellID=9953, LogicRNCID=501', \
'Label=UH0611_U3, CellID=6113, LogicRNCID=501', \
'Label=UH0995_U2, CellID=9952, LogicRNCID=501', \
'Label=UH3990_U3, CellID=39903, LogicRNCID=501', \
'Label=UH0995_U1, CellID=9951, LogicRNCID=501', \
'Label=UH0611_U1, CellID=6111, LogicRNCID=501', \
'Label=UH3990_U2, CellID=39902, LogicRNCID=501', \
'Label=UH3990_U1, CellID=39901, LogicRNCID=501', \
'Label=UH0841_U3, CellID=8413, LogicRNCID=501', \
'Label=UH0841_U2, CellID=8412, LogicRNCID=501', \
'Label=UH1916_U3, CellID=19163, LogicRNCID=501', \
'Label=UH0841_U1, CellID=8411, LogicRNCID=501', \
'Label=UH0636_U3, CellID=6363, LogicRNCID=501', \
'Label=UH0508_U3, CellID=5083, LogicRNCID=501', \
'Label=UH1916_U2, CellID=19162, LogicRNCID=501', \
'Label=UH0636_U2, CellID=6362, LogicRNCID=501', \
'Label=UH0508_U2, CellID=5082, LogicRNCID=501', \
'Label=UH1907_U66, CellID=65497, LogicRNCID=501', \
'Label=UH1916_U1, CellID=19161, LogicRNCID=501', \
'Label=UH1907_U55, CellID=65496, LogicRNCID=501', \
'Label=UH1907_U44, CellID=65495, LogicRNCID=501', \
'Label=UH0994_U3, CellID=9943, LogicRNCID=501', \
'Label=UH0738_U3, CellID=7383, LogicRNCID=501', \
'Label=UH0610_U3, CellID=6103, LogicRNCID=501', \
'Label=UH0994_U2, CellID=9942, LogicRNCID=501', \
'Label=UH0738_U2, CellID=7382, LogicRNCID=501', \
'Label=UH0610_U2, CellID=6102, LogicRNCID=501', \
'Label=UH3989_U3, CellID=39893, LogicRNCID=501', \
'Label=UH2709_U3, CellID=27093, LogicRNCID=501', \
'Label=UH0994_U1, CellID=9941, LogicRNCID=501', \
'Label=UH0738_U1, CellID=7381, LogicRNCID=501', \
'Label=UH3989_U2, CellID=39892, LogicRNCID=501', \
'Label=UH2709_U2, CellID=27092, LogicRNCID=501', \
'Label=UH3989_U1, CellID=39891, LogicRNCID=501', \
'Label=UH2709_U1, CellID=27091, LogicRNCID=501', \
'Label=UH0840_U3, CellID=8403, LogicRNCID=501', \
'Label=UH0840_U2, CellID=8402, LogicRNCID=501', \
'Label=UH0840_U1, CellID=8401, LogicRNCID=501', \
'Label=UH0763_U3, CellID=7633, LogicRNCID=501', \
'Label=UH0763_U2, CellID=7632, LogicRNCID=501', \
'Label=UH0763_U1, CellID=7631, LogicRNCID=501', \
'Label=UH2913_U3, CellID=29133, LogicRNCID=501', \
'Label=UH2913_U2, CellID=29132, LogicRNCID=501', \
'Label=UH2913_U1, CellID=29131, LogicRNCID=501', \
'Label=UH0839_U3, CellID=8393, LogicRNCID=501', \
'Label=UH0711_U3, CellID=7113, LogicRNCID=501', \
'Label=UH0839_U2, CellID=8392, LogicRNCID=501', \
'Label=UH0711_U2, CellID=7112, LogicRNCID=501', \
'Label=UH2938_U3, CellID=29383, LogicRNCID=501', \
'Label=UH0839_U1, CellID=8391, LogicRNCID=501', \
'Label=UH0711_U1, CellID=7111, LogicRNCID=501', \
'Label=UH2938_U2, CellID=29382, LogicRNCID=501', \
'Label=UH2938_U1, CellID=29381, LogicRNCID=501', \
'Label=UH1939_U3, CellID=19393, LogicRNCID=501', \
'Label=UH1939_U2, CellID=19392, LogicRNCID=501', \
'Label=UH1907_U3, CellID=19073, LogicRNCID=501', \
'Label=UH0960_U1, CellID=9601, LogicRNCID=501', \
'Label=UH0832_U1, CellID=8321, LogicRNCID=501', \
'Label=UH1907_U2, CellID=19072, LogicRNCID=501', \
'Label=UH1705_U2, CellID=17052, LogicRNCID=501', \
'Label=UH1833_U1, CellID=18331, LogicRNCID=501', \
'Label=UH1705_U1, CellID=17051, LogicRNCID=501', \
'Label=UH0988_U3, CellID=9883, LogicRNCID=501', \
'Label=UH2983_U3, CellID=29833, LogicRNCID=501', \
'Label=UH2932_U1, CellID=29321, LogicRNCID=501', \
'Label=UH1908_U1, CellID=19081, LogicRNCID=501', \
'Label=UH0963_U2, CellID=9632, LogicRNCID=501', \
'Label=UH0835_U2, CellID=8352, LogicRNCID=501', \
'Label=UH0966_U1, CellID=9661, LogicRNCID=501', \
'Label=UH0838_U1, CellID=8381, LogicRNCID=501', \
'Label=UH1935_U2, CellID=19352, LogicRNCID=501', \
'Label=UH1935_U1, CellID=19351, LogicRNCID=501', \
'Label=UH0962_U3, CellID=9623, LogicRNCID=501', \
'Label=UH1707_U1, CellID=17071, LogicRNCID=501', \
'Label=UH0734_U3, CellID=7343, LogicRNCID=501', \
'Label=UH0734_U2, CellID=7342, LogicRNCID=501', \
'Label=UH0734_U1, CellID=7341, LogicRNCID=501', \
'Label=UH0961_U1, CellID=9611, LogicRNCID=501', \
'Label=UH2932_U2, CellID=29322, LogicRNCID=501', \
'Label=UH1908_U2, CellID=19082, LogicRNCID=501', \
'Label=UH1960_U3, CellID=19603, LogicRNCID=501', \
'Label=UH1704_U3, CellID=17043, LogicRNCID=501', \
'Label=UH2984_U2, CellID=29842, LogicRNCID=501', \
'Label=UH1960_U2, CellID=19602, LogicRNCID=501', \
'Label=UH1939_U1, CellID=19391, LogicRNCID=501', \
'Label=UH0966_U3, CellID=9663, LogicRNCID=501', \
'Label=UH0838_U3, CellID=8383, LogicRNCID=501', \
'Label=UH0966_U2, CellID=9662, LogicRNCID=501', \
'Label=UH0838_U2, CellID=8382, LogicRNCID=501', \
'Label=UH0961_U3, CellID=9613, LogicRNCID=501', \
'Label=UH0602_U1, CellID=6021, LogicRNCID=501', \
'Label=UH1805_U2, CellID=18052, LogicRNCID=501', \
'Label=UH1704_U2, CellID=17042, LogicRNCID=501', \
'Label=UH2984_U1, CellID=29841, LogicRNCID=501', \
'Label=UH1960_U1, CellID=19601, LogicRNCID=501', \
'Label=UH1704_U1, CellID=17041, LogicRNCID=501', \
'Label=UH4600_U1, CellID=46001, LogicRNCID=501', \
'Label=UH1707_U3, CellID=17073, LogicRNCID=501', \
'Label=UH1707_U2, CellID=17072, LogicRNCID=501', \
'Label=UH1936_U2, CellID=19362, LogicRNCID=501', \
'Label=UH1936_U1, CellID=19361, LogicRNCID=501', \
'Label=UH0963_U3, CellID=9633, LogicRNCID=501', \
'Label=UH0835_U3, CellID=8353, LogicRNCID=501', \
'Label=UH1833_U3, CellID=18333, LogicRNCID=501', \
'Label=UH1705_U3, CellID=17053, LogicRNCID=501', \
'Label=UH0758_U1, CellID=7581, LogicRNCID=501', \
'Label=UH1833_U2, CellID=18332, LogicRNCID=501', \
'Label=UH1984_U2, CellID=19842, LogicRNCID=501', \
'Label=UH0770_U2, CellID=7702, LogicRNCID=501', \
'Label=UH0770_U1, CellID=7701, LogicRNCID=501', \
'Label=UH0821_U3, CellID=8213, LogicRNCID=501', \
'Label=UH0950_U1, CellID=9501, LogicRNCID=501', \
'Label=UH0617_U3, CellID=6173, LogicRNCID=501', \
'Label=UH1976_U2, CellID=19762, LogicRNCID=501', \
'Label=UH0952_U1, CellID=9521, LogicRNCID=501', \
'Label=UH3995_U3, CellID=39953, LogicRNCID=501', \
'Label=UH0618_U2, CellID=6182, LogicRNCID=501', \
'Label=UH0618_U1, CellID=6181, LogicRNCID=501', \
'Label=UH0820_U2, CellID=8202, LogicRNCID=501', \
'Label=UH1976_U3, CellID=19763, LogicRNCID=501', \
'Label=UH0845_U3, CellID=8453, LogicRNCID=501', \
'Label=UH0845_U2, CellID=8452, LogicRNCID=501', \
'Label=UH0973_U1, CellID=9731, LogicRNCID=501', \
'Label=UH0845_U1, CellID=8451, LogicRNCID=501', \
'Label=UH0717_U2, CellID=7172, LogicRNCID=501', \
'Label=UH0973_U2, CellID=9732, LogicRNCID=501', \
'Label=UH0717_U3, CellID=7173, LogicRNCID=501', \
'Label=UH0820_U1, CellID=8201, LogicRNCID=501', \
'Label=UH0743_U2, CellID=7432, LogicRNCID=501', \
'Label=UH0823_U2, CellID=8232, LogicRNCID=501', \
'Label=UH0951_U1, CellID=9511, LogicRNCID=501', \
'Label=UH0618_U3, CellID=6183, LogicRNCID=501', \
'Label=UH0822_U2, CellID=8222, LogicRNCID=501', \
'Label=UH0770_U3, CellID=7703, LogicRNCID=501', \
'Label=UH0743_U1, CellID=7431, LogicRNCID=501', \
'Label=UH0973_U3, CellID=9733, LogicRNCID=501', \
'Label=UH3995_U2, CellID=39952, LogicRNCID=501', \
'Label=UH1947_U2, CellID=19472, LogicRNCID=501', \
'Label=UH3995_U1, CellID=39951, LogicRNCID=501', \
'Label=UH0823_U3, CellID=8233, LogicRNCID=501', \
'Label=UH0821_U2, CellID=8212, LogicRNCID=501', \
'Label=UH0821_U1, CellID=8211, LogicRNCID=501', \
'Label=UH0950_U3, CellID=9503, LogicRNCID=501', \
'Label=UH0822_U3, CellID=8223, LogicRNCID=501', \
'Label=UH0951_U3, CellID=9513, LogicRNCID=501', \
'Label=UH0849_U3, CellID=8493, LogicRNCID=501', \
'Label=UH3921_U2, CellID=39212, LogicRNCID=501', \
'Label=UH0849_U2, CellID=8492, LogicRNCID=501', \
'Label=UH3921_U1, CellID=39211, LogicRNCID=501', \
'Label=UH3996_U3, CellID=39963, LogicRNCID=501', \
'Label=UH3996_U1, CellID=39961, LogicRNCID=501', \
'Label=UH3996_U2, CellID=39962, LogicRNCID=501', \
'Label=UH0617_U1, CellID=6171, LogicRNCID=501', \
'Label=UH0717_U1, CellID=7171, LogicRNCID=501', \
'Label=UH1947_U1, CellID=19471, LogicRNCID=501', \
'Label=UH0974_U1, CellID=9741, LogicRNCID=501', \
'Label=UH3920_U1, CellID=39201, LogicRNCID=501', \
'Label=UH0747_U2, CellID=7472, LogicRNCID=501', \
'Label=UH3921_U3, CellID=39213, LogicRNCID=501', \
'Label=UH3920_U2, CellID=39202, LogicRNCID=501', \
'Label=UH3918_U1, CellID=39181, LogicRNCID=501', \
'Label=UH0974_U2, CellID=9742, LogicRNCID=501', \
'Label=UH0974_U3, CellID=9743, LogicRNCID=501', \
'Label=UH0849_U1, CellID=8491, LogicRNCID=501', \
'Label=UH3918_U3, CellID=39183, LogicRNCID=501', \
'Label=UH0950_U2, CellID=9502, LogicRNCID=501', \
'Label=UH0823_U1, CellID=8231, LogicRNCID=501', \
'Label=UH0743_U3, CellID=7433, LogicRNCID=501', \
'Label=UH0952_U3, CellID=9523, LogicRNCID=501', \
'Label=UH1947_U3, CellID=19473, LogicRNCID=501', \
'Label=UH1976_U1, CellID=19761, LogicRNCID=501', \
'Label=UH0952_U2, CellID=9522, LogicRNCID=501', \
'Label=UH0617_U2, CellID=6172, LogicRNCID=501', \
'Label=UH0822_U1, CellID=8221, LogicRNCID=501', \
'Label=UH0960_U2, CellID=9602, LogicRNCID=501', \
'Label=UH0832_U2, CellID=8322, LogicRNCID=501', \
'Label=UH1984_U1, CellID=19841, LogicRNCID=501', \
'Label=UH0962_U2, CellID=9622, LogicRNCID=501', \
'Label=UH0962_U1, CellID=9621, LogicRNCID=501', \
'Label=UH0965_U1, CellID=9651, LogicRNCID=501', \
'Label=UH4600_U2, CellID=46002, LogicRNCID=501', \
'Label=UH0988_U2, CellID=9882, LogicRNCID=501', \
'Label=UH1935_U3, CellID=19353, LogicRNCID=501', \
'Label=UH0988_U1, CellID=9881, LogicRNCID=501', \
'Label=UH1703_U3, CellID=17033, LogicRNCID=501', \
'Label=UH0807_U3, CellID=8073, LogicRNCID=501', \
'Label=UH2983_U2, CellID=29832, LogicRNCID=501', \
'Label=UH1703_U2, CellID=17032, LogicRNCID=501', \
'Label=UH1936_U3, CellID=19363, LogicRNCID=501', \
'Label=UH1805_U1, CellID=18051, LogicRNCID=501', \
'Label=UH0960_U3, CellID=9603, LogicRNCID=501', \
'Label=UH0832_U3, CellID=8323, LogicRNCID=501', \
'Label=UH1882_U2, CellID=18822, LogicRNCID=501', \
'Label=UH0602_U2, CellID=6022, LogicRNCID=501', \
'Label=UH1882_U1, CellID=18821, LogicRNCID=501', \
'Label=UH1805_U3, CellID=18053, LogicRNCID=501', \
'Label=UH0965_U3, CellID=9653, LogicRNCID=501', \
'Label=UH0965_U2, CellID=9652, LogicRNCID=501', \
'Label=UH4600_U3, CellID=46003, LogicRNCID=501', \
'Label=UH0963_U1, CellID=9631, LogicRNCID=501', \
'Label=UH0835_U1, CellID=8351, LogicRNCID=501', \
'Label=UH0758_U3, CellID=7583, LogicRNCID=501', \
'Label=UH0758_U2, CellID=7582, LogicRNCID=501', \
'Label=UH0807_U2, CellID=8072, LogicRNCID=501', \
'Label=UH2983_U1, CellID=29831, LogicRNCID=501', \
'Label=UH1882_U3, CellID=18823, LogicRNCID=501', \
'Label=UH1703_U1, CellID=17031, LogicRNCID=501', \
'Label=UH0807_U1, CellID=8071, LogicRNCID=501', \
'Label=UH0602_U3, CellID=6023, LogicRNCID=501', \
'Label=UH0961_U2, CellID=9612, LogicRNCID=501', \
'Label=UH2932_U3, CellID=29323, LogicRNCID=501', \
'Label=UH1908_U3, CellID=19083, LogicRNCID=501', \
'Label=UH0706_U1, CellID=7061, LogicRNCID=501', \
'Label=UH2984_U3, CellID=29843, LogicRNCID=501', \
'Label=UH0708_U3, CellID=7083, LogicRNCID=501', \
'Label=UH0708_U2, CellID=7082, LogicRNCID=501', \
'Label=UH0708_U1, CellID=7081, LogicRNCID=501', \
'Label=UH1964_U2, CellID=19642, LogicRNCID=501', \
'Label=UH1964_U1, CellID=19641, LogicRNCID=501', \
'Label=UH3922_U1, CellID=39221, LogicRNCID=501', \
'Label=UH0880_U3, CellID=8803, LogicRNCID=501', \
'Label=UH0880_U1, CellID=8801, LogicRNCID=501', \
'Label=UH0980_U3, CellID=9803, LogicRNCID=501', \
'Label=UH0980_U2, CellID=9802, LogicRNCID=501', \
'Label=UH0980_U1, CellID=9801, LogicRNCID=501', \
'Label=UH1905_U2, CellID=19052, LogicRNCID=501', \
'Label=UH0881_U2, CellID=8812, LogicRNCID=501', \
'Label=UH1905_U1, CellID=19051, LogicRNCID=501', \
'Label=UH0881_U1, CellID=8811, LogicRNCID=501', \
'Label=UH1702_U2, CellID=17022, LogicRNCID=501', \
'Label=UH1702_U1, CellID=17021, LogicRNCID=501', \
'Label=UH0601_U3, CellID=6013, LogicRNCID=501', \
'Label=UH0601_U2, CellID=6012, LogicRNCID=501', \
'Label=UH0601_U1, CellID=6011, LogicRNCID=501', \
'Label=UH1801_U3, CellID=18013, LogicRNCID=501', \
'Label=UH0854_U1, CellID=8541, LogicRNCID=501', \
'Label=UH3928_U2, CellID=39282, LogicRNCID=501', \
'Label=UH0600_U2, CellID=6002, LogicRNCID=501', \
'Label=UH3928_U1, CellID=39281, LogicRNCID=501', \
'Label=UH0622_U2, CellID=6222, LogicRNCID=501', \
'Label=UH0622_U1, CellID=6221, LogicRNCID=501', \
'Label=UH0879_U2, CellID=8792, LogicRNCID=501', \
'Label=UH0623_U2, CellID=6232, LogicRNCID=501', \
'Label=UH0623_U1, CellID=6231, LogicRNCID=501', \
'Label=UH1958_U3, CellID=19583, LogicRNCID=501', \
'Label=UH1907_U1, CellID=19071, LogicRNCID=501', \
'Label=UH1702_U3, CellID=17023, LogicRNCID=501', \
'Label=UH1958_U2, CellID=19582, LogicRNCID=501', \
'Label=UH1928_U2, CellID=19282, LogicRNCID=501', \
'Label=UH1928_U1, CellID=19281, LogicRNCID=501', \
'Label=UH0955_U3, CellID=9553, LogicRNCID=501', \
'Label=UH2950_U1, CellID=29501, LogicRNCID=501', \
'Label=UH0825_U3, CellID=8253, LogicRNCID=501', \
'Label=UH0825_U2, CellID=8252, LogicRNCID=501', \
'Label=UH3925_U3, CellID=39253, LogicRNCID=501', \
'Label=UH0981_U3, CellID=9813, LogicRNCID=501', \
'Label=UH3928_U3, CellID=39283, LogicRNCID=501', \
'Label=UH2981_U1, CellID=29811, LogicRNCID=501', \
'Label=UH1854_U2, CellID=18542, LogicRNCID=501', \
'Label=UH0830_U2, CellID=8302, LogicRNCID=501', \
'Label=UH1854_U1, CellID=18541, LogicRNCID=501', \
'Label=UH0881_U3, CellID=8813, LogicRNCID=501', \
'Label=UH0830_U1, CellID=8301, LogicRNCID=501', \
'Label=UH3925_U2, CellID=39252, LogicRNCID=501', \
'Label=UH0981_U2, CellID=9812, LogicRNCID=501', \
'Label=UH3925_U1, CellID=39251, LogicRNCID=501', \
'Label=UH1928_U3, CellID=19283, LogicRNCID=501', \
'Label=UH0981_U1, CellID=9811, LogicRNCID=501', \
'Label=UH1951_U2, CellID=19512, LogicRNCID=501', \
'Label=UH1951_U1, CellID=19511, LogicRNCID=501', \
'Label=UH3922_U2, CellID=39222, LogicRNCID=501', \
'Label=UH1801_U1, CellID=18011, LogicRNCID=501', \
'Label=UH2981_U3, CellID=29813, LogicRNCID=501', \
'Label=UH1906_U1, CellID=19061, LogicRNCID=501', \
'Label=UH2981_U2, CellID=29812, LogicRNCID=501', \
'Label=UH0854_U3, CellID=8543, LogicRNCID=501', \
'Label=UH0854_U2, CellID=8542, LogicRNCID=501', \
'Label=UH1854_U3, CellID=18543, LogicRNCID=501', \
'Label=UH1803_U1, CellID=18031, LogicRNCID=501', \
'Label=UH0830_U3, CellID=8303, LogicRNCID=501', \
'Label=UH0825_U1, CellID=8251, LogicRNCID=501', \
'Label=UH1951_U3, CellID=19513, LogicRNCID=501', \
'Label=UH2950_U3, CellID=29503, LogicRNCID=501', \
'Label=UH2950_U2, CellID=29502, LogicRNCID=501', \
'Label=UH1906_U3, CellID=19063, LogicRNCID=501', \
'Label=UH0831_U1, CellID=8311, LogicRNCID=501', \
'Label=UH1906_U2, CellID=19062, LogicRNCID=501', \
'Label=UH0955_U2, CellID=9552, LogicRNCID=501', \
'Label=UH0955_U1, CellID=9551, LogicRNCID=501', \
'Label=UH0622_U3, CellID=6223, LogicRNCID=501', \
'Label=UH0879_U3, CellID=8793, LogicRNCID=501', \
'Label=UH0623_U3, CellID=6233, LogicRNCID=501', \
'Label=UH1803_U3, CellID=18033, LogicRNCID=501', \
'Label=UH0600_U1, CellID=6001, LogicRNCID=501', \
'Label=UH1803_U2, CellID=18032, LogicRNCID=501', \
'Label=UH0831_U3, CellID=8313, LogicRNCID=501', \
'Label=UH0831_U2, CellID=8312, LogicRNCID=501', \
'Label=UH0748_U3, CellID=7483, LogicRNCID=501', \
'Label=UH0748_U2, CellID=7482, LogicRNCID=501', \
'Label=UH0748_U1, CellID=7481, LogicRNCID=501'
]  # кластер число активных сот 235
list_F2_10637 = [
'Label=UH1945_U5, CellID=19455, LogicRNCID=501', \
'Label=UH1945_U4, CellID=19454, LogicRNCID=501', \
'Label=UH0972_U6, CellID=9726, LogicRNCID=501', \
'Label=UH0844_U6, CellID=8446, LogicRNCID=501', \
'Label=UH0972_U5, CellID=9725, LogicRNCID=501', \
'Label=UH0844_U5, CellID=8445, LogicRNCID=501', \
'Label=UH0972_U4, CellID=9724, LogicRNCID=501', \
'Label=UH0844_U4, CellID=8444, LogicRNCID=501', \
'Label=UH2763_U6, CellID=27636, LogicRNCID=501', \
'Label=UH2763_U5, CellID=27635, LogicRNCID=501', \
'Label=UH2763_U4, CellID=27634, LogicRNCID=501', \
'Label=UH1918_U6, CellID=19186, LogicRNCID=501', \
'Label=UH0880_U14, CellID=65521, LogicRNCID=501', \
'Label=UH1918_U5, CellID=19185, LogicRNCID=501', \
'Label=UH1918_U4, CellID=19184, LogicRNCID=501', \
'Label=UH1892_U6, CellID=18926, LogicRNCID=501', \
'Label=UH1892_U5, CellID=18925, LogicRNCID=501', \
'Label=UH0970_U6, CellID=9706, LogicRNCID=501', \
'Label=UH0842_U6, CellID=8426, LogicRNCID=501', \
'Label=UH0714_U6, CellID=7146, LogicRNCID=501', \
'Label=UH0970_U5, CellID=9705, LogicRNCID=501', \
'Label=UH0842_U5, CellID=8425, LogicRNCID=501', \
'Label=UH0714_U5, CellID=7145, LogicRNCID=501', \
'Label=UH0970_U4, CellID=9704, LogicRNCID=501', \
'Label=UH0842_U4, CellID=8424, LogicRNCID=501', \
'Label=UH0765_U6, CellID=7656, LogicRNCID=501', \
'Label=UH0714_U4, CellID=7144, LogicRNCID=501', \
'Label=UH0765_U5, CellID=7655, LogicRNCID=501', \
'Label=UH0765_U4, CellID=7654, LogicRNCID=501', \
'Label=UH0995_U6, CellID=9956, LogicRNCID=501', \
'Label=UH0611_U6, CellID=6116, LogicRNCID=501', \
'Label=UH0995_U5, CellID=9955, LogicRNCID=501', \
'Label=UH3990_U6, CellID=39906, LogicRNCID=501', \
'Label=UH0995_U4, CellID=9954, LogicRNCID=501', \
'Label=UH0611_U4, CellID=6114, LogicRNCID=501', \
'Label=UH3990_U5, CellID=39905, LogicRNCID=501', \
'Label=UH3990_U4, CellID=39904, LogicRNCID=501', \
'Label=UH0841_U6, CellID=8416, LogicRNCID=501', \
'Label=UH0841_U5, CellID=8415, LogicRNCID=501', \
'Label=UH1916_U6, CellID=19166, LogicRNCID=501', \
'Label=UH0841_U4, CellID=8414, LogicRNCID=501', \
'Label=UH0636_U6, CellID=6366, LogicRNCID=501', \
'Label=UH0508_U6, CellID=5086, LogicRNCID=501', \
'Label=UH1916_U5, CellID=19165, LogicRNCID=501', \
'Label=UH0636_U5, CellID=6365, LogicRNCID=501', \
'Label=UH0508_U5, CellID=5085, LogicRNCID=501', \
'Label=UH1916_U4, CellID=19164, LogicRNCID=501', \
'Label=UH0994_U6, CellID=9946, LogicRNCID=501', \
'Label=UH0738_U6, CellID=7386, LogicRNCID=501', \
'Label=UH0610_U6, CellID=6106, LogicRNCID=501', \
'Label=UH0994_U5, CellID=9945, LogicRNCID=501', \
'Label=UH0738_U5, CellID=7385, LogicRNCID=501', \
'Label=UH0610_U5, CellID=6105, LogicRNCID=501', \
'Label=UH3989_U6, CellID=39896, LogicRNCID=501', \
'Label=UH2709_U6, CellID=27096, LogicRNCID=501', \
'Label=UH0994_U4, CellID=9944, LogicRNCID=501', \
'Label=UH0738_U4, CellID=7384, LogicRNCID=501', \
'Label=UH3989_U5, CellID=39895, LogicRNCID=501', \
'Label=UH2709_U5, CellID=27095, LogicRNCID=501', \
'Label=UH1907_U33, CellID=65494, LogicRNCID=501', \
'Label=UH3989_U4, CellID=39894, LogicRNCID=501', \
'Label=UH2709_U4, CellID=27094, LogicRNCID=501', \
'Label=UH0840_U6, CellID=8406, LogicRNCID=501', \
'Label=UH1907_U22, CellID=65493, LogicRNCID=501', \
'Label=UH0840_U5, CellID=8405, LogicRNCID=501', \
'Label=UH1907_U11, CellID=65492, LogicRNCID=501', \
'Label=UH0840_U4, CellID=8404, LogicRNCID=501', \
'Label=UH0763_U6, CellID=7636, LogicRNCID=501', \
'Label=UH0763_U5, CellID=7635, LogicRNCID=501', \
'Label=UH0763_U4, CellID=7634, LogicRNCID=501', \
'Label=UH2913_U6, CellID=29136, LogicRNCID=501', \
'Label=UH2913_U5, CellID=29135, LogicRNCID=501', \
'Label=UH2913_U4, CellID=29134, LogicRNCID=501', \
'Label=UH0839_U6, CellID=8396, LogicRNCID=501', \
'Label=UH0711_U6, CellID=7116, LogicRNCID=501', \
'Label=UH0839_U5, CellID=8395, LogicRNCID=501', \
'Label=UH0711_U5, CellID=7115, LogicRNCID=501', \
'Label=UH2938_U6, CellID=29386, LogicRNCID=501', \
'Label=UH0839_U4, CellID=8394, LogicRNCID=501', \
'Label=UH0711_U4, CellID=7114, LogicRNCID=501', \
'Label=UH2938_U5, CellID=29385, LogicRNCID=501', \
'Label=UH2938_U4, CellID=29384, LogicRNCID=501', \
'Label=UH1939_U6, CellID=19396, LogicRNCID=501', \
'Label=UH1939_U5, CellID=19395, LogicRNCID=501', \
'Label=UH1939_U4, CellID=19394, LogicRNCID=501', \
'Label=UH0966_U6, CellID=9666, LogicRNCID=501', \
'Label=UH0838_U6, CellID=8386, LogicRNCID=501', \
'Label=UH0966_U5, CellID=9665, LogicRNCID=501', \
'Label=UH0838_U5, CellID=8385, LogicRNCID=501', \
'Label=UH0966_U4, CellID=9664, LogicRNCID=501', \
'Label=UH0838_U4, CellID=8384, LogicRNCID=501', \
'Label=UH1958_U5, CellID=19585, LogicRNCID=501', \
'Label=UH1702_U5, CellID=17025, LogicRNCID=501', \
'Label=UH1702_U4, CellID=17024, LogicRNCID=501', \
'Label=UH0988_U4, CellID=9884, LogicRNCID=501', \
'Label=UH1935_U5, CellID=19355, LogicRNCID=501', \
'Label=UH1703_U4, CellID=17034, LogicRNCID=501', \
'Label=UH0807_U4, CellID=8074, LogicRNCID=501', \
'Label=UH0602_U6, CellID=6026, LogicRNCID=501', \
'Label=UH1882_U5, CellID=18825, LogicRNCID=501', \
'Label=UH1833_U6, CellID=18336, LogicRNCID=501', \
'Label=UH1705_U6, CellID=17056, LogicRNCID=501', \
'Label=UH0758_U4, CellID=7584, LogicRNCID=501', \
'Label=UH1833_U5, CellID=18335, LogicRNCID=501', \
'Label=UH1705_U5, CellID=17055, LogicRNCID=501', \
'Label=UH1964_U4, CellID=19644, LogicRNCID=501', \
'Label=UH0962_U5, CellID=9625, LogicRNCID=501', \
'Label=UH0962_U4, CellID=9624, LogicRNCID=501', \
'Label=UH0706_U4, CellID=7064, LogicRNCID=501', \
'Label=UH0734_U4, CellID=7344, LogicRNCID=501', \
'Label=UH0708_U6, CellID=7086, LogicRNCID=501', \
'Label=UH1703_U5, CellID=17035, LogicRNCID=501', \
'Label=UH0807_U5, CellID=8075, LogicRNCID=501', \
'Label=UH2983_U4, CellID=29834, LogicRNCID=501', \
'Label=UH1882_U6, CellID=18826, LogicRNCID=501', \
'Label=UH1964_U5, CellID=19645, LogicRNCID=501', \
'Label=UH0961_U4, CellID=9614, LogicRNCID=501', \
'Label=UH2932_U5, CellID=29325, LogicRNCID=501', \
'Label=UH1908_U5, CellID=19085, LogicRNCID=501', \
'Label=UH2983_U6, CellID=29836, LogicRNCID=501', \
'Label=UH2932_U4, CellID=29324, LogicRNCID=501', \
'Label=UH0960_U5, CellID=9605, LogicRNCID=501', \
'Label=UH0832_U5, CellID=8325, LogicRNCID=501', \
'Label=UH1984_U4, CellID=19844, LogicRNCID=501', \
'Label=UH1907_U6, CellID=19076, LogicRNCID=501', \
'Label=UH0960_U4, CellID=9604, LogicRNCID=501', \
'Label=UH0734_U5, CellID=7345, LogicRNCID=501', \
'Label=UH0963_U4, CellID=9634, LogicRNCID=501', \
'Label=UH0835_U4, CellID=8354, LogicRNCID=501', \
'Label=UH0758_U6, CellID=7586, LogicRNCID=501', \
'Label=UH0758_U5, CellID=7585, LogicRNCID=501', \
'Label=UH0988_U5, CellID=9885, LogicRNCID=501', \
'Label=UH1935_U6, CellID=19356, LogicRNCID=501', \
'Label=UH0961_U6, CellID=9616, LogicRNCID=501', \
'Label=UH0961_U5, CellID=9615, LogicRNCID=501', \
'Label=UH2932_U6, CellID=29326, LogicRNCID=501', \
'Label=UH1908_U6, CellID=19086, LogicRNCID=501', \
'Label=UH1958_U6, CellID=19586, LogicRNCID=501', \
'Label=UH1907_U4, CellID=19074, LogicRNCID=501', \
'Label=UH1702_U6, CellID=17026, LogicRNCID=501', \
'Label=UH3996_U4, CellID=39964, LogicRNCID=501', \
'Label=UH3921_U6, CellID=39216, LogicRNCID=501', \
'Label=UH3995_U4, CellID=39954, LogicRNCID=501', \
'Label=UH1947_U4, CellID=19474, LogicRNCID=501', \
'Label=UH0974_U5, CellID=9745, LogicRNCID=501', \
'Label=UH0742_U4, CellID=7424, LogicRNCID=501', \
'Label=UH3920_U4, CellID=39204, LogicRNCID=501', \
'Label=UH0743_U6, CellID=7436, LogicRNCID=501', \
'Label=UH0743_U5, CellID=7435, LogicRNCID=501', \
'Label=UH0747_U5, CellID=7475, LogicRNCID=501', \
'Label=UH0973_U6, CellID=9736, LogicRNCID=501', \
'Label=UH0973_U5, CellID=9735, LogicRNCID=501', \
'Label=UH0617_U5, CellID=6175, LogicRNCID=501', \
'Label=UH3996_U6, CellID=39966, LogicRNCID=501', \
'Label=UH0770_U4, CellID=7704, LogicRNCID=501', \
'Label=UH0845_U4, CellID=8454, LogicRNCID=501', \
'Label=UH0717_U4, CellID=7174, LogicRNCID=501', \
'Label=UH0973_U4, CellID=9734, LogicRNCID=501', \
'Label=UH0717_U5, CellID=7175, LogicRNCID=501', \
'Label=UH0823_U5, CellID=8235, LogicRNCID=501', \
'Label=UH0823_U4, CellID=8234, LogicRNCID=501', \
'Label=UH3995_U6, CellID=39956, LogicRNCID=501', \
'Label=UH3995_U5, CellID=39955, LogicRNCID=501', \
'Label=UH1947_U5, CellID=19475, LogicRNCID=501', \
'Label=UH1947_U6, CellID=19476, LogicRNCID=501', \
'Label=UH0822_U4, CellID=8224, LogicRNCID=501', \
'Label=UH0617_U6, CellID=6176, LogicRNCID=501', \
'Label=UH0950_U4, CellID=9504, LogicRNCID=501', \
'Label=UH0618_U5, CellID=6185, LogicRNCID=501', \
'Label=UH0618_U6, CellID=6186, LogicRNCID=501', \
'Label=UH0951_U4, CellID=9514, LogicRNCID=501', \
'Label=UH0770_U6, CellID=7706, LogicRNCID=501', \
'Label=UH0770_U5, CellID=7705, LogicRNCID=501', \
'Label=UH0742_U5, CellID=7425, LogicRNCID=501', \
'Label=UH0950_U6, CellID=9506, LogicRNCID=501', \
'Label=UH0822_U5, CellID=8225, LogicRNCID=501', \
'Label=UH0849_U5, CellID=8495, LogicRNCID=501', \
'Label=UH3921_U4, CellID=39214, LogicRNCID=501', \
'Label=UH0849_U4, CellID=8494, LogicRNCID=501', \
'Label=UH3921_U5, CellID=39215, LogicRNCID=501', \
'Label=UH0950_U5, CellID=9505, LogicRNCID=501', \
'Label=UH0822_U6, CellID=8226, LogicRNCID=501', \
'Label=UH0820_U5, CellID=8205, LogicRNCID=501', \
'Label=UH1945_U6, CellID=19456, LogicRNCID=501', \
'Label=UH0823_U6, CellID=8236, LogicRNCID=501', \
'Label=UH0951_U6, CellID=9516, LogicRNCID=501', \
'Label=UH0974_U4, CellID=9744, LogicRNCID=501', \
'Label=UH3918_U4, CellID=39184, LogicRNCID=501', \
'Label=UH0821_U4, CellID=8214, LogicRNCID=501', \
'Label=UH0821_U5, CellID=8215, LogicRNCID=501', \
'Label=UH0821_U6, CellID=8216, LogicRNCID=501', \
'Label=UH0617_U4, CellID=6174, LogicRNCID=501', \
'Label=UH0618_U4, CellID=6184, LogicRNCID=501', \
'Label=UH0845_U5, CellID=8455, LogicRNCID=501', \
'Label=UH0717_U6, CellID=7176, LogicRNCID=501', \
'Label=UH0845_U6, CellID=8456, LogicRNCID=501', \
'Label=UH0743_U4, CellID=7434, LogicRNCID=501', \
'Label=UH0820_U4, CellID=8204, LogicRNCID=501', \
'Label=UH3920_U5, CellID=39205, LogicRNCID=501', \
'Label=UH0974_U6, CellID=9746, LogicRNCID=501', \
'Label=UH3918_U6, CellID=39186, LogicRNCID=501', \
'Label=UH0849_U6, CellID=8496, LogicRNCID=501', \
'Label=UH3996_U5, CellID=39965, LogicRNCID=501', \
'Label=UH2984_U6, CellID=29846, LogicRNCID=501', \
'Label=UH1960_U6, CellID=19606, LogicRNCID=501', \
'Label=UH1704_U6, CellID=17046, LogicRNCID=501', \
'Label=UH2984_U5, CellID=29845, LogicRNCID=501', \
'Label=UH1960_U5, CellID=19605, LogicRNCID=501', \
'Label=UH1704_U5, CellID=17045, LogicRNCID=501', \
'Label=UH1707_U4, CellID=17074, LogicRNCID=501', \
'Label=UH0734_U6, CellID=7346, LogicRNCID=501', \
'Label=UH1935_U4, CellID=19354, LogicRNCID=501', \
'Label=UH0962_U6, CellID=9626, LogicRNCID=501', \
'Label=UH0602_U5, CellID=6025, LogicRNCID=501', \
'Label=UH1882_U4, CellID=18824, LogicRNCID=501', \
'Label=UH1805_U6, CellID=18056, LogicRNCID=501', \
'Label=UH1936_U5, CellID=19365, LogicRNCID=501', \
'Label=UH1936_U4, CellID=19364, LogicRNCID=501', \
'Label=UH0963_U6, CellID=9636, LogicRNCID=501', \
'Label=UH0835_U6, CellID=8356, LogicRNCID=501', \
'Label=UH0963_U5, CellID=9635, LogicRNCID=501', \
'Label=UH0835_U5, CellID=8355, LogicRNCID=501', \
'Label=UH0832_U4, CellID=8324, LogicRNCID=501', \
'Label=UH1907_U5, CellID=19075, LogicRNCID=501', \
'Label=UH1805_U4, CellID=18054, LogicRNCID=501', \
'Label=UH0960_U6, CellID=9606, LogicRNCID=501', \
'Label=UH0832_U6, CellID=8326, LogicRNCID=501', \
'Label=UH1984_U5, CellID=19845, LogicRNCID=501', \
'Label=UH0965_U4, CellID=9654, LogicRNCID=501', \
'Label=UH4600_U5, CellID=46005, LogicRNCID=501', \
'Label=UH4600_U4, CellID=46004, LogicRNCID=501', \
'Label=UH1707_U6, CellID=17076, LogicRNCID=501', \
'Label=UH1707_U5, CellID=17075, LogicRNCID=501', \
'Label=UH1833_U4, CellID=18334, LogicRNCID=501', \
'Label=UH1705_U4, CellID=17054, LogicRNCID=501', \
'Label=UH0988_U6, CellID=9886, LogicRNCID=501', \
'Label=UH0602_U4, CellID=6024, LogicRNCID=501', \
'Label=UH1805_U5, CellID=18055, LogicRNCID=501', \
'Label=UH1908_U4, CellID=19084, LogicRNCID=501', \
'Label=UH1703_U6, CellID=17036, LogicRNCID=501', \
'Label=UH0807_U6, CellID=8076, LogicRNCID=501', \
'Label=UH2983_U5, CellID=29835, LogicRNCID=501', \
'Label=UH2984_U4, CellID=29844, LogicRNCID=501', \
'Label=UH1960_U4, CellID=19604, LogicRNCID=501', \
'Label=UH1704_U4, CellID=17044, LogicRNCID=501', \
'Label=UH0708_U5, CellID=7085, LogicRNCID=501', \
'Label=UH0708_U4, CellID=7084, LogicRNCID=501', \
'Label=UH1936_U6, CellID=19366, LogicRNCID=501', \
'Label=UH0965_U6, CellID=9656, LogicRNCID=501', \
'Label=UH0965_U5, CellID=9655, LogicRNCID=501', \
'Label=UH4600_U6, CellID=46006, LogicRNCID=501', \
'Label=UH0952_U6, CellID=9526, LogicRNCID=501', \
'Label=UH1976_U5, CellID=19765, LogicRNCID=501', \
'Label=UH0952_U5, CellID=9525, LogicRNCID=501', \
'Label=UH1976_U4, CellID=19764, LogicRNCID=501', \
'Label=UH0952_U4, CellID=9524, LogicRNCID=501', \
'Label=UH0854_U6, CellID=8546, LogicRNCID=501', \
'Label=UH0854_U5, CellID=8545, LogicRNCID=501', \
'Label=UH1801_U6, CellID=18016, LogicRNCID=501', \
'Label=UH0854_U4, CellID=8544, LogicRNCID=501', \
'Label=UH0601_U4, CellID=6014, LogicRNCID=501', \
'Label=UH0831_U6, CellID=8316, LogicRNCID=501', \
'Label=UH0831_U5, CellID=8315, LogicRNCID=501', \
'Label=UH0879_U6, CellID=8796, LogicRNCID=501', \
'Label=UH0623_U6, CellID=6236, LogicRNCID=501', \
'Label=UH1854_U6, CellID=18546, LogicRNCID=501', \
'Label=UH1803_U4, CellID=18034, LogicRNCID=501', \
'Label=UH0830_U6, CellID=8306, LogicRNCID=501', \
'Label=UH1854_U5, CellID=18545, LogicRNCID=501', \
'Label=UH0980_U5, CellID=9805, LogicRNCID=501', \
'Label=UH0980_U4, CellID=9804, LogicRNCID=501', \
'Label=UH0981_U6, CellID=9816, LogicRNCID=501', \
'Label=UH3925_U5, CellID=39255, LogicRNCID=501', \
'Label=UH0981_U5, CellID=9815, LogicRNCID=501', \
'Label=UH0601_U6, CellID=6016, LogicRNCID=501', \
'Label=UH0601_U5, CellID=6015, LogicRNCID=501', \
'Label=UH0955_U5, CellID=9555, LogicRNCID=501', \
'Label=UH0955_U4, CellID=9554, LogicRNCID=501', \
'Label=UH0622_U6, CellID=6226, LogicRNCID=501', \
'Label=UH0622_U5, CellID=6225, LogicRNCID=501', \
'Label=UH0825_U4, CellID=8254, LogicRNCID=501', \
'Label=UH1951_U6, CellID=19516, LogicRNCID=501', \
'Label=UH3925_U4, CellID=39254, LogicRNCID=501', \
'Label=UH1928_U6, CellID=19286, LogicRNCID=501', \
'Label=UH0981_U4, CellID=9814, LogicRNCID=501', \
'Label=UH1928_U5, CellID=19285, LogicRNCID=501', \
'Label=UH0600_U4, CellID=6004, LogicRNCID=501', \
'Label=UH1803_U5, CellID=18035, LogicRNCID=501', \
'Label=UH1905_U4, CellID=19054, LogicRNCID=501', \
'Label=UH0881_U4, CellID=8814, LogicRNCID=501', \
'Label=UH0880_U6, CellID=8806, LogicRNCID=501', \
'Label=UH0880_U4, CellID=8804, LogicRNCID=501', \
'Label=UH1928_U4, CellID=19284, LogicRNCID=501', \
'Label=UH0955_U6, CellID=9556, LogicRNCID=501', \
'Label=UH3922_U4, CellID=39224, LogicRNCID=501', \
'Label=UH1976_U6, CellID=19766, LogicRNCID=501', \
'Label=UH0879_U5, CellID=8795, LogicRNCID=501', \
'Label=UH0623_U5, CellID=6235, LogicRNCID=501', \
'Label=UH3928_U5, CellID=39285, LogicRNCID=501', \
'Label=UH0600_U5, CellID=6005, LogicRNCID=501', \
'Label=UH3928_U4, CellID=39284, LogicRNCID=501', \
'Label=UH1803_U6, CellID=18036, LogicRNCID=501', \
'Label=UH1801_U4, CellID=18014, LogicRNCID=501', \
'Label=UH0830_U4, CellID=8304, LogicRNCID=501', \
'Label=UH1905_U5, CellID=19055, LogicRNCID=501', \
'Label=UH0881_U5, CellID=8815, LogicRNCID=501', \
'Label=UH1951_U5, CellID=19515, LogicRNCID=501', \
'Label=UH1951_U4, CellID=19514, LogicRNCID=501', \
'Label=UH3922_U5, CellID=39225, LogicRNCID=501', \
'Label=UH2950_U4, CellID=29504, LogicRNCID=501', \
'Label=UH0825_U6, CellID=8256, LogicRNCID=501', \
'Label=UH0825_U5, CellID=8255, LogicRNCID=501', \
'Label=UH2981_U5, CellID=29815, LogicRNCID=501', \
'Label=UH3928_U6, CellID=39286, LogicRNCID=501', \
'Label=UH2981_U4, CellID=29814, LogicRNCID=501', \
'Label=UH2950_U6, CellID=29506, LogicRNCID=501', \
'Label=UH2950_U5, CellID=29505, LogicRNCID=501', \
'Label=UH0622_U4, CellID=6224, LogicRNCID=501', \
'Label=UH0980_U6, CellID=9806, LogicRNCID=501', \
'Label=UH0623_U4, CellID=6234, LogicRNCID=501', \
'Label=UH3925_U6, CellID=39256, LogicRNCID=501', \
'Label=UH0830_U5, CellID=8305, LogicRNCID=501', \
'Label=UH1854_U4, CellID=18544, LogicRNCID=501', \
'Label=UH0881_U6, CellID=8816, LogicRNCID=501', \
'Label=UH1906_U6, CellID=19066, LogicRNCID=501', \
'Label=UH0831_U4, CellID=8314, LogicRNCID=501', \
'Label=UH1906_U5, CellID=19065, LogicRNCID=501', \
'Label=UH2981_U6, CellID=29816, LogicRNCID=501', \
'Label=UH1906_U4, CellID=19064, LogicRNCID=501', \
'Label=UH0748_U5, CellID=7485, LogicRNCID=501', \
'Label=UH0748_U6, CellID=7486, LogicRNCID=501', \
'Label=UH0748_U4, CellID=7484, LogicRNCID=501'
]  # кластер число активных сот 236
list_F3_2937 = [
'Label=UH0761_U94, CellID=47614, LogicRNCID=501', \
'Label=UH0735_U96, CellID=47356, LogicRNCID=501', \
'Label=UH0735_U95, CellID=47355, LogicRNCID=501', \
'Label=UH0735_U94, CellID=47354, LogicRNCID=501', \
'Label=UH0965_U96, CellID=49656, LogicRNCID=501', \
'Label=UH0965_U95, CellID=49655, LogicRNCID=501', \
'Label=UH0965_U94, CellID=49654, LogicRNCID=501', \
'Label=UH0734_U96, CellID=47346, LogicRNCID=501', \
'Label=UH0734_U95, CellID=47345, LogicRNCID=501', \
'Label=UH0734_U94, CellID=47344, LogicRNCID=501', \
'Label=UH0708_U96, CellID=47086, LogicRNCID=501', \
'Label=UH0708_U95, CellID=47085, LogicRNCID=501', \
'Label=UH0759_U96, CellID=47596, LogicRNCID=501', \
'Label=UH0708_U94, CellID=47084, LogicRNCID=501', \
'Label=UH0759_U95, CellID=47595, LogicRNCID=501', \
'Label=UH1936_U96, CellID=59366, LogicRNCID=501', \
'Label=UH1936_U95, CellID=59365, LogicRNCID=501', \
'Label=UH0963_U96, CellID=49636, LogicRNCID=501', \
'Label=UH0835_U96, CellID=48356, LogicRNCID=501', \
'Label=UH0963_U95, CellID=49635, LogicRNCID=501', \
'Label=UH0835_U95, CellID=48355, LogicRNCID=501', \
'Label=UH0963_U94, CellID=49634, LogicRNCID=501', \
'Label=UH0835_U94, CellID=48354, LogicRNCID=501', \
'Label=UH0758_U96, CellID=47586, LogicRNCID=501', \
'Label=UH0758_U95, CellID=47585, LogicRNCID=501', \
'Label=UH0758_U94, CellID=47584, LogicRNCID=501', \
'Label=UH0732_U96, CellID=47326, LogicRNCID=501', \
'Label=UH0732_U95, CellID=47325, LogicRNCID=501', \
'Label=UH0732_U94, CellID=47324, LogicRNCID=501', \
'Label=UH0962_U96, CellID=49626, LogicRNCID=501', \
'Label=UH0706_U96, CellID=47066, LogicRNCID=501', \
'Label=UH0962_U95, CellID=49625, LogicRNCID=501', \
'Label=UH0706_U95, CellID=47065, LogicRNCID=501', \
'Label=UH0962_U94, CellID=49624, LogicRNCID=501', \
'Label=UH0706_U94, CellID=47064, LogicRNCID=501', \
'Label=UH0731_U96, CellID=47316, LogicRNCID=501', \
'Label=UH0731_U95, CellID=47315, LogicRNCID=501', \
'Label=UH0731_U94, CellID=47314, LogicRNCID=501', \
'Label=UH0833_U96, CellID=48336, LogicRNCID=501', \
'Label=UH0705_U96, CellID=47056, LogicRNCID=501', \
'Label=UH0833_U95, CellID=48335, LogicRNCID=501', \
'Label=UH0705_U95, CellID=47055, LogicRNCID=501', \
'Label=UH0833_U94, CellID=48334, LogicRNCID=501', \
'Label=UH0756_U96, CellID=47566, LogicRNCID=501', \
'Label=UH0705_U94, CellID=47054, LogicRNCID=501', \
'Label=UH0807_U96, CellID=48076, LogicRNCID=501', \
'Label=UH0756_U94, CellID=47564, LogicRNCID=501', \
'Label=UH0807_U95, CellID=48075, LogicRNCID=501', \
'Label=UH0807_U94, CellID=48074, LogicRNCID=501', \
'Label=UH0602_U96, CellID=46026, LogicRNCID=501', \
'Label=UH0602_U95, CellID=46025, LogicRNCID=501', \
'Label=UH0602_U94, CellID=46024, LogicRNCID=501', \
'Label=UH0960_U96, CellID=49606, LogicRNCID=501', \
'Label=UH0832_U96, CellID=48326, LogicRNCID=501', \
'Label=UH0704_U96, CellID=47046, LogicRNCID=501', \
'Label=UH0960_U95, CellID=49605, LogicRNCID=501', \
'Label=UH0832_U95, CellID=48325, LogicRNCID=501', \
'Label=UH0704_U95, CellID=47045, LogicRNCID=501', \
'Label=UH0960_U94, CellID=49604, LogicRNCID=501', \
'Label=UH0832_U94, CellID=48324, LogicRNCID=501', \
'Label=UH0601_U96, CellID=46016, LogicRNCID=501', \
'Label=UH0825_U96, CellID=48256, LogicRNCID=501', \
'Label=UH0623_U95, CellID=46235, LogicRNCID=501', \
'Label=UH0760_U97, CellID=50074, LogicRNCID=501', \
'Label=UH1945_U95, CellID=50057, LogicRNCID=501', \
'Label=UH1801_U96, CellID=58016, LogicRNCID=501', \
'Label=UH0831_U96, CellID=48316, LogicRNCID=501', \
'Label=UH0703_U96, CellID=47036, LogicRNCID=501', \
'Label=UH0831_U95, CellID=48315, LogicRNCID=501', \
'Label=UH0703_U95, CellID=47035, LogicRNCID=501', \
'Label=UH0831_U94, CellID=48314, LogicRNCID=501', \
'Label=UH0760_U95, CellID=50072, LogicRNCID=501', \
'Label=UH0760_U94, CellID=50071, LogicRNCID=501', \
'Label=UH0881_U95, CellID=48815, LogicRNCID=501', \
'Label=UH0881_U94, CellID=48814, LogicRNCID=501', \
'Label=UH1945_U96, CellID=50058, LogicRNCID=501', \
'Label=UH2709_U95, CellID=50066, LogicRNCID=501', \
'Label=UH0955_U94, CellID=49554, LogicRNCID=501', \
'Label=UH0750_U96, CellID=47506, LogicRNCID=501', \
'Label=UH0622_U96, CellID=46226, LogicRNCID=501', \
'Label=UH0601_U95, CellID=46015, LogicRNCID=501', \
'Label=UH0601_U94, CellID=46014, LogicRNCID=501', \
'Label=UH0723_U96, CellID=47236, LogicRNCID=501', \
'Label=UH2709_U94, CellID=50065, LogicRNCID=501', \
'Label=UH0750_U95, CellID=47505, LogicRNCID=501', \
'Label=UH0622_U95, CellID=46225, LogicRNCID=501', \
'Label=UH1951_U96, CellID=50064, LogicRNCID=501', \
'Label=UH0702_U95, CellID=47025, LogicRNCID=501', \
'Label=UH0881_U96, CellID=48816, LogicRNCID=501', \
'Label=UH0830_U94, CellID=48304, LogicRNCID=501', \
'Label=UH0702_U94, CellID=47024, LogicRNCID=501', \
'Label=UH0751_U96, CellID=47516, LogicRNCID=501', \
'Label=UH0623_U96, CellID=46236, LogicRNCID=501', \
'Label=UH0750_U94, CellID=47504, LogicRNCID=501', \
'Label=UH0622_U94, CellID=46224, LogicRNCID=501', \
'Label=UH1951_U95, CellID=50063, LogicRNCID=501', \
'Label=UH1951_U94, CellID=50062, LogicRNCID=501', \
'Label=UH0840_U95, CellID=48405, LogicRNCID=501', \
'Label=UH0840_U94, CellID=48404, LogicRNCID=501', \
'Label=UH0763_U96, CellID=47636, LogicRNCID=501', \
'Label=UH0879_U96, CellID=55060, LogicRNCID=501', \
'Label=UH0879_U95, CellID=55061, LogicRNCID=501', \
'Label=UH0636_U95, CellID=46365, LogicRNCID=501', \
'Label=UH0766_U95, CellID=47665, LogicRNCID=501', \
'Label=UH0611_U96, CellID=46116, LogicRNCID=501', \
'Label=UH0739_U95, CellID=47395, LogicRNCID=501', \
'Label=UH0611_U95, CellID=46115, LogicRNCID=501', \
'Label=UH0739_U96, CellID=47396, LogicRNCID=501', \
'Label=UH0839_U96, CellID=48396, LogicRNCID=501', \
'Label=UH0843_U95, CellID=48435, LogicRNCID=501', \
'Label=UH0766_U94, CellID=47666, LogicRNCID=501', \
'Label=UH0736_U94, CellID=47364, LogicRNCID=501', \
'Label=UH0765_U96, CellID=47656, LogicRNCID=501', \
'Label=UH0713_U94, CellID=47134, LogicRNCID=501', \
'Label=UH0738_U94, CellID=47384, LogicRNCID=501', \
'Label=UH0840_U96, CellID=48406, LogicRNCID=501', \
'Label=UH0736_U96, CellID=47366, LogicRNCID=501', \
'Label=UH0736_U95, CellID=47365, LogicRNCID=501', \
'Label=UH0714_U95, CellID=47145, LogicRNCID=501', \
'Label=UH0763_U94, CellID=47634, LogicRNCID=501', \
'Label=UH0763_U95, CellID=47635, LogicRNCID=501', \
'Label=UH0713_U96, CellID=47136, LogicRNCID=501', \
'Label=UH0841_U95, CellID=48415, LogicRNCID=501', \
'Label=UH0713_U95, CellID=47135, LogicRNCID=501', \
'Label=UH0738_U96, CellID=47386, LogicRNCID=501', \
'Label=UH0738_U95, CellID=47385, LogicRNCID=501', \
'Label=UH0838_U95, CellID=48385, LogicRNCID=501', \
'Label=UH0761_U96, CellID=47616, LogicRNCID=501', \
'Label=UH0737_U94, CellID=47374, LogicRNCID=501', \
'Label=UH0611_U94, CellID=46114, LogicRNCID=501', \
'Label=UH0740_U95, CellID=47405, LogicRNCID=501', \
'Label=UH0740_U96, CellID=47406, LogicRNCID=501', \
'Label=UH0841_U96, CellID=48416, LogicRNCID=501', \
'Label=UH0838_U94, CellID=48384, LogicRNCID=501', \
'Label=UH0838_U96, CellID=48386, LogicRNCID=501', \
'Label=UH0740_U94, CellID=47404, LogicRNCID=501', \
'Label=UH0737_U95, CellID=47375, LogicRNCID=501', \
'Label=UH0879_U94, CellID=55062, LogicRNCID=501', \
'Label=UH0636_U96, CellID=46366, LogicRNCID=501', \
'Label=UH0841_U94, CellID=48414, LogicRNCID=501', \
'Label=UH0765_U94, CellID=47654, LogicRNCID=501', \
'Label=UH0714_U94, CellID=47144, LogicRNCID=501', \
'Label=UH0715_U94, CellID=47154, LogicRNCID=501', \
'Label=UH0843_U94, CellID=48434, LogicRNCID=501', \
'Label=UH0715_U95, CellID=47155, LogicRNCID=501', \
'Label=UH0839_U94, CellID=48394, LogicRNCID=501', \
'Label=UH0839_U95, CellID=48395, LogicRNCID=501', \
'Label=UH0737_U96, CellID=47376, LogicRNCID=501', \
'Label=UH0766_U96, CellID=47664, LogicRNCID=501', \
'Label=UH0636_U94, CellID=46364, LogicRNCID=501', \
'Label=UH1928_U95, CellID=50049, LogicRNCID=501', \
'Label=UH0830_U96, CellID=48306, LogicRNCID=501', \
'Label=UH0702_U96, CellID=47026, LogicRNCID=501', \
'Label=UH0830_U95, CellID=48305, LogicRNCID=501', \
'Label=UH0751_U94, CellID=47514, LogicRNCID=501', \
'Label=UH0623_U94, CellID=46234, LogicRNCID=501', \
'Label=UH0760_U96, CellID=50073, LogicRNCID=501', \
'Label=UH1945_U94, CellID=50056, LogicRNCID=501', \
'Label=UH0752_U94, CellID=47524, LogicRNCID=501', \
'Label=UH0723_U95, CellID=47235, LogicRNCID=501', \
'Label=UH1928_U96, CellID=50050, LogicRNCID=501', \
'Label=UH0723_U94, CellID=47234, LogicRNCID=501', \
'Label=UH1801_U94, CellID=58014, LogicRNCID=501', \
'Label=UH0955_U96, CellID=49556, LogicRNCID=501', \
'Label=UH2709_U96, CellID=50067, LogicRNCID=501', \
'Label=UH0955_U95, CellID=49555, LogicRNCID=501', \
'Label=UH0752_U96, CellID=47526, LogicRNCID=501', \
'Label=UH0752_U95, CellID=47525, LogicRNCID=501', \
'Label=UH0703_U94, CellID=47034, LogicRNCID=501', \
'Label=UH0754_U95, CellID=47545, LogicRNCID=501', \
'Label=UH0754_U94, CellID=47544, LogicRNCID=501', \
'Label=UH0843_U96, CellID=48436, LogicRNCID=501', \
'Label=UH0715_U96, CellID=47156, LogicRNCID=501', \
'Label=UH1803_U94, CellID=50019, LogicRNCID=501', \
'Label=UH0745_U96, CellID=47456, LogicRNCID=501', \
'Label=UH0973_U96, CellID=49736, LogicRNCID=501', \
'Label=UH0845_U96, CellID=48456, LogicRNCID=501', \
'Label=UH0717_U96, CellID=47176, LogicRNCID=501', \
'Label=UH0973_U95, CellID=49735, LogicRNCID=501', \
'Label=UH0845_U95, CellID=48455, LogicRNCID=501', \
'Label=UH0951_U96, CellID=49516, LogicRNCID=501', \
'Label=UH0823_U96, CellID=48236, LogicRNCID=501', \
'Label=UH1805_U96, CellID=50027, LogicRNCID=501', \
'Label=UH0823_U95, CellID=48235, LogicRNCID=501', \
'Label=UH1604_U94, CellID=50013, LogicRNCID=501', \
'Label=UH1964_U95, CellID=55388, LogicRNCID=501', \
'Label=UH1603_U96, CellID=50012, LogicRNCID=501', \
'Label=UH1964_U94, CellID=55387, LogicRNCID=501', \
'Label=UH1882_U96, CellID=50033, LogicRNCID=501', \
'Label=UH0820_U95, CellID=48205, LogicRNCID=501', \
'Label=UH0820_U94, CellID=48204, LogicRNCID=501', \
'Label=UH0743_U96, CellID=47436, LogicRNCID=501', \
'Label=UH0743_U95, CellID=47435, LogicRNCID=501', \
'Label=UH0821_U95, CellID=48215, LogicRNCID=501', \
'Label=UH0821_U94, CellID=48214, LogicRNCID=501', \
'Label=UH0825_U95, CellID=48255, LogicRNCID=501', \
'Label=UH0825_U94, CellID=48254, LogicRNCID=501', \
'Label=UH0844_U95, CellID=48445, LogicRNCID=501', \
'Label=UH0844_U94, CellID=48444, LogicRNCID=501', \
'Label=UH0767_U96, CellID=47676, LogicRNCID=501', \
'Label=UH0744_U96, CellID=47446, LogicRNCID=501', \
'Label=UH0744_U95, CellID=47445, LogicRNCID=501', \
'Label=UH1902_U95, CellID=50035, LogicRNCID=501', \
'Label=UH0747_U95, CellID=47475, LogicRNCID=501', \
'Label=UH1902_U94, CellID=50034, LogicRNCID=501', \
'Label=UH0747_U94, CellID=47474, LogicRNCID=501', \
'Label=UH0849_U94, CellID=48494, LogicRNCID=501', \
'Label=UH1804_U96, CellID=50024, LogicRNCID=501', \
'Label=UH0618_U94, CellID=46184, LogicRNCID=501', \
'Label=UH1804_U95, CellID=50023, LogicRNCID=501', \
'Label=UH1804_U94, CellID=50022, LogicRNCID=501', \
'Label=UH1803_U96, CellID=50021, LogicRNCID=501', \
'Label=UH1803_U95, CellID=50020, LogicRNCID=501', \
'Label=UH0744_U94, CellID=47444, LogicRNCID=501', \
'Label=UH0869_U96, CellID=48696, LogicRNCID=501', \
'Label=UH0869_U95, CellID=48695, LogicRNCID=501', \
'Label=UH0869_U94, CellID=48694, LogicRNCID=501', \
'Label=UH1603_U95, CellID=50011, LogicRNCID=501', \
'Label=UH1603_U94, CellID=50010, LogicRNCID=501', \
'Label=UH0770_U96, CellID=47706, LogicRNCID=501', \
'Label=UH1902_U96, CellID=50036, LogicRNCID=501', \
'Label=UH1604_U96, CellID=50015, LogicRNCID=501', \
'Label=UH0745_U95, CellID=47455, LogicRNCID=501', \
'Label=UH1604_U95, CellID=50014, LogicRNCID=501', \
'Label=UH0745_U94, CellID=47454, LogicRNCID=501', \
'Label=UH1964_U96, CellID=55389, LogicRNCID=501', \
'Label=UH0717_U95, CellID=47175, LogicRNCID=501', \
'Label=UH0973_U94, CellID=49734, LogicRNCID=501', \
'Label=UH0845_U94, CellID=48454, LogicRNCID=501', \
'Label=UH0717_U94, CellID=47174, LogicRNCID=501', \
'Label=UH0646_U96, CellID=51012, LogicRNCID=501', \
'Label=UH0646_U95, CellID=51011, LogicRNCID=501', \
'Label=UH0646_U94, CellID=51010, LogicRNCID=501', \
'Label=UH0742_U96, CellID=47426, LogicRNCID=501', \
'Label=UH1882_U94, CellID=50031, LogicRNCID=501', \
'Label=UH0849_U95, CellID=48495, LogicRNCID=501', \
'Label=UH0767_U95, CellID=47675, LogicRNCID=501', \
'Label=UH0767_U94, CellID=47674, LogicRNCID=501', \
'Label=UH0614_U94, CellID=46144, LogicRNCID=501', \
'Label=UH1602_U94, CellID=51007, LogicRNCID=501', \
'Label=UH0844_U96, CellID=48446, LogicRNCID=501', \
'Label=UH1907_U95, CellID=50038, LogicRNCID=501', \
'Label=UH1907_U94, CellID=50037, LogicRNCID=501', \
'Label=UH1805_U95, CellID=50026, LogicRNCID=501', \
'Label=UH0951_U94, CellID=49514, LogicRNCID=501', \
'Label=UH0823_U94, CellID=48234, LogicRNCID=501', \
'Label=UH0746_U96, CellID=47466, LogicRNCID=501', \
'Label=UH0618_U96, CellID=46186, LogicRNCID=501', \
'Label=UH1805_U94, CellID=50025, LogicRNCID=501', \
'Label=UH0746_U95, CellID=47465, LogicRNCID=501', \
'Label=UH0618_U95, CellID=46185, LogicRNCID=501', \
'Label=UH0614_U96, CellID=46146, LogicRNCID=501', \
'Label=UH1602_U96, CellID=51009, LogicRNCID=501', \
'Label=UH0742_U95, CellID=47425, LogicRNCID=501', \
'Label=UH0614_U95, CellID=46145, LogicRNCID=501', \
'Label=UH1602_U95, CellID=51008, LogicRNCID=501', \
'Label=UH0742_U94, CellID=47424, LogicRNCID=501', \
'Label=UH0820_U96, CellID=48206, LogicRNCID=501', \
'Label=UH0770_U95, CellID=47705, LogicRNCID=501', \
'Label=UH0821_U96, CellID=48216, LogicRNCID=501', \
'Label=UH0770_U94, CellID=47704, LogicRNCID=501', \
'Label=UH1882_U95, CellID=50032, LogicRNCID=501', \
'Label=UH0849_U96, CellID=48496, LogicRNCID=501', \
'Label=UH1907_U96, CellID=50039, LogicRNCID=501', \
'Label=UH0711_U96, CellID=47116, LogicRNCID=501', \
'Label=UH0711_U95, CellID=47115, LogicRNCID=501', \
'Label=UH0711_U94, CellID=47114, LogicRNCID=501', \
'Label=UH0748_U94, CellID=47484, LogicRNCID=501', \
'Label=UH0748_U95, CellID=47485, LogicRNCID=501', \
'Label=UH0748_U96, CellID=47486, LogicRNCID=501'
]  # кластер число активных сот 204
list_U2100N = [
'NodeB Function Name=UH3920, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH3920, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH3920, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1935, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1935, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1935, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1935, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1935, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1935, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH3920, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0844, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0844, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0844, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0844, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0844, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0844, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0845, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0845, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0845, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0845, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0845, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0845, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1803, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1803, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1803, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1803, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1803, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1803, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0734, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0734, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0734, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0734, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0734, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0734, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0763, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0763, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0763, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0763, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0763, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0763, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0770, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0770, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0770, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0770, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0770, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0770, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1704, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1704, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1704, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1704, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1704, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1704, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1984, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1984, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1984, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1984, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH3922, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH3922, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH3922, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH3922, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0738, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0738, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0738, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0738, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0738, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0738, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0831, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0831, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0831, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0831, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0831, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0831, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0960, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0960, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0960, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0960, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0960, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0960, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0961, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0961, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0961, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0961, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0961, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0961, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1892, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1892, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1892, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1892, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1905, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1905, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1905, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1905, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1960, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1960, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1960, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1960, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1960, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1960, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH2709, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH2709, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH2709, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH2709, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH2709, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH2709, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH2763, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH2763, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH2763, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH2763, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH2763, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH2763, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1702, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1702, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1702, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1702, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1702, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1702, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH3921, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH3921, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH3921, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH3921, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH3921, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH3921, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1707, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1707, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1707, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1707, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1707, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1707, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0508, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0508, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0508, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0508, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0600, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0600, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0600, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0600, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0611, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0611, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0611, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0611, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0636, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0636, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0636, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0636, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0747, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0747, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0830, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0830, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0830, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0830, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0830, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0830, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0832, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0832, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0832, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0832, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0832, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0832, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0849, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0849, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0849, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0849, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0849, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0849, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0980, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0980, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0980, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0980, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0980, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0980, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0963, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0963, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0963, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0963, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0963, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0963, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0974, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0974, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0974, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0974, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0974, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0974, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1805, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1805, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1805, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1805, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1805, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1805, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1906, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1906, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1906, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1906, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1906, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1906, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1964, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1964, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1964, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1964, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH2932, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH2932, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH2932, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH2932, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH2932, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH2932, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH2984, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH2984, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH2984, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH2984, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH2984, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH2984, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH3928, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH3928, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH3928, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH3928, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH3928, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH3928, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH3996, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH3996, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH3996, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH3996, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH3996, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH3996, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0711, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0711, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0711, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0711, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0711, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0711, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0842, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0842, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0842, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0842, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0842, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0842, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1928, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1928, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1928, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1928, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1928, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1928, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0708, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0708, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0708, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0708, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0708, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0708, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0952, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0952, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0952, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0952, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0952, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0952, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0807, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0807, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0807, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0807, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0807, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0807, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0838, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0838, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0838, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0838, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0838, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0838, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0841, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0841, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0841, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0841, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0841, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0841, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=15, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=14, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=10, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=11, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=12, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=13, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0622, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0622, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0622, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0622, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0622, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0622, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0742, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0742, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0742, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0742, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0743, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0743, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0743, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0743, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0743, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0743, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0950, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0950, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0950, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0950, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0950, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0950, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0981, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0981, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0981, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0981, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0981, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0981, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0610, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0610, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0610, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0610, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0610, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0610, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0623, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0623, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0623, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0623, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0623, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0623, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0717, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0717, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0717, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0717, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0717, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0717, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH2938, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH2938, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH2938, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH2938, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH2938, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH2938, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0840, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0840, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0840, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0840, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0840, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0840, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0951, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0951, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0951, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0951, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0962, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0962, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0962, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0962, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0962, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0962, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0966, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0966, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0966, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0966, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0966, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0966, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0972, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0972, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0972, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0972, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0972, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0972, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1703, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1703, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1703, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1703, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1703, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1703, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1801, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1801, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1801, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1801, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0988, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0988, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0988, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0988, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0988, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0988, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1936, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1936, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1936, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1936, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1936, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1936, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1939, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1939, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1939, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1939, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1939, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1939, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0765, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0765, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0765, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0765, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0822, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0822, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0822, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0822, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0822, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0822, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0973, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0973, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0973, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0973, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0973, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0973, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0994, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0994, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0994, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0994, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0994, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0994, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1854, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1854, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1854, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1854, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1854, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1854, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1705, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1705, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1705, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1705, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1705, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1705, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1916, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1916, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1916, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1916, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1916, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1916, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH3990, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH3990, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH3990, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH3990, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH3990, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH3990, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0601, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0601, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0601, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0601, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0601, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0601, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0602, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0602, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0602, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0602, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0602, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0602, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0617, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0617, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0617, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0617, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0617, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0617, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0618, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0618, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0618, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0618, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0618, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0618, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0714, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0714, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0714, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0714, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0714, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0714, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0758, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0758, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0758, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0758, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0758, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0758, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1908, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1908, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1908, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1908, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1908, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1908, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1951, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1951, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1951, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1951, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1951, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1951, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH1958, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1958, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1958, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1958, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1976, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH1976, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH1976, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH1976, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH1976, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH1976, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH3989, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH3989, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH3989, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH3989, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH3989, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH3989, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH3995, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH3995, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH3995, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH3995, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH3995, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH3995, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH4600, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH4600, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH4600, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH4600, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH4600, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH4600, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0748, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0748, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0748, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0748, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0748, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0748, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0880, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0880, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0880, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0880, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0880, Local Cell ID=1, Cell Name=CELLNAME', \
'NodeB Function Name=UH0880, Local Cell ID=0, Cell Name=CELLNAME', \
'NodeB Function Name=UH0879, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0879, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0879, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0879, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0766, Local Cell ID=6, Cell Name=CELLNAME', \
'NodeB Function Name=UH0766, Local Cell ID=5, Cell Name=CELLNAME', \
'NodeB Function Name=UH0766, Local Cell ID=4, Cell Name=CELLNAME', \
'NodeB Function Name=UH0766, Local Cell ID=3, Cell Name=CELLNAME', \
'NodeB Function Name=UH0766, Local Cell ID=2, Cell Name=CELLNAME', \
'NodeB Function Name=UH0766, Local Cell ID=1, Cell Name=CELLNAME']
list_U900N = ['NodeB Function Name=UH0744, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0744, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0744, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0760, Local Cell ID=97, Cell Name=CELLNAME', \
'NodeB Function Name=UH0760, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0760, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0760, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0737, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0737, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0737, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0752, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0752, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0752, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0843, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0843, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0843, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0844, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0844, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0844, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0845, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0845, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0845, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1803, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1803, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1803, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0702, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0702, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0702, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0715, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0715, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0715, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0734, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0734, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0734, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0770, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0770, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0770, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0745, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0745, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0745, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0703, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0703, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0703, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0738, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0738, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0738, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0831, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0831, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0831, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0960, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0960, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0960, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH2709, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH2709, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH2709, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0600, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0600, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0611, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0611, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0611, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0614, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0614, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0614, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0636, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0636, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0636, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0736, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0736, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0736, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0732, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0732, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0732, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0739, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0739, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0740, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0740, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0740, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0747, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0747, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0750, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0750, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0750, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0761, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0761, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0830, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0830, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0830, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0832, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0832, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0832, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0849, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0849, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0849, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0963, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0963, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0963, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1805, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1805, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1805, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1925, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1925, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1925, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1964, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1964, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1964, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0746, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0746, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0711, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0711, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0711, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0731, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0731, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0731, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0735, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0735, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0735, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1928, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1928, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0708, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0708, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0708, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0807, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0807, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0807, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0838, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0838, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0838, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0841, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0841, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0841, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1907, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0622, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0622, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0622, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0869, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0869, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0869, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0646, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0646, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0646, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0742, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0742, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0742, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0743, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0743, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0751, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0751, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0756, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0756, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1902, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1902, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1902, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0623, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0623, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0623, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0713, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0713, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0713, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0717, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0717, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0717, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0833, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0833, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0833, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0840, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0840, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0840, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0951, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0951, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0962, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0962, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0962, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1801, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1801, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1804, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1804, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1804, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1936, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1936, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0765, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0765, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0822, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0822, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0822, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0973, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0973, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0973, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0601, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0601, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0601, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0602, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0602, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0602, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0618, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0618, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0618, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0714, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0714, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0758, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0758, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0758, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1913, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1913, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1913, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1951, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1951, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1951, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0748, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0748, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0748, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0763, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0763, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0763, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0881, Local Cell ID=97, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=99, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=98, Cell Name=CELLNAME', \
'NodeB Function Name=UH0821, Local Cell ID=97, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=99, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=98, Cell Name=CELLNAME', \
'NodeB Function Name=UH0820, Local Cell ID=97, Cell Name=CELLNAME', \
'NodeB Function Name=UH0761, Local Cell ID=97, Cell Name=CELLNAME', \
'NodeB Function Name=UH0869, Local Cell ID=97, Cell Name=CELLNAME', \
'NodeB Function Name=UH0965, Local Cell ID=97, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=99, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=98, Cell Name=CELLNAME', \
'NodeB Function Name=UH1945, Local Cell ID=97, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=99, Cell Name=CELLNAME', \
'NodeB Function Name=UH0839, Local Cell ID=97, Cell Name=CELLNAME', \
'NodeB Function Name=UH0879, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0879, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0879, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0767, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0767, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0767, Local Cell ID=94, Cell Name=CELLNAME', \
'NodeB Function Name=UH0766, Local Cell ID=96, Cell Name=CELLNAME', \
'NodeB Function Name=UH0766, Local Cell ID=95, Cell Name=CELLNAME', \
'NodeB Function Name=UH0766, Local Cell ID=94, Cell Name=CELLNAME']

# фильтрация по кластеру:
cluster_UH1952cell = [
'Label=UH1952_U1, CellID=19521, LogicRNCID=501',
'Label=UH1952_U6, CellID=19526, LogicRNCID=501',
'Label=UH1952_U5, CellID=19525, LogicRNCID=501',
'Label=UH1952_U4, CellID=19524, LogicRNCID=501',
'Label=UH1952_U3, CellID=19523, LogicRNCID=501',
'Label=UH1952_U2, CellID=19522, LogicRNCID=501',
]
cluster_Empty = ['UH0960']  # пустой список
cluster_UH = [
'Label=UH1945_U5, CellID=19455, LogicRNCID=501',
'Label=UH0761_U94, CellID=47614, LogicRNCID=501',
'Label=UH1945_U4, CellID=19454, LogicRNCID=501',
'Label=UH0972_U6, CellID=9726, LogicRNCID=501',
'Label=UH0844_U6, CellID=8446, LogicRNCID=501',
'Label=UH0742_U2, CellID=7422, LogicRNCID=501',
'Label=UH1945_U3, CellID=19453, LogicRNCID=501',
'Label=UH0972_U5, CellID=9725, LogicRNCID=501',
'Label=UH0844_U5, CellID=8445, LogicRNCID=501',
'Label=UH0742_U1, CellID=7421, LogicRNCID=501',
'Label=UH0735_U96, CellID=47356, LogicRNCID=501',
'Label=UH1945_U2, CellID=19452, LogicRNCID=501',
'Label=UH0972_U4, CellID=9724, LogicRNCID=501',
'Label=UH0844_U4, CellID=8444, LogicRNCID=501',
'Label=UH0735_U95, CellID=47355, LogicRNCID=501',
'Label=UH1945_U1, CellID=19451, LogicRNCID=501',
'Label=UH0972_U3, CellID=9723, LogicRNCID=501',
'Label=UH0844_U3, CellID=8443, LogicRNCID=501',
'Label=UH0735_U94, CellID=47354, LogicRNCID=501',
'Label=UH0972_U2, CellID=9722, LogicRNCID=501',
'Label=UH0844_U2, CellID=8442, LogicRNCID=501',
'Label=UH0965_U97, CellID=49657, LogicRNCID=501',
'Label=UH0972_U1, CellID=9721, LogicRNCID=501',
'Label=UH0844_U1, CellID=8441, LogicRNCID=501',
'Label=UH0965_U96, CellID=49656, LogicRNCID=501',
'Label=UH0965_U95, CellID=49655, LogicRNCID=501',
'Label=UH0965_U94, CellID=49654, LogicRNCID=501',
'Label=UH2763_U6, CellID=27636, LogicRNCID=501',
'Label=UH2763_U5, CellID=27635, LogicRNCID=501',
'Label=UH0734_U96, CellID=47346, LogicRNCID=501',
'Label=UH2763_U4, CellID=27634, LogicRNCID=501',
'Label=UH1918_U6, CellID=19186, LogicRNCID=501',
'Label=UH0766_U6, CellID=7666, LogicRNCID=501',
'Label=UH0880_U14, CellID=65521, LogicRNCID=501',
'Label=UH0734_U95, CellID=47345, LogicRNCID=501',
'Label=UH2763_U3, CellID=27633, LogicRNCID=501',
'Label=UH1918_U5, CellID=19185, LogicRNCID=501',
'Label=UH0766_U5, CellID=7665, LogicRNCID=501',
'Label=UH0880_U11, CellID=65520, LogicRNCID=501',
'Label=UH0734_U94, CellID=47344, LogicRNCID=501',
'Label=UH2763_U2, CellID=27632, LogicRNCID=501',
'Label=UH1918_U4, CellID=19184, LogicRNCID=501',
'Label=UH0766_U4, CellID=7664, LogicRNCID=501',
'Label=UH2763_U1, CellID=27631, LogicRNCID=501',
'Label=UH1918_U3, CellID=19183, LogicRNCID=501',
'Label=UH0766_U3, CellID=7663, LogicRNCID=501',
'Label=UH0708_U96, CellID=47086, LogicRNCID=501',
'Label=UH1918_U2, CellID=19182, LogicRNCID=501',
'Label=UH1892_U6, CellID=18926, LogicRNCID=501',
'Label=UH0766_U2, CellID=7662, LogicRNCID=501',
'Label=UH0708_U95, CellID=47085, LogicRNCID=501',
'Label=UH1918_U1, CellID=19181, LogicRNCID=501',
'Label=UH1892_U5, CellID=18925, LogicRNCID=501',
'Label=UH0766_U1, CellID=7661, LogicRNCID=501',
'Label=UH0759_U96, CellID=47596, LogicRNCID=501',
'Label=UH0708_U94, CellID=47084, LogicRNCID=501',
'Label=UH0759_U95, CellID=47595, LogicRNCID=501',
'Label=UH1892_U3, CellID=18923, LogicRNCID=501',
'Label=UH1892_U2, CellID=18922, LogicRNCID=501',
'Label=UH0970_U6, CellID=9706, LogicRNCID=501',
'Label=UH0842_U6, CellID=8426, LogicRNCID=501',
'Label=UH0714_U6, CellID=7146, LogicRNCID=501',
'Label=UH0970_U5, CellID=9705, LogicRNCID=501',
'Label=UH0842_U5, CellID=8425, LogicRNCID=501',
'Label=UH0714_U5, CellID=7145, LogicRNCID=501',
'Label=UH0970_U4, CellID=9704, LogicRNCID=501',
'Label=UH0842_U4, CellID=8424, LogicRNCID=501',
'Label=UH0765_U6, CellID=7656, LogicRNCID=501',
'Label=UH0714_U4, CellID=7144, LogicRNCID=501',
'Label=UH0970_U3, CellID=9703, LogicRNCID=501',
'Label=UH0842_U3, CellID=8423, LogicRNCID=501',
'Label=UH0765_U5, CellID=7655, LogicRNCID=501',
'Label=UH0714_U3, CellID=7143, LogicRNCID=501',
'Label=UH1936_U96, CellID=59366, LogicRNCID=501',
'Label=UH0970_U2, CellID=9702, LogicRNCID=501',
'Label=UH0842_U2, CellID=8422, LogicRNCID=501',
'Label=UH0765_U4, CellID=7654, LogicRNCID=501',
'Label=UH0714_U2, CellID=7142, LogicRNCID=501',
'Label=UH1936_U95, CellID=59365, LogicRNCID=501',
'Label=UH0970_U1, CellID=9701, LogicRNCID=501',
'Label=UH0842_U1, CellID=8421, LogicRNCID=501',
'Label=UH0765_U3, CellID=7653, LogicRNCID=501',
'Label=UH0714_U1, CellID=7141, LogicRNCID=501',
'Label=UH0963_U96, CellID=49636, LogicRNCID=501',
'Label=UH0835_U96, CellID=48356, LogicRNCID=501',
'Label=UH0995_U6, CellID=9956, LogicRNCID=501',
'Label=UH0765_U2, CellID=7652, LogicRNCID=501',
'Label=UH0611_U6, CellID=6116, LogicRNCID=501',
'Label=UH0963_U95, CellID=49635, LogicRNCID=501',
'Label=UH0835_U95, CellID=48355, LogicRNCID=501',
'Label=UH0995_U5, CellID=9955, LogicRNCID=501',
'Label=UH0765_U1, CellID=7651, LogicRNCID=501',
'Label=UH0963_U94, CellID=49634, LogicRNCID=501',
'Label=UH0835_U94, CellID=48354, LogicRNCID=501',
'Label=UH0758_U96, CellID=47586, LogicRNCID=501',
'Label=UH3990_U6, CellID=39906, LogicRNCID=501',
'Label=UH0995_U4, CellID=9954, LogicRNCID=501',
'Label=UH0611_U4, CellID=6114, LogicRNCID=501',
'Label=UH0758_U95, CellID=47585, LogicRNCID=501',
'Label=UH3990_U5, CellID=39905, LogicRNCID=501',
'Label=UH0995_U3, CellID=9953, LogicRNCID=501',
'Label=UH0611_U3, CellID=6113, LogicRNCID=501',
'Label=UH0758_U94, CellID=47584, LogicRNCID=501',
'Label=UH3990_U4, CellID=39904, LogicRNCID=501',
'Label=UH0995_U2, CellID=9952, LogicRNCID=501',
'Label=UH0841_U6, CellID=8416, LogicRNCID=501',
'Label=UH3990_U3, CellID=39903, LogicRNCID=501',
'Label=UH0995_U1, CellID=9951, LogicRNCID=501',
'Label=UH0841_U5, CellID=8415, LogicRNCID=501',
'Label=UH0611_U1, CellID=6111, LogicRNCID=501',
'Label=UH0732_U96, CellID=47326, LogicRNCID=501',
'Label=UH3990_U2, CellID=39902, LogicRNCID=501',
'Label=UH1916_U6, CellID=19166, LogicRNCID=501',
'Label=UH0841_U4, CellID=8414, LogicRNCID=501',
'Label=UH0636_U6, CellID=6366, LogicRNCID=501',
'Label=UH0508_U6, CellID=5086, LogicRNCID=501',
'Label=UH0732_U95, CellID=47325, LogicRNCID=501',
'Label=UH3990_U1, CellID=39901, LogicRNCID=501',
'Label=UH1916_U5, CellID=19165, LogicRNCID=501',
'Label=UH0841_U3, CellID=8413, LogicRNCID=501',
'Label=UH0636_U5, CellID=6365, LogicRNCID=501',
'Label=UH0508_U5, CellID=5085, LogicRNCID=501',
'Label=UH0732_U94, CellID=47324, LogicRNCID=501',
'Label=UH1916_U4, CellID=19164, LogicRNCID=501',
'Label=UH0841_U2, CellID=8412, LogicRNCID=501',
'Label=UH1916_U3, CellID=19163, LogicRNCID=501',
'Label=UH0841_U1, CellID=8411, LogicRNCID=501',
'Label=UH0636_U3, CellID=6363, LogicRNCID=501',
'Label=UH0508_U3, CellID=5083, LogicRNCID=501',
'Label=UH0962_U96, CellID=49626, LogicRNCID=501',
'Label=UH0706_U96, CellID=47066, LogicRNCID=501',
'Label=UH1916_U2, CellID=19162, LogicRNCID=501',
'Label=UH0994_U6, CellID=9946, LogicRNCID=501',
'Label=UH0738_U6, CellID=7386, LogicRNCID=501',
'Label=UH0636_U2, CellID=6362, LogicRNCID=501',
'Label=UH0610_U6, CellID=6106, LogicRNCID=501',
'Label=UH0508_U2, CellID=5082, LogicRNCID=501',
'Label=UH1907_U66, CellID=65497, LogicRNCID=501',
'Label=UH0962_U95, CellID=49625, LogicRNCID=501',
'Label=UH0706_U95, CellID=47065, LogicRNCID=501',
'Label=UH1916_U1, CellID=19161, LogicRNCID=501',
'Label=UH0994_U5, CellID=9945, LogicRNCID=501',
'Label=UH0738_U5, CellID=7385, LogicRNCID=501',
'Label=UH0610_U5, CellID=6105, LogicRNCID=501',
'Label=UH1907_U55, CellID=65496, LogicRNCID=501',
'Label=UH0962_U94, CellID=49624, LogicRNCID=501',
'Label=UH0706_U94, CellID=47064, LogicRNCID=501',
'Label=UH3989_U6, CellID=39896, LogicRNCID=501',
'Label=UH2709_U6, CellID=27096, LogicRNCID=501',
'Label=UH0994_U4, CellID=9944, LogicRNCID=501',
'Label=UH0738_U4, CellID=7384, LogicRNCID=501',
'Label=UH1907_U44, CellID=65495, LogicRNCID=501',
'Label=UH3989_U5, CellID=39895, LogicRNCID=501',
'Label=UH2709_U5, CellID=27095, LogicRNCID=501',
'Label=UH0994_U3, CellID=9943, LogicRNCID=501',
'Label=UH0738_U3, CellID=7383, LogicRNCID=501',
'Label=UH0610_U3, CellID=6103, LogicRNCID=501',
'Label=UH1907_U33, CellID=65494, LogicRNCID=501',
'Label=UH3989_U4, CellID=39894, LogicRNCID=501',
'Label=UH2709_U4, CellID=27094, LogicRNCID=501',
'Label=UH0994_U2, CellID=9942, LogicRNCID=501',
'Label=UH0840_U6, CellID=8406, LogicRNCID=501',
'Label=UH0738_U2, CellID=7382, LogicRNCID=501',
'Label=UH0610_U2, CellID=6102, LogicRNCID=501',
'Label=UH1907_U22, CellID=65493, LogicRNCID=501',
'Label=UH3989_U3, CellID=39893, LogicRNCID=501',
'Label=UH2709_U3, CellID=27093, LogicRNCID=501',
'Label=UH0994_U1, CellID=9941, LogicRNCID=501',
'Label=UH0840_U5, CellID=8405, LogicRNCID=501',
'Label=UH0738_U1, CellID=7381, LogicRNCID=501',
'Label=UH1907_U11, CellID=65492, LogicRNCID=501',
'Label=UH0731_U96, CellID=47316, LogicRNCID=501',
'Label=UH3989_U2, CellID=39892, LogicRNCID=501',
'Label=UH2709_U2, CellID=27092, LogicRNCID=501',
'Label=UH0840_U4, CellID=8404, LogicRNCID=501',
'Label=UH0763_U6, CellID=7636, LogicRNCID=501',
'Label=UH0731_U95, CellID=47315, LogicRNCID=501',
'Label=UH3989_U1, CellID=39891, LogicRNCID=501',
'Label=UH2709_U1, CellID=27091, LogicRNCID=501',
'Label=UH0840_U3, CellID=8403, LogicRNCID=501',
'Label=UH0763_U5, CellID=7635, LogicRNCID=501',
'Label=UH0731_U94, CellID=47314, LogicRNCID=501',
'Label=UH0840_U2, CellID=8402, LogicRNCID=501',
'Label=UH0763_U4, CellID=7634, LogicRNCID=501',
'Label=UH0840_U1, CellID=8401, LogicRNCID=501',
'Label=UH0763_U3, CellID=7633, LogicRNCID=501',
'Label=UH0833_U96, CellID=48336, LogicRNCID=501',
'Label=UH0705_U96, CellID=47056, LogicRNCID=501',
'Label=UH2913_U6, CellID=29136, LogicRNCID=501',
'Label=UH0763_U2, CellID=7632, LogicRNCID=501',
'Label=UH0833_U95, CellID=48335, LogicRNCID=501',
'Label=UH0705_U95, CellID=47055, LogicRNCID=501',
'Label=UH2913_U5, CellID=29135, LogicRNCID=501',
'Label=UH0763_U1, CellID=7631, LogicRNCID=501',
'Label=UH0833_U94, CellID=48334, LogicRNCID=501',
'Label=UH0756_U96, CellID=47566, LogicRNCID=501',
'Label=UH0705_U94, CellID=47054, LogicRNCID=501',
'Label=UH2913_U4, CellID=29134, LogicRNCID=501',
'Label=UH2913_U3, CellID=29133, LogicRNCID=501',
'Label=UH0807_U96, CellID=48076, LogicRNCID=501',
'Label=UH0756_U94, CellID=47564, LogicRNCID=501',
'Label=UH2913_U2, CellID=29132, LogicRNCID=501',
'Label=UH0839_U6, CellID=8396, LogicRNCID=501',
'Label=UH0711_U6, CellID=7116, LogicRNCID=501',
'Label=UH0807_U95, CellID=48075, LogicRNCID=501',
'Label=UH2913_U1, CellID=29131, LogicRNCID=501',
'Label=UH0839_U5, CellID=8395, LogicRNCID=501',
'Label=UH0711_U5, CellID=7115, LogicRNCID=501',
'Label=UH0807_U94, CellID=48074, LogicRNCID=501',
'Label=UH0602_U96, CellID=46026, LogicRNCID=501',
'Label=UH2938_U6, CellID=29386, LogicRNCID=501',
'Label=UH0839_U4, CellID=8394, LogicRNCID=501',
'Label=UH0711_U4, CellID=7114, LogicRNCID=501',
'Label=UH0602_U95, CellID=46025, LogicRNCID=501',
'Label=UH2938_U5, CellID=29385, LogicRNCID=501',
'Label=UH0839_U3, CellID=8393, LogicRNCID=501',
'Label=UH0711_U3, CellID=7113, LogicRNCID=501',
'Label=UH0602_U94, CellID=46024, LogicRNCID=501',
'Label=UH2938_U4, CellID=29384, LogicRNCID=501',
'Label=UH0839_U2, CellID=8392, LogicRNCID=501',
'Label=UH0711_U2, CellID=7112, LogicRNCID=501',
'Label=UH2938_U3, CellID=29383, LogicRNCID=501',
'Label=UH0839_U1, CellID=8391, LogicRNCID=501',
'Label=UH0711_U1, CellID=7111, LogicRNCID=501',
'Label=UH0960_U96, CellID=49606, LogicRNCID=501',
'Label=UH0832_U96, CellID=48326, LogicRNCID=501',
'Label=UH0704_U96, CellID=47046, LogicRNCID=501',
'Label=UH2938_U2, CellID=29382, LogicRNCID=501',
'Label=UH0960_U95, CellID=49605, LogicRNCID=501',
'Label=UH0832_U95, CellID=48325, LogicRNCID=501',
'Label=UH0704_U95, CellID=47045, LogicRNCID=501',
'Label=UH2938_U1, CellID=29381, LogicRNCID=501',
'Label=UH0960_U94, CellID=49604, LogicRNCID=501',
'Label=UH0832_U94, CellID=48324, LogicRNCID=501',
'Label=UH1939_U6, CellID=19396, LogicRNCID=501',
'Label=UH1939_U5, CellID=19395, LogicRNCID=501',
'Label=UH1939_U4, CellID=19394, LogicRNCID=501',
'Label=UH0966_U6, CellID=9666, LogicRNCID=501',
'Label=UH0838_U6, CellID=8386, LogicRNCID=501',
'Label=UH1939_U3, CellID=19393, LogicRNCID=501',
'Label=UH0961_U6, CellID=9616, LogicRNCID=501',
'Label=UH1951_U95, CellID=50063, LogicRNCID=501',
'Label=UH0961_U5, CellID=9615, LogicRNCID=501',
'Label=UH1951_U94, CellID=50062, LogicRNCID=501',
'Label=UH2932_U6, CellID=29326, LogicRNCID=501',
'Label=UH1908_U6, CellID=19086, LogicRNCID=501',
'Label=UH0961_U4, CellID=9614, LogicRNCID=501',
'Label=UH1945_U99, CellID=50061, LogicRNCID=501',
'Label=UH1907_U4, CellID=19074, LogicRNCID=501',
'Label=UH1702_U6, CellID=17026, LogicRNCID=501',
'Label=UH0960_U2, CellID=9602, LogicRNCID=501',
'Label=UH0832_U2, CellID=8322, LogicRNCID=501',
'Label=UH1928_U95, CellID=50049, LogicRNCID=501',
'Label=UH1984_U1, CellID=19841, LogicRNCID=501',
'Label=UH1958_U5, CellID=19585, LogicRNCID=501',
'Label=UH1907_U3, CellID=19073, LogicRNCID=501',
'Label=UH1907_U5, CellID=19075, LogicRNCID=501',
'Label=UH1805_U1, CellID=18051, LogicRNCID=501',
'Label=UH0960_U3, CellID=9603, LogicRNCID=501',
'Label=UH0832_U3, CellID=8323, LogicRNCID=501',
'Label=UH1928_U96, CellID=50050, LogicRNCID=501',
'Label=UH0723_U94, CellID=47234, LogicRNCID=501',
'Label=UH1984_U2, CellID=19842, LogicRNCID=501',
'Label=UH1958_U6, CellID=19586, LogicRNCID=501',
'Label=UH0966_U3, CellID=9663, LogicRNCID=501',
'Label=UH0838_U3, CellID=8383, LogicRNCID=501',
'Label=UH0601_U94, CellID=46014, LogicRNCID=501',
'Label=UH0966_U2, CellID=9662, LogicRNCID=501',
'Label=UH0838_U2, CellID=8382, LogicRNCID=501',
'Label=UH1964_U5, CellID=19645, LogicRNCID=501',
'Label=UH0966_U1, CellID=9661, LogicRNCID=501',
'Label=UH0838_U1, CellID=8381, LogicRNCID=501',
'Label=UH0602_U1, CellID=6021, LogicRNCID=501',
'Label=UH0723_U96, CellID=47236, LogicRNCID=501',
'Label=UH1984_U4, CellID=19844, LogicRNCID=501',
'Label=UH1907_U6, CellID=19076, LogicRNCID=501',
'Label=UH1805_U2, CellID=18052, LogicRNCID=501',
'Label=UH0960_U4, CellID=9604, LogicRNCID=501',
'Label=UH0832_U4, CellID=8324, LogicRNCID=501',
'Label=UH0723_U95, CellID=47235, LogicRNCID=501',
'Label=UH1707_U1, CellID=17071, LogicRNCID=501',
'Label=UH0734_U3, CellID=7343, LogicRNCID=501',
'Label=UH0881_U94, CellID=48814, LogicRNCID=501',
'Label=UH0734_U2, CellID=7342, LogicRNCID=501',
'Label=UH0708_U6, CellID=7086, LogicRNCID=501',
'Label=UH0734_U1, CellID=7341, LogicRNCID=501',
'Label=UH0708_U5, CellID=7085, LogicRNCID=501',
'Label=UH0708_U4, CellID=7084, LogicRNCID=501',
'Label=UH0960_U6, CellID=9606, LogicRNCID=501',
'Label=UH0832_U6, CellID=8326, LogicRNCID=501',
'Label=UH0602_U2, CellID=6022, LogicRNCID=501',
'Label=UH1984_U5, CellID=19845, LogicRNCID=501',
'Label=UH1882_U1, CellID=18821, LogicRNCID=501',
'Label=UH1805_U3, CellID=18053, LogicRNCID=501',
'Label=UH0960_U5, CellID=9605, LogicRNCID=501',
'Label=UH0832_U5, CellID=8325, LogicRNCID=501',
'Label=UH0966_U5, CellID=9665, LogicRNCID=501',
'Label=UH0838_U5, CellID=8385, LogicRNCID=501',
'Label=UH0601_U96, CellID=46016, LogicRNCID=501',
'Label=UH1939_U2, CellID=19392, LogicRNCID=501',
'Label=UH0966_U4, CellID=9664, LogicRNCID=501',
'Label=UH0838_U4, CellID=8384, LogicRNCID=501',
'Label=UH0601_U95, CellID=46015, LogicRNCID=501',
'Label=UH1939_U1, CellID=19391, LogicRNCID=501',
'Label=UH1705_U6, CellID=17056, LogicRNCID=501',
'Label=UH0963_U2, CellID=9632, LogicRNCID=501',
'Label=UH0835_U2, CellID=8352, LogicRNCID=501',
'Label=UH0758_U4, CellID=7584, LogicRNCID=501',
'Label=UH1833_U5, CellID=18335, LogicRNCID=501',
'Label=UH1705_U5, CellID=17055, LogicRNCID=501',
'Label=UH0963_U1, CellID=9631, LogicRNCID=501',
'Label=UH0835_U1, CellID=8351, LogicRNCID=501',
'Label=UH0623_U95, CellID=46235, LogicRNCID=501',
'Label=UH1935_U5, CellID=19355, LogicRNCID=501',
'Label=UH1833_U1, CellID=18331, LogicRNCID=501',
'Label=UH1705_U1, CellID=17051, LogicRNCID=501',
'Label=UH0988_U3, CellID=9883, LogicRNCID=501',
'Label=UH0760_U97, CellID=50074, LogicRNCID=501',
'Label=UH0751_U94, CellID=47514, LogicRNCID=501',
'Label=UH0623_U94, CellID=46234, LogicRNCID=501',
'Label=UH2984_U4, CellID=29844, LogicRNCID=501',
'Label=UH1960_U4, CellID=19604, LogicRNCID=501',
'Label=UH1704_U4, CellID=17044, LogicRNCID=501',
'Label=UH2709_U96, CellID=50067, LogicRNCID=501',
'Label=UH0955_U95, CellID=49555, LogicRNCID=501',
'Label=UH2984_U3, CellID=29843, LogicRNCID=501',
'Label=UH1960_U3, CellID=19603, LogicRNCID=501',
'Label=UH1704_U3, CellID=17043, LogicRNCID=501',
'Label=UH1707_U3, CellID=17073, LogicRNCID=501',
'Label=UH0734_U5, CellID=7345, LogicRNCID=501',
'Label=UH0881_U96, CellID=48816, LogicRNCID=501',
'Label=UH0830_U94, CellID=48304, LogicRNCID=501',
'Label=UH0702_U94, CellID=47024, LogicRNCID=501',
'Label=UH1707_U2, CellID=17072, LogicRNCID=501',
'Label=UH0734_U4, CellID=7344, LogicRNCID=501',
'Label=UH0881_U95, CellID=48815, LogicRNCID=501',
'Label=UH2983_U1, CellID=29831, LogicRNCID=501',
'Label=UH1882_U3, CellID=18823, LogicRNCID=501',
'Label=UH1805_U5, CellID=18055, LogicRNCID=501',
'Label=UH1703_U1, CellID=17031, LogicRNCID=501',
'Label=UH0807_U1, CellID=8071, LogicRNCID=501',
'Label=UH0602_U3, CellID=6023, LogicRNCID=501',
'Label=UH1882_U2, CellID=18822, LogicRNCID=501',
'Label=UH1805_U4, CellID=18054, LogicRNCID=501',
'Label=UH1935_U2, CellID=19352, LogicRNCID=501',
'Label=UH0962_U4, CellID=9624, LogicRNCID=501',
'Label=UH0706_U4, CellID=7064, LogicRNCID=501',
'Label=UH0760_U94, CellID=50071, LogicRNCID=501',
'Label=UH1935_U1, CellID=19351, LogicRNCID=501',
'Label=UH0962_U3, CellID=9623, LogicRNCID=501',
'Label=UH2984_U6, CellID=29846, LogicRNCID=501',
'Label=UH1960_U6, CellID=19606, LogicRNCID=501',
'Label=UH0961_U2, CellID=9612, LogicRNCID=501',
'Label=UH0807_U6, CellID=8076, LogicRNCID=501',
'Label=UH1945_U97, CellID=50059, LogicRNCID=501',
'Label=UH2983_U5, CellID=29835, LogicRNCID=501',
'Label=UH2932_U3, CellID=29323, LogicRNCID=501',
'Label=UH1908_U3, CellID=19083, LogicRNCID=501',
'Label=UH1703_U5, CellID=17035, LogicRNCID=501',
'Label=UH0961_U1, CellID=9611, LogicRNCID=501',
'Label=UH0835_U4, CellID=8354, LogicRNCID=501',
'Label=UH0758_U6, CellID=7586, LogicRNCID=501',
'Label=UH1936_U1, CellID=19361, LogicRNCID=501',
'Label=UH0963_U3, CellID=9633, LogicRNCID=501',
'Label=UH0835_U3, CellID=8353, LogicRNCID=501',
'Label=UH0758_U5, CellID=7585, LogicRNCID=501',
'Label=UH1801_U96, CellID=58016, LogicRNCID=501',
'Label=UH1833_U6, CellID=18336, LogicRNCID=501',
'Label=UH0702_U96, CellID=47026, LogicRNCID=501',
'Label=UH4600_U2, CellID=46002, LogicRNCID=501',
'Label=UH1707_U4, CellID=17074, LogicRNCID=501',
'Label=UH0734_U6, CellID=7346, LogicRNCID=501',
'Label=UH0881_U97, CellID=48817, LogicRNCID=501',
'Label=UH0830_U95, CellID=48305, LogicRNCID=501',
'Label=UH0702_U95, CellID=47025, LogicRNCID=501',
'Label=UH4600_U1, CellID=46001, LogicRNCID=501',
'Label=UH0602_U6, CellID=6026, LogicRNCID=501',
'Label=UH1945_U95, CellID=50057, LogicRNCID=501',
'Label=UH2983_U3, CellID=29833, LogicRNCID=501',
'Label=UH2932_U1, CellID=29321, LogicRNCID=501',
'Label=UH1908_U1, CellID=19081, LogicRNCID=501',
'Label=UH1882_U5, CellID=18825, LogicRNCID=501',
'Label=UH1703_U3, CellID=17033, LogicRNCID=501',
'Label=UH0807_U3, CellID=8073, LogicRNCID=501',
'Label=UH0602_U5, CellID=6025, LogicRNCID=501',
'Label=UH1945_U94, CellID=50056, LogicRNCID=501',
'Label=UH2983_U2, CellID=29832, LogicRNCID=501',
'Label=UH1882_U4, CellID=18824, LogicRNCID=501',
'Label=UH0748_U5, CellID=7485, LogicRNCID=501',
'Label=UH0767_U96, CellID=47676, LogicRNCID=501',
'Label=UH0825_U2, CellID=8252, LogicRNCID=501',
'Label=UH0767_U95, CellID=47675, LogicRNCID=501',
'Label=UH0742_U4, CellID=7424, LogicRNCID=501',
'Label=UH1951_U5, CellID=19515, LogicRNCID=501',
'Label=UH0748_U4, CellID=7484, LogicRNCID=501',
'Label=UH1951_U6, CellID=19516, LogicRNCID=501',
'Label=UH0849_U6, CellID=8496, LogicRNCID=501',
'Label=UH3921_U5, CellID=39215, LogicRNCID=501',
'Label=UH0740_U96, CellID=47406, LogicRNCID=501',
'Label=UH0849_U4, CellID=8494, LogicRNCID=501',
'Label=UH0717_U1, CellID=7171, LogicRNCID=501',
'Label=UH0838_U95, CellID=48385, LogicRNCID=501',
'Label=UH0742_U5, CellID=7425, LogicRNCID=501',
'Label=UH0761_U96, CellID=47616, LogicRNCID=501',
'Label=UH0825_U5, CellID=8255, LogicRNCID=501',
'Label=UH2950_U2, CellID=29502, LogicRNCID=501',
'Label=UH0748_U6, CellID=7486, LogicRNCID=501',
'Label=UH2950_U1, CellID=29501, LogicRNCID=501',
'Label=UH0765_U94, CellID=47654, LogicRNCID=501',
'Label=UH3920_U5, CellID=39205, LogicRNCID=501',
'Label=UH0739_U96, CellID=47396, LogicRNCID=501',
'Label=UH3920_U4, CellID=39204, LogicRNCID=501',
'Label=UH0822_U3, CellID=8223, LogicRNCID=501',
'Label=UH0841_U94, CellID=48414, LogicRNCID=501',
'Label=UH0636_U96, CellID=46366, LogicRNCID=501',
'Label=UH0950_U2, CellID=9502, LogicRNCID=501',
'Label=UH0821_U4, CellID=8214, LogicRNCID=501',
'Label=UH0879_U95, CellID=55061, LogicRNCID=501',
'Label=UH0821_U3, CellID=8213, LogicRNCID=501',
'Label=UH0879_U96, CellID=55060, LogicRNCID=501',
'Label=UH0715_U94, CellID=47154, LogicRNCID=501',
'Label=UH0952_U2, CellID=9522, LogicRNCID=501',
'Label=UH1976_U1, CellID=19761, LogicRNCID=501',
'Label=UH0766_U96, CellID=47664, LogicRNCID=501',
'Label=UH0845_U3, CellID=8453, LogicRNCID=501',
'Label=UH0736_U94, CellID=47364, LogicRNCID=501',
'Label=UH0845_U2, CellID=8452, LogicRNCID=501',
'Label=UH0973_U1, CellID=9731, LogicRNCID=501',
'Label=UH0617_U2, CellID=6172, LogicRNCID=501',
'Label=UH0617_U1, CellID=6171, LogicRNCID=501',
'Label=UH3996_U2, CellID=39962, LogicRNCID=501',
'Label=UH0738_U95, CellID=47385, LogicRNCID=501',
'Label=UH0839_U97, CellID=48397, LogicRNCID=501',
'Label=UH0974_U1, CellID=9741, LogicRNCID=501',
'Label=UH0820_U5, CellID=8205, LogicRNCID=501',
'Label=UH0711_U96, CellID=47116, LogicRNCID=501',
'Label=UH0618_U5, CellID=6185, LogicRNCID=501',
'Label=UH0714_U94, CellID=47144, LogicRNCID=501',
'Label=UH0618_U4, CellID=6184, LogicRNCID=501',
'Label=UH0823_U1, CellID=8231, LogicRNCID=501',
'Label=UH0952_U4, CellID=9524, LogicRNCID=501',
'Label=UH0715_U95, CellID=47155, LogicRNCID=501',
'Label=UH0952_U3, CellID=9523, LogicRNCID=501',
'Label=UH0843_U94, CellID=48434, LogicRNCID=501',
'Label=UH0820_U1, CellID=8201, LogicRNCID=501',
'Label=UH0973_U6, CellID=9736, LogicRNCID=501',
'Label=UH0743_U2, CellID=7432, LogicRNCID=501',
'Label=UH0973_U5, CellID=9735, LogicRNCID=501',
'Label=UH0743_U1, CellID=7431, LogicRNCID=501',
'Label=UH0736_U96, CellID=47366, LogicRNCID=501',
'Label=UH0845_U4, CellID=8454, LogicRNCID=501',
'Label=UH0736_U95, CellID=47365, LogicRNCID=501',
'Label=UH0841_U96, CellID=48416, LogicRNCID=501',
'Label=UH0950_U4, CellID=9504, LogicRNCID=501',
'Label=UH0617_U6, CellID=6176, LogicRNCID=501',
'Label=UH0713_U95, CellID=47135, LogicRNCID=501',
'Label=UH0763_U96, CellID=47636, LogicRNCID=501',
'Label=UH1947_U6, CellID=19476, LogicRNCID=501',
'Label=UH0763_U95, CellID=47635, LogicRNCID=501',
'Label=UH1947_U5, CellID=19475, LogicRNCID=501',
'Label=UH1976_U6, CellID=19766, LogicRNCID=501',
'Label=UH3922_U1, CellID=39221, LogicRNCID=501',
'Label=UH0952_U5, CellID=9525, LogicRNCID=501',
'Label=UH0715_U96, CellID=47156, LogicRNCID=501',
'Label=UH0743_U6, CellID=7436, LogicRNCID=501',
'Label=UH0711_U95, CellID=47115, LogicRNCID=501',
'Label=UH0839_U94, CellID=48394, LogicRNCID=501',
'Label=UH0820_U2, CellID=8202, LogicRNCID=501',
'Label=UH0849_U1, CellID=8491, LogicRNCID=501',
'Label=UH0951_U4, CellID=9514, LogicRNCID=501',
'Label=UH0618_U6, CellID=6186, LogicRNCID=501',
'Label=UH0951_U3, CellID=9513, LogicRNCID=501',
'Label=UH0737_U95, CellID=47375, LogicRNCID=501',
'Label=UH3918_U3, CellID=39183, LogicRNCID=501',
'Label=UH0974_U3, CellID=9743, LogicRNCID=501',
'Label=UH0737_U94, CellID=47374, LogicRNCID=501',
'Label=UH0763_U94, CellID=47634, LogicRNCID=501',
'Label=UH3918_U6, CellID=39186, LogicRNCID=501',
'Label=UH0974_U6, CellID=9746, LogicRNCID=501',
'Label=UH3995_U3, CellID=39953, LogicRNCID=501',
'Label=UH0748_U1, CellID=7481, LogicRNCID=501',
'Label=UH3922_U4, CellID=39224, LogicRNCID=501',
'Label=UH0869_U95, CellID=48695, LogicRNCID=501',
'Label=UH0869_U94, CellID=48694, LogicRNCID=501',
'Label=UH0617_U4, CellID=6174, LogicRNCID=501',
'Label=UH3996_U5, CellID=39965, LogicRNCID=501',
'Label=UH0822_U1, CellID=8221, LogicRNCID=501',
'Label=UH0636_U94, CellID=46364, LogicRNCID=501',
'Label=UH0770_U5, CellID=7705, LogicRNCID=501',
'Label=UH0821_U6, CellID=8216, LogicRNCID=501',
'Label=UH0821_U5, CellID=8215, LogicRNCID=501',
'Label=UH0879_U94, CellID=55062, LogicRNCID=501',
'Label=UH0611_U95, CellID=46115, LogicRNCID=501',
'Label=UH3920_U2, CellID=39202, LogicRNCID=501',
'Label=UH0822_U6, CellID=8226, LogicRNCID=501',
'Label=UH0950_U5, CellID=9505, LogicRNCID=501',
'Label=UH3921_U3, CellID=39213, LogicRNCID=501',
'Label=UH0740_U94, CellID=47404, LogicRNCID=501',
'Label=UH0951_U6, CellID=9516, LogicRNCID=501',
'Label=UH0823_U6, CellID=8236, LogicRNCID=501',
'Label=UH0825_U1, CellID=8251, LogicRNCID=501',
'Label=UH0767_U94, CellID=47674, LogicRNCID=501',
'Label=UH0748_U2, CellID=7482, LogicRNCID=501',
'Label=UH3922_U5, CellID=39225, LogicRNCID=501',
'Label=UH1951_U3, CellID=19513, LogicRNCID=501',
'Label=UH0869_U97, CellID=48697, LogicRNCID=501',
'Label=UH1951_U4, CellID=19514, LogicRNCID=501',
'Label=UH0748_U3, CellID=7483, LogicRNCID=501',
'Label=UH3921_U1, CellID=39211, LogicRNCID=501',
'Label=UH0849_U2, CellID=8492, LogicRNCID=501',
'Label=UH3921_U2, CellID=39212, LogicRNCID=501',
'Label=UH0849_U3, CellID=8493, LogicRNCID=501',
'Label=UH0822_U5, CellID=8225, LogicRNCID=501',
'Label=UH3920_U1, CellID=39201, LogicRNCID=501',
'Label=UH0950_U6, CellID=9506, LogicRNCID=501',
'Label=UH0611_U94, CellID=46114, LogicRNCID=501',
'Label=UH0840_U96, CellID=48406, LogicRNCID=501',
'Label=UH0770_U3, CellID=7703, LogicRNCID=501',
'Label=UH0770_U4, CellID=7704, LogicRNCID=501',
'Label=UH0738_U94, CellID=47384, LogicRNCID=501',
'Label=UH3996_U4, CellID=39964, LogicRNCID=501',
'Label=UH0617_U3, CellID=6173, LogicRNCID=501',
'Label=UH0950_U1, CellID=9501, LogicRNCID=501',
'Label=UH0636_U95, CellID=46365, LogicRNCID=501',
'Label=UH3922_U2, CellID=39222, LogicRNCID=501',
'Label=UH1951_U1, CellID=19511, LogicRNCID=501',
'Label=UH1951_U2, CellID=19512, LogicRNCID=501',
'Label=UH0869_U96, CellID=48696, LogicRNCID=501',
'Label=UH1947_U3, CellID=19473, LogicRNCID=501',
'Label=UH0846_U6, CellID=8466, LogicRNCID=501',
'Label=UH1947_U4, CellID=19474, LogicRNCID=501',
'Label=UH3995_U4, CellID=39954, LogicRNCID=501',
'Label=UH0974_U2, CellID=9742, LogicRNCID=501',
'Label=UH0846_U3, CellID=8463, LogicRNCID=501',
'Label=UH1947_U1, CellID=19471, LogicRNCID=501',
'Label=UH3995_U1, CellID=39951, LogicRNCID=501',
'Label=UH0823_U3, CellID=8233, LogicRNCID=501',
'Label=UH0714_U95, CellID=47145, LogicRNCID=501',
'Label=UH0823_U4, CellID=8234, LogicRNCID=501',
'Label=UH0823_U5, CellID=8235, LogicRNCID=501',
'Label=UH0743_U4, CellID=7434, LogicRNCID=501',
'Label=UH0711_U94, CellID=47114, LogicRNCID=501',
'Label=UH0743_U5, CellID=7435, LogicRNCID=501',
'Label=UH0839_U95, CellID=48395, LogicRNCID=501',
'Label=UH1976_U4, CellID=19764, LogicRNCID=501',
'Label=UH0843_U96, CellID=48436, LogicRNCID=501',
'Label=UH1976_U5, CellID=19765, LogicRNCID=501',
'Label=UH0952_U6, CellID=9526, LogicRNCID=501',
'Label=UH0821_U1, CellID=8211, LogicRNCID=501',
'Label=UH3995_U5, CellID=39955, LogicRNCID=501',
'Label=UH0821_U2, CellID=8212, LogicRNCID=501',
'Label=UH3995_U6, CellID=39956, LogicRNCID=501',
'Label=UH0950_U3, CellID=9503, LogicRNCID=501',
'Label=UH0841_U95, CellID=48415, LogicRNCID=501',
'Label=UH0822_U4, CellID=8224, LogicRNCID=501',
'Label=UH0713_U96, CellID=47136, LogicRNCID=501',
'Label=UH0973_U3, CellID=9733, LogicRNCID=501',
'Label=UH0717_U4, CellID=7174, LogicRNCID=501',
'Label=UH0973_U4, CellID=9734, LogicRNCID=501',
'Label=UH0717_U5, CellID=7175, LogicRNCID=501',
'Label=UH0845_U5, CellID=8455, LogicRNCID=501',
'Label=UH0717_U6, CellID=7176, LogicRNCID=501',
'Label=UH0845_U6, CellID=8456, LogicRNCID=501',
'Label=UH0743_U3, CellID=7433, LogicRNCID=501',
'Label=UH0766_U94, CellID=47666, LogicRNCID=501',
'Label=UH0747_U5, CellID=7475, LogicRNCID=501',
'Label=UH1976_U3, CellID=19763, LogicRNCID=501',
'Label=UH0843_U95, CellID=48435, LogicRNCID=501',
'Label=UH0618_U3, CellID=6183, LogicRNCID=501',
'Label=UH0951_U1, CellID=9511, LogicRNCID=501',
'Label=UH0823_U2, CellID=8232, LogicRNCID=501',
'Label=UH0765_U96, CellID=47656, LogicRNCID=501',
'Label=UH0820_U4, CellID=8204, LogicRNCID=501',
'Label=UH0839_U96, CellID=48396, LogicRNCID=501',
'Label=UH0846_U1, CellID=8461, LogicRNCID=501',
'Label=UH3918_U1, CellID=39181, LogicRNCID=501',
'Label=UH3996_U1, CellID=39961, LogicRNCID=501',
'Label=UH0770_U6, CellID=7706, LogicRNCID=501',
'Label=UH0738_U96, CellID=47386, LogicRNCID=501',
'Label=UH3996_U3, CellID=39963, LogicRNCID=501',
'Label=UH0845_U1, CellID=8451, LogicRNCID=501',
'Label=UH0717_U2, CellID=7172, LogicRNCID=501',
'Label=UH0973_U2, CellID=9732, LogicRNCID=501',
'Label=UH0717_U3, CellID=7173, LogicRNCID=501',
'Label=UH3921_U6, CellID=39216, LogicRNCID=501',
'Label=UH0952_U1, CellID=9521, LogicRNCID=501',
'Label=UH0766_U95, CellID=47665, LogicRNCID=501',
'Label=UH1976_U2, CellID=19762, LogicRNCID=501',
'Label=UH0840_U94, CellID=48404, LogicRNCID=501',
'Label=UH0770_U1, CellID=7701, LogicRNCID=501',
'Label=UH0840_U95, CellID=48405, LogicRNCID=501',
'Label=UH0770_U2, CellID=7702, LogicRNCID=501',
'Label=UH0822_U2, CellID=8222, LogicRNCID=501',
'Label=UH3996_U6, CellID=39966, LogicRNCID=501',
'Label=UH0713_U94, CellID=47134, LogicRNCID=501',
'Label=UH0617_U5, CellID=6175, LogicRNCID=501',
'Label=UH0739_U95, CellID=47395, LogicRNCID=501',
'Label=UH0611_U96, CellID=46116, LogicRNCID=501',
'Label=UH0618_U1, CellID=6181, LogicRNCID=501',
'Label=UH0618_U2, CellID=6182, LogicRNCID=501',
'Label=UH0825_U3, CellID=8253, LogicRNCID=501',
'Label=UH0844_U95, CellID=48445, LogicRNCID=501',
'Label=UH0825_U4, CellID=8254, LogicRNCID=501',
'Label=UH0844_U96, CellID=48446, LogicRNCID=501',
'Label=UH1945_U6, CellID=19456, LogicRNCID=501',
'Label=UH0838_U94, CellID=48384, LogicRNCID=501',
'Label=UH0761_U97, CellID=47617, LogicRNCID=501',
'Label=UH0838_U96, CellID=48386, LogicRNCID=501',
'Label=UH0740_U95, CellID=47405, LogicRNCID=501',
'Label=UH3921_U4, CellID=39214, LogicRNCID=501',
'Label=UH0849_U5, CellID=8495, LogicRNCID=501',
'Label=UH0747_U2, CellID=7472, LogicRNCID=501',
'Label=UH0844_U94, CellID=48444, LogicRNCID=501',
'Label=UH1805_U6, CellID=18056, LogicRNCID=501',
'Label=UH1703_U2, CellID=17032, LogicRNCID=501',
'Label=UH0807_U2, CellID=8072, LogicRNCID=501',
'Label=UH0602_U4, CellID=6024, LogicRNCID=501',
'Label=UH0988_U5, CellID=9885, LogicRNCID=501',
'Label=UH0758_U1, CellID=7581, LogicRNCID=501',
'Label=UH0751_U96, CellID=47516, LogicRNCID=501',
'Label=UH0623_U96, CellID=46236, LogicRNCID=501',
'Label=UH1935_U6, CellID=19356, LogicRNCID=501',
'Label=UH1833_U2, CellID=18332, LogicRNCID=501',
'Label=UH1705_U2, CellID=17052, LogicRNCID=501',
'Label=UH0988_U4, CellID=9884, LogicRNCID=501',
'Label=UH2709_U95, CellID=50066, LogicRNCID=501',
'Label=UH0955_U94, CellID=49554, LogicRNCID=501',
'Label=UH0750_U96, CellID=47506, LogicRNCID=501',
'Label=UH0622_U96, CellID=46226, LogicRNCID=501',
'Label=UH2984_U2, CellID=29842, LogicRNCID=501',
'Label=UH1960_U2, CellID=19602, LogicRNCID=501',
'Label=UH1704_U2, CellID=17042, LogicRNCID=501',
'Label=UH2709_U94, CellID=50065, LogicRNCID=501',
'Label=UH0965_U3, CellID=9653, LogicRNCID=501',
'Label=UH4600_U4, CellID=46004, LogicRNCID=501',
'Label=UH1707_U6, CellID=17076, LogicRNCID=501',
'Label=UH0965_U2, CellID=9652, LogicRNCID=501',
'Label=UH4600_U3, CellID=46003, LogicRNCID=501',
'Label=UH1707_U5, CellID=17075, LogicRNCID=501',
'Label=UH0965_U1, CellID=9651, LogicRNCID=501',
'Label=UH0830_U96, CellID=48306, LogicRNCID=501',
'Label=UH0807_U5, CellID=8075, LogicRNCID=501',
'Label=UH1945_U96, CellID=50058, LogicRNCID=501',
'Label=UH2983_U4, CellID=29834, LogicRNCID=501',
'Label=UH2932_U2, CellID=29322, LogicRNCID=501',
'Label=UH1908_U2, CellID=19082, LogicRNCID=501',
'Label=UH1882_U6, CellID=18826, LogicRNCID=501',
'Label=UH1703_U4, CellID=17034, LogicRNCID=501',
'Label=UH0807_U4, CellID=8074, LogicRNCID=501',
'Label=UH1936_U4, CellID=19364, LogicRNCID=501',
'Label=UH0963_U6, CellID=9636, LogicRNCID=501',
'Label=UH0835_U6, CellID=8356, LogicRNCID=501',
'Label=UH1936_U3, CellID=19363, LogicRNCID=501',
'Label=UH0963_U5, CellID=9635, LogicRNCID=501',
'Label=UH0835_U5, CellID=8355, LogicRNCID=501',
'Label=UH1936_U2, CellID=19362, LogicRNCID=501',
'Label=UH0963_U4, CellID=9634, LogicRNCID=501',
'Label=UH2932_U5, CellID=29325, LogicRNCID=501',
'Label=UH1908_U5, CellID=19085, LogicRNCID=501',
'Label=UH0961_U3, CellID=9613, LogicRNCID=501',
'Label=UH1945_U98, CellID=50060, LogicRNCID=501',
'Label=UH2983_U6, CellID=29836, LogicRNCID=501',
'Label=UH2932_U4, CellID=29324, LogicRNCID=501',
'Label=UH1908_U4, CellID=19084, LogicRNCID=501',
'Label=UH1703_U6, CellID=17036, LogicRNCID=501',
'Label=UH0750_U95, CellID=47505, LogicRNCID=501',
'Label=UH0622_U95, CellID=46225, LogicRNCID=501',
'Label=UH2984_U1, CellID=29841, LogicRNCID=501',
'Label=UH1960_U1, CellID=19601, LogicRNCID=501',
'Label=UH1704_U1, CellID=17041, LogicRNCID=501',
'Label=UH1951_U96, CellID=50064, LogicRNCID=501',
'Label=UH0750_U94, CellID=47504, LogicRNCID=501',
'Label=UH0622_U94, CellID=46224, LogicRNCID=501',
'Label=UH0754_U95, CellID=47545, LogicRNCID=501',
'Label=UH1964_U1, CellID=19641, LogicRNCID=501',
'Label=UH0754_U94, CellID=47544, LogicRNCID=501',
'Label=UH0965_U6, CellID=9656, LogicRNCID=501',
'Label=UH0965_U5, CellID=9655, LogicRNCID=501',
'Label=UH4600_U6, CellID=46006, LogicRNCID=501',
'Label=UH0965_U4, CellID=9654, LogicRNCID=501',
'Label=UH4600_U5, CellID=46005, LogicRNCID=501',
'Label=UH1935_U4, CellID=19354, LogicRNCID=501',
'Label=UH0988_U2, CellID=9882, LogicRNCID=501',
'Label=UH0962_U6, CellID=9626, LogicRNCID=501',
'Label=UH0760_U96, CellID=50073, LogicRNCID=501',
'Label=UH1935_U3, CellID=19353, LogicRNCID=501',
'Label=UH0988_U1, CellID=9881, LogicRNCID=501',
'Label=UH0962_U5, CellID=9625, LogicRNCID=501',
'Label=UH0760_U95, CellID=50072, LogicRNCID=501',
'Label=UH1704_U6, CellID=17046, LogicRNCID=501',
'Label=UH0962_U2, CellID=9622, LogicRNCID=501',
'Label=UH2984_U5, CellID=29845, LogicRNCID=501',
'Label=UH1960_U5, CellID=19605, LogicRNCID=501',
'Label=UH1704_U5, CellID=17045, LogicRNCID=501',
'Label=UH0962_U1, CellID=9621, LogicRNCID=501',
'Label=UH0706_U1, CellID=7061, LogicRNCID=501',
'Label=UH0955_U96, CellID=49556, LogicRNCID=501',
'Label=UH0758_U3, CellID=7583, LogicRNCID=501',
'Label=UH1801_U94, CellID=58014, LogicRNCID=501',
'Label=UH1833_U4, CellID=18334, LogicRNCID=501',
'Label=UH1705_U4, CellID=17054, LogicRNCID=501',
'Label=UH0988_U6, CellID=9886, LogicRNCID=501',
'Label=UH0758_U2, CellID=7582, LogicRNCID=501',
'Label=UH1833_U3, CellID=18333, LogicRNCID=501',
'Label=UH1705_U3, CellID=17053, LogicRNCID=501',
'Label=UH0708_U3, CellID=7083, LogicRNCID=501',
'Label=UH0708_U2, CellID=7082, LogicRNCID=501',
'Label=UH0708_U1, CellID=7081, LogicRNCID=501',
'Label=UH0752_U96, CellID=47526, LogicRNCID=501',
'Label=UH1936_U6, CellID=19366, LogicRNCID=501',
'Label=UH0752_U95, CellID=47525, LogicRNCID=501',
'Label=UH1936_U5, CellID=19365, LogicRNCID=501',
'Label=UH0752_U94, CellID=47524, LogicRNCID=501',
'Label=UH0831_U96, CellID=48316, LogicRNCID=501',
'Label=UH0703_U96, CellID=47036, LogicRNCID=501',
'Label=UH1964_U4, CellID=19644, LogicRNCID=501',
'Label=UH0831_U95, CellID=48315, LogicRNCID=501',
'Label=UH0703_U95, CellID=47035, LogicRNCID=501',
'Label=UH0831_U94, CellID=48314, LogicRNCID=501',
'Label=UH0703_U94, CellID=47034, LogicRNCID=501',
'Label=UH1964_U2, CellID=19642, LogicRNCID=501',
'Label=UH3925_U5, CellID=39255, LogicRNCID=501',
'Label=UH0981_U5, CellID=9815, LogicRNCID=501',
'Label=UH0623_U1, CellID=6231, LogicRNCID=501',
'Label=UH0821_U94, CellID=48214, LogicRNCID=501',
'Label=UH0744_U96, CellID=47446, LogicRNCID=501',
'Label=UH3925_U4, CellID=39254, LogicRNCID=501',
'Label=UH1928_U6, CellID=19286, LogicRNCID=501',
'Label=UH0981_U4, CellID=9814, LogicRNCID=501',
'Label=UH1952_U1, CellID=19521, LogicRNCID=501',
'Label=UH1602_U95, CellID=51008, LogicRNCID=501',
'Label=UH0742_U94, CellID=47424, LogicRNCID=501',
'Label=UH0614_U94, CellID=46144, LogicRNCID=501',
'Label=UH2950_U4, CellID=29504, LogicRNCID=501',
'Label=UH0825_U6, CellID=8256, LogicRNCID=501',
'Label=UH1602_U94, CellID=51007, LogicRNCID=501',
'Label=UH2950_U3, CellID=29503, LogicRNCID=501',
'Label=UH0742_U96, CellID=47426, LogicRNCID=501',
'Label=UH0614_U96, CellID=46146, LogicRNCID=501',
'Label=UH2950_U6, CellID=29506, LogicRNCID=501',
'Label=UH1952_U2, CellID=19522, LogicRNCID=501',
'Label=UH1602_U96, CellID=51009, LogicRNCID=501',
'Label=UH0742_U95, CellID=47425, LogicRNCID=501',
'Label=UH0614_U95, CellID=46145, LogicRNCID=501',
'Label=UH2950_U5, CellID=29505, LogicRNCID=501',
'Label=UH1925_U95, CellID=50047, LogicRNCID=501',
'Label=UH0825_U95, CellID=48255, LogicRNCID=501',
'Label=UH1958_U3, CellID=19583, LogicRNCID=501',
'Label=UH1907_U1, CellID=19071, LogicRNCID=501',
'Label=UH1702_U3, CellID=17023, LogicRNCID=501',
'Label=UH0601_U5, CellID=6015, LogicRNCID=501',
'Label=UH1925_U96, CellID=50046, LogicRNCID=501',
'Label=UH0825_U94, CellID=48254, LogicRNCID=501',
'Label=UH1952_U6, CellID=19526, LogicRNCID=501',
'Label=UH1952_U5, CellID=19525, LogicRNCID=501',
'Label=UH0646_U96, CellID=51012, LogicRNCID=501',
'Label=UH1952_U4, CellID=19524, LogicRNCID=501',
'Label=UH0646_U95, CellID=51011, LogicRNCID=501',
'Label=UH1952_U3, CellID=19523, LogicRNCID=501',
'Label=UH0646_U94, CellID=51010, LogicRNCID=501',
'Label=UH1902_U95, CellID=50035, LogicRNCID=501',
'Label=UH0747_U95, CellID=47475, LogicRNCID=501',
'Label=UH3928_U3, CellID=39283, LogicRNCID=501',
'Label=UH2981_U1, CellID=29811, LogicRNCID=501',
'Label=UH1803_U5, CellID=18035, LogicRNCID=501',
'Label=UH1902_U94, CellID=50034, LogicRNCID=501',
'Label=UH0747_U94, CellID=47474, LogicRNCID=501',
'Label=UH3928_U2, CellID=39282, LogicRNCID=501',
'Label=UH0973_U95, CellID=49735, LogicRNCID=501',
'Label=UH0845_U95, CellID=48455, LogicRNCID=501',
'Label=UH0717_U95, CellID=47175, LogicRNCID=501',
'Label=UH0973_U94, CellID=49734, LogicRNCID=501',
'Label=UH0845_U94, CellID=48454, LogicRNCID=501',
'Label=UH0717_U94, CellID=47174, LogicRNCID=501',
'Label=UH1702_U5, CellID=17025, LogicRNCID=501',
'Label=UH0960_U1, CellID=9601, LogicRNCID=501',
'Label=UH0832_U1, CellID=8321, LogicRNCID=501',
'Label=UH1925_U94, CellID=50048, LogicRNCID=501',
'Label=UH0825_U96, CellID=48256, LogicRNCID=501',
'Label=UH1907_U2, CellID=19072, LogicRNCID=501',
'Label=UH1702_U4, CellID=17024, LogicRNCID=501',
'Label=UH0601_U6, CellID=6016, LogicRNCID=501',
'Label=UH1905_U4, CellID=19054, LogicRNCID=501',
'Label=UH1854_U2, CellID=18542, LogicRNCID=501',
'Label=UH0881_U4, CellID=8814, LogicRNCID=501',
'Label=UH0830_U2, CellID=8302, LogicRNCID=501',
'Label=UH1854_U1, CellID=18541, LogicRNCID=501',
'Label=UH0881_U3, CellID=8813, LogicRNCID=501',
'Label=UH0830_U1, CellID=8301, LogicRNCID=501',
'Label=UH0951_U96, CellID=49516, LogicRNCID=501',
'Label=UH0618_U94, CellID=46184, LogicRNCID=501',
'Label=UH1804_U95, CellID=50023, LogicRNCID=501',
'Label=UH1804_U94, CellID=50022, LogicRNCID=501',
'Label=UH0880_U6, CellID=8806, LogicRNCID=501',
'Label=UH1803_U96, CellID=50021, LogicRNCID=501',
'Label=UH1803_U95, CellID=50020, LogicRNCID=501',
'Label=UH1603_U96, CellID=50012, LogicRNCID=501',
'Label=UH0879_U6, CellID=8796, LogicRNCID=501',
'Label=UH0623_U6, CellID=6236, LogicRNCID=501',
'Label=UH1964_U94, CellID=55387, LogicRNCID=501',
'Label=UH1603_U95, CellID=50011, LogicRNCID=501',
'Label=UH0821_U99, CellID=48219, LogicRNCID=501',
'Label=UH1801_U1, CellID=18011, LogicRNCID=501',
'Label=UH0879_U5, CellID=8795, LogicRNCID=501',
'Label=UH2981_U3, CellID=29813, LogicRNCID=501',
'Label=UH1906_U1, CellID=19061, LogicRNCID=501',
'Label=UH0600_U5, CellID=6005, LogicRNCID=501',
'Label=UH1902_U96, CellID=50036, LogicRNCID=501',
'Label=UH3928_U4, CellID=39284, LogicRNCID=501',
'Label=UH2981_U2, CellID=29812, LogicRNCID=501',
'Label=UH1803_U6, CellID=18036, LogicRNCID=501',
'Label=UH0600_U4, CellID=6004, LogicRNCID=501',
'Label=UH0980_U2, CellID=9802, LogicRNCID=501',
'Label=UH0980_U1, CellID=9801, LogicRNCID=501',
'Label=UH0973_U96, CellID=49736, LogicRNCID=501',
'Label=UH0845_U96, CellID=48456, LogicRNCID=501',
'Label=UH0717_U96, CellID=47176, LogicRNCID=501',
'Label=UH0822_U94, CellID=48224, LogicRNCID=501',
'Label=UH0745_U96, CellID=47456, LogicRNCID=501',
'Label=UH1801_U6, CellID=18016, LogicRNCID=501',
'Label=UH0854_U4, CellID=8544, LogicRNCID=501',
'Label=UH1604_U96, CellID=50015, LogicRNCID=501',
'Label=UH0745_U95, CellID=47455, LogicRNCID=501',
'Label=UH0854_U3, CellID=8543, LogicRNCID=501',
'Label=UH1604_U95, CellID=50014, LogicRNCID=501',
'Label=UH0955_U6, CellID=9556, LogicRNCID=501',
'Label=UH3925_U1, CellID=39251, LogicRNCID=501',
'Label=UH1928_U3, CellID=19283, LogicRNCID=501',
'Label=UH0981_U1, CellID=9811, LogicRNCID=501',
'Label=UH0955_U5, CellID=9555, LogicRNCID=501',
'Label=UH0846_U96, CellID=48466, LogicRNCID=501',
'Label=UH1928_U2, CellID=19282, LogicRNCID=501',
'Label=UH0955_U4, CellID=9554, LogicRNCID=501',
'Label=UH1882_U94, CellID=50031, LogicRNCID=501',
'Label=UH0849_U95, CellID=48495, LogicRNCID=501',
'Label=UH1905_U5, CellID=19055, LogicRNCID=501',
'Label=UH1854_U3, CellID=18543, LogicRNCID=501',
'Label=UH1803_U1, CellID=18031, LogicRNCID=501',
'Label=UH0881_U5, CellID=8815, LogicRNCID=501',
'Label=UH0830_U3, CellID=8303, LogicRNCID=501',
'Label=UH0849_U94, CellID=48494, LogicRNCID=501',
'Label=UH1906_U3, CellID=19063, LogicRNCID=501',
'Label=UH0831_U1, CellID=8311, LogicRNCID=501',
'Label=UH1907_U95, CellID=50038, LogicRNCID=501',
'Label=UH3928_U6, CellID=39286, LogicRNCID=501',
'Label=UH2981_U4, CellID=29814, LogicRNCID=501',
'Label=UH1906_U2, CellID=19062, LogicRNCID=501',
'Label=UH1907_U94, CellID=50037, LogicRNCID=501',
'Label=UH3928_U5, CellID=39285, LogicRNCID=501',
'Label=UH0955_U2, CellID=9552, LogicRNCID=501',
'Label=UH0622_U4, CellID=6224, LogicRNCID=501',
'Label=UH0820_U97, CellID=48207, LogicRNCID=501',
'Label=UH0955_U1, CellID=9551, LogicRNCID=501',
'Label=UH0622_U3, CellID=6223, LogicRNCID=501',
'Label=UH0820_U96, CellID=48206, LogicRNCID=501',
'Label=UH0980_U6, CellID=9806, LogicRNCID=501',
'Label=UH0622_U2, CellID=6222, LogicRNCID=501',
'Label=UH0820_U95, CellID=48205, LogicRNCID=501',
'Label=UH0980_U5, CellID=9805, LogicRNCID=501',
'Label=UH0622_U1, CellID=6221, LogicRNCID=501',
'Label=UH0820_U94, CellID=48204, LogicRNCID=501',
'Label=UH0743_U96, CellID=47436, LogicRNCID=501',
'Label=UH0980_U4, CellID=9804, LogicRNCID=501',
'Label=UH0743_U95, CellID=47435, LogicRNCID=501',
'Label=UH0980_U3, CellID=9803, LogicRNCID=501',
'Label=UH0951_U94, CellID=49514, LogicRNCID=501',
'Label=UH0823_U94, CellID=48234, LogicRNCID=501',
'Label=UH0746_U96, CellID=47466, LogicRNCID=501',
'Label=UH0618_U96, CellID=46186, LogicRNCID=501',
'Label=UH1805_U94, CellID=50025, LogicRNCID=501',
'Label=UH0746_U95, CellID=47465, LogicRNCID=501',
'Label=UH0618_U95, CellID=46185, LogicRNCID=501',
'Label=UH1804_U96, CellID=50024, LogicRNCID=501',
'Label=UH0623_U5, CellID=6235, LogicRNCID=501',
'Label=UH1603_U94, CellID=50010, LogicRNCID=501',
'Label=UH0821_U98, CellID=48218, LogicRNCID=501',
'Label=UH0770_U96, CellID=47706, LogicRNCID=501',
'Label=UH0623_U4, CellID=6234, LogicRNCID=501',
'Label=UH0821_U97, CellID=48217, LogicRNCID=501',
'Label=UH0770_U95, CellID=47705, LogicRNCID=501',
'Label=UH0879_U3, CellID=8793, LogicRNCID=501',
'Label=UH1906_U5, CellID=19065, LogicRNCID=501',
'Label=UH0831_U3, CellID=8313, LogicRNCID=501',
'Label=UH1913_U94, CellID=50040, LogicRNCID=501',
'Label=UH2981_U6, CellID=29816, LogicRNCID=501',
'Label=UH1906_U4, CellID=19064, LogicRNCID=501',
'Label=UH0831_U2, CellID=8312, LogicRNCID=501',
'Label=UH1907_U96, CellID=50039, LogicRNCID=501',
'Label=UH2981_U5, CellID=29815, LogicRNCID=501',
'Label=UH0622_U6, CellID=6226, LogicRNCID=501',
'Label=UH0846_U95, CellID=48465, LogicRNCID=501',
'Label=UH0820_U99, CellID=48209, LogicRNCID=501',
'Label=UH1928_U1, CellID=19281, LogicRNCID=501',
'Label=UH0955_U3, CellID=9553, LogicRNCID=501',
'Label=UH0622_U5, CellID=6225, LogicRNCID=501',
'Label=UH0846_U94, CellID=48464, LogicRNCID=501',
'Label=UH0820_U98, CellID=48208, LogicRNCID=501',
'Label=UH0830_U5, CellID=8305, LogicRNCID=501',
'Label=UH0600_U1, CellID=6001, LogicRNCID=501',
'Label=UH1882_U95, CellID=50032, LogicRNCID=501',
'Label=UH0849_U96, CellID=48496, LogicRNCID=501',
'Label=UH1854_U4, CellID=18544, LogicRNCID=501',
'Label=UH1803_U2, CellID=18032, LogicRNCID=501',
'Label=UH0881_U6, CellID=8816, LogicRNCID=501',
'Label=UH0830_U4, CellID=8304, LogicRNCID=501',
'Label=UH0744_U95, CellID=47445, LogicRNCID=501',
'Label=UH3925_U3, CellID=39253, LogicRNCID=501',
'Label=UH1928_U5, CellID=19285, LogicRNCID=501',
'Label=UH0981_U3, CellID=9813, LogicRNCID=501',
'Label=UH0744_U94, CellID=47444, LogicRNCID=501',
'Label=UH3925_U2, CellID=39252, LogicRNCID=501',
'Label=UH1928_U4, CellID=19284, LogicRNCID=501',
'Label=UH0981_U2, CellID=9812, LogicRNCID=501',
'Label=UH0623_U3, CellID=6233, LogicRNCID=501',
'Label=UH0821_U96, CellID=48216, LogicRNCID=501',
'Label=UH0770_U94, CellID=47704, LogicRNCID=501',
'Label=UH3925_U6, CellID=39256, LogicRNCID=501',
'Label=UH0981_U6, CellID=9816, LogicRNCID=501',
'Label=UH0879_U2, CellID=8792, LogicRNCID=501',
'Label=UH0623_U2, CellID=6232, LogicRNCID=501',
'Label=UH0821_U95, CellID=48215, LogicRNCID=501',
'Label=UH0831_U6, CellID=8316, LogicRNCID=501',
'Label=UH0601_U2, CellID=6012, LogicRNCID=501',
'Label=UH0831_U5, CellID=8315, LogicRNCID=501',
'Label=UH0601_U1, CellID=6011, LogicRNCID=501',
'Label=UH1913_U96, CellID=50042, LogicRNCID=501',
'Label=UH1906_U6, CellID=19066, LogicRNCID=501',
'Label=UH0831_U4, CellID=8314, LogicRNCID=501',
'Label=UH1913_U95, CellID=50041, LogicRNCID=501',
'Label=UH0880_U4, CellID=8804, LogicRNCID=501',
'Label=UH1803_U94, CellID=50019, LogicRNCID=501',
'Label=UH0880_U3, CellID=8803, LogicRNCID=501',
'Label=UH0822_U96, CellID=48226, LogicRNCID=501',
'Label=UH0854_U6, CellID=8546, LogicRNCID=501',
'Label=UH0822_U95, CellID=48225, LogicRNCID=501',
'Label=UH0880_U1, CellID=8801, LogicRNCID=501',
'Label=UH0854_U5, CellID=8545, LogicRNCID=501',
'Label=UH0745_U94, CellID=47454, LogicRNCID=501',
'Label=UH1801_U4, CellID=18014, LogicRNCID=501',
'Label=UH0854_U2, CellID=8542, LogicRNCID=501',
'Label=UH1964_U96, CellID=55389, LogicRNCID=501',
'Label=UH1604_U94, CellID=50013, LogicRNCID=501',
'Label=UH1801_U3, CellID=18013, LogicRNCID=501',
'Label=UH0854_U1, CellID=8541, LogicRNCID=501',
'Label=UH1964_U95, CellID=55388, LogicRNCID=501',
'Label=UH0823_U96, CellID=48236, LogicRNCID=501',
'Label=UH1905_U2, CellID=19052, LogicRNCID=501',
'Label=UH0881_U2, CellID=8812, LogicRNCID=501',
'Label=UH1805_U96, CellID=50027, LogicRNCID=501',
'Label=UH0823_U95, CellID=48235, LogicRNCID=501',
'Label=UH1905_U1, CellID=19051, LogicRNCID=501',
'Label=UH0881_U1, CellID=8811, LogicRNCID=501',
'Label=UH1805_U95, CellID=50026, LogicRNCID=501',
'Label=UH1854_U6, CellID=18546, LogicRNCID=501',
'Label=UH1803_U4, CellID=18034, LogicRNCID=501',
'Label=UH0830_U6, CellID=8306, LogicRNCID=501',
'Label=UH0600_U2, CellID=6002, LogicRNCID=501',
'Label=UH1882_U96, CellID=50033, LogicRNCID=501',
'Label=UH3928_U1, CellID=39281, LogicRNCID=501',
'Label=UH1854_U5, CellID=18545, LogicRNCID=501',
'Label=UH1803_U3, CellID=18033, LogicRNCID=501',
'Label=UH0748_U96, CellID=47486, LogicRNCID=501',
'Label=UH1958_U2, CellID=19582, LogicRNCID=501',
'Label=UH1702_U2, CellID=17022, LogicRNCID=501',
'Label=UH0601_U4, CellID=6014, LogicRNCID=501',
'Label=UH0748_U95, CellID=47485, LogicRNCID=501',
'Label=UH1702_U1, CellID=17021, LogicRNCID=501',
'Label=UH0601_U3, CellID=6013, LogicRNCID=501',
'Label=UH0748_U94, CellID=47484, LogicRNCID=501',
'Label=UH0974_U5, CellID=9745, LogicRNCID=501',
'Label=UH0737_U96, CellID=47376, LogicRNCID=501',
'Label=UH3995_U2, CellID=39952, LogicRNCID=501',
'Label=UH3918_U4, CellID=39184, LogicRNCID=501',
'Label=UH1947_U2, CellID=19472, LogicRNCID=501',
'Label=UH0974_U4, CellID=9744, LogicRNCID=501',
'Label=UH0846_U4, CellID=8464, LogicRNCID=501',
'Label=UH0839_U99, CellID=48399, LogicRNCID=501',
]
cluster_UH_N = [
'UH0737',
'UH0752',
'UH0843',
'UH0844',
'UH0845',
'UH1803',
'UH0702',
'UH0715',
'UH0734',
'UH0744',
'UH0763',
'UH0770',
'UH0745',
'UH1704',
'UH1984',
'UH3922',
'UH0703',
'UH0738',
'UH0831',
'UH0960',
'UH0961',
'UH0965',
'UH1892',
'UH1905',
'UH1960',
'UH2709',
'UH2763',
'UH1702',
'UH3921',
'UH1707',
'UH0508',
'UH0600',
'UH0611',
'UH0614',
'UH0636',
'UH0736',
'UH0732',
'UH0739',
'UH0740',
'UH0747',
'UH0750',
'UH0760',
'UH0761',
'UH0830',
'UH0832',
'UH0849',
'UH0980',
'UH0963',
'UH0974',
'UH1805',
'UH1906',
'UH1925',
'UH1964',
'UH2932',
'UH2984',
'UH3928',
'UH3996',
'UH0746',
'UH0711',
'UH0731',
'UH0735',
'UH0842',
'UH1928',
'UH0708',
'UH0952',
'UH0838',
'UH0839',
'UH0841',
'UH1907',
'UH0622',
'UH0869',
'UH0646',
'UH0742',
'UH0743',
'UH0751',
'UH0756',
'UH0981',
'UH0610',
'UH1902',
'UH0623',
'UH0713',
'UH1945',
'UH0717',
'UH2938',
'UH0833',
'UH0840',
'UH0825',
'UH0951',
'UH0962',
'UH0966',
'UH0972',
'UH1703',
'UH1801',
'UH1804',
'UH0988',
'UH1936',
'UH1939',
'UH0765',
'UH0820',
'UH0973',
'UH0994',
'UH1854',
'UH1705',
'UH3990',
'UH0601',
'UH0602',
'UH0617',
'UH0618',
'UH0714',
'UH0754',
'UH0758',
'UH1908',
'UH1913',
'UH1951',
'UH1958',
'UH1976',
'UH3989',
'UH3995',
'UH4600',
'UH0748',
'UH0970',
'UH3925',
'UH1947',
'UH0880',
'UH3920',
'UH0879',
'UH0767',
'UH1935',
'UH0766',
'UH1952',
'UH0846',
]
cluster_CR = [
'Label=CR0835_U5, CellID=8355, LogicRNCID=501',
'Label=CR0835_U4, CellID=8354, LogicRNCID=501',
'Label=CR0835_U3, CellID=8353, LogicRNCID=501',
'Label=CR0835_U1, CellID=8351, LogicRNCID=501',
'Label=CR0835_U2, CellID=8352, LogicRNCID=501',
'Label=CR0835_U6, CellID=8356, LogicRNCID=501',
'Label=CR1857_U5, CellID=18575, LogicRNCID=501',
'Label=CR1857_U3, CellID=18573, LogicRNCID=501',
'Label=CR1857_U2, CellID=18572, LogicRNCID=501',
'Label=KR1858_U6, CellID=18586, LogicRNCID=501',
'Label=KR1858_U4, CellID=18584, LogicRNCID=501',
'Label=KR1858_U3, CellID=18583, LogicRNCID=501',
'Label=KR1858_U1, CellID=18581, LogicRNCID=501',
'Label=KR1852_U6, CellID=18526, LogicRNCID=501',
'Label=KR1852_U4, CellID=18524, LogicRNCID=501',
'Label=KR1852_U2, CellID=18522, LogicRNCID=501',
'Label=KR1852_U1, CellID=18521, LogicRNCID=501',
'Label=CR0297_U5, CellID=2975, LogicRNCID=501',
'Label=CR0297_U4, CellID=2974, LogicRNCID=501',
'Label=CR0297_U2, CellID=2972, LogicRNCID=501',
'Label=CR1266_U6, CellID=12666, LogicRNCID=501',
'Label=CR1266_U5, CellID=12665, LogicRNCID=501',
'Label=CR1266_U3, CellID=12663, LogicRNCID=501',
'Label=CR1266_U2, CellID=12662, LogicRNCID=501',
'Label=CR0456_U6, CellID=4566, LogicRNCID=501',
'Label=CR0456_U4, CellID=4564, LogicRNCID=501',
'Label=CR0456_U3, CellID=4563, LogicRNCID=501',
'Label=CR0456_U1, CellID=4561, LogicRNCID=501',
'Label=CR1749_U6, CellID=17496, LogicRNCID=501',
'Label=CR1749_U4, CellID=17494, LogicRNCID=501',
'Label=CR1749_U2, CellID=17492, LogicRNCID=501',
'Label=CR1749_U1, CellID=17491, LogicRNCID=501',
'Label=CR1329_U5, CellID=13295, LogicRNCID=501',
'Label=CR1329_U4, CellID=13294, LogicRNCID=501',
'Label=CR1329_U2, CellID=13292, LogicRNCID=501',
'Label=CR0776_U6, CellID=7766, LogicRNCID=501',
'Label=CR0776_U5, CellID=7765, LogicRNCID=501',
'Label=CR0776_U3, CellID=7763, LogicRNCID=501',
'Label=CR0776_U2, CellID=7762, LogicRNCID=501',
'Label=CR0473_U6, CellID=4736, LogicRNCID=501',
'Label=CR0473_U4, CellID=4734, LogicRNCID=501',
'Label=CR0473_U3, CellID=4733, LogicRNCID=501',
'Label=CR0473_U1, CellID=4731, LogicRNCID=501',
'Label=CR1390_U6, CellID=13906, LogicRNCID=501',
'Label=CR1390_U4, CellID=13904, LogicRNCID=501',
'Label=CR1390_U2, CellID=13902, LogicRNCID=501',
'Label=CR1390_U1, CellID=13901, LogicRNCID=501',
'Label=CR1408_U5, CellID=14085, LogicRNCID=501',
'Label=CR1408_U4, CellID=14084, LogicRNCID=501',
'Label=CR1408_U2, CellID=14082, LogicRNCID=501',
'Label=CR0592_U6, CellID=5926, LogicRNCID=501',
'Label=CR0592_U5, CellID=5925, LogicRNCID=501',
'Label=CR0592_U3, CellID=5923, LogicRNCID=501',
'Label=CR0592_U2, CellID=5922, LogicRNCID=501',
'Label=CR1102_U6, CellID=11026, LogicRNCID=501',
'Label=CR1102_U4, CellID=11024, LogicRNCID=501',
'Label=CR1102_U3, CellID=11023, LogicRNCID=501',
'Label=CR1102_U1, CellID=11021, LogicRNCID=501',
'Label=CR0465_U6, CellID=4656, LogicRNCID=501',
'Label=CR0465_U4, CellID=4654, LogicRNCID=501',
'Label=CR0465_U2, CellID=4652, LogicRNCID=501',
'Label=CR0465_U1, CellID=4651, LogicRNCID=501',
'Label=CR0472_U5, CellID=4725, LogicRNCID=501',
'Label=CR0472_U4, CellID=4724, LogicRNCID=501',
'Label=CR0472_U2, CellID=4722, LogicRNCID=501',
'Label=KR1851_U6, CellID=18516, LogicRNCID=501',
'Label=KR1851_U5, CellID=18515, LogicRNCID=501',
'Label=KR1851_U3, CellID=18513, LogicRNCID=501',
'Label=KR1851_U2, CellID=18512, LogicRNCID=501',
'Label=CR0264_U6, CellID=2646, LogicRNCID=501',
'Label=CR0264_U4, CellID=2644, LogicRNCID=501',
'Label=CR0264_U3, CellID=2643, LogicRNCID=501',
'Label=CR0264_U1, CellID=2641, LogicRNCID=501',
'Label=CR0258_U6, CellID=2586, LogicRNCID=501',
'Label=CR0258_U4, CellID=2584, LogicRNCID=501',
'Label=CR0258_U2, CellID=2582, LogicRNCID=501',
'Label=CR0258_U1, CellID=2581, LogicRNCID=501',
'Label=CR1245_U5, CellID=12455, LogicRNCID=501',
'Label=CR1245_U4, CellID=12454, LogicRNCID=501',
'Label=CR1245_U2, CellID=12452, LogicRNCID=501',
'Label=CR1377_U6, CellID=13776, LogicRNCID=501',
'Label=CR1377_U5, CellID=13775, LogicRNCID=501',
'Label=CR1377_U3, CellID=13773, LogicRNCID=501',
'Label=CR1377_U2, CellID=13772, LogicRNCID=501',
'Label=CR1380_U6, CellID=13806, LogicRNCID=501',
'Label=CR1380_U4, CellID=13804, LogicRNCID=501',
'Label=CR1380_U3, CellID=13803, LogicRNCID=501',
'Label=CR1380_U1, CellID=13801, LogicRNCID=501',
'Label=CR0463_U6, CellID=4636, LogicRNCID=501',
'Label=CR0463_U4, CellID=4634, LogicRNCID=501',
'Label=CR0463_U2, CellID=4632, LogicRNCID=501',
'Label=CR0463_U1, CellID=4631, LogicRNCID=501',
'Label=CR0478_U5, CellID=4785, LogicRNCID=501',
'Label=CR0478_U4, CellID=4784, LogicRNCID=501',
'Label=CR0478_U2, CellID=4782, LogicRNCID=501',
'Label=CR1371_U6, CellID=13716, LogicRNCID=501',
'Label=CR1371_U5, CellID=13715, LogicRNCID=501',
'Label=CR1371_U3, CellID=13713, LogicRNCID=501',
'Label=CR1371_U2, CellID=13712, LogicRNCID=501',
'Label=CR1376_U6, CellID=13766, LogicRNCID=501',
'Label=CR1376_U4, CellID=13764, LogicRNCID=501',
'Label=CR1376_U3, CellID=13763, LogicRNCID=501',
'Label=CR1376_U1, CellID=13761, LogicRNCID=501',
'Label=CR1383_U6, CellID=13836, LogicRNCID=501',
'Label=CR1383_U4, CellID=13834, LogicRNCID=501',
'Label=CR1383_U2, CellID=13832, LogicRNCID=501',
'Label=CR1383_U1, CellID=13831, LogicRNCID=501',
'Label=CR1365_U5, CellID=13655, LogicRNCID=501',
'Label=CR1365_U4, CellID=13654, LogicRNCID=501',
'Label=CR1365_U2, CellID=13652, LogicRNCID=501',
'Label=CR1366_U6, CellID=13666, LogicRNCID=501',
'Label=CR1366_U5, CellID=13665, LogicRNCID=501',
'Label=CR1366_U3, CellID=13663, LogicRNCID=501',
'Label=CR1366_U2, CellID=13662, LogicRNCID=501',
'Label=CR1382_U6, CellID=13826, LogicRNCID=501',
'Label=CR1382_U4, CellID=13824, LogicRNCID=501',
'Label=CR1382_U3, CellID=13823, LogicRNCID=501',
'Label=CR1382_U1, CellID=13821, LogicRNCID=501',
'Label=CR1343_U6, CellID=13436, LogicRNCID=501',
'Label=CR1343_U4, CellID=13434, LogicRNCID=501',
'Label=CR1343_U2, CellID=13432, LogicRNCID=501',
'Label=CR1343_U1, CellID=13431, LogicRNCID=501',
'Label=CR1344_U5, CellID=13445, LogicRNCID=501',
'Label=CR1344_U4, CellID=13444, LogicRNCID=501',
'Label=CR1344_U2, CellID=13442, LogicRNCID=501',
'Label=CR1415_U6, CellID=14156, LogicRNCID=501',
'Label=CR1415_U5, CellID=14155, LogicRNCID=501',
'Label=CR1415_U3, CellID=14153, LogicRNCID=501',
'Label=CR1415_U2, CellID=14152, LogicRNCID=501',
'Label=CR1411_U6, CellID=14116, LogicRNCID=501',
'Label=CR1411_U4, CellID=14114, LogicRNCID=501',
'Label=CR1411_U3, CellID=14113, LogicRNCID=501',
'Label=CR1411_U1, CellID=14111, LogicRNCID=501',
'Label=CR1238_U6, CellID=12386, LogicRNCID=501',
'Label=CR1238_U4, CellID=12384, LogicRNCID=501',
'Label=CR1238_U2, CellID=12382, LogicRNCID=501',
'Label=CR1238_U1, CellID=12381, LogicRNCID=501',
'Label=CR0469_U5, CellID=4695, LogicRNCID=501',
'Label=CR0469_U4, CellID=4694, LogicRNCID=501',
'Label=CR0469_U2, CellID=4692, LogicRNCID=501',
'Label=CR1211_U6, CellID=12116, LogicRNCID=501',
'Label=CR1211_U5, CellID=12115, LogicRNCID=501',
'Label=CR1211_U3, CellID=12113, LogicRNCID=501',
'Label=CR1211_U2, CellID=12112, LogicRNCID=501',
'Label=CR1310_U6, CellID=13106, LogicRNCID=501',
'Label=CR1310_U4, CellID=13104, LogicRNCID=501',
'Label=CR1310_U3, CellID=13103, LogicRNCID=501',
'Label=CR1310_U1, CellID=13101, LogicRNCID=501',
'Label=CR1334_U6, CellID=13346, LogicRNCID=501',
'Label=CR1334_U4, CellID=13344, LogicRNCID=501',
'Label=CR1334_U2, CellID=13342, LogicRNCID=501',
'Label=CR1334_U1, CellID=13341, LogicRNCID=501',
'Label=CR1319_U5, CellID=13195, LogicRNCID=501',
'Label=CR1319_U4, CellID=13194, LogicRNCID=501',
'Label=CR1319_U2, CellID=13192, LogicRNCID=501',
'Label=CR1247_U6, CellID=12476, LogicRNCID=501',
'Label=CR1247_U5, CellID=12475, LogicRNCID=501',
'Label=CR1247_U3, CellID=12473, LogicRNCID=501',
'Label=CR1247_U2, CellID=12472, LogicRNCID=501',
'Label=CR1321_U6, CellID=13216, LogicRNCID=501',
'Label=CR1321_U4, CellID=13214, LogicRNCID=501',
'Label=CR1321_U3, CellID=13213, LogicRNCID=501',
'Label=CR1321_U1, CellID=13211, LogicRNCID=501',
'Label=CR1308_U6, CellID=13086, LogicRNCID=501',
'Label=CR1308_U4, CellID=13084, LogicRNCID=501',
'Label=CR1308_U2, CellID=13082, LogicRNCID=501',
'Label=CR1308_U1, CellID=13081, LogicRNCID=501',
'Label=CR0479_U5, CellID=4795, LogicRNCID=501',
'Label=CR0479_U3, CellID=4793, LogicRNCID=501',
'Label=CR0758_U6, CellID=64556, LogicRNCID=501',
'Label=CR0758_U4, CellID=64554, LogicRNCID=501',
'Label=CR0758_U3, CellID=64553, LogicRNCID=501',
'Label=CR0758_U1, CellID=64551, LogicRNCID=501',
'Label=CR0477_U6, CellID=4776, LogicRNCID=501',
'Label=CR0477_U4, CellID=4774, LogicRNCID=501',
'Label=CR0477_U2, CellID=4772, LogicRNCID=501',
'Label=CR0477_U1, CellID=4771, LogicRNCID=501',
'Label=CR1254_U5, CellID=12545, LogicRNCID=501',
'Label=CR1254_U4, CellID=12544, LogicRNCID=501',
'Label=CR1254_U2, CellID=12542, LogicRNCID=501',
'Label=CR0466_U6, CellID=4666, LogicRNCID=501',
'Label=CR0466_U5, CellID=4665, LogicRNCID=501',
'Label=CR0466_U3, CellID=4663, LogicRNCID=501',
'Label=CR0466_U2, CellID=4662, LogicRNCID=501',
'Label=CR1217_U6, CellID=12176, LogicRNCID=501',
'Label=CR1217_U4, CellID=12174, LogicRNCID=501',
'Label=CR1217_U3, CellID=12173, LogicRNCID=501',
'Label=CR1217_U1, CellID=12171, LogicRNCID=501',
'Label=CR0468_U6, CellID=4686, LogicRNCID=501',
'Label=CR0468_U4, CellID=4684, LogicRNCID=501',
'Label=CR0468_U2, CellID=4682, LogicRNCID=501',
'Label=CR0468_U1, CellID=4681, LogicRNCID=501',
'Label=CR1218_U5, CellID=12185, LogicRNCID=501',
'Label=CR1218_U4, CellID=12184, LogicRNCID=501',
'Label=CR1218_U2, CellID=12182, LogicRNCID=501',
'Label=CR0457_U6, CellID=4576, LogicRNCID=501',
'Label=CR0457_U5, CellID=4575, LogicRNCID=501',
'Label=CR0457_U3, CellID=4573, LogicRNCID=501',
'Label=CR0457_U2, CellID=4572, LogicRNCID=501',
'Label=CR1256_U6, CellID=12566, LogicRNCID=501',
'Label=CR1256_U4, CellID=12564, LogicRNCID=501',
'Label=CR1256_U3, CellID=12563, LogicRNCID=501',
'Label=CR1256_U1, CellID=12561, LogicRNCID=501',
'Label=CR0474_U6, CellID=4746, LogicRNCID=501',
'Label=CR0474_U4, CellID=4744, LogicRNCID=501',
'Label=CR0474_U2, CellID=4742, LogicRNCID=501',
'Label=CR0474_U1, CellID=4741, LogicRNCID=501',
'Label=KR1680_U5, CellID=16805, LogicRNCID=501',
'Label=KR1680_U4, CellID=16804, LogicRNCID=501',
'Label=KR1680_U2, CellID=16802, LogicRNCID=501',
'Label=CR0192_U6, CellID=1926, LogicRNCID=501',
'Label=CR0192_U5, CellID=1925, LogicRNCID=501',
'Label=CR0192_U3, CellID=1923, LogicRNCID=501',
'Label=CR0192_U2, CellID=1922, LogicRNCID=501',
'Label=CR1103_U6, CellID=11036, LogicRNCID=501',
'Label=CR1103_U4, CellID=11034, LogicRNCID=501',
'Label=CR1103_U3, CellID=11033, LogicRNCID=501',
'Label=CR1103_U1, CellID=11031, LogicRNCID=501',
'Label=CR1279_U8, CellID=12798, LogicRNCID=501',
'Label=CR1279_U6, CellID=12796, LogicRNCID=501',
'Label=CR1279_U4, CellID=12794, LogicRNCID=501',
'Label=CR1279_U3, CellID=12793, LogicRNCID=501',
'Label=CR1279_U1, CellID=12791, LogicRNCID=501',
'Label=CR1101_U6, CellID=11016, LogicRNCID=501',
'Label=CR1101_U4, CellID=11014, LogicRNCID=501',
'Label=CR1101_U2, CellID=11012, LogicRNCID=501',
'Label=CR1101_U1, CellID=11011, LogicRNCID=501',
'Label=CR1580_U5, CellID=15805, LogicRNCID=501',
'Label=CR1580_U4, CellID=15804, LogicRNCID=501',
'Label=CR1580_U2, CellID=15802, LogicRNCID=501',
'Label=CR0475_U6, CellID=4756, LogicRNCID=501',
'Label=CR0475_U5, CellID=4755, LogicRNCID=501',
'Label=CR0475_U3, CellID=4753, LogicRNCID=501',
'Label=CR0475_U2, CellID=4752, LogicRNCID=501',
'Label=CR1228_U6, CellID=12286, LogicRNCID=501',
'Label=CR1228_U4, CellID=12284, LogicRNCID=501',
'Label=CR1228_U3, CellID=12283, LogicRNCID=501',
'Label=CR1228_U1, CellID=12281, LogicRNCID=501',
'Label=CR1301_U3, CellID=13013, LogicRNCID=501',
'Label=CR1301_U1, CellID=13011, LogicRNCID=501',
'Label=CR1293_U5, CellID=12935, LogicRNCID=501',
'Label=CR1293_U4, CellID=12934, LogicRNCID=501',
'Label=CR1293_U2, CellID=12932, LogicRNCID=501',
'Label=CR1293_U1, CellID=12931, LogicRNCID=501',
'Label=CR1248_U5, CellID=12485, LogicRNCID=501',
'Label=CR1248_U3, CellID=12483, LogicRNCID=501',
'Label=CR1248_U2, CellID=12482, LogicRNCID=501',
'Label=CR1249_U6, CellID=12496, LogicRNCID=501',
'Label=CR1249_U5, CellID=12495, LogicRNCID=501',
'Label=CR1249_U3, CellID=12493, LogicRNCID=501',
'Label=CR1249_U1, CellID=12491, LogicRNCID=501',
'Label=CR1295_U6, CellID=12956, LogicRNCID=501',
'Label=CR1295_U4, CellID=12954, LogicRNCID=501',
'Label=CR1295_U3, CellID=12953, LogicRNCID=501',
'Label=CR1295_U1, CellID=12951, LogicRNCID=501',
'Label=CR1301_U5, CellID=13015, LogicRNCID=501',
'Label=CR1301_U4, CellID=13014, LogicRNCID=501',
'Label=CR1285_U5, CellID=12855, LogicRNCID=501',
'Label=CR1285_U4, CellID=12854, LogicRNCID=501',
'Label=CR1285_U2, CellID=12852, LogicRNCID=501',
'Label=CR1087_U6, CellID=10876, LogicRNCID=501',
'Label=CR1087_U5, CellID=10875, LogicRNCID=501',
'Label=CR1087_U3, CellID=10873, LogicRNCID=501',
'Label=CR1087_U2, CellID=10872, LogicRNCID=501',
'Label=KR0108_U6, CellID=1086, LogicRNCID=501',
'Label=KR0108_U4, CellID=1084, LogicRNCID=501',
'Label=KR0108_U3, CellID=1083, LogicRNCID=501',
'Label=KR0108_U1, CellID=1081, LogicRNCID=501',
'Label=CR0101_U6, CellID=1016, LogicRNCID=501',
'Label=CR0101_U4, CellID=1014, LogicRNCID=501',
'Label=CR0101_U2, CellID=1012, LogicRNCID=501',
'Label=CR0101_U1, CellID=1011, LogicRNCID=501',
'Label=CR0841_U5, CellID=61015, LogicRNCID=501',
'Label=CR0841_U4, CellID=61014, LogicRNCID=501',
'Label=CR0841_U2, CellID=61012, LogicRNCID=501',
'Label=KR1854_U6, CellID=60596, LogicRNCID=501',
'Label=KR1854_U5, CellID=60595, LogicRNCID=501',
'Label=KR1854_U3, CellID=60593, LogicRNCID=501',
'Label=KR1854_U2, CellID=60592, LogicRNCID=501',
'Label=CR0610_U6, CellID=60626, LogicRNCID=501',
'Label=CR0610_U4, CellID=60624, LogicRNCID=501',
'Label=CR0610_U3, CellID=60623, LogicRNCID=501',
'Label=CR0610_U1, CellID=60621, LogicRNCID=501',
'Label=CR0577_U6, CellID=5776, LogicRNCID=501',
'Label=CR0577_U4, CellID=5774, LogicRNCID=501',
'Label=CR0577_U2, CellID=5772, LogicRNCID=501',
'Label=CR0577_U1, CellID=5771, LogicRNCID=501',
'Label=CR0831_U5, CellID=63985, LogicRNCID=501',
'Label=CR0831_U4, CellID=63984, LogicRNCID=501',
'Label=CR0831_U2, CellID=63982, LogicRNCID=501',
'Label=CR0847_U6, CellID=8476, LogicRNCID=501',
'Label=CR0847_U5, CellID=8475, LogicRNCID=501',
'Label=CR0847_U3, CellID=8473, LogicRNCID=501',
'Label=CR0847_U2, CellID=8472, LogicRNCID=501',
'Label=CR0788_U6, CellID=7886, LogicRNCID=501',
'Label=CR0788_U4, CellID=7884, LogicRNCID=501',
'Label=CR0788_U3, CellID=7883, LogicRNCID=501',
'Label=CR0788_U1, CellID=7881, LogicRNCID=501',
'Label=CR1857_U6, CellID=18576, LogicRNCID=501',
'Label=CR0792_U5, CellID=7925, LogicRNCID=501',
'Label=CR0792_U3, CellID=7923, LogicRNCID=501',
'Label=CR0792_U2, CellID=7922, LogicRNCID=501',
'Label=CR0309_U6, CellID=3096, LogicRNCID=501',
'Label=CR0309_U5, CellID=3095, LogicRNCID=501',
'Label=CR0309_U3, CellID=3093, LogicRNCID=501',
'Label=CR0309_U1, CellID=3091, LogicRNCID=501',
'Label=CR0791_U5, CellID=7915, LogicRNCID=501',
'Label=CR0791_U3, CellID=7913, LogicRNCID=501',
'Label=CR0791_U2, CellID=7912, LogicRNCID=501',
'Label=CR1888_U6, CellID=64086, LogicRNCID=501',
'Label=CR1888_U4, CellID=64084, LogicRNCID=501',
'Label=CR1888_U3, CellID=64083, LogicRNCID=501',
'Label=CR1888_U1, CellID=64081, LogicRNCID=501',
'Label=CR0571_U6, CellID=5716, LogicRNCID=501',
'Label=CR0571_U4, CellID=5714, LogicRNCID=501',
'Label=CR0571_U2, CellID=5712, LogicRNCID=501',
'Label=CR0571_U1, CellID=5711, LogicRNCID=501',
'Label=CR0813_U5, CellID=64905, LogicRNCID=501',
'Label=CR0813_U4, CellID=64904, LogicRNCID=501',
'Label=CR0813_U2, CellID=64902, LogicRNCID=501',
'Label=CR0804_U6, CellID=8046, LogicRNCID=501',
'Label=CR0804_U5, CellID=8045, LogicRNCID=501',
'Label=CR0804_U3, CellID=8043, LogicRNCID=501',
'Label=CR0804_U2, CellID=8042, LogicRNCID=501',
'Label=CR1690_U6, CellID=16906, LogicRNCID=501',
'Label=CR1690_U4, CellID=16904, LogicRNCID=501',
'Label=CR1690_U3, CellID=16903, LogicRNCID=501',
'Label=CR1690_U1, CellID=16901, LogicRNCID=501',
'Label=CR0811_U6, CellID=8116, LogicRNCID=501',
'Label=CR0811_U4, CellID=8114, LogicRNCID=501',
'Label=CR0811_U2, CellID=8112, LogicRNCID=501',
'Label=CR0811_U1, CellID=8111, LogicRNCID=501',
'Label=CR0798_U5, CellID=7985, LogicRNCID=501',
'Label=CR0798_U4, CellID=7984, LogicRNCID=501',
'Label=CR0798_U2, CellID=7982, LogicRNCID=501',
'Label=CR1518_U6, CellID=15186, LogicRNCID=501',
'Label=CR1518_U5, CellID=15185, LogicRNCID=501',
'Label=CR1518_U3, CellID=15183, LogicRNCID=501',
'Label=CR1518_U2, CellID=15182, LogicRNCID=501',
'Label=CR0557_U6, CellID=5576, LogicRNCID=501',
'Label=CR0557_U4, CellID=5574, LogicRNCID=501',
'Label=CR0557_U3, CellID=5573, LogicRNCID=501',
'Label=CR0557_U1, CellID=5571, LogicRNCID=501',
'Label=CR0408_U6, CellID=4086, LogicRNCID=501',
'Label=CR0408_U4, CellID=4084, LogicRNCID=501',
'Label=CR0408_U2, CellID=4082, LogicRNCID=501',
'Label=CR0408_U1, CellID=4081, LogicRNCID=501',
'Label=CR0794_U5, CellID=7945, LogicRNCID=501',
'Label=CR0794_U4, CellID=7944, LogicRNCID=501',
'Label=CR0794_U2, CellID=7942, LogicRNCID=501',
'Label=CR1517_U6, CellID=15176, LogicRNCID=501',
'Label=CR1517_U5, CellID=15175, LogicRNCID=501',
'Label=CR1517_U3, CellID=15173, LogicRNCID=501',
'Label=CR1517_U2, CellID=15172, LogicRNCID=501',
'Label=CR1686_U5, CellID=16865, LogicRNCID=501',
'Label=CR1686_U2, CellID=16862, LogicRNCID=501',
'Label=CR1686_U1, CellID=16861, LogicRNCID=501',
'Label=CR0305_U5, CellID=3055, LogicRNCID=501',
'Label=CR0305_U4, CellID=3054, LogicRNCID=501',
'Label=CR0305_U2, CellID=3052, LogicRNCID=501',
'Label=CR0304_U6, CellID=3046, LogicRNCID=501',
'Label=CR0304_U5, CellID=3045, LogicRNCID=501',
'Label=CR0304_U3, CellID=3043, LogicRNCID=501',
'Label=CR0304_U2, CellID=3042, LogicRNCID=501',
'Label=CR0406_U6, CellID=4066, LogicRNCID=501',
'Label=CR0406_U4, CellID=4064, LogicRNCID=501',
'Label=CR0406_U3, CellID=4063, LogicRNCID=501',
'Label=CR0406_U1, CellID=4061, LogicRNCID=501',
'Label=KR1020_U6, CellID=10206, LogicRNCID=501',
'Label=KR1020_U4, CellID=10204, LogicRNCID=501',
'Label=KR1020_U2, CellID=10202, LogicRNCID=501',
'Label=KR1020_U1, CellID=10201, LogicRNCID=501',
'Label=CR0768_U5, CellID=7685, LogicRNCID=501',
'Label=CR0768_U3, CellID=7683, LogicRNCID=501',
'Label=CR0767_U6, CellID=7676, LogicRNCID=501',
'Label=CR0767_U4, CellID=7674, LogicRNCID=501',
'Label=CR0767_U3, CellID=7673, LogicRNCID=501',
'Label=CR0767_U1, CellID=7671, LogicRNCID=501',
'Label=CR0476_U6, CellID=4766, LogicRNCID=501',
'Label=CR0476_U3, CellID=4763, LogicRNCID=501',
'Label=CR0306_U6, CellID=3066, LogicRNCID=501',
'Label=CR0306_U5, CellID=3065, LogicRNCID=501',
'Label=CR0306_U3, CellID=3063, LogicRNCID=501',
'Label=CR0306_U2, CellID=3062, LogicRNCID=501',
'Label=CR1005_U6, CellID=64656, LogicRNCID=501',
'Label=CR1005_U4, CellID=64654, LogicRNCID=501',
'Label=CR1005_U3, CellID=64653, LogicRNCID=501',
'Label=CR1005_U1, CellID=64651, LogicRNCID=501',
'Label=CR0499_U6, CellID=4996, LogicRNCID=501',
'Label=CR0499_U4, CellID=4994, LogicRNCID=501',
'Label=CR0499_U2, CellID=4992, LogicRNCID=501',
'Label=CR0499_U1, CellID=4991, LogicRNCID=501',
'Label=CR1106_U5, CellID=11065, LogicRNCID=501',
'Label=CR1106_U4, CellID=11064, LogicRNCID=501',
'Label=CR1106_U2, CellID=11062, LogicRNCID=501',
'Label=KR1001_U6, CellID=64756, LogicRNCID=501',
'Label=KR1001_U5, CellID=64755, LogicRNCID=501',
'Label=KR1001_U3, CellID=64753, LogicRNCID=501',
'Label=KR1001_U2, CellID=64752, LogicRNCID=501',
'Label=CR1026_U6, CellID=10266, LogicRNCID=501',
'Label=CR1026_U4, CellID=10264, LogicRNCID=501',
'Label=CR1026_U3, CellID=10263, LogicRNCID=501',
'Label=CR1026_U1, CellID=10261, LogicRNCID=501',
'Label=CR0993_U6, CellID=9936, LogicRNCID=501',
'Label=CR0993_U4, CellID=9934, LogicRNCID=501',
'Label=CR0993_U2, CellID=9932, LogicRNCID=501',
'Label=CR0993_U1, CellID=9931, LogicRNCID=501',
'Label=CR0591_U5, CellID=5915, LogicRNCID=501',
'Label=CR0591_U4, CellID=5914, LogicRNCID=501',
'Label=CR0591_U2, CellID=5912, LogicRNCID=501',
'Label=CR0567_U6, CellID=5676, LogicRNCID=501',
'Label=CR0567_U5, CellID=5675, LogicRNCID=501',
'Label=CR0567_U3, CellID=5673, LogicRNCID=501',
'Label=CR0567_U2, CellID=5672, LogicRNCID=501',
'Label=CR0445_U6, CellID=4456, LogicRNCID=501',
'Label=CR0445_U4, CellID=4454, LogicRNCID=501',
'Label=CR0445_U3, CellID=4453, LogicRNCID=501',
'Label=CR0445_U1, CellID=4451, LogicRNCID=501',
'Label=CR0310_U6, CellID=3106, LogicRNCID=501',
'Label=CR0310_U4, CellID=3104, LogicRNCID=501',
'Label=CR0310_U2, CellID=3102, LogicRNCID=501',
'Label=CR0310_U1, CellID=3101, LogicRNCID=501',
'Label=CR0190_U5, CellID=1905, LogicRNCID=501',
'Label=CR0190_U4, CellID=1904, LogicRNCID=501',
'Label=CR0190_U2, CellID=1902, LogicRNCID=501',
'Label=CR0572_U6, CellID=5726, LogicRNCID=501',
'Label=CR0572_U5, CellID=5725, LogicRNCID=501',
'Label=CR0572_U3, CellID=5723, LogicRNCID=501',
'Label=CR0572_U2, CellID=5722, LogicRNCID=501',
'Label=CR1078_U6, CellID=10786, LogicRNCID=501',
'Label=CR1078_U4, CellID=10784, LogicRNCID=501',
'Label=CR1078_U3, CellID=10783, LogicRNCID=501',
'Label=CR1078_U1, CellID=10781, LogicRNCID=501',
'Label=CR0947_U6, CellID=9476, LogicRNCID=501',
'Label=CR0947_U4, CellID=9474, LogicRNCID=501',
'Label=CR0947_U2, CellID=9472, LogicRNCID=501',
'Label=CR0947_U1, CellID=9471, LogicRNCID=501',
'Label=KR1850_U5, CellID=18505, LogicRNCID=501',
'Label=KR1850_U4, CellID=18504, LogicRNCID=501',
'Label=KR1850_U2, CellID=18502, LogicRNCID=501',
'Label=CR0411_U6, CellID=4116, LogicRNCID=501',
'Label=CR0411_U5, CellID=4115, LogicRNCID=501',
'Label=CR0411_U3, CellID=4113, LogicRNCID=501',
'Label=CR0411_U2, CellID=4112, LogicRNCID=501',
'Label=KR1564_U6, CellID=15646, LogicRNCID=501',
'Label=KR1564_U4, CellID=15644, LogicRNCID=501',
'Label=KR1564_U3, CellID=15643, LogicRNCID=501',
'Label=KR1564_U1, CellID=15641, LogicRNCID=501',
'Label=CR1862_U6, CellID=18626, LogicRNCID=501',
'Label=CR1862_U4, CellID=18624, LogicRNCID=501',
'Label=CR1862_U2, CellID=18622, LogicRNCID=501',
'Label=CR1862_U1, CellID=18621, LogicRNCID=501',
'Label=KR0125_U1, CellID=1251, LogicRNCID=501',
'Label=CR1762_U6, CellID=17626, LogicRNCID=501',
'Label=CR1762_U4, CellID=17624, LogicRNCID=501',
'Label=CR1762_U2, CellID=17622, LogicRNCID=501',
'Label=CR1762_U1, CellID=17621, LogicRNCID=501',
'Label=CR0403_U5, CellID=4035, LogicRNCID=501',
'Label=CR0403_U4, CellID=4034, LogicRNCID=501',
'Label=CR0403_U2, CellID=4032, LogicRNCID=501',
'Label=CR0640_U6, CellID=6406, LogicRNCID=501',
'Label=CR0640_U5, CellID=6405, LogicRNCID=501',
'Label=CR0640_U3, CellID=6403, LogicRNCID=501',
'Label=CR0640_U2, CellID=6402, LogicRNCID=501',
'Label=CR0448_U6, CellID=4486, LogicRNCID=501',
'Label=CR0448_U4, CellID=4484, LogicRNCID=501',
'Label=CR0448_U3, CellID=4483, LogicRNCID=501',
'Label=CR0448_U1, CellID=4481, LogicRNCID=501',
'Label=CR1105_U6, CellID=11056, LogicRNCID=501',
'Label=CR1105_U4, CellID=11054, LogicRNCID=501',
'Label=CR1105_U2, CellID=11052, LogicRNCID=501',
'Label=CR1105_U1, CellID=11051, LogicRNCID=501',
'Label=KR1861_U5, CellID=18615, LogicRNCID=501',
'Label=KR1861_U4, CellID=18614, LogicRNCID=501',
'Label=KR1861_U2, CellID=18612, LogicRNCID=501',
'Label=CR0174_U6, CellID=1746, LogicRNCID=501',
'Label=CR0174_U5, CellID=1745, LogicRNCID=501',
'Label=CR0174_U3, CellID=1743, LogicRNCID=501',
'Label=CR0174_U2, CellID=1742, LogicRNCID=501',
'Label=CR1104_U6, CellID=11046, LogicRNCID=501',
'Label=CR1104_U4, CellID=11044, LogicRNCID=501',
'Label=CR1104_U3, CellID=11043, LogicRNCID=501',
'Label=CR1104_U1, CellID=11041, LogicRNCID=501',
'Label=CR0143_U6, CellID=1436, LogicRNCID=501',
'Label=CR0143_U4, CellID=1434, LogicRNCID=501',
'Label=CR0143_U2, CellID=1432, LogicRNCID=501',
'Label=CR0143_U1, CellID=1431, LogicRNCID=501',
'Label=CR0447_U5, CellID=4475, LogicRNCID=501',
'Label=CR0447_U4, CellID=4474, LogicRNCID=501',
'Label=CR0447_U2, CellID=4472, LogicRNCID=501',
'Label=KR0642_U6, CellID=6426, LogicRNCID=501',
'Label=KR0642_U5, CellID=6425, LogicRNCID=501',
'Label=KR0642_U3, CellID=6423, LogicRNCID=501',
'Label=KR0642_U2, CellID=6422, LogicRNCID=501',
'Label=CR1100_U6, CellID=11006, LogicRNCID=501',
'Label=CR1100_U4, CellID=11004, LogicRNCID=501',
'Label=CR1100_U3, CellID=11003, LogicRNCID=501',
'Label=CR1100_U1, CellID=11001, LogicRNCID=501',
'Label=CR0410_U6, CellID=4106, LogicRNCID=501',
'Label=CR0410_U4, CellID=4104, LogicRNCID=501',
'Label=CR0410_U2, CellID=4102, LogicRNCID=501',
'Label=CR0410_U1, CellID=4101, LogicRNCID=501',
'Label=KR0664_U5, CellID=6645, LogicRNCID=501',
'Label=KR0664_U4, CellID=6644, LogicRNCID=501',
'Label=KR0664_U2, CellID=6642, LogicRNCID=501',
'Label=KR0732_U6, CellID=7326, LogicRNCID=501',
'Label=KR0732_U5, CellID=7325, LogicRNCID=501',
'Label=KR0732_U3, CellID=7323, LogicRNCID=501',
'Label=KR0732_U2, CellID=7322, LogicRNCID=501',
'Label=CR1188_U6, CellID=11886, LogicRNCID=501',
'Label=CR1188_U4, CellID=11884, LogicRNCID=501',
'Label=CR1188_U3, CellID=11883, LogicRNCID=501',
'Label=CR1188_U1, CellID=11881, LogicRNCID=501',
'Label=KR0624_U6, CellID=6246, LogicRNCID=501',
'Label=KR0624_U4, CellID=6244, LogicRNCID=501',
'Label=KR0624_U2, CellID=6242, LogicRNCID=501',
'Label=KR0624_U1, CellID=6241, LogicRNCID=501',
'Label=CR1881_U5, CellID=18815, LogicRNCID=501',
'Label=CR1881_U4, CellID=18814, LogicRNCID=501',
'Label=CR1881_U2, CellID=18812, LogicRNCID=501',
'Label=CR1886_U6, CellID=60666, LogicRNCID=501',
'Label=CR1886_U5, CellID=60665, LogicRNCID=501',
'Label=CR1886_U3, CellID=60663, LogicRNCID=501',
'Label=CR1886_U2, CellID=60662, LogicRNCID=501',
'Label=KR1855_U6, CellID=18556, LogicRNCID=501',
'Label=KR1855_U4, CellID=18554, LogicRNCID=501',
'Label=KR1855_U3, CellID=18553, LogicRNCID=501',
'Label=KR1855_U1, CellID=18551, LogicRNCID=501',
'Label=CR1678_U6, CellID=16786, LogicRNCID=501',
'Label=CR1678_U4, CellID=16784, LogicRNCID=501',
'Label=CR1678_U2, CellID=16782, LogicRNCID=501',
'Label=CR1678_U1, CellID=16781, LogicRNCID=501',
'Label=CR0668_U4, CellID=6684, LogicRNCID=501',
'Label=CR0668_U2, CellID=6682, LogicRNCID=501',
'Label=KR1437_U6, CellID=14376, LogicRNCID=501',
'Label=KR1437_U4, CellID=14374, LogicRNCID=501',
'Label=KR1437_U3, CellID=14373, LogicRNCID=501',
'Label=KR1437_U1, CellID=14371, LogicRNCID=501',
'Label=KR1439_U6, CellID=14396, LogicRNCID=501',
'Label=KR1439_U4, CellID=14394, LogicRNCID=501',
'Label=KR1439_U2, CellID=14392, LogicRNCID=501',
'Label=KR1439_U1, CellID=14391, LogicRNCID=501',
'Label=KR1429_U5, CellID=14295, LogicRNCID=501',
'Label=KR1429_U4, CellID=14294, LogicRNCID=501',
'Label=KR1429_U2, CellID=14292, LogicRNCID=501',
'Label=KR1433_U6, CellID=14336, LogicRNCID=501',
'Label=KR1433_U5, CellID=14335, LogicRNCID=501',
'Label=KR1433_U3, CellID=14333, LogicRNCID=501',
'Label=KR1433_U2, CellID=14332, LogicRNCID=501',
'Label=CR1759_U6, CellID=17596, LogicRNCID=501',
'Label=CR1759_U4, CellID=17594, LogicRNCID=501',
'Label=CR1759_U3, CellID=17593, LogicRNCID=501',
'Label=CR1759_U1, CellID=17591, LogicRNCID=501',
'Label=CR1122_U6, CellID=11226, LogicRNCID=501',
'Label=CR1122_U4, CellID=11224, LogicRNCID=501',
'Label=CR1122_U2, CellID=11222, LogicRNCID=501',
'Label=CR1122_U1, CellID=11221, LogicRNCID=501',
'Label=CR1200_U5, CellID=12005, LogicRNCID=501',
'Label=CR1200_U4, CellID=12004, LogicRNCID=501',
'Label=CR1200_U2, CellID=12002, LogicRNCID=501',
'Label=KR1860_U6, CellID=18606, LogicRNCID=501',
'Label=KR1860_U5, CellID=18605, LogicRNCID=501',
'Label=KR1860_U3, CellID=18603, LogicRNCID=501',
'Label=KR1860_U2, CellID=18602, LogicRNCID=501',
'Label=CR1885_U6, CellID=18856, LogicRNCID=501',
'Label=CR1885_U5, CellID=18855, LogicRNCID=501',
'Label=KR1860_U1, CellID=18601, LogicRNCID=501',
'Label=KR1860_U4, CellID=18604, LogicRNCID=501',
'Label=CR1200_U1, CellID=12001, LogicRNCID=501',
'Label=CR1200_U3, CellID=12003, LogicRNCID=501',
'Label=CR1200_U6, CellID=12006, LogicRNCID=501',
'Label=CR1122_U3, CellID=11223, LogicRNCID=501',
'Label=CR1122_U5, CellID=11225, LogicRNCID=501',
'Label=CR1759_U2, CellID=17592, LogicRNCID=501',
'Label=CR1759_U5, CellID=17595, LogicRNCID=501',
'Label=KR1433_U1, CellID=14331, LogicRNCID=501',
'Label=KR1433_U4, CellID=14334, LogicRNCID=501',
'Label=KR1429_U1, CellID=14291, LogicRNCID=501',
'Label=KR1429_U3, CellID=14293, LogicRNCID=501',
'Label=KR1429_U6, CellID=14296, LogicRNCID=501',
'Label=KR1439_U3, CellID=14393, LogicRNCID=501',
'Label=KR1439_U5, CellID=14395, LogicRNCID=501',
'Label=KR1437_U2, CellID=14372, LogicRNCID=501',
'Label=KR1437_U5, CellID=14375, LogicRNCID=501',
'Label=CR0668_U1, CellID=6681, LogicRNCID=501',
'Label=CR0668_U5, CellID=6685, LogicRNCID=501',
'Label=CR1678_U3, CellID=16783, LogicRNCID=501',
'Label=CR1678_U5, CellID=16785, LogicRNCID=501',
'Label=KR1855_U2, CellID=18552, LogicRNCID=501',
'Label=KR1855_U5, CellID=18555, LogicRNCID=501',
'Label=CR1886_U1, CellID=60661, LogicRNCID=501',
'Label=CR1886_U4, CellID=60664, LogicRNCID=501',
'Label=CR1881_U1, CellID=18811, LogicRNCID=501',
'Label=CR1881_U3, CellID=18813, LogicRNCID=501',
'Label=CR1881_U6, CellID=18816, LogicRNCID=501',
'Label=KR0624_U3, CellID=6243, LogicRNCID=501',
'Label=KR0624_U5, CellID=6245, LogicRNCID=501',
'Label=CR1188_U2, CellID=11882, LogicRNCID=501',
'Label=CR1188_U5, CellID=11885, LogicRNCID=501',
'Label=KR0732_U1, CellID=7321, LogicRNCID=501',
'Label=KR0732_U4, CellID=7324, LogicRNCID=501',
'Label=KR0664_U1, CellID=6641, LogicRNCID=501',
'Label=KR0664_U3, CellID=6643, LogicRNCID=501',
'Label=KR0664_U6, CellID=6646, LogicRNCID=501',
'Label=CR0410_U3, CellID=4103, LogicRNCID=501',
'Label=CR0410_U5, CellID=4105, LogicRNCID=501',
'Label=CR1100_U2, CellID=11002, LogicRNCID=501',
'Label=CR1100_U5, CellID=11005, LogicRNCID=501',
'Label=KR0642_U1, CellID=6421, LogicRNCID=501',
'Label=KR0642_U4, CellID=6424, LogicRNCID=501',
'Label=CR0447_U1, CellID=4471, LogicRNCID=501',
'Label=CR0447_U3, CellID=4473, LogicRNCID=501',
'Label=CR0447_U6, CellID=4476, LogicRNCID=501',
'Label=CR0143_U3, CellID=1433, LogicRNCID=501',
'Label=CR0143_U5, CellID=1435, LogicRNCID=501',
'Label=CR1104_U2, CellID=11042, LogicRNCID=501',
'Label=CR1104_U5, CellID=11045, LogicRNCID=501',
'Label=CR0174_U1, CellID=1741, LogicRNCID=501',
'Label=CR0174_U4, CellID=1744, LogicRNCID=501',
'Label=KR1861_U1, CellID=18611, LogicRNCID=501',
'Label=KR1861_U3, CellID=18613, LogicRNCID=501',
'Label=KR1861_U6, CellID=18616, LogicRNCID=501',
'Label=CR1105_U3, CellID=11053, LogicRNCID=501',
'Label=CR1105_U5, CellID=11055, LogicRNCID=501',
'Label=CR0448_U2, CellID=4482, LogicRNCID=501',
'Label=CR0448_U5, CellID=4485, LogicRNCID=501',
'Label=CR0640_U1, CellID=6401, LogicRNCID=501',
'Label=CR0640_U4, CellID=6404, LogicRNCID=501',
'Label=CR0403_U1, CellID=4031, LogicRNCID=501',
'Label=CR0403_U3, CellID=4033, LogicRNCID=501',
'Label=CR0403_U6, CellID=4036, LogicRNCID=501',
'Label=CR1762_U3, CellID=17623, LogicRNCID=501',
'Label=CR1762_U5, CellID=17625, LogicRNCID=501',
'Label=KR0125_U4, CellID=1254, LogicRNCID=501',
'Label=CR1862_U3, CellID=18623, LogicRNCID=501',
'Label=CR1862_U5, CellID=18625, LogicRNCID=501',
'Label=KR1564_U2, CellID=15642, LogicRNCID=501',
'Label=KR1564_U5, CellID=15645, LogicRNCID=501',
'Label=CR0411_U1, CellID=4111, LogicRNCID=501',
'Label=CR0411_U4, CellID=4114, LogicRNCID=501',
'Label=KR1850_U1, CellID=18501, LogicRNCID=501',
'Label=KR1850_U3, CellID=18503, LogicRNCID=501',
'Label=KR1850_U6, CellID=18506, LogicRNCID=501',
'Label=CR0947_U3, CellID=9473, LogicRNCID=501',
'Label=CR0947_U5, CellID=9475, LogicRNCID=501',
'Label=CR1078_U2, CellID=10782, LogicRNCID=501',
'Label=CR1078_U5, CellID=10785, LogicRNCID=501',
'Label=CR0572_U1, CellID=5721, LogicRNCID=501',
'Label=CR0572_U4, CellID=5724, LogicRNCID=501',
'Label=CR0190_U1, CellID=1901, LogicRNCID=501',
'Label=CR0190_U3, CellID=1903, LogicRNCID=501',
'Label=CR0190_U6, CellID=1906, LogicRNCID=501',
'Label=CR0310_U3, CellID=3103, LogicRNCID=501',
'Label=CR0310_U5, CellID=3105, LogicRNCID=501',
'Label=CR0445_U2, CellID=4452, LogicRNCID=501',
'Label=CR0445_U5, CellID=4455, LogicRNCID=501',
'Label=CR0567_U1, CellID=5671, LogicRNCID=501',
'Label=CR0567_U4, CellID=5674, LogicRNCID=501',
'Label=CR0591_U1, CellID=5911, LogicRNCID=501',
'Label=CR0591_U3, CellID=5913, LogicRNCID=501',
'Label=CR0591_U6, CellID=5916, LogicRNCID=501',
'Label=CR0993_U3, CellID=9933, LogicRNCID=501',
'Label=CR0993_U5, CellID=9935, LogicRNCID=501',
'Label=CR1026_U2, CellID=10262, LogicRNCID=501',
'Label=CR1026_U5, CellID=10265, LogicRNCID=501',
'Label=KR1001_U1, CellID=64751, LogicRNCID=501',
'Label=KR1001_U4, CellID=64754, LogicRNCID=501',
'Label=CR1106_U1, CellID=11061, LogicRNCID=501',
'Label=CR1106_U3, CellID=11063, LogicRNCID=501',
'Label=CR1106_U6, CellID=11066, LogicRNCID=501',
'Label=CR0499_U3, CellID=4993, LogicRNCID=501',
'Label=CR0499_U5, CellID=4995, LogicRNCID=501',
'Label=CR1005_U2, CellID=64652, LogicRNCID=501',
'Label=CR1005_U5, CellID=64655, LogicRNCID=501',
'Label=CR0306_U1, CellID=3061, LogicRNCID=501',
'Label=CR0306_U4, CellID=3064, LogicRNCID=501',
'Label=CR0476_U2, CellID=4762, LogicRNCID=501',
'Label=CR0476_U5, CellID=4765, LogicRNCID=501',
'Label=CR0767_U2, CellID=7672, LogicRNCID=501',
'Label=CR0767_U5, CellID=7675, LogicRNCID=501',
'Label=CR0768_U2, CellID=7682, LogicRNCID=501',
'Label=CR0768_U6, CellID=7686, LogicRNCID=501',
'Label=KR1020_U3, CellID=10203, LogicRNCID=501',
'Label=KR1020_U5, CellID=10205, LogicRNCID=501',
'Label=CR0406_U2, CellID=4062, LogicRNCID=501',
'Label=CR0406_U5, CellID=4065, LogicRNCID=501',
'Label=CR0304_U1, CellID=3041, LogicRNCID=501',
'Label=CR0304_U4, CellID=3044, LogicRNCID=501',
'Label=CR0305_U1, CellID=3051, LogicRNCID=501',
'Label=CR0305_U3, CellID=3053, LogicRNCID=501',
'Label=CR0305_U6, CellID=3056, LogicRNCID=501',
'Label=CR1686_U4, CellID=16864, LogicRNCID=501',
'Label=CR1517_U1, CellID=15171, LogicRNCID=501',
'Label=CR1517_U4, CellID=15174, LogicRNCID=501',
'Label=CR0794_U1, CellID=7941, LogicRNCID=501',
'Label=CR0794_U3, CellID=7943, LogicRNCID=501',
'Label=CR0794_U6, CellID=7946, LogicRNCID=501',
'Label=CR0408_U3, CellID=4083, LogicRNCID=501',
'Label=CR0408_U5, CellID=4085, LogicRNCID=501',
'Label=CR0557_U2, CellID=5572, LogicRNCID=501',
'Label=CR0557_U5, CellID=5575, LogicRNCID=501',
'Label=CR1518_U1, CellID=15181, LogicRNCID=501',
'Label=CR1518_U4, CellID=15184, LogicRNCID=501',
'Label=CR0798_U1, CellID=7981, LogicRNCID=501',
'Label=CR0798_U3, CellID=7983, LogicRNCID=501',
'Label=CR0798_U6, CellID=7986, LogicRNCID=501',
'Label=CR0811_U3, CellID=8113, LogicRNCID=501',
'Label=CR0811_U5, CellID=8115, LogicRNCID=501',
'Label=CR1690_U2, CellID=16902, LogicRNCID=501',
'Label=CR1690_U5, CellID=16905, LogicRNCID=501',
'Label=CR0804_U1, CellID=8041, LogicRNCID=501',
'Label=CR0804_U4, CellID=8044, LogicRNCID=501',
'Label=CR0813_U1, CellID=64901, LogicRNCID=501',
'Label=CR0813_U3, CellID=64903, LogicRNCID=501',
'Label=CR0813_U6, CellID=64906, LogicRNCID=501',
'Label=CR0571_U3, CellID=5713, LogicRNCID=501',
'Label=CR0571_U5, CellID=5715, LogicRNCID=501',
'Label=CR1888_U2, CellID=64082, LogicRNCID=501',
'Label=CR1888_U5, CellID=64085, LogicRNCID=501',
'Label=CR0791_U1, CellID=7911, LogicRNCID=501',
'Label=CR0791_U4, CellID=7914, LogicRNCID=501',
'Label=CR0309_U2, CellID=3092, LogicRNCID=501',
'Label=CR0309_U4, CellID=3094, LogicRNCID=501',
'Label=CR0792_U1, CellID=7921, LogicRNCID=501',
'Label=CR0792_U4, CellID=7924, LogicRNCID=501',
'Label=CR0792_U6, CellID=7926, LogicRNCID=501',
'Label=CR0788_U2, CellID=7882, LogicRNCID=501',
'Label=CR0788_U5, CellID=7885, LogicRNCID=501',
'Label=CR0847_U1, CellID=8471, LogicRNCID=501',
'Label=CR0847_U4, CellID=8474, LogicRNCID=501',
'Label=CR0831_U1, CellID=63981, LogicRNCID=501',
'Label=CR0831_U3, CellID=63983, LogicRNCID=501',
'Label=CR0831_U6, CellID=63986, LogicRNCID=501',
'Label=CR0577_U3, CellID=5773, LogicRNCID=501',
'Label=CR0577_U5, CellID=5775, LogicRNCID=501',
'Label=CR0610_U2, CellID=60622, LogicRNCID=501',
'Label=CR0610_U5, CellID=60625, LogicRNCID=501',
'Label=KR1854_U1, CellID=60591, LogicRNCID=501',
'Label=KR1854_U4, CellID=60594, LogicRNCID=501',
'Label=CR0841_U1, CellID=61011, LogicRNCID=501',
'Label=CR0841_U3, CellID=61013, LogicRNCID=501',
'Label=CR0841_U6, CellID=61016, LogicRNCID=501',
'Label=CR0101_U3, CellID=1013, LogicRNCID=501',
'Label=CR0101_U5, CellID=1015, LogicRNCID=501',
'Label=KR0108_U2, CellID=1082, LogicRNCID=501',
'Label=KR0108_U5, CellID=1085, LogicRNCID=501',
'Label=CR1087_U1, CellID=10871, LogicRNCID=501',
'Label=CR1087_U4, CellID=10874, LogicRNCID=501',
'Label=CR1285_U1, CellID=12851, LogicRNCID=501',
'Label=CR1285_U3, CellID=12853, LogicRNCID=501',
'Label=CR1285_U6, CellID=12856, LogicRNCID=501',
'Label=CR1301_U6, CellID=13016, LogicRNCID=501',
'Label=CR1295_U2, CellID=12952, LogicRNCID=501',
'Label=CR1295_U5, CellID=12955, LogicRNCID=501',
'Label=CR1249_U2, CellID=12492, LogicRNCID=501',
'Label=CR1249_U4, CellID=12494, LogicRNCID=501',
'Label=CR1248_U1, CellID=12481, LogicRNCID=501',
'Label=CR1248_U4, CellID=12484, LogicRNCID=501',
'Label=CR1248_U6, CellID=12486, LogicRNCID=501',
'Label=CR1293_U3, CellID=12933, LogicRNCID=501',
'Label=CR1293_U6, CellID=12936, LogicRNCID=501',
'Label=CR1301_U2, CellID=13012, LogicRNCID=501',
'Label=CR1228_U2, CellID=12282, LogicRNCID=501',
'Label=CR1228_U5, CellID=12285, LogicRNCID=501',
'Label=CR0475_U1, CellID=4751, LogicRNCID=501',
'Label=CR0475_U4, CellID=4754, LogicRNCID=501',
'Label=CR1580_U1, CellID=15801, LogicRNCID=501',
'Label=CR1580_U3, CellID=15803, LogicRNCID=501',
'Label=CR1580_U6, CellID=15806, LogicRNCID=501',
'Label=CR1101_U3, CellID=11013, LogicRNCID=501',
'Label=CR1101_U5, CellID=11015, LogicRNCID=501',
'Label=CR1279_U2, CellID=12792, LogicRNCID=501',
'Label=CR1279_U5, CellID=12795, LogicRNCID=501',
'Label=CR1279_U7, CellID=12797, LogicRNCID=501',
'Label=CR1103_U2, CellID=11032, LogicRNCID=501',
'Label=CR1103_U5, CellID=11035, LogicRNCID=501',
'Label=CR0192_U1, CellID=1921, LogicRNCID=501',
'Label=CR0192_U4, CellID=1924, LogicRNCID=501',
'Label=KR1680_U1, CellID=16801, LogicRNCID=501',
'Label=KR1680_U3, CellID=16803, LogicRNCID=501',
'Label=KR1680_U6, CellID=16806, LogicRNCID=501',
'Label=CR0474_U3, CellID=4743, LogicRNCID=501',
'Label=CR0474_U5, CellID=4745, LogicRNCID=501',
'Label=CR1256_U2, CellID=12562, LogicRNCID=501',
'Label=CR1256_U5, CellID=12565, LogicRNCID=501',
'Label=CR0457_U1, CellID=4571, LogicRNCID=501',
'Label=CR0457_U4, CellID=4574, LogicRNCID=501',
'Label=CR1218_U1, CellID=12181, LogicRNCID=501',
'Label=CR1218_U3, CellID=12183, LogicRNCID=501',
'Label=CR1218_U6, CellID=12186, LogicRNCID=501',
'Label=CR0468_U3, CellID=4683, LogicRNCID=501',
'Label=CR0468_U5, CellID=4685, LogicRNCID=501',
'Label=CR1217_U2, CellID=12172, LogicRNCID=501',
'Label=CR1217_U5, CellID=12175, LogicRNCID=501',
'Label=CR0466_U1, CellID=4661, LogicRNCID=501',
'Label=CR0466_U4, CellID=4664, LogicRNCID=501',
'Label=CR1254_U1, CellID=12541, LogicRNCID=501',
'Label=CR1254_U3, CellID=12543, LogicRNCID=501',
'Label=CR1254_U6, CellID=12546, LogicRNCID=501',
'Label=CR0477_U3, CellID=4773, LogicRNCID=501',
'Label=CR0477_U5, CellID=4775, LogicRNCID=501',
'Label=CR0758_U2, CellID=64552, LogicRNCID=501',
'Label=CR0758_U5, CellID=64555, LogicRNCID=501',
'Label=CR0479_U2, CellID=4792, LogicRNCID=501',
'Label=CR0479_U6, CellID=4796, LogicRNCID=501',
'Label=CR1308_U3, CellID=13083, LogicRNCID=501',
'Label=CR1308_U5, CellID=13085, LogicRNCID=501',
'Label=CR1321_U2, CellID=13212, LogicRNCID=501',
'Label=CR1321_U5, CellID=13215, LogicRNCID=501',
'Label=CR1247_U1, CellID=12471, LogicRNCID=501',
'Label=CR1247_U4, CellID=12474, LogicRNCID=501',
'Label=CR1319_U1, CellID=13191, LogicRNCID=501',
'Label=CR1319_U3, CellID=13193, LogicRNCID=501',
'Label=CR1319_U6, CellID=13196, LogicRNCID=501',
'Label=CR1334_U3, CellID=13343, LogicRNCID=501',
'Label=CR1334_U5, CellID=13345, LogicRNCID=501',
'Label=CR1310_U2, CellID=13102, LogicRNCID=501',
'Label=CR1310_U5, CellID=13105, LogicRNCID=501',
'Label=CR1211_U1, CellID=12111, LogicRNCID=501',
'Label=CR1211_U4, CellID=12114, LogicRNCID=501',
'Label=CR0469_U1, CellID=4691, LogicRNCID=501',
'Label=CR0469_U3, CellID=4693, LogicRNCID=501',
'Label=CR0469_U6, CellID=4696, LogicRNCID=501',
'Label=CR1238_U3, CellID=12383, LogicRNCID=501',
'Label=CR1238_U5, CellID=12385, LogicRNCID=501',
'Label=CR1411_U2, CellID=14112, LogicRNCID=501',
'Label=CR1411_U5, CellID=14115, LogicRNCID=501',
'Label=CR1415_U1, CellID=14151, LogicRNCID=501',
'Label=CR1415_U4, CellID=14154, LogicRNCID=501',
'Label=CR1344_U1, CellID=13441, LogicRNCID=501',
'Label=CR1344_U3, CellID=13443, LogicRNCID=501',
'Label=CR1344_U6, CellID=13446, LogicRNCID=501',
'Label=CR1343_U3, CellID=13433, LogicRNCID=501',
'Label=CR1343_U5, CellID=13435, LogicRNCID=501',
'Label=CR1382_U2, CellID=13822, LogicRNCID=501',
'Label=CR1382_U5, CellID=13825, LogicRNCID=501',
'Label=CR1366_U1, CellID=13661, LogicRNCID=501',
'Label=CR1366_U4, CellID=13664, LogicRNCID=501',
'Label=CR1365_U1, CellID=13651, LogicRNCID=501',
'Label=CR1365_U3, CellID=13653, LogicRNCID=501',
'Label=CR1365_U6, CellID=13656, LogicRNCID=501',
'Label=CR1383_U3, CellID=13833, LogicRNCID=501',
'Label=CR1383_U5, CellID=13835, LogicRNCID=501',
'Label=CR1376_U2, CellID=13762, LogicRNCID=501',
'Label=CR1376_U5, CellID=13765, LogicRNCID=501',
'Label=CR1371_U1, CellID=13711, LogicRNCID=501',
'Label=CR1371_U4, CellID=13714, LogicRNCID=501',
'Label=CR0478_U1, CellID=4781, LogicRNCID=501',
'Label=CR0478_U3, CellID=4783, LogicRNCID=501',
'Label=CR0478_U6, CellID=4786, LogicRNCID=501',
'Label=CR0463_U3, CellID=4633, LogicRNCID=501',
'Label=CR0463_U5, CellID=4635, LogicRNCID=501',
'Label=CR1380_U2, CellID=13802, LogicRNCID=501',
'Label=CR1380_U5, CellID=13805, LogicRNCID=501',
'Label=CR1377_U1, CellID=13771, LogicRNCID=501',
'Label=CR1377_U4, CellID=13774, LogicRNCID=501',
'Label=CR1245_U1, CellID=12451, LogicRNCID=501',
'Label=CR1245_U3, CellID=12453, LogicRNCID=501',
'Label=CR1245_U6, CellID=12456, LogicRNCID=501',
'Label=CR0258_U3, CellID=2583, LogicRNCID=501',
'Label=CR0258_U5, CellID=2585, LogicRNCID=501',
'Label=CR0264_U2, CellID=2642, LogicRNCID=501',
'Label=CR0264_U5, CellID=2645, LogicRNCID=501',
'Label=KR1851_U1, CellID=18511, LogicRNCID=501',
'Label=KR1851_U4, CellID=18514, LogicRNCID=501',
'Label=CR0472_U1, CellID=4721, LogicRNCID=501',
'Label=CR0472_U3, CellID=4723, LogicRNCID=501',
'Label=CR0472_U6, CellID=4726, LogicRNCID=501',
'Label=CR0465_U3, CellID=4653, LogicRNCID=501',
'Label=CR0465_U5, CellID=4655, LogicRNCID=501',
'Label=CR1102_U2, CellID=11022, LogicRNCID=501',
'Label=CR1102_U5, CellID=11025, LogicRNCID=501',
'Label=CR0592_U1, CellID=5921, LogicRNCID=501',
'Label=CR0592_U4, CellID=5924, LogicRNCID=501',
'Label=CR1408_U1, CellID=14081, LogicRNCID=501',
'Label=CR1408_U3, CellID=14083, LogicRNCID=501',
'Label=CR1408_U6, CellID=14086, LogicRNCID=501',
'Label=CR1390_U3, CellID=13903, LogicRNCID=501',
'Label=CR1390_U5, CellID=13905, LogicRNCID=501',
'Label=CR0473_U2, CellID=4732, LogicRNCID=501',
'Label=CR0473_U5, CellID=4735, LogicRNCID=501',
'Label=CR0776_U1, CellID=7761, LogicRNCID=501',
'Label=CR0776_U4, CellID=7764, LogicRNCID=501',
'Label=CR1329_U1, CellID=13291, LogicRNCID=501',
'Label=CR1329_U3, CellID=13293, LogicRNCID=501',
'Label=CR1329_U6, CellID=13296, LogicRNCID=501',
'Label=CR1749_U3, CellID=17493, LogicRNCID=501',
'Label=CR1749_U5, CellID=17495, LogicRNCID=501',
'Label=CR0456_U2, CellID=4562, LogicRNCID=501',
'Label=CR0456_U5, CellID=4565, LogicRNCID=501',
'Label=CR1266_U1, CellID=12661, LogicRNCID=501',
'Label=CR1266_U4, CellID=12664, LogicRNCID=501',
'Label=CR0297_U1, CellID=2971, LogicRNCID=501',
'Label=CR0297_U3, CellID=2973, LogicRNCID=501',
'Label=CR0297_U6, CellID=2976, LogicRNCID=501',
'Label=KR1852_U3, CellID=18523, LogicRNCID=501',
'Label=KR1852_U5, CellID=18525, LogicRNCID=501',
'Label=KR1858_U2, CellID=18582, LogicRNCID=501',
'Label=KR1858_U5, CellID=18585, LogicRNCID=501',
'Label=CR1857_U1, CellID=18571, LogicRNCID=501',
'Label=CR1857_U4, CellID=18574, LogicRNCID=501',
'Label=CR1885_U4, CellID=18854, LogicRNCID=501',
'Label=CR1885_U3, CellID=18853, LogicRNCID=501',
'Label=CR1885_U1, CellID=18851, LogicRNCID=501',
'Label=CR0485_U6, CellID=4856, LogicRNCID=501',
'Label=CR0485_U4, CellID=4854, LogicRNCID=501',
'Label=CR0485_U2, CellID=4852, LogicRNCID=501',
'Label=CR0485_U1, CellID=4851, LogicRNCID=501',
'Label=CR1209_U5, CellID=12095, LogicRNCID=501',
'Label=CR1209_U4, CellID=12094, LogicRNCID=501',
'Label=CR1209_U2, CellID=12092, LogicRNCID=501',
'Label=CR1128_U6, CellID=11286, LogicRNCID=501',
'Label=CR1128_U5, CellID=11285, LogicRNCID=501',
'Label=CR1128_U3, CellID=11283, LogicRNCID=501',
'Label=CR1128_U2, CellID=11282, LogicRNCID=501',
'Label=CR1883_U6, CellID=65466, LogicRNCID=501',
'Label=CR1883_U4, CellID=65464, LogicRNCID=501',
'Label=CR1883_U3, CellID=65463, LogicRNCID=501',
'Label=CR1883_U1, CellID=65461, LogicRNCID=501',
'Label=CR0454_U6, CellID=4546, LogicRNCID=501',
'Label=CR0454_U4, CellID=4544, LogicRNCID=501',
'Label=CR0454_U2, CellID=4542, LogicRNCID=501',
'Label=CR0454_U1, CellID=4541, LogicRNCID=501',
'Label=CR0484_U5, CellID=4845, LogicRNCID=501',
'Label=CR0484_U4, CellID=4844, LogicRNCID=501',
'Label=CR0484_U2, CellID=4842, LogicRNCID=501',
'Label=CR1107_U5, CellID=11075, LogicRNCID=501',
'Label=CR1107_U4, CellID=11074, LogicRNCID=501',
'Label=CR1107_U1, CellID=11071, LogicRNCID=501',
'Label=KR0632_U6, CellID=6326, LogicRNCID=501',
'Label=KR0632_U4, CellID=6324, LogicRNCID=501',
'Label=KR0632_U2, CellID=6322, LogicRNCID=501',
'Label=KR0632_U1, CellID=6321, LogicRNCID=501',
'Label=KR0631_U5, CellID=6315, LogicRNCID=501',
'Label=KR0631_U4, CellID=6314, LogicRNCID=501',
'Label=KR0631_U2, CellID=6312, LogicRNCID=501',
'Label=KR0633_U6, CellID=6336, LogicRNCID=501',
'Label=KR0633_U5, CellID=6335, LogicRNCID=501',
'Label=KR0633_U3, CellID=6333, LogicRNCID=501',
'Label=KR0633_U2, CellID=6332, LogicRNCID=501',
'Label=CR0486_U6, CellID=4866, LogicRNCID=501',
'Label=CR0486_U4, CellID=4864, LogicRNCID=501',
'Label=CR0486_U3, CellID=4863, LogicRNCID=501',
'Label=CR0486_U1, CellID=4861, LogicRNCID=501',
'Label=CR1756_U6, CellID=17566, LogicRNCID=501',
'Label=CR1756_U4, CellID=17564, LogicRNCID=501',
'Label=CR1756_U2, CellID=17562, LogicRNCID=501',
'Label=CR1756_U1, CellID=17561, LogicRNCID=501',
'Label=KR0553_U5, CellID=5535, LogicRNCID=501',
'Label=KR0553_U4, CellID=5534, LogicRNCID=501',
'Label=KR0553_U2, CellID=5532, LogicRNCID=501',
'Label=KR1870_U6, CellID=18706, LogicRNCID=501',
'Label=KR1870_U5, CellID=18705, LogicRNCID=501',
'Label=KR1870_U3, CellID=18703, LogicRNCID=501',
'Label=KR1870_U2, CellID=18702, LogicRNCID=501',
'Label=CR1077_U6, CellID=10776, LogicRNCID=501',
'Label=CR1077_U4, CellID=10774, LogicRNCID=501',
'Label=CR1077_U3, CellID=10773, LogicRNCID=501',
'Label=CR1077_U1, CellID=10771, LogicRNCID=501',
'Label=KR1869_U6, CellID=18696, LogicRNCID=501',
'Label=KR1869_U4, CellID=18694, LogicRNCID=501',
'Label=KR1869_U2, CellID=18692, LogicRNCID=501',
'Label=KR1869_U1, CellID=18691, LogicRNCID=501',
'Label=CR1051_U5, CellID=10515, LogicRNCID=501',
'Label=CR1051_U4, CellID=10514, LogicRNCID=501',
'Label=CR1051_U2, CellID=10512, LogicRNCID=501',
'Label=CR0507_U6, CellID=5076, LogicRNCID=501',
'Label=CR0507_U5, CellID=5075, LogicRNCID=501',
'Label=CR0507_U3, CellID=5073, LogicRNCID=501',
'Label=CR0507_U2, CellID=5072, LogicRNCID=501',
'Label=KR0778_U6, CellID=7786, LogicRNCID=501',
'Label=KR0778_U4, CellID=7784, LogicRNCID=501',
'Label=KR0778_U3, CellID=7783, LogicRNCID=501',
'Label=KR0778_U1, CellID=7781, LogicRNCID=501',
'Label=CR1894_U6, CellID=63096, LogicRNCID=501',
'Label=CR1894_U4, CellID=63094, LogicRNCID=501',
'Label=CR1894_U2, CellID=63092, LogicRNCID=501',
'Label=CR1894_U1, CellID=63091, LogicRNCID=501',
'Label=CR0725_U5, CellID=7255, LogicRNCID=501',
'Label=CR0725_U4, CellID=7254, LogicRNCID=501',
'Label=CR0725_U2, CellID=7252, LogicRNCID=501',
'Label=CR1169_U6, CellID=11696, LogicRNCID=501',
'Label=CR1169_U5, CellID=11695, LogicRNCID=501',
'Label=CR1169_U3, CellID=11693, LogicRNCID=501',
'Label=CR1169_U2, CellID=11692, LogicRNCID=501',
'Label=CR0184_U6, CellID=1846, LogicRNCID=501',
'Label=CR0184_U4, CellID=1844, LogicRNCID=501',
'Label=CR0184_U3, CellID=1843, LogicRNCID=501',
'Label=CR0184_U1, CellID=1841, LogicRNCID=501',
'Label=CR1160_U6, CellID=11606, LogicRNCID=501',
'Label=CR1160_U4, CellID=11604, LogicRNCID=501',
'Label=CR1160_U2, CellID=11602, LogicRNCID=501',
'Label=CR1160_U1, CellID=11601, LogicRNCID=501',
'Label=CR0414_U5, CellID=4145, LogicRNCID=501',
'Label=CR0414_U4, CellID=4144, LogicRNCID=501',
'Label=CR0414_U2, CellID=4142, LogicRNCID=501',
'Label=CR1142_U6, CellID=11426, LogicRNCID=501',
'Label=CR1142_U5, CellID=11425, LogicRNCID=501',
'Label=CR1142_U3, CellID=11423, LogicRNCID=501',
'Label=CR1142_U2, CellID=11422, LogicRNCID=501',
'Label=CR1895_U6, CellID=63146, LogicRNCID=501',
'Label=CR1895_U4, CellID=63144, LogicRNCID=501',
'Label=CR1895_U3, CellID=63143, LogicRNCID=501',
'Label=CR1895_U1, CellID=63141, LogicRNCID=501',
'Label=CR1146_U6, CellID=11466, LogicRNCID=501',
'Label=CR1146_U4, CellID=11464, LogicRNCID=501',
'Label=CR1146_U2, CellID=11462, LogicRNCID=501',
'Label=CR1146_U1, CellID=11461, LogicRNCID=501',
'Label=CR0325_U5, CellID=3255, LogicRNCID=501',
'Label=CR0325_U4, CellID=3254, LogicRNCID=501',
'Label=CR0325_U2, CellID=3252, LogicRNCID=501',
'Label=CR0827_U6, CellID=64436, LogicRNCID=501',
'Label=CR0827_U5, CellID=64435, LogicRNCID=501',
'Label=CR0827_U3, CellID=64433, LogicRNCID=501',
'Label=CR0827_U2, CellID=64432, LogicRNCID=501',
'Label=CR1031_U6, CellID=10316, LogicRNCID=501',
'Label=CR1031_U4, CellID=10314, LogicRNCID=501',
'Label=CR1031_U3, CellID=10313, LogicRNCID=501',
'Label=CR1031_U1, CellID=10311, LogicRNCID=501',
'Label=CR0489_U6, CellID=4896, LogicRNCID=501',
'Label=CR0489_U4, CellID=4894, LogicRNCID=501',
'Label=CR0489_U2, CellID=4892, LogicRNCID=501',
'Label=CR0489_U1, CellID=4891, LogicRNCID=501',
'Label=CR1533_U4, CellID=15334, LogicRNCID=501',
'Label=CR1533_U2, CellID=15332, LogicRNCID=501',
'Label=KR1471_U6, CellID=14716, LogicRNCID=501',
'Label=KR1471_U4, CellID=14714, LogicRNCID=501',
'Label=KR1471_U3, CellID=14713, LogicRNCID=501',
'Label=KR1471_U1, CellID=14711, LogicRNCID=501',
'Label=CR1474_U6, CellID=14746, LogicRNCID=501',
'Label=CR1474_U4, CellID=14744, LogicRNCID=501',
'Label=CR1474_U2, CellID=14742, LogicRNCID=501',
'Label=CR1474_U1, CellID=14741, LogicRNCID=501',
'Label=KR1504_U5, CellID=15045, LogicRNCID=501',
'Label=KR1504_U4, CellID=15044, LogicRNCID=501',
'Label=KR1504_U2, CellID=15042, LogicRNCID=501',
'Label=CR1079_U6, CellID=10796, LogicRNCID=501',
'Label=CR1079_U5, CellID=10795, LogicRNCID=501',
'Label=CR1079_U3, CellID=10793, LogicRNCID=501',
'Label=CR1079_U2, CellID=10792, LogicRNCID=501',
'Label=KR1865_U6, CellID=18656, LogicRNCID=501',
'Label=KR1865_U4, CellID=18654, LogicRNCID=501',
'Label=KR1865_U3, CellID=18653, LogicRNCID=501',
'Label=KR1865_U1, CellID=18651, LogicRNCID=501',
'Label=KR1873_U6, CellID=18736, LogicRNCID=501',
'Label=KR1873_U4, CellID=18734, LogicRNCID=501',
'Label=KR1873_U2, CellID=18732, LogicRNCID=501',
'Label=KR1873_U1, CellID=18731, LogicRNCID=501',
'Label=CR1082_U5, CellID=10825, LogicRNCID=501',
'Label=CR1082_U4, CellID=10824, LogicRNCID=501',
'Label=CR1082_U2, CellID=10822, LogicRNCID=501',
'Label=CR1080_U6, CellID=10806, LogicRNCID=501',
'Label=CR1080_U5, CellID=10805, LogicRNCID=501',
'Label=CR1080_U3, CellID=10803, LogicRNCID=501',
'Label=CR1080_U2, CellID=10802, LogicRNCID=501',
'Label=CR0675_U6, CellID=6756, LogicRNCID=501',
'Label=CR0675_U4, CellID=6754, LogicRNCID=501',
'Label=CR0675_U3, CellID=6753, LogicRNCID=501',
'Label=CR0675_U1, CellID=6751, LogicRNCID=501',
'Label=CR1081_U6, CellID=10816, LogicRNCID=501',
'Label=CR1081_U4, CellID=10814, LogicRNCID=501',
'Label=CR1081_U2, CellID=10812, LogicRNCID=501',
'Label=CR1081_U1, CellID=10811, LogicRNCID=501',
'Label=CR0502_U5, CellID=5025, LogicRNCID=501',
'Label=CR0502_U4, CellID=5024, LogicRNCID=501',
'Label=CR0502_U2, CellID=5022, LogicRNCID=501',
'Label=CR0955_U6, CellID=65096, LogicRNCID=501',
'Label=CR0955_U5, CellID=65095, LogicRNCID=501',
'Label=CR0955_U3, CellID=65093, LogicRNCID=501',
'Label=CR0955_U2, CellID=65092, LogicRNCID=501',
'Label=KR1872_U6, CellID=18726, LogicRNCID=501',
'Label=KR1872_U4, CellID=18724, LogicRNCID=501',
'Label=KR1872_U3, CellID=18723, LogicRNCID=501',
'Label=KR1872_U1, CellID=18721, LogicRNCID=501',
'Label=CR0718_U2, CellID=7182, LogicRNCID=501',
'Label=KR0488_U4, CellID=4884, LogicRNCID=501',
'Label=KR0488_U2, CellID=4882, LogicRNCID=501',
'Label=KR0488_U1, CellID=4881, LogicRNCID=501',
'Label=CR0718_U5, CellID=7185, LogicRNCID=501',
'Label=CR0718_U4, CellID=7184, LogicRNCID=501',
'Label=CR0650_U6, CellID=6506, LogicRNCID=501',
'Label=CR0650_U4, CellID=6504, LogicRNCID=501',
'Label=CR0650_U3, CellID=6503, LogicRNCID=501',
'Label=CR0650_U1, CellID=6501, LogicRNCID=501',
'Label=CR1479_U6, CellID=14796, LogicRNCID=501',
'Label=CR1479_U4, CellID=14794, LogicRNCID=501',
'Label=CR1479_U2, CellID=14792, LogicRNCID=501',
'Label=CR1479_U1, CellID=14791, LogicRNCID=501',
'Label=CR1360_U5, CellID=13605, LogicRNCID=501',
'Label=CR1360_U4, CellID=13604, LogicRNCID=501',
'Label=CR1360_U2, CellID=13602, LogicRNCID=501',
'Label=CR1360_U1, CellID=13601, LogicRNCID=501',
'Label=CR1360_U3, CellID=13603, LogicRNCID=501',
'Label=CR1360_U6, CellID=13606, LogicRNCID=501',
'Label=CR1479_U3, CellID=14793, LogicRNCID=501',
'Label=CR1479_U5, CellID=14795, LogicRNCID=501',
'Label=CR0650_U2, CellID=6502, LogicRNCID=501',
'Label=CR0650_U5, CellID=6505, LogicRNCID=501',
'Label=CR0718_U3, CellID=7183, LogicRNCID=501',
'Label=CR0718_U6, CellID=7186, LogicRNCID=501',
'Label=KR0488_U3, CellID=4883, LogicRNCID=501',
'Label=CR0718_U1, CellID=7181, LogicRNCID=501',
'Label=KR1872_U2, CellID=18722, LogicRNCID=501',
'Label=KR1872_U5, CellID=18725, LogicRNCID=501',
'Label=CR0955_U1, CellID=65091, LogicRNCID=501',
'Label=CR0955_U4, CellID=65094, LogicRNCID=501',
'Label=CR0502_U1, CellID=5021, LogicRNCID=501',
'Label=CR0502_U3, CellID=5023, LogicRNCID=501',
'Label=CR0502_U6, CellID=5026, LogicRNCID=501',
'Label=CR1081_U3, CellID=10813, LogicRNCID=501',
'Label=CR1081_U5, CellID=10815, LogicRNCID=501',
'Label=CR0675_U2, CellID=6752, LogicRNCID=501',
'Label=CR0675_U5, CellID=6755, LogicRNCID=501',
'Label=CR1080_U1, CellID=10801, LogicRNCID=501',
'Label=CR1080_U4, CellID=10804, LogicRNCID=501',
'Label=CR1082_U1, CellID=10821, LogicRNCID=501',
'Label=CR1082_U3, CellID=10823, LogicRNCID=501',
'Label=CR1082_U6, CellID=10826, LogicRNCID=501',
'Label=KR1873_U3, CellID=18733, LogicRNCID=501',
'Label=KR1873_U5, CellID=18735, LogicRNCID=501',
'Label=KR1865_U2, CellID=18652, LogicRNCID=501',
'Label=KR1865_U5, CellID=18655, LogicRNCID=501',
'Label=CR1079_U1, CellID=10791, LogicRNCID=501',
'Label=CR1079_U4, CellID=10794, LogicRNCID=501',
'Label=KR1504_U1, CellID=15041, LogicRNCID=501',
'Label=KR1504_U3, CellID=15043, LogicRNCID=501',
'Label=KR1504_U6, CellID=15046, LogicRNCID=501',
'Label=CR1474_U3, CellID=14743, LogicRNCID=501',
'Label=CR1474_U5, CellID=14745, LogicRNCID=501',
'Label=KR1471_U2, CellID=14712, LogicRNCID=501',
'Label=KR1471_U5, CellID=14715, LogicRNCID=501',
'Label=CR1533_U1, CellID=15331, LogicRNCID=501',
'Label=CR1533_U5, CellID=15335, LogicRNCID=501',
'Label=CR0489_U3, CellID=4893, LogicRNCID=501',
'Label=CR0489_U5, CellID=4895, LogicRNCID=501',
'Label=CR1031_U2, CellID=10312, LogicRNCID=501',
'Label=CR1031_U5, CellID=10315, LogicRNCID=501',
'Label=CR0827_U1, CellID=64431, LogicRNCID=501',
'Label=CR0827_U4, CellID=64434, LogicRNCID=501',
'Label=CR0325_U1, CellID=3251, LogicRNCID=501',
'Label=CR0325_U3, CellID=3253, LogicRNCID=501',
'Label=CR0325_U6, CellID=3256, LogicRNCID=501',
'Label=CR1146_U3, CellID=11463, LogicRNCID=501',
'Label=CR1146_U5, CellID=11465, LogicRNCID=501',
'Label=CR1895_U2, CellID=63142, LogicRNCID=501',
'Label=CR1895_U5, CellID=63145, LogicRNCID=501',
'Label=CR1142_U1, CellID=11421, LogicRNCID=501',
'Label=CR1142_U4, CellID=11424, LogicRNCID=501',
'Label=CR0414_U1, CellID=4141, LogicRNCID=501',
'Label=CR0414_U3, CellID=4143, LogicRNCID=501',
'Label=CR0414_U6, CellID=4146, LogicRNCID=501',
'Label=CR1160_U3, CellID=11603, LogicRNCID=501',
'Label=CR1160_U5, CellID=11605, LogicRNCID=501',
'Label=CR0184_U2, CellID=1842, LogicRNCID=501',
'Label=CR0184_U5, CellID=1845, LogicRNCID=501',
'Label=CR1169_U1, CellID=11691, LogicRNCID=501',
'Label=CR1169_U4, CellID=11694, LogicRNCID=501',
'Label=CR0725_U1, CellID=7251, LogicRNCID=501',
'Label=CR0725_U3, CellID=7253, LogicRNCID=501',
'Label=CR0725_U6, CellID=7256, LogicRNCID=501',
'Label=CR1894_U3, CellID=63093, LogicRNCID=501',
'Label=CR1894_U5, CellID=63095, LogicRNCID=501',
'Label=KR0778_U2, CellID=7782, LogicRNCID=501',
'Label=KR0778_U5, CellID=7785, LogicRNCID=501',
'Label=CR0507_U1, CellID=5071, LogicRNCID=501',
'Label=CR0507_U4, CellID=5074, LogicRNCID=501',
'Label=CR1051_U1, CellID=10511, LogicRNCID=501',
'Label=CR1051_U3, CellID=10513, LogicRNCID=501',
'Label=CR1051_U6, CellID=10516, LogicRNCID=501',
'Label=KR1869_U3, CellID=18693, LogicRNCID=501',
'Label=KR1869_U5, CellID=18695, LogicRNCID=501',
'Label=CR1077_U2, CellID=10772, LogicRNCID=501',
'Label=CR1077_U5, CellID=10775, LogicRNCID=501',
'Label=KR1870_U1, CellID=18701, LogicRNCID=501',
'Label=KR1870_U4, CellID=18704, LogicRNCID=501',
'Label=KR0553_U1, CellID=5531, LogicRNCID=501',
'Label=KR0553_U3, CellID=5533, LogicRNCID=501',
'Label=KR0553_U6, CellID=5536, LogicRNCID=501',
'Label=CR1756_U3, CellID=17563, LogicRNCID=501',
'Label=CR1756_U5, CellID=17565, LogicRNCID=501',
'Label=CR0486_U2, CellID=4862, LogicRNCID=501',
'Label=CR0486_U5, CellID=4865, LogicRNCID=501',
'Label=KR0633_U1, CellID=6331, LogicRNCID=501',
'Label=KR0633_U4, CellID=6334, LogicRNCID=501',
'Label=KR0631_U1, CellID=6311, LogicRNCID=501',
'Label=KR0631_U3, CellID=6313, LogicRNCID=501',
'Label=KR0631_U6, CellID=6316, LogicRNCID=501',
'Label=KR0632_U3, CellID=6323, LogicRNCID=501',
'Label=KR0632_U5, CellID=6325, LogicRNCID=501',
'Label=CR1107_U2, CellID=11072, LogicRNCID=501',
'Label=CR0484_U1, CellID=4841, LogicRNCID=501',
'Label=CR0484_U3, CellID=4843, LogicRNCID=501',
'Label=CR0484_U6, CellID=4846, LogicRNCID=501',
'Label=CR0454_U3, CellID=4543, LogicRNCID=501',
'Label=CR0454_U5, CellID=4545, LogicRNCID=501',
'Label=CR1883_U2, CellID=65462, LogicRNCID=501',
'Label=CR1883_U5, CellID=65465, LogicRNCID=501',
'Label=CR1128_U1, CellID=11281, LogicRNCID=501',
'Label=CR1128_U4, CellID=11284, LogicRNCID=501',
'Label=CR1209_U1, CellID=12091, LogicRNCID=501',
'Label=CR1209_U3, CellID=12093, LogicRNCID=501',
'Label=CR1209_U6, CellID=12096, LogicRNCID=501',
'Label=CR0485_U3, CellID=4853, LogicRNCID=501',
'Label=CR0485_U5, CellID=4855, LogicRNCID=501',
'Label=CR1885_U2, CellID=18852, LogicRNCID=501',
'Label=KR1876_U6, CellID=18766, LogicRNCID=501',
'Label=KR1876_U4, CellID=18764, LogicRNCID=501',
'Label=KR1876_U2, CellID=18762, LogicRNCID=501',
'Label=KR1876_U1, CellID=18761, LogicRNCID=501',
'Label=KR1876_U3, CellID=18763, LogicRNCID=501',
'Label=KR1876_U5, CellID=18765, LogicRNCID=501',
'Label=CR1879_U6, CellID=18796, LogicRNCID=501',
'Label=CR1879_U5, CellID=18795, LogicRNCID=501',
'Label=CR1879_U4, CellID=18794, LogicRNCID=501',
'Label=CR1879_U3, CellID=18793, LogicRNCID=501',
'Label=CR1879_U2, CellID=18792, LogicRNCID=501',
'Label=CR1879_U1, CellID=18791, LogicRNCID=501',
'Label=CR1529_U6, CellID=64156, LogicRNCID=501',
'Label=CR1529_U5, CellID=64155, LogicRNCID=501',
'Label=CR1529_U4, CellID=64154, LogicRNCID=501',
'Label=CR1529_U3, CellID=64153, LogicRNCID=501',
'Label=CR1529_U2, CellID=64152, LogicRNCID=501',
'Label=CR1529_U1, CellID=64151, LogicRNCID=501',
'Label=CR1882_U5, CellID=60275, LogicRNCID=501',
'Label=CR1882_U4, CellID=60274, LogicRNCID=501',
'Label=CR1882_U2, CellID=60272, LogicRNCID=501',
'Label=CR1882_U1, CellID=60271, LogicRNCID=501',
'Label=CR1882_U3, CellID=60273, LogicRNCID=501',
'Label=CR1882_U6, CellID=60276, LogicRNCID=501',
] # рехоуминг
cluster_CR_N = [
'CR0101',
'CR0192',
'CR0484',
'CR0502',
'CR0640',
'CR0831',
'CR0947',
'CR1077',
'CR1078',
'CR1100',
'CR1103',
'CR1104',
'CR1105',
'CR1107',
'CR1376',
'CR1382',
'CR1882',
'CR1885',
'KR0553',
'KR1471',
'KR1852',
'KR1854',
'KR1855',
'KR1858',
'KR1861',
'KR1870',
'KR1876',
] # NodeB СR KR

cluster_Kahovka = [
'Label=UH1882_U96, CellID=50033, LogicRNCID=501',
'Label=UH1882_U95, CellID=50032, LogicRNCID=501',
'Label=UH1882_U5, CellID=18825, LogicRNCID=501',
'Label=UH1882_U4, CellID=18824, LogicRNCID=501',
'Label=UH1882_U1, CellID=18821, LogicRNCID=501',
'Label=UH1882_U2, CellID=18822, LogicRNCID=501',
'Label=UH0995_U4, CellID=9954, LogicRNCID=501',
'Label=UH0995_U5, CellID=9955, LogicRNCID=501',
'Label=UH3918_U6, CellID=39186, LogicRNCID=501',
'Label=UH3918_U3, CellID=39183, LogicRNCID=501',
'Label=UH3918_U1, CellID=39181, LogicRNCID=501',
'Label=UH0995_U3, CellID=9953, LogicRNCID=501',
'Label=UH0995_U1, CellID=9951, LogicRNCID=501',
'Label=UH0988_U1, CellID=9881, LogicRNCID=501',
'Label=UH0988_U4, CellID=9884, LogicRNCID=501',
'Label=UH0988_U3, CellID=9883, LogicRNCID=501',
'Label=UH0988_U2, CellID=9882, LogicRNCID=501',
'Label=UH0820_U95, CellID=48205, LogicRNCID=501',
'Label=UH0820_U94, CellID=48204, LogicRNCID=501',
'Label=UH0820_U96, CellID=48206, LogicRNCID=501',
'Label=UH0988_U5, CellID=9885, LogicRNCID=501',
'Label=UH0988_U6, CellID=9886, LogicRNCID=501',
'Label=UH0995_U2, CellID=9952, LogicRNCID=501',
'Label=UH3918_U4, CellID=39184, LogicRNCID=501',
'Label=UH0995_U6, CellID=9956, LogicRNCID=501',
'Label=UH1882_U3, CellID=18823, LogicRNCID=501',
'Label=UH1882_U6, CellID=18826, LogicRNCID=501',
'Label=UH1882_U94, CellID=50031, LogicRNCID=501',
'Label=UH0820_U5, CellID=8205, LogicRNCID=501',
'Label=UH0820_U1, CellID=8201, LogicRNCID=501',
'Label=UH0820_U4, CellID=8204, LogicRNCID=501',
'Label=UH0820_U2, CellID=8202, LogicRNCID=501',
'Label=UH0820_U98, CellID=48208, LogicRNCID=501',
'Label=UH0820_U99, CellID=48209, LogicRNCID=501',
'Label=UH0820_U97, CellID=48207, LogicRNCID=501',
]  # новый кластер для отчёта
cluster_N_Kahovka = [
'Label=UH0881_U94, CellID=48814, LogicRNCID=501',
'Label=UH0881_U95, CellID=48815, LogicRNCID=501',
'Label=UH0821_U95, CellID=48215, LogicRNCID=501',
'Label=UH0821_U94, CellID=48214, LogicRNCID=501',
'Label=UH2981_U4, CellID=29814, LogicRNCID=501',
'Label=UH2981_U3, CellID=29813, LogicRNCID=501',
'Label=UH2981_U6, CellID=29816, LogicRNCID=501',
'Label=UH2981_U5, CellID=29815, LogicRNCID=501',
'Label=UH1947_U3, CellID=19473, LogicRNCID=501',
'Label=UH2981_U1, CellID=29811, LogicRNCID=501',
'Label=UH1947_U6, CellID=19476, LogicRNCID=501',
'Label=UH1947_U5, CellID=19475, LogicRNCID=501',
'Label=UH1947_U4, CellID=19474, LogicRNCID=501',
'Label=UH1947_U1, CellID=19471, LogicRNCID=501',
'Label=UH1947_U2, CellID=19472, LogicRNCID=501',
'Label=UH2981_U2, CellID=29812, LogicRNCID=501',
'Label=UH0821_U96, CellID=48216, LogicRNCID=501',
'Label=UH0881_U96, CellID=48816, LogicRNCID=501',
'Label=UH3925_U6, CellID=39256, LogicRNCID=501',
'Label=UH3925_U4, CellID=39254, LogicRNCID=501',
'Label=UH3925_U3, CellID=39253, LogicRNCID=501',
'Label=UH0970_U3, CellID=9703, LogicRNCID=501',
'Label=UH0970_U6, CellID=9706, LogicRNCID=501',
'Label=UH0970_U4, CellID=9704, LogicRNCID=501',
'Label=UH0881_U3, CellID=8813, LogicRNCID=501',
'Label=UH0970_U1, CellID=9701, LogicRNCID=501',
'Label=UH0881_U6, CellID=8816, LogicRNCID=501',
'Label=UH0821_U4, CellID=8214, LogicRNCID=501',
'Label=UH0821_U2, CellID=8212, LogicRNCID=501',
'Label=UH0881_U4, CellID=8814, LogicRNCID=501',
'Label=UH0881_U2, CellID=8812, LogicRNCID=501',
'Label=UH0881_U1, CellID=8811, LogicRNCID=501',
'Label=UH0821_U6, CellID=8216, LogicRNCID=501',
'Label=UH0821_U5, CellID=8215, LogicRNCID=501',
'Label=UH0821_U3, CellID=8213, LogicRNCID=501',
'Label=UH0821_U1, CellID=8211, LogicRNCID=501',
'Label=UH0881_U5, CellID=8815, LogicRNCID=501',
'Label=UH0970_U2, CellID=9702, LogicRNCID=501',
'Label=UH0970_U5, CellID=9705, LogicRNCID=501',
'Label=UH3925_U2, CellID=39252, LogicRNCID=501',
'Label=UH3925_U5, CellID=39255, LogicRNCID=501',
'Label=UH3925_U1, CellID=39251, LogicRNCID=501',
'Label=UH0881_U97, CellID=48817, LogicRNCID=501',
'Label=UH0821_U98, CellID=48218, LogicRNCID=501',
'Label=UH0821_U99, CellID=48219, LogicRNCID=501',
'Label=UH0821_U97, CellID=48217, LogicRNCID=501',
]  # новый кластер для отчёта
cluster_Genichesk = [
'Label=UH0602_U94, CellID=46024, LogicRNCID=501',
'Label=UH0602_U96, CellID=46026, LogicRNCID=501',
'Label=UH0602_U95, CellID=46025, LogicRNCID=501',
'Label=UH4600_U6, CellID=46006, LogicRNCID=501',
'Label=UH4600_U5, CellID=46005, LogicRNCID=501',
'Label=UH4600_U1, CellID=46001, LogicRNCID=501',
'Label=UH4600_U2, CellID=46002, LogicRNCID=501',
'Label=UH0966_U6, CellID=9666, LogicRNCID=501',
'Label=UH0966_U3, CellID=9663, LogicRNCID=501',
'Label=UH0966_U2, CellID=9662, LogicRNCID=501',
'Label=UH0966_U1, CellID=9661, LogicRNCID=501',
'Label=UH0952_U4, CellID=9524, LogicRNCID=501',
'Label=UH0952_U3, CellID=9523, LogicRNCID=501',
'Label=UH0952_U1, CellID=9521, LogicRNCID=501',
'Label=UH0966_U4, CellID=9664, LogicRNCID=501',
'Label=UH0952_U6, CellID=9526, LogicRNCID=501',
'Label=UH0952_U2, CellID=9522, LogicRNCID=501',
'Label=UH0952_U5, CellID=9525, LogicRNCID=501',
'Label=UH0966_U5, CellID=9665, LogicRNCID=501',
'Label=UH4600_U4, CellID=46004, LogicRNCID=501',
'Label=UH4600_U3, CellID=46003, LogicRNCID=501',
'Label=UH0602_U3, CellID=6023, LogicRNCID=501',
'Label=UH0602_U6, CellID=6026, LogicRNCID=501',
'Label=UH0602_U5, CellID=6025, LogicRNCID=501',
'Label=UH0602_U1, CellID=6021, LogicRNCID=501',
'Label=UH0602_U4, CellID=6024, LogicRNCID=501',
'Label=UH0602_U2, CellID=6022, LogicRNCID=501',
]  # новый кластер для отчёта
cluster_Zhel_Port = [
'Label=UH1907_U22, CellID=65493, LogicRNCID=501',
'Label=UH1907_U11, CellID=65492, LogicRNCID=501',
'Label=UH1907_U55, CellID=65496, LogicRNCID=501',
'Label=UH1907_U96, CellID=50039, LogicRNCID=501',
'Label=UH1907_U95, CellID=50038, LogicRNCID=501',
'Label=UH1907_U44, CellID=65495, LogicRNCID=501',
'Label=UH0832_U95, CellID=48325, LogicRNCID=501',
'Label=UH0832_U94, CellID=48324, LogicRNCID=501',
'Label=UH0960_U95, CellID=49605, LogicRNCID=501',
'Label=UH1704_U5, CellID=17045, LogicRNCID=501',
'Label=UH1704_U4, CellID=17044, LogicRNCID=501',
'Label=UH1704_U2, CellID=17042, LogicRNCID=501',
'Label=UH2932_U3, CellID=29323, LogicRNCID=501',
'Label=UH2932_U6, CellID=29326, LogicRNCID=501',
'Label=UH2932_U1, CellID=29321, LogicRNCID=501',
'Label=UH2932_U4, CellID=29324, LogicRNCID=501',
'Label=UH0832_U6, CellID=8326, LogicRNCID=501',
'Label=UH0832_U5, CellID=8325, LogicRNCID=501',
'Label=UH0832_U3, CellID=8323, LogicRNCID=501',
'Label=UH0832_U2, CellID=8322, LogicRNCID=501',
'Label=UH0832_U4, CellID=8324, LogicRNCID=501',
'Label=UH0832_U1, CellID=8321, LogicRNCID=501',
'Label=UH2932_U5, CellID=29325, LogicRNCID=501',
'Label=UH2932_U2, CellID=29322, LogicRNCID=501',
'Label=UH1704_U3, CellID=17043, LogicRNCID=501',
'Label=UH1704_U6, CellID=17046, LogicRNCID=501',
'Label=UH1704_U1, CellID=17041, LogicRNCID=501',
'Label=UH0960_U94, CellID=49604, LogicRNCID=501',
'Label=UH0960_U96, CellID=49606, LogicRNCID=501',
'Label=UH0832_U96, CellID=48326, LogicRNCID=501',
'Label=UH1907_U33, CellID=65494, LogicRNCID=501',
'Label=UH1907_U94, CellID=50037, LogicRNCID=501',
'Label=UH1907_U66, CellID=65497, LogicRNCID=501',
'Label=UH1960_U2, CellID=19602, LogicRNCID=501',
'Label=UH1905_U5, CellID=19055, LogicRNCID=501',
'Label=UH1905_U4, CellID=19054, LogicRNCID=501',
'Label=UH1960_U3, CellID=19603, LogicRNCID=501',
'Label=UH1960_U1, CellID=19601, LogicRNCID=501',
'Label=UH1905_U2, CellID=19052, LogicRNCID=501',
'Label=UH1960_U5, CellID=19605, LogicRNCID=501',
'Label=UH0960_U6, CellID=9606, LogicRNCID=501',
'Label=UH0960_U4, CellID=9604, LogicRNCID=501',
'Label=UH0960_U3, CellID=9603, LogicRNCID=501',
'Label=UH0960_U1, CellID=9601, LogicRNCID=501',
'Label=UH1703_U1, CellID=17031, LogicRNCID=501',
'Label=UH1907_U1, CellID=19071, LogicRNCID=501',
'Label=UH1958_U6, CellID=19586, LogicRNCID=501',
'Label=UH1907_U6, CellID=19076, LogicRNCID=501',
'Label=UH1907_U5, CellID=19075, LogicRNCID=501',
'Label=UH1907_U4, CellID=19074, LogicRNCID=501',
'Label=UH1907_U2, CellID=19072, LogicRNCID=501',
'Label=UH1958_U5, CellID=19585, LogicRNCID=501',
'Label=UH1703_U6, CellID=17036, LogicRNCID=501',
'Label=UH1958_U2, CellID=19582, LogicRNCID=501',
'Label=UH1703_U4, CellID=17034, LogicRNCID=501',
'Label=UH1703_U5, CellID=17035, LogicRNCID=501',
'Label=UH1958_U3, CellID=19583, LogicRNCID=501',
'Label=UH1703_U3, CellID=17033, LogicRNCID=501',
'Label=UH1703_U2, CellID=17032, LogicRNCID=501',
'Label=UH1907_U3, CellID=19073, LogicRNCID=501',
'Label=UH0960_U2, CellID=9602, LogicRNCID=501',
'Label=UH0960_U5, CellID=9605, LogicRNCID=501',
'Label=UH1960_U6, CellID=19606, LogicRNCID=501',
'Label=UH1905_U1, CellID=19051, LogicRNCID=501',
'Label=UH1960_U4, CellID=19604, LogicRNCID=501',
]  # новый кластер для отчёта
cluster_Lazurnoye = [
'Label=UH1964_U5, CellID=19645, LogicRNCID=501',
'Label=UH1964_U2, CellID=19642, LogicRNCID=501',
'Label=UH1964_U1, CellID=19641, LogicRNCID=501',
'Label=UH1964_U94, CellID=55387, LogicRNCID=501',
'Label=UH1964_U95, CellID=55388, LogicRNCID=501',
'Label=UH0963_U96, CellID=49636, LogicRNCID=501',
'Label=UH0831_U96, CellID=48316, LogicRNCID=501',
'Label=UH0831_U95, CellID=48315, LogicRNCID=501',
'Label=UH0831_U94, CellID=48314, LogicRNCID=501',
'Label=UH0963_U95, CellID=49635, LogicRNCID=501',
'Label=UH0963_U94, CellID=49634, LogicRNCID=501',
'Label=UH1964_U96, CellID=55389, LogicRNCID=501',
'Label=UH1964_U4, CellID=19644, LogicRNCID=501',
'Label=UH0963_U3, CellID=9633, LogicRNCID=501',
'Label=UH0963_U6, CellID=9636, LogicRNCID=501',
'Label=UH0963_U4, CellID=9634, LogicRNCID=501',
'Label=UH0963_U1, CellID=9631, LogicRNCID=501',
'Label=UH0963_U2, CellID=9632, LogicRNCID=501',
'Label=UH1705_U1, CellID=17051, LogicRNCID=501',
'Label=UH1705_U6, CellID=17056, LogicRNCID=501',
'Label=UH1705_U5, CellID=17055, LogicRNCID=501',
'Label=UH1705_U4, CellID=17054, LogicRNCID=501',
'Label=UH1705_U3, CellID=17053, LogicRNCID=501',
'Label=UH1707_U1, CellID=17071, LogicRNCID=501',
'Label=UH1892_U3, CellID=18923, LogicRNCID=501',
'Label=UH1707_U6, CellID=17076, LogicRNCID=501',
'Label=UH1707_U5, CellID=17075, LogicRNCID=501',
'Label=UH1707_U2, CellID=17072, LogicRNCID=501',
'Label=UH0831_U2, CellID=8312, LogicRNCID=501',
'Label=UH0831_U1, CellID=8311, LogicRNCID=501',
'Label=UH0831_U6, CellID=8316, LogicRNCID=501',
'Label=UH0831_U4, CellID=8314, LogicRNCID=501',
'Label=UH0831_U3, CellID=8313, LogicRNCID=501',
'Label=UH3989_U5, CellID=39895, LogicRNCID=501',
'Label=UH3989_U4, CellID=39894, LogicRNCID=501',
'Label=UH3989_U2, CellID=39892, LogicRNCID=501',
'Label=UH3989_U3, CellID=39893, LogicRNCID=501',
'Label=UH3989_U6, CellID=39896, LogicRNCID=501',
'Label=UH3989_U1, CellID=39891, LogicRNCID=501',
'Label=UH0831_U5, CellID=8315, LogicRNCID=501',
'Label=UH1707_U4, CellID=17074, LogicRNCID=501',
'Label=UH1892_U5, CellID=18925, LogicRNCID=501',
'Label=UH1892_U2, CellID=18922, LogicRNCID=501',
'Label=UH1892_U6, CellID=18926, LogicRNCID=501',
'Label=UH1707_U3, CellID=17073, LogicRNCID=501',
'Label=UH1705_U2, CellID=17052, LogicRNCID=501',
'Label=UH0963_U5, CellID=9635, LogicRNCID=501',
]  # новый кластер для отчёта
cluster_Skadovsk = [
'Label=UH0830_U95, CellID=48305, LogicRNCID=501',
'Label=UH0830_U94, CellID=48304, LogicRNCID=501',
'Label=UH0962_U96, CellID=49626, LogicRNCID=501',
'Label=UH0962_U94, CellID=49624, LogicRNCID=501',
'Label=UH0830_U6, CellID=8306, LogicRNCID=501',
'Label=UH0830_U5, CellID=8305, LogicRNCID=501',
'Label=UH0830_U3, CellID=8303, LogicRNCID=501',
'Label=UH0830_U2, CellID=8302, LogicRNCID=501',
'Label=UH0830_U1, CellID=8301, LogicRNCID=501',
'Label=UH0830_U4, CellID=8304, LogicRNCID=501',
'Label=UH0962_U95, CellID=49625, LogicRNCID=501',
'Label=UH0830_U96, CellID=48306, LogicRNCID=501',
'Label=UH3922_U5, CellID=39225, LogicRNCID=501',
'Label=UH3922_U4, CellID=39224, LogicRNCID=501',
'Label=UH3922_U2, CellID=39222, LogicRNCID=501',
'Label=UH3922_U1, CellID=39221, LogicRNCID=501',
'Label=UH3990_U4, CellID=39904, LogicRNCID=501',
'Label=UH3990_U2, CellID=39902, LogicRNCID=501',
'Label=UH3990_U6, CellID=39906, LogicRNCID=501',
'Label=UH2938_U6, CellID=29386, LogicRNCID=501',
'Label=UH2938_U5, CellID=29385, LogicRNCID=501',
'Label=UH2938_U3, CellID=29383, LogicRNCID=501',
'Label=UH2938_U2, CellID=29382, LogicRNCID=501',
'Label=UH0961_U4, CellID=9614, LogicRNCID=501',
'Label=UH0961_U3, CellID=9613, LogicRNCID=501',
'Label=UH0961_U1, CellID=9611, LogicRNCID=501',
'Label=UH0961_U6, CellID=9616, LogicRNCID=501',
'Label=UH0962_U3, CellID=9623, LogicRNCID=501',
'Label=UH0962_U5, CellID=9625, LogicRNCID=501',
'Label=UH0962_U2, CellID=9622, LogicRNCID=501',
'Label=UH0962_U1, CellID=9621, LogicRNCID=501',
'Label=UH0962_U4, CellID=9624, LogicRNCID=501',
'Label=UH0962_U6, CellID=9626, LogicRNCID=501',
'Label=UH0961_U2, CellID=9612, LogicRNCID=501',
'Label=UH0961_U5, CellID=9615, LogicRNCID=501',
'Label=UH2938_U1, CellID=29381, LogicRNCID=501',
'Label=UH3990_U1, CellID=39901, LogicRNCID=501',
'Label=UH2938_U4, CellID=29384, LogicRNCID=501',
'Label=UH3990_U5, CellID=39905, LogicRNCID=501',
'Label=UH3990_U3, CellID=39903, LogicRNCID=501',
]  # новый кластер для отчёта
cluster_VsyoOstalnoye = [
'Label=UH2709_U96, CellID=50067, LogicRNCID=501',
'Label=UH2709_U95, CellID=50066, LogicRNCID=501',
'Label=UH2709_U94, CellID=50065, LogicRNCID=501',
'Label=UH2709_U5, CellID=27095, LogicRNCID=501',
'Label=UH2709_U4, CellID=27094, LogicRNCID=501',
'Label=UH2709_U2, CellID=27092, LogicRNCID=501',
'Label=UH2709_U1, CellID=27091, LogicRNCID=501',
'Label=UH1804_U95, CellID=50023, LogicRNCID=501',
'Label=UH1804_U94, CellID=50022, LogicRNCID=501',
'Label=UH0761_U94, CellID=47614, LogicRNCID=501',
'Label=UH0746_U96, CellID=47466, LogicRNCID=501',
'Label=UH1602_U94, CellID=51007, LogicRNCID=501',
'Label=UH0869_U95, CellID=48695, LogicRNCID=501',
'Label=UH0869_U94, CellID=48694, LogicRNCID=501',
'Label=UH1936_U95, CellID=59365, LogicRNCID=501',
'Label=UH0980_U6, CellID=9806, LogicRNCID=501',
'Label=UH0980_U5, CellID=9805, LogicRNCID=501',
'Label=UH0980_U3, CellID=9803, LogicRNCID=501',
'Label=UH0980_U1, CellID=9801, LogicRNCID=501',
'Label=UH1902_U96, CellID=50036, LogicRNCID=501',
'Label=UH1902_U94, CellID=50034, LogicRNCID=501',
'Label=UH0980_U2, CellID=9802, LogicRNCID=501',
'Label=UH0974_U3, CellID=9743, LogicRNCID=501',
'Label=UH0974_U1, CellID=9741, LogicRNCID=501',
'Label=UH0974_U6, CellID=9746, LogicRNCID=501',
'Label=UH0974_U4, CellID=9744, LogicRNCID=501',
'Label=UH0854_U4, CellID=8544, LogicRNCID=501',
'Label=UH0854_U3, CellID=8543, LogicRNCID=501',
'Label=UH0854_U6, CellID=8546, LogicRNCID=501',
'Label=UH1604_U96, CellID=50015, LogicRNCID=501',
'Label=UH1604_U95, CellID=50014, LogicRNCID=501',
'Label=UH0854_U1, CellID=8541, LogicRNCID=501',
'Label=UH0717_U96, CellID=47176, LogicRNCID=501',
'Label=UH0717_U94, CellID=47174, LogicRNCID=501',
'Label=UH1603_U96, CellID=50012, LogicRNCID=501',
'Label=UH1603_U95, CellID=50011, LogicRNCID=501',
'Label=UH1801_U96, CellID=58016, LogicRNCID=501',
'Label=UH1801_U94, CellID=58014, LogicRNCID=501',
'Label=UH3928_U5, CellID=39285, LogicRNCID=501',
'Label=UH3928_U4, CellID=39284, LogicRNCID=501',
'Label=UH3928_U2, CellID=39282, LogicRNCID=501',
'Label=UH3928_U1, CellID=39281, LogicRNCID=501',
'Label=UH1984_U4, CellID=19844, LogicRNCID=501',
'Label=UH1984_U2, CellID=19842, LogicRNCID=501',
'Label=UH1918_U6, CellID=19186, LogicRNCID=501',
'Label=UH1918_U5, CellID=19185, LogicRNCID=501',
'Label=UH1918_U4, CellID=19184, LogicRNCID=501',
'Label=UH1918_U2, CellID=19182, LogicRNCID=501',
'Label=UH1918_U1, CellID=19181, LogicRNCID=501',
'Label=UH0955_U95, CellID=49555, LogicRNCID=501',
'Label=UH0955_U94, CellID=49554, LogicRNCID=501',
'Label=UH0955_U3, CellID=9553, LogicRNCID=501',
'Label=UH0955_U1, CellID=9551, LogicRNCID=501',
'Label=UH0955_U6, CellID=9556, LogicRNCID=501',
'Label=UH0955_U4, CellID=9554, LogicRNCID=501',
'Label=UH0723_U95, CellID=47235, LogicRNCID=501',
'Label=UH0723_U94, CellID=47234, LogicRNCID=501',
'Label=UH3995_U2, CellID=39952, LogicRNCID=501',
'Label=UH3996_U5, CellID=39965, LogicRNCID=501',
'Label=UH3996_U4, CellID=39964, LogicRNCID=501',
'Label=UH3996_U2, CellID=39962, LogicRNCID=501',
'Label=UH1928_U95, CellID=50049, LogicRNCID=501',
'Label=UH1928_U5, CellID=19285, LogicRNCID=501',
'Label=UH1928_U4, CellID=19284, LogicRNCID=501',
'Label=UH0636_U96, CellID=46366, LogicRNCID=501',
'Label=UH1928_U3, CellID=19283, LogicRNCID=501',
'Label=UH1928_U2, CellID=19282, LogicRNCID=501',
'Label=UH3995_U1, CellID=39951, LogicRNCID=501',
'Label=UH0636_U94, CellID=46364, LogicRNCID=501',
'Label=UH0636_U95, CellID=46365, LogicRNCID=501',
'Label=UH0636_U6, CellID=6366, LogicRNCID=501',
'Label=UH0636_U3, CellID=6363, LogicRNCID=501',
'Label=UH3995_U5, CellID=39955, LogicRNCID=501',
'Label=UH1928_U96, CellID=50050, LogicRNCID=501',
'Label=UH3995_U3, CellID=39953, LogicRNCID=501',
'Label=UH1805_U96, CellID=50027, LogicRNCID=501',
'Label=UH1805_U95, CellID=50026, LogicRNCID=501',
'Label=UH1805_U94, CellID=50025, LogicRNCID=501',
'Label=UH0849_U95, CellID=48495, LogicRNCID=501',
'Label=UH0849_U94, CellID=48494, LogicRNCID=501',
'Label=UH0849_U96, CellID=48496, LogicRNCID=501',
'Label=UH0611_U96, CellID=46116, LogicRNCID=501',
'Label=UH0611_U95, CellID=46115, LogicRNCID=501',
'Label=UH0611_U94, CellID=46114, LogicRNCID=501',
'Label=UH0747_U94, CellID=47474, LogicRNCID=501',
'Label=UH0760_U97, CellID=50074, LogicRNCID=501',
'Label=UH0750_U96, CellID=47506, LogicRNCID=501',
'Label=UH0750_U95, CellID=47505, LogicRNCID=501',
'Label=UH0760_U95, CellID=50072, LogicRNCID=501',
'Label=UH0760_U94, CellID=50071, LogicRNCID=501',
'Label=UH0736_U94, CellID=47364, LogicRNCID=501',
'Label=UH0739_U96, CellID=47396, LogicRNCID=501',
'Label=UH0732_U96, CellID=47326, LogicRNCID=501',
'Label=UH0745_U94, CellID=47454, LogicRNCID=501',
'Label=UH0740_U96, CellID=47406, LogicRNCID=501',
'Label=UH0739_U95, CellID=47395, LogicRNCID=501',
'Label=UH0736_U96, CellID=47366, LogicRNCID=501',
'Label=UH0704_U96, CellID=47046, LogicRNCID=501',
'Label=UH0736_U95, CellID=47365, LogicRNCID=501',
'Label=UH0745_U96, CellID=47456, LogicRNCID=501',
'Label=UH0704_U95, CellID=47045, LogicRNCID=501',
'Label=UH1803_U96, CellID=50021, LogicRNCID=501',
'Label=UH1803_U94, CellID=50019, LogicRNCID=501',
'Label=UH1803_U95, CellID=50020, LogicRNCID=501',
'Label=UH0965_U95, CellID=49655, LogicRNCID=501',
'Label=UH0965_U94, CellID=49654, LogicRNCID=501',
'Label=UH0844_U96, CellID=48446, LogicRNCID=501',
'Label=UH0845_U96, CellID=48456, LogicRNCID=501',
'Label=UH0845_U95, CellID=48455, LogicRNCID=501',
'Label=UH0770_U94, CellID=47704, LogicRNCID=501',
'Label=UH0744_U95, CellID=47445, LogicRNCID=501',
'Label=UH0738_U96, CellID=47386, LogicRNCID=501',
'Label=UH0738_U95, CellID=47385, LogicRNCID=501',
'Label=UH0734_U95, CellID=47345, LogicRNCID=501',
'Label=UH0770_U96, CellID=47706, LogicRNCID=501',
'Label=UH0823_U95, CellID=48235, LogicRNCID=501',
'Label=UH0845_U94, CellID=48454, LogicRNCID=501',
'Label=UH1951_U94, CellID=50062, LogicRNCID=501',
'Label=UH0734_U94, CellID=47344, LogicRNCID=501',
'Label=UH0767_U96, CellID=47676, LogicRNCID=501',
'Label=UH0844_U94, CellID=48444, LogicRNCID=501',
'Label=UH0738_U94, CellID=47384, LogicRNCID=501',
'Label=UH0767_U94, CellID=47674, LogicRNCID=501',
'Label=UH1951_U96, CellID=50064, LogicRNCID=501',
'Label=UH0840_U94, CellID=48404, LogicRNCID=501',
'Label=UH0843_U95, CellID=48435, LogicRNCID=501',
'Label=UH1602_U95, CellID=51008, LogicRNCID=501',
'Label=UH0973_U95, CellID=49735, LogicRNCID=501',
'Label=UH0737_U95, CellID=47375, LogicRNCID=501',
'Label=UH0737_U94, CellID=47374, LogicRNCID=501',
'Label=UH0973_U96, CellID=49736, LogicRNCID=501',
'Label=UH0752_U96, CellID=47526, LogicRNCID=501',
'Label=UH0752_U95, CellID=47525, LogicRNCID=501',
'Label=UH0843_U94, CellID=48434, LogicRNCID=501',
'Label=UH0702_U96, CellID=47026, LogicRNCID=501',
'Label=UH0973_U94, CellID=49734, LogicRNCID=501',
'Label=UH0951_U3, CellID=9513, LogicRNCID=501',
'Label=UH0715_U94, CellID=47154, LogicRNCID=501',
'Label=UH0703_U96, CellID=47036, LogicRNCID=501',
'Label=UH0703_U95, CellID=47035, LogicRNCID=501',
'Label=UH0841_U96, CellID=48416, LogicRNCID=501',
'Label=UH0841_U95, CellID=48415, LogicRNCID=501',
'Label=UH0752_U94, CellID=47524, LogicRNCID=501',
'Label=UH0715_U95, CellID=47155, LogicRNCID=501',
'Label=UH0951_U6, CellID=9516, LogicRNCID=501',
'Label=UH0841_U94, CellID=48414, LogicRNCID=501',
'Label=UH0840_U95, CellID=48405, LogicRNCID=501',
'Label=UH0839_U96, CellID=48396, LogicRNCID=501',
'Label=UH0839_U95, CellID=48395, LogicRNCID=501',
'Label=UH0838_U96, CellID=48386, LogicRNCID=501',
'Label=UH0838_U95, CellID=48385, LogicRNCID=501',
'Label=UH0765_U94, CellID=47654, LogicRNCID=501',
'Label=UH0758_U94, CellID=47584, LogicRNCID=501',
'Label=UH0835_U96, CellID=48356, LogicRNCID=501',
'Label=UH0835_U95, CellID=48355, LogicRNCID=501',
'Label=UH0833_U96, CellID=48336, LogicRNCID=501',
'Label=UH0825_U96, CellID=48256, LogicRNCID=501',
'Label=UH0825_U94, CellID=48254, LogicRNCID=501',
'Label=UH0766_U94, CellID=47666, LogicRNCID=501',
'Label=UH0825_U95, CellID=48255, LogicRNCID=501',
'Label=UH0833_U95, CellID=48335, LogicRNCID=501',
'Label=UH0833_U94, CellID=48334, LogicRNCID=501',
'Label=UH0766_U96, CellID=47664, LogicRNCID=501',
'Label=UH0807_U94, CellID=48074, LogicRNCID=501',
'Label=UH0600_U1, CellID=6001, LogicRNCID=501',
'Label=UH0758_U95, CellID=47585, LogicRNCID=501',
'Label=UH0754_U95, CellID=47545, LogicRNCID=501',
'Label=UH0714_U95, CellID=47145, LogicRNCID=501',
'Label=UH0714_U94, CellID=47144, LogicRNCID=501',
'Label=UH0614_U95, CellID=46145, LogicRNCID=501',
'Label=UH0614_U94, CellID=46144, LogicRNCID=501',
'Label=UH0713_U95, CellID=47135, LogicRNCID=501',
'Label=UH0600_U5, CellID=6005, LogicRNCID=501',
'Label=UH0756_U94, CellID=47564, LogicRNCID=501',
'Label=UH0600_U2, CellID=6002, LogicRNCID=501',
'Label=UH0708_U95, CellID=47085, LogicRNCID=501',
'Label=UH0708_U94, CellID=47084, LogicRNCID=501',
'Label=UH0706_U96, CellID=47066, LogicRNCID=501',
'Label=UH0614_U96, CellID=46146, LogicRNCID=501',
'Label=UH0705_U94, CellID=47054, LogicRNCID=501',
'Label=UH0731_U95, CellID=47315, LogicRNCID=501',
'Label=UH0756_U96, CellID=47566, LogicRNCID=501',
'Label=UH0706_U94, CellID=47064, LogicRNCID=501',
'Label=UH0705_U96, CellID=47056, LogicRNCID=501',
'Label=UH0751_U96, CellID=47516, LogicRNCID=501',
'Label=UH0751_U94, CellID=47514, LogicRNCID=501',
'Label=UH0742_U96, CellID=47426, LogicRNCID=501',
'Label=UH0742_U95, CellID=47425, LogicRNCID=501',
'Label=UH0742_U94, CellID=47424, LogicRNCID=501',
'Label=UH0735_U95, CellID=47355, LogicRNCID=501',
'Label=UH0743_U96, CellID=47436, LogicRNCID=501',
'Label=UH0623_U96, CellID=46236, LogicRNCID=501',
'Label=UH0623_U95, CellID=46235, LogicRNCID=501',
'Label=UH0622_U96, CellID=46226, LogicRNCID=501',
'Label=UH0623_U94, CellID=46234, LogicRNCID=501',
'Label=UH0623_U1, CellID=6231, LogicRNCID=501',
'Label=UH0618_U96, CellID=46186, LogicRNCID=501',
'Label=UH0623_U6, CellID=6236, LogicRNCID=501',
'Label=UH0623_U5, CellID=6235, LogicRNCID=501',
'Label=UH0601_U95, CellID=46015, LogicRNCID=501',
'Label=UH0623_U4, CellID=6234, LogicRNCID=501',
'Label=UH0623_U2, CellID=6232, LogicRNCID=501',
'Label=UH0601_U94, CellID=46014, LogicRNCID=501',
'Label=UH0622_U95, CellID=46225, LogicRNCID=501',
'Label=UH0508_U6, CellID=5086, LogicRNCID=501',
'Label=UH0508_U3, CellID=5083, LogicRNCID=501',
'Label=UH0508_U2, CellID=5082, LogicRNCID=501',
'Label=UH0742_U2, CellID=7422, LogicRNCID=501',
'Label=UH0842_U3, CellID=8423, LogicRNCID=501',
'Label=UH0842_U2, CellID=8422, LogicRNCID=501',
'Label=UH0842_U1, CellID=8421, LogicRNCID=501',
'Label=UH0822_U6, CellID=8226, LogicRNCID=501',
'Label=UH0822_U4, CellID=8224, LogicRNCID=501',
'Label=UH0841_U6, CellID=8416, LogicRNCID=501',
'Label=UH0841_U4, CellID=8414, LogicRNCID=501',
'Label=UH0841_U3, CellID=8413, LogicRNCID=501',
'Label=UH0841_U2, CellID=8412, LogicRNCID=501',
'Label=UH0842_U6, CellID=8426, LogicRNCID=501',
'Label=UH0842_U5, CellID=8425, LogicRNCID=501',
'Label=UH0822_U3, CellID=8223, LogicRNCID=501',
'Label=UH0822_U2, CellID=8222, LogicRNCID=501',
'Label=UH0841_U1, CellID=8411, LogicRNCID=501',
'Label=UH0742_U5, CellID=7425, LogicRNCID=501',
'Label=UH0841_U5, CellID=8415, LogicRNCID=501',
'Label=UH0822_U5, CellID=8225, LogicRNCID=501',
'Label=UH0742_U1, CellID=7421, LogicRNCID=501',
'Label=UH0842_U4, CellID=8424, LogicRNCID=501',
'Label=UH0822_U1, CellID=8221, LogicRNCID=501',
'Label=UH0742_U4, CellID=7424, LogicRNCID=501',
'Label=UH0508_U5, CellID=5085, LogicRNCID=501',
'Label=UH0623_U3, CellID=6233, LogicRNCID=501',
'Label=UH0601_U96, CellID=46016, LogicRNCID=501',
'Label=UH0618_U95, CellID=46185, LogicRNCID=501',
'Label=UH0618_U94, CellID=46184, LogicRNCID=501',
'Label=UH0735_U94, CellID=47354, LogicRNCID=501',
'Label=UH0622_U94, CellID=46224, LogicRNCID=501',
'Label=UH0735_U96, CellID=47356, LogicRNCID=501',
'Label=UH0743_U95, CellID=47435, LogicRNCID=501',
'Label=UH0705_U95, CellID=47055, LogicRNCID=501',
'Label=UH0731_U94, CellID=47314, LogicRNCID=501',
'Label=UH0731_U96, CellID=47316, LogicRNCID=501',
'Label=UH0706_U95, CellID=47065, LogicRNCID=501',
'Label=UH0708_U96, CellID=47086, LogicRNCID=501',
'Label=UH0600_U4, CellID=6004, LogicRNCID=501',
'Label=UH0713_U96, CellID=47136, LogicRNCID=501',
'Label=UH0713_U94, CellID=47134, LogicRNCID=501',
'Label=UH0754_U94, CellID=47544, LogicRNCID=501',
'Label=UH0766_U95, CellID=47665, LogicRNCID=501',
'Label=UH0765_U96, CellID=47656, LogicRNCID=501',
'Label=UH0807_U96, CellID=48076, LogicRNCID=501',
'Label=UH0807_U95, CellID=48075, LogicRNCID=501',
'Label=UH0838_U94, CellID=48384, LogicRNCID=501',
'Label=UH0758_U96, CellID=47586, LogicRNCID=501',
'Label=UH0835_U94, CellID=48354, LogicRNCID=501',
'Label=UH0839_U94, CellID=48394, LogicRNCID=501',
'Label=UH0840_U96, CellID=48406, LogicRNCID=501',
'Label=UH0703_U94, CellID=47034, LogicRNCID=501',
'Label=UH0715_U96, CellID=47156, LogicRNCID=501',
'Label=UH0951_U94, CellID=49514, LogicRNCID=501',
'Label=UH0951_U96, CellID=49516, LogicRNCID=501',
'Label=UH0702_U95, CellID=47025, LogicRNCID=501',
'Label=UH0702_U94, CellID=47024, LogicRNCID=501',
'Label=UH0843_U96, CellID=48436, LogicRNCID=501',
'Label=UH0737_U96, CellID=47376, LogicRNCID=501',
'Label=UH1602_U96, CellID=51009, LogicRNCID=501',
'Label=UH1951_U95, CellID=50063, LogicRNCID=501',
'Label=UH0744_U96, CellID=47446, LogicRNCID=501',
'Label=UH0767_U95, CellID=47675, LogicRNCID=501',
'Label=UH0744_U94, CellID=47444, LogicRNCID=501',
'Label=UH0770_U95, CellID=47705, LogicRNCID=501',
'Label=UH0734_U96, CellID=47346, LogicRNCID=501',
'Label=UH0823_U94, CellID=48234, LogicRNCID=501',
'Label=UH0823_U96, CellID=48236, LogicRNCID=501',
'Label=UH0844_U95, CellID=48445, LogicRNCID=501',
'Label=UH0965_U96, CellID=49656, LogicRNCID=501',
'Label=UH0759_U95, CellID=47595, LogicRNCID=501',
'Label=UH0750_U94, CellID=47504, LogicRNCID=501',
'Label=UH0732_U94, CellID=47324, LogicRNCID=501',
'Label=UH0740_U95, CellID=47405, LogicRNCID=501',
'Label=UH0732_U95, CellID=47325, LogicRNCID=501',
'Label=UH0740_U94, CellID=47404, LogicRNCID=501',
'Label=UH0760_U96, CellID=50073, LogicRNCID=501',
'Label=UH0759_U96, CellID=47596, LogicRNCID=501',
'Label=UH0745_U95, CellID=47455, LogicRNCID=501',
'Label=UH0747_U95, CellID=47475, LogicRNCID=501',
'Label=UH0636_U2, CellID=6362, LogicRNCID=501',
'Label=UH3995_U4, CellID=39954, LogicRNCID=501',
'Label=UH0636_U5, CellID=6365, LogicRNCID=501',
'Label=UH3995_U6, CellID=39956, LogicRNCID=501',
'Label=UH3996_U1, CellID=39961, LogicRNCID=501',
'Label=UH1928_U1, CellID=19281, LogicRNCID=501',
'Label=UH1928_U6, CellID=19286, LogicRNCID=501',
'Label=UH3996_U3, CellID=39963, LogicRNCID=501',
'Label=UH3996_U6, CellID=39966, LogicRNCID=501',
'Label=UH0723_U96, CellID=47236, LogicRNCID=501',
'Label=UH0955_U5, CellID=9555, LogicRNCID=501',
'Label=UH0955_U2, CellID=9552, LogicRNCID=501',
'Label=UH0955_U96, CellID=49556, LogicRNCID=501',
'Label=UH1918_U3, CellID=19183, LogicRNCID=501',
'Label=UH1984_U1, CellID=19841, LogicRNCID=501',
'Label=UH1984_U5, CellID=19845, LogicRNCID=501',
'Label=UH3928_U3, CellID=39283, LogicRNCID=501',
'Label=UH3928_U6, CellID=39286, LogicRNCID=501',
'Label=UH1603_U94, CellID=50010, LogicRNCID=501',
'Label=UH0717_U95, CellID=47175, LogicRNCID=501',
'Label=UH1604_U94, CellID=50013, LogicRNCID=501',
'Label=UH0854_U5, CellID=8545, LogicRNCID=501',
'Label=UH0854_U2, CellID=8542, LogicRNCID=501',
'Label=UH0974_U5, CellID=9745, LogicRNCID=501',
'Label=UH0974_U2, CellID=9742, LogicRNCID=501',
'Label=UH1902_U95, CellID=50035, LogicRNCID=501',
'Label=UH0980_U4, CellID=9804, LogicRNCID=501',
'Label=UH1936_U96, CellID=59366, LogicRNCID=501',
'Label=UH0869_U96, CellID=48696, LogicRNCID=501',
'Label=UH0746_U95, CellID=47465, LogicRNCID=501',
'Label=UH0761_U96, CellID=47616, LogicRNCID=501',
'Label=UH1804_U96, CellID=50024, LogicRNCID=501',
'Label=UH2709_U3, CellID=27093, LogicRNCID=501',
'Label=UH2709_U6, CellID=27096, LogicRNCID=501',
'Label=UH0849_U5, CellID=8495, LogicRNCID=501',
'Label=UH0849_U3, CellID=8493, LogicRNCID=501',
'Label=UH0849_U2, CellID=8492, LogicRNCID=501',
'Label=UH0849_U1, CellID=8491, LogicRNCID=501',
'Label=UH3920_U4, CellID=39204, LogicRNCID=501',
'Label=UH3920_U2, CellID=39202, LogicRNCID=501',
'Label=UH2984_U6, CellID=29846, LogicRNCID=501',
'Label=UH2984_U5, CellID=29845, LogicRNCID=501',
'Label=UH2984_U2, CellID=29842, LogicRNCID=501',
'Label=UH2984_U1, CellID=29841, LogicRNCID=501',
'Label=UH0839_U6, CellID=8396, LogicRNCID=501',
'Label=UH0839_U5, CellID=8395, LogicRNCID=501',
'Label=UH0965_U2, CellID=9652, LogicRNCID=501',
'Label=UH0965_U1, CellID=9651, LogicRNCID=501',
'Label=UH0839_U2, CellID=8392, LogicRNCID=501',
'Label=UH1906_U2, CellID=19062, LogicRNCID=501',
'Label=UH0965_U6, CellID=9656, LogicRNCID=501',
'Label=UH0965_U3, CellID=9653, LogicRNCID=501',
'Label=UH1805_U4, CellID=18054, LogicRNCID=501',
'Label=UH1805_U1, CellID=18051, LogicRNCID=501',
'Label=UH0839_U3, CellID=8393, LogicRNCID=501',
'Label=UH1906_U5, CellID=19065, LogicRNCID=501',
'Label=UH1906_U4, CellID=19064, LogicRNCID=501',
'Label=UH1805_U5, CellID=18055, LogicRNCID=501',
'Label=UH1805_U3, CellID=18053, LogicRNCID=501',
'Label=UH1906_U3, CellID=19063, LogicRNCID=501',
'Label=UH1916_U6, CellID=19166, LogicRNCID=501',
'Label=UH1916_U5, CellID=19165, LogicRNCID=501',
'Label=UH1916_U4, CellID=19164, LogicRNCID=501',
'Label=UH1945_U3, CellID=19453, LogicRNCID=501',
'Label=UH1916_U1, CellID=19161, LogicRNCID=501',
'Label=UH1945_U6, CellID=19456, LogicRNCID=501',
'Label=UH0823_U5, CellID=8235, LogicRNCID=501',
'Label=UH0823_U3, CellID=8233, LogicRNCID=501',
'Label=UH0823_U2, CellID=8232, LogicRNCID=501',
'Label=UH1945_U1, CellID=19451, LogicRNCID=501',
'Label=UH0823_U1, CellID=8231, LogicRNCID=501',
'Label=UH0738_U5, CellID=7385, LogicRNCID=501',
'Label=UH0738_U4, CellID=7384, LogicRNCID=501',
'Label=UH0738_U3, CellID=7383, LogicRNCID=501',
'Label=UH0743_U6, CellID=7436, LogicRNCID=501',
'Label=UH0743_U5, CellID=7435, LogicRNCID=501',
'Label=UH0743_U3, CellID=7433, LogicRNCID=501',
'Label=UH0743_U2, CellID=7432, LogicRNCID=501',
'Label=UH1916_U3, CellID=19163, LogicRNCID=501',
'Label=UH1945_U4, CellID=19454, LogicRNCID=501',
'Label=UH0743_U1, CellID=7431, LogicRNCID=501',
'Label=UH2763_U2, CellID=27632, LogicRNCID=501',
'Label=UH3921_U2, CellID=39212, LogicRNCID=501',
'Label=UH2763_U1, CellID=27631, LogicRNCID=501',
'Label=UH3921_U6, CellID=39216, LogicRNCID=501',
'Label=UH2763_U6, CellID=27636, LogicRNCID=501',
'Label=UH2763_U4, CellID=27634, LogicRNCID=501',
'Label=UH2763_U3, CellID=27633, LogicRNCID=501',
'Label=UH3921_U1, CellID=39211, LogicRNCID=501',
'Label=UH1951_U1, CellID=19511, LogicRNCID=501',
'Label=UH3921_U5, CellID=39215, LogicRNCID=501',
'Label=UH3921_U4, CellID=39214, LogicRNCID=501',
'Label=UH1702_U1, CellID=17021, LogicRNCID=501',
'Label=UH1951_U5, CellID=19515, LogicRNCID=501',
'Label=UH1951_U4, CellID=19514, LogicRNCID=501',
'Label=UH1702_U5, CellID=17025, LogicRNCID=501',
'Label=UH1702_U4, CellID=17024, LogicRNCID=501',
'Label=UH3921_U3, CellID=39213, LogicRNCID=501',
'Label=UH0717_U4, CellID=7174, LogicRNCID=501',
'Label=UH0717_U3, CellID=7173, LogicRNCID=501',
'Label=UH1951_U3, CellID=19513, LogicRNCID=501',
'Label=UH0717_U6, CellID=7176, LogicRNCID=501',
'Label=UH0717_U5, CellID=7175, LogicRNCID=501',
'Label=UH0747_U5, CellID=7475, LogicRNCID=501',
'Label=UH0622_U1, CellID=6221, LogicRNCID=501',
'Label=UH0622_U6, CellID=6226, LogicRNCID=501',
'Label=UH0622_U4, CellID=6224, LogicRNCID=501',
'Label=UH0622_U3, CellID=6223, LogicRNCID=501',
'Label=UH0711_U4, CellID=7114, LogicRNCID=501',
'Label=UH0711_U3, CellID=7113, LogicRNCID=501',
'Label=UH0711_U6, CellID=7116, LogicRNCID=501',
'Label=UH0734_U2, CellID=7342, LogicRNCID=501',
'Label=UH0770_U3, CellID=7703, LogicRNCID=501',
'Label=UH0770_U2, CellID=7702, LogicRNCID=501',
'Label=UH0763_U4, CellID=7634, LogicRNCID=501',
'Label=UH0763_U2, CellID=7632, LogicRNCID=501',
'Label=UH0763_U1, CellID=7631, LogicRNCID=501',
'Label=UH0763_U6, CellID=7636, LogicRNCID=501',
'Label=UH0770_U5, CellID=7705, LogicRNCID=501',
'Label=UH0734_U5, CellID=7345, LogicRNCID=501',
'Label=UH0763_U5, CellID=7635, LogicRNCID=501',
'Label=UH0734_U1, CellID=7341, LogicRNCID=501',
'Label=UH0763_U3, CellID=7633, LogicRNCID=501',
'Label=UH0611_U4, CellID=6114, LogicRNCID=501',
'Label=UH2983_U3, CellID=29833, LogicRNCID=501',
'Label=UH2983_U2, CellID=29832, LogicRNCID=501',
'Label=UH0994_U6, CellID=9946, LogicRNCID=501',
'Label=UH0994_U5, CellID=9945, LogicRNCID=501',
'Label=UH0994_U4, CellID=9944, LogicRNCID=501',
'Label=UH0840_U1, CellID=8401, LogicRNCID=501',
'Label=UH0807_U2, CellID=8072, LogicRNCID=501',
'Label=UH2983_U5, CellID=29835, LogicRNCID=501',
'Label=UH0840_U4, CellID=8404, LogicRNCID=501',
'Label=UH0994_U3, CellID=9943, LogicRNCID=501',
'Label=UH0994_U1, CellID=9941, LogicRNCID=501',
'Label=UH0840_U6, CellID=8406, LogicRNCID=501',
'Label=UH0807_U1, CellID=8071, LogicRNCID=501',
'Label=UH2983_U1, CellID=29831, LogicRNCID=501',
'Label=UH1939_U4, CellID=19394, LogicRNCID=501',
'Label=UH0807_U3, CellID=8073, LogicRNCID=501',
'Label=UH0611_U1, CellID=6111, LogicRNCID=501',
'Label=UH1939_U6, CellID=19396, LogicRNCID=501',
'Label=UH1939_U1, CellID=19391, LogicRNCID=501',
'Label=UH0611_U3, CellID=6113, LogicRNCID=501',
'Label=UH0807_U6, CellID=8076, LogicRNCID=501',
'Label=UH0807_U4, CellID=8074, LogicRNCID=501',
'Label=UH1939_U2, CellID=19392, LogicRNCID=501',
'Label=UH1976_U4, CellID=19764, LogicRNCID=501',
'Label=UH1976_U3, CellID=19763, LogicRNCID=501',
'Label=UH1908_U3, CellID=19083, LogicRNCID=501',
'Label=UH1908_U2, CellID=19082, LogicRNCID=501',
'Label=UH1976_U1, CellID=19761, LogicRNCID=501',
'Label=UH1908_U6, CellID=19086, LogicRNCID=501',
'Label=UH0706_U1, CellID=7061, LogicRNCID=501',
'Label=UH1908_U4, CellID=19084, LogicRNCID=501',
'Label=UH1976_U6, CellID=19766, LogicRNCID=501',
'Label=UH0714_U3, CellID=7143, LogicRNCID=501',
'Label=UH0714_U5, CellID=7145, LogicRNCID=501',
'Label=UH0838_U3, CellID=8383, LogicRNCID=501',
'Label=UH0838_U2, CellID=8382, LogicRNCID=501',
'Label=UH0972_U2, CellID=9722, LogicRNCID=501',
'Label=UH0714_U1, CellID=7141, LogicRNCID=501',
'Label=UH0838_U5, CellID=8385, LogicRNCID=501',
'Label=UH0714_U2, CellID=7142, LogicRNCID=501',
'Label=UH0838_U1, CellID=8381, LogicRNCID=501',
'Label=UH0981_U1, CellID=9811, LogicRNCID=501',
'Label=UH1935_U5, CellID=19355, LogicRNCID=501',
'Label=UH0972_U3, CellID=9723, LogicRNCID=501',
'Label=UH0981_U4, CellID=9814, LogicRNCID=501',
'Label=UH1935_U3, CellID=19353, LogicRNCID=501',
'Label=UH1935_U2, CellID=19352, LogicRNCID=501',
'Label=UH0981_U6, CellID=9816, LogicRNCID=501',
'Label=UH0981_U5, CellID=9815, LogicRNCID=501',
'Label=UH0972_U4, CellID=9724, LogicRNCID=501',
'Label=UH0973_U3, CellID=9733, LogicRNCID=501',
'Label=UH0973_U2, CellID=9732, LogicRNCID=501',
'Label=UH0972_U6, CellID=9726, LogicRNCID=501',
'Label=UH0972_U5, CellID=9725, LogicRNCID=501',
'Label=UH0973_U4, CellID=9734, LogicRNCID=501',
'Label=UH0981_U3, CellID=9813, LogicRNCID=501',
'Label=UH0973_U6, CellID=9736, LogicRNCID=501',
'Label=UH0618_U3, CellID=6183, LogicRNCID=501',
'Label=UH0618_U1, CellID=6181, LogicRNCID=501',
'Label=UH0825_U6, CellID=8256, LogicRNCID=501',
'Label=UH0825_U3, CellID=8253, LogicRNCID=501',
'Label=UH0825_U2, CellID=8252, LogicRNCID=501',
'Label=UH0601_U6, CellID=6016, LogicRNCID=501',
'Label=UH0618_U6, CellID=6186, LogicRNCID=501',
'Label=UH0601_U4, CellID=6014, LogicRNCID=501',
'Label=UH0601_U3, CellID=6013, LogicRNCID=501',
'Label=UH0601_U1, CellID=6011, LogicRNCID=501',
'Label=UH1803_U6, CellID=18036, LogicRNCID=501',
'Label=UH0825_U4, CellID=8254, LogicRNCID=501',
'Label=UH1803_U1, CellID=18031, LogicRNCID=501',
'Label=UH1803_U4, CellID=18034, LogicRNCID=501',
'Label=UH0618_U5, CellID=6185, LogicRNCID=501',
'Label=UH1833_U1, CellID=18331, LogicRNCID=501',
'Label=UH0845_U6, CellID=8456, LogicRNCID=501',
'Label=UH0845_U5, CellID=8455, LogicRNCID=501',
'Label=UH0845_U3, CellID=8453, LogicRNCID=501',
'Label=UH1803_U3, CellID=18033, LogicRNCID=501',
'Label=UH0844_U5, CellID=8445, LogicRNCID=501',
'Label=UH0845_U2, CellID=8452, LogicRNCID=501',
'Label=UH0835_U6, CellID=8356, LogicRNCID=501',
'Label=UH0835_U5, CellID=8355, LogicRNCID=501',
'Label=UH0844_U4, CellID=8444, LogicRNCID=501',
'Label=UH0844_U3, CellID=8443, LogicRNCID=501',
'Label=UH1833_U6, CellID=18336, LogicRNCID=501',
'Label=UH0845_U1, CellID=8451, LogicRNCID=501',
'Label=UH0844_U6, CellID=8446, LogicRNCID=501',
'Label=UH0835_U3, CellID=8353, LogicRNCID=501',
'Label=UH0835_U2, CellID=8352, LogicRNCID=501',
'Label=UH1833_U3, CellID=18333, LogicRNCID=501',
'Label=UH1833_U2, CellID=18332, LogicRNCID=501',
'Label=UH0844_U1, CellID=8441, LogicRNCID=501',
'Label=UH2950_U1, CellID=29501, LogicRNCID=501',
'Label=UH0758_U3, CellID=7583, LogicRNCID=501',
'Label=UH2950_U6, CellID=29506, LogicRNCID=501',
'Label=UH2950_U4, CellID=29504, LogicRNCID=501',
'Label=UH2950_U3, CellID=29503, LogicRNCID=501',
'Label=UH0758_U2, CellID=7582, LogicRNCID=501',
'Label=UH0758_U1, CellID=7581, LogicRNCID=501',
'Label=UH0758_U6, CellID=7586, LogicRNCID=501',
'Label=UH0758_U4, CellID=7584, LogicRNCID=501',
'Label=UH0708_U2, CellID=7082, LogicRNCID=501',
'Label=UH2913_U5, CellID=29135, LogicRNCID=501',
'Label=UH2913_U4, CellID=29134, LogicRNCID=501',
'Label=UH2913_U3, CellID=29133, LogicRNCID=501',
'Label=UH0950_U3, CellID=9503, LogicRNCID=501',
'Label=UH2913_U1, CellID=29131, LogicRNCID=501',
'Label=UH2913_U2, CellID=29132, LogicRNCID=501',
'Label=UH0950_U4, CellID=9504, LogicRNCID=501',
'Label=UH2913_U6, CellID=29136, LogicRNCID=501',
'Label=UH0758_U5, CellID=7585, LogicRNCID=501',
'Label=UH2950_U2, CellID=29502, LogicRNCID=501',
'Label=UH2950_U5, CellID=29505, LogicRNCID=501',
'Label=UH0835_U1, CellID=8351, LogicRNCID=501',
'Label=UH1833_U4, CellID=18334, LogicRNCID=501',
'Label=UH1833_U5, CellID=18335, LogicRNCID=501',
'Label=UH0835_U4, CellID=8354, LogicRNCID=501',
'Label=UH0844_U2, CellID=8442, LogicRNCID=501',
'Label=UH1803_U2, CellID=18032, LogicRNCID=501',
'Label=UH0845_U4, CellID=8454, LogicRNCID=501',
'Label=UH0618_U4, CellID=6184, LogicRNCID=501',
'Label=UH1803_U5, CellID=18035, LogicRNCID=501',
'Label=UH0825_U5, CellID=8255, LogicRNCID=501',
'Label=UH0601_U2, CellID=6012, LogicRNCID=501',
'Label=UH0601_U5, CellID=6015, LogicRNCID=501',
'Label=UH0825_U1, CellID=8251, LogicRNCID=501',
'Label=UH0618_U2, CellID=6182, LogicRNCID=501',
'Label=UH0981_U2, CellID=9812, LogicRNCID=501',
'Label=UH0838_U4, CellID=8384, LogicRNCID=501',
'Label=UH0973_U1, CellID=9731, LogicRNCID=501',
'Label=UH1935_U1, CellID=19351, LogicRNCID=501',
'Label=UH0973_U5, CellID=9735, LogicRNCID=501',
'Label=UH1935_U6, CellID=19356, LogicRNCID=501',
'Label=UH1935_U4, CellID=19354, LogicRNCID=501',
'Label=UH0972_U1, CellID=9721, LogicRNCID=501',
'Label=UH0838_U6, CellID=8386, LogicRNCID=501',
'Label=UH0714_U6, CellID=7146, LogicRNCID=501',
'Label=UH0714_U4, CellID=7144, LogicRNCID=501',
'Label=UH0706_U4, CellID=7064, LogicRNCID=501',
'Label=UH1908_U5, CellID=19085, LogicRNCID=501',
'Label=UH1908_U1, CellID=19081, LogicRNCID=501',
'Label=UH1976_U2, CellID=19762, LogicRNCID=501',
'Label=UH1976_U5, CellID=19765, LogicRNCID=501',
'Label=UH0807_U5, CellID=8075, LogicRNCID=501',
'Label=UH1939_U5, CellID=19395, LogicRNCID=501',
'Label=UH0611_U6, CellID=6116, LogicRNCID=501',
'Label=UH0840_U3, CellID=8403, LogicRNCID=501',
'Label=UH0994_U2, CellID=9942, LogicRNCID=501',
'Label=UH2983_U6, CellID=29836, LogicRNCID=501',
'Label=UH1939_U3, CellID=19393, LogicRNCID=501',
'Label=UH0840_U5, CellID=8405, LogicRNCID=501',
'Label=UH2983_U4, CellID=29834, LogicRNCID=501',
'Label=UH0840_U2, CellID=8402, LogicRNCID=501',
'Label=UH0711_U1, CellID=7111, LogicRNCID=501',
'Label=UH0734_U3, CellID=7343, LogicRNCID=501',
'Label=UH0734_U4, CellID=7344, LogicRNCID=501',
'Label=UH0734_U6, CellID=7346, LogicRNCID=501',
'Label=UH0770_U1, CellID=7701, LogicRNCID=501',
'Label=UH0770_U4, CellID=7704, LogicRNCID=501',
'Label=UH0770_U6, CellID=7706, LogicRNCID=501',
'Label=UH0711_U5, CellID=7115, LogicRNCID=501',
'Label=UH0711_U2, CellID=7112, LogicRNCID=501',
'Label=UH0622_U2, CellID=6222, LogicRNCID=501',
'Label=UH0622_U5, CellID=6225, LogicRNCID=501',
'Label=UH0747_U2, CellID=7472, LogicRNCID=501',
'Label=UH0717_U1, CellID=7171, LogicRNCID=501',
'Label=UH1951_U2, CellID=19512, LogicRNCID=501',
'Label=UH0717_U2, CellID=7172, LogicRNCID=501',
'Label=UH1702_U3, CellID=17023, LogicRNCID=501',
'Label=UH1702_U6, CellID=17026, LogicRNCID=501',
'Label=UH1951_U6, CellID=19516, LogicRNCID=501',
'Label=UH2763_U5, CellID=27635, LogicRNCID=501',
'Label=UH1702_U2, CellID=17022, LogicRNCID=501',
'Label=UH0823_U6, CellID=8236, LogicRNCID=501',
'Label=UH0743_U4, CellID=7434, LogicRNCID=501',
'Label=UH0738_U2, CellID=7382, LogicRNCID=501',
'Label=UH0738_U1, CellID=7381, LogicRNCID=501',
'Label=UH0738_U6, CellID=7386, LogicRNCID=501',
'Label=UH1945_U5, CellID=19455, LogicRNCID=501',
'Label=UH0823_U4, CellID=8234, LogicRNCID=501',
'Label=UH1945_U2, CellID=19452, LogicRNCID=501',
'Label=UH1916_U2, CellID=19162, LogicRNCID=501',
'Label=UH1906_U1, CellID=19061, LogicRNCID=501',
'Label=UH1805_U6, CellID=18056, LogicRNCID=501',
'Label=UH1906_U6, CellID=19066, LogicRNCID=501',
'Label=UH1805_U2, CellID=18052, LogicRNCID=501',
'Label=UH0965_U5, CellID=9655, LogicRNCID=501',
'Label=UH0839_U1, CellID=8391, LogicRNCID=501',
'Label=UH0965_U4, CellID=9654, LogicRNCID=501',
'Label=UH0839_U4, CellID=8394, LogicRNCID=501',
'Label=UH2984_U3, CellID=29843, LogicRNCID=501',
'Label=UH2984_U4, CellID=29844, LogicRNCID=501',
'Label=UH3920_U1, CellID=39201, LogicRNCID=501',
'Label=UH3920_U5, CellID=39205, LogicRNCID=501',
'Label=UH0849_U4, CellID=8494, LogicRNCID=501',
'Label=UH0849_U6, CellID=8496, LogicRNCID=501',
'Label=UH0950_U1, CellID=9501, LogicRNCID=501',
'Label=UH0950_U6, CellID=9506, LogicRNCID=501',
'Label=UH0708_U3, CellID=7083, LogicRNCID=501',
'Label=UH0708_U5, CellID=7085, LogicRNCID=501',
'Label=UH0708_U1, CellID=7081, LogicRNCID=501',
'Label=UH1936_U5, CellID=19365, LogicRNCID=501',
'Label=UH1936_U4, CellID=19364, LogicRNCID=501',
'Label=UH1936_U2, CellID=19362, LogicRNCID=501',
'Label=UH1936_U1, CellID=19361, LogicRNCID=501',
'Label=UH0950_U5, CellID=9505, LogicRNCID=501',
'Label=UH0950_U2, CellID=9502, LogicRNCID=501',
'Label=UH1801_U3, CellID=18013, LogicRNCID=501',
'Label=UH1801_U6, CellID=18016, LogicRNCID=501',
'Label=UH1801_U4, CellID=18014, LogicRNCID=501',
'Label=UH1801_U1, CellID=18011, LogicRNCID=501',
'Label=UH0765_U3, CellID=7653, LogicRNCID=501',
'Label=UH0617_U3, CellID=6173, LogicRNCID=501',
'Label=UH0765_U5, CellID=7655, LogicRNCID=501',
'Label=UH0765_U1, CellID=7651, LogicRNCID=501',
'Label=UH0617_U6, CellID=6176, LogicRNCID=501',
'Label=UH0617_U4, CellID=6174, LogicRNCID=501',
'Label=UH0617_U2, CellID=6172, LogicRNCID=501',
'Label=UH0610_U3, CellID=6103, LogicRNCID=501',
'Label=UH0610_U6, CellID=6106, LogicRNCID=501',
'Label=UH0610_U2, CellID=6102, LogicRNCID=501',
'Label=UH1854_U6, CellID=18546, LogicRNCID=501',
'Label=UH1854_U5, CellID=18545, LogicRNCID=501',
'Label=UH1854_U3, CellID=18543, LogicRNCID=501',
'Label=UH0951_U1, CellID=9511, LogicRNCID=501',
'Label=UH1854_U2, CellID=18542, LogicRNCID=501',
'Label=UH1945_U96, CellID=50058, LogicRNCID=501',
'Label=UH1945_U95, CellID=50057, LogicRNCID=501',
'Label=UH0646_U96, CellID=51012, LogicRNCID=501',
'Label=UH0646_U95, CellID=51011, LogicRNCID=501',
'Label=UH0763_U96, CellID=47636, LogicRNCID=501',
'Label=UH0763_U95, CellID=47635, LogicRNCID=501',
'Label=UH0761_U97, CellID=47617, LogicRNCID=501',
'Label=UH0869_U97, CellID=48697, LogicRNCID=501',
'Label=UH1945_U99, CellID=50061, LogicRNCID=501',
'Label=UH1945_U98, CellID=50060, LogicRNCID=501',
'Label=UH0839_U99, CellID=48399, LogicRNCID=501',
'Label=UH0839_U97, CellID=48397, LogicRNCID=501',
'Label=UH0880_U6, CellID=8806, LogicRNCID=501',
'Label=UH0880_U4, CellID=8804, LogicRNCID=501',
'Label=UH0880_U1, CellID=8801, LogicRNCID=501',
'Label=UH0880_U11, CellID=65520, LogicRNCID=501',
'Label=UH0879_U96, CellID=55060, LogicRNCID=501',
'Label=UH0879_U5, CellID=8795, LogicRNCID=501',
'Label=UH0879_U3, CellID=8793, LogicRNCID=501',
'Label=UH0879_U94, CellID=55062, LogicRNCID=501',
'Label=UH0879_U95, CellID=55061, LogicRNCID=501',
'Label=UH0766_U5, CellID=7665, LogicRNCID=501',
'Label=UH0766_U4, CellID=7664, LogicRNCID=501',
'Label=UH0766_U2, CellID=7662, LogicRNCID=501',
'Label=UH0766_U1, CellID=7661, LogicRNCID=501',
'Label=UH0711_U95, CellID=47115, LogicRNCID=501',
'Label=UH0711_U94, CellID=47114, LogicRNCID=501',
'Label=test_rtk_U2, CellID=11112, LogicRNCID=501',
'Label=UH0748_U94, CellID=47484, LogicRNCID=501',
'Label=UH0748_U96, CellID=47486, LogicRNCID=501',
'Label=UH0748_U1, CellID=7481, LogicRNCID=501',
'Label=UH0748_U6, CellID=7486, LogicRNCID=501',
'Label=UH0748_U4, CellID=7484, LogicRNCID=501',
'Label=UH0748_U3, CellID=7483, LogicRNCID=501',
'Label=UH1925_U94, CellID=50048, LogicRNCID=501',
'Label=UH1925_U96, CellID=50046, LogicRNCID=501',
'Label=UH1913_U96, CellID=50042, LogicRNCID=501',
'Label=UH1913_U95, CellID=50041, LogicRNCID=501',
'Label=UH1913_U94, CellID=50040, LogicRNCID=501',
'Label=UH0822_U94, CellID=48224, LogicRNCID=501',
'Label=UH0822_U96, CellID=48226, LogicRNCID=501',
'Label=UH0846_U4, CellID=8464, LogicRNCID=501',
'Label=UH0846_U3, CellID=8463, LogicRNCID=501',
'Label=UH0846_U96, CellID=48466, LogicRNCID=501',
'Label=UH0846_U95, CellID=48465, LogicRNCID=501',
'Label=UH1952_U6, CellID=19526, LogicRNCID=501',
'Label=UH1952_U3, CellID=19523, LogicRNCID=501',
'Label=UH1952_U2, CellID=19522, LogicRNCID=501',
'Label=UH1952_U5, CellID=19525, LogicRNCID=501',
'Label=UH1952_U1, CellID=19521, LogicRNCID=501',
'Label=CR1852_U5, CellID=60745, LogicRNCID=501',
'Label=CR1852_U4, CellID=60744, LogicRNCID=501',
'Label=CR1852_U2, CellID=60742, LogicRNCID=501',
'Label=CR1852_U1, CellID=60741, LogicRNCID=501',
'Label=CR1858_U5, CellID=60415, LogicRNCID=501',
'Label=CR1858_U4, CellID=60414, LogicRNCID=501',
'Label=CR1858_U2, CellID=60412, LogicRNCID=501',
'Label=CR1858_U1, CellID=60411, LogicRNCID=501',
'Label=CR1858_U3, CellID=60413, LogicRNCID=501',
'Label=CR1858_U6, CellID=60416, LogicRNCID=501',
'Label=CR1852_U3, CellID=60743, LogicRNCID=501',
'Label=CR1852_U6, CellID=60746, LogicRNCID=501',
'Label=UH1952_U4, CellID=19524, LogicRNCID=501',
'Label=UH0846_U94, CellID=48464, LogicRNCID=501',
'Label=UH0846_U1, CellID=8461, LogicRNCID=501',
'Label=UH0846_U6, CellID=8466, LogicRNCID=501',
'Label=UH0822_U95, CellID=48225, LogicRNCID=501',
'Label=UH1925_U95, CellID=50047, LogicRNCID=501',
'Label=UH0748_U2, CellID=7482, LogicRNCID=501',
'Label=UH0748_U5, CellID=7485, LogicRNCID=501',
'Label=UH0748_U95, CellID=47485, LogicRNCID=501',
'Label=test_rtk_U1, CellID=11111, LogicRNCID=501',
'Label=UH0711_U96, CellID=47116, LogicRNCID=501',
'Label=UH0766_U3, CellID=7663, LogicRNCID=501',
'Label=UH0766_U6, CellID=7666, LogicRNCID=501',
'Label=UH0879_U2, CellID=8792, LogicRNCID=501',
'Label=UH0879_U6, CellID=8796, LogicRNCID=501',
'Label=UH0880_U3, CellID=8803, LogicRNCID=501',
'Label=UH0880_U14, CellID=65521, LogicRNCID=501',
'Label=UH1945_U97, CellID=50059, LogicRNCID=501',
'Label=UH0965_U97, CellID=49657, LogicRNCID=501',
'Label=UH0763_U94, CellID=47634, LogicRNCID=501',
'Label=UH0646_U94, CellID=51010, LogicRNCID=501',
'Label=UH1945_U94, CellID=50056, LogicRNCID=501',
'Label=UH0951_U4, CellID=9514, LogicRNCID=501',
'Label=UH1854_U4, CellID=18544, LogicRNCID=501',
'Label=UH1854_U1, CellID=18541, LogicRNCID=501',
'Label=UH0610_U5, CellID=6105, LogicRNCID=501',
'Label=UH0617_U1, CellID=6171, LogicRNCID=501',
'Label=UH0617_U5, CellID=6175, LogicRNCID=501',
'Label=UH0765_U6, CellID=7656, LogicRNCID=501',
'Label=UH0765_U2, CellID=7652, LogicRNCID=501',
'Label=UH0765_U4, CellID=7654, LogicRNCID=501',
'Label=UH0708_U4, CellID=7084, LogicRNCID=501',
'Label=UH1936_U3, CellID=19363, LogicRNCID=501',
'Label=UH1936_U6, CellID=19366, LogicRNCID=501',
'Label=UH0708_U6, CellID=7086, LogicRNCID=501',
]  # новый кластер для отчёта
cell_49626 = ['Label=UH0962_U96, CellID=49626, LogicRNCID=501',]
cell_49605 = ['Label=UH0960_U95, CellID=49605, LogicRNCID=501',]
cell_UH0766 = [
'Label=UH0766_U94, CellID=47666, LogicRNCID=501',
'Label=UH0766_U95, CellID=47665, LogicRNCID=501',
'Label=UH0766_U96, CellID=47664, LogicRNCID=501',
'Label=UH0766_U6, CellID=7666, LogicRNCID=501',
'Label=UH0766_U4, CellID=7664, LogicRNCID=501',
'Label=UH0766_U3, CellID=7663, LogicRNCID=501',
'Label=UH0766_U1, CellID=7661, LogicRNCID=501',
'Label=UH0766_U2, CellID=7662, LogicRNCID=501',
'Label=UH0766_U5, CellID=7665, LogicRNCID=501',
]
site_UH0766 = ['UH0766']
site_KR1858 = [
'Label=KR1858_U6, CellID=18586, LogicRNCID=501',
'Label=KR1858_U4, CellID=18584, LogicRNCID=501',
'Label=KR1858_U3, CellID=18583, LogicRNCID=501',
'Label=KR1858_U1, CellID=18581, LogicRNCID=501',
'Label=KR1858_U2, CellID=18582, LogicRNCID=501',
'Label=KR1858_U5, CellID=18585, LogicRNCID=501',
]
site_KR1858_N = ['KR1858']
cluster_UH_noUH0845 = [
'Label=UH1945_U5, CellID=19455, LogicRNCID=501',
'Label=UH0761_U94, CellID=47614, LogicRNCID=501',
'Label=UH1945_U4, CellID=19454, LogicRNCID=501',
'Label=UH0972_U6, CellID=9726, LogicRNCID=501',
'Label=UH0844_U6, CellID=8446, LogicRNCID=501',
'Label=UH0742_U2, CellID=7422, LogicRNCID=501',
'Label=UH1945_U3, CellID=19453, LogicRNCID=501',
'Label=UH0972_U5, CellID=9725, LogicRNCID=501',
'Label=UH0844_U5, CellID=8445, LogicRNCID=501',
'Label=UH0742_U1, CellID=7421, LogicRNCID=501',
'Label=UH0735_U96, CellID=47356, LogicRNCID=501',
'Label=UH1945_U2, CellID=19452, LogicRNCID=501',
'Label=UH0972_U4, CellID=9724, LogicRNCID=501',
'Label=UH0844_U4, CellID=8444, LogicRNCID=501',
'Label=UH0735_U95, CellID=47355, LogicRNCID=501',
'Label=UH1945_U1, CellID=19451, LogicRNCID=501',
'Label=UH0972_U3, CellID=9723, LogicRNCID=501',
'Label=UH0844_U3, CellID=8443, LogicRNCID=501',
'Label=UH0735_U94, CellID=47354, LogicRNCID=501',
'Label=UH0972_U2, CellID=9722, LogicRNCID=501',
'Label=UH0844_U2, CellID=8442, LogicRNCID=501',
'Label=UH0965_U97, CellID=49657, LogicRNCID=501',
'Label=UH0972_U1, CellID=9721, LogicRNCID=501',
'Label=UH0844_U1, CellID=8441, LogicRNCID=501',
'Label=UH0965_U96, CellID=49656, LogicRNCID=501',
'Label=UH0965_U95, CellID=49655, LogicRNCID=501',
'Label=UH0965_U94, CellID=49654, LogicRNCID=501',
'Label=UH2763_U6, CellID=27636, LogicRNCID=501',
'Label=UH2763_U5, CellID=27635, LogicRNCID=501',
'Label=UH0734_U96, CellID=47346, LogicRNCID=501',
'Label=UH2763_U4, CellID=27634, LogicRNCID=501',
'Label=UH1918_U6, CellID=19186, LogicRNCID=501',
'Label=UH0766_U6, CellID=7666, LogicRNCID=501',
'Label=UH0880_U14, CellID=65521, LogicRNCID=501',
'Label=UH0734_U95, CellID=47345, LogicRNCID=501',
'Label=UH2763_U3, CellID=27633, LogicRNCID=501',
'Label=UH1918_U5, CellID=19185, LogicRNCID=501',
'Label=UH0766_U5, CellID=7665, LogicRNCID=501',
'Label=UH0880_U11, CellID=65520, LogicRNCID=501',
'Label=UH0734_U94, CellID=47344, LogicRNCID=501',
'Label=UH2763_U2, CellID=27632, LogicRNCID=501',
'Label=UH1918_U4, CellID=19184, LogicRNCID=501',
'Label=UH0766_U4, CellID=7664, LogicRNCID=501',
'Label=UH2763_U1, CellID=27631, LogicRNCID=501',
'Label=UH1918_U3, CellID=19183, LogicRNCID=501',
'Label=UH0766_U3, CellID=7663, LogicRNCID=501',
'Label=UH0708_U96, CellID=47086, LogicRNCID=501',
'Label=UH1918_U2, CellID=19182, LogicRNCID=501',
'Label=UH1892_U6, CellID=18926, LogicRNCID=501',
'Label=UH0766_U2, CellID=7662, LogicRNCID=501',
'Label=UH0708_U95, CellID=47085, LogicRNCID=501',
'Label=UH1918_U1, CellID=19181, LogicRNCID=501',
'Label=UH1892_U5, CellID=18925, LogicRNCID=501',
'Label=UH0766_U1, CellID=7661, LogicRNCID=501',
'Label=UH0759_U96, CellID=47596, LogicRNCID=501',
'Label=UH0708_U94, CellID=47084, LogicRNCID=501',
'Label=UH0759_U95, CellID=47595, LogicRNCID=501',
'Label=UH1892_U3, CellID=18923, LogicRNCID=501',
'Label=UH1892_U2, CellID=18922, LogicRNCID=501',
'Label=UH0970_U6, CellID=9706, LogicRNCID=501',
'Label=UH0842_U6, CellID=8426, LogicRNCID=501',
'Label=UH0714_U6, CellID=7146, LogicRNCID=501',
'Label=UH0970_U5, CellID=9705, LogicRNCID=501',
'Label=UH0842_U5, CellID=8425, LogicRNCID=501',
'Label=UH0714_U5, CellID=7145, LogicRNCID=501',
'Label=UH0970_U4, CellID=9704, LogicRNCID=501',
'Label=UH0842_U4, CellID=8424, LogicRNCID=501',
'Label=UH0765_U6, CellID=7656, LogicRNCID=501',
'Label=UH0714_U4, CellID=7144, LogicRNCID=501',
'Label=UH0970_U3, CellID=9703, LogicRNCID=501',
'Label=UH0842_U3, CellID=8423, LogicRNCID=501',
'Label=UH0765_U5, CellID=7655, LogicRNCID=501',
'Label=UH0714_U3, CellID=7143, LogicRNCID=501',
'Label=UH1936_U96, CellID=59366, LogicRNCID=501',
'Label=UH0970_U2, CellID=9702, LogicRNCID=501',
'Label=UH0842_U2, CellID=8422, LogicRNCID=501',
'Label=UH0765_U4, CellID=7654, LogicRNCID=501',
'Label=UH0714_U2, CellID=7142, LogicRNCID=501',
'Label=UH1936_U95, CellID=59365, LogicRNCID=501',
'Label=UH0970_U1, CellID=9701, LogicRNCID=501',
'Label=UH0842_U1, CellID=8421, LogicRNCID=501',
'Label=UH0765_U3, CellID=7653, LogicRNCID=501',
'Label=UH0714_U1, CellID=7141, LogicRNCID=501',
'Label=UH0963_U96, CellID=49636, LogicRNCID=501',
'Label=UH0835_U96, CellID=48356, LogicRNCID=501',
'Label=UH0995_U6, CellID=9956, LogicRNCID=501',
'Label=UH0765_U2, CellID=7652, LogicRNCID=501',
'Label=UH0611_U6, CellID=6116, LogicRNCID=501',
'Label=UH0963_U95, CellID=49635, LogicRNCID=501',
'Label=UH0835_U95, CellID=48355, LogicRNCID=501',
'Label=UH0995_U5, CellID=9955, LogicRNCID=501',
'Label=UH0765_U1, CellID=7651, LogicRNCID=501',
'Label=UH0963_U94, CellID=49634, LogicRNCID=501',
'Label=UH0835_U94, CellID=48354, LogicRNCID=501',
'Label=UH0758_U96, CellID=47586, LogicRNCID=501',
'Label=UH3990_U6, CellID=39906, LogicRNCID=501',
'Label=UH0995_U4, CellID=9954, LogicRNCID=501',
'Label=UH0611_U4, CellID=6114, LogicRNCID=501',
'Label=UH0758_U95, CellID=47585, LogicRNCID=501',
'Label=UH3990_U5, CellID=39905, LogicRNCID=501',
'Label=UH0995_U3, CellID=9953, LogicRNCID=501',
'Label=UH0611_U3, CellID=6113, LogicRNCID=501',
'Label=UH0758_U94, CellID=47584, LogicRNCID=501',
'Label=UH3990_U4, CellID=39904, LogicRNCID=501',
'Label=UH0995_U2, CellID=9952, LogicRNCID=501',
'Label=UH0841_U6, CellID=8416, LogicRNCID=501',
'Label=UH3990_U3, CellID=39903, LogicRNCID=501',
'Label=UH0995_U1, CellID=9951, LogicRNCID=501',
'Label=UH0841_U5, CellID=8415, LogicRNCID=501',
'Label=UH0611_U1, CellID=6111, LogicRNCID=501',
'Label=UH0732_U96, CellID=47326, LogicRNCID=501',
'Label=UH3990_U2, CellID=39902, LogicRNCID=501',
'Label=UH1916_U6, CellID=19166, LogicRNCID=501',
'Label=UH0841_U4, CellID=8414, LogicRNCID=501',
'Label=UH0636_U6, CellID=6366, LogicRNCID=501',
'Label=UH0508_U6, CellID=5086, LogicRNCID=501',
'Label=UH0732_U95, CellID=47325, LogicRNCID=501',
'Label=UH3990_U1, CellID=39901, LogicRNCID=501',
'Label=UH1916_U5, CellID=19165, LogicRNCID=501',
'Label=UH0841_U3, CellID=8413, LogicRNCID=501',
'Label=UH0636_U5, CellID=6365, LogicRNCID=501',
'Label=UH0508_U5, CellID=5085, LogicRNCID=501',
'Label=UH0732_U94, CellID=47324, LogicRNCID=501',
'Label=UH1916_U4, CellID=19164, LogicRNCID=501',
'Label=UH0841_U2, CellID=8412, LogicRNCID=501',
'Label=UH1916_U3, CellID=19163, LogicRNCID=501',
'Label=UH0841_U1, CellID=8411, LogicRNCID=501',
'Label=UH0636_U3, CellID=6363, LogicRNCID=501',
'Label=UH0508_U3, CellID=5083, LogicRNCID=501',
'Label=UH0962_U96, CellID=49626, LogicRNCID=501',
'Label=UH0706_U96, CellID=47066, LogicRNCID=501',
'Label=UH1916_U2, CellID=19162, LogicRNCID=501',
'Label=UH0994_U6, CellID=9946, LogicRNCID=501',
'Label=UH0738_U6, CellID=7386, LogicRNCID=501',
'Label=UH0636_U2, CellID=6362, LogicRNCID=501',
'Label=UH0610_U6, CellID=6106, LogicRNCID=501',
'Label=UH0508_U2, CellID=5082, LogicRNCID=501',
'Label=UH1907_U66, CellID=65497, LogicRNCID=501',
'Label=UH0962_U95, CellID=49625, LogicRNCID=501',
'Label=UH0706_U95, CellID=47065, LogicRNCID=501',
'Label=UH1916_U1, CellID=19161, LogicRNCID=501',
'Label=UH0994_U5, CellID=9945, LogicRNCID=501',
'Label=UH0738_U5, CellID=7385, LogicRNCID=501',
'Label=UH0610_U5, CellID=6105, LogicRNCID=501',
'Label=UH1907_U55, CellID=65496, LogicRNCID=501',
'Label=UH0962_U94, CellID=49624, LogicRNCID=501',
'Label=UH0706_U94, CellID=47064, LogicRNCID=501',
'Label=UH3989_U6, CellID=39896, LogicRNCID=501',
'Label=UH2709_U6, CellID=27096, LogicRNCID=501',
'Label=UH0994_U4, CellID=9944, LogicRNCID=501',
'Label=UH0738_U4, CellID=7384, LogicRNCID=501',
'Label=UH1907_U44, CellID=65495, LogicRNCID=501',
'Label=UH3989_U5, CellID=39895, LogicRNCID=501',
'Label=UH2709_U5, CellID=27095, LogicRNCID=501',
'Label=UH0994_U3, CellID=9943, LogicRNCID=501',
'Label=UH0738_U3, CellID=7383, LogicRNCID=501',
'Label=UH0610_U3, CellID=6103, LogicRNCID=501',
'Label=UH1907_U33, CellID=65494, LogicRNCID=501',
'Label=UH3989_U4, CellID=39894, LogicRNCID=501',
'Label=UH2709_U4, CellID=27094, LogicRNCID=501',
'Label=UH0994_U2, CellID=9942, LogicRNCID=501',
'Label=UH0840_U6, CellID=8406, LogicRNCID=501',
'Label=UH0738_U2, CellID=7382, LogicRNCID=501',
'Label=UH0610_U2, CellID=6102, LogicRNCID=501',
'Label=UH1907_U22, CellID=65493, LogicRNCID=501',
'Label=UH3989_U3, CellID=39893, LogicRNCID=501',
'Label=UH2709_U3, CellID=27093, LogicRNCID=501',
'Label=UH0994_U1, CellID=9941, LogicRNCID=501',
'Label=UH0840_U5, CellID=8405, LogicRNCID=501',
'Label=UH0738_U1, CellID=7381, LogicRNCID=501',
'Label=UH1907_U11, CellID=65492, LogicRNCID=501',
'Label=UH0731_U96, CellID=47316, LogicRNCID=501',
'Label=UH3989_U2, CellID=39892, LogicRNCID=501',
'Label=UH2709_U2, CellID=27092, LogicRNCID=501',
'Label=UH0840_U4, CellID=8404, LogicRNCID=501',
'Label=UH0763_U6, CellID=7636, LogicRNCID=501',
'Label=UH0731_U95, CellID=47315, LogicRNCID=501',
'Label=UH3989_U1, CellID=39891, LogicRNCID=501',
'Label=UH2709_U1, CellID=27091, LogicRNCID=501',
'Label=UH0840_U3, CellID=8403, LogicRNCID=501',
'Label=UH0763_U5, CellID=7635, LogicRNCID=501',
'Label=UH0731_U94, CellID=47314, LogicRNCID=501',
'Label=UH0840_U2, CellID=8402, LogicRNCID=501',
'Label=UH0763_U4, CellID=7634, LogicRNCID=501',
'Label=UH0840_U1, CellID=8401, LogicRNCID=501',
'Label=UH0763_U3, CellID=7633, LogicRNCID=501',
'Label=UH0833_U96, CellID=48336, LogicRNCID=501',
'Label=UH0705_U96, CellID=47056, LogicRNCID=501',
'Label=UH2913_U6, CellID=29136, LogicRNCID=501',
'Label=UH0763_U2, CellID=7632, LogicRNCID=501',
'Label=UH0833_U95, CellID=48335, LogicRNCID=501',
'Label=UH0705_U95, CellID=47055, LogicRNCID=501',
'Label=UH2913_U5, CellID=29135, LogicRNCID=501',
'Label=UH0763_U1, CellID=7631, LogicRNCID=501',
'Label=UH0833_U94, CellID=48334, LogicRNCID=501',
'Label=UH0756_U96, CellID=47566, LogicRNCID=501',
'Label=UH0705_U94, CellID=47054, LogicRNCID=501',
'Label=UH2913_U4, CellID=29134, LogicRNCID=501',
'Label=UH2913_U3, CellID=29133, LogicRNCID=501',
'Label=UH0807_U96, CellID=48076, LogicRNCID=501',
'Label=UH0756_U94, CellID=47564, LogicRNCID=501',
'Label=UH2913_U2, CellID=29132, LogicRNCID=501',
'Label=UH0839_U6, CellID=8396, LogicRNCID=501',
'Label=UH0711_U6, CellID=7116, LogicRNCID=501',
'Label=UH0807_U95, CellID=48075, LogicRNCID=501',
'Label=UH2913_U1, CellID=29131, LogicRNCID=501',
'Label=UH0839_U5, CellID=8395, LogicRNCID=501',
'Label=UH0711_U5, CellID=7115, LogicRNCID=501',
'Label=UH0807_U94, CellID=48074, LogicRNCID=501',
'Label=UH0602_U96, CellID=46026, LogicRNCID=501',
'Label=UH2938_U6, CellID=29386, LogicRNCID=501',
'Label=UH0839_U4, CellID=8394, LogicRNCID=501',
'Label=UH0711_U4, CellID=7114, LogicRNCID=501',
'Label=UH0602_U95, CellID=46025, LogicRNCID=501',
'Label=UH2938_U5, CellID=29385, LogicRNCID=501',
'Label=UH0839_U3, CellID=8393, LogicRNCID=501',
'Label=UH0711_U3, CellID=7113, LogicRNCID=501',
'Label=UH0602_U94, CellID=46024, LogicRNCID=501',
'Label=UH2938_U4, CellID=29384, LogicRNCID=501',
'Label=UH0839_U2, CellID=8392, LogicRNCID=501',
'Label=UH0711_U2, CellID=7112, LogicRNCID=501',
'Label=UH2938_U3, CellID=29383, LogicRNCID=501',
'Label=UH0839_U1, CellID=8391, LogicRNCID=501',
'Label=UH0711_U1, CellID=7111, LogicRNCID=501',
'Label=UH0960_U96, CellID=49606, LogicRNCID=501',
'Label=UH0832_U96, CellID=48326, LogicRNCID=501',
'Label=UH0704_U96, CellID=47046, LogicRNCID=501',
'Label=UH2938_U2, CellID=29382, LogicRNCID=501',
'Label=UH0960_U95, CellID=49605, LogicRNCID=501',
'Label=UH0832_U95, CellID=48325, LogicRNCID=501',
'Label=UH0704_U95, CellID=47045, LogicRNCID=501',
'Label=UH2938_U1, CellID=29381, LogicRNCID=501',
'Label=UH0960_U94, CellID=49604, LogicRNCID=501',
'Label=UH0832_U94, CellID=48324, LogicRNCID=501',
'Label=UH1939_U6, CellID=19396, LogicRNCID=501',
'Label=UH1939_U5, CellID=19395, LogicRNCID=501',
'Label=UH1939_U4, CellID=19394, LogicRNCID=501',
'Label=UH0966_U6, CellID=9666, LogicRNCID=501',
'Label=UH0838_U6, CellID=8386, LogicRNCID=501',
'Label=UH1939_U3, CellID=19393, LogicRNCID=501',
'Label=UH0961_U6, CellID=9616, LogicRNCID=501',
'Label=UH1951_U95, CellID=50063, LogicRNCID=501',
'Label=UH0961_U5, CellID=9615, LogicRNCID=501',
'Label=UH1951_U94, CellID=50062, LogicRNCID=501',
'Label=UH2932_U6, CellID=29326, LogicRNCID=501',
'Label=UH1908_U6, CellID=19086, LogicRNCID=501',
'Label=UH0961_U4, CellID=9614, LogicRNCID=501',
'Label=UH1945_U99, CellID=50061, LogicRNCID=501',
'Label=UH1907_U4, CellID=19074, LogicRNCID=501',
'Label=UH1702_U6, CellID=17026, LogicRNCID=501',
'Label=UH0960_U2, CellID=9602, LogicRNCID=501',
'Label=UH0832_U2, CellID=8322, LogicRNCID=501',
'Label=UH1928_U95, CellID=50049, LogicRNCID=501',
'Label=UH1984_U1, CellID=19841, LogicRNCID=501',
'Label=UH1958_U5, CellID=19585, LogicRNCID=501',
'Label=UH1907_U3, CellID=19073, LogicRNCID=501',
'Label=UH1907_U5, CellID=19075, LogicRNCID=501',
'Label=UH1805_U1, CellID=18051, LogicRNCID=501',
'Label=UH0960_U3, CellID=9603, LogicRNCID=501',
'Label=UH0832_U3, CellID=8323, LogicRNCID=501',
'Label=UH1928_U96, CellID=50050, LogicRNCID=501',
'Label=UH0723_U94, CellID=47234, LogicRNCID=501',
'Label=UH1984_U2, CellID=19842, LogicRNCID=501',
'Label=UH1958_U6, CellID=19586, LogicRNCID=501',
'Label=UH0966_U3, CellID=9663, LogicRNCID=501',
'Label=UH0838_U3, CellID=8383, LogicRNCID=501',
'Label=UH0601_U94, CellID=46014, LogicRNCID=501',
'Label=UH0966_U2, CellID=9662, LogicRNCID=501',
'Label=UH0838_U2, CellID=8382, LogicRNCID=501',
'Label=UH1964_U5, CellID=19645, LogicRNCID=501',
'Label=UH0966_U1, CellID=9661, LogicRNCID=501',
'Label=UH0838_U1, CellID=8381, LogicRNCID=501',
'Label=UH0602_U1, CellID=6021, LogicRNCID=501',
'Label=UH0723_U96, CellID=47236, LogicRNCID=501',
'Label=UH1984_U4, CellID=19844, LogicRNCID=501',
'Label=UH1907_U6, CellID=19076, LogicRNCID=501',
'Label=UH1805_U2, CellID=18052, LogicRNCID=501',
'Label=UH0960_U4, CellID=9604, LogicRNCID=501',
'Label=UH0832_U4, CellID=8324, LogicRNCID=501',
'Label=UH0723_U95, CellID=47235, LogicRNCID=501',
'Label=UH1707_U1, CellID=17071, LogicRNCID=501',
'Label=UH0734_U3, CellID=7343, LogicRNCID=501',
'Label=UH0881_U94, CellID=48814, LogicRNCID=501',
'Label=UH0734_U2, CellID=7342, LogicRNCID=501',
'Label=UH0708_U6, CellID=7086, LogicRNCID=501',
'Label=UH0734_U1, CellID=7341, LogicRNCID=501',
'Label=UH0708_U5, CellID=7085, LogicRNCID=501',
'Label=UH0708_U4, CellID=7084, LogicRNCID=501',
'Label=UH0960_U6, CellID=9606, LogicRNCID=501',
'Label=UH0832_U6, CellID=8326, LogicRNCID=501',
'Label=UH0602_U2, CellID=6022, LogicRNCID=501',
'Label=UH1984_U5, CellID=19845, LogicRNCID=501',
'Label=UH1882_U1, CellID=18821, LogicRNCID=501',
'Label=UH1805_U3, CellID=18053, LogicRNCID=501',
'Label=UH0960_U5, CellID=9605, LogicRNCID=501',
'Label=UH0832_U5, CellID=8325, LogicRNCID=501',
'Label=UH0966_U5, CellID=9665, LogicRNCID=501',
'Label=UH0838_U5, CellID=8385, LogicRNCID=501',
'Label=UH0601_U96, CellID=46016, LogicRNCID=501',
'Label=UH1939_U2, CellID=19392, LogicRNCID=501',
'Label=UH0966_U4, CellID=9664, LogicRNCID=501',
'Label=UH0838_U4, CellID=8384, LogicRNCID=501',
'Label=UH0601_U95, CellID=46015, LogicRNCID=501',
'Label=UH1939_U1, CellID=19391, LogicRNCID=501',
'Label=UH1705_U6, CellID=17056, LogicRNCID=501',
'Label=UH0963_U2, CellID=9632, LogicRNCID=501',
'Label=UH0835_U2, CellID=8352, LogicRNCID=501',
'Label=UH0758_U4, CellID=7584, LogicRNCID=501',
'Label=UH1833_U5, CellID=18335, LogicRNCID=501',
'Label=UH1705_U5, CellID=17055, LogicRNCID=501',
'Label=UH0963_U1, CellID=9631, LogicRNCID=501',
'Label=UH0835_U1, CellID=8351, LogicRNCID=501',
'Label=UH0623_U95, CellID=46235, LogicRNCID=501',
'Label=UH1935_U5, CellID=19355, LogicRNCID=501',
'Label=UH1833_U1, CellID=18331, LogicRNCID=501',
'Label=UH1705_U1, CellID=17051, LogicRNCID=501',
'Label=UH0988_U3, CellID=9883, LogicRNCID=501',
'Label=UH0760_U97, CellID=50074, LogicRNCID=501',
'Label=UH0751_U94, CellID=47514, LogicRNCID=501',
'Label=UH0623_U94, CellID=46234, LogicRNCID=501',
'Label=UH2984_U4, CellID=29844, LogicRNCID=501',
'Label=UH1960_U4, CellID=19604, LogicRNCID=501',
'Label=UH1704_U4, CellID=17044, LogicRNCID=501',
'Label=UH2709_U96, CellID=50067, LogicRNCID=501',
'Label=UH0955_U95, CellID=49555, LogicRNCID=501',
'Label=UH2984_U3, CellID=29843, LogicRNCID=501',
'Label=UH1960_U3, CellID=19603, LogicRNCID=501',
'Label=UH1704_U3, CellID=17043, LogicRNCID=501',
'Label=UH1707_U3, CellID=17073, LogicRNCID=501',
'Label=UH0734_U5, CellID=7345, LogicRNCID=501',
'Label=UH0881_U96, CellID=48816, LogicRNCID=501',
'Label=UH0830_U94, CellID=48304, LogicRNCID=501',
'Label=UH0702_U94, CellID=47024, LogicRNCID=501',
'Label=UH1707_U2, CellID=17072, LogicRNCID=501',
'Label=UH0734_U4, CellID=7344, LogicRNCID=501',
'Label=UH0881_U95, CellID=48815, LogicRNCID=501',
'Label=UH2983_U1, CellID=29831, LogicRNCID=501',
'Label=UH1882_U3, CellID=18823, LogicRNCID=501',
'Label=UH1805_U5, CellID=18055, LogicRNCID=501',
'Label=UH1703_U1, CellID=17031, LogicRNCID=501',
'Label=UH0807_U1, CellID=8071, LogicRNCID=501',
'Label=UH0602_U3, CellID=6023, LogicRNCID=501',
'Label=UH1882_U2, CellID=18822, LogicRNCID=501',
'Label=UH1805_U4, CellID=18054, LogicRNCID=501',
'Label=UH1935_U2, CellID=19352, LogicRNCID=501',
'Label=UH0962_U4, CellID=9624, LogicRNCID=501',
'Label=UH0706_U4, CellID=7064, LogicRNCID=501',
'Label=UH0760_U94, CellID=50071, LogicRNCID=501',
'Label=UH1935_U1, CellID=19351, LogicRNCID=501',
'Label=UH0962_U3, CellID=9623, LogicRNCID=501',
'Label=UH2984_U6, CellID=29846, LogicRNCID=501',
'Label=UH1960_U6, CellID=19606, LogicRNCID=501',
'Label=UH0961_U2, CellID=9612, LogicRNCID=501',
'Label=UH0807_U6, CellID=8076, LogicRNCID=501',
'Label=UH1945_U97, CellID=50059, LogicRNCID=501',
'Label=UH2983_U5, CellID=29835, LogicRNCID=501',
'Label=UH2932_U3, CellID=29323, LogicRNCID=501',
'Label=UH1908_U3, CellID=19083, LogicRNCID=501',
'Label=UH1703_U5, CellID=17035, LogicRNCID=501',
'Label=UH0961_U1, CellID=9611, LogicRNCID=501',
'Label=UH0835_U4, CellID=8354, LogicRNCID=501',
'Label=UH0758_U6, CellID=7586, LogicRNCID=501',
'Label=UH1936_U1, CellID=19361, LogicRNCID=501',
'Label=UH0963_U3, CellID=9633, LogicRNCID=501',
'Label=UH0835_U3, CellID=8353, LogicRNCID=501',
'Label=UH0758_U5, CellID=7585, LogicRNCID=501',
'Label=UH1801_U96, CellID=58016, LogicRNCID=501',
'Label=UH1833_U6, CellID=18336, LogicRNCID=501',
'Label=UH0702_U96, CellID=47026, LogicRNCID=501',
'Label=UH4600_U2, CellID=46002, LogicRNCID=501',
'Label=UH1707_U4, CellID=17074, LogicRNCID=501',
'Label=UH0734_U6, CellID=7346, LogicRNCID=501',
'Label=UH0881_U97, CellID=48817, LogicRNCID=501',
'Label=UH0830_U95, CellID=48305, LogicRNCID=501',
'Label=UH0702_U95, CellID=47025, LogicRNCID=501',
'Label=UH4600_U1, CellID=46001, LogicRNCID=501',
'Label=UH0602_U6, CellID=6026, LogicRNCID=501',
'Label=UH1945_U95, CellID=50057, LogicRNCID=501',
'Label=UH2983_U3, CellID=29833, LogicRNCID=501',
'Label=UH2932_U1, CellID=29321, LogicRNCID=501',
'Label=UH1908_U1, CellID=19081, LogicRNCID=501',
'Label=UH1882_U5, CellID=18825, LogicRNCID=501',
'Label=UH1703_U3, CellID=17033, LogicRNCID=501',
'Label=UH0807_U3, CellID=8073, LogicRNCID=501',
'Label=UH0602_U5, CellID=6025, LogicRNCID=501',
'Label=UH1945_U94, CellID=50056, LogicRNCID=501',
'Label=UH2983_U2, CellID=29832, LogicRNCID=501',
'Label=UH1882_U4, CellID=18824, LogicRNCID=501',
'Label=UH0748_U5, CellID=7485, LogicRNCID=501',
'Label=UH0767_U96, CellID=47676, LogicRNCID=501',
'Label=UH0825_U2, CellID=8252, LogicRNCID=501',
'Label=UH0767_U95, CellID=47675, LogicRNCID=501',
'Label=UH0742_U4, CellID=7424, LogicRNCID=501',
'Label=UH1951_U5, CellID=19515, LogicRNCID=501',
'Label=UH0748_U4, CellID=7484, LogicRNCID=501',
'Label=UH1951_U6, CellID=19516, LogicRNCID=501',
'Label=UH0849_U6, CellID=8496, LogicRNCID=501',
'Label=UH3921_U5, CellID=39215, LogicRNCID=501',
'Label=UH0740_U96, CellID=47406, LogicRNCID=501',
'Label=UH0849_U4, CellID=8494, LogicRNCID=501',
'Label=UH0717_U1, CellID=7171, LogicRNCID=501',
'Label=UH0838_U95, CellID=48385, LogicRNCID=501',
'Label=UH0742_U5, CellID=7425, LogicRNCID=501',
'Label=UH0761_U96, CellID=47616, LogicRNCID=501',
'Label=UH0825_U5, CellID=8255, LogicRNCID=501',
'Label=UH2950_U2, CellID=29502, LogicRNCID=501',
'Label=UH0748_U6, CellID=7486, LogicRNCID=501',
'Label=UH2950_U1, CellID=29501, LogicRNCID=501',
'Label=UH0765_U94, CellID=47654, LogicRNCID=501',
'Label=UH3920_U5, CellID=39205, LogicRNCID=501',
'Label=UH0739_U96, CellID=47396, LogicRNCID=501',
'Label=UH3920_U4, CellID=39204, LogicRNCID=501',
'Label=UH0822_U3, CellID=8223, LogicRNCID=501',
'Label=UH0841_U94, CellID=48414, LogicRNCID=501',
'Label=UH0636_U96, CellID=46366, LogicRNCID=501',
'Label=UH0950_U2, CellID=9502, LogicRNCID=501',
'Label=UH0821_U4, CellID=8214, LogicRNCID=501',
'Label=UH0879_U95, CellID=55061, LogicRNCID=501',
'Label=UH0821_U3, CellID=8213, LogicRNCID=501',
'Label=UH0879_U96, CellID=55060, LogicRNCID=501',
'Label=UH0715_U94, CellID=47154, LogicRNCID=501',
'Label=UH0952_U2, CellID=9522, LogicRNCID=501',
'Label=UH1976_U1, CellID=19761, LogicRNCID=501',
'Label=UH0766_U96, CellID=47664, LogicRNCID=501',
'Label=UH0736_U94, CellID=47364, LogicRNCID=501',
'Label=UH0973_U1, CellID=9731, LogicRNCID=501',
'Label=UH0617_U2, CellID=6172, LogicRNCID=501',
'Label=UH0617_U1, CellID=6171, LogicRNCID=501',
'Label=UH3996_U2, CellID=39962, LogicRNCID=501',
'Label=UH0738_U95, CellID=47385, LogicRNCID=501',
'Label=UH0839_U97, CellID=48397, LogicRNCID=501',
'Label=UH0974_U1, CellID=9741, LogicRNCID=501',
'Label=UH0820_U5, CellID=8205, LogicRNCID=501',
'Label=UH0711_U96, CellID=47116, LogicRNCID=501',
'Label=UH0618_U5, CellID=6185, LogicRNCID=501',
'Label=UH0714_U94, CellID=47144, LogicRNCID=501',
'Label=UH0618_U4, CellID=6184, LogicRNCID=501',
'Label=UH0823_U1, CellID=8231, LogicRNCID=501',
'Label=UH0952_U4, CellID=9524, LogicRNCID=501',
'Label=UH0715_U95, CellID=47155, LogicRNCID=501',
'Label=UH0952_U3, CellID=9523, LogicRNCID=501',
'Label=UH0843_U94, CellID=48434, LogicRNCID=501',
'Label=UH0820_U1, CellID=8201, LogicRNCID=501',
'Label=UH0973_U6, CellID=9736, LogicRNCID=501',
'Label=UH0743_U2, CellID=7432, LogicRNCID=501',
'Label=UH0973_U5, CellID=9735, LogicRNCID=501',
'Label=UH0743_U1, CellID=7431, LogicRNCID=501',
'Label=UH0736_U96, CellID=47366, LogicRNCID=501',
'Label=UH0736_U95, CellID=47365, LogicRNCID=501',
'Label=UH0841_U96, CellID=48416, LogicRNCID=501',
'Label=UH0950_U4, CellID=9504, LogicRNCID=501',
'Label=UH0617_U6, CellID=6176, LogicRNCID=501',
'Label=UH0713_U95, CellID=47135, LogicRNCID=501',
'Label=UH0763_U96, CellID=47636, LogicRNCID=501',
'Label=UH1947_U6, CellID=19476, LogicRNCID=501',
'Label=UH0763_U95, CellID=47635, LogicRNCID=501',
'Label=UH1947_U5, CellID=19475, LogicRNCID=501',
'Label=UH1976_U6, CellID=19766, LogicRNCID=501',
'Label=UH3922_U1, CellID=39221, LogicRNCID=501',
'Label=UH0952_U5, CellID=9525, LogicRNCID=501',
'Label=UH0715_U96, CellID=47156, LogicRNCID=501',
'Label=UH0743_U6, CellID=7436, LogicRNCID=501',
'Label=UH0711_U95, CellID=47115, LogicRNCID=501',
'Label=UH0839_U94, CellID=48394, LogicRNCID=501',
'Label=UH0820_U2, CellID=8202, LogicRNCID=501',
'Label=UH0849_U1, CellID=8491, LogicRNCID=501',
'Label=UH0951_U4, CellID=9514, LogicRNCID=501',
'Label=UH0618_U6, CellID=6186, LogicRNCID=501',
'Label=UH0951_U3, CellID=9513, LogicRNCID=501',
'Label=UH0737_U95, CellID=47375, LogicRNCID=501',
'Label=UH3918_U3, CellID=39183, LogicRNCID=501',
'Label=UH0974_U3, CellID=9743, LogicRNCID=501',
'Label=UH0737_U94, CellID=47374, LogicRNCID=501',
'Label=UH0763_U94, CellID=47634, LogicRNCID=501',
'Label=UH3918_U6, CellID=39186, LogicRNCID=501',
'Label=UH0974_U6, CellID=9746, LogicRNCID=501',
'Label=UH3995_U3, CellID=39953, LogicRNCID=501',
'Label=UH0748_U1, CellID=7481, LogicRNCID=501',
'Label=UH3922_U4, CellID=39224, LogicRNCID=501',
'Label=UH0869_U95, CellID=48695, LogicRNCID=501',
'Label=UH0869_U94, CellID=48694, LogicRNCID=501',
'Label=UH0617_U4, CellID=6174, LogicRNCID=501',
'Label=UH3996_U5, CellID=39965, LogicRNCID=501',
'Label=UH0822_U1, CellID=8221, LogicRNCID=501',
'Label=UH0636_U94, CellID=46364, LogicRNCID=501',
'Label=UH0770_U5, CellID=7705, LogicRNCID=501',
'Label=UH0821_U6, CellID=8216, LogicRNCID=501',
'Label=UH0821_U5, CellID=8215, LogicRNCID=501',
'Label=UH0879_U94, CellID=55062, LogicRNCID=501',
'Label=UH0611_U95, CellID=46115, LogicRNCID=501',
'Label=UH3920_U2, CellID=39202, LogicRNCID=501',
'Label=UH0822_U6, CellID=8226, LogicRNCID=501',
'Label=UH0950_U5, CellID=9505, LogicRNCID=501',
'Label=UH3921_U3, CellID=39213, LogicRNCID=501',
'Label=UH0740_U94, CellID=47404, LogicRNCID=501',
'Label=UH0951_U6, CellID=9516, LogicRNCID=501',
'Label=UH0823_U6, CellID=8236, LogicRNCID=501',
'Label=UH0825_U1, CellID=8251, LogicRNCID=501',
'Label=UH0767_U94, CellID=47674, LogicRNCID=501',
'Label=UH0748_U2, CellID=7482, LogicRNCID=501',
'Label=UH3922_U5, CellID=39225, LogicRNCID=501',
'Label=UH1951_U3, CellID=19513, LogicRNCID=501',
'Label=UH0869_U97, CellID=48697, LogicRNCID=501',
'Label=UH1951_U4, CellID=19514, LogicRNCID=501',
'Label=UH0748_U3, CellID=7483, LogicRNCID=501',
'Label=UH3921_U1, CellID=39211, LogicRNCID=501',
'Label=UH0849_U2, CellID=8492, LogicRNCID=501',
'Label=UH3921_U2, CellID=39212, LogicRNCID=501',
'Label=UH0849_U3, CellID=8493, LogicRNCID=501',
'Label=UH0822_U5, CellID=8225, LogicRNCID=501',
'Label=UH3920_U1, CellID=39201, LogicRNCID=501',
'Label=UH0950_U6, CellID=9506, LogicRNCID=501',
'Label=UH0611_U94, CellID=46114, LogicRNCID=501',
'Label=UH0840_U96, CellID=48406, LogicRNCID=501',
'Label=UH0770_U3, CellID=7703, LogicRNCID=501',
'Label=UH0770_U4, CellID=7704, LogicRNCID=501',
'Label=UH0738_U94, CellID=47384, LogicRNCID=501',
'Label=UH3996_U4, CellID=39964, LogicRNCID=501',
'Label=UH0617_U3, CellID=6173, LogicRNCID=501',
'Label=UH0950_U1, CellID=9501, LogicRNCID=501',
'Label=UH0636_U95, CellID=46365, LogicRNCID=501',
'Label=UH3922_U2, CellID=39222, LogicRNCID=501',
'Label=UH1951_U1, CellID=19511, LogicRNCID=501',
'Label=UH1951_U2, CellID=19512, LogicRNCID=501',
'Label=UH0869_U96, CellID=48696, LogicRNCID=501',
'Label=UH1947_U3, CellID=19473, LogicRNCID=501',
'Label=UH0846_U6, CellID=8466, LogicRNCID=501',
'Label=UH1947_U4, CellID=19474, LogicRNCID=501',
'Label=UH3995_U4, CellID=39954, LogicRNCID=501',
'Label=UH0974_U2, CellID=9742, LogicRNCID=501',
'Label=UH0846_U3, CellID=8463, LogicRNCID=501',
'Label=UH1947_U1, CellID=19471, LogicRNCID=501',
'Label=UH3995_U1, CellID=39951, LogicRNCID=501',
'Label=UH0823_U3, CellID=8233, LogicRNCID=501',
'Label=UH0714_U95, CellID=47145, LogicRNCID=501',
'Label=UH0823_U4, CellID=8234, LogicRNCID=501',
'Label=UH0823_U5, CellID=8235, LogicRNCID=501',
'Label=UH0743_U4, CellID=7434, LogicRNCID=501',
'Label=UH0711_U94, CellID=47114, LogicRNCID=501',
'Label=UH0743_U5, CellID=7435, LogicRNCID=501',
'Label=UH0839_U95, CellID=48395, LogicRNCID=501',
'Label=UH1976_U4, CellID=19764, LogicRNCID=501',
'Label=UH0843_U96, CellID=48436, LogicRNCID=501',
'Label=UH1976_U5, CellID=19765, LogicRNCID=501',
'Label=UH0952_U6, CellID=9526, LogicRNCID=501',
'Label=UH0821_U1, CellID=8211, LogicRNCID=501',
'Label=UH3995_U5, CellID=39955, LogicRNCID=501',
'Label=UH0821_U2, CellID=8212, LogicRNCID=501',
'Label=UH3995_U6, CellID=39956, LogicRNCID=501',
'Label=UH0950_U3, CellID=9503, LogicRNCID=501',
'Label=UH0841_U95, CellID=48415, LogicRNCID=501',
'Label=UH0822_U4, CellID=8224, LogicRNCID=501',
'Label=UH0713_U96, CellID=47136, LogicRNCID=501',
'Label=UH0973_U3, CellID=9733, LogicRNCID=501',
'Label=UH0717_U4, CellID=7174, LogicRNCID=501',
'Label=UH0973_U4, CellID=9734, LogicRNCID=501',
'Label=UH0717_U5, CellID=7175, LogicRNCID=501',
'Label=UH0717_U6, CellID=7176, LogicRNCID=501',
'Label=UH0743_U3, CellID=7433, LogicRNCID=501',
'Label=UH0766_U94, CellID=47666, LogicRNCID=501',
'Label=UH0747_U5, CellID=7475, LogicRNCID=501',
'Label=UH1976_U3, CellID=19763, LogicRNCID=501',
'Label=UH0843_U95, CellID=48435, LogicRNCID=501',
'Label=UH0618_U3, CellID=6183, LogicRNCID=501',
'Label=UH0951_U1, CellID=9511, LogicRNCID=501',
'Label=UH0823_U2, CellID=8232, LogicRNCID=501',
'Label=UH0765_U96, CellID=47656, LogicRNCID=501',
'Label=UH0820_U4, CellID=8204, LogicRNCID=501',
'Label=UH0839_U96, CellID=48396, LogicRNCID=501',
'Label=UH0846_U1, CellID=8461, LogicRNCID=501',
'Label=UH3918_U1, CellID=39181, LogicRNCID=501',
'Label=UH3996_U1, CellID=39961, LogicRNCID=501',
'Label=UH0770_U6, CellID=7706, LogicRNCID=501',
'Label=UH0738_U96, CellID=47386, LogicRNCID=501',
'Label=UH3996_U3, CellID=39963, LogicRNCID=501',
'Label=UH0717_U2, CellID=7172, LogicRNCID=501',
'Label=UH0973_U2, CellID=9732, LogicRNCID=501',
'Label=UH0717_U3, CellID=7173, LogicRNCID=501',
'Label=UH3921_U6, CellID=39216, LogicRNCID=501',
'Label=UH0952_U1, CellID=9521, LogicRNCID=501',
'Label=UH0766_U95, CellID=47665, LogicRNCID=501',
'Label=UH1976_U2, CellID=19762, LogicRNCID=501',
'Label=UH0840_U94, CellID=48404, LogicRNCID=501',
'Label=UH0770_U1, CellID=7701, LogicRNCID=501',
'Label=UH0840_U95, CellID=48405, LogicRNCID=501',
'Label=UH0770_U2, CellID=7702, LogicRNCID=501',
'Label=UH0822_U2, CellID=8222, LogicRNCID=501',
'Label=UH3996_U6, CellID=39966, LogicRNCID=501',
'Label=UH0713_U94, CellID=47134, LogicRNCID=501',
'Label=UH0617_U5, CellID=6175, LogicRNCID=501',
'Label=UH0739_U95, CellID=47395, LogicRNCID=501',
'Label=UH0611_U96, CellID=46116, LogicRNCID=501',
'Label=UH0618_U1, CellID=6181, LogicRNCID=501',
'Label=UH0618_U2, CellID=6182, LogicRNCID=501',
'Label=UH0825_U3, CellID=8253, LogicRNCID=501',
'Label=UH0844_U95, CellID=48445, LogicRNCID=501',
'Label=UH0825_U4, CellID=8254, LogicRNCID=501',
'Label=UH0844_U96, CellID=48446, LogicRNCID=501',
'Label=UH1945_U6, CellID=19456, LogicRNCID=501',
'Label=UH0838_U94, CellID=48384, LogicRNCID=501',
'Label=UH0761_U97, CellID=47617, LogicRNCID=501',
'Label=UH0838_U96, CellID=48386, LogicRNCID=501',
'Label=UH0740_U95, CellID=47405, LogicRNCID=501',
'Label=UH3921_U4, CellID=39214, LogicRNCID=501',
'Label=UH0849_U5, CellID=8495, LogicRNCID=501',
'Label=UH0747_U2, CellID=7472, LogicRNCID=501',
'Label=UH0844_U94, CellID=48444, LogicRNCID=501',
'Label=UH1805_U6, CellID=18056, LogicRNCID=501',
'Label=UH1703_U2, CellID=17032, LogicRNCID=501',
'Label=UH0807_U2, CellID=8072, LogicRNCID=501',
'Label=UH0602_U4, CellID=6024, LogicRNCID=501',
'Label=UH0988_U5, CellID=9885, LogicRNCID=501',
'Label=UH0758_U1, CellID=7581, LogicRNCID=501',
'Label=UH0751_U96, CellID=47516, LogicRNCID=501',
'Label=UH0623_U96, CellID=46236, LogicRNCID=501',
'Label=UH1935_U6, CellID=19356, LogicRNCID=501',
'Label=UH1833_U2, CellID=18332, LogicRNCID=501',
'Label=UH1705_U2, CellID=17052, LogicRNCID=501',
'Label=UH0988_U4, CellID=9884, LogicRNCID=501',
'Label=UH2709_U95, CellID=50066, LogicRNCID=501',
'Label=UH0955_U94, CellID=49554, LogicRNCID=501',
'Label=UH0750_U96, CellID=47506, LogicRNCID=501',
'Label=UH0622_U96, CellID=46226, LogicRNCID=501',
'Label=UH2984_U2, CellID=29842, LogicRNCID=501',
'Label=UH1960_U2, CellID=19602, LogicRNCID=501',
'Label=UH1704_U2, CellID=17042, LogicRNCID=501',
'Label=UH2709_U94, CellID=50065, LogicRNCID=501',
'Label=UH0965_U3, CellID=9653, LogicRNCID=501',
'Label=UH4600_U4, CellID=46004, LogicRNCID=501',
'Label=UH1707_U6, CellID=17076, LogicRNCID=501',
'Label=UH0965_U2, CellID=9652, LogicRNCID=501',
'Label=UH4600_U3, CellID=46003, LogicRNCID=501',
'Label=UH1707_U5, CellID=17075, LogicRNCID=501',
'Label=UH0965_U1, CellID=9651, LogicRNCID=501',
'Label=UH0830_U96, CellID=48306, LogicRNCID=501',
'Label=UH0807_U5, CellID=8075, LogicRNCID=501',
'Label=UH1945_U96, CellID=50058, LogicRNCID=501',
'Label=UH2983_U4, CellID=29834, LogicRNCID=501',
'Label=UH2932_U2, CellID=29322, LogicRNCID=501',
'Label=UH1908_U2, CellID=19082, LogicRNCID=501',
'Label=UH1882_U6, CellID=18826, LogicRNCID=501',
'Label=UH1703_U4, CellID=17034, LogicRNCID=501',
'Label=UH0807_U4, CellID=8074, LogicRNCID=501',
'Label=UH1936_U4, CellID=19364, LogicRNCID=501',
'Label=UH0963_U6, CellID=9636, LogicRNCID=501',
'Label=UH0835_U6, CellID=8356, LogicRNCID=501',
'Label=UH1936_U3, CellID=19363, LogicRNCID=501',
'Label=UH0963_U5, CellID=9635, LogicRNCID=501',
'Label=UH0835_U5, CellID=8355, LogicRNCID=501',
'Label=UH1936_U2, CellID=19362, LogicRNCID=501',
'Label=UH0963_U4, CellID=9634, LogicRNCID=501',
'Label=UH2932_U5, CellID=29325, LogicRNCID=501',
'Label=UH1908_U5, CellID=19085, LogicRNCID=501',
'Label=UH0961_U3, CellID=9613, LogicRNCID=501',
'Label=UH1945_U98, CellID=50060, LogicRNCID=501',
'Label=UH2983_U6, CellID=29836, LogicRNCID=501',
'Label=UH2932_U4, CellID=29324, LogicRNCID=501',
'Label=UH1908_U4, CellID=19084, LogicRNCID=501',
'Label=UH1703_U6, CellID=17036, LogicRNCID=501',
'Label=UH0750_U95, CellID=47505, LogicRNCID=501',
'Label=UH0622_U95, CellID=46225, LogicRNCID=501',
'Label=UH2984_U1, CellID=29841, LogicRNCID=501',
'Label=UH1960_U1, CellID=19601, LogicRNCID=501',
'Label=UH1704_U1, CellID=17041, LogicRNCID=501',
'Label=UH1951_U96, CellID=50064, LogicRNCID=501',
'Label=UH0750_U94, CellID=47504, LogicRNCID=501',
'Label=UH0622_U94, CellID=46224, LogicRNCID=501',
'Label=UH0754_U95, CellID=47545, LogicRNCID=501',
'Label=UH1964_U1, CellID=19641, LogicRNCID=501',
'Label=UH0754_U94, CellID=47544, LogicRNCID=501',
'Label=UH0965_U6, CellID=9656, LogicRNCID=501',
'Label=UH0965_U5, CellID=9655, LogicRNCID=501',
'Label=UH4600_U6, CellID=46006, LogicRNCID=501',
'Label=UH0965_U4, CellID=9654, LogicRNCID=501',
'Label=UH4600_U5, CellID=46005, LogicRNCID=501',
'Label=UH1935_U4, CellID=19354, LogicRNCID=501',
'Label=UH0988_U2, CellID=9882, LogicRNCID=501',
'Label=UH0962_U6, CellID=9626, LogicRNCID=501',
'Label=UH0760_U96, CellID=50073, LogicRNCID=501',
'Label=UH1935_U3, CellID=19353, LogicRNCID=501',
'Label=UH0988_U1, CellID=9881, LogicRNCID=501',
'Label=UH0962_U5, CellID=9625, LogicRNCID=501',
'Label=UH0760_U95, CellID=50072, LogicRNCID=501',
'Label=UH1704_U6, CellID=17046, LogicRNCID=501',
'Label=UH0962_U2, CellID=9622, LogicRNCID=501',
'Label=UH2984_U5, CellID=29845, LogicRNCID=501',
'Label=UH1960_U5, CellID=19605, LogicRNCID=501',
'Label=UH1704_U5, CellID=17045, LogicRNCID=501',
'Label=UH0962_U1, CellID=9621, LogicRNCID=501',
'Label=UH0706_U1, CellID=7061, LogicRNCID=501',
'Label=UH0955_U96, CellID=49556, LogicRNCID=501',
'Label=UH0758_U3, CellID=7583, LogicRNCID=501',
'Label=UH1801_U94, CellID=58014, LogicRNCID=501',
'Label=UH1833_U4, CellID=18334, LogicRNCID=501',
'Label=UH1705_U4, CellID=17054, LogicRNCID=501',
'Label=UH0988_U6, CellID=9886, LogicRNCID=501',
'Label=UH0758_U2, CellID=7582, LogicRNCID=501',
'Label=UH1833_U3, CellID=18333, LogicRNCID=501',
'Label=UH1705_U3, CellID=17053, LogicRNCID=501',
'Label=UH0708_U3, CellID=7083, LogicRNCID=501',
'Label=UH0708_U2, CellID=7082, LogicRNCID=501',
'Label=UH0708_U1, CellID=7081, LogicRNCID=501',
'Label=UH0752_U96, CellID=47526, LogicRNCID=501',
'Label=UH1936_U6, CellID=19366, LogicRNCID=501',
'Label=UH0752_U95, CellID=47525, LogicRNCID=501',
'Label=UH1936_U5, CellID=19365, LogicRNCID=501',
'Label=UH0752_U94, CellID=47524, LogicRNCID=501',
'Label=UH0831_U96, CellID=48316, LogicRNCID=501',
'Label=UH0703_U96, CellID=47036, LogicRNCID=501',
'Label=UH1964_U4, CellID=19644, LogicRNCID=501',
'Label=UH0831_U95, CellID=48315, LogicRNCID=501',
'Label=UH0703_U95, CellID=47035, LogicRNCID=501',
'Label=UH0831_U94, CellID=48314, LogicRNCID=501',
'Label=UH0703_U94, CellID=47034, LogicRNCID=501',
'Label=UH1964_U2, CellID=19642, LogicRNCID=501',
'Label=UH3925_U5, CellID=39255, LogicRNCID=501',
'Label=UH0981_U5, CellID=9815, LogicRNCID=501',
'Label=UH0623_U1, CellID=6231, LogicRNCID=501',
'Label=UH0821_U94, CellID=48214, LogicRNCID=501',
'Label=UH0744_U96, CellID=47446, LogicRNCID=501',
'Label=UH3925_U4, CellID=39254, LogicRNCID=501',
'Label=UH1928_U6, CellID=19286, LogicRNCID=501',
'Label=UH0981_U4, CellID=9814, LogicRNCID=501',
'Label=UH1952_U1, CellID=19521, LogicRNCID=501',
'Label=UH1602_U95, CellID=51008, LogicRNCID=501',
'Label=UH0742_U94, CellID=47424, LogicRNCID=501',
'Label=UH0614_U94, CellID=46144, LogicRNCID=501',
'Label=UH2950_U4, CellID=29504, LogicRNCID=501',
'Label=UH0825_U6, CellID=8256, LogicRNCID=501',
'Label=UH1602_U94, CellID=51007, LogicRNCID=501',
'Label=UH2950_U3, CellID=29503, LogicRNCID=501',
'Label=UH0742_U96, CellID=47426, LogicRNCID=501',
'Label=UH0614_U96, CellID=46146, LogicRNCID=501',
'Label=UH2950_U6, CellID=29506, LogicRNCID=501',
'Label=UH1952_U2, CellID=19522, LogicRNCID=501',
'Label=UH1602_U96, CellID=51009, LogicRNCID=501',
'Label=UH0742_U95, CellID=47425, LogicRNCID=501',
'Label=UH0614_U95, CellID=46145, LogicRNCID=501',
'Label=UH2950_U5, CellID=29505, LogicRNCID=501',
'Label=UH1925_U95, CellID=50047, LogicRNCID=501',
'Label=UH0825_U95, CellID=48255, LogicRNCID=501',
'Label=UH1958_U3, CellID=19583, LogicRNCID=501',
'Label=UH1907_U1, CellID=19071, LogicRNCID=501',
'Label=UH1702_U3, CellID=17023, LogicRNCID=501',
'Label=UH0601_U5, CellID=6015, LogicRNCID=501',
'Label=UH1925_U96, CellID=50046, LogicRNCID=501',
'Label=UH0825_U94, CellID=48254, LogicRNCID=501',
'Label=UH1952_U6, CellID=19526, LogicRNCID=501',
'Label=UH1952_U5, CellID=19525, LogicRNCID=501',
'Label=UH0646_U96, CellID=51012, LogicRNCID=501',
'Label=UH1952_U4, CellID=19524, LogicRNCID=501',
'Label=UH0646_U95, CellID=51011, LogicRNCID=501',
'Label=UH1952_U3, CellID=19523, LogicRNCID=501',
'Label=UH0646_U94, CellID=51010, LogicRNCID=501',
'Label=UH1902_U95, CellID=50035, LogicRNCID=501',
'Label=UH0747_U95, CellID=47475, LogicRNCID=501',
'Label=UH3928_U3, CellID=39283, LogicRNCID=501',
'Label=UH2981_U1, CellID=29811, LogicRNCID=501',
'Label=UH1803_U5, CellID=18035, LogicRNCID=501',
'Label=UH1902_U94, CellID=50034, LogicRNCID=501',
'Label=UH0747_U94, CellID=47474, LogicRNCID=501',
'Label=UH3928_U2, CellID=39282, LogicRNCID=501',
'Label=UH0973_U95, CellID=49735, LogicRNCID=501',
'Label=UH0717_U95, CellID=47175, LogicRNCID=501',
'Label=UH0973_U94, CellID=49734, LogicRNCID=501',
'Label=UH0717_U94, CellID=47174, LogicRNCID=501',
'Label=UH1702_U5, CellID=17025, LogicRNCID=501',
'Label=UH0960_U1, CellID=9601, LogicRNCID=501',
'Label=UH0832_U1, CellID=8321, LogicRNCID=501',
'Label=UH1925_U94, CellID=50048, LogicRNCID=501',
'Label=UH0825_U96, CellID=48256, LogicRNCID=501',
'Label=UH1907_U2, CellID=19072, LogicRNCID=501',
'Label=UH1702_U4, CellID=17024, LogicRNCID=501',
'Label=UH0601_U6, CellID=6016, LogicRNCID=501',
'Label=UH1905_U4, CellID=19054, LogicRNCID=501',
'Label=UH1854_U2, CellID=18542, LogicRNCID=501',
'Label=UH0881_U4, CellID=8814, LogicRNCID=501',
'Label=UH0830_U2, CellID=8302, LogicRNCID=501',
'Label=UH1854_U1, CellID=18541, LogicRNCID=501',
'Label=UH0881_U3, CellID=8813, LogicRNCID=501',
'Label=UH0830_U1, CellID=8301, LogicRNCID=501',
'Label=UH0951_U96, CellID=49516, LogicRNCID=501',
'Label=UH0618_U94, CellID=46184, LogicRNCID=501',
'Label=UH1804_U95, CellID=50023, LogicRNCID=501',
'Label=UH1804_U94, CellID=50022, LogicRNCID=501',
'Label=UH0880_U6, CellID=8806, LogicRNCID=501',
'Label=UH1803_U96, CellID=50021, LogicRNCID=501',
'Label=UH1803_U95, CellID=50020, LogicRNCID=501',
'Label=UH1603_U96, CellID=50012, LogicRNCID=501',
'Label=UH0879_U6, CellID=8796, LogicRNCID=501',
'Label=UH0623_U6, CellID=6236, LogicRNCID=501',
'Label=UH1964_U94, CellID=55387, LogicRNCID=501',
'Label=UH1603_U95, CellID=50011, LogicRNCID=501',
'Label=UH0821_U99, CellID=48219, LogicRNCID=501',
'Label=UH1801_U1, CellID=18011, LogicRNCID=501',
'Label=UH0879_U5, CellID=8795, LogicRNCID=501',
'Label=UH2981_U3, CellID=29813, LogicRNCID=501',
'Label=UH1906_U1, CellID=19061, LogicRNCID=501',
'Label=UH0600_U5, CellID=6005, LogicRNCID=501',
'Label=UH1902_U96, CellID=50036, LogicRNCID=501',
'Label=UH3928_U4, CellID=39284, LogicRNCID=501',
'Label=UH2981_U2, CellID=29812, LogicRNCID=501',
'Label=UH1803_U6, CellID=18036, LogicRNCID=501',
'Label=UH0600_U4, CellID=6004, LogicRNCID=501',
'Label=UH0980_U2, CellID=9802, LogicRNCID=501',
'Label=UH0980_U1, CellID=9801, LogicRNCID=501',
'Label=UH0973_U96, CellID=49736, LogicRNCID=501',
'Label=UH0717_U96, CellID=47176, LogicRNCID=501',
'Label=UH0822_U94, CellID=48224, LogicRNCID=501',
'Label=UH0745_U96, CellID=47456, LogicRNCID=501',
'Label=UH1801_U6, CellID=18016, LogicRNCID=501',
'Label=UH0854_U4, CellID=8544, LogicRNCID=501',
'Label=UH1604_U96, CellID=50015, LogicRNCID=501',
'Label=UH0745_U95, CellID=47455, LogicRNCID=501',
'Label=UH0854_U3, CellID=8543, LogicRNCID=501',
'Label=UH1604_U95, CellID=50014, LogicRNCID=501',
'Label=UH0955_U6, CellID=9556, LogicRNCID=501',
'Label=UH3925_U1, CellID=39251, LogicRNCID=501',
'Label=UH1928_U3, CellID=19283, LogicRNCID=501',
'Label=UH0981_U1, CellID=9811, LogicRNCID=501',
'Label=UH0955_U5, CellID=9555, LogicRNCID=501',
'Label=UH0846_U96, CellID=48466, LogicRNCID=501',
'Label=UH1928_U2, CellID=19282, LogicRNCID=501',
'Label=UH0955_U4, CellID=9554, LogicRNCID=501',
'Label=UH1882_U94, CellID=50031, LogicRNCID=501',
'Label=UH0849_U95, CellID=48495, LogicRNCID=501',
'Label=UH1905_U5, CellID=19055, LogicRNCID=501',
'Label=UH1854_U3, CellID=18543, LogicRNCID=501',
'Label=UH1803_U1, CellID=18031, LogicRNCID=501',
'Label=UH0881_U5, CellID=8815, LogicRNCID=501',
'Label=UH0830_U3, CellID=8303, LogicRNCID=501',
'Label=UH0849_U94, CellID=48494, LogicRNCID=501',
'Label=UH1906_U3, CellID=19063, LogicRNCID=501',
'Label=UH0831_U1, CellID=8311, LogicRNCID=501',
'Label=UH1907_U95, CellID=50038, LogicRNCID=501',
'Label=UH3928_U6, CellID=39286, LogicRNCID=501',
'Label=UH2981_U4, CellID=29814, LogicRNCID=501',
'Label=UH1906_U2, CellID=19062, LogicRNCID=501',
'Label=UH1907_U94, CellID=50037, LogicRNCID=501',
'Label=UH3928_U5, CellID=39285, LogicRNCID=501',
'Label=UH0955_U2, CellID=9552, LogicRNCID=501',
'Label=UH0622_U4, CellID=6224, LogicRNCID=501',
'Label=UH0820_U97, CellID=48207, LogicRNCID=501',
'Label=UH0955_U1, CellID=9551, LogicRNCID=501',
'Label=UH0622_U3, CellID=6223, LogicRNCID=501',
'Label=UH0820_U96, CellID=48206, LogicRNCID=501',
'Label=UH0980_U6, CellID=9806, LogicRNCID=501',
'Label=UH0622_U2, CellID=6222, LogicRNCID=501',
'Label=UH0820_U95, CellID=48205, LogicRNCID=501',
'Label=UH0980_U5, CellID=9805, LogicRNCID=501',
'Label=UH0622_U1, CellID=6221, LogicRNCID=501',
'Label=UH0820_U94, CellID=48204, LogicRNCID=501',
'Label=UH0743_U96, CellID=47436, LogicRNCID=501',
'Label=UH0980_U4, CellID=9804, LogicRNCID=501',
'Label=UH0743_U95, CellID=47435, LogicRNCID=501',
'Label=UH0980_U3, CellID=9803, LogicRNCID=501',
'Label=UH0951_U94, CellID=49514, LogicRNCID=501',
'Label=UH0823_U94, CellID=48234, LogicRNCID=501',
'Label=UH0746_U96, CellID=47466, LogicRNCID=501',
'Label=UH0618_U96, CellID=46186, LogicRNCID=501',
'Label=UH1805_U94, CellID=50025, LogicRNCID=501',
'Label=UH0746_U95, CellID=47465, LogicRNCID=501',
'Label=UH0618_U95, CellID=46185, LogicRNCID=501',
'Label=UH1804_U96, CellID=50024, LogicRNCID=501',
'Label=UH0623_U5, CellID=6235, LogicRNCID=501',
'Label=UH1603_U94, CellID=50010, LogicRNCID=501',
'Label=UH0821_U98, CellID=48218, LogicRNCID=501',
'Label=UH0770_U96, CellID=47706, LogicRNCID=501',
'Label=UH0623_U4, CellID=6234, LogicRNCID=501',
'Label=UH0821_U97, CellID=48217, LogicRNCID=501',
'Label=UH0770_U95, CellID=47705, LogicRNCID=501',
'Label=UH0879_U3, CellID=8793, LogicRNCID=501',
'Label=UH1906_U5, CellID=19065, LogicRNCID=501',
'Label=UH0831_U3, CellID=8313, LogicRNCID=501',
'Label=UH1913_U94, CellID=50040, LogicRNCID=501',
'Label=UH2981_U6, CellID=29816, LogicRNCID=501',
'Label=UH1906_U4, CellID=19064, LogicRNCID=501',
'Label=UH0831_U2, CellID=8312, LogicRNCID=501',
'Label=UH1907_U96, CellID=50039, LogicRNCID=501',
'Label=UH2981_U5, CellID=29815, LogicRNCID=501',
'Label=UH0622_U6, CellID=6226, LogicRNCID=501',
'Label=UH0846_U95, CellID=48465, LogicRNCID=501',
'Label=UH0820_U99, CellID=48209, LogicRNCID=501',
'Label=UH1928_U1, CellID=19281, LogicRNCID=501',
'Label=UH0955_U3, CellID=9553, LogicRNCID=501',
'Label=UH0622_U5, CellID=6225, LogicRNCID=501',
'Label=UH0846_U94, CellID=48464, LogicRNCID=501',
'Label=UH0820_U98, CellID=48208, LogicRNCID=501',
'Label=UH0830_U5, CellID=8305, LogicRNCID=501',
'Label=UH0600_U1, CellID=6001, LogicRNCID=501',
'Label=UH1882_U95, CellID=50032, LogicRNCID=501',
'Label=UH0849_U96, CellID=48496, LogicRNCID=501',
'Label=UH1854_U4, CellID=18544, LogicRNCID=501',
'Label=UH1803_U2, CellID=18032, LogicRNCID=501',
'Label=UH0881_U6, CellID=8816, LogicRNCID=501',
'Label=UH0830_U4, CellID=8304, LogicRNCID=501',
'Label=UH0744_U95, CellID=47445, LogicRNCID=501',
'Label=UH3925_U3, CellID=39253, LogicRNCID=501',
'Label=UH1928_U5, CellID=19285, LogicRNCID=501',
'Label=UH0981_U3, CellID=9813, LogicRNCID=501',
'Label=UH0744_U94, CellID=47444, LogicRNCID=501',
'Label=UH3925_U2, CellID=39252, LogicRNCID=501',
'Label=UH1928_U4, CellID=19284, LogicRNCID=501',
'Label=UH0981_U2, CellID=9812, LogicRNCID=501',
'Label=UH0623_U3, CellID=6233, LogicRNCID=501',
'Label=UH0821_U96, CellID=48216, LogicRNCID=501',
'Label=UH0770_U94, CellID=47704, LogicRNCID=501',
'Label=UH3925_U6, CellID=39256, LogicRNCID=501',
'Label=UH0981_U6, CellID=9816, LogicRNCID=501',
'Label=UH0879_U2, CellID=8792, LogicRNCID=501',
'Label=UH0623_U2, CellID=6232, LogicRNCID=501',
'Label=UH0821_U95, CellID=48215, LogicRNCID=501',
'Label=UH0831_U6, CellID=8316, LogicRNCID=501',
'Label=UH0601_U2, CellID=6012, LogicRNCID=501',
'Label=UH0831_U5, CellID=8315, LogicRNCID=501',
'Label=UH0601_U1, CellID=6011, LogicRNCID=501',
'Label=UH1913_U96, CellID=50042, LogicRNCID=501',
'Label=UH1906_U6, CellID=19066, LogicRNCID=501',
'Label=UH0831_U4, CellID=8314, LogicRNCID=501',
'Label=UH1913_U95, CellID=50041, LogicRNCID=501',
'Label=UH0880_U4, CellID=8804, LogicRNCID=501',
'Label=UH1803_U94, CellID=50019, LogicRNCID=501',
'Label=UH0880_U3, CellID=8803, LogicRNCID=501',
'Label=UH0822_U96, CellID=48226, LogicRNCID=501',
'Label=UH0854_U6, CellID=8546, LogicRNCID=501',
'Label=UH0822_U95, CellID=48225, LogicRNCID=501',
'Label=UH0880_U1, CellID=8801, LogicRNCID=501',
'Label=UH0854_U5, CellID=8545, LogicRNCID=501',
'Label=UH0745_U94, CellID=47454, LogicRNCID=501',
'Label=UH1801_U4, CellID=18014, LogicRNCID=501',
'Label=UH0854_U2, CellID=8542, LogicRNCID=501',
'Label=UH1964_U96, CellID=55389, LogicRNCID=501',
'Label=UH1604_U94, CellID=50013, LogicRNCID=501',
'Label=UH1801_U3, CellID=18013, LogicRNCID=501',
'Label=UH0854_U1, CellID=8541, LogicRNCID=501',
'Label=UH1964_U95, CellID=55388, LogicRNCID=501',
'Label=UH0823_U96, CellID=48236, LogicRNCID=501',
'Label=UH1905_U2, CellID=19052, LogicRNCID=501',
'Label=UH0881_U2, CellID=8812, LogicRNCID=501',
'Label=UH1805_U96, CellID=50027, LogicRNCID=501',
'Label=UH0823_U95, CellID=48235, LogicRNCID=501',
'Label=UH1905_U1, CellID=19051, LogicRNCID=501',
'Label=UH0881_U1, CellID=8811, LogicRNCID=501',
'Label=UH1805_U95, CellID=50026, LogicRNCID=501',
'Label=UH1854_U6, CellID=18546, LogicRNCID=501',
'Label=UH1803_U4, CellID=18034, LogicRNCID=501',
'Label=UH0830_U6, CellID=8306, LogicRNCID=501',
'Label=UH0600_U2, CellID=6002, LogicRNCID=501',
'Label=UH1882_U96, CellID=50033, LogicRNCID=501',
'Label=UH3928_U1, CellID=39281, LogicRNCID=501',
'Label=UH1854_U5, CellID=18545, LogicRNCID=501',
'Label=UH1803_U3, CellID=18033, LogicRNCID=501',
'Label=UH0748_U96, CellID=47486, LogicRNCID=501',
'Label=UH1958_U2, CellID=19582, LogicRNCID=501',
'Label=UH1702_U2, CellID=17022, LogicRNCID=501',
'Label=UH0601_U4, CellID=6014, LogicRNCID=501',
'Label=UH0748_U95, CellID=47485, LogicRNCID=501',
'Label=UH1702_U1, CellID=17021, LogicRNCID=501',
'Label=UH0601_U3, CellID=6013, LogicRNCID=501',
'Label=UH0748_U94, CellID=47484, LogicRNCID=501',
'Label=UH0974_U5, CellID=9745, LogicRNCID=501',
'Label=UH0737_U96, CellID=47376, LogicRNCID=501',
'Label=UH3995_U2, CellID=39952, LogicRNCID=501',
'Label=UH3918_U4, CellID=39184, LogicRNCID=501',
'Label=UH1947_U2, CellID=19472, LogicRNCID=501',
'Label=UH0974_U4, CellID=9744, LogicRNCID=501',
'Label=UH0846_U4, CellID=8464, LogicRNCID=501',
'Label=UH0839_U99, CellID=48399, LogicRNCID=501',
]
cluster_UH_noUH0845_top = [
'Label=UH1945_U5, CellID=19455, LogicRNCID=501',
'Label=UH0761_U94, CellID=47614, LogicRNCID=501',
'Label=UH1945_U4, CellID=19454, LogicRNCID=501',
'Label=UH0972_U6, CellID=9726, LogicRNCID=501',
'Label=UH0844_U6, CellID=8446, LogicRNCID=501',
'Label=UH0742_U2, CellID=7422, LogicRNCID=501',
'Label=UH1945_U3, CellID=19453, LogicRNCID=501',
'Label=UH0972_U5, CellID=9725, LogicRNCID=501',
'Label=UH0844_U5, CellID=8445, LogicRNCID=501',
'Label=UH0742_U1, CellID=7421, LogicRNCID=501',
'Label=UH0735_U96, CellID=47356, LogicRNCID=501',
'Label=UH1945_U2, CellID=19452, LogicRNCID=501',
'Label=UH0972_U4, CellID=9724, LogicRNCID=501',
'Label=UH0844_U4, CellID=8444, LogicRNCID=501',
'Label=UH0735_U95, CellID=47355, LogicRNCID=501',
'Label=UH1945_U1, CellID=19451, LogicRNCID=501',
'Label=UH0972_U3, CellID=9723, LogicRNCID=501',
'Label=UH0844_U3, CellID=8443, LogicRNCID=501',
'Label=UH0735_U94, CellID=47354, LogicRNCID=501',
'Label=UH0972_U2, CellID=9722, LogicRNCID=501',
'Label=UH0844_U2, CellID=8442, LogicRNCID=501',
'Label=UH0965_U97, CellID=49657, LogicRNCID=501',
'Label=UH0972_U1, CellID=9721, LogicRNCID=501',
'Label=UH0844_U1, CellID=8441, LogicRNCID=501',
'Label=UH0965_U96, CellID=49656, LogicRNCID=501',
'Label=UH0965_U95, CellID=49655, LogicRNCID=501',
'Label=UH0965_U94, CellID=49654, LogicRNCID=501',
'Label=UH2763_U6, CellID=27636, LogicRNCID=501',
'Label=UH2763_U5, CellID=27635, LogicRNCID=501',
'Label=UH0734_U96, CellID=47346, LogicRNCID=501',
'Label=UH2763_U4, CellID=27634, LogicRNCID=501',
'Label=UH1918_U6, CellID=19186, LogicRNCID=501',
'Label=UH0766_U6, CellID=7666, LogicRNCID=501',
'Label=UH0880_U14, CellID=65521, LogicRNCID=501',
'Label=UH0734_U95, CellID=47345, LogicRNCID=501',
'Label=UH2763_U3, CellID=27633, LogicRNCID=501',
'Label=UH1918_U5, CellID=19185, LogicRNCID=501',
'Label=UH0766_U5, CellID=7665, LogicRNCID=501',
'Label=UH0880_U11, CellID=65520, LogicRNCID=501',
'Label=UH0734_U94, CellID=47344, LogicRNCID=501',
'Label=UH2763_U2, CellID=27632, LogicRNCID=501',
'Label=UH1918_U4, CellID=19184, LogicRNCID=501',
'Label=UH0766_U4, CellID=7664, LogicRNCID=501',
'Label=UH2763_U1, CellID=27631, LogicRNCID=501',
'Label=UH1918_U3, CellID=19183, LogicRNCID=501',
'Label=UH0766_U3, CellID=7663, LogicRNCID=501',
'Label=UH0708_U96, CellID=47086, LogicRNCID=501',
'Label=UH1918_U2, CellID=19182, LogicRNCID=501',
'Label=UH1892_U6, CellID=18926, LogicRNCID=501',
'Label=UH0766_U2, CellID=7662, LogicRNCID=501',
'Label=UH0708_U95, CellID=47085, LogicRNCID=501',
'Label=UH1918_U1, CellID=19181, LogicRNCID=501',
'Label=UH1892_U5, CellID=18925, LogicRNCID=501',
'Label=UH0766_U1, CellID=7661, LogicRNCID=501',
'Label=UH0759_U96, CellID=47596, LogicRNCID=501',
'Label=UH0708_U94, CellID=47084, LogicRNCID=501',
'Label=UH0759_U95, CellID=47595, LogicRNCID=501',
'Label=UH1892_U3, CellID=18923, LogicRNCID=501',
'Label=UH1892_U2, CellID=18922, LogicRNCID=501',
'Label=UH0970_U6, CellID=9706, LogicRNCID=501',
'Label=UH0842_U6, CellID=8426, LogicRNCID=501',
'Label=UH0714_U6, CellID=7146, LogicRNCID=501',
'Label=UH0970_U5, CellID=9705, LogicRNCID=501',
'Label=UH0842_U5, CellID=8425, LogicRNCID=501',
'Label=UH0714_U5, CellID=7145, LogicRNCID=501',
'Label=UH0970_U4, CellID=9704, LogicRNCID=501',
'Label=UH0842_U4, CellID=8424, LogicRNCID=501',
'Label=UH0765_U6, CellID=7656, LogicRNCID=501',
'Label=UH0714_U4, CellID=7144, LogicRNCID=501',
'Label=UH0970_U3, CellID=9703, LogicRNCID=501',
'Label=UH0842_U3, CellID=8423, LogicRNCID=501',
'Label=UH0765_U5, CellID=7655, LogicRNCID=501',
'Label=UH0714_U3, CellID=7143, LogicRNCID=501',
'Label=UH1936_U96, CellID=59366, LogicRNCID=501',
'Label=UH0970_U2, CellID=9702, LogicRNCID=501',
'Label=UH0842_U2, CellID=8422, LogicRNCID=501',
'Label=UH0765_U4, CellID=7654, LogicRNCID=501',
'Label=UH0714_U2, CellID=7142, LogicRNCID=501',
'Label=UH1936_U95, CellID=59365, LogicRNCID=501',
'Label=UH0970_U1, CellID=9701, LogicRNCID=501',
'Label=UH0842_U1, CellID=8421, LogicRNCID=501',
'Label=UH0765_U3, CellID=7653, LogicRNCID=501',
'Label=UH0714_U1, CellID=7141, LogicRNCID=501',
'Label=UH0963_U96, CellID=49636, LogicRNCID=501',
'Label=UH0835_U96, CellID=48356, LogicRNCID=501',
'Label=UH0995_U6, CellID=9956, LogicRNCID=501',
'Label=UH0765_U2, CellID=7652, LogicRNCID=501',
'Label=UH0611_U6, CellID=6116, LogicRNCID=501',
'Label=UH0963_U95, CellID=49635, LogicRNCID=501',
'Label=UH0835_U95, CellID=48355, LogicRNCID=501',
'Label=UH0995_U5, CellID=9955, LogicRNCID=501',
'Label=UH0765_U1, CellID=7651, LogicRNCID=501',
'Label=UH0963_U94, CellID=49634, LogicRNCID=501',
'Label=UH0835_U94, CellID=48354, LogicRNCID=501',
'Label=UH0758_U96, CellID=47586, LogicRNCID=501',
'Label=UH3990_U6, CellID=39906, LogicRNCID=501',
'Label=UH0995_U4, CellID=9954, LogicRNCID=501',
'Label=UH0611_U4, CellID=6114, LogicRNCID=501',
'Label=UH0758_U95, CellID=47585, LogicRNCID=501',
'Label=UH3990_U5, CellID=39905, LogicRNCID=501',
'Label=UH0995_U3, CellID=9953, LogicRNCID=501',
'Label=UH0611_U3, CellID=6113, LogicRNCID=501',
'Label=UH0758_U94, CellID=47584, LogicRNCID=501',
'Label=UH3990_U4, CellID=39904, LogicRNCID=501',
'Label=UH0995_U2, CellID=9952, LogicRNCID=501',
'Label=UH0841_U6, CellID=8416, LogicRNCID=501',
'Label=UH3990_U3, CellID=39903, LogicRNCID=501',
'Label=UH0995_U1, CellID=9951, LogicRNCID=501',
'Label=UH0841_U5, CellID=8415, LogicRNCID=501',
'Label=UH0611_U1, CellID=6111, LogicRNCID=501',
'Label=UH0732_U96, CellID=47326, LogicRNCID=501',
'Label=UH3990_U2, CellID=39902, LogicRNCID=501',
'Label=UH1916_U6, CellID=19166, LogicRNCID=501',
'Label=UH0841_U4, CellID=8414, LogicRNCID=501',
'Label=UH0636_U6, CellID=6366, LogicRNCID=501',
'Label=UH0508_U6, CellID=5086, LogicRNCID=501',
'Label=UH0732_U95, CellID=47325, LogicRNCID=501',
'Label=UH3990_U1, CellID=39901, LogicRNCID=501',
'Label=UH1916_U5, CellID=19165, LogicRNCID=501',
'Label=UH0841_U3, CellID=8413, LogicRNCID=501',
'Label=UH0636_U5, CellID=6365, LogicRNCID=501',
'Label=UH0508_U5, CellID=5085, LogicRNCID=501',
'Label=UH0732_U94, CellID=47324, LogicRNCID=501',
'Label=UH1916_U4, CellID=19164, LogicRNCID=501',
'Label=UH0841_U2, CellID=8412, LogicRNCID=501',
'Label=UH1916_U3, CellID=19163, LogicRNCID=501',
'Label=UH0841_U1, CellID=8411, LogicRNCID=501',
'Label=UH0636_U3, CellID=6363, LogicRNCID=501',
'Label=UH0508_U3, CellID=5083, LogicRNCID=501',
'Label=UH0962_U96, CellID=49626, LogicRNCID=501',
'Label=UH0706_U96, CellID=47066, LogicRNCID=501',
'Label=UH1916_U2, CellID=19162, LogicRNCID=501',
'Label=UH0994_U6, CellID=9946, LogicRNCID=501',
'Label=UH0738_U6, CellID=7386, LogicRNCID=501',
'Label=UH0636_U2, CellID=6362, LogicRNCID=501',
'Label=UH0610_U6, CellID=6106, LogicRNCID=501',
'Label=UH0508_U2, CellID=5082, LogicRNCID=501',
'Label=UH1907_U66, CellID=65497, LogicRNCID=501',
'Label=UH0962_U95, CellID=49625, LogicRNCID=501',
'Label=UH0706_U95, CellID=47065, LogicRNCID=501',
'Label=UH1916_U1, CellID=19161, LogicRNCID=501',
'Label=UH0994_U5, CellID=9945, LogicRNCID=501',
'Label=UH0738_U5, CellID=7385, LogicRNCID=501',
'Label=UH0610_U5, CellID=6105, LogicRNCID=501',
'Label=UH1907_U55, CellID=65496, LogicRNCID=501',
'Label=UH0962_U94, CellID=49624, LogicRNCID=501',
'Label=UH0706_U94, CellID=47064, LogicRNCID=501',
'Label=UH3989_U6, CellID=39896, LogicRNCID=501',
'Label=UH2709_U6, CellID=27096, LogicRNCID=501',
'Label=UH0994_U4, CellID=9944, LogicRNCID=501',
'Label=UH0738_U4, CellID=7384, LogicRNCID=501',
'Label=UH1907_U44, CellID=65495, LogicRNCID=501',
'Label=UH3989_U5, CellID=39895, LogicRNCID=501',
'Label=UH2709_U5, CellID=27095, LogicRNCID=501',
'Label=UH0994_U3, CellID=9943, LogicRNCID=501',
'Label=UH0738_U3, CellID=7383, LogicRNCID=501',
'Label=UH0610_U3, CellID=6103, LogicRNCID=501',
'Label=UH1907_U33, CellID=65494, LogicRNCID=501',
'Label=UH3989_U4, CellID=39894, LogicRNCID=501',
'Label=UH2709_U4, CellID=27094, LogicRNCID=501',
'Label=UH0994_U2, CellID=9942, LogicRNCID=501',
'Label=UH0840_U6, CellID=8406, LogicRNCID=501',
'Label=UH0738_U2, CellID=7382, LogicRNCID=501',
'Label=UH0610_U2, CellID=6102, LogicRNCID=501',
'Label=UH1907_U22, CellID=65493, LogicRNCID=501',
'Label=UH3989_U3, CellID=39893, LogicRNCID=501',
'Label=UH2709_U3, CellID=27093, LogicRNCID=501',
'Label=UH0994_U1, CellID=9941, LogicRNCID=501',
'Label=UH0840_U5, CellID=8405, LogicRNCID=501',
'Label=UH0738_U1, CellID=7381, LogicRNCID=501',
'Label=UH1907_U11, CellID=65492, LogicRNCID=501',
'Label=UH0731_U96, CellID=47316, LogicRNCID=501',
'Label=UH3989_U2, CellID=39892, LogicRNCID=501',
'Label=UH2709_U2, CellID=27092, LogicRNCID=501',
'Label=UH0840_U4, CellID=8404, LogicRNCID=501',
'Label=UH0763_U6, CellID=7636, LogicRNCID=501',
'Label=UH0731_U95, CellID=47315, LogicRNCID=501',
'Label=UH3989_U1, CellID=39891, LogicRNCID=501',
'Label=UH2709_U1, CellID=27091, LogicRNCID=501',
'Label=UH0840_U3, CellID=8403, LogicRNCID=501',
'Label=UH0763_U5, CellID=7635, LogicRNCID=501',
'Label=UH0731_U94, CellID=47314, LogicRNCID=501',
'Label=UH0840_U2, CellID=8402, LogicRNCID=501',
'Label=UH0763_U4, CellID=7634, LogicRNCID=501',
'Label=UH0840_U1, CellID=8401, LogicRNCID=501',
'Label=UH0763_U3, CellID=7633, LogicRNCID=501',
'Label=UH0833_U96, CellID=48336, LogicRNCID=501',
'Label=UH0705_U96, CellID=47056, LogicRNCID=501',
'Label=UH2913_U6, CellID=29136, LogicRNCID=501',
'Label=UH0763_U2, CellID=7632, LogicRNCID=501',
'Label=UH0833_U95, CellID=48335, LogicRNCID=501',
'Label=UH0705_U95, CellID=47055, LogicRNCID=501',
'Label=UH2913_U5, CellID=29135, LogicRNCID=501',
'Label=UH0763_U1, CellID=7631, LogicRNCID=501',
'Label=UH0833_U94, CellID=48334, LogicRNCID=501',
'Label=UH0756_U96, CellID=47566, LogicRNCID=501',
'Label=UH0705_U94, CellID=47054, LogicRNCID=501',
'Label=UH2913_U4, CellID=29134, LogicRNCID=501',
'Label=UH2913_U3, CellID=29133, LogicRNCID=501',
'Label=UH0807_U96, CellID=48076, LogicRNCID=501',
'Label=UH0756_U94, CellID=47564, LogicRNCID=501',
'Label=UH2913_U2, CellID=29132, LogicRNCID=501',
'Label=UH0839_U6, CellID=8396, LogicRNCID=501',
'Label=UH0711_U6, CellID=7116, LogicRNCID=501',
'Label=UH0807_U95, CellID=48075, LogicRNCID=501',
'Label=UH2913_U1, CellID=29131, LogicRNCID=501',
'Label=UH0839_U5, CellID=8395, LogicRNCID=501',
'Label=UH0711_U5, CellID=7115, LogicRNCID=501',
'Label=UH0807_U94, CellID=48074, LogicRNCID=501',
'Label=UH0602_U96, CellID=46026, LogicRNCID=501',
'Label=UH2938_U6, CellID=29386, LogicRNCID=501',
'Label=UH0839_U4, CellID=8394, LogicRNCID=501',
'Label=UH0711_U4, CellID=7114, LogicRNCID=501',
'Label=UH0602_U95, CellID=46025, LogicRNCID=501',
'Label=UH2938_U5, CellID=29385, LogicRNCID=501',
'Label=UH0839_U3, CellID=8393, LogicRNCID=501',
'Label=UH0711_U3, CellID=7113, LogicRNCID=501',
'Label=UH0602_U94, CellID=46024, LogicRNCID=501',
'Label=UH2938_U4, CellID=29384, LogicRNCID=501',
'Label=UH0839_U2, CellID=8392, LogicRNCID=501',
'Label=UH0711_U2, CellID=7112, LogicRNCID=501',
'Label=UH2938_U3, CellID=29383, LogicRNCID=501',
'Label=UH0839_U1, CellID=8391, LogicRNCID=501',
'Label=UH0711_U1, CellID=7111, LogicRNCID=501',
'Label=UH0960_U96, CellID=49606, LogicRNCID=501',
'Label=UH0832_U96, CellID=48326, LogicRNCID=501',
'Label=UH0704_U96, CellID=47046, LogicRNCID=501',
'Label=UH2938_U2, CellID=29382, LogicRNCID=501',
'Label=UH0960_U95, CellID=49605, LogicRNCID=501',
'Label=UH0832_U95, CellID=48325, LogicRNCID=501',
'Label=UH0704_U95, CellID=47045, LogicRNCID=501',
'Label=UH2938_U1, CellID=29381, LogicRNCID=501',
'Label=UH0960_U94, CellID=49604, LogicRNCID=501',
'Label=UH0832_U94, CellID=48324, LogicRNCID=501',
'Label=UH1939_U6, CellID=19396, LogicRNCID=501',
'Label=UH1939_U5, CellID=19395, LogicRNCID=501',
'Label=UH1939_U4, CellID=19394, LogicRNCID=501',
'Label=UH0966_U6, CellID=9666, LogicRNCID=501',
'Label=UH0838_U6, CellID=8386, LogicRNCID=501',
'Label=UH1939_U3, CellID=19393, LogicRNCID=501',
'Label=UH0961_U6, CellID=9616, LogicRNCID=501',
'Label=UH1951_U95, CellID=50063, LogicRNCID=501',
'Label=UH0961_U5, CellID=9615, LogicRNCID=501',
'Label=UH1951_U94, CellID=50062, LogicRNCID=501',
'Label=UH2932_U6, CellID=29326, LogicRNCID=501',
'Label=UH1908_U6, CellID=19086, LogicRNCID=501',
'Label=UH0961_U4, CellID=9614, LogicRNCID=501',
'Label=UH1945_U99, CellID=50061, LogicRNCID=501',
'Label=UH1907_U4, CellID=19074, LogicRNCID=501',
'Label=UH1702_U6, CellID=17026, LogicRNCID=501',
'Label=UH0960_U2, CellID=9602, LogicRNCID=501',
'Label=UH0832_U2, CellID=8322, LogicRNCID=501',
'Label=UH1928_U95, CellID=50049, LogicRNCID=501',
'Label=UH1984_U1, CellID=19841, LogicRNCID=501',
'Label=UH1958_U5, CellID=19585, LogicRNCID=501',
'Label=UH1907_U3, CellID=19073, LogicRNCID=501',
'Label=UH1907_U5, CellID=19075, LogicRNCID=501',
'Label=UH1805_U1, CellID=18051, LogicRNCID=501',
'Label=UH0960_U3, CellID=9603, LogicRNCID=501',
'Label=UH0832_U3, CellID=8323, LogicRNCID=501',
'Label=UH1928_U96, CellID=50050, LogicRNCID=501',
'Label=UH0723_U94, CellID=47234, LogicRNCID=501',
'Label=UH1984_U2, CellID=19842, LogicRNCID=501',
'Label=UH1958_U6, CellID=19586, LogicRNCID=501',
'Label=UH0966_U3, CellID=9663, LogicRNCID=501',
'Label=UH0838_U3, CellID=8383, LogicRNCID=501',
'Label=UH0601_U94, CellID=46014, LogicRNCID=501',
'Label=UH0966_U2, CellID=9662, LogicRNCID=501',
'Label=UH0838_U2, CellID=8382, LogicRNCID=501',
'Label=UH1964_U5, CellID=19645, LogicRNCID=501',
'Label=UH0966_U1, CellID=9661, LogicRNCID=501',
'Label=UH0838_U1, CellID=8381, LogicRNCID=501',
'Label=UH0602_U1, CellID=6021, LogicRNCID=501',
'Label=UH0723_U96, CellID=47236, LogicRNCID=501',
'Label=UH1984_U4, CellID=19844, LogicRNCID=501',
'Label=UH1907_U6, CellID=19076, LogicRNCID=501',
'Label=UH1805_U2, CellID=18052, LogicRNCID=501',
'Label=UH0960_U4, CellID=9604, LogicRNCID=501',
'Label=UH0832_U4, CellID=8324, LogicRNCID=501',
'Label=UH0723_U95, CellID=47235, LogicRNCID=501',
'Label=UH1707_U1, CellID=17071, LogicRNCID=501',
'Label=UH0734_U3, CellID=7343, LogicRNCID=501',
'Label=UH0881_U94, CellID=48814, LogicRNCID=501',
'Label=UH0734_U2, CellID=7342, LogicRNCID=501',
'Label=UH0708_U6, CellID=7086, LogicRNCID=501',
'Label=UH0734_U1, CellID=7341, LogicRNCID=501',
'Label=UH0708_U5, CellID=7085, LogicRNCID=501',
'Label=UH0708_U4, CellID=7084, LogicRNCID=501',
'Label=UH0960_U6, CellID=9606, LogicRNCID=501',
'Label=UH0832_U6, CellID=8326, LogicRNCID=501',
'Label=UH0602_U2, CellID=6022, LogicRNCID=501',
'Label=UH1984_U5, CellID=19845, LogicRNCID=501',
'Label=UH1882_U1, CellID=18821, LogicRNCID=501',
'Label=UH1805_U3, CellID=18053, LogicRNCID=501',
'Label=UH0960_U5, CellID=9605, LogicRNCID=501',
'Label=UH0832_U5, CellID=8325, LogicRNCID=501',
'Label=UH0966_U5, CellID=9665, LogicRNCID=501',
'Label=UH0838_U5, CellID=8385, LogicRNCID=501',
'Label=UH0601_U96, CellID=46016, LogicRNCID=501',
'Label=UH1939_U2, CellID=19392, LogicRNCID=501',
'Label=UH0966_U4, CellID=9664, LogicRNCID=501',
'Label=UH0838_U4, CellID=8384, LogicRNCID=501',
'Label=UH0601_U95, CellID=46015, LogicRNCID=501',
'Label=UH1939_U1, CellID=19391, LogicRNCID=501',
'Label=UH1705_U6, CellID=17056, LogicRNCID=501',
'Label=UH0963_U2, CellID=9632, LogicRNCID=501',
'Label=UH0835_U2, CellID=8352, LogicRNCID=501',
'Label=UH0758_U4, CellID=7584, LogicRNCID=501',
'Label=UH1833_U5, CellID=18335, LogicRNCID=501',
'Label=UH1705_U5, CellID=17055, LogicRNCID=501',
'Label=UH0963_U1, CellID=9631, LogicRNCID=501',
'Label=UH0835_U1, CellID=8351, LogicRNCID=501',
'Label=UH0623_U95, CellID=46235, LogicRNCID=501',
'Label=UH1935_U5, CellID=19355, LogicRNCID=501',
'Label=UH1833_U1, CellID=18331, LogicRNCID=501',
'Label=UH1705_U1, CellID=17051, LogicRNCID=501',
'Label=UH0988_U3, CellID=9883, LogicRNCID=501',
'Label=UH0760_U97, CellID=50074, LogicRNCID=501',
'Label=UH0751_U94, CellID=47514, LogicRNCID=501',
'Label=UH0623_U94, CellID=46234, LogicRNCID=501',
'Label=UH2984_U4, CellID=29844, LogicRNCID=501',
'Label=UH1960_U4, CellID=19604, LogicRNCID=501',
'Label=UH1704_U4, CellID=17044, LogicRNCID=501',
'Label=UH2709_U96, CellID=50067, LogicRNCID=501',
'Label=UH0955_U95, CellID=49555, LogicRNCID=501',
'Label=UH2984_U3, CellID=29843, LogicRNCID=501',
'Label=UH1960_U3, CellID=19603, LogicRNCID=501',
'Label=UH1704_U3, CellID=17043, LogicRNCID=501',
'Label=UH1707_U3, CellID=17073, LogicRNCID=501',
'Label=UH0734_U5, CellID=7345, LogicRNCID=501',
'Label=UH0881_U96, CellID=48816, LogicRNCID=501',
'Label=UH0830_U94, CellID=48304, LogicRNCID=501',
'Label=UH0702_U94, CellID=47024, LogicRNCID=501',
'Label=UH1707_U2, CellID=17072, LogicRNCID=501',
'Label=UH0734_U4, CellID=7344, LogicRNCID=501',
'Label=UH0881_U95, CellID=48815, LogicRNCID=501',
'Label=UH2983_U1, CellID=29831, LogicRNCID=501',
'Label=UH1882_U3, CellID=18823, LogicRNCID=501',
'Label=UH1805_U5, CellID=18055, LogicRNCID=501',
'Label=UH1703_U1, CellID=17031, LogicRNCID=501',
'Label=UH0807_U1, CellID=8071, LogicRNCID=501',
'Label=UH0602_U3, CellID=6023, LogicRNCID=501',
'Label=UH1882_U2, CellID=18822, LogicRNCID=501',
'Label=UH1805_U4, CellID=18054, LogicRNCID=501',
'Label=UH1935_U2, CellID=19352, LogicRNCID=501',
'Label=UH0962_U4, CellID=9624, LogicRNCID=501',
'Label=UH0706_U4, CellID=7064, LogicRNCID=501',
'Label=UH0760_U94, CellID=50071, LogicRNCID=501',
'Label=UH1935_U1, CellID=19351, LogicRNCID=501',
'Label=UH0962_U3, CellID=9623, LogicRNCID=501',
'Label=UH2984_U6, CellID=29846, LogicRNCID=501',
'Label=UH1960_U6, CellID=19606, LogicRNCID=501',
'Label=UH0961_U2, CellID=9612, LogicRNCID=501',
'Label=UH0807_U6, CellID=8076, LogicRNCID=501',
'Label=UH2983_U5, CellID=29835, LogicRNCID=501',
'Label=UH2932_U3, CellID=29323, LogicRNCID=501',
'Label=UH1908_U3, CellID=19083, LogicRNCID=501',
'Label=UH1703_U5, CellID=17035, LogicRNCID=501',
'Label=UH0961_U1, CellID=9611, LogicRNCID=501',
'Label=UH0835_U4, CellID=8354, LogicRNCID=501',
'Label=UH0758_U6, CellID=7586, LogicRNCID=501',
'Label=UH1936_U1, CellID=19361, LogicRNCID=501',
'Label=UH0963_U3, CellID=9633, LogicRNCID=501',
'Label=UH0835_U3, CellID=8353, LogicRNCID=501',
'Label=UH0758_U5, CellID=7585, LogicRNCID=501',
'Label=UH1801_U96, CellID=58016, LogicRNCID=501',
'Label=UH1833_U6, CellID=18336, LogicRNCID=501',
'Label=UH0702_U96, CellID=47026, LogicRNCID=501',
'Label=UH4600_U2, CellID=46002, LogicRNCID=501',
'Label=UH1707_U4, CellID=17074, LogicRNCID=501',
'Label=UH0734_U6, CellID=7346, LogicRNCID=501',
'Label=UH0881_U97, CellID=48817, LogicRNCID=501',
'Label=UH0830_U95, CellID=48305, LogicRNCID=501',
'Label=UH0702_U95, CellID=47025, LogicRNCID=501',
'Label=UH4600_U1, CellID=46001, LogicRNCID=501',
'Label=UH0602_U6, CellID=6026, LogicRNCID=501',
'Label=UH1945_U95, CellID=50057, LogicRNCID=501',
'Label=UH2983_U3, CellID=29833, LogicRNCID=501',
'Label=UH2932_U1, CellID=29321, LogicRNCID=501',
'Label=UH1908_U1, CellID=19081, LogicRNCID=501',
'Label=UH1882_U5, CellID=18825, LogicRNCID=501',
'Label=UH1703_U3, CellID=17033, LogicRNCID=501',
'Label=UH0807_U3, CellID=8073, LogicRNCID=501',
'Label=UH0602_U5, CellID=6025, LogicRNCID=501',
'Label=UH2983_U2, CellID=29832, LogicRNCID=501',
'Label=UH1882_U4, CellID=18824, LogicRNCID=501',
'Label=UH0748_U5, CellID=7485, LogicRNCID=501',
'Label=UH0767_U96, CellID=47676, LogicRNCID=501',
'Label=UH0825_U2, CellID=8252, LogicRNCID=501',
'Label=UH0767_U95, CellID=47675, LogicRNCID=501',
'Label=UH0742_U4, CellID=7424, LogicRNCID=501',
'Label=UH1951_U5, CellID=19515, LogicRNCID=501',
'Label=UH0748_U4, CellID=7484, LogicRNCID=501',
'Label=UH1951_U6, CellID=19516, LogicRNCID=501',
'Label=UH0849_U6, CellID=8496, LogicRNCID=501',
'Label=UH3921_U5, CellID=39215, LogicRNCID=501',
'Label=UH0740_U96, CellID=47406, LogicRNCID=501',
'Label=UH0849_U4, CellID=8494, LogicRNCID=501',
'Label=UH0717_U1, CellID=7171, LogicRNCID=501',
'Label=UH0838_U95, CellID=48385, LogicRNCID=501',
'Label=UH0742_U5, CellID=7425, LogicRNCID=501',
'Label=UH0761_U96, CellID=47616, LogicRNCID=501',
'Label=UH0825_U5, CellID=8255, LogicRNCID=501',
'Label=UH2950_U2, CellID=29502, LogicRNCID=501',
'Label=UH0748_U6, CellID=7486, LogicRNCID=501',
'Label=UH2950_U1, CellID=29501, LogicRNCID=501',
'Label=UH0765_U94, CellID=47654, LogicRNCID=501',
'Label=UH3920_U5, CellID=39205, LogicRNCID=501',
'Label=UH0739_U96, CellID=47396, LogicRNCID=501',
'Label=UH3920_U4, CellID=39204, LogicRNCID=501',
'Label=UH0822_U3, CellID=8223, LogicRNCID=501',
'Label=UH0841_U94, CellID=48414, LogicRNCID=501',
'Label=UH0636_U96, CellID=46366, LogicRNCID=501',
'Label=UH0950_U2, CellID=9502, LogicRNCID=501',
'Label=UH0821_U4, CellID=8214, LogicRNCID=501',
'Label=UH0879_U95, CellID=55061, LogicRNCID=501',
'Label=UH0821_U3, CellID=8213, LogicRNCID=501',
'Label=UH0879_U96, CellID=55060, LogicRNCID=501',
'Label=UH0715_U94, CellID=47154, LogicRNCID=501',
'Label=UH0952_U2, CellID=9522, LogicRNCID=501',
'Label=UH1976_U1, CellID=19761, LogicRNCID=501',
'Label=UH0766_U96, CellID=47664, LogicRNCID=501',
'Label=UH0736_U94, CellID=47364, LogicRNCID=501',
'Label=UH0973_U1, CellID=9731, LogicRNCID=501',
'Label=UH0617_U2, CellID=6172, LogicRNCID=501',
'Label=UH0617_U1, CellID=6171, LogicRNCID=501',
'Label=UH3996_U2, CellID=39962, LogicRNCID=501',
'Label=UH0738_U95, CellID=47385, LogicRNCID=501',
'Label=UH0839_U97, CellID=48397, LogicRNCID=501',
'Label=UH0974_U1, CellID=9741, LogicRNCID=501',
'Label=UH0820_U5, CellID=8205, LogicRNCID=501',
'Label=UH0711_U96, CellID=47116, LogicRNCID=501',
'Label=UH0618_U5, CellID=6185, LogicRNCID=501',
'Label=UH0714_U94, CellID=47144, LogicRNCID=501',
'Label=UH0618_U4, CellID=6184, LogicRNCID=501',
'Label=UH0823_U1, CellID=8231, LogicRNCID=501',
'Label=UH0952_U4, CellID=9524, LogicRNCID=501',
'Label=UH0715_U95, CellID=47155, LogicRNCID=501',
'Label=UH0952_U3, CellID=9523, LogicRNCID=501',
'Label=UH0843_U94, CellID=48434, LogicRNCID=501',
'Label=UH0820_U1, CellID=8201, LogicRNCID=501',
'Label=UH0973_U6, CellID=9736, LogicRNCID=501',
'Label=UH0743_U2, CellID=7432, LogicRNCID=501',
'Label=UH0973_U5, CellID=9735, LogicRNCID=501',
'Label=UH0743_U1, CellID=7431, LogicRNCID=501',
'Label=UH0736_U96, CellID=47366, LogicRNCID=501',
'Label=UH0736_U95, CellID=47365, LogicRNCID=501',
'Label=UH0841_U96, CellID=48416, LogicRNCID=501',
'Label=UH0950_U4, CellID=9504, LogicRNCID=501',
'Label=UH0617_U6, CellID=6176, LogicRNCID=501',
'Label=UH0713_U95, CellID=47135, LogicRNCID=501',
'Label=UH0763_U96, CellID=47636, LogicRNCID=501',
'Label=UH1947_U6, CellID=19476, LogicRNCID=501',
'Label=UH0763_U95, CellID=47635, LogicRNCID=501',
'Label=UH1947_U5, CellID=19475, LogicRNCID=501',
'Label=UH1976_U6, CellID=19766, LogicRNCID=501',
'Label=UH3922_U1, CellID=39221, LogicRNCID=501',
'Label=UH0952_U5, CellID=9525, LogicRNCID=501',
'Label=UH0715_U96, CellID=47156, LogicRNCID=501',
'Label=UH0743_U6, CellID=7436, LogicRNCID=501',
'Label=UH0711_U95, CellID=47115, LogicRNCID=501',
'Label=UH0839_U94, CellID=48394, LogicRNCID=501',
'Label=UH0820_U2, CellID=8202, LogicRNCID=501',
'Label=UH0849_U1, CellID=8491, LogicRNCID=501',
'Label=UH0951_U4, CellID=9514, LogicRNCID=501',
'Label=UH0618_U6, CellID=6186, LogicRNCID=501',
'Label=UH0951_U3, CellID=9513, LogicRNCID=501',
'Label=UH0737_U95, CellID=47375, LogicRNCID=501',
'Label=UH3918_U3, CellID=39183, LogicRNCID=501',
'Label=UH0974_U3, CellID=9743, LogicRNCID=501',
'Label=UH0737_U94, CellID=47374, LogicRNCID=501',
'Label=UH0763_U94, CellID=47634, LogicRNCID=501',
'Label=UH3918_U6, CellID=39186, LogicRNCID=501',
'Label=UH0974_U6, CellID=9746, LogicRNCID=501',
'Label=UH3995_U3, CellID=39953, LogicRNCID=501',
'Label=UH0748_U1, CellID=7481, LogicRNCID=501',
'Label=UH3922_U4, CellID=39224, LogicRNCID=501',
'Label=UH0869_U95, CellID=48695, LogicRNCID=501',
'Label=UH0869_U94, CellID=48694, LogicRNCID=501',
'Label=UH0617_U4, CellID=6174, LogicRNCID=501',
'Label=UH3996_U5, CellID=39965, LogicRNCID=501',
'Label=UH0822_U1, CellID=8221, LogicRNCID=501',
'Label=UH0636_U94, CellID=46364, LogicRNCID=501',
'Label=UH0770_U5, CellID=7705, LogicRNCID=501',
'Label=UH0821_U6, CellID=8216, LogicRNCID=501',
'Label=UH0821_U5, CellID=8215, LogicRNCID=501',
'Label=UH0879_U94, CellID=55062, LogicRNCID=501',
'Label=UH0611_U95, CellID=46115, LogicRNCID=501',
'Label=UH3920_U2, CellID=39202, LogicRNCID=501',
'Label=UH0822_U6, CellID=8226, LogicRNCID=501',
'Label=UH0950_U5, CellID=9505, LogicRNCID=501',
'Label=UH3921_U3, CellID=39213, LogicRNCID=501',
'Label=UH0740_U94, CellID=47404, LogicRNCID=501',
'Label=UH0951_U6, CellID=9516, LogicRNCID=501',
'Label=UH0823_U6, CellID=8236, LogicRNCID=501',
'Label=UH0825_U1, CellID=8251, LogicRNCID=501',
'Label=UH0767_U94, CellID=47674, LogicRNCID=501',
'Label=UH0748_U2, CellID=7482, LogicRNCID=501',
'Label=UH3922_U5, CellID=39225, LogicRNCID=501',
'Label=UH1951_U3, CellID=19513, LogicRNCID=501',
'Label=UH0869_U97, CellID=48697, LogicRNCID=501',
'Label=UH1951_U4, CellID=19514, LogicRNCID=501',
'Label=UH0748_U3, CellID=7483, LogicRNCID=501',
'Label=UH3921_U1, CellID=39211, LogicRNCID=501',
'Label=UH0849_U2, CellID=8492, LogicRNCID=501',
'Label=UH3921_U2, CellID=39212, LogicRNCID=501',
'Label=UH0849_U3, CellID=8493, LogicRNCID=501',
'Label=UH0822_U5, CellID=8225, LogicRNCID=501',
'Label=UH3920_U1, CellID=39201, LogicRNCID=501',
'Label=UH0950_U6, CellID=9506, LogicRNCID=501',
'Label=UH0611_U94, CellID=46114, LogicRNCID=501',
'Label=UH0840_U96, CellID=48406, LogicRNCID=501',
'Label=UH0770_U3, CellID=7703, LogicRNCID=501',
'Label=UH0770_U4, CellID=7704, LogicRNCID=501',
'Label=UH0738_U94, CellID=47384, LogicRNCID=501',
'Label=UH3996_U4, CellID=39964, LogicRNCID=501',
'Label=UH0617_U3, CellID=6173, LogicRNCID=501',
'Label=UH0950_U1, CellID=9501, LogicRNCID=501',
'Label=UH0636_U95, CellID=46365, LogicRNCID=501',
'Label=UH3922_U2, CellID=39222, LogicRNCID=501',
'Label=UH1951_U1, CellID=19511, LogicRNCID=501',
'Label=UH1951_U2, CellID=19512, LogicRNCID=501',
'Label=UH0869_U96, CellID=48696, LogicRNCID=501',
'Label=UH1947_U3, CellID=19473, LogicRNCID=501',
'Label=UH0846_U6, CellID=8466, LogicRNCID=501',
'Label=UH1947_U4, CellID=19474, LogicRNCID=501',
'Label=UH3995_U4, CellID=39954, LogicRNCID=501',
'Label=UH0974_U2, CellID=9742, LogicRNCID=501',
'Label=UH0846_U3, CellID=8463, LogicRNCID=501',
'Label=UH1947_U1, CellID=19471, LogicRNCID=501',
'Label=UH3995_U1, CellID=39951, LogicRNCID=501',
'Label=UH0823_U3, CellID=8233, LogicRNCID=501',
'Label=UH0714_U95, CellID=47145, LogicRNCID=501',
'Label=UH0823_U4, CellID=8234, LogicRNCID=501',
'Label=UH0823_U5, CellID=8235, LogicRNCID=501',
'Label=UH0743_U4, CellID=7434, LogicRNCID=501',
'Label=UH0743_U5, CellID=7435, LogicRNCID=501',
'Label=UH0839_U95, CellID=48395, LogicRNCID=501',
'Label=UH1976_U4, CellID=19764, LogicRNCID=501',
'Label=UH0843_U96, CellID=48436, LogicRNCID=501',
'Label=UH1976_U5, CellID=19765, LogicRNCID=501',
'Label=UH0952_U6, CellID=9526, LogicRNCID=501',
'Label=UH0821_U1, CellID=8211, LogicRNCID=501',
'Label=UH3995_U5, CellID=39955, LogicRNCID=501',
'Label=UH0821_U2, CellID=8212, LogicRNCID=501',
'Label=UH3995_U6, CellID=39956, LogicRNCID=501',
'Label=UH0950_U3, CellID=9503, LogicRNCID=501',
'Label=UH0841_U95, CellID=48415, LogicRNCID=501',
'Label=UH0822_U4, CellID=8224, LogicRNCID=501',
'Label=UH0713_U96, CellID=47136, LogicRNCID=501',
'Label=UH0973_U3, CellID=9733, LogicRNCID=501',
'Label=UH0717_U4, CellID=7174, LogicRNCID=501',
'Label=UH0973_U4, CellID=9734, LogicRNCID=501',
'Label=UH0717_U5, CellID=7175, LogicRNCID=501',
'Label=UH0717_U6, CellID=7176, LogicRNCID=501',
'Label=UH0743_U3, CellID=7433, LogicRNCID=501',
'Label=UH0766_U94, CellID=47666, LogicRNCID=501',
'Label=UH0747_U5, CellID=7475, LogicRNCID=501',
'Label=UH1976_U3, CellID=19763, LogicRNCID=501',
'Label=UH0843_U95, CellID=48435, LogicRNCID=501',
'Label=UH0618_U3, CellID=6183, LogicRNCID=501',
'Label=UH0951_U1, CellID=9511, LogicRNCID=501',
'Label=UH0823_U2, CellID=8232, LogicRNCID=501',
'Label=UH0765_U96, CellID=47656, LogicRNCID=501',
'Label=UH0820_U4, CellID=8204, LogicRNCID=501',
'Label=UH0839_U96, CellID=48396, LogicRNCID=501',
'Label=UH0846_U1, CellID=8461, LogicRNCID=501',
'Label=UH3918_U1, CellID=39181, LogicRNCID=501',
'Label=UH3996_U1, CellID=39961, LogicRNCID=501',
'Label=UH0770_U6, CellID=7706, LogicRNCID=501',
'Label=UH0738_U96, CellID=47386, LogicRNCID=501',
'Label=UH3996_U3, CellID=39963, LogicRNCID=501',
'Label=UH0717_U2, CellID=7172, LogicRNCID=501',
'Label=UH0973_U2, CellID=9732, LogicRNCID=501',
'Label=UH0717_U3, CellID=7173, LogicRNCID=501',
'Label=UH3921_U6, CellID=39216, LogicRNCID=501',
'Label=UH0952_U1, CellID=9521, LogicRNCID=501',
'Label=UH0766_U95, CellID=47665, LogicRNCID=501',
'Label=UH1976_U2, CellID=19762, LogicRNCID=501',
'Label=UH0840_U94, CellID=48404, LogicRNCID=501',
'Label=UH0770_U1, CellID=7701, LogicRNCID=501',
'Label=UH0840_U95, CellID=48405, LogicRNCID=501',
'Label=UH0770_U2, CellID=7702, LogicRNCID=501',
'Label=UH0822_U2, CellID=8222, LogicRNCID=501',
'Label=UH3996_U6, CellID=39966, LogicRNCID=501',
'Label=UH0713_U94, CellID=47134, LogicRNCID=501',
'Label=UH0617_U5, CellID=6175, LogicRNCID=501',
'Label=UH0739_U95, CellID=47395, LogicRNCID=501',
'Label=UH0611_U96, CellID=46116, LogicRNCID=501',
'Label=UH0618_U1, CellID=6181, LogicRNCID=501',
'Label=UH0618_U2, CellID=6182, LogicRNCID=501',
'Label=UH0825_U3, CellID=8253, LogicRNCID=501',
'Label=UH0844_U95, CellID=48445, LogicRNCID=501',
'Label=UH0825_U4, CellID=8254, LogicRNCID=501',
'Label=UH0844_U96, CellID=48446, LogicRNCID=501',
'Label=UH1945_U6, CellID=19456, LogicRNCID=501',
'Label=UH0838_U94, CellID=48384, LogicRNCID=501',
'Label=UH0761_U97, CellID=47617, LogicRNCID=501',
'Label=UH0838_U96, CellID=48386, LogicRNCID=501',
'Label=UH0740_U95, CellID=47405, LogicRNCID=501',
'Label=UH3921_U4, CellID=39214, LogicRNCID=501',
'Label=UH0849_U5, CellID=8495, LogicRNCID=501',
'Label=UH0747_U2, CellID=7472, LogicRNCID=501',
'Label=UH0844_U94, CellID=48444, LogicRNCID=501',
'Label=UH1805_U6, CellID=18056, LogicRNCID=501',
'Label=UH1703_U2, CellID=17032, LogicRNCID=501',
'Label=UH0807_U2, CellID=8072, LogicRNCID=501',
'Label=UH0602_U4, CellID=6024, LogicRNCID=501',
'Label=UH0988_U5, CellID=9885, LogicRNCID=501',
'Label=UH0758_U1, CellID=7581, LogicRNCID=501',
'Label=UH0751_U96, CellID=47516, LogicRNCID=501',
'Label=UH0623_U96, CellID=46236, LogicRNCID=501',
'Label=UH1935_U6, CellID=19356, LogicRNCID=501',
'Label=UH1833_U2, CellID=18332, LogicRNCID=501',
'Label=UH1705_U2, CellID=17052, LogicRNCID=501',
'Label=UH0988_U4, CellID=9884, LogicRNCID=501',
'Label=UH2709_U95, CellID=50066, LogicRNCID=501',
'Label=UH0955_U94, CellID=49554, LogicRNCID=501',
'Label=UH0750_U96, CellID=47506, LogicRNCID=501',
'Label=UH0622_U96, CellID=46226, LogicRNCID=501',
'Label=UH2984_U2, CellID=29842, LogicRNCID=501',
'Label=UH1960_U2, CellID=19602, LogicRNCID=501',
'Label=UH1704_U2, CellID=17042, LogicRNCID=501',
'Label=UH2709_U94, CellID=50065, LogicRNCID=501',
'Label=UH0965_U3, CellID=9653, LogicRNCID=501',
'Label=UH4600_U4, CellID=46004, LogicRNCID=501',
'Label=UH1707_U6, CellID=17076, LogicRNCID=501',
'Label=UH0965_U2, CellID=9652, LogicRNCID=501',
'Label=UH4600_U3, CellID=46003, LogicRNCID=501',
'Label=UH1707_U5, CellID=17075, LogicRNCID=501',
'Label=UH0965_U1, CellID=9651, LogicRNCID=501',
'Label=UH0830_U96, CellID=48306, LogicRNCID=501',
'Label=UH0807_U5, CellID=8075, LogicRNCID=501',
'Label=UH1945_U96, CellID=50058, LogicRNCID=501',
'Label=UH2983_U4, CellID=29834, LogicRNCID=501',
'Label=UH2932_U2, CellID=29322, LogicRNCID=501',
'Label=UH1908_U2, CellID=19082, LogicRNCID=501',
'Label=UH1882_U6, CellID=18826, LogicRNCID=501',
'Label=UH1703_U4, CellID=17034, LogicRNCID=501',
'Label=UH0807_U4, CellID=8074, LogicRNCID=501',
'Label=UH1936_U4, CellID=19364, LogicRNCID=501',
'Label=UH0963_U6, CellID=9636, LogicRNCID=501',
'Label=UH0835_U6, CellID=8356, LogicRNCID=501',
'Label=UH1936_U3, CellID=19363, LogicRNCID=501',
'Label=UH0963_U5, CellID=9635, LogicRNCID=501',
'Label=UH0835_U5, CellID=8355, LogicRNCID=501',
'Label=UH1936_U2, CellID=19362, LogicRNCID=501',
'Label=UH0963_U4, CellID=9634, LogicRNCID=501',
'Label=UH2932_U5, CellID=29325, LogicRNCID=501',
'Label=UH1908_U5, CellID=19085, LogicRNCID=501',
'Label=UH0961_U3, CellID=9613, LogicRNCID=501',
'Label=UH1945_U98, CellID=50060, LogicRNCID=501',
'Label=UH2983_U6, CellID=29836, LogicRNCID=501',
'Label=UH2932_U4, CellID=29324, LogicRNCID=501',
'Label=UH1908_U4, CellID=19084, LogicRNCID=501',
'Label=UH1703_U6, CellID=17036, LogicRNCID=501',
'Label=UH0750_U95, CellID=47505, LogicRNCID=501',
'Label=UH0622_U95, CellID=46225, LogicRNCID=501',
'Label=UH2984_U1, CellID=29841, LogicRNCID=501',
'Label=UH1960_U1, CellID=19601, LogicRNCID=501',
'Label=UH1704_U1, CellID=17041, LogicRNCID=501',
'Label=UH1951_U96, CellID=50064, LogicRNCID=501',
'Label=UH0750_U94, CellID=47504, LogicRNCID=501',
'Label=UH0622_U94, CellID=46224, LogicRNCID=501',
'Label=UH0754_U95, CellID=47545, LogicRNCID=501',
'Label=UH1964_U1, CellID=19641, LogicRNCID=501',
'Label=UH0754_U94, CellID=47544, LogicRNCID=501',
'Label=UH0965_U6, CellID=9656, LogicRNCID=501',
'Label=UH0965_U5, CellID=9655, LogicRNCID=501',
'Label=UH4600_U6, CellID=46006, LogicRNCID=501',
'Label=UH0965_U4, CellID=9654, LogicRNCID=501',
'Label=UH4600_U5, CellID=46005, LogicRNCID=501',
'Label=UH1935_U4, CellID=19354, LogicRNCID=501',
'Label=UH0988_U2, CellID=9882, LogicRNCID=501',
'Label=UH0962_U6, CellID=9626, LogicRNCID=501',
'Label=UH0760_U96, CellID=50073, LogicRNCID=501',
'Label=UH1935_U3, CellID=19353, LogicRNCID=501',
'Label=UH0988_U1, CellID=9881, LogicRNCID=501',
'Label=UH0962_U5, CellID=9625, LogicRNCID=501',
'Label=UH0760_U95, CellID=50072, LogicRNCID=501',
'Label=UH1704_U6, CellID=17046, LogicRNCID=501',
'Label=UH0962_U2, CellID=9622, LogicRNCID=501',
'Label=UH2984_U5, CellID=29845, LogicRNCID=501',
'Label=UH1960_U5, CellID=19605, LogicRNCID=501',
'Label=UH1704_U5, CellID=17045, LogicRNCID=501',
'Label=UH0962_U1, CellID=9621, LogicRNCID=501',
'Label=UH0706_U1, CellID=7061, LogicRNCID=501',
'Label=UH0955_U96, CellID=49556, LogicRNCID=501',
'Label=UH0758_U3, CellID=7583, LogicRNCID=501',
'Label=UH1801_U94, CellID=58014, LogicRNCID=501',
'Label=UH1833_U4, CellID=18334, LogicRNCID=501',
'Label=UH1705_U4, CellID=17054, LogicRNCID=501',
'Label=UH0988_U6, CellID=9886, LogicRNCID=501',
'Label=UH0758_U2, CellID=7582, LogicRNCID=501',
'Label=UH1833_U3, CellID=18333, LogicRNCID=501',
'Label=UH1705_U3, CellID=17053, LogicRNCID=501',
'Label=UH0708_U3, CellID=7083, LogicRNCID=501',
'Label=UH0708_U2, CellID=7082, LogicRNCID=501',
'Label=UH0708_U1, CellID=7081, LogicRNCID=501',
'Label=UH0752_U96, CellID=47526, LogicRNCID=501',
'Label=UH1936_U6, CellID=19366, LogicRNCID=501',
'Label=UH0752_U95, CellID=47525, LogicRNCID=501',
'Label=UH1936_U5, CellID=19365, LogicRNCID=501',
'Label=UH0752_U94, CellID=47524, LogicRNCID=501',
'Label=UH0831_U96, CellID=48316, LogicRNCID=501',
'Label=UH0703_U96, CellID=47036, LogicRNCID=501',
'Label=UH1964_U4, CellID=19644, LogicRNCID=501',
'Label=UH0831_U95, CellID=48315, LogicRNCID=501',
'Label=UH0703_U95, CellID=47035, LogicRNCID=501',
'Label=UH0831_U94, CellID=48314, LogicRNCID=501',
'Label=UH0703_U94, CellID=47034, LogicRNCID=501',
'Label=UH1964_U2, CellID=19642, LogicRNCID=501',
'Label=UH3925_U5, CellID=39255, LogicRNCID=501',
'Label=UH0981_U5, CellID=9815, LogicRNCID=501',
'Label=UH0623_U1, CellID=6231, LogicRNCID=501',
'Label=UH0821_U94, CellID=48214, LogicRNCID=501',
'Label=UH0744_U96, CellID=47446, LogicRNCID=501',
'Label=UH3925_U4, CellID=39254, LogicRNCID=501',
'Label=UH1928_U6, CellID=19286, LogicRNCID=501',
'Label=UH0981_U4, CellID=9814, LogicRNCID=501',
'Label=UH1952_U1, CellID=19521, LogicRNCID=501',
'Label=UH1602_U95, CellID=51008, LogicRNCID=501',
'Label=UH0742_U94, CellID=47424, LogicRNCID=501',
'Label=UH0614_U94, CellID=46144, LogicRNCID=501',
'Label=UH2950_U4, CellID=29504, LogicRNCID=501',
'Label=UH0825_U6, CellID=8256, LogicRNCID=501',
'Label=UH1602_U94, CellID=51007, LogicRNCID=501',
'Label=UH2950_U3, CellID=29503, LogicRNCID=501',
'Label=UH0742_U96, CellID=47426, LogicRNCID=501',
'Label=UH0614_U96, CellID=46146, LogicRNCID=501',
'Label=UH2950_U6, CellID=29506, LogicRNCID=501',
'Label=UH1952_U2, CellID=19522, LogicRNCID=501',
'Label=UH1602_U96, CellID=51009, LogicRNCID=501',
'Label=UH0742_U95, CellID=47425, LogicRNCID=501',
'Label=UH0614_U95, CellID=46145, LogicRNCID=501',
'Label=UH2950_U5, CellID=29505, LogicRNCID=501',
'Label=UH1925_U95, CellID=50047, LogicRNCID=501',
'Label=UH0825_U95, CellID=48255, LogicRNCID=501',
'Label=UH1958_U3, CellID=19583, LogicRNCID=501',
'Label=UH1907_U1, CellID=19071, LogicRNCID=501',
'Label=UH1702_U3, CellID=17023, LogicRNCID=501',
'Label=UH0601_U5, CellID=6015, LogicRNCID=501',
'Label=UH1925_U96, CellID=50046, LogicRNCID=501',
'Label=UH0825_U94, CellID=48254, LogicRNCID=501',
'Label=UH1952_U6, CellID=19526, LogicRNCID=501',
'Label=UH1952_U5, CellID=19525, LogicRNCID=501',
'Label=UH0646_U96, CellID=51012, LogicRNCID=501',
'Label=UH1952_U4, CellID=19524, LogicRNCID=501',
'Label=UH0646_U95, CellID=51011, LogicRNCID=501',
'Label=UH1952_U3, CellID=19523, LogicRNCID=501',
'Label=UH0646_U94, CellID=51010, LogicRNCID=501',
'Label=UH1902_U95, CellID=50035, LogicRNCID=501',
'Label=UH0747_U95, CellID=47475, LogicRNCID=501',
'Label=UH3928_U3, CellID=39283, LogicRNCID=501',
'Label=UH2981_U1, CellID=29811, LogicRNCID=501',
'Label=UH1803_U5, CellID=18035, LogicRNCID=501',
'Label=UH1902_U94, CellID=50034, LogicRNCID=501',
'Label=UH0747_U94, CellID=47474, LogicRNCID=501',
'Label=UH3928_U2, CellID=39282, LogicRNCID=501',
'Label=UH0973_U95, CellID=49735, LogicRNCID=501',
'Label=UH0717_U95, CellID=47175, LogicRNCID=501',
'Label=UH0973_U94, CellID=49734, LogicRNCID=501',
'Label=UH0717_U94, CellID=47174, LogicRNCID=501',
'Label=UH1702_U5, CellID=17025, LogicRNCID=501',
'Label=UH0960_U1, CellID=9601, LogicRNCID=501',
'Label=UH0832_U1, CellID=8321, LogicRNCID=501',
'Label=UH1925_U94, CellID=50048, LogicRNCID=501',
'Label=UH0825_U96, CellID=48256, LogicRNCID=501',
'Label=UH1907_U2, CellID=19072, LogicRNCID=501',
'Label=UH1702_U4, CellID=17024, LogicRNCID=501',
'Label=UH0601_U6, CellID=6016, LogicRNCID=501',
'Label=UH1905_U4, CellID=19054, LogicRNCID=501',
'Label=UH1854_U2, CellID=18542, LogicRNCID=501',
'Label=UH0881_U4, CellID=8814, LogicRNCID=501',
'Label=UH0830_U2, CellID=8302, LogicRNCID=501',
'Label=UH1854_U1, CellID=18541, LogicRNCID=501',
'Label=UH0881_U3, CellID=8813, LogicRNCID=501',
'Label=UH0830_U1, CellID=8301, LogicRNCID=501',
'Label=UH0618_U94, CellID=46184, LogicRNCID=501',
'Label=UH1804_U95, CellID=50023, LogicRNCID=501',
'Label=UH1804_U94, CellID=50022, LogicRNCID=501',
'Label=UH0880_U6, CellID=8806, LogicRNCID=501',
'Label=UH1803_U96, CellID=50021, LogicRNCID=501',
'Label=UH1603_U96, CellID=50012, LogicRNCID=501',
'Label=UH0879_U6, CellID=8796, LogicRNCID=501',
'Label=UH0623_U6, CellID=6236, LogicRNCID=501',
'Label=UH1964_U94, CellID=55387, LogicRNCID=501',
'Label=UH1603_U95, CellID=50011, LogicRNCID=501',
'Label=UH0821_U99, CellID=48219, LogicRNCID=501',
'Label=UH1801_U1, CellID=18011, LogicRNCID=501',
'Label=UH0879_U5, CellID=8795, LogicRNCID=501',
'Label=UH2981_U3, CellID=29813, LogicRNCID=501',
'Label=UH1906_U1, CellID=19061, LogicRNCID=501',
'Label=UH0600_U5, CellID=6005, LogicRNCID=501',
'Label=UH1902_U96, CellID=50036, LogicRNCID=501',
'Label=UH3928_U4, CellID=39284, LogicRNCID=501',
'Label=UH2981_U2, CellID=29812, LogicRNCID=501',
'Label=UH1803_U6, CellID=18036, LogicRNCID=501',
'Label=UH0600_U4, CellID=6004, LogicRNCID=501',
'Label=UH0980_U2, CellID=9802, LogicRNCID=501',
'Label=UH0980_U1, CellID=9801, LogicRNCID=501',
'Label=UH0973_U96, CellID=49736, LogicRNCID=501',
'Label=UH0717_U96, CellID=47176, LogicRNCID=501',
'Label=UH0822_U94, CellID=48224, LogicRNCID=501',
'Label=UH0745_U96, CellID=47456, LogicRNCID=501',
'Label=UH1801_U6, CellID=18016, LogicRNCID=501',
'Label=UH0854_U4, CellID=8544, LogicRNCID=501',
'Label=UH1604_U96, CellID=50015, LogicRNCID=501',
'Label=UH0745_U95, CellID=47455, LogicRNCID=501',
'Label=UH0854_U3, CellID=8543, LogicRNCID=501',
'Label=UH1604_U95, CellID=50014, LogicRNCID=501',
'Label=UH0955_U6, CellID=9556, LogicRNCID=501',
'Label=UH3925_U1, CellID=39251, LogicRNCID=501',
'Label=UH1928_U3, CellID=19283, LogicRNCID=501',
'Label=UH0981_U1, CellID=9811, LogicRNCID=501',
'Label=UH0955_U5, CellID=9555, LogicRNCID=501',
'Label=UH0846_U96, CellID=48466, LogicRNCID=501',
'Label=UH1928_U2, CellID=19282, LogicRNCID=501',
'Label=UH0955_U4, CellID=9554, LogicRNCID=501',
'Label=UH1882_U94, CellID=50031, LogicRNCID=501',
'Label=UH0849_U95, CellID=48495, LogicRNCID=501',
'Label=UH1905_U5, CellID=19055, LogicRNCID=501',
'Label=UH1854_U3, CellID=18543, LogicRNCID=501',
'Label=UH1803_U1, CellID=18031, LogicRNCID=501',
'Label=UH0881_U5, CellID=8815, LogicRNCID=501',
'Label=UH0830_U3, CellID=8303, LogicRNCID=501',
'Label=UH0849_U94, CellID=48494, LogicRNCID=501',
'Label=UH1906_U3, CellID=19063, LogicRNCID=501',
'Label=UH0831_U1, CellID=8311, LogicRNCID=501',
'Label=UH1907_U95, CellID=50038, LogicRNCID=501',
'Label=UH3928_U6, CellID=39286, LogicRNCID=501',
'Label=UH2981_U4, CellID=29814, LogicRNCID=501',
'Label=UH1906_U2, CellID=19062, LogicRNCID=501',
'Label=UH1907_U94, CellID=50037, LogicRNCID=501',
'Label=UH3928_U5, CellID=39285, LogicRNCID=501',
'Label=UH0955_U2, CellID=9552, LogicRNCID=501',
'Label=UH0622_U4, CellID=6224, LogicRNCID=501',
'Label=UH0820_U97, CellID=48207, LogicRNCID=501',
'Label=UH0955_U1, CellID=9551, LogicRNCID=501',
'Label=UH0622_U3, CellID=6223, LogicRNCID=501',
'Label=UH0820_U96, CellID=48206, LogicRNCID=501',
'Label=UH0980_U6, CellID=9806, LogicRNCID=501',
'Label=UH0622_U2, CellID=6222, LogicRNCID=501',
'Label=UH0820_U95, CellID=48205, LogicRNCID=501',
'Label=UH0980_U5, CellID=9805, LogicRNCID=501',
'Label=UH0622_U1, CellID=6221, LogicRNCID=501',
'Label=UH0820_U94, CellID=48204, LogicRNCID=501',
'Label=UH0743_U96, CellID=47436, LogicRNCID=501',
'Label=UH0980_U4, CellID=9804, LogicRNCID=501',
'Label=UH0743_U95, CellID=47435, LogicRNCID=501',
'Label=UH0980_U3, CellID=9803, LogicRNCID=501',
'Label=UH0951_U94, CellID=49514, LogicRNCID=501',
'Label=UH0823_U94, CellID=48234, LogicRNCID=501',
'Label=UH0746_U96, CellID=47466, LogicRNCID=501',
'Label=UH0618_U96, CellID=46186, LogicRNCID=501',
'Label=UH1805_U94, CellID=50025, LogicRNCID=501',
'Label=UH0746_U95, CellID=47465, LogicRNCID=501',
'Label=UH0618_U95, CellID=46185, LogicRNCID=501',
'Label=UH1804_U96, CellID=50024, LogicRNCID=501',
'Label=UH0623_U5, CellID=6235, LogicRNCID=501',
'Label=UH1603_U94, CellID=50010, LogicRNCID=501',
'Label=UH0821_U98, CellID=48218, LogicRNCID=501',
'Label=UH0770_U96, CellID=47706, LogicRNCID=501',
'Label=UH0623_U4, CellID=6234, LogicRNCID=501',
'Label=UH0821_U97, CellID=48217, LogicRNCID=501',
'Label=UH0770_U95, CellID=47705, LogicRNCID=501',
'Label=UH0879_U3, CellID=8793, LogicRNCID=501',
'Label=UH1906_U5, CellID=19065, LogicRNCID=501',
'Label=UH0831_U3, CellID=8313, LogicRNCID=501',
'Label=UH1913_U94, CellID=50040, LogicRNCID=501',
'Label=UH2981_U6, CellID=29816, LogicRNCID=501',
'Label=UH1906_U4, CellID=19064, LogicRNCID=501',
'Label=UH0831_U2, CellID=8312, LogicRNCID=501',
'Label=UH1907_U96, CellID=50039, LogicRNCID=501',
'Label=UH2981_U5, CellID=29815, LogicRNCID=501',
'Label=UH0622_U6, CellID=6226, LogicRNCID=501',
'Label=UH0846_U95, CellID=48465, LogicRNCID=501',
'Label=UH0820_U99, CellID=48209, LogicRNCID=501',
'Label=UH1928_U1, CellID=19281, LogicRNCID=501',
'Label=UH0955_U3, CellID=9553, LogicRNCID=501',
'Label=UH0622_U5, CellID=6225, LogicRNCID=501',
'Label=UH0846_U94, CellID=48464, LogicRNCID=501',
'Label=UH0820_U98, CellID=48208, LogicRNCID=501',
'Label=UH0830_U5, CellID=8305, LogicRNCID=501',
'Label=UH0600_U1, CellID=6001, LogicRNCID=501',
'Label=UH1882_U95, CellID=50032, LogicRNCID=501',
'Label=UH0849_U96, CellID=48496, LogicRNCID=501',
'Label=UH1854_U4, CellID=18544, LogicRNCID=501',
'Label=UH1803_U2, CellID=18032, LogicRNCID=501',
'Label=UH0881_U6, CellID=8816, LogicRNCID=501',
'Label=UH0830_U4, CellID=8304, LogicRNCID=501',
'Label=UH0744_U95, CellID=47445, LogicRNCID=501',
'Label=UH3925_U3, CellID=39253, LogicRNCID=501',
'Label=UH1928_U5, CellID=19285, LogicRNCID=501',
'Label=UH0981_U3, CellID=9813, LogicRNCID=501',
'Label=UH0744_U94, CellID=47444, LogicRNCID=501',
'Label=UH3925_U2, CellID=39252, LogicRNCID=501',
'Label=UH1928_U4, CellID=19284, LogicRNCID=501',
'Label=UH0981_U2, CellID=9812, LogicRNCID=501',
'Label=UH0623_U3, CellID=6233, LogicRNCID=501',
'Label=UH0821_U96, CellID=48216, LogicRNCID=501',
'Label=UH0770_U94, CellID=47704, LogicRNCID=501',
'Label=UH3925_U6, CellID=39256, LogicRNCID=501',
'Label=UH0981_U6, CellID=9816, LogicRNCID=501',
'Label=UH0879_U2, CellID=8792, LogicRNCID=501',
'Label=UH0623_U2, CellID=6232, LogicRNCID=501',
'Label=UH0821_U95, CellID=48215, LogicRNCID=501',
'Label=UH0831_U6, CellID=8316, LogicRNCID=501',
'Label=UH0601_U2, CellID=6012, LogicRNCID=501',
'Label=UH0831_U5, CellID=8315, LogicRNCID=501',
'Label=UH0601_U1, CellID=6011, LogicRNCID=501',
'Label=UH1913_U96, CellID=50042, LogicRNCID=501',
'Label=UH1906_U6, CellID=19066, LogicRNCID=501',
'Label=UH0831_U4, CellID=8314, LogicRNCID=501',
'Label=UH1913_U95, CellID=50041, LogicRNCID=501',
'Label=UH0880_U4, CellID=8804, LogicRNCID=501',
'Label=UH1803_U94, CellID=50019, LogicRNCID=501',
'Label=UH0880_U3, CellID=8803, LogicRNCID=501',
'Label=UH0822_U96, CellID=48226, LogicRNCID=501',
'Label=UH0854_U6, CellID=8546, LogicRNCID=501',
'Label=UH0822_U95, CellID=48225, LogicRNCID=501',
'Label=UH0880_U1, CellID=8801, LogicRNCID=501',
'Label=UH0854_U5, CellID=8545, LogicRNCID=501',
'Label=UH0745_U94, CellID=47454, LogicRNCID=501',
'Label=UH1801_U4, CellID=18014, LogicRNCID=501',
'Label=UH0854_U2, CellID=8542, LogicRNCID=501',
'Label=UH1964_U96, CellID=55389, LogicRNCID=501',
'Label=UH1604_U94, CellID=50013, LogicRNCID=501',
'Label=UH1801_U3, CellID=18013, LogicRNCID=501',
'Label=UH0854_U1, CellID=8541, LogicRNCID=501',
'Label=UH1964_U95, CellID=55388, LogicRNCID=501',
'Label=UH0823_U96, CellID=48236, LogicRNCID=501',
'Label=UH1905_U2, CellID=19052, LogicRNCID=501',
'Label=UH0881_U2, CellID=8812, LogicRNCID=501',
'Label=UH1805_U96, CellID=50027, LogicRNCID=501',
'Label=UH0823_U95, CellID=48235, LogicRNCID=501',
'Label=UH1905_U1, CellID=19051, LogicRNCID=501',
'Label=UH0881_U1, CellID=8811, LogicRNCID=501',
'Label=UH1805_U95, CellID=50026, LogicRNCID=501',
'Label=UH1854_U6, CellID=18546, LogicRNCID=501',
'Label=UH1803_U4, CellID=18034, LogicRNCID=501',
'Label=UH0830_U6, CellID=8306, LogicRNCID=501',
'Label=UH0600_U2, CellID=6002, LogicRNCID=501',
'Label=UH1882_U96, CellID=50033, LogicRNCID=501',
'Label=UH3928_U1, CellID=39281, LogicRNCID=501',
'Label=UH1854_U5, CellID=18545, LogicRNCID=501',
'Label=UH1803_U3, CellID=18033, LogicRNCID=501',
'Label=UH0748_U96, CellID=47486, LogicRNCID=501',
'Label=UH1958_U2, CellID=19582, LogicRNCID=501',
'Label=UH1702_U2, CellID=17022, LogicRNCID=501',
'Label=UH0601_U4, CellID=6014, LogicRNCID=501',
'Label=UH0748_U95, CellID=47485, LogicRNCID=501',
'Label=UH1702_U1, CellID=17021, LogicRNCID=501',
'Label=UH0601_U3, CellID=6013, LogicRNCID=501',
'Label=UH0748_U94, CellID=47484, LogicRNCID=501',
'Label=UH0974_U5, CellID=9745, LogicRNCID=501',
'Label=UH0737_U96, CellID=47376, LogicRNCID=501',
'Label=UH3995_U2, CellID=39952, LogicRNCID=501',
'Label=UH3918_U4, CellID=39184, LogicRNCID=501',
'Label=UH1947_U2, CellID=19472, LogicRNCID=501',
'Label=UH0974_U4, CellID=9744, LogicRNCID=501',
'Label=UH0846_U4, CellID=8464, LogicRNCID=501',
'Label=UH0839_U99, CellID=48399, LogicRNCID=501',
]
sts_df = sts_df[sts_df['BSC6910UCell'].isin(cluster_UH_noUH0845_top)]
print('sts_df отфильтрована')
stsN_df = stsN_df[stsN_df['NE Name'].isin(cluster_UH_N)]

active_cell_number = sts_df['BSC6910UCell'].nunique()
print('active_cell_number= ',  active_cell_number)

# ===обработка weekly  для всей сети без разбивки на кластера===
weekly_df = sts_df.groupby(['week'])[list_1]. sum().reset_index()
weekly_df['CS traffic 3G, Erl'] = weekly_df['CS Voice Traffic Volume (Erl)']
weekly_df['PS traffic 3G UL+DL, GB'] = (weekly_df['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + weekly_df['VS.PS.Bkg.DL.8.Traffic (bit)'] + weekly_df['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Bkg.DL.32.Traffic (bit)'] + weekly_df['VS.PS.Bkg.DL.64.Traffic (bit)'] + weekly_df['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Bkg.DL.144.Traffic (bit)'] + weekly_df['VS.PS.Bkg.DL.256.Traffic (bit)'] + weekly_df['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Bkg.UL.8.Traffic (bit)'] + weekly_df['VS.PS.Bkg.UL.16.Traffic (bit)'] + weekly_df['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Bkg.UL.64.Traffic (bit)'] + weekly_df['VS.PS.Bkg.UL.128.Traffic (bit)'] + weekly_df['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Bkg.UL.256.Traffic (bit)'] + weekly_df['VS.PS.Bkg.UL.384.Traffic (bit)'] + weekly_df['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Int.DL.16.Traffic (bit)'] + weekly_df['VS.PS.Int.DL.32.Traffic (bit)'] + weekly_df['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Int.DL.128.Traffic (bit)'] + weekly_df['VS.PS.Int.DL.144.Traffic (bit)'] + weekly_df['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Int.DL.384.Traffic (bit)'] + weekly_df['VS.PS.Int.UL.8.Traffic (bit)'] + weekly_df['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Int.UL.32.Traffic (bit)'] + weekly_df['VS.PS.Int.UL.64.Traffic (bit)'] + weekly_df['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Int.UL.144.Traffic (bit)'] + weekly_df['VS.PS.Int.UL.256.Traffic (bit)'] + weekly_df['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Str.DL.32.Traffic (bit)'] + weekly_df['VS.PS.Str.DL.64.Traffic (bit)'] + weekly_df['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Str.DL.144.Traffic (bit)'] + weekly_df['VS.PS.Str.UL.16.Traffic (bit)'] + weekly_df['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      weekly_df['VS.PS.Str.UL.64.Traffic (bit)'] + weekly_df['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
weekly_df['CS RAB Drop Rate (%)'] = weekly_df['VS.RAB.AbnormRel.CS (None)'] / (weekly_df['VS.RAB.AbnormRel.CS (None)'] + weekly_df['VS.RAB.NormRel.CS (None)']) * 100
weekly_df['PS Blocking Rate (%)'] = (weekly_df['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + weekly_df['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + weekly_df['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    weekly_df['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + weekly_df['VS.RAB.FailEstabPS.Code.Cong (None)'] + weekly_df['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    weekly_df['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + weekly_df['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + weekly_df['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (weekly_df['VS.RAB.AttEstabPS.Str (None)'] + weekly_df['VS.RAB.AttEstabPS.Int (None)'] + weekly_df['VS.RAB.AttEstabPS.Bkg (None)']) *100

# weekly_df['PS RAB Drop Rate (%)'] = (weekly_df['VS.RAB.AbnormRel.PS (None)'] + weekly_df['VS.RAB.AbnormRel.PS.PCH (None)'] + weekly_df['VS.RAB.AbnormRel.PS.D2P (None)'] + \
#                                     weekly_df['VS.RAB.AbnormRel.PS.F2P (None)']) / \
#                                    (weekly_df['VS.RAB.AbnormRel.PS (None)'] + weekly_df['VS.RAB.NormRel.PS (None)'] + weekly_df['VS.RAB.AbnormRel.PS.PCH (None)'] + \
#                                     weekly_df['VS.RAB.NormRel.PS.PCH (None)']) * 100
weekly_df['PS RAB Drop Rate (%)'] = (weekly_df['VS.RAB.AbnormRel.PS (None)'] - weekly_df['VS.RAB.AbnormRel.PS.PCH (None)'] - weekly_df['VS.RAB.AbnormRel.PS.D2P (None)'] - \
                                    weekly_df['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (weekly_df['VS.RAB.AbnormRel.PS (None)'] + weekly_df['VS.RAB.NormRel.PS (None)'] - weekly_df['VS.RAB.AbnormRel.PS.PCH (None)'] - \
                                    weekly_df['VS.RAB.NormRel.PS.PCH (None)'] + weekly_df['VS.DCCC.D2P.Succ (None)'] + weekly_df['VS.DCCC.Succ.F2P (None)']+weekly_df['VS.DCCC.Succ.F2U (None)'] + weekly_df['VS.DCCC.Succ.D2U (None)']) * 100
weekly_df['PS HS- Drop Rate (%)'] =  weekly_df['VS.HSDPA.RAB.AbnormRel (None)'] / (weekly_df['VS.HSDPA.RAB.AbnormRel (None)'] + weekly_df['VS.HSDPA.RAB.NormRel (None)'] + weekly_df['VS.HSDPA.H2D.Succ (None)'] + weekly_df['VS.HSDPA.H2F.Succ (None)'] +weekly_df['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + weekly_df['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
weekly_df['HSDPA Throughput, kbps'] = weekly_df['VS.HSDPA.MeanChThroughput (kbit/s)'] / active_cell_number / 24 / 7 # количество сот
weekly_df['HSUPA Throughput, kbps'] = weekly_df['VS.HSUPA.MeanChThroughput (kbit/s)'] / active_cell_number / 24 / 7 # количество сот
weekly_df['Soft Handover Success rate, %'] = (weekly_df['VS.SHO.SuccRLAdd (None)'] + weekly_df['VS.SHO.SuccRLDel (None)']) / (weekly_df['VS.SHO.AttRLAdd (None)'] + weekly_df['VS.SHO.AttRLDel (None)']) * 100
weekly_df['Hard Handover Success rate, %'] = weekly_df['VS.HHO.SuccInterFreqOut (None)'] / weekly_df['VS.HHO.AttInterFreqOut (None)'] * 100
weekly_df['CS W2G Inter-RAT Handover Out SR'] = weekly_df['IRATHO.SuccOutCS (None)'] / (weekly_df['IRATHO.AttOutCS (None)'] - weekly_df['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
weekly_df['RRC Assignment SucessRate (CS BH), %'] = weekly_df['RRC.SuccConnEstab.sum (None)'] / weekly_df['VS.RRC.AttConnEstab.Sum (None)'] * 100
weekly_df['RRC Assignment SucessRate (PS BH), %'] = weekly_df['RRC.SuccConnEstab.sum (None)'] / weekly_df['VS.RRC.AttConnEstab.Sum (None)'] * 100
weekly_df['RRC Drop Rate (CS BH), %'] = (weekly_df['RRC.AttConnRelCCCH.Cong (None)'] + weekly_df['RRC.AttConnRelCCCH.Preempt (None)'] + weekly_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             weekly_df['RRC.AttConnRelCCCH.Unspec (None)'] + weekly_df['RRC.AttConnRelDCCH.Cong (None)'] + weekly_df['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             weekly_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + weekly_df['RRC.AttConnRelDCCH.Unspec (None)'] + weekly_df['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (weekly_df['RRC.AttConnRelDCCH.Cong (None)'] + weekly_df['RRC.AttConnRelDCCH.Preempt (None)'] + weekly_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           weekly_df['RRC.AttConnRelDCCH.DSCR (None)'] + weekly_df['RRC.AttConnRelDCCH.UsrInact (None)'] + weekly_df['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           weekly_df['RRC.AttConnRelCCCH.Cong (None)'] + weekly_df['RRC.AttConnRelCCCH.Preempt (None)'] + weekly_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           weekly_df['RRC.AttConnRelCCCH.DSCR (None)'] + weekly_df['RRC.AttConnRelDCCH.Norm (None)'] + weekly_df['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           weekly_df['RRC.AttConnRelCCCH.UsrInact (None)'] + weekly_df['RRC.AttConnRelCCCH.Unspec (None)'] + weekly_df['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           weekly_df['VS.DCCC.Succ.F2P (None)'] + weekly_df['IRATHO.SuccOutCS (None)'] + weekly_df['IRATHO.SuccOutPSUTRAN (None)'] + weekly_df['VS.DCCC.Succ.F2U (None)'] + \
                                           weekly_df['VS.DCCC.Succ.D2U (None)']) * 100
weekly_df['RRC Drop Rate (PS BH), %'] = (weekly_df['RRC.AttConnRelCCCH.Cong (None)'] + weekly_df['RRC.AttConnRelCCCH.Preempt (None)'] + weekly_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             weekly_df['RRC.AttConnRelCCCH.Unspec (None)'] + weekly_df['RRC.AttConnRelDCCH.Cong (None)'] + weekly_df['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             weekly_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + weekly_df['RRC.AttConnRelDCCH.Unspec (None)'] + weekly_df['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (weekly_df['RRC.AttConnRelDCCH.Cong (None)'] + weekly_df['RRC.AttConnRelDCCH.Preempt (None)'] + weekly_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           weekly_df['RRC.AttConnRelDCCH.DSCR (None)'] + weekly_df['RRC.AttConnRelDCCH.UsrInact (None)'] + weekly_df['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           weekly_df['RRC.AttConnRelCCCH.Cong (None)'] + weekly_df['RRC.AttConnRelCCCH.Preempt (None)'] + weekly_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           weekly_df['RRC.AttConnRelCCCH.DSCR (None)'] + weekly_df['RRC.AttConnRelDCCH.Norm (None)'] + weekly_df['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           weekly_df['RRC.AttConnRelCCCH.UsrInact (None)'] + weekly_df['RRC.AttConnRelCCCH.Unspec (None)'] + weekly_df['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           weekly_df['VS.DCCC.Succ.F2P (None)'] + weekly_df['IRATHO.SuccOutCS (None)'] + weekly_df['IRATHO.SuccOutPSUTRAN (None)'] + weekly_df['VS.DCCC.Succ.F2U (None)'] + \
                                           weekly_df['VS.DCCC.Succ.D2U (None)']) * 100
weekly_df['RAB Assignment Success Rate (CS), %'] = weekly_df['VS.RAB.SuccEstabCS.AMR (None)'] / weekly_df['VS.RAB.AttEstab.AMR (None)'] * 100
weekly_df['RAB Assignment Success Rate (PS), %'] = (weekly_df['VS.RAB.SuccEstabPS.Conv (None)'] + weekly_df['VS.RAB.SuccEstabPS.Bkg (None)'] + weekly_df['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   weekly_df['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (weekly_df['VS.RAB.AttEstabPS.Bkg (None)'] + weekly_df['VS.RAB.AttEstabPS.Int (None)'] + weekly_df['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   weekly_df['VS.RAB.AttEstabPS.Conv (None)']) * 100
weekly_df['CCSR3G, %'] = weekly_df['RRC Assignment SucessRate (CS BH), %'] * (100 - weekly_df['RRC Drop Rate (CS BH), %']) * weekly_df['RAB Assignment Success Rate (CS), %'] * (100 - weekly_df['CS RAB Drop Rate (%)'])/ 1000000
weekly_df['DCSR3G, %'] = weekly_df['RRC Assignment SucessRate (PS BH), %'] * (100 - weekly_df['RRC Drop Rate (PS BH), %']) * weekly_df['RAB Assignment Success Rate (PS), %'] * (100 - weekly_df['PS RAB Drop Rate (%)'])/ 1000000
weekly_df['UcellUnavail'] = weekly_df['VS.Cell.UnavailTime (s)'] + weekly_df['VS.Cell.UnavailTime.Sys (s)'] / 8600 / 24 / 7
weekly_df = weekly_df.drop(list_1, axis=1)
weekly_df_trans = weekly_df.transpose()

weeklyN_df = stsN_df.groupby(['week'])[list_1N]. sum().reset_index()
weeklyN_df['MeanThrHSDPA,kbps'] = weeklyN_df['VS.HSDPA.DataOutput.Traffic (bit)']/weeklyN_df['VS.HSDPA.DataTtiNum.User (None)'] / 2
weeklyN_df['MeanThrHSDPA DC,kbps'] = weeklyN_df['VS.DataOutput.AllHSDPA.Traffic (bit)'] / weeklyN_df['VS.AllHSDPA.DataTtiNum.User (None)'] / 2
weeklyN_df['MeanThrHSUPA,kbps'] = (weeklyN_df['VS.HSUPA.2msTTI.Traffic (kbit)'] + weeklyN_df['VS.HSUPA.10msTTI.Traffic (kbit)']) / (weeklyN_df['VS.HSUPA.2msPDU.TTI.Num (None)'] * 0.002 + weeklyN_df['VS.HSUPA.10msPDU.TTI.Num (None)'] * 0.01)
weeklyN_df = weeklyN_df.drop(list_1N, axis=1)
#weeklyN_df_trans = weeklyN_df.transpose()


# ===обработка daily===
daily_df = sts_df.groupby(['date'])[list_1]. sum().reset_index()
daily_df['CS traffic 3G, Erl'] = daily_df['CS Voice Traffic Volume (Erl)']
daily_df['PS traffic 3G UL+DL, GB'] = (daily_df['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + daily_df['VS.PS.Bkg.DL.8.Traffic (bit)'] + daily_df['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      daily_df['VS.PS.Bkg.DL.32.Traffic (bit)'] + daily_df['VS.PS.Bkg.DL.64.Traffic (bit)'] + daily_df['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      daily_df['VS.PS.Bkg.DL.144.Traffic (bit)'] + daily_df['VS.PS.Bkg.DL.256.Traffic (bit)'] + daily_df['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      daily_df['VS.PS.Bkg.UL.8.Traffic (bit)'] + daily_df['VS.PS.Bkg.UL.16.Traffic (bit)'] + daily_df['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      daily_df['VS.PS.Bkg.UL.64.Traffic (bit)'] + daily_df['VS.PS.Bkg.UL.128.Traffic (bit)'] + daily_df['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      daily_df['VS.PS.Bkg.UL.256.Traffic (bit)'] + daily_df['VS.PS.Bkg.UL.384.Traffic (bit)'] + daily_df['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      daily_df['VS.PS.Int.DL.16.Traffic (bit)'] + daily_df['VS.PS.Int.DL.32.Traffic (bit)'] + daily_df['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      daily_df['VS.PS.Int.DL.128.Traffic (bit)'] + daily_df['VS.PS.Int.DL.144.Traffic (bit)'] + daily_df['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      daily_df['VS.PS.Int.DL.384.Traffic (bit)'] + daily_df['VS.PS.Int.UL.8.Traffic (bit)'] + daily_df['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      daily_df['VS.PS.Int.UL.32.Traffic (bit)'] + daily_df['VS.PS.Int.UL.64.Traffic (bit)'] + daily_df['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      daily_df['VS.PS.Int.UL.144.Traffic (bit)'] + daily_df['VS.PS.Int.UL.256.Traffic (bit)'] + daily_df['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      daily_df['VS.PS.Str.DL.32.Traffic (bit)'] + daily_df['VS.PS.Str.DL.64.Traffic (bit)'] + daily_df['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      daily_df['VS.PS.Str.DL.144.Traffic (bit)'] + daily_df['VS.PS.Str.UL.16.Traffic (bit)'] + daily_df['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      daily_df['VS.PS.Str.UL.64.Traffic (bit)'] + daily_df['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
daily_df['CS RAB Drop Rate (%)'] = daily_df['VS.RAB.AbnormRel.CS (None)'] / (daily_df['VS.RAB.AbnormRel.CS (None)'] + daily_df['VS.RAB.NormRel.CS (None)']) * 100
daily_df['PS Blocking Rate (%)'] = (daily_df['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + daily_df['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + daily_df['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    daily_df['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + daily_df['VS.RAB.FailEstabPS.Code.Cong (None)'] + daily_df['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    daily_df['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + daily_df['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + daily_df['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (daily_df['VS.RAB.AttEstabPS.Str (None)'] + daily_df['VS.RAB.AttEstabPS.Int (None)'] + daily_df['VS.RAB.AttEstabPS.Bkg (None)']) *100
# daily_df['PS RAB Drop Rate (%)'] = (daily_df['VS.RAB.AbnormRel.PS (None)'] + daily_df['VS.RAB.AbnormRel.PS.PCH (None)'] + daily_df['VS.RAB.AbnormRel.PS.D2P (None)'] + \
#                                     daily_df['VS.RAB.AbnormRel.PS.F2P (None)']) / \
#                                    (daily_df['VS.RAB.AbnormRel.PS (None)'] + daily_df['VS.RAB.NormRel.PS (None)'] + daily_df['VS.RAB.AbnormRel.PS.PCH (None)'] + \
#                                     daily_df['VS.RAB.NormRel.PS.PCH (None)']) * 100
daily_df['PS RAB Drop Rate (%)'] = (daily_df['VS.RAB.AbnormRel.PS (None)'] - daily_df['VS.RAB.AbnormRel.PS.PCH (None)'] - daily_df['VS.RAB.AbnormRel.PS.D2P (None)'] - \
                                    daily_df['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (daily_df['VS.RAB.AbnormRel.PS (None)'] + daily_df['VS.RAB.NormRel.PS (None)'] - daily_df['VS.RAB.AbnormRel.PS.PCH (None)'] - \
                                    daily_df['VS.RAB.NormRel.PS.PCH (None)'] + daily_df['VS.DCCC.D2P.Succ (None)']+daily_df['VS.DCCC.Succ.F2P (None)']+daily_df['VS.DCCC.Succ.F2U (None)']+daily_df['VS.DCCC.Succ.D2U (None)']) * 100

daily_df['PS HS- Drop Rate (%)'] =  daily_df['VS.HSDPA.RAB.AbnormRel (None)'] / (daily_df['VS.HSDPA.RAB.AbnormRel (None)'] + daily_df['VS.HSDPA.RAB.NormRel (None)'] + daily_df['VS.HSDPA.H2D.Succ (None)'] + daily_df['VS.HSDPA.H2F.Succ (None)'] +daily_df['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + daily_df['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
daily_df['HSDPA Throughput, kbps'] = daily_df['VS.HSDPA.MeanChThroughput (kbit/s)'] / 538 / 24 # количество сот 538
daily_df['HSUPA Throughput, kbps'] = daily_df['VS.HSUPA.MeanChThroughput (kbit/s)'] / 538 / 24# количество сот 538
daily_df['Soft Handover Success rate, %'] = (daily_df['VS.SHO.SuccRLAdd (None)'] + daily_df['VS.SHO.SuccRLDel (None)']) / (daily_df['VS.SHO.AttRLAdd (None)'] + daily_df['VS.SHO.AttRLDel (None)']) * 100
daily_df['Hard Handover Success rate, %'] = daily_df['VS.HHO.SuccInterFreqOut (None)'] / daily_df['VS.HHO.AttInterFreqOut (None)'] * 100
daily_df['CS W2G Inter-RAT Handover Out SR'] = daily_df['IRATHO.SuccOutCS (None)'] / (daily_df['IRATHO.AttOutCS (None)'] - daily_df['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
daily_df['RRC Assignment SucessRate (CS BH), %'] = daily_df['RRC.SuccConnEstab.sum (None)'] / daily_df['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_df['RRC Assignment SucessRate (PS BH), %'] = daily_df['RRC.SuccConnEstab.sum (None)'] / daily_df['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_df['RRC Drop Rate (CS BH), %'] = (daily_df['RRC.AttConnRelCCCH.Cong (None)'] + daily_df['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_df['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df['RRC.AttConnRelDCCH.Cong (None)'] + daily_df['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_df['RRC.AttConnRelDCCH.Unspec (None)'] + daily_df['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_df['RRC.AttConnRelDCCH.Cong (None)'] + daily_df['RRC.AttConnRelDCCH.Preempt (None)'] + daily_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_df['RRC.AttConnRelDCCH.DSCR (None)'] + daily_df['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_df['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_df['RRC.AttConnRelCCCH.Cong (None)'] + daily_df['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_df['RRC.AttConnRelCCCH.DSCR (None)'] + daily_df['RRC.AttConnRelDCCH.Norm (None)'] + daily_df['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_df['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_df['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_df['VS.DCCC.Succ.F2P (None)'] + daily_df['IRATHO.SuccOutCS (None)'] + daily_df['IRATHO.SuccOutPSUTRAN (None)'] + daily_df['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_df['VS.DCCC.Succ.D2U (None)']) * 100
daily_df['RRC Drop Rate (PS BH), %'] = (daily_df['RRC.AttConnRelCCCH.Cong (None)'] + daily_df['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_df['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df['RRC.AttConnRelDCCH.Cong (None)'] + daily_df['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_df['RRC.AttConnRelDCCH.Unspec (None)'] + daily_df['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_df['RRC.AttConnRelDCCH.Cong (None)'] + daily_df['RRC.AttConnRelDCCH.Preempt (None)'] + daily_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_df['RRC.AttConnRelDCCH.DSCR (None)'] + daily_df['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_df['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_df['RRC.AttConnRelCCCH.Cong (None)'] + daily_df['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_df['RRC.AttConnRelCCCH.DSCR (None)'] + daily_df['RRC.AttConnRelDCCH.Norm (None)'] + daily_df['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_df['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_df['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_df['VS.DCCC.Succ.F2P (None)'] + daily_df['IRATHO.SuccOutCS (None)'] + daily_df['IRATHO.SuccOutPSUTRAN (None)'] + daily_df['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_df['VS.DCCC.Succ.D2U (None)']) * 100
daily_df['RAB Assignment Success Rate (CS), %'] = daily_df['VS.RAB.SuccEstabCS.AMR (None)'] / daily_df['VS.RAB.AttEstab.AMR (None)'] * 100
daily_df['RAB Assignment Success Rate (PS), %'] = (daily_df['VS.RAB.SuccEstabPS.Conv (None)'] + daily_df['VS.RAB.SuccEstabPS.Bkg (None)'] + daily_df['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   daily_df['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (daily_df['VS.RAB.AttEstabPS.Bkg (None)'] + daily_df['VS.RAB.AttEstabPS.Int (None)'] + daily_df['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   daily_df['VS.RAB.AttEstabPS.Conv (None)']) * 100
daily_df['CCSR3G, %'] = daily_df['RRC Assignment SucessRate (CS BH), %'] * (100 - daily_df['RRC Drop Rate (CS BH), %']) * daily_df['RAB Assignment Success Rate (CS), %'] * (100 - daily_df['CS RAB Drop Rate (%)'])/ 1000000
daily_df['DCSR3G, %'] = daily_df['RRC Assignment SucessRate (PS BH), %'] * (100 - daily_df['RRC Drop Rate (PS BH), %']) * daily_df['RAB Assignment Success Rate (PS), %'] * (100 - daily_df['PS RAB Drop Rate (%)'])/ 1000000
daily_df = daily_df.drop(list_1, axis=1)

# фильтрация по U2100
daily_dfU2100 = sts_df[sts_df['BSC6910UCell'].isin(list_U2100)]
daily_dfU2100 = daily_dfU2100.groupby(['date'])[list_1]. sum().reset_index()
daily_dfU2100['CS traffic 3G, Erl_U2100'] = daily_dfU2100['CS Voice Traffic Volume (Erl)']
daily_dfU2100['PS traffic 3G UL+DL, GB_U2100'] = (daily_dfU2100['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + daily_dfU2100['VS.PS.Bkg.DL.8.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Bkg.DL.32.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.DL.64.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Bkg.DL.144.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.DL.256.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Bkg.UL.8.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.UL.16.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Bkg.UL.64.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.UL.128.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Bkg.UL.256.Traffic (bit)'] + daily_dfU2100['VS.PS.Bkg.UL.384.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Int.DL.16.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.DL.32.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Int.DL.128.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.DL.144.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Int.DL.384.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.UL.8.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Int.UL.32.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.UL.64.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Int.UL.144.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.UL.256.Traffic (bit)'] + daily_dfU2100['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Str.DL.32.Traffic (bit)'] + daily_dfU2100['VS.PS.Str.DL.64.Traffic (bit)'] + daily_dfU2100['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Str.DL.144.Traffic (bit)'] + daily_dfU2100['VS.PS.Str.UL.16.Traffic (bit)'] + daily_dfU2100['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      daily_dfU2100['VS.PS.Str.UL.64.Traffic (bit)'] + daily_dfU2100['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
daily_dfU2100['CS RAB Drop Rate (%)_U2100'] = daily_dfU2100['VS.RAB.AbnormRel.CS (None)'] / (daily_dfU2100['VS.RAB.AbnormRel.CS (None)'] + daily_dfU2100['VS.RAB.NormRel.CS (None)']) * 100
daily_dfU2100['PS Blocking Rate (%)_U2100'] = (daily_dfU2100['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + daily_dfU2100['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + daily_dfU2100['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    daily_dfU2100['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + daily_dfU2100['VS.RAB.FailEstabPS.Code.Cong (None)'] + daily_dfU2100['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    daily_dfU2100['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + daily_dfU2100['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + daily_dfU2100['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (daily_dfU2100['VS.RAB.AttEstabPS.Str (None)'] + daily_dfU2100['VS.RAB.AttEstabPS.Int (None)'] + daily_dfU2100['VS.RAB.AttEstabPS.Bkg (None)']) *100
# daily_dfU2100['PS RAB Drop Rate (%)_U2100'] = (daily_dfU2100['VS.RAB.AbnormRel.PS (None)'] + daily_dfU2100['VS.RAB.AbnormRel.PS.PCH (None)'] + daily_dfU2100['VS.RAB.AbnormRel.PS.D2P (None)'] + \
#                                     daily_dfU2100['VS.RAB.AbnormRel.PS.F2P (None)']) / \
#                                    (daily_dfU2100['VS.RAB.AbnormRel.PS (None)'] + daily_dfU2100['VS.RAB.NormRel.PS (None)'] + daily_dfU2100['VS.RAB.AbnormRel.PS.PCH (None)'] + \
#                                     daily_dfU2100['VS.RAB.NormRel.PS.PCH (None)']) * 100

daily_dfU2100['PS RAB Drop Rate (%)_U2100'] = (daily_dfU2100['VS.RAB.AbnormRel.PS (None)'] - daily_dfU2100['VS.RAB.AbnormRel.PS.PCH (None)'] - daily_dfU2100['VS.RAB.AbnormRel.PS.D2P (None)'] - \
                                    daily_dfU2100['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (daily_dfU2100['VS.RAB.AbnormRel.PS (None)'] + daily_dfU2100['VS.RAB.NormRel.PS (None)'] - daily_dfU2100['VS.RAB.AbnormRel.PS.PCH (None)'] - \
                                    daily_dfU2100['VS.RAB.NormRel.PS.PCH (None)'] + daily_dfU2100['VS.DCCC.D2P.Succ (None)']+daily_dfU2100['VS.DCCC.Succ.F2P (None)']+daily_dfU2100['VS.DCCC.Succ.F2U (None)']+daily_dfU2100['VS.DCCC.Succ.D2U (None)']) * 100


daily_dfU2100['PS HS- Drop Rate (%)_U2100'] =  daily_dfU2100['VS.HSDPA.RAB.AbnormRel (None)'] / (daily_dfU2100['VS.HSDPA.RAB.AbnormRel (None)'] + daily_dfU2100['VS.HSDPA.RAB.NormRel (None)'] + daily_dfU2100['VS.HSDPA.H2D.Succ (None)'] + daily_dfU2100['VS.HSDPA.H2F.Succ (None)'] +daily_dfU2100['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + daily_dfU2100['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
daily_dfU2100['HSDPA Throughput, kbps_U2100'] = daily_dfU2100['VS.HSDPA.MeanChThroughput (kbit/s)'] / 471 / 24 # количество сот 471!!!
daily_dfU2100['HSUPA Throughput, kbps_U2100'] = daily_dfU2100['VS.HSUPA.MeanChThroughput (kbit/s)'] / 471 / 24# количество сот 471!!!
daily_dfU2100['Soft Handover Success rate, %_U2100'] = (daily_dfU2100['VS.SHO.SuccRLAdd (None)'] + daily_dfU2100['VS.SHO.SuccRLDel (None)']) / (daily_dfU2100['VS.SHO.AttRLAdd (None)'] + daily_dfU2100['VS.SHO.AttRLDel (None)']) * 100
daily_dfU2100['Hard Handover Success rate, %_U2100'] = daily_dfU2100['VS.HHO.SuccInterFreqOut (None)'] / daily_dfU2100['VS.HHO.AttInterFreqOut (None)'] * 100
daily_dfU2100['CS W2G Inter-RAT Handover Out SR_U2100'] = daily_dfU2100['IRATHO.SuccOutCS (None)'] / (daily_dfU2100['IRATHO.AttOutCS (None)'] - daily_dfU2100['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
daily_dfU2100['RRC Assignment SucessRate (CS BH), %_U2100'] = daily_dfU2100['RRC.SuccConnEstab.sum (None)'] / daily_dfU2100['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_dfU2100['RRC Assignment SucessRate (PS BH), %_U2100'] = daily_dfU2100['RRC.SuccConnEstab.sum (None)'] / daily_dfU2100['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_dfU2100['RRC Drop Rate (CS BH), %_U2100'] = (daily_dfU2100['RRC.AttConnRelCCCH.Cong (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.Preempt (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_dfU2100['RRC.AttConnRelCCCH.Unspec (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Cong (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_dfU2100['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Unspec (None)'] + daily_dfU2100['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_dfU2100['RRC.AttConnRelDCCH.Cong (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Preempt (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_dfU2100['RRC.AttConnRelDCCH.DSCR (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_dfU2100['RRC.AttConnRelCCCH.Cong (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.Preempt (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_dfU2100['RRC.AttConnRelCCCH.DSCR (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Norm (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_dfU2100['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.Unspec (None)'] + daily_dfU2100['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_dfU2100['VS.DCCC.Succ.F2P (None)'] + daily_dfU2100['IRATHO.SuccOutCS (None)'] + daily_dfU2100['IRATHO.SuccOutPSUTRAN (None)'] + daily_dfU2100['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_dfU2100['VS.DCCC.Succ.D2U (None)']) * 100
daily_dfU2100['RRC Drop Rate (PS BH), %_U2100'] = (daily_dfU2100['RRC.AttConnRelCCCH.Cong (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.Preempt (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_dfU2100['RRC.AttConnRelCCCH.Unspec (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Cong (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_dfU2100['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Unspec (None)'] + daily_dfU2100['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_dfU2100['RRC.AttConnRelDCCH.Cong (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Preempt (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_dfU2100['RRC.AttConnRelDCCH.DSCR (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_dfU2100['RRC.AttConnRelCCCH.Cong (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.Preempt (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_dfU2100['RRC.AttConnRelCCCH.DSCR (None)'] + daily_dfU2100['RRC.AttConnRelDCCH.Norm (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_dfU2100['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_dfU2100['RRC.AttConnRelCCCH.Unspec (None)'] + daily_dfU2100['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_dfU2100['VS.DCCC.Succ.F2P (None)'] + daily_dfU2100['IRATHO.SuccOutCS (None)'] + daily_dfU2100['IRATHO.SuccOutPSUTRAN (None)'] + daily_dfU2100['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_dfU2100['VS.DCCC.Succ.D2U (None)']) * 100
daily_dfU2100['RAB Assignment Success Rate (CS), %_U2100'] = daily_dfU2100['VS.RAB.SuccEstabCS.AMR (None)'] / daily_dfU2100['VS.RAB.AttEstab.AMR (None)'] * 100
daily_dfU2100['RAB Assignment Success Rate (PS), %_U2100'] = (daily_dfU2100['VS.RAB.SuccEstabPS.Conv (None)'] + daily_dfU2100['VS.RAB.SuccEstabPS.Bkg (None)'] + daily_dfU2100['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   daily_dfU2100['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (daily_dfU2100['VS.RAB.AttEstabPS.Bkg (None)'] + daily_dfU2100['VS.RAB.AttEstabPS.Int (None)'] + daily_dfU2100['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   daily_dfU2100['VS.RAB.AttEstabPS.Conv (None)']) * 100
daily_dfU2100['CCSR3G, %_U2100'] = daily_dfU2100['RRC Assignment SucessRate (CS BH), %_U2100'] * (100 - daily_dfU2100['RRC Drop Rate (CS BH), %_U2100']) * daily_dfU2100['RAB Assignment Success Rate (CS), %_U2100'] * (100 - daily_dfU2100['CS RAB Drop Rate (%)_U2100'])/ 1000000
daily_dfU2100['DCSR3G, %_U2100'] = daily_dfU2100['RRC Assignment SucessRate (PS BH), %_U2100'] * (100 - daily_dfU2100['RRC Drop Rate (PS BH), %_U2100']) * daily_dfU2100['RAB Assignment Success Rate (PS), %_U2100'] * (100 - daily_dfU2100['PS RAB Drop Rate (%)_U2100'])/ 1000000
daily_dfU2100 = daily_dfU2100.drop(list_1, axis=1)

# фильтрация по U900
daily_dfU900 = sts_df[sts_df['BSC6910UCell'].isin(list_U900)]
daily_dfU900 = daily_dfU900.groupby(['date'])[list_1]. sum().reset_index()
daily_dfU900['CS traffic 3G, Erl_U900'] = daily_dfU900['CS Voice Traffic Volume (Erl)']
daily_dfU900['PS traffic 3G UL+DL, GB_U900'] = (daily_dfU900['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + daily_dfU900['VS.PS.Bkg.DL.8.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Bkg.DL.32.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.DL.64.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Bkg.DL.144.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.DL.256.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Bkg.UL.8.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.UL.16.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Bkg.UL.64.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.UL.128.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Bkg.UL.256.Traffic (bit)'] + daily_dfU900['VS.PS.Bkg.UL.384.Traffic (bit)'] + daily_dfU900['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Int.DL.16.Traffic (bit)'] + daily_dfU900['VS.PS.Int.DL.32.Traffic (bit)'] + daily_dfU900['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Int.DL.128.Traffic (bit)'] + daily_dfU900['VS.PS.Int.DL.144.Traffic (bit)'] + daily_dfU900['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Int.DL.384.Traffic (bit)'] + daily_dfU900['VS.PS.Int.UL.8.Traffic (bit)'] + daily_dfU900['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Int.UL.32.Traffic (bit)'] + daily_dfU900['VS.PS.Int.UL.64.Traffic (bit)'] + daily_dfU900['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Int.UL.144.Traffic (bit)'] + daily_dfU900['VS.PS.Int.UL.256.Traffic (bit)'] + daily_dfU900['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Str.DL.32.Traffic (bit)'] + daily_dfU900['VS.PS.Str.DL.64.Traffic (bit)'] + daily_dfU900['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Str.DL.144.Traffic (bit)'] + daily_dfU900['VS.PS.Str.UL.16.Traffic (bit)'] + daily_dfU900['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      daily_dfU900['VS.PS.Str.UL.64.Traffic (bit)'] + daily_dfU900['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
daily_dfU900['CS RAB Drop Rate (%)_U900'] = daily_dfU900['VS.RAB.AbnormRel.CS (None)'] / (daily_dfU900['VS.RAB.AbnormRel.CS (None)'] + daily_dfU900['VS.RAB.NormRel.CS (None)']) * 100
daily_dfU900['PS Blocking Rate (%)_U900'] = (daily_dfU900['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + daily_dfU900['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + daily_dfU900['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    daily_dfU900['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + daily_dfU900['VS.RAB.FailEstabPS.Code.Cong (None)'] + daily_dfU900['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    daily_dfU900['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + daily_dfU900['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + daily_dfU900['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (daily_dfU900['VS.RAB.AttEstabPS.Str (None)'] + daily_dfU900['VS.RAB.AttEstabPS.Int (None)'] + daily_dfU900['VS.RAB.AttEstabPS.Bkg (None)']) *100
# daily_dfU900['PS RAB Drop Rate (%)_U900'] = (daily_dfU900['VS.RAB.AbnormRel.PS (None)'] + daily_dfU900['VS.RAB.AbnormRel.PS.PCH (None)'] + daily_dfU900['VS.RAB.AbnormRel.PS.D2P (None)'] + \
#                                     daily_dfU900['VS.RAB.AbnormRel.PS.F2P (None)']) / \
#                                    (daily_dfU900['VS.RAB.AbnormRel.PS (None)'] + daily_dfU900['VS.RAB.NormRel.PS (None)'] + daily_dfU900['VS.RAB.AbnormRel.PS.PCH (None)'] + \
#                                     daily_dfU900['VS.RAB.NormRel.PS.PCH (None)']) * 100
daily_dfU900['PS RAB Drop Rate (%)_U900'] = (daily_dfU900['VS.RAB.AbnormRel.PS (None)'] - daily_dfU900['VS.RAB.AbnormRel.PS.PCH (None)'] - daily_dfU900['VS.RAB.AbnormRel.PS.D2P (None)'] - \
                                    daily_dfU900['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (daily_dfU900['VS.RAB.AbnormRel.PS (None)'] + daily_dfU900['VS.RAB.NormRel.PS (None)'] - daily_dfU900['VS.RAB.AbnormRel.PS.PCH (None)'] - \
                                    daily_dfU900['VS.RAB.NormRel.PS.PCH (None)'] + daily_dfU900['VS.DCCC.D2P.Succ (None)']+daily_dfU900['VS.DCCC.Succ.F2P (None)']+daily_dfU900['VS.DCCC.Succ.F2U (None)']+daily_dfU900['VS.DCCC.Succ.D2U (None)']) * 100

daily_dfU900['PS HS- Drop Rate (%)_U900'] =  daily_dfU900['VS.HSDPA.RAB.AbnormRel (None)'] / (daily_dfU900['VS.HSDPA.RAB.AbnormRel (None)'] + daily_dfU900['VS.HSDPA.RAB.NormRel (None)'] + daily_dfU900['VS.HSDPA.H2D.Succ (None)'] + daily_dfU900['VS.HSDPA.H2F.Succ (None)'] +daily_dfU900['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + daily_dfU900['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
daily_dfU900['HSDPA Throughput, kbps_U900'] = daily_dfU900['VS.HSDPA.MeanChThroughput (kbit/s)'] / 216 / 24 # количество сот 216!!!
daily_dfU900['HSUPA Throughput, kbps_U900'] = daily_dfU900['VS.HSUPA.MeanChThroughput (kbit/s)'] / 216 / 24# количество сот 216
daily_dfU900['Soft Handover Success rate, %_U900'] = (daily_dfU900['VS.SHO.SuccRLAdd (None)'] + daily_dfU900['VS.SHO.SuccRLDel (None)']) / (daily_dfU900['VS.SHO.AttRLAdd (None)'] + daily_dfU900['VS.SHO.AttRLDel (None)']) * 100
daily_dfU900['Hard Handover Success rate, %_U900'] = daily_dfU900['VS.HHO.SuccInterFreqOut (None)'] / daily_dfU900['VS.HHO.AttInterFreqOut (None)'] * 100
daily_dfU900['CS W2G Inter-RAT Handover Out SR_U900'] = daily_dfU900['IRATHO.SuccOutCS (None)'] / (daily_dfU900['IRATHO.AttOutCS (None)'] - daily_dfU900['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
daily_dfU900['RRC Assignment SucessRate (CS BH), %_U900'] = daily_dfU900['RRC.SuccConnEstab.sum (None)'] / daily_dfU900['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_dfU900['RRC Assignment SucessRate (PS BH), %_U900'] = daily_dfU900['RRC.SuccConnEstab.sum (None)'] / daily_dfU900['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_dfU900['RRC Drop Rate (CS BH), %_U900'] = (daily_dfU900['RRC.AttConnRelCCCH.Cong (None)'] + daily_dfU900['RRC.AttConnRelCCCH.Preempt (None)'] + daily_dfU900['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_dfU900['RRC.AttConnRelCCCH.Unspec (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Cong (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_dfU900['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Unspec (None)'] + daily_dfU900['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_dfU900['RRC.AttConnRelDCCH.Cong (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Preempt (None)'] + daily_dfU900['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_dfU900['RRC.AttConnRelDCCH.DSCR (None)'] + daily_dfU900['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_dfU900['RRC.AttConnRelCCCH.Cong (None)'] + daily_dfU900['RRC.AttConnRelCCCH.Preempt (None)'] + daily_dfU900['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_dfU900['RRC.AttConnRelCCCH.DSCR (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Norm (None)'] + daily_dfU900['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_dfU900['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_dfU900['RRC.AttConnRelCCCH.Unspec (None)'] + daily_dfU900['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_dfU900['VS.DCCC.Succ.F2P (None)'] + daily_dfU900['IRATHO.SuccOutCS (None)'] + daily_dfU900['IRATHO.SuccOutPSUTRAN (None)'] + daily_dfU900['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_dfU900['VS.DCCC.Succ.D2U (None)']) * 100
daily_dfU900['RRC Drop Rate (PS BH), %_U900'] = (daily_dfU900['RRC.AttConnRelCCCH.Cong (None)'] + daily_dfU900['RRC.AttConnRelCCCH.Preempt (None)'] + daily_dfU900['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_dfU900['RRC.AttConnRelCCCH.Unspec (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Cong (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_dfU900['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Unspec (None)'] + daily_dfU900['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_dfU900['RRC.AttConnRelDCCH.Cong (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Preempt (None)'] + daily_dfU900['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_dfU900['RRC.AttConnRelDCCH.DSCR (None)'] + daily_dfU900['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_dfU900['RRC.AttConnRelCCCH.Cong (None)'] + daily_dfU900['RRC.AttConnRelCCCH.Preempt (None)'] + daily_dfU900['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_dfU900['RRC.AttConnRelCCCH.DSCR (None)'] + daily_dfU900['RRC.AttConnRelDCCH.Norm (None)'] + daily_dfU900['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_dfU900['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_dfU900['RRC.AttConnRelCCCH.Unspec (None)'] + daily_dfU900['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_dfU900['VS.DCCC.Succ.F2P (None)'] + daily_dfU900['IRATHO.SuccOutCS (None)'] + daily_dfU900['IRATHO.SuccOutPSUTRAN (None)'] + daily_dfU900['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_dfU900['VS.DCCC.Succ.D2U (None)']) * 100
daily_dfU900['RAB Assignment Success Rate (CS), %_U900'] = daily_dfU900['VS.RAB.SuccEstabCS.AMR (None)'] / daily_dfU900['VS.RAB.AttEstab.AMR (None)'] * 100
daily_dfU900['RAB Assignment Success Rate (PS), %_U900'] = (daily_dfU900['VS.RAB.SuccEstabPS.Conv (None)'] + daily_dfU900['VS.RAB.SuccEstabPS.Bkg (None)'] + daily_dfU900['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   daily_dfU900['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (daily_dfU900['VS.RAB.AttEstabPS.Bkg (None)'] + daily_dfU900['VS.RAB.AttEstabPS.Int (None)'] + daily_dfU900['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   daily_dfU900['VS.RAB.AttEstabPS.Conv (None)']) * 100
daily_dfU900['CCSR3G, %_U900'] = daily_dfU900['RRC Assignment SucessRate (CS BH), %_U900'] * (100 - daily_dfU900['RRC Drop Rate (CS BH), %_U900']) * daily_dfU900['RAB Assignment Success Rate (CS), %_U900'] * (100 - daily_dfU900['CS RAB Drop Rate (%)_U900'])/ 1000000
daily_dfU900['DCSR3G, %_U900'] = daily_dfU900['RRC Assignment SucessRate (PS BH), %_U900'] * (100 - daily_dfU900['RRC Drop Rate (PS BH), %_U900']) * daily_dfU900['RAB Assignment Success Rate (PS), %_U900'] * (100 - daily_dfU900['PS RAB Drop Rate (%)_U900'])/ 1000000
daily_dfU900 = daily_dfU900.drop(list_1, axis=1)
# фильтрация по 10612
daily_df10612 = sts_df[sts_df['BSC6910UCell'].isin(list_F1_10612)]
daily_df10612 = daily_df10612.groupby(['date'])[list_1]. sum().reset_index()
daily_df10612['CS traffic 3G, Erl_10612'] = daily_df10612['CS Voice Traffic Volume (Erl)']
daily_df10612['PS traffic 3G UL+DL, GB_10612'] = (daily_df10612['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + daily_df10612['VS.PS.Bkg.DL.8.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Bkg.DL.32.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.DL.64.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Bkg.DL.144.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.DL.256.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Bkg.UL.8.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.UL.16.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Bkg.UL.64.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.UL.128.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Bkg.UL.256.Traffic (bit)'] + daily_df10612['VS.PS.Bkg.UL.384.Traffic (bit)'] + daily_df10612['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Int.DL.16.Traffic (bit)'] + daily_df10612['VS.PS.Int.DL.32.Traffic (bit)'] + daily_df10612['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Int.DL.128.Traffic (bit)'] + daily_df10612['VS.PS.Int.DL.144.Traffic (bit)'] + daily_df10612['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Int.DL.384.Traffic (bit)'] + daily_df10612['VS.PS.Int.UL.8.Traffic (bit)'] + daily_df10612['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Int.UL.32.Traffic (bit)'] + daily_df10612['VS.PS.Int.UL.64.Traffic (bit)'] + daily_df10612['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Int.UL.144.Traffic (bit)'] + daily_df10612['VS.PS.Int.UL.256.Traffic (bit)'] + daily_df10612['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Str.DL.32.Traffic (bit)'] + daily_df10612['VS.PS.Str.DL.64.Traffic (bit)'] + daily_df10612['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Str.DL.144.Traffic (bit)'] + daily_df10612['VS.PS.Str.UL.16.Traffic (bit)'] + daily_df10612['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      daily_df10612['VS.PS.Str.UL.64.Traffic (bit)'] + daily_df10612['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
daily_df10612['CS RAB Drop Rate (%)_10612'] = daily_df10612['VS.RAB.AbnormRel.CS (None)'] / (daily_df10612['VS.RAB.AbnormRel.CS (None)'] + daily_df10612['VS.RAB.NormRel.CS (None)']) * 100
daily_df10612['PS Blocking Rate (%)_10612'] = (daily_df10612['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + daily_df10612['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + daily_df10612['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    daily_df10612['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + daily_df10612['VS.RAB.FailEstabPS.Code.Cong (None)'] + daily_df10612['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    daily_df10612['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + daily_df10612['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + daily_df10612['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (daily_df10612['VS.RAB.AttEstabPS.Str (None)'] + daily_df10612['VS.RAB.AttEstabPS.Int (None)'] + daily_df10612['VS.RAB.AttEstabPS.Bkg (None)']) *100
daily_df10612['PS RAB Drop Rate (%)_10612'] = (daily_df10612['VS.RAB.AbnormRel.PS (None)'] + daily_df10612['VS.RAB.AbnormRel.PS.PCH (None)'] + daily_df10612['VS.RAB.AbnormRel.PS.D2P (None)'] + \
                                    daily_df10612['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (daily_df10612['VS.RAB.AbnormRel.PS (None)'] + daily_df10612['VS.RAB.NormRel.PS (None)'] + daily_df10612['VS.RAB.AbnormRel.PS.PCH (None)'] + \
                                    daily_df10612['VS.RAB.NormRel.PS.PCH (None)']) * 100
daily_df10612['PS HS- Drop Rate (%)_10612'] =  daily_df10612['VS.HSDPA.RAB.AbnormRel (None)'] / (daily_df10612['VS.HSDPA.RAB.AbnormRel (None)'] + daily_df10612['VS.HSDPA.RAB.NormRel (None)'] + daily_df10612['VS.HSDPA.H2D.Succ (None)'] + daily_df10612['VS.HSDPA.H2F.Succ (None)'] +daily_df10612['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + daily_df10612['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
daily_df10612['HSDPA Throughput, kbps_10612'] = daily_df10612['VS.HSDPA.MeanChThroughput (kbit/s)'] / 235 / 24 # количество сот 235!!!
daily_df10612['HSUPA Throughput, kbps_10612'] = daily_df10612['VS.HSUPA.MeanChThroughput (kbit/s)'] / 235 / 24# количество сот 235
daily_df10612['Soft Handover Success rate, %_10612'] = (daily_df10612['VS.SHO.SuccRLAdd (None)'] + daily_df10612['VS.SHO.SuccRLDel (None)']) / (daily_df10612['VS.SHO.AttRLAdd (None)'] + daily_df10612['VS.SHO.AttRLDel (None)']) * 100
daily_df10612['Hard Handover Success rate, %_10612'] = daily_df10612['VS.HHO.SuccInterFreqOut (None)'] / daily_df10612['VS.HHO.AttInterFreqOut (None)'] * 100
daily_df10612['CS W2G Inter-RAT Handover Out SR_10612'] = daily_df10612['IRATHO.SuccOutCS (None)'] / (daily_df10612['IRATHO.AttOutCS (None)'] - daily_df10612['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
daily_df10612['RRC Assignment SucessRate (CS BH), %_10612'] = daily_df10612['RRC.SuccConnEstab.sum (None)'] / daily_df10612['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_df10612['RRC Assignment SucessRate (PS BH), %_10612'] = daily_df10612['RRC.SuccConnEstab.sum (None)'] / daily_df10612['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_df10612['RRC Drop Rate (CS BH), %_10612'] = (daily_df10612['RRC.AttConnRelCCCH.Cong (None)'] + daily_df10612['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df10612['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_df10612['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df10612['RRC.AttConnRelDCCH.Cong (None)'] + daily_df10612['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_df10612['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_df10612['RRC.AttConnRelDCCH.Unspec (None)'] + daily_df10612['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_df10612['RRC.AttConnRelDCCH.Cong (None)'] + daily_df10612['RRC.AttConnRelDCCH.Preempt (None)'] + daily_df10612['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_df10612['RRC.AttConnRelDCCH.DSCR (None)'] + daily_df10612['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_df10612['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_df10612['RRC.AttConnRelCCCH.Cong (None)'] + daily_df10612['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df10612['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_df10612['RRC.AttConnRelCCCH.DSCR (None)'] + daily_df10612['RRC.AttConnRelDCCH.Norm (None)'] + daily_df10612['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_df10612['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_df10612['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df10612['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_df10612['VS.DCCC.Succ.F2P (None)'] + daily_df10612['IRATHO.SuccOutCS (None)'] + daily_df10612['IRATHO.SuccOutPSUTRAN (None)'] + daily_df10612['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_df10612['VS.DCCC.Succ.D2U (None)']) * 100
daily_df10612['RRC Drop Rate (PS BH), %_10612'] = (daily_df10612['RRC.AttConnRelCCCH.Cong (None)'] + daily_df10612['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df10612['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_df10612['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df10612['RRC.AttConnRelDCCH.Cong (None)'] + daily_df10612['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_df10612['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_df10612['RRC.AttConnRelDCCH.Unspec (None)'] + daily_df10612['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_df10612['RRC.AttConnRelDCCH.Cong (None)'] + daily_df10612['RRC.AttConnRelDCCH.Preempt (None)'] + daily_df10612['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_df10612['RRC.AttConnRelDCCH.DSCR (None)'] + daily_df10612['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_df10612['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_df10612['RRC.AttConnRelCCCH.Cong (None)'] + daily_df10612['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df10612['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_df10612['RRC.AttConnRelCCCH.DSCR (None)'] + daily_df10612['RRC.AttConnRelDCCH.Norm (None)'] + daily_df10612['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_df10612['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_df10612['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df10612['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_df10612['VS.DCCC.Succ.F2P (None)'] + daily_df10612['IRATHO.SuccOutCS (None)'] + daily_df10612['IRATHO.SuccOutPSUTRAN (None)'] + daily_df10612['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_df10612['VS.DCCC.Succ.D2U (None)']) * 100
daily_df10612['RAB Assignment Success Rate (CS), %_10612'] = daily_df10612['VS.RAB.SuccEstabCS.AMR (None)'] / daily_df10612['VS.RAB.AttEstab.AMR (None)'] * 100
daily_df10612['RAB Assignment Success Rate (PS), %_10612'] = (daily_df10612['VS.RAB.SuccEstabPS.Conv (None)'] + daily_df10612['VS.RAB.SuccEstabPS.Bkg (None)'] + daily_df10612['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   daily_df10612['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (daily_df10612['VS.RAB.AttEstabPS.Bkg (None)'] + daily_df10612['VS.RAB.AttEstabPS.Int (None)'] + daily_df10612['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   daily_df10612['VS.RAB.AttEstabPS.Conv (None)']) * 100
daily_df10612['CCSR3G, %_10612'] = daily_df10612['RRC Assignment SucessRate (CS BH), %_10612'] * (100 - daily_df10612['RRC Drop Rate (CS BH), %_10612']) * daily_df10612['RAB Assignment Success Rate (CS), %_10612'] * (100 - daily_df10612['CS RAB Drop Rate (%)_10612'])/ 1000000
daily_df10612['DCSR3G, %_10612'] = daily_df10612['RRC Assignment SucessRate (PS BH), %_10612'] * (100 - daily_df10612['RRC Drop Rate (PS BH), %_10612']) * daily_df10612['RAB Assignment Success Rate (PS), %_10612'] * (100 - daily_df10612['PS RAB Drop Rate (%)_10612'])/ 1000000
daily_df10612 = daily_df10612.drop(list_1, axis=1)
# фильтрация по 10637
daily_df10637 = sts_df[sts_df['BSC6910UCell'].isin(list_F2_10637)]
daily_df10637 = daily_df10637.groupby(['date'])[list_1]. sum().reset_index()
daily_df10637['CS traffic 3G, Erl_10637'] = daily_df10637['CS Voice Traffic Volume (Erl)']
daily_df10637['PS traffic 3G UL+DL, GB_10637'] = (daily_df10637['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + daily_df10637['VS.PS.Bkg.DL.8.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Bkg.DL.32.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.DL.64.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Bkg.DL.144.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.DL.256.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Bkg.UL.8.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.UL.16.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Bkg.UL.64.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.UL.128.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Bkg.UL.256.Traffic (bit)'] + daily_df10637['VS.PS.Bkg.UL.384.Traffic (bit)'] + daily_df10637['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Int.DL.16.Traffic (bit)'] + daily_df10637['VS.PS.Int.DL.32.Traffic (bit)'] + daily_df10637['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Int.DL.128.Traffic (bit)'] + daily_df10637['VS.PS.Int.DL.144.Traffic (bit)'] + daily_df10637['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Int.DL.384.Traffic (bit)'] + daily_df10637['VS.PS.Int.UL.8.Traffic (bit)'] + daily_df10637['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Int.UL.32.Traffic (bit)'] + daily_df10637['VS.PS.Int.UL.64.Traffic (bit)'] + daily_df10637['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Int.UL.144.Traffic (bit)'] + daily_df10637['VS.PS.Int.UL.256.Traffic (bit)'] + daily_df10637['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Str.DL.32.Traffic (bit)'] + daily_df10637['VS.PS.Str.DL.64.Traffic (bit)'] + daily_df10637['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Str.DL.144.Traffic (bit)'] + daily_df10637['VS.PS.Str.UL.16.Traffic (bit)'] + daily_df10637['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      daily_df10637['VS.PS.Str.UL.64.Traffic (bit)'] + daily_df10637['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
daily_df10637['CS RAB Drop Rate (%)_10637'] = daily_df10637['VS.RAB.AbnormRel.CS (None)'] / (daily_df10637['VS.RAB.AbnormRel.CS (None)'] + daily_df10637['VS.RAB.NormRel.CS (None)']) * 100
daily_df10637['PS Blocking Rate (%)_10637'] = (daily_df10637['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + daily_df10637['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + daily_df10637['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    daily_df10637['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + daily_df10637['VS.RAB.FailEstabPS.Code.Cong (None)'] + daily_df10637['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    daily_df10637['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + daily_df10637['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + daily_df10637['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (daily_df10637['VS.RAB.AttEstabPS.Str (None)'] + daily_df10637['VS.RAB.AttEstabPS.Int (None)'] + daily_df10637['VS.RAB.AttEstabPS.Bkg (None)']) *100
daily_df10637['PS RAB Drop Rate (%)_10637'] = (daily_df10637['VS.RAB.AbnormRel.PS (None)'] + daily_df10637['VS.RAB.AbnormRel.PS.PCH (None)'] + daily_df10637['VS.RAB.AbnormRel.PS.D2P (None)'] + \
                                    daily_df10637['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (daily_df10637['VS.RAB.AbnormRel.PS (None)'] + daily_df10637['VS.RAB.NormRel.PS (None)'] + daily_df10637['VS.RAB.AbnormRel.PS.PCH (None)'] + \
                                    daily_df10637['VS.RAB.NormRel.PS.PCH (None)']) * 100
daily_df10637['PS HS- Drop Rate (%)_10637'] =  daily_df10637['VS.HSDPA.RAB.AbnormRel (None)'] / (daily_df10637['VS.HSDPA.RAB.AbnormRel (None)'] + daily_df10637['VS.HSDPA.RAB.NormRel (None)'] + daily_df10637['VS.HSDPA.H2D.Succ (None)'] + daily_df10637['VS.HSDPA.H2F.Succ (None)'] +daily_df10637['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + daily_df10637['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
daily_df10637['HSDPA Throughput, kbps_10637'] = daily_df10637['VS.HSDPA.MeanChThroughput (kbit/s)'] / 236 / 24 # количество сот 236!!!
daily_df10637['HSUPA Throughput, kbps_10637'] = daily_df10637['VS.HSUPA.MeanChThroughput (kbit/s)'] / 236 / 24# количество сот 236
daily_df10637['Soft Handover Success rate, %_10637'] = (daily_df10637['VS.SHO.SuccRLAdd (None)'] + daily_df10637['VS.SHO.SuccRLDel (None)']) / (daily_df10637['VS.SHO.AttRLAdd (None)'] + daily_df10637['VS.SHO.AttRLDel (None)']) * 100
daily_df10637['Hard Handover Success rate, %_10637'] = daily_df10637['VS.HHO.SuccInterFreqOut (None)'] / daily_df10637['VS.HHO.AttInterFreqOut (None)'] * 100
daily_df10637['CS W2G Inter-RAT Handover Out SR_10637'] = daily_df10637['IRATHO.SuccOutCS (None)'] / (daily_df10637['IRATHO.AttOutCS (None)'] - daily_df10637['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
daily_df10637['RRC Assignment SucessRate (CS BH), %_10637'] = daily_df10637['RRC.SuccConnEstab.sum (None)'] / daily_df10637['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_df10637['RRC Assignment SucessRate (PS BH), %_10637'] = daily_df10637['RRC.SuccConnEstab.sum (None)'] / daily_df10637['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_df10637['RRC Drop Rate (CS BH), %_10637'] = (daily_df10637['RRC.AttConnRelCCCH.Cong (None)'] + daily_df10637['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df10637['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_df10637['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df10637['RRC.AttConnRelDCCH.Cong (None)'] + daily_df10637['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_df10637['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_df10637['RRC.AttConnRelDCCH.Unspec (None)'] + daily_df10637['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_df10637['RRC.AttConnRelDCCH.Cong (None)'] + daily_df10637['RRC.AttConnRelDCCH.Preempt (None)'] + daily_df10637['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_df10637['RRC.AttConnRelDCCH.DSCR (None)'] + daily_df10637['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_df10637['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_df10637['RRC.AttConnRelCCCH.Cong (None)'] + daily_df10637['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df10637['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_df10637['RRC.AttConnRelCCCH.DSCR (None)'] + daily_df10637['RRC.AttConnRelDCCH.Norm (None)'] + daily_df10637['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_df10637['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_df10637['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df10637['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_df10637['VS.DCCC.Succ.F2P (None)'] + daily_df10637['IRATHO.SuccOutCS (None)'] + daily_df10637['IRATHO.SuccOutPSUTRAN (None)'] + daily_df10637['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_df10637['VS.DCCC.Succ.D2U (None)']) * 100
daily_df10637['RRC Drop Rate (PS BH), %_10637'] = (daily_df10637['RRC.AttConnRelCCCH.Cong (None)'] + daily_df10637['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df10637['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_df10637['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df10637['RRC.AttConnRelDCCH.Cong (None)'] + daily_df10637['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_df10637['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_df10637['RRC.AttConnRelDCCH.Unspec (None)'] + daily_df10637['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_df10637['RRC.AttConnRelDCCH.Cong (None)'] + daily_df10637['RRC.AttConnRelDCCH.Preempt (None)'] + daily_df10637['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_df10637['RRC.AttConnRelDCCH.DSCR (None)'] + daily_df10637['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_df10637['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_df10637['RRC.AttConnRelCCCH.Cong (None)'] + daily_df10637['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df10637['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_df10637['RRC.AttConnRelCCCH.DSCR (None)'] + daily_df10637['RRC.AttConnRelDCCH.Norm (None)'] + daily_df10637['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_df10637['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_df10637['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df10637['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_df10637['VS.DCCC.Succ.F2P (None)'] + daily_df10637['IRATHO.SuccOutCS (None)'] + daily_df10637['IRATHO.SuccOutPSUTRAN (None)'] + daily_df10637['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_df10637['VS.DCCC.Succ.D2U (None)']) * 100
daily_df10637['RAB Assignment Success Rate (CS), %_10637'] = daily_df10637['VS.RAB.SuccEstabCS.AMR (None)'] / daily_df10637['VS.RAB.AttEstab.AMR (None)'] * 100
daily_df10637['RAB Assignment Success Rate (PS), %_10637'] = (daily_df10637['VS.RAB.SuccEstabPS.Conv (None)'] + daily_df10637['VS.RAB.SuccEstabPS.Bkg (None)'] + daily_df10637['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   daily_df10637['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (daily_df10637['VS.RAB.AttEstabPS.Bkg (None)'] + daily_df10637['VS.RAB.AttEstabPS.Int (None)'] + daily_df10637['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   daily_df10637['VS.RAB.AttEstabPS.Conv (None)']) * 100
daily_df10637['CCSR3G, %_10637'] = daily_df10637['RRC Assignment SucessRate (CS BH), %_10637'] * (100 - daily_df10637['RRC Drop Rate (CS BH), %_10637']) * daily_df10637['RAB Assignment Success Rate (CS), %_10637'] * (100 - daily_df10637['CS RAB Drop Rate (%)_10637'])/ 1000000
daily_df10637['DCSR3G, %_10637'] = daily_df10637['RRC Assignment SucessRate (PS BH), %_10637'] * (100 - daily_df10637['RRC Drop Rate (PS BH), %_10637']) * daily_df10637['RAB Assignment Success Rate (PS), %_10637'] * (100 - daily_df10637['PS RAB Drop Rate (%)_10637'])/ 1000000
daily_df10637 = daily_df10637.drop(list_1, axis=1)
# фильтрация по 2937
daily_df2937 = sts_df[sts_df['BSC6910UCell'].isin(list_F3_2937)]
daily_df2937 = daily_df2937.groupby(['date'])[list_1]. sum().reset_index()
daily_df2937['CS traffic 3G, Erl_2937'] = daily_df2937['CS Voice Traffic Volume (Erl)']
daily_df2937['PS traffic 3G UL+DL, GB_2937'] = (daily_df2937['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + daily_df2937['VS.PS.Bkg.DL.8.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Bkg.DL.32.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.DL.64.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Bkg.DL.144.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.DL.256.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Bkg.UL.8.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.UL.16.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Bkg.UL.64.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.UL.128.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Bkg.UL.256.Traffic (bit)'] + daily_df2937['VS.PS.Bkg.UL.384.Traffic (bit)'] + daily_df2937['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Int.DL.16.Traffic (bit)'] + daily_df2937['VS.PS.Int.DL.32.Traffic (bit)'] + daily_df2937['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Int.DL.128.Traffic (bit)'] + daily_df2937['VS.PS.Int.DL.144.Traffic (bit)'] + daily_df2937['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Int.DL.384.Traffic (bit)'] + daily_df2937['VS.PS.Int.UL.8.Traffic (bit)'] + daily_df2937['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Int.UL.32.Traffic (bit)'] + daily_df2937['VS.PS.Int.UL.64.Traffic (bit)'] + daily_df2937['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Int.UL.144.Traffic (bit)'] + daily_df2937['VS.PS.Int.UL.256.Traffic (bit)'] + daily_df2937['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Str.DL.32.Traffic (bit)'] + daily_df2937['VS.PS.Str.DL.64.Traffic (bit)'] + daily_df2937['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Str.DL.144.Traffic (bit)'] + daily_df2937['VS.PS.Str.UL.16.Traffic (bit)'] + daily_df2937['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      daily_df2937['VS.PS.Str.UL.64.Traffic (bit)'] + daily_df2937['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
daily_df2937['CS RAB Drop Rate (%)_2937'] = daily_df2937['VS.RAB.AbnormRel.CS (None)'] / (daily_df2937['VS.RAB.AbnormRel.CS (None)'] + daily_df2937['VS.RAB.NormRel.CS (None)']) * 100
daily_df2937['PS Blocking Rate (%)_2937'] = (daily_df2937['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + daily_df2937['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + daily_df2937['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    daily_df2937['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + daily_df2937['VS.RAB.FailEstabPS.Code.Cong (None)'] + daily_df2937['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    daily_df2937['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + daily_df2937['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + daily_df2937['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (daily_df2937['VS.RAB.AttEstabPS.Str (None)'] + daily_df2937['VS.RAB.AttEstabPS.Int (None)'] + daily_df2937['VS.RAB.AttEstabPS.Bkg (None)']) *100
daily_df2937['PS RAB Drop Rate (%)_2937'] = (daily_df2937['VS.RAB.AbnormRel.PS (None)'] + daily_df2937['VS.RAB.AbnormRel.PS.PCH (None)'] + daily_df2937['VS.RAB.AbnormRel.PS.D2P (None)'] + \
                                    daily_df2937['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (daily_df2937['VS.RAB.AbnormRel.PS (None)'] + daily_df2937['VS.RAB.NormRel.PS (None)'] + daily_df2937['VS.RAB.AbnormRel.PS.PCH (None)'] + \
                                    daily_df2937['VS.RAB.NormRel.PS.PCH (None)']) * 100
daily_df2937['PS HS- Drop Rate (%)_2937'] =  daily_df2937['VS.HSDPA.RAB.AbnormRel (None)'] / (daily_df2937['VS.HSDPA.RAB.AbnormRel (None)'] + daily_df2937['VS.HSDPA.RAB.NormRel (None)'] + daily_df2937['VS.HSDPA.H2D.Succ (None)'] + daily_df2937['VS.HSDPA.H2F.Succ (None)'] +daily_df2937['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + daily_df2937['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
daily_df2937['HSDPA Throughput, kbps_2937'] = daily_df2937['VS.HSDPA.MeanChThroughput (kbit/s)'] / 204 / 24 # количество сот 204!!!
daily_df2937['HSUPA Throughput, kbps_2937'] = daily_df2937['VS.HSUPA.MeanChThroughput (kbit/s)'] / 204 / 24# количество сот 204
daily_df2937['Soft Handover Success rate, %_2937'] = (daily_df2937['VS.SHO.SuccRLAdd (None)'] + daily_df2937['VS.SHO.SuccRLDel (None)']) / (daily_df2937['VS.SHO.AttRLAdd (None)'] + daily_df2937['VS.SHO.AttRLDel (None)']) * 100
daily_df2937['Hard Handover Success rate, %_2937'] = daily_df2937['VS.HHO.SuccInterFreqOut (None)'] / daily_df2937['VS.HHO.AttInterFreqOut (None)'] * 100
daily_df2937['CS W2G Inter-RAT Handover Out SR_2937'] = daily_df2937['IRATHO.SuccOutCS (None)'] / (daily_df2937['IRATHO.AttOutCS (None)'] - daily_df2937['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
daily_df2937['RRC Assignment SucessRate (CS BH), %_2937'] = daily_df2937['RRC.SuccConnEstab.sum (None)'] / daily_df2937['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_df2937['RRC Assignment SucessRate (PS BH), %_2937'] = daily_df2937['RRC.SuccConnEstab.sum (None)'] / daily_df2937['VS.RRC.AttConnEstab.Sum (None)'] * 100
daily_df2937['RRC Drop Rate (CS BH), %_2937'] = (daily_df2937['RRC.AttConnRelCCCH.Cong (None)'] + daily_df2937['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df2937['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_df2937['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df2937['RRC.AttConnRelDCCH.Cong (None)'] + daily_df2937['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_df2937['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_df2937['RRC.AttConnRelDCCH.Unspec (None)'] + daily_df2937['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_df2937['RRC.AttConnRelDCCH.Cong (None)'] + daily_df2937['RRC.AttConnRelDCCH.Preempt (None)'] + daily_df2937['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_df2937['RRC.AttConnRelDCCH.DSCR (None)'] + daily_df2937['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_df2937['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_df2937['RRC.AttConnRelCCCH.Cong (None)'] + daily_df2937['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df2937['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_df2937['RRC.AttConnRelCCCH.DSCR (None)'] + daily_df2937['RRC.AttConnRelDCCH.Norm (None)'] + daily_df2937['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_df2937['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_df2937['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df2937['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_df2937['VS.DCCC.Succ.F2P (None)'] + daily_df2937['IRATHO.SuccOutCS (None)'] + daily_df2937['IRATHO.SuccOutPSUTRAN (None)'] + daily_df2937['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_df2937['VS.DCCC.Succ.D2U (None)']) * 100
daily_df2937['RRC Drop Rate (PS BH), %_2937'] = (daily_df2937['RRC.AttConnRelCCCH.Cong (None)'] + daily_df2937['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df2937['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             daily_df2937['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df2937['RRC.AttConnRelDCCH.Cong (None)'] + daily_df2937['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             daily_df2937['RRC.AttConnRelDCCH.ReEstRej (None)'] + daily_df2937['RRC.AttConnRelDCCH.Unspec (None)'] + daily_df2937['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (daily_df2937['RRC.AttConnRelDCCH.Cong (None)'] + daily_df2937['RRC.AttConnRelDCCH.Preempt (None)'] + daily_df2937['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           daily_df2937['RRC.AttConnRelDCCH.DSCR (None)'] + daily_df2937['RRC.AttConnRelDCCH.UsrInact (None)'] + daily_df2937['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           daily_df2937['RRC.AttConnRelCCCH.Cong (None)'] + daily_df2937['RRC.AttConnRelCCCH.Preempt (None)'] + daily_df2937['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           daily_df2937['RRC.AttConnRelCCCH.DSCR (None)'] + daily_df2937['RRC.AttConnRelDCCH.Norm (None)'] + daily_df2937['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           daily_df2937['RRC.AttConnRelCCCH.UsrInact (None)'] + daily_df2937['RRC.AttConnRelCCCH.Unspec (None)'] + daily_df2937['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           daily_df2937['VS.DCCC.Succ.F2P (None)'] + daily_df2937['IRATHO.SuccOutCS (None)'] + daily_df2937['IRATHO.SuccOutPSUTRAN (None)'] + daily_df2937['VS.DCCC.Succ.F2U (None)'] + \
                                           daily_df2937['VS.DCCC.Succ.D2U (None)']) * 100
daily_df2937['RAB Assignment Success Rate (CS), %_2937'] = daily_df2937['VS.RAB.SuccEstabCS.AMR (None)'] / daily_df2937['VS.RAB.AttEstab.AMR (None)'] * 100
daily_df2937['RAB Assignment Success Rate (PS), %_2937'] = (daily_df2937['VS.RAB.SuccEstabPS.Conv (None)'] + daily_df2937['VS.RAB.SuccEstabPS.Bkg (None)'] + daily_df2937['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   daily_df2937['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (daily_df2937['VS.RAB.AttEstabPS.Bkg (None)'] + daily_df2937['VS.RAB.AttEstabPS.Int (None)'] + daily_df2937['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   daily_df2937['VS.RAB.AttEstabPS.Conv (None)']) * 100
daily_df2937['CCSR3G, %_2937'] = daily_df2937['RRC Assignment SucessRate (CS BH), %_2937'] * (100 - daily_df2937['RRC Drop Rate (CS BH), %_2937']) * daily_df2937['RAB Assignment Success Rate (CS), %_2937'] * (100 - daily_df2937['CS RAB Drop Rate (%)_2937'])/ 1000000
daily_df2937['DCSR3G, %_2937'] = daily_df2937['RRC Assignment SucessRate (PS BH), %_2937'] * (100 - daily_df2937['RRC Drop Rate (PS BH), %_2937']) * daily_df2937['RAB Assignment Success Rate (PS), %_2937'] * (100 - daily_df2937['PS RAB Drop Rate (%)_2937'])/ 1000000
daily_df2937 = daily_df2937.drop(list_1, axis=1)

daily_dfall = pd.merge(daily_df, daily_dfU2100, how="left")
daily_dfall = pd.merge(daily_dfall, daily_dfU900, how="left")
daily_dfall = pd.merge(daily_dfall, daily_df10612, how="left")
daily_dfall = pd.merge(daily_dfall, daily_df10637, how="left")
daily_dfall = pd.merge(daily_dfall, daily_df2937, how="left")
daily_dfall_trans = daily_dfall.transpose()


#NodeB дневная
dailyN_df = stsN_df.groupby(['date'])[list_1N]. sum().reset_index()
dailyN_df['MeanThrHSDPA,kbps'] = dailyN_df['VS.HSDPA.DataOutput.Traffic (bit)']/dailyN_df['VS.HSDPA.DataTtiNum.User (None)'] / 2
dailyN_df['MeanThrHSDPA DC,kbps'] = dailyN_df['VS.DataOutput.AllHSDPA.Traffic (bit)'] / dailyN_df['VS.AllHSDPA.DataTtiNum.User (None)'] / 2
dailyN_df['MeanThrHSUPA,kbps'] = (dailyN_df['VS.HSUPA.2msTTI.Traffic (kbit)'] + dailyN_df['VS.HSUPA.10msTTI.Traffic (kbit)']) / (dailyN_df['VS.HSUPA.2msPDU.TTI.Num (None)'] * 0.002 + dailyN_df['VS.HSUPA.10msPDU.TTI.Num (None)'] * 0.01)
dailyN_df = dailyN_df.drop(list_1N, axis=1)

# сортировка по диапазонам
dailyN_dfU2100 = stsN_df[stsN_df['ULoCell'].isin(list_U2100N)]
dailyN_dfU2100 = dailyN_dfU2100.groupby(['date'])[list_1N]. sum().reset_index()
dailyN_dfU2100['MeanThrHSDPAU2100,kbps'] = dailyN_dfU2100['VS.HSDPA.DataOutput.Traffic (bit)']/dailyN_dfU2100['VS.HSDPA.DataTtiNum.User (None)'] / 2
dailyN_dfU2100['MeanThrHSDPAU2100 DC,kbps'] = dailyN_dfU2100['VS.DataOutput.AllHSDPA.Traffic (bit)'] / dailyN_dfU2100['VS.AllHSDPA.DataTtiNum.User (None)'] / 2
dailyN_dfU2100['MeanThrHSUPAU2100,kbps'] = (dailyN_dfU2100['VS.HSUPA.2msTTI.Traffic (kbit)'] + dailyN_dfU2100['VS.HSUPA.10msTTI.Traffic (kbit)']) / (dailyN_dfU2100['VS.HSUPA.2msPDU.TTI.Num (None)'] * 0.002 + dailyN_dfU2100['VS.HSUPA.10msPDU.TTI.Num (None)'] * 0.01)
dailyN_dfU2100 = dailyN_dfU2100.drop(list_1N, axis=1)

dailyN_dfU900 = stsN_df[stsN_df['ULoCell'].isin(list_U900N)]
dailyN_dfU900 = dailyN_dfU900.groupby(['date'])[list_1N]. sum().reset_index()
dailyN_dfU900['MeanThrHSDPAU900,kbps'] = dailyN_dfU900['VS.HSDPA.DataOutput.Traffic (bit)']/dailyN_dfU900['VS.HSDPA.DataTtiNum.User (None)'] / 2
dailyN_dfU900['MeanThrHSDPAU900 DC,kbps'] = dailyN_dfU900['VS.DataOutput.AllHSDPA.Traffic (bit)'] / dailyN_dfU900['VS.AllHSDPA.DataTtiNum.User (None)'] / 2
dailyN_dfU900['MeanThrHSUPAU900,kbps'] = (dailyN_dfU900['VS.HSUPA.2msTTI.Traffic (kbit)'] + dailyN_dfU900['VS.HSUPA.10msTTI.Traffic (kbit)']) / (dailyN_dfU900['VS.HSUPA.2msPDU.TTI.Num (None)'] * 0.002 + dailyN_dfU900['VS.HSUPA.10msPDU.TTI.Num (None)'] * 0.01)
dailyN_dfU900 = dailyN_dfU900.drop(list_1N, axis=1)

dailyN_df = pd.merge(dailyN_df, dailyN_dfU2100, how="left")
dailyN_df = pd.merge(dailyN_df, dailyN_dfU900, how="left")
#dailyN_df_trans = dailyN_df.transpose()



# ===обработка часовая===
hourly_df = sts_df.groupby(['date', 'hour'])[list_1]. sum().reset_index()
hourly_df['CS traffic 3G, Erl'] = hourly_df['CS Voice Traffic Volume (Erl)']
hourly_df['PS traffic 3G UL+DL, GB'] = (hourly_df['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + hourly_df['VS.PS.Bkg.DL.8.Traffic (bit)'] + hourly_df['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Bkg.DL.32.Traffic (bit)'] + hourly_df['VS.PS.Bkg.DL.64.Traffic (bit)'] + hourly_df['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Bkg.DL.144.Traffic (bit)'] + hourly_df['VS.PS.Bkg.DL.256.Traffic (bit)'] + hourly_df['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Bkg.UL.8.Traffic (bit)'] + hourly_df['VS.PS.Bkg.UL.16.Traffic (bit)'] + hourly_df['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Bkg.UL.64.Traffic (bit)'] + hourly_df['VS.PS.Bkg.UL.128.Traffic (bit)'] + hourly_df['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Bkg.UL.256.Traffic (bit)'] + hourly_df['VS.PS.Bkg.UL.384.Traffic (bit)'] + hourly_df['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Int.DL.16.Traffic (bit)'] + hourly_df['VS.PS.Int.DL.32.Traffic (bit)'] + hourly_df['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Int.DL.128.Traffic (bit)'] + hourly_df['VS.PS.Int.DL.144.Traffic (bit)'] + hourly_df['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Int.DL.384.Traffic (bit)'] + hourly_df['VS.PS.Int.UL.8.Traffic (bit)'] + hourly_df['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Int.UL.32.Traffic (bit)'] + hourly_df['VS.PS.Int.UL.64.Traffic (bit)'] + hourly_df['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Int.UL.144.Traffic (bit)'] + hourly_df['VS.PS.Int.UL.256.Traffic (bit)'] + hourly_df['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Str.DL.32.Traffic (bit)'] + hourly_df['VS.PS.Str.DL.64.Traffic (bit)'] + hourly_df['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Str.DL.144.Traffic (bit)'] + hourly_df['VS.PS.Str.UL.16.Traffic (bit)'] + hourly_df['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      hourly_df['VS.PS.Str.UL.64.Traffic (bit)'] + hourly_df['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
hourly_df['CS RAB Drop Rate (%)'] = hourly_df['VS.RAB.AbnormRel.CS (None)'] / (hourly_df['VS.RAB.AbnormRel.CS (None)'] + hourly_df['VS.RAB.NormRel.CS (None)']) * 100
hourly_df['PS Blocking Rate (%)'] = (hourly_df['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + hourly_df['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + hourly_df['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    hourly_df['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + hourly_df['VS.RAB.FailEstabPS.Code.Cong (None)'] + hourly_df['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    hourly_df['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + hourly_df['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + hourly_df['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (hourly_df['VS.RAB.AttEstabPS.Str (None)'] + hourly_df['VS.RAB.AttEstabPS.Int (None)'] + hourly_df['VS.RAB.AttEstabPS.Bkg (None)']) *100
# hourly_df['PS RAB Drop Rate (%)'] = (hourly_df['VS.RAB.AbnormRel.PS (None)'] + hourly_df['VS.RAB.AbnormRel.PS.PCH (None)'] + hourly_df['VS.RAB.AbnormRel.PS.D2P (None)'] + \
#                                     hourly_df['VS.RAB.AbnormRel.PS.F2P (None)']) / \
#                                    (hourly_df['VS.RAB.AbnormRel.PS (None)'] + hourly_df['VS.RAB.NormRel.PS (None)'] + hourly_df['VS.RAB.AbnormRel.PS.PCH (None)'] + \
#                                     hourly_df['VS.RAB.NormRel.PS.PCH (None)']) * 100

hourly_df['PS RAB Drop Rate (%)'] = (hourly_df['VS.RAB.AbnormRel.PS (None)'] - hourly_df['VS.RAB.AbnormRel.PS.PCH (None)'] - hourly_df['VS.RAB.AbnormRel.PS.D2P (None)'] - \
                                    hourly_df['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (hourly_df['VS.RAB.AbnormRel.PS (None)'] + hourly_df['VS.RAB.NormRel.PS (None)'] - hourly_df['VS.RAB.AbnormRel.PS.PCH (None)'] - \
                                    hourly_df['VS.RAB.NormRel.PS.PCH (None)'] + hourly_df['VS.DCCC.D2P.Succ (None)']+hourly_df['VS.DCCC.Succ.F2P (None)']+hourly_df['VS.DCCC.Succ.F2U (None)']+hourly_df['VS.DCCC.Succ.D2U (None)']) * 100


hourly_df['PS HS- Drop Rate (%)'] =  hourly_df['VS.HSDPA.RAB.AbnormRel (None)'] / (hourly_df['VS.HSDPA.RAB.AbnormRel (None)'] + hourly_df['VS.HSDPA.RAB.NormRel (None)'] + hourly_df['VS.HSDPA.H2D.Succ (None)'] + hourly_df['VS.HSDPA.H2F.Succ (None)'] +hourly_df['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + hourly_df['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
hourly_df['HSDPA Throughput, kbps'] = hourly_df['VS.HSDPA.MeanChThroughput (kbit/s)'] / 538 / 24 # количество сот 538
hourly_df['HSUPA Throughput, kbps'] = hourly_df['VS.HSUPA.MeanChThroughput (kbit/s)'] / 538 / 24# количество сот 538
hourly_df['Soft Handover Success rate, %'] = (hourly_df['VS.SHO.SuccRLAdd (None)'] + hourly_df['VS.SHO.SuccRLDel (None)']) / (hourly_df['VS.SHO.AttRLAdd (None)'] + hourly_df['VS.SHO.AttRLDel (None)']) * 100
hourly_df['Hard Handover Success rate, %'] = hourly_df['VS.HHO.SuccInterFreqOut (None)'] / hourly_df['VS.HHO.AttInterFreqOut (None)'] * 100
hourly_df['CS W2G Inter-RAT Handover Out SR'] = hourly_df['IRATHO.SuccOutCS (None)'] / (hourly_df['IRATHO.AttOutCS (None)'] - hourly_df['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
hourly_df['RRC Assignment SucessRate (CS BH), %'] = hourly_df['RRC.SuccConnEstab.sum (None)'] / hourly_df['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_df['RRC Assignment SucessRate (PS BH), %'] = hourly_df['RRC.SuccConnEstab.sum (None)'] / hourly_df['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_df['RRC Drop Rate (CS BH), %'] = (hourly_df['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_df['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_df['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_df['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_df['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_df['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_df['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_df['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_df['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_df['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_df['RRC.AttConnRelDCCH.Norm (None)'] + hourly_df['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_df['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_df['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_df['VS.DCCC.Succ.F2P (None)'] + hourly_df['IRATHO.SuccOutCS (None)'] + hourly_df['IRATHO.SuccOutPSUTRAN (None)'] + hourly_df['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_df['VS.DCCC.Succ.D2U (None)']) * 100
hourly_df['RRC Drop Rate (PS BH), %'] = (hourly_df['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_df['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_df['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_df['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_df['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_df['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_df['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_df['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_df['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_df['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_df['RRC.AttConnRelDCCH.Norm (None)'] + hourly_df['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_df['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_df['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_df['VS.DCCC.Succ.F2P (None)'] + hourly_df['IRATHO.SuccOutCS (None)'] + hourly_df['IRATHO.SuccOutPSUTRAN (None)'] + hourly_df['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_df['VS.DCCC.Succ.D2U (None)']) * 100
hourly_df['RAB Assignment Success Rate (CS), %'] = hourly_df['VS.RAB.SuccEstabCS.AMR (None)'] / hourly_df['VS.RAB.AttEstab.AMR (None)'] * 100
hourly_df['RAB Assignment Success Rate (PS), %'] = (hourly_df['VS.RAB.SuccEstabPS.Conv (None)'] + hourly_df['VS.RAB.SuccEstabPS.Bkg (None)'] + hourly_df['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   hourly_df['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (hourly_df['VS.RAB.AttEstabPS.Bkg (None)'] + hourly_df['VS.RAB.AttEstabPS.Int (None)'] + hourly_df['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   hourly_df['VS.RAB.AttEstabPS.Conv (None)']) * 100
hourly_df['CCSR3G, %'] = hourly_df['RRC Assignment SucessRate (CS BH), %'] * (100 - hourly_df['RRC Drop Rate (CS BH), %']) * hourly_df['RAB Assignment Success Rate (CS), %'] * (100 - hourly_df['CS RAB Drop Rate (%)'])/ 1000000
hourly_df['DCSR3G, %'] = hourly_df['RRC Assignment SucessRate (PS BH), %'] * (100 - hourly_df['RRC Drop Rate (PS BH), %']) * hourly_df['RAB Assignment Success Rate (PS), %'] * (100 - hourly_df['PS RAB Drop Rate (%)'])/ 1000000
hourly_df = hourly_df.drop(list_1, axis=1)

# фильтрация по U2100
hourly_dfU2100 = sts_df[sts_df['BSC6910UCell'].isin(list_U2100)]
hourly_dfU2100 = hourly_dfU2100.groupby(['date', 'hour'])[list_1]. sum().reset_index()
hourly_dfU2100['CS traffic 3G, Erl_U2100'] = hourly_dfU2100['CS Voice Traffic Volume (Erl)']
hourly_dfU2100['PS traffic 3G UL+DL, GB_U2100'] = (hourly_dfU2100['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + hourly_dfU2100['VS.PS.Bkg.DL.8.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Bkg.DL.32.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.DL.64.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Bkg.DL.144.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.DL.256.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Bkg.UL.8.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.UL.16.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Bkg.UL.64.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.UL.128.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Bkg.UL.256.Traffic (bit)'] + hourly_dfU2100['VS.PS.Bkg.UL.384.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Int.DL.16.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.DL.32.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Int.DL.128.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.DL.144.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Int.DL.384.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.UL.8.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Int.UL.32.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.UL.64.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Int.UL.144.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.UL.256.Traffic (bit)'] + hourly_dfU2100['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Str.DL.32.Traffic (bit)'] + hourly_dfU2100['VS.PS.Str.DL.64.Traffic (bit)'] + hourly_dfU2100['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Str.DL.144.Traffic (bit)'] + hourly_dfU2100['VS.PS.Str.UL.16.Traffic (bit)'] + hourly_dfU2100['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      hourly_dfU2100['VS.PS.Str.UL.64.Traffic (bit)'] + hourly_dfU2100['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
hourly_dfU2100['CS RAB Drop Rate (%)_U2100'] = hourly_dfU2100['VS.RAB.AbnormRel.CS (None)'] / (hourly_dfU2100['VS.RAB.AbnormRel.CS (None)'] + hourly_dfU2100['VS.RAB.NormRel.CS (None)']) * 100
hourly_dfU2100['PS Blocking Rate (%)_U2100'] = (hourly_dfU2100['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + hourly_dfU2100['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + hourly_dfU2100['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    hourly_dfU2100['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + hourly_dfU2100['VS.RAB.FailEstabPS.Code.Cong (None)'] + hourly_dfU2100['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    hourly_dfU2100['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + hourly_dfU2100['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + hourly_dfU2100['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (hourly_dfU2100['VS.RAB.AttEstabPS.Str (None)'] + hourly_dfU2100['VS.RAB.AttEstabPS.Int (None)'] + hourly_dfU2100['VS.RAB.AttEstabPS.Bkg (None)']) *100
# hourly_dfU2100['PS RAB Drop Rate (%)_U2100'] = (hourly_dfU2100['VS.RAB.AbnormRel.PS (None)'] + hourly_dfU2100['VS.RAB.AbnormRel.PS.PCH (None)'] + hourly_dfU2100['VS.RAB.AbnormRel.PS.D2P (None)'] + \
#                                     hourly_dfU2100['VS.RAB.AbnormRel.PS.F2P (None)']) / \
#                                    (hourly_dfU2100['VS.RAB.AbnormRel.PS (None)'] + hourly_dfU2100['VS.RAB.NormRel.PS (None)'] + hourly_dfU2100['VS.RAB.AbnormRel.PS.PCH (None)'] + \
#                                     hourly_dfU2100['VS.RAB.NormRel.PS.PCH (None)']) * 100
hourly_dfU2100['PS RAB Drop Rate (%)_U2100'] = (hourly_dfU2100['VS.RAB.AbnormRel.PS (None)'] - hourly_dfU2100['VS.RAB.AbnormRel.PS.PCH (None)'] - hourly_dfU2100['VS.RAB.AbnormRel.PS.D2P (None)'] - \
                                    hourly_dfU2100['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (hourly_dfU2100['VS.RAB.AbnormRel.PS (None)'] + hourly_dfU2100['VS.RAB.NormRel.PS (None)'] - hourly_dfU2100['VS.RAB.AbnormRel.PS.PCH (None)'] - \
                                    hourly_dfU2100['VS.RAB.NormRel.PS.PCH (None)'] + hourly_dfU2100['VS.DCCC.D2P.Succ (None)']+hourly_dfU2100['VS.DCCC.Succ.F2P (None)']+hourly_dfU2100['VS.DCCC.Succ.F2U (None)']+hourly_dfU2100['VS.DCCC.Succ.D2U (None)']) * 100


hourly_dfU2100['PS HS- Drop Rate (%)_U2100'] =  hourly_dfU2100['VS.HSDPA.RAB.AbnormRel (None)'] / (hourly_dfU2100['VS.HSDPA.RAB.AbnormRel (None)'] + hourly_dfU2100['VS.HSDPA.RAB.NormRel (None)'] + hourly_dfU2100['VS.HSDPA.H2D.Succ (None)'] + hourly_dfU2100['VS.HSDPA.H2F.Succ (None)'] +hourly_dfU2100['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + hourly_dfU2100['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
hourly_dfU2100['HSDPA Throughput, kbps_U2100'] = hourly_dfU2100['VS.HSDPA.MeanChThroughput (kbit/s)'] / 471 / 24 # количество сот 471!!!
hourly_dfU2100['HSUPA Throughput, kbps_U2100'] = hourly_dfU2100['VS.HSUPA.MeanChThroughput (kbit/s)'] / 471 / 24# количество сот 471!!!
hourly_dfU2100['Soft Handover Success rate, %_U2100'] = (hourly_dfU2100['VS.SHO.SuccRLAdd (None)'] + hourly_dfU2100['VS.SHO.SuccRLDel (None)']) / (hourly_dfU2100['VS.SHO.AttRLAdd (None)'] + hourly_dfU2100['VS.SHO.AttRLDel (None)']) * 100
hourly_dfU2100['Hard Handover Success rate, %_U2100'] = hourly_dfU2100['VS.HHO.SuccInterFreqOut (None)'] / hourly_dfU2100['VS.HHO.AttInterFreqOut (None)'] * 100
hourly_dfU2100['CS W2G Inter-RAT Handover Out SR_U2100'] = hourly_dfU2100['IRATHO.SuccOutCS (None)'] / (hourly_dfU2100['IRATHO.AttOutCS (None)'] - hourly_dfU2100['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
hourly_dfU2100['RRC Assignment SucessRate (CS BH), %_U2100'] = hourly_dfU2100['RRC.SuccConnEstab.sum (None)'] / hourly_dfU2100['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_dfU2100['RRC Assignment SucessRate (PS BH), %_U2100'] = hourly_dfU2100['RRC.SuccConnEstab.sum (None)'] / hourly_dfU2100['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_dfU2100['RRC Drop Rate (CS BH), %_U2100'] = (hourly_dfU2100['RRC.AttConnRelCCCH.Cong (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_dfU2100['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Cong (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_dfU2100['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_dfU2100['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_dfU2100['RRC.AttConnRelDCCH.Cong (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_dfU2100['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_dfU2100['RRC.AttConnRelCCCH.Cong (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_dfU2100['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Norm (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_dfU2100['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_dfU2100['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_dfU2100['VS.DCCC.Succ.F2P (None)'] + hourly_dfU2100['IRATHO.SuccOutCS (None)'] + hourly_dfU2100['IRATHO.SuccOutPSUTRAN (None)'] + hourly_dfU2100['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_dfU2100['VS.DCCC.Succ.D2U (None)']) * 100
hourly_dfU2100['RRC Drop Rate (PS BH), %_U2100'] = (hourly_dfU2100['RRC.AttConnRelCCCH.Cong (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_dfU2100['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Cong (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_dfU2100['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_dfU2100['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_dfU2100['RRC.AttConnRelDCCH.Cong (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_dfU2100['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_dfU2100['RRC.AttConnRelCCCH.Cong (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_dfU2100['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_dfU2100['RRC.AttConnRelDCCH.Norm (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_dfU2100['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_dfU2100['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_dfU2100['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_dfU2100['VS.DCCC.Succ.F2P (None)'] + hourly_dfU2100['IRATHO.SuccOutCS (None)'] + hourly_dfU2100['IRATHO.SuccOutPSUTRAN (None)'] + hourly_dfU2100['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_dfU2100['VS.DCCC.Succ.D2U (None)']) * 100
hourly_dfU2100['RAB Assignment Success Rate (CS), %_U2100'] = hourly_dfU2100['VS.RAB.SuccEstabCS.AMR (None)'] / hourly_dfU2100['VS.RAB.AttEstab.AMR (None)'] * 100
hourly_dfU2100['RAB Assignment Success Rate (PS), %_U2100'] = (hourly_dfU2100['VS.RAB.SuccEstabPS.Conv (None)'] + hourly_dfU2100['VS.RAB.SuccEstabPS.Bkg (None)'] + hourly_dfU2100['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   hourly_dfU2100['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (hourly_dfU2100['VS.RAB.AttEstabPS.Bkg (None)'] + hourly_dfU2100['VS.RAB.AttEstabPS.Int (None)'] + hourly_dfU2100['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   hourly_dfU2100['VS.RAB.AttEstabPS.Conv (None)']) * 100
hourly_dfU2100['CCSR3G, %_U2100'] = hourly_dfU2100['RRC Assignment SucessRate (CS BH), %_U2100'] * (100 - hourly_dfU2100['RRC Drop Rate (CS BH), %_U2100']) * hourly_dfU2100['RAB Assignment Success Rate (CS), %_U2100'] * (100 - hourly_dfU2100['CS RAB Drop Rate (%)_U2100'])/ 1000000
hourly_dfU2100['DCSR3G, %_U2100'] = hourly_dfU2100['RRC Assignment SucessRate (PS BH), %_U2100'] * (100 - hourly_dfU2100['RRC Drop Rate (PS BH), %_U2100']) * hourly_dfU2100['RAB Assignment Success Rate (PS), %_U2100'] * (100 - hourly_dfU2100['PS RAB Drop Rate (%)_U2100'])/ 1000000
hourly_dfU2100 = hourly_dfU2100.drop(list_1, axis=1)

# фильтрация по U900
hourly_dfU900 = sts_df[sts_df['BSC6910UCell'].isin(list_U900)]
hourly_dfU900 = hourly_dfU900.groupby(['date', 'hour'])[list_1]. sum().reset_index()
hourly_dfU900['CS traffic 3G, Erl_U900'] = hourly_dfU900['CS Voice Traffic Volume (Erl)']
hourly_dfU900['PS traffic 3G UL+DL, GB_U900'] = (hourly_dfU900['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + hourly_dfU900['VS.PS.Bkg.DL.8.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Bkg.DL.32.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.DL.64.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Bkg.DL.144.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.DL.256.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Bkg.UL.8.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.UL.16.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Bkg.UL.64.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.UL.128.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Bkg.UL.256.Traffic (bit)'] + hourly_dfU900['VS.PS.Bkg.UL.384.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Int.DL.16.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.DL.32.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Int.DL.128.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.DL.144.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Int.DL.384.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.UL.8.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Int.UL.32.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.UL.64.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Int.UL.144.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.UL.256.Traffic (bit)'] + hourly_dfU900['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Str.DL.32.Traffic (bit)'] + hourly_dfU900['VS.PS.Str.DL.64.Traffic (bit)'] + hourly_dfU900['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Str.DL.144.Traffic (bit)'] + hourly_dfU900['VS.PS.Str.UL.16.Traffic (bit)'] + hourly_dfU900['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      hourly_dfU900['VS.PS.Str.UL.64.Traffic (bit)'] + hourly_dfU900['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
hourly_dfU900['CS RAB Drop Rate (%)_U900'] = hourly_dfU900['VS.RAB.AbnormRel.CS (None)'] / (hourly_dfU900['VS.RAB.AbnormRel.CS (None)'] + hourly_dfU900['VS.RAB.NormRel.CS (None)']) * 100
hourly_dfU900['PS Blocking Rate (%)_U900'] = (hourly_dfU900['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + hourly_dfU900['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + hourly_dfU900['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    hourly_dfU900['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + hourly_dfU900['VS.RAB.FailEstabPS.Code.Cong (None)'] + hourly_dfU900['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    hourly_dfU900['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + hourly_dfU900['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + hourly_dfU900['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (hourly_dfU900['VS.RAB.AttEstabPS.Str (None)'] + hourly_dfU900['VS.RAB.AttEstabPS.Int (None)'] + hourly_dfU900['VS.RAB.AttEstabPS.Bkg (None)']) *100
# hourly_dfU900['PS RAB Drop Rate (%)_U900'] = (hourly_dfU900['VS.RAB.AbnormRel.PS (None)'] + hourly_dfU900['VS.RAB.AbnormRel.PS.PCH (None)'] + hourly_dfU900['VS.RAB.AbnormRel.PS.D2P (None)'] + \
#                                     hourly_dfU900['VS.RAB.AbnormRel.PS.F2P (None)']) / \
#                                    (hourly_dfU900['VS.RAB.AbnormRel.PS (None)'] + hourly_dfU900['VS.RAB.NormRel.PS (None)'] + hourly_dfU900['VS.RAB.AbnormRel.PS.PCH (None)'] + \
#                                     hourly_dfU900['VS.RAB.NormRel.PS.PCH (None)']) * 100
hourly_dfU900['PS RAB Drop Rate (%)_U900'] = (hourly_dfU900['VS.RAB.AbnormRel.PS (None)'] - hourly_dfU900['VS.RAB.AbnormRel.PS.PCH (None)'] - hourly_dfU900['VS.RAB.AbnormRel.PS.D2P (None)'] - \
                                    hourly_dfU900['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (hourly_dfU900['VS.RAB.AbnormRel.PS (None)'] + hourly_dfU900['VS.RAB.NormRel.PS (None)'] - hourly_dfU900['VS.RAB.AbnormRel.PS.PCH (None)'] - \
                                    hourly_dfU900['VS.RAB.NormRel.PS.PCH (None)'] + hourly_dfU900['VS.DCCC.D2P.Succ (None)']+hourly_dfU900['VS.DCCC.Succ.F2P (None)']+hourly_dfU900['VS.DCCC.Succ.F2U (None)']+hourly_dfU900['VS.DCCC.Succ.D2U (None)']) * 100


hourly_dfU900['PS HS- Drop Rate (%)_U900'] =  hourly_dfU900['VS.HSDPA.RAB.AbnormRel (None)'] / (hourly_dfU900['VS.HSDPA.RAB.AbnormRel (None)'] + hourly_dfU900['VS.HSDPA.RAB.NormRel (None)'] + hourly_dfU900['VS.HSDPA.H2D.Succ (None)'] + hourly_dfU900['VS.HSDPA.H2F.Succ (None)'] +hourly_dfU900['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + hourly_dfU900['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
hourly_dfU900['HSDPA Throughput, kbps_U900'] = hourly_dfU900['VS.HSDPA.MeanChThroughput (kbit/s)'] / 216 / 24 # количество сот 216!!!
hourly_dfU900['HSUPA Throughput, kbps_U900'] = hourly_dfU900['VS.HSUPA.MeanChThroughput (kbit/s)'] / 216 / 24# количество сот 216
hourly_dfU900['Soft Handover Success rate, %_U900'] = (hourly_dfU900['VS.SHO.SuccRLAdd (None)'] + hourly_dfU900['VS.SHO.SuccRLDel (None)']) / (hourly_dfU900['VS.SHO.AttRLAdd (None)'] + hourly_dfU900['VS.SHO.AttRLDel (None)']) * 100
hourly_dfU900['Hard Handover Success rate, %_U900'] = hourly_dfU900['VS.HHO.SuccInterFreqOut (None)'] / hourly_dfU900['VS.HHO.AttInterFreqOut (None)'] * 100
hourly_dfU900['CS W2G Inter-RAT Handover Out SR_U900'] = hourly_dfU900['IRATHO.SuccOutCS (None)'] / (hourly_dfU900['IRATHO.AttOutCS (None)'] - hourly_dfU900['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
hourly_dfU900['RRC Assignment SucessRate (CS BH), %_U900'] = hourly_dfU900['RRC.SuccConnEstab.sum (None)'] / hourly_dfU900['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_dfU900['RRC Assignment SucessRate (PS BH), %_U900'] = hourly_dfU900['RRC.SuccConnEstab.sum (None)'] / hourly_dfU900['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_dfU900['RRC Drop Rate (CS BH), %_U900'] = (hourly_dfU900['RRC.AttConnRelCCCH.Cong (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_dfU900['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Cong (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_dfU900['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_dfU900['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_dfU900['RRC.AttConnRelDCCH.Cong (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_dfU900['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_dfU900['RRC.AttConnRelCCCH.Cong (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_dfU900['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Norm (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_dfU900['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_dfU900['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_dfU900['VS.DCCC.Succ.F2P (None)'] + hourly_dfU900['IRATHO.SuccOutCS (None)'] + hourly_dfU900['IRATHO.SuccOutPSUTRAN (None)'] + hourly_dfU900['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_dfU900['VS.DCCC.Succ.D2U (None)']) * 100
hourly_dfU900['RRC Drop Rate (PS BH), %_U900'] = (hourly_dfU900['RRC.AttConnRelCCCH.Cong (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_dfU900['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Cong (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_dfU900['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_dfU900['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_dfU900['RRC.AttConnRelDCCH.Cong (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_dfU900['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_dfU900['RRC.AttConnRelCCCH.Cong (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_dfU900['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_dfU900['RRC.AttConnRelDCCH.Norm (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_dfU900['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_dfU900['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_dfU900['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_dfU900['VS.DCCC.Succ.F2P (None)'] + hourly_dfU900['IRATHO.SuccOutCS (None)'] + hourly_dfU900['IRATHO.SuccOutPSUTRAN (None)'] + hourly_dfU900['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_dfU900['VS.DCCC.Succ.D2U (None)']) * 100
hourly_dfU900['RAB Assignment Success Rate (CS), %_U900'] = hourly_dfU900['VS.RAB.SuccEstabCS.AMR (None)'] / hourly_dfU900['VS.RAB.AttEstab.AMR (None)'] * 100
hourly_dfU900['RAB Assignment Success Rate (PS), %_U900'] = (hourly_dfU900['VS.RAB.SuccEstabPS.Conv (None)'] + hourly_dfU900['VS.RAB.SuccEstabPS.Bkg (None)'] + hourly_dfU900['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   hourly_dfU900['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (hourly_dfU900['VS.RAB.AttEstabPS.Bkg (None)'] + hourly_dfU900['VS.RAB.AttEstabPS.Int (None)'] + hourly_dfU900['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   hourly_dfU900['VS.RAB.AttEstabPS.Conv (None)']) * 100
hourly_dfU900['CCSR3G, %_U900'] = hourly_dfU900['RRC Assignment SucessRate (CS BH), %_U900'] * (100 - hourly_dfU900['RRC Drop Rate (CS BH), %_U900']) * hourly_dfU900['RAB Assignment Success Rate (CS), %_U900'] * (100 - hourly_dfU900['CS RAB Drop Rate (%)_U900'])/ 1000000
hourly_dfU900['DCSR3G, %_U900'] = hourly_dfU900['RRC Assignment SucessRate (PS BH), %_U900'] * (100 - hourly_dfU900['RRC Drop Rate (PS BH), %_U900']) * hourly_dfU900['RAB Assignment Success Rate (PS), %_U900'] * (100 - hourly_dfU900['PS RAB Drop Rate (%)_U900'])/ 1000000
hourly_dfU900 = hourly_dfU900.drop(list_1, axis=1)
# фильтрация по 10612
hourly_df10612 = sts_df[sts_df['BSC6910UCell'].isin(list_F1_10612)]
hourly_df10612 = hourly_df10612.groupby(['date', 'hour'])[list_1]. sum().reset_index()
hourly_df10612['CS traffic 3G, Erl_10612'] = hourly_df10612['CS Voice Traffic Volume (Erl)']
hourly_df10612['PS traffic 3G UL+DL, GB_10612'] = (hourly_df10612['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + hourly_df10612['VS.PS.Bkg.DL.8.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Bkg.DL.32.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.DL.64.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Bkg.DL.144.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.DL.256.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Bkg.UL.8.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.UL.16.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Bkg.UL.64.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.UL.128.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Bkg.UL.256.Traffic (bit)'] + hourly_df10612['VS.PS.Bkg.UL.384.Traffic (bit)'] + hourly_df10612['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Int.DL.16.Traffic (bit)'] + hourly_df10612['VS.PS.Int.DL.32.Traffic (bit)'] + hourly_df10612['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Int.DL.128.Traffic (bit)'] + hourly_df10612['VS.PS.Int.DL.144.Traffic (bit)'] + hourly_df10612['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Int.DL.384.Traffic (bit)'] + hourly_df10612['VS.PS.Int.UL.8.Traffic (bit)'] + hourly_df10612['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Int.UL.32.Traffic (bit)'] + hourly_df10612['VS.PS.Int.UL.64.Traffic (bit)'] + hourly_df10612['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Int.UL.144.Traffic (bit)'] + hourly_df10612['VS.PS.Int.UL.256.Traffic (bit)'] + hourly_df10612['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Str.DL.32.Traffic (bit)'] + hourly_df10612['VS.PS.Str.DL.64.Traffic (bit)'] + hourly_df10612['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Str.DL.144.Traffic (bit)'] + hourly_df10612['VS.PS.Str.UL.16.Traffic (bit)'] + hourly_df10612['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      hourly_df10612['VS.PS.Str.UL.64.Traffic (bit)'] + hourly_df10612['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
hourly_df10612['CS RAB Drop Rate (%)_10612'] = hourly_df10612['VS.RAB.AbnormRel.CS (None)'] / (hourly_df10612['VS.RAB.AbnormRel.CS (None)'] + hourly_df10612['VS.RAB.NormRel.CS (None)']) * 100
hourly_df10612['PS Blocking Rate (%)_10612'] = (hourly_df10612['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + hourly_df10612['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + hourly_df10612['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    hourly_df10612['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + hourly_df10612['VS.RAB.FailEstabPS.Code.Cong (None)'] + hourly_df10612['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    hourly_df10612['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + hourly_df10612['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + hourly_df10612['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (hourly_df10612['VS.RAB.AttEstabPS.Str (None)'] + hourly_df10612['VS.RAB.AttEstabPS.Int (None)'] + hourly_df10612['VS.RAB.AttEstabPS.Bkg (None)']) *100
hourly_df10612['PS RAB Drop Rate (%)_10612'] = (hourly_df10612['VS.RAB.AbnormRel.PS (None)'] + hourly_df10612['VS.RAB.AbnormRel.PS.PCH (None)'] + hourly_df10612['VS.RAB.AbnormRel.PS.D2P (None)'] + \
                                    hourly_df10612['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (hourly_df10612['VS.RAB.AbnormRel.PS (None)'] + hourly_df10612['VS.RAB.NormRel.PS (None)'] + hourly_df10612['VS.RAB.AbnormRel.PS.PCH (None)'] + \
                                    hourly_df10612['VS.RAB.NormRel.PS.PCH (None)']) * 100
hourly_df10612['PS HS- Drop Rate (%)_10612'] =  hourly_df10612['VS.HSDPA.RAB.AbnormRel (None)'] / (hourly_df10612['VS.HSDPA.RAB.AbnormRel (None)'] + hourly_df10612['VS.HSDPA.RAB.NormRel (None)'] + hourly_df10612['VS.HSDPA.H2D.Succ (None)'] + hourly_df10612['VS.HSDPA.H2F.Succ (None)'] +hourly_df10612['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + hourly_df10612['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
hourly_df10612['HSDPA Throughput, kbps_10612'] = hourly_df10612['VS.HSDPA.MeanChThroughput (kbit/s)'] / 235 / 24 # количество сот 235!!!
hourly_df10612['HSUPA Throughput, kbps_10612'] = hourly_df10612['VS.HSUPA.MeanChThroughput (kbit/s)'] / 235 / 24# количество сот 235
hourly_df10612['Soft Handover Success rate, %_10612'] = (hourly_df10612['VS.SHO.SuccRLAdd (None)'] + hourly_df10612['VS.SHO.SuccRLDel (None)']) / (hourly_df10612['VS.SHO.AttRLAdd (None)'] + hourly_df10612['VS.SHO.AttRLDel (None)']) * 100
hourly_df10612['Hard Handover Success rate, %_10612'] = hourly_df10612['VS.HHO.SuccInterFreqOut (None)'] / hourly_df10612['VS.HHO.AttInterFreqOut (None)'] * 100
hourly_df10612['CS W2G Inter-RAT Handover Out SR_10612'] = hourly_df10612['IRATHO.SuccOutCS (None)'] / (hourly_df10612['IRATHO.AttOutCS (None)'] - hourly_df10612['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
hourly_df10612['RRC Assignment SucessRate (CS BH), %_10612'] = hourly_df10612['RRC.SuccConnEstab.sum (None)'] / hourly_df10612['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_df10612['RRC Assignment SucessRate (PS BH), %_10612'] = hourly_df10612['RRC.SuccConnEstab.sum (None)'] / hourly_df10612['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_df10612['RRC Drop Rate (CS BH), %_10612'] = (hourly_df10612['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df10612['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df10612['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_df10612['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_df10612['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_df10612['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_df10612['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_df10612['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_df10612['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_df10612['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_df10612['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df10612['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df10612['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_df10612['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Norm (None)'] + hourly_df10612['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_df10612['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_df10612['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df10612['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_df10612['VS.DCCC.Succ.F2P (None)'] + hourly_df10612['IRATHO.SuccOutCS (None)'] + hourly_df10612['IRATHO.SuccOutPSUTRAN (None)'] + hourly_df10612['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_df10612['VS.DCCC.Succ.D2U (None)']) * 100
hourly_df10612['RRC Drop Rate (PS BH), %_10612'] = (hourly_df10612['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df10612['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df10612['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_df10612['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_df10612['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_df10612['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_df10612['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_df10612['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_df10612['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_df10612['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_df10612['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df10612['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df10612['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_df10612['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_df10612['RRC.AttConnRelDCCH.Norm (None)'] + hourly_df10612['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_df10612['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_df10612['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df10612['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_df10612['VS.DCCC.Succ.F2P (None)'] + hourly_df10612['IRATHO.SuccOutCS (None)'] + hourly_df10612['IRATHO.SuccOutPSUTRAN (None)'] + hourly_df10612['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_df10612['VS.DCCC.Succ.D2U (None)']) * 100
hourly_df10612['RAB Assignment Success Rate (CS), %_10612'] = hourly_df10612['VS.RAB.SuccEstabCS.AMR (None)'] / hourly_df10612['VS.RAB.AttEstab.AMR (None)'] * 100
hourly_df10612['RAB Assignment Success Rate (PS), %_10612'] = (hourly_df10612['VS.RAB.SuccEstabPS.Conv (None)'] + hourly_df10612['VS.RAB.SuccEstabPS.Bkg (None)'] + hourly_df10612['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   hourly_df10612['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (hourly_df10612['VS.RAB.AttEstabPS.Bkg (None)'] + hourly_df10612['VS.RAB.AttEstabPS.Int (None)'] + hourly_df10612['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   hourly_df10612['VS.RAB.AttEstabPS.Conv (None)']) * 100
hourly_df10612['CCSR3G, %_10612'] = hourly_df10612['RRC Assignment SucessRate (CS BH), %_10612'] * (100 - hourly_df10612['RRC Drop Rate (CS BH), %_10612']) * hourly_df10612['RAB Assignment Success Rate (CS), %_10612'] * (100 - hourly_df10612['CS RAB Drop Rate (%)_10612'])/ 1000000
hourly_df10612['DCSR3G, %_10612'] = hourly_df10612['RRC Assignment SucessRate (PS BH), %_10612'] * (100 - hourly_df10612['RRC Drop Rate (PS BH), %_10612']) * hourly_df10612['RAB Assignment Success Rate (PS), %_10612'] * (100 - hourly_df10612['PS RAB Drop Rate (%)_10612'])/ 1000000
hourly_df10612 = hourly_df10612.drop(list_1, axis=1)
# фильтрация по 10637
hourly_df10637 = sts_df[sts_df['BSC6910UCell'].isin(list_F2_10637)]
hourly_df10637 = hourly_df10637.groupby(['date', 'hour'])[list_1]. sum().reset_index()
hourly_df10637['CS traffic 3G, Erl_10637'] = hourly_df10637['CS Voice Traffic Volume (Erl)']
hourly_df10637['PS traffic 3G UL+DL, GB_10637'] = (hourly_df10637['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + hourly_df10637['VS.PS.Bkg.DL.8.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Bkg.DL.32.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.DL.64.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Bkg.DL.144.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.DL.256.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Bkg.UL.8.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.UL.16.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Bkg.UL.64.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.UL.128.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Bkg.UL.256.Traffic (bit)'] + hourly_df10637['VS.PS.Bkg.UL.384.Traffic (bit)'] + hourly_df10637['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Int.DL.16.Traffic (bit)'] + hourly_df10637['VS.PS.Int.DL.32.Traffic (bit)'] + hourly_df10637['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Int.DL.128.Traffic (bit)'] + hourly_df10637['VS.PS.Int.DL.144.Traffic (bit)'] + hourly_df10637['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Int.DL.384.Traffic (bit)'] + hourly_df10637['VS.PS.Int.UL.8.Traffic (bit)'] + hourly_df10637['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Int.UL.32.Traffic (bit)'] + hourly_df10637['VS.PS.Int.UL.64.Traffic (bit)'] + hourly_df10637['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Int.UL.144.Traffic (bit)'] + hourly_df10637['VS.PS.Int.UL.256.Traffic (bit)'] + hourly_df10637['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Str.DL.32.Traffic (bit)'] + hourly_df10637['VS.PS.Str.DL.64.Traffic (bit)'] + hourly_df10637['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Str.DL.144.Traffic (bit)'] + hourly_df10637['VS.PS.Str.UL.16.Traffic (bit)'] + hourly_df10637['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      hourly_df10637['VS.PS.Str.UL.64.Traffic (bit)'] + hourly_df10637['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
hourly_df10637['CS RAB Drop Rate (%)_10637'] = hourly_df10637['VS.RAB.AbnormRel.CS (None)'] / (hourly_df10637['VS.RAB.AbnormRel.CS (None)'] + hourly_df10637['VS.RAB.NormRel.CS (None)']) * 100
hourly_df10637['PS Blocking Rate (%)_10637'] = (hourly_df10637['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + hourly_df10637['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + hourly_df10637['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    hourly_df10637['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + hourly_df10637['VS.RAB.FailEstabPS.Code.Cong (None)'] + hourly_df10637['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    hourly_df10637['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + hourly_df10637['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + hourly_df10637['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (hourly_df10637['VS.RAB.AttEstabPS.Str (None)'] + hourly_df10637['VS.RAB.AttEstabPS.Int (None)'] + hourly_df10637['VS.RAB.AttEstabPS.Bkg (None)']) *100
hourly_df10637['PS RAB Drop Rate (%)_10637'] = (hourly_df10637['VS.RAB.AbnormRel.PS (None)'] + hourly_df10637['VS.RAB.AbnormRel.PS.PCH (None)'] + hourly_df10637['VS.RAB.AbnormRel.PS.D2P (None)'] + \
                                    hourly_df10637['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (hourly_df10637['VS.RAB.AbnormRel.PS (None)'] + hourly_df10637['VS.RAB.NormRel.PS (None)'] + hourly_df10637['VS.RAB.AbnormRel.PS.PCH (None)'] + \
                                    hourly_df10637['VS.RAB.NormRel.PS.PCH (None)']) * 100
hourly_df10637['PS HS- Drop Rate (%)_10637'] =  hourly_df10637['VS.HSDPA.RAB.AbnormRel (None)'] / (hourly_df10637['VS.HSDPA.RAB.AbnormRel (None)'] + hourly_df10637['VS.HSDPA.RAB.NormRel (None)'] + hourly_df10637['VS.HSDPA.H2D.Succ (None)'] + hourly_df10637['VS.HSDPA.H2F.Succ (None)'] +hourly_df10637['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + hourly_df10637['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
hourly_df10637['HSDPA Throughput, kbps_10637'] = hourly_df10637['VS.HSDPA.MeanChThroughput (kbit/s)'] / 236 / 24 # количество сот 236!!!
hourly_df10637['HSUPA Throughput, kbps_10637'] = hourly_df10637['VS.HSUPA.MeanChThroughput (kbit/s)'] / 236 / 24# количество сот 236
hourly_df10637['Soft Handover Success rate, %_10637'] = (hourly_df10637['VS.SHO.SuccRLAdd (None)'] + hourly_df10637['VS.SHO.SuccRLDel (None)']) / (hourly_df10637['VS.SHO.AttRLAdd (None)'] + hourly_df10637['VS.SHO.AttRLDel (None)']) * 100
hourly_df10637['Hard Handover Success rate, %_10637'] = hourly_df10637['VS.HHO.SuccInterFreqOut (None)'] / hourly_df10637['VS.HHO.AttInterFreqOut (None)'] * 100
hourly_df10637['CS W2G Inter-RAT Handover Out SR_10637'] = hourly_df10637['IRATHO.SuccOutCS (None)'] / (hourly_df10637['IRATHO.AttOutCS (None)'] - hourly_df10637['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
hourly_df10637['RRC Assignment SucessRate (CS BH), %_10637'] = hourly_df10637['RRC.SuccConnEstab.sum (None)'] / hourly_df10637['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_df10637['RRC Assignment SucessRate (PS BH), %_10637'] = hourly_df10637['RRC.SuccConnEstab.sum (None)'] / hourly_df10637['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_df10637['RRC Drop Rate (CS BH), %_10637'] = (hourly_df10637['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df10637['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df10637['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_df10637['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_df10637['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_df10637['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_df10637['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_df10637['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_df10637['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_df10637['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_df10637['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df10637['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df10637['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_df10637['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Norm (None)'] + hourly_df10637['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_df10637['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_df10637['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df10637['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_df10637['VS.DCCC.Succ.F2P (None)'] + hourly_df10637['IRATHO.SuccOutCS (None)'] + hourly_df10637['IRATHO.SuccOutPSUTRAN (None)'] + hourly_df10637['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_df10637['VS.DCCC.Succ.D2U (None)']) * 100
hourly_df10637['RRC Drop Rate (PS BH), %_10637'] = (hourly_df10637['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df10637['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df10637['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_df10637['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_df10637['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_df10637['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_df10637['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_df10637['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_df10637['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_df10637['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_df10637['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df10637['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df10637['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_df10637['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_df10637['RRC.AttConnRelDCCH.Norm (None)'] + hourly_df10637['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_df10637['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_df10637['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df10637['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_df10637['VS.DCCC.Succ.F2P (None)'] + hourly_df10637['IRATHO.SuccOutCS (None)'] + hourly_df10637['IRATHO.SuccOutPSUTRAN (None)'] + hourly_df10637['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_df10637['VS.DCCC.Succ.D2U (None)']) * 100
hourly_df10637['RAB Assignment Success Rate (CS), %_10637'] = hourly_df10637['VS.RAB.SuccEstabCS.AMR (None)'] / hourly_df10637['VS.RAB.AttEstab.AMR (None)'] * 100
hourly_df10637['RAB Assignment Success Rate (PS), %_10637'] = (hourly_df10637['VS.RAB.SuccEstabPS.Conv (None)'] + hourly_df10637['VS.RAB.SuccEstabPS.Bkg (None)'] + hourly_df10637['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   hourly_df10637['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (hourly_df10637['VS.RAB.AttEstabPS.Bkg (None)'] + hourly_df10637['VS.RAB.AttEstabPS.Int (None)'] + hourly_df10637['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   hourly_df10637['VS.RAB.AttEstabPS.Conv (None)']) * 100
hourly_df10637['CCSR3G, %_10637'] = hourly_df10637['RRC Assignment SucessRate (CS BH), %_10637'] * (100 - hourly_df10637['RRC Drop Rate (CS BH), %_10637']) * hourly_df10637['RAB Assignment Success Rate (CS), %_10637'] * (100 - hourly_df10637['CS RAB Drop Rate (%)_10637'])/ 1000000
hourly_df10637['DCSR3G, %_10637'] = hourly_df10637['RRC Assignment SucessRate (PS BH), %_10637'] * (100 - hourly_df10637['RRC Drop Rate (PS BH), %_10637']) * hourly_df10637['RAB Assignment Success Rate (PS), %_10637'] * (100 - hourly_df10637['PS RAB Drop Rate (%)_10637'])/ 1000000
hourly_df10637 = hourly_df10637.drop(list_1, axis=1)
# фильтрация по 2937
hourly_df2937 = sts_df[sts_df['BSC6910UCell'].isin(list_F3_2937)]
hourly_df2937 = hourly_df2937.groupby(['date', 'hour'])[list_1]. sum().reset_index()
hourly_df2937['CS traffic 3G, Erl_2937'] = hourly_df2937['CS Voice Traffic Volume (Erl)']
hourly_df2937['PS traffic 3G UL+DL, GB_2937'] = (hourly_df2937['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + hourly_df2937['VS.PS.Bkg.DL.8.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Bkg.DL.32.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.DL.64.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Bkg.DL.144.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.DL.256.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Bkg.UL.8.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.UL.16.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Bkg.UL.64.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.UL.128.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Bkg.UL.256.Traffic (bit)'] + hourly_df2937['VS.PS.Bkg.UL.384.Traffic (bit)'] + hourly_df2937['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Int.DL.16.Traffic (bit)'] + hourly_df2937['VS.PS.Int.DL.32.Traffic (bit)'] + hourly_df2937['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Int.DL.128.Traffic (bit)'] + hourly_df2937['VS.PS.Int.DL.144.Traffic (bit)'] + hourly_df2937['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Int.DL.384.Traffic (bit)'] + hourly_df2937['VS.PS.Int.UL.8.Traffic (bit)'] + hourly_df2937['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Int.UL.32.Traffic (bit)'] + hourly_df2937['VS.PS.Int.UL.64.Traffic (bit)'] + hourly_df2937['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Int.UL.144.Traffic (bit)'] + hourly_df2937['VS.PS.Int.UL.256.Traffic (bit)'] + hourly_df2937['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Str.DL.32.Traffic (bit)'] + hourly_df2937['VS.PS.Str.DL.64.Traffic (bit)'] + hourly_df2937['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Str.DL.144.Traffic (bit)'] + hourly_df2937['VS.PS.Str.UL.16.Traffic (bit)'] + hourly_df2937['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      hourly_df2937['VS.PS.Str.UL.64.Traffic (bit)'] + hourly_df2937['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
hourly_df2937['CS RAB Drop Rate (%)_2937'] = hourly_df2937['VS.RAB.AbnormRel.CS (None)'] / (hourly_df2937['VS.RAB.AbnormRel.CS (None)'] + hourly_df2937['VS.RAB.NormRel.CS (None)']) * 100
hourly_df2937['PS Blocking Rate (%)_2937'] = (hourly_df2937['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + hourly_df2937['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + hourly_df2937['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    hourly_df2937['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + hourly_df2937['VS.RAB.FailEstabPS.Code.Cong (None)'] + hourly_df2937['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    hourly_df2937['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + hourly_df2937['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + hourly_df2937['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (hourly_df2937['VS.RAB.AttEstabPS.Str (None)'] + hourly_df2937['VS.RAB.AttEstabPS.Int (None)'] + hourly_df2937['VS.RAB.AttEstabPS.Bkg (None)']) *100
hourly_df2937['PS RAB Drop Rate (%)_2937'] = (hourly_df2937['VS.RAB.AbnormRel.PS (None)'] + hourly_df2937['VS.RAB.AbnormRel.PS.PCH (None)'] + hourly_df2937['VS.RAB.AbnormRel.PS.D2P (None)'] + \
                                    hourly_df2937['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (hourly_df2937['VS.RAB.AbnormRel.PS (None)'] + hourly_df2937['VS.RAB.NormRel.PS (None)'] + hourly_df2937['VS.RAB.AbnormRel.PS.PCH (None)'] + \
                                    hourly_df2937['VS.RAB.NormRel.PS.PCH (None)']) * 100
hourly_df2937['PS HS- Drop Rate (%)_2937'] =  hourly_df2937['VS.HSDPA.RAB.AbnormRel (None)'] / (hourly_df2937['VS.HSDPA.RAB.AbnormRel (None)'] + hourly_df2937['VS.HSDPA.RAB.NormRel (None)'] + hourly_df2937['VS.HSDPA.H2D.Succ (None)'] + hourly_df2937['VS.HSDPA.H2F.Succ (None)'] +hourly_df2937['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + hourly_df2937['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
hourly_df2937['HSDPA Throughput, kbps_2937'] = hourly_df2937['VS.HSDPA.MeanChThroughput (kbit/s)'] / 204 / 24 # количество сот 204!!!
hourly_df2937['HSUPA Throughput, kbps_2937'] = hourly_df2937['VS.HSUPA.MeanChThroughput (kbit/s)'] / 204 / 24# количество сот 204
hourly_df2937['Soft Handover Success rate, %_2937'] = (hourly_df2937['VS.SHO.SuccRLAdd (None)'] + hourly_df2937['VS.SHO.SuccRLDel (None)']) / (hourly_df2937['VS.SHO.AttRLAdd (None)'] + hourly_df2937['VS.SHO.AttRLDel (None)']) * 100
hourly_df2937['Hard Handover Success rate, %_2937'] = hourly_df2937['VS.HHO.SuccInterFreqOut (None)'] / hourly_df2937['VS.HHO.AttInterFreqOut (None)'] * 100
hourly_df2937['CS W2G Inter-RAT Handover Out SR_2937'] = hourly_df2937['IRATHO.SuccOutCS (None)'] / (hourly_df2937['IRATHO.AttOutCS (None)'] - hourly_df2937['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
hourly_df2937['RRC Assignment SucessRate (CS BH), %_2937'] = hourly_df2937['RRC.SuccConnEstab.sum (None)'] / hourly_df2937['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_df2937['RRC Assignment SucessRate (PS BH), %_2937'] = hourly_df2937['RRC.SuccConnEstab.sum (None)'] / hourly_df2937['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourly_df2937['RRC Drop Rate (CS BH), %_2937'] = (hourly_df2937['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df2937['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df2937['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_df2937['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_df2937['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_df2937['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_df2937['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_df2937['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_df2937['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_df2937['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_df2937['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df2937['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df2937['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_df2937['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Norm (None)'] + hourly_df2937['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_df2937['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_df2937['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df2937['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_df2937['VS.DCCC.Succ.F2P (None)'] + hourly_df2937['IRATHO.SuccOutCS (None)'] + hourly_df2937['IRATHO.SuccOutPSUTRAN (None)'] + hourly_df2937['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_df2937['VS.DCCC.Succ.D2U (None)']) * 100
hourly_df2937['RRC Drop Rate (PS BH), %_2937'] = (hourly_df2937['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df2937['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df2937['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourly_df2937['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourly_df2937['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Unspec (None)'] + hourly_df2937['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourly_df2937['RRC.AttConnRelDCCH.Cong (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Preempt (None)'] + hourly_df2937['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourly_df2937['RRC.AttConnRelDCCH.DSCR (None)'] + hourly_df2937['RRC.AttConnRelDCCH.UsrInact (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourly_df2937['RRC.AttConnRelCCCH.Cong (None)'] + hourly_df2937['RRC.AttConnRelCCCH.Preempt (None)'] + hourly_df2937['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourly_df2937['RRC.AttConnRelCCCH.DSCR (None)'] + hourly_df2937['RRC.AttConnRelDCCH.Norm (None)'] + hourly_df2937['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourly_df2937['RRC.AttConnRelCCCH.UsrInact (None)'] + hourly_df2937['RRC.AttConnRelCCCH.Unspec (None)'] + hourly_df2937['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourly_df2937['VS.DCCC.Succ.F2P (None)'] + hourly_df2937['IRATHO.SuccOutCS (None)'] + hourly_df2937['IRATHO.SuccOutPSUTRAN (None)'] + hourly_df2937['VS.DCCC.Succ.F2U (None)'] + \
                                           hourly_df2937['VS.DCCC.Succ.D2U (None)']) * 100
hourly_df2937['RAB Assignment Success Rate (CS), %_2937'] = hourly_df2937['VS.RAB.SuccEstabCS.AMR (None)'] / hourly_df2937['VS.RAB.AttEstab.AMR (None)'] * 100
hourly_df2937['RAB Assignment Success Rate (PS), %_2937'] = (hourly_df2937['VS.RAB.SuccEstabPS.Conv (None)'] + hourly_df2937['VS.RAB.SuccEstabPS.Bkg (None)'] + hourly_df2937['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   hourly_df2937['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (hourly_df2937['VS.RAB.AttEstabPS.Bkg (None)'] + hourly_df2937['VS.RAB.AttEstabPS.Int (None)'] + hourly_df2937['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   hourly_df2937['VS.RAB.AttEstabPS.Conv (None)']) * 100
hourly_df2937['CCSR3G, %_2937'] = hourly_df2937['RRC Assignment SucessRate (CS BH), %_2937'] * (100 - hourly_df2937['RRC Drop Rate (CS BH), %_2937']) * hourly_df2937['RAB Assignment Success Rate (CS), %_2937'] * (100 - hourly_df2937['CS RAB Drop Rate (%)_2937'])/ 1000000
hourly_df2937['DCSR3G, %_2937'] = hourly_df2937['RRC Assignment SucessRate (PS BH), %_2937'] * (100 - hourly_df2937['RRC Drop Rate (PS BH), %_2937']) * hourly_df2937['RAB Assignment Success Rate (PS), %_2937'] * (100 - hourly_df2937['PS RAB Drop Rate (%)_2937'])/ 1000000
hourly_df2937 = hourly_df2937.drop(list_1, axis=1)

hourly_dfall = pd.merge(hourly_df, hourly_dfU2100, how="left")
hourly_dfall = pd.merge(hourly_dfall, hourly_dfU900, how="left")
hourly_dfall = pd.merge(hourly_dfall, hourly_df10612, how="left")
hourly_dfall = pd.merge(hourly_dfall, hourly_df10637, how="left")
hourly_dfall = pd.merge(hourly_dfall, hourly_df2937, how="left")
hourly_dfall_trans = hourly_dfall.transpose()

# NodeB часовая
hourlyN_df = stsN_df.groupby(['date', 'hour'])[list_1N]. sum().reset_index()
hourlyN_df['MeanThrHSDPA,kbps'] = hourlyN_df['VS.HSDPA.DataOutput.Traffic (bit)']/hourlyN_df['VS.HSDPA.DataTtiNum.User (None)'] / 2
hourlyN_df['MeanThrHSDPA DC,kbps'] = hourlyN_df['VS.DataOutput.AllHSDPA.Traffic (bit)'] / hourlyN_df['VS.AllHSDPA.DataTtiNum.User (None)'] / 2
hourlyN_df['MeanThrHSUPA,kbps'] = (hourlyN_df['VS.HSUPA.2msTTI.Traffic (kbit)'] + hourlyN_df['VS.HSUPA.10msTTI.Traffic (kbit)']) / (hourlyN_df['VS.HSUPA.2msPDU.TTI.Num (None)'] * 0.002 + hourlyN_df['VS.HSUPA.10msPDU.TTI.Num (None)'] * 0.01)
hourlyN_df = hourlyN_df.drop(list_1N, axis=1)

# сортировка по диапазонам
hourlyN_dfU2100 = stsN_df[stsN_df['ULoCell'].isin(list_U2100N)]
hourlyN_dfU2100 = hourlyN_dfU2100.groupby(['date', 'hour'])[list_1N]. sum().reset_index()
hourlyN_dfU2100['MeanThrHSDPAU2100,kbps'] = hourlyN_dfU2100['VS.HSDPA.DataOutput.Traffic (bit)']/hourlyN_dfU2100['VS.HSDPA.DataTtiNum.User (None)'] / 2
hourlyN_dfU2100['MeanThrHSDPAU2100 DC,kbps'] = hourlyN_dfU2100['VS.DataOutput.AllHSDPA.Traffic (bit)'] / hourlyN_dfU2100['VS.AllHSDPA.DataTtiNum.User (None)'] / 2
hourlyN_dfU2100['MeanThrHSUPAU2100,kbps'] = (hourlyN_dfU2100['VS.HSUPA.2msTTI.Traffic (kbit)'] + hourlyN_dfU2100['VS.HSUPA.10msTTI.Traffic (kbit)']) / (hourlyN_dfU2100['VS.HSUPA.2msPDU.TTI.Num (None)'] * 0.002 + hourlyN_dfU2100['VS.HSUPA.10msPDU.TTI.Num (None)'] * 0.01)
hourlyN_dfU2100 = hourlyN_dfU2100.drop(list_1N, axis=1)

hourlyN_dfU900 = stsN_df[stsN_df['ULoCell'].isin(list_U900N)]
hourlyN_dfU900 = hourlyN_dfU900.groupby(['date', 'hour'])[list_1N]. sum().reset_index()
hourlyN_dfU900['MeanThrHSDPAU900,kbps'] = hourlyN_dfU900['VS.HSDPA.DataOutput.Traffic (bit)']/hourlyN_dfU900['VS.HSDPA.DataTtiNum.User (None)'] / 2
hourlyN_dfU900['MeanThrHSDPAU900 DC,kbps'] = hourlyN_dfU900['VS.DataOutput.AllHSDPA.Traffic (bit)'] / hourlyN_dfU900['VS.AllHSDPA.DataTtiNum.User (None)'] / 2
hourlyN_dfU900['MeanThrHSUPAU900,kbps'] = (hourlyN_dfU900['VS.HSUPA.2msTTI.Traffic (kbit)'] + hourlyN_dfU900['VS.HSUPA.10msTTI.Traffic (kbit)']) / (hourlyN_dfU900['VS.HSUPA.2msPDU.TTI.Num (None)'] * 0.002 + hourlyN_dfU900['VS.HSUPA.10msPDU.TTI.Num (None)'] * 0.01)
hourlyN_dfU900 = hourlyN_dfU900.drop(list_1N, axis=1)

hourlyN_df = pd.merge(hourlyN_df, hourlyN_dfU2100, how="left")
hourlyN_df = pd.merge(hourlyN_df, hourlyN_dfU900, how="left")
#dailyN_df_trans = dailyN_df.transpose()



#####
# обработка busy hour
hourly1_df = sts_df.groupby(['date', 'hour'])[list_1].sum().reset_index()
max_index_PS = hourly1_df.groupby('date')['VS.HSDPA.MeanChThroughput.TotalBytes (byte)'].idxmax()
hourlyPS_df = hourly1_df.loc[max_index_PS]
max_index_CS = hourly1_df.groupby('date')['CS Voice Traffic Volume (Erl)'].idxmax()
hourlyCS_df = hourly1_df.loc[max_index_CS]
hourlyCS_df['CS traffic 3G, Erl'] = hourlyCS_df['CS Voice Traffic Volume (Erl)']
hourlyPS_df['PS traffic 3G UL+DL, GB'] = (hourlyPS_df['VS.HSUPA.MeanChThroughput.TotalBytes (byte)'] + hourlyPS_df['VS.PS.Bkg.DL.8.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.DL.16.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Bkg.DL.32.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.DL.64.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.DL.128.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Bkg.DL.144.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.DL.256.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.DL.384.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Bkg.UL.8.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.UL.16.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.UL.32.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Bkg.UL.64.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.UL.128.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.UL.144.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Bkg.UL.256.Traffic (bit)'] + hourlyPS_df['VS.PS.Bkg.UL.384.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.DL.8.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Int.DL.16.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.DL.32.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.DL.64.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Int.DL.128.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.DL.144.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.DL.256.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Int.DL.384.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.UL.8.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.UL.16.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Int.UL.32.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.UL.64.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.UL.128.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Int.UL.144.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.UL.256.Traffic (bit)'] + hourlyPS_df['VS.PS.Int.UL.384.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Str.DL.32.Traffic (bit)'] + hourlyPS_df['VS.PS.Str.DL.64.Traffic (bit)'] + hourlyPS_df['VS.PS.Str.DL.128.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Str.DL.144.Traffic (bit)'] + hourlyPS_df['VS.PS.Str.UL.16.Traffic (bit)'] + hourlyPS_df['VS.PS.Str.UL.32.Traffic (bit)'] + \
                                      hourlyPS_df['VS.PS.Str.UL.64.Traffic (bit)'] + hourlyPS_df['VS.HSDPA.MeanChThroughput.TotalBytes (byte)']) / 1024/1024/1024
hourlyCS_df['CS RAB Drop Rate (%)'] = hourlyCS_df['VS.RAB.AbnormRel.CS (None)'] / (hourlyCS_df['VS.RAB.AbnormRel.CS (None)'] + hourlyCS_df['VS.RAB.NormRel.CS (None)']) * 100
hourlyPS_df['PS Blocking Rate (%)'] = (hourlyPS_df['VS.RAB.FailEstabPS.DLIUBBand.Cong (None)'] + hourlyPS_df['VS.RAB.FailEstabPS.ULIUBBand.Cong (None)'] + hourlyPS_df['VS.RAB.FailEstabPS.ULCE.Cong (None)'] + \
                                    hourlyPS_df['VS.RAB.FailEstabPS.DLCE.Cong (None)'] + hourlyPS_df['VS.RAB.FailEstabPS.Code.Cong (None)'] + hourlyPS_df['VS.RAB.FailEstabPS.ULPower.Cong (None)'] + \
                                    hourlyPS_df['VS.RAB.FailEstabPS.DLPower.Cong (None)'] + hourlyPS_df['VS.RAB.FailEstabPS.HSDPAUser.Cong (None)'] + hourlyPS_df['VS.RAB.FailEstabPS.HSUPAUser.Cong (None)']) / \
                                    (hourlyPS_df['VS.RAB.AttEstabPS.Str (None)'] + hourlyPS_df['VS.RAB.AttEstabPS.Int (None)'] + hourlyPS_df['VS.RAB.AttEstabPS.Bkg (None)']) *100
# hourlyPS_df['PS RAB Drop Rate (%)'] = (hourlyPS_df['VS.RAB.AbnormRel.PS (None)'] + hourlyPS_df['VS.RAB.AbnormRel.PS.PCH (None)'] + hourlyPS_df['VS.RAB.AbnormRel.PS.D2P (None)'] + \
#                                     hourlyPS_df['VS.RAB.AbnormRel.PS.F2P (None)']) / \
#                                    (hourlyPS_df['VS.RAB.AbnormRel.PS (None)'] + hourlyPS_df['VS.RAB.NormRel.PS (None)'] + hourlyPS_df['VS.RAB.AbnormRel.PS.PCH (None)'] + \
#                                     hourlyPS_df['VS.RAB.NormRel.PS.PCH (None)']) * 100
hourlyPS_df['PS RAB Drop Rate (%)'] = (hourlyPS_df['VS.RAB.AbnormRel.PS (None)'] - hourlyPS_df['VS.RAB.AbnormRel.PS.PCH (None)'] - hourlyPS_df['VS.RAB.AbnormRel.PS.D2P (None)'] - \
                                    hourlyPS_df['VS.RAB.AbnormRel.PS.F2P (None)']) / \
                                   (hourlyPS_df['VS.RAB.AbnormRel.PS (None)'] + hourlyPS_df['VS.RAB.NormRel.PS (None)'] - hourlyPS_df['VS.RAB.AbnormRel.PS.PCH (None)'] - \
                                    hourlyPS_df['VS.RAB.NormRel.PS.PCH (None)'] + hourlyPS_df['VS.DCCC.D2P.Succ (None)']+hourlyPS_df['VS.DCCC.Succ.F2P (None)']+hourlyPS_df['VS.DCCC.Succ.F2U (None)']+hourlyPS_df['VS.DCCC.Succ.D2U (None)']) * 100


hourlyPS_df['PS HS- Drop Rate (%)'] =  hourlyPS_df['VS.HSDPA.RAB.AbnormRel (None)'] / (hourlyPS_df['VS.HSDPA.RAB.AbnormRel (None)'] + hourlyPS_df['VS.HSDPA.RAB.NormRel (None)'] + hourlyPS_df['VS.HSDPA.H2D.Succ (None)'] + hourlyPS_df['VS.HSDPA.H2F.Succ (None)'] +hourlyPS_df['VS.HSDPA.HHO.H2D.SuccOutIntraFreq (None)'] + hourlyPS_df['VS.HSDPA.HHO.H2D.SuccOutInterFreq (None)']) * 100
hourlyPS_df['HSDPA Throughput, kbps'] = hourlyPS_df['VS.HSDPA.MeanChThroughput (kbit/s)'] / active_cell_number # количество сот
hourlyPS_df['HSUPA Throughput, kbps'] = hourlyPS_df['VS.HSUPA.MeanChThroughput (kbit/s)'] / active_cell_number # количество сот
hourlyCS_df['Soft Handover Success rate, %'] = (hourlyCS_df['VS.SHO.SuccRLAdd (None)'] + hourlyCS_df['VS.SHO.SuccRLDel (None)']) / (hourlyCS_df['VS.SHO.AttRLAdd (None)'] + hourlyCS_df['VS.SHO.AttRLDel (None)']) * 100
hourlyCS_df['Hard Handover Success rate, %'] = hourlyCS_df['VS.HHO.SuccInterFreqOut (None)'] / hourlyCS_df['VS.HHO.AttInterFreqOut (None)'] * 100
hourlyCS_df['CS W2G Inter-RAT Handover Out SR'] = hourlyCS_df['IRATHO.SuccOutCS (None)'] / (hourlyCS_df['IRATHO.AttOutCS (None)'] - hourlyCS_df['VS.IRATHOCS.Cancel.ReEstab (None)']) * 100
hourlyCS_df['RRC Assignment SucessRate (CS BH), %'] = hourlyCS_df['RRC.SuccConnEstab.sum (None)'] / hourlyCS_df['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourlyPS_df['RRC Assignment SucessRate (PS BH), %'] = hourlyPS_df['RRC.SuccConnEstab.sum (None)'] / hourlyPS_df['VS.RRC.AttConnEstab.Sum (None)'] * 100
hourlyCS_df['RRC Drop Rate (CS BH), %'] = (hourlyCS_df['RRC.AttConnRelCCCH.Cong (None)'] + hourlyCS_df['RRC.AttConnRelCCCH.Preempt (None)'] + hourlyCS_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourlyCS_df['RRC.AttConnRelCCCH.Unspec (None)'] + hourlyCS_df['RRC.AttConnRelDCCH.Cong (None)'] + hourlyCS_df['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourlyCS_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourlyCS_df['RRC.AttConnRelDCCH.Unspec (None)'] + hourlyCS_df['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourlyCS_df['RRC.AttConnRelDCCH.Cong (None)'] + hourlyCS_df['RRC.AttConnRelDCCH.Preempt (None)'] + hourlyCS_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourlyCS_df['RRC.AttConnRelDCCH.DSCR (None)'] + hourlyCS_df['RRC.AttConnRelDCCH.UsrInact (None)'] + hourlyCS_df['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourlyCS_df['RRC.AttConnRelCCCH.Cong (None)'] + hourlyCS_df['RRC.AttConnRelCCCH.Preempt (None)'] + hourlyCS_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourlyCS_df['RRC.AttConnRelCCCH.DSCR (None)'] + hourlyCS_df['RRC.AttConnRelDCCH.Norm (None)'] + hourlyCS_df['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourlyCS_df['RRC.AttConnRelCCCH.UsrInact (None)'] + hourlyCS_df['RRC.AttConnRelCCCH.Unspec (None)'] + hourlyCS_df['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourlyCS_df['VS.DCCC.Succ.F2P (None)'] + hourlyCS_df['IRATHO.SuccOutCS (None)'] + hourlyCS_df['IRATHO.SuccOutPSUTRAN (None)'] + hourlyCS_df['VS.DCCC.Succ.F2U (None)'] + \
                                           hourlyCS_df['VS.DCCC.Succ.D2U (None)']) * 100
hourlyPS_df['RRC Drop Rate (PS BH), %'] = (hourlyPS_df['RRC.AttConnRelCCCH.Cong (None)'] + hourlyPS_df['RRC.AttConnRelCCCH.Preempt (None)'] + hourlyPS_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                             hourlyPS_df['RRC.AttConnRelCCCH.Unspec (None)'] + hourlyPS_df['RRC.AttConnRelDCCH.Cong (None)'] + hourlyPS_df['RRC.AttConnRelDCCH.Preempt (None)'] + \
                                             hourlyPS_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + hourlyPS_df['RRC.AttConnRelDCCH.Unspec (None)'] + hourlyPS_df['VS.RRC.ConnRel.CellUpd (None)']) \
                                        / (hourlyPS_df['RRC.AttConnRelDCCH.Cong (None)'] + hourlyPS_df['RRC.AttConnRelDCCH.Preempt (None)'] + hourlyPS_df['RRC.AttConnRelDCCH.ReEstRej (None)'] + \
                                           hourlyPS_df['RRC.AttConnRelDCCH.DSCR (None)'] + hourlyPS_df['RRC.AttConnRelDCCH.UsrInact (None)'] + hourlyPS_df['RRC.AttConnRelDCCH.Unspec (None)'] + \
                                           hourlyPS_df['RRC.AttConnRelCCCH.Cong (None)'] + hourlyPS_df['RRC.AttConnRelCCCH.Preempt (None)'] + hourlyPS_df['RRC.AttConnRelCCCH.ReEstRej (None)'] + \
                                           hourlyPS_df['RRC.AttConnRelCCCH.DSCR (None)'] + hourlyPS_df['RRC.AttConnRelDCCH.Norm (None)'] + hourlyPS_df['RRC.AttConnRelCCCH.Norm (None)'] + \
                                           hourlyPS_df['RRC.AttConnRelCCCH.UsrInact (None)'] + hourlyPS_df['RRC.AttConnRelCCCH.Unspec (None)'] + hourlyPS_df['VS.RRC.ConnRel.CellUpd (None)'] + \
                                           hourlyPS_df['VS.DCCC.Succ.F2P (None)'] + hourlyPS_df['IRATHO.SuccOutCS (None)'] + hourlyPS_df['IRATHO.SuccOutPSUTRAN (None)'] + hourlyPS_df['VS.DCCC.Succ.F2U (None)'] + \
                                           hourlyPS_df['VS.DCCC.Succ.D2U (None)']) * 100
hourlyCS_df['RAB Assignment Success Rate (CS), %'] = hourlyCS_df['VS.RAB.SuccEstabCS.AMR (None)'] / hourlyCS_df['VS.RAB.AttEstab.AMR (None)'] * 100
hourlyPS_df['RAB Assignment Success Rate (PS), %'] = (hourlyPS_df['VS.RAB.SuccEstabPS.Conv (None)'] + hourlyPS_df['VS.RAB.SuccEstabPS.Bkg (None)'] + hourlyPS_df['VS.RAB.SuccEstabPS.Int (None)'] + \
                                                   hourlyPS_df['VS.RAB.SuccEstabPS.Str (None)']) / \
                                                  (hourlyPS_df['VS.RAB.AttEstabPS.Bkg (None)'] + hourlyPS_df['VS.RAB.AttEstabPS.Int (None)'] + hourlyPS_df['VS.RAB.AttEstabPS.Str (None)'] + \
                                                   hourlyPS_df['VS.RAB.AttEstabPS.Conv (None)']) * 100
hourlyCS_df['CCSR3G, %'] = hourlyCS_df['RRC Assignment SucessRate (CS BH), %'] * (100 - hourlyCS_df['RRC Drop Rate (CS BH), %']) * hourlyCS_df['RAB Assignment Success Rate (CS), %'] * (100 - hourlyCS_df['CS RAB Drop Rate (%)'])/ 1000000
hourlyPS_df['DCSR3G, %'] = hourlyPS_df['RRC Assignment SucessRate (PS BH), %'] * (100 - hourlyPS_df['RRC Drop Rate (PS BH), %']) * hourlyPS_df['RAB Assignment Success Rate (PS), %'] * (100 - hourlyPS_df['PS RAB Drop Rate (%)'])/ 1000000
hourlyCS_df = hourlyCS_df.drop(list_1, axis=1)
hourlyCS_df_trans = hourlyCS_df.transpose()
hourlyPS_df = hourlyPS_df.drop(list_1, axis=1)
hourlyPS_df_trans = hourlyPS_df.transpose()

# NodeB статистика в чнн
hourlyPSN_df = stsN_df.groupby(['date', 'hour'])[list_1N]. sum().reset_index()
max_index_PS = hourlyPSN_df.groupby('date')['VS.HSDPA.DataOutput.Traffic (bit)'].idxmax()
hourlyPSN_df = hourlyPSN_df.loc[max_index_PS]
hourlyPSN_df['MeanThrHSDPA,kbps'] = hourlyPSN_df['VS.HSDPA.DataOutput.Traffic (bit)']/hourlyPSN_df['VS.HSDPA.DataTtiNum.User (None)'] / 2
hourlyPSN_df['MeanThrHSDPA DC,kbps'] = hourlyPSN_df['VS.DataOutput.AllHSDPA.Traffic (bit)'] / hourlyPSN_df['VS.AllHSDPA.DataTtiNum.User (None)'] / 2
hourlyPSN_df['MeanThrHSUPA,kbps'] = (hourlyPSN_df['VS.HSUPA.2msTTI.Traffic (kbit)'] + hourlyPSN_df['VS.HSUPA.10msTTI.Traffic (kbit)']) / (hourlyPSN_df['VS.HSUPA.2msPDU.TTI.Num (None)'] * 0.002 + hourlyPSN_df['VS.HSUPA.10msPDU.TTI.Num (None)'] * 0.01)
hourlyPSN_df = hourlyPSN_df.drop(list_1N, axis=1)
#hourlyPSN_df_trans = hourlyPSN_df.transpose()



#daily_df.to_excel("C:/test/sts3G/daily_df.xls", engine='openpyxl', sheet_name='Book2')
#hourlyCS_df.to_excel("C:/test/sts3G/hourly_df.xls", engine='openpyxl', sheet_name='Book2')

with pd.ExcelWriter(f"{directory}{csv_name1}{output_comment}.xlsx", engine='openpyxl') as writer:
    weekly_df.to_excel(writer, sheet_name='weekly')
    daily_dfall.to_excel(writer, sheet_name='daily')
    hourly_dfall.to_excel(writer, sheet_name='hourly')
    hourlyCS_df.to_excel(writer, sheet_name='busy_hourCS')
    hourlyPS_df.to_excel(writer, sheet_name='busy_hourPS')
    weeklyN_df.to_excel(writer, sheet_name='NodeBweekly')
    dailyN_df.to_excel(writer, sheet_name='NodeBdaily')
    hourlyN_df.to_excel(writer, sheet_name='NodeBhourly')
    hourlyPSN_df.to_excel(writer, sheet_name='NodeB_BH')
    weekly_df_trans.to_excel(writer, sheet_name='weekly_trans')
#    daily_dfall_trans.to_excel(writer, sheet_name='daily_trans')
#    hourly_dfall_trans.to_excel(writer, sheet_name='hourly_trans')
#     hourlyCS_df_trans.to_excel(writer, sheet_name='busy_hourCS_trans')
#     hourlyPS_df_trans.to_excel(writer, sheet_name='busy_hourPS_trans')

''' переходим к работе с эксель файлом - форматирование строк и добавление графиков
    используем модуль openpyxl'''

my_file = openpyxl.load_workbook(f"{directory}{csv_name1}{output_comment}.xlsx")

weekly_sheet = my_file["weekly"]
daily_sheet = my_file["daily"]
hourly_sheet = my_file["hourly"]
busy_hourCS_sheet = my_file["busy_hourCS"]
busy_hourPS_sheet = my_file["busy_hourPS"]
# weekly_sheet_trans = my_file["weekly_trans"]
# daily_sheet_trans = my_file["daily_trans"]
# hourly_sheet_trans = my_file["hourly_trans"]
# busy_hourCS_sheet_trans = my_file["busy_hourCS_trans"]
# busy_hourPS_sheet_trans = my_file["busy_hourPS_trans"]
weeklyN_sheet = my_file["NodeBweekly"]
dailyN_sheet = my_file["NodeBdaily"]
hourlyN_sheet = my_file["NodeBhourly"]
hourlyPSN_sheet = my_file["NodeB_BH"]


weekly_sheet.column_dimensions["A"].width = 2
weekly_sheet.column_dimensions["B"].width = 7
daily_sheet.column_dimensions["A"].width = 2
daily_sheet.column_dimensions["B"].width = 11
hourly_sheet.column_dimensions["A"].width = 4
busy_hourCS_sheet.column_dimensions["A"].width = 11
busy_hourCS_sheet.column_dimensions["B"].width = 11
busy_hourPS_sheet.column_dimensions["A"].width = 11
busy_hourPS_sheet.column_dimensions["B"].width = 11

weeklyN_sheet.column_dimensions["A"].width = 2
weeklyN_sheet.column_dimensions["B"].width = 7
dailyN_sheet.column_dimensions["A"].width = 2
dailyN_sheet.column_dimensions["B"].width = 11
hourlyN_sheet.column_dimensions["A"].width = 4
hourlyPSN_sheet.column_dimensions["A"].width = 11
hourlyPSN_sheet.column_dimensions["B"].width = 11


# weekly_sheet_trans.column_dimensions["A"].width = 35
# daily_sheet_trans.column_dimensions["A"].width = 35
# hourly_sheet_trans.column_dimensions["A"].width = 35
# busy_hourCS_sheet_trans.column_dimensions["A"].width = 35
# busy_hourPS_sheet_trans.column_dimensions["A"].width = 35

hourly_sheet.delete_cols(1) # удаляем первые столбцы чтобы номера столбцов для всех KPI были одинаковыми как в дневной статистике
busy_hourCS_sheet.delete_cols(1) # удаляем первые столбцы чтобы номера столбцов для всех KPI были одинаковыми как в дневной статистике
busy_hourPS_sheet.delete_cols(1) # удаляем первые столбцы чтобы номера столбцов для всех KPI были одинаковыми как в дневной статистике

hourlyN_sheet.delete_cols(1) # удаляем первые столбцы чтобы номера столбцов для всех KPI были одинаковыми как в дневной статистике
hourlyPSN_sheet.delete_cols(1) # удаляем первые столбцы чтобы номера столбцов для всех KPI были одинаковыми как в дневной статистике


# определение количества строк в таблицах
last_row_weekly = weekly_sheet.max_row
last_row_daily = daily_sheet.max_row
last_row_hourly = hourly_sheet.max_row
last_row_BHCS = busy_hourCS_sheet.max_row
last_row_BHPS = busy_hourPS_sheet.max_row

last_row_weeklyN = weeklyN_sheet.max_row
last_row_dailyN = dailyN_sheet.max_row
last_row_hourlyN = hourlyN_sheet.max_row
last_row_BHPSN = hourlyPSN_sheet.max_row

# выставление правильного формата для столбцов с датами
for r in range(2,(last_row_daily+1)):
    daily_sheet[f'B{r}'].number_format ='DD.MM.YYYY'
for r in range(2, (last_row_hourly+1)):
    hourly_sheet[f'A{r}'].number_format ='DD'
for r in range(2,(last_row_BHCS+1)):
    busy_hourCS_sheet[f'A{r}'].number_format ='DD.MM.YYYY'
for r in range(2,(last_row_BHPS+1)):
    busy_hourPS_sheet[f'A{r}'].number_format = 'DD.MM.YYYY'

for r in range(2,(last_row_dailyN+1)):
    dailyN_sheet[f'B{r}'].number_format ='DD.MM.YYYY'
for r in range(2, (last_row_hourlyN+1)):
    hourlyN_sheet[f'A{r}'].number_format ='DD'
for r in range(2,(last_row_BHPSN+1)):
    hourlyPSN_sheet[f'A{r}'].number_format ='DD.MM.YYYY'

# for cell in daily_sheet_trans[2]:
#     cell.number_format ='DD.MM.YYYY'

#row = daily_sheet_trans[2]
#row_dimension = row[0].row_dimension
#row_dimension.width = 20


# выставление переноса строк для названий KPI
for cell in weekly_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in daily_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in hourly_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in busy_hourCS_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in busy_hourPS_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

for cell in weeklyN_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in dailyN_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in hourlyN_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in hourlyPSN_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

#  графики в недельной таблице weekly_sheet
x_values = Reference(weekly_sheet, range_string=(f"weekly!$B$2:$B${last_row_weekly}"))

CStraffic3GErl= Reference(weekly_sheet, min_col=3, min_row=1, max_row=last_row_weekly)
PStraffic3GULDLGB= Reference(weekly_sheet, min_col=4, min_row=1, max_row=last_row_weekly)
CSRABDropRate= Reference(weekly_sheet, min_col=5, min_row=1, max_row=last_row_weekly)
PSBlockingRate= Reference(weekly_sheet, min_col=6, min_row=1, max_row=last_row_weekly)
PSRABDropRate= Reference(weekly_sheet, min_col=7, min_row=1, max_row=last_row_weekly)
PSHSDropRate= Reference(weekly_sheet, min_col=8, min_row=1, max_row=last_row_weekly)
HSDPAThroughputkbps= Reference(weekly_sheet, min_col=9, min_row=1, max_row=last_row_weekly)
HSUPAThroughputkbps= Reference(weekly_sheet, min_col=10, min_row=1, max_row=last_row_weekly)
SoftHandoverSuccessrate= Reference(weekly_sheet, min_col=11, min_row=1, max_row=last_row_weekly)
HardHandoverSuccessrate= Reference(weekly_sheet, min_col=12, min_row=1, max_row=last_row_weekly)
CSW2GInterRATHandoverOutSR= Reference(weekly_sheet, min_col=13, min_row=1, max_row=last_row_weekly)
RRCAssignmentSucessRateCSBH= Reference(weekly_sheet, min_col=14, min_row=1, max_row=last_row_weekly)
RRCAssignmentSucessRatePSBH= Reference(weekly_sheet, min_col=15, min_row=1, max_row=last_row_weekly)
RRCDropRateCSBH= Reference(weekly_sheet, min_col=16, min_row=1, max_row=last_row_weekly)
RRCDropRatePSBH= Reference(weekly_sheet, min_col=17, min_row=1, max_row=last_row_weekly)
RABAssignmentSuccessRateCS= Reference(weekly_sheet, min_col=18, min_row=1, max_row=last_row_weekly)
RABAssignmentSuccessRatePS= Reference(weekly_sheet, min_col=19, min_row=1, max_row=last_row_weekly)
CCSR3G= Reference(weekly_sheet, min_col=20, min_row=1, max_row=last_row_weekly)
DCSR3G= Reference(weekly_sheet, min_col=21, min_row=1, max_row=last_row_weekly)

CStraffic_chart = LineChart()
CStraffic_chart.width = 40
CStraffic_chart.height = 10
CStraffic_chart.add_data(CStraffic3GErl, titles_from_data = True)  #
CStraffic_chart.add_data(PStraffic3GULDLGB, titles_from_data = True)
CStraffic_chart.set_categories(x_values)
CStraffic_chart.legend.position = 'b'
weekly_sheet.add_chart(CStraffic_chart, "A18")

PStraffic_chart = LineChart()
PStraffic_chart.width = 40
PStraffic_chart.height = 10
PStraffic_chart.add_data(PStraffic3GULDLGB, titles_from_data = True)  #
PStraffic_chart.set_categories(x_values)
PStraffic_chart.legend.position = 'b'
weekly_sheet.add_chart(PStraffic_chart, "A38")

CSdrop_chart = LineChart()
CSdrop_chart.width = 40
CSdrop_chart.height = 10
CSdrop_chart.add_data(CSRABDropRate, titles_from_data = True)  #
CSdrop_chart.set_categories(x_values)
CSdrop_chart.legend.position = 'b'
weekly_sheet.add_chart(CSdrop_chart, "A58")

PSdrop_chart = LineChart()
PSdrop_chart.width = 40
PSdrop_chart.height = 10
PSdrop_chart.add_data(PSRABDropRate, titles_from_data = True)  #
PSdrop_chart.set_categories(x_values)
PSdrop_chart.legend.position = 'b'
weekly_sheet.add_chart(PSdrop_chart, "A78")

RRCdrop_chart = LineChart()
RRCdrop_chart.width = 40
RRCdrop_chart.height = 10
RRCdrop_chart.add_data(RRCDropRateCSBH, titles_from_data = True)  #
RRCdrop_chart.set_categories(x_values)
RRCdrop_chart.legend.position = 'b'
weekly_sheet.add_chart(RRCdrop_chart, "A98")

HSDPAThroughput_chart = LineChart()
HSDPAThroughput_chart.width = 40
HSDPAThroughput_chart.height = 10
HSDPAThroughput_chart.add_data(HSDPAThroughputkbps, titles_from_data = True)  #
HSDPAThroughput_chart.set_categories(x_values)
HSDPAThroughput_chart.legend.position = 'b'
weekly_sheet.add_chart(HSDPAThroughput_chart, "A118")

HSUPAThroughput_chart = LineChart()
HSUPAThroughput_chart.width = 40
HSUPAThroughput_chart.height = 10
HSUPAThroughput_chart.add_data(HSUPAThroughputkbps, titles_from_data = True)  #
HSUPAThroughput_chart.set_categories(x_values)
HSUPAThroughput_chart.legend.position = 'b'
weekly_sheet.add_chart(HSUPAThroughput_chart, "A138")

Handover_chart = LineChart()
Handover_chart.width = 40
Handover_chart.height = 10
Handover_chart.add_data(SoftHandoverSuccessrate, titles_from_data = True)  #
Handover_chart.add_data(HardHandoverSuccessrate, titles_from_data = True)  #
Handover_chart.add_data(CSW2GInterRATHandoverOutSR, titles_from_data = True)  #
Handover_chart.set_categories(x_values)
Handover_chart.legend.position = 'b'
weekly_sheet.add_chart(Handover_chart, "A158")

RRCassign_chart = LineChart()
RRCassign_chart.width = 40
RRCassign_chart.height = 10
RRCassign_chart.add_data(RRCAssignmentSucessRateCSBH, titles_from_data = True)  #
RRCassign_chart.set_categories(x_values)
RRCassign_chart.legend.position = 'b'
weekly_sheet.add_chart(RRCassign_chart, "A178")

RABCSassign_chart = LineChart()
RABCSassign_chart.width = 40
RABCSassign_chart.height = 10
RABCSassign_chart.add_data(RABAssignmentSuccessRateCS, titles_from_data = True)  #
RABCSassign_chart.set_categories(x_values)
RABCSassign_chart.legend.position = 'b'
weekly_sheet.add_chart(RABCSassign_chart, "A198")

RABPSassign_chart = LineChart()
RABPSassign_chart.width = 40
RABPSassign_chart.height = 10
RABPSassign_chart.add_data(RABAssignmentSuccessRatePS, titles_from_data = True)  #
RABPSassign_chart.set_categories(x_values)
RABPSassign_chart.legend.position = 'b'
weekly_sheet.add_chart(RABPSassign_chart, "A218")

CCSR3G_chart = LineChart()
CCSR3G_chart.width = 40
CCSR3G_chart.height = 10
CCSR3G_chart.add_data(CCSR3G, titles_from_data = True)  #
CCSR3G_chart.set_categories(x_values)
CCSR3G_chart.legend.position = 'b'
weekly_sheet.add_chart(CCSR3G_chart, "A238")

DCSR3G_chart = LineChart()
DCSR3G_chart.width = 40
DCSR3G_chart.height = 10
DCSR3G_chart.add_data(DCSR3G, titles_from_data = True)  #
DCSR3G_chart.add_data(CCSR3G, titles_from_data = True)  #
DCSR3G_chart.set_categories(x_values)
DCSR3G_chart.legend.position = 'b'
weekly_sheet.add_chart(DCSR3G_chart, "A258")

# weekle NodeB
x_valuesN = Reference(weeklyN_sheet, range_string=(f"NodeBweekly!$B$2:$B${last_row_weeklyN}"))
MeanThrHSDPAkbps= Reference(weeklyN_sheet, min_col=3, min_row=1, max_row=last_row_weeklyN)
MeanThrHSDPADCkbps= Reference(weeklyN_sheet, min_col=4, min_row=1, max_row=last_row_weeklyN)
MeanThrHSUPAkbps= Reference(weeklyN_sheet, min_col=5, min_row=1, max_row=last_row_weeklyN)

MeanThrHSDPA_chart = LineChart()
MeanThrHSDPA_chart.width = 40
MeanThrHSDPA_chart.height = 10
MeanThrHSDPA_chart.add_data(MeanThrHSDPAkbps, titles_from_data = True)  #
MeanThrHSDPA_chart.add_data(MeanThrHSDPADCkbps, titles_from_data = True)  #
MeanThrHSDPA_chart.set_categories(x_valuesN)
MeanThrHSDPA_chart.legend.position = 'b'
weeklyN_sheet.add_chart(MeanThrHSDPA_chart, "A18")

MeanThrHSUPAkbps_chart = LineChart()
MeanThrHSUPAkbps_chart.width = 40
MeanThrHSUPAkbps_chart.height = 10
MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAkbps, titles_from_data = True)  #
MeanThrHSUPAkbps_chart.set_categories(x_values)
MeanThrHSUPAkbps_chart.legend.position = 'b'
weeklyN_sheet.add_chart(MeanThrHSUPAkbps_chart, "A38")



#  графики в суточной таблице daily_sheet   last_row_daily
x_values = Reference(daily_sheet, range_string=(f"daily!$B$2:$B${last_row_daily}"))

CStraffic3GErl= Reference(daily_sheet, min_col=3, min_row=1, max_row=last_row_daily)
PStraffic3GULDLGB= Reference(daily_sheet, min_col=4, min_row=1, max_row=last_row_daily)
CSRABDropRate= Reference(daily_sheet, min_col=5, min_row=1, max_row=last_row_daily)
PSBlockingRate= Reference(daily_sheet, min_col=6, min_row=1, max_row=last_row_daily)
PSRABDropRate= Reference(daily_sheet, min_col=7, min_row=1, max_row=last_row_daily)
PSHSDropRate= Reference(daily_sheet, min_col=8, min_row=1, max_row=last_row_daily)
HSDPAThroughputkbps= Reference(daily_sheet, min_col=9, min_row=1, max_row=last_row_daily)
HSUPAThroughputkbps= Reference(daily_sheet, min_col=10, min_row=1, max_row=last_row_daily)
SoftHandoverSuccessrate= Reference(daily_sheet, min_col=11, min_row=1, max_row=last_row_daily)
HardHandoverSuccessrate= Reference(daily_sheet, min_col=12, min_row=1, max_row=last_row_daily)
CSW2GInterRATHandoverOutSR= Reference(daily_sheet, min_col=13, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRateCSBH= Reference(daily_sheet, min_col=14, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRatePSBH= Reference(daily_sheet, min_col=15, min_row=1, max_row=last_row_daily)
RRCDropRateCSBH= Reference(daily_sheet, min_col=16, min_row=1, max_row=last_row_daily)
RRCDropRatePSBH= Reference(daily_sheet, min_col=17, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRateCS= Reference(daily_sheet, min_col=18, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRatePS= Reference(daily_sheet, min_col=19, min_row=1, max_row=last_row_daily)
CCSR3G= Reference(daily_sheet, min_col=20, min_row=1, max_row=last_row_daily)
DCSR3G= Reference(daily_sheet, min_col=21, min_row=1, max_row=last_row_daily)
CStraffic3GErlU2100= Reference(daily_sheet, min_col=22, min_row=1, max_row=last_row_daily)
PStraffic3GULDLGBU2100= Reference(daily_sheet, min_col=23, min_row=1, max_row=last_row_daily)
CSRABDropRateU2100= Reference(daily_sheet, min_col=24, min_row=1, max_row=last_row_daily)
PSBlockingRateU2100= Reference(daily_sheet, min_col=25, min_row=1, max_row=last_row_daily)
PSRABDropRateU2100= Reference(daily_sheet, min_col=26, min_row=1, max_row=last_row_daily)
PSHSDropRateU2100= Reference(daily_sheet, min_col=27, min_row=1, max_row=last_row_daily)
HSDPAThroughputkbpsU2100= Reference(daily_sheet, min_col=28, min_row=1, max_row=last_row_daily)
HSUPAThroughputkbpsU2100= Reference(daily_sheet, min_col=29, min_row=1, max_row=last_row_daily)
SoftHandoverSuccessrateU2100= Reference(daily_sheet, min_col=30, min_row=1, max_row=last_row_daily)
HardHandoverSuccessrateU2100= Reference(daily_sheet, min_col=31, min_row=1, max_row=last_row_daily)
CSW2GInterRATHandoverOutSRU2100= Reference(daily_sheet, min_col=32, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRateCSBHU2100= Reference(daily_sheet, min_col=33, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRatePSBHU2100= Reference(daily_sheet, min_col=34, min_row=1, max_row=last_row_daily)
RRCDropRateCSBHU2100= Reference(daily_sheet, min_col=35, min_row=1, max_row=last_row_daily)
RRCDropRatePSBHU2100= Reference(daily_sheet, min_col=36, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRateCSU2100= Reference(daily_sheet, min_col=37, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRatePSU2100= Reference(daily_sheet, min_col=38, min_row=1, max_row=last_row_daily)
CCSR3GU2100= Reference(daily_sheet, min_col=39, min_row=1, max_row=last_row_daily)
DCSR3GU2100= Reference(daily_sheet, min_col=40, min_row=1, max_row=last_row_daily)
CStraffic3GErlU900= Reference(daily_sheet, min_col=41, min_row=1, max_row=last_row_daily)
PStraffic3GULDLGBU900= Reference(daily_sheet, min_col=42, min_row=1, max_row=last_row_daily)
CSRABDropRateU900= Reference(daily_sheet, min_col=43, min_row=1, max_row=last_row_daily)
PSBlockingRateU900= Reference(daily_sheet, min_col=44, min_row=1, max_row=last_row_daily)
PSRABDropRateU900= Reference(daily_sheet, min_col=45, min_row=1, max_row=last_row_daily)
PSHSDropRateU900= Reference(daily_sheet, min_col=46, min_row=1, max_row=last_row_daily)
HSDPAThroughputkbpsU900= Reference(daily_sheet, min_col=47, min_row=1, max_row=last_row_daily)
HSUPAThroughputkbpsU900= Reference(daily_sheet, min_col=48, min_row=1, max_row=last_row_daily)
SoftHandoverSuccessrateU900= Reference(daily_sheet, min_col=49, min_row=1, max_row=last_row_daily)
HardHandoverSuccessrateU900= Reference(daily_sheet, min_col=50, min_row=1, max_row=last_row_daily)
CSW2GInterRATHandoverOutSRU900= Reference(daily_sheet, min_col=51, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRateCSBHU900= Reference(daily_sheet, min_col=52, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRatePSBHU900= Reference(daily_sheet, min_col=53, min_row=1, max_row=last_row_daily)
RRCDropRateCSBHU900= Reference(daily_sheet, min_col=54, min_row=1, max_row=last_row_daily)
RRCDropRatePSBHU900= Reference(daily_sheet, min_col=55, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRateCSU900= Reference(daily_sheet, min_col=56, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRatePSU900= Reference(daily_sheet, min_col=57, min_row=1, max_row=last_row_daily)
CCSR3GU900= Reference(daily_sheet, min_col=58, min_row=1, max_row=last_row_daily)
DCSR3GU900= Reference(daily_sheet, min_col=59, min_row=1, max_row=last_row_daily)
CStraffic3GErl10612= Reference(daily_sheet, min_col=60, min_row=1, max_row=last_row_daily)
PStraffic3GULDLGB10612= Reference(daily_sheet, min_col=61, min_row=1, max_row=last_row_daily)
CSRABDropRate10612= Reference(daily_sheet, min_col=62, min_row=1, max_row=last_row_daily)
PSBlockingRate10612= Reference(daily_sheet, min_col=63, min_row=1, max_row=last_row_daily)
PSRABDropRate10612= Reference(daily_sheet, min_col=64, min_row=1, max_row=last_row_daily)
PSHSDropRate10612= Reference(daily_sheet, min_col=65, min_row=1, max_row=last_row_daily)
HSDPAThroughputkbps10612= Reference(daily_sheet, min_col=66, min_row=1, max_row=last_row_daily)
HSUPAThroughputkbps10612= Reference(daily_sheet, min_col=67, min_row=1, max_row=last_row_daily)
SoftHandoverSuccessrate10612= Reference(daily_sheet, min_col=68, min_row=1, max_row=last_row_daily)
HardHandoverSuccessrate10612= Reference(daily_sheet, min_col=69, min_row=1, max_row=last_row_daily)
CSW2GInterRATHandoverOutSR10612= Reference(daily_sheet, min_col=70, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRateCSBH10612= Reference(daily_sheet, min_col=71, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRatePSBH10612= Reference(daily_sheet, min_col=72, min_row=1, max_row=last_row_daily)
RRCDropRateCSBH10612= Reference(daily_sheet, min_col=73, min_row=1, max_row=last_row_daily)
RRCDropRatePSBH10612= Reference(daily_sheet, min_col=74, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRateCS10612= Reference(daily_sheet, min_col=75, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRatePS10612= Reference(daily_sheet, min_col=76, min_row=1, max_row=last_row_daily)
CCSR3G10612= Reference(daily_sheet, min_col=77, min_row=1, max_row=last_row_daily)
DCSR3G10612= Reference(daily_sheet, min_col=78, min_row=1, max_row=last_row_daily)
CStraffic3GErl10637= Reference(daily_sheet, min_col=79, min_row=1, max_row=last_row_daily)
PStraffic3GULDLGB10637= Reference(daily_sheet, min_col=80, min_row=1, max_row=last_row_daily)
CSRABDropRate10637= Reference(daily_sheet, min_col=81, min_row=1, max_row=last_row_daily)
PSBlockingRate10637= Reference(daily_sheet, min_col=82, min_row=1, max_row=last_row_daily)
PSRABDropRate10637= Reference(daily_sheet, min_col=83, min_row=1, max_row=last_row_daily)
PSHSDropRate10637= Reference(daily_sheet, min_col=84, min_row=1, max_row=last_row_daily)
HSDPAThroughputkbps10637= Reference(daily_sheet, min_col=85, min_row=1, max_row=last_row_daily)
HSUPAThroughputkbps10637= Reference(daily_sheet, min_col=86, min_row=1, max_row=last_row_daily)
SoftHandoverSuccessrate10637= Reference(daily_sheet, min_col=87, min_row=1, max_row=last_row_daily)
HardHandoverSuccessrate10637= Reference(daily_sheet, min_col=88, min_row=1, max_row=last_row_daily)
CSW2GInterRATHandoverOutSR10637= Reference(daily_sheet, min_col=89, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRateCSBH10637= Reference(daily_sheet, min_col=90, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRatePSBH10637= Reference(daily_sheet, min_col=91, min_row=1, max_row=last_row_daily)
RRCDropRateCSBH10637= Reference(daily_sheet, min_col=92, min_row=1, max_row=last_row_daily)
RRCDropRatePSBH10637= Reference(daily_sheet, min_col=93, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRateCS10637= Reference(daily_sheet, min_col=94, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRatePS10637= Reference(daily_sheet, min_col=95, min_row=1, max_row=last_row_daily)
CCSR3G10637= Reference(daily_sheet, min_col=96, min_row=1, max_row=last_row_daily)
DCSR3G10637= Reference(daily_sheet, min_col=97, min_row=1, max_row=last_row_daily)
CStraffic3GErl2937= Reference(daily_sheet, min_col=98, min_row=1, max_row=last_row_daily)
PStraffic3GULDLGB2937= Reference(daily_sheet, min_col=99, min_row=1, max_row=last_row_daily)
CSRABDropRate2937= Reference(daily_sheet, min_col=100, min_row=1, max_row=last_row_daily)
PSBlockingRate2937= Reference(daily_sheet, min_col=101, min_row=1, max_row=last_row_daily)
PSRABDropRate2937= Reference(daily_sheet, min_col=102, min_row=1, max_row=last_row_daily)
PSHSDropRate2937= Reference(daily_sheet, min_col=103, min_row=1, max_row=last_row_daily)
HSDPAThroughputkbps2937= Reference(daily_sheet, min_col=104, min_row=1, max_row=last_row_daily)
HSUPAThroughputkbps2937= Reference(daily_sheet, min_col=105, min_row=1, max_row=last_row_daily)
SoftHandoverSuccessrate2937= Reference(daily_sheet, min_col=106, min_row=1, max_row=last_row_daily)
HardHandoverSuccessrate2937= Reference(daily_sheet, min_col=107, min_row=1, max_row=last_row_daily)
CSW2GInterRATHandoverOutSR2937= Reference(daily_sheet, min_col=108, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRateCSBH2937= Reference(daily_sheet, min_col=109, min_row=1, max_row=last_row_daily)
RRCAssignmentSucessRatePSBH2937= Reference(daily_sheet, min_col=110, min_row=1, max_row=last_row_daily)
RRCDropRateCSBH2937= Reference(daily_sheet, min_col=111, min_row=1, max_row=last_row_daily)
RRCDropRatePSBH2937= Reference(daily_sheet, min_col=112, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRateCS2937= Reference(daily_sheet, min_col=113, min_row=1, max_row=last_row_daily)
RABAssignmentSuccessRatePS2937= Reference(daily_sheet, min_col=114, min_row=1, max_row=last_row_daily)
CCSR3G2937= Reference(daily_sheet, min_col=115, min_row=1, max_row=last_row_daily)
DCSR3G2937= Reference(daily_sheet, min_col=116, min_row=1, max_row=last_row_daily)

CStraffic_chart = LineChart()
CStraffic_chart.width = 40
CStraffic_chart.height = 10
CStraffic_chart.add_data(CStraffic3GErl, titles_from_data = True)  #
CStraffic_chart.add_data(CStraffic3GErlU2100, titles_from_data = True)
CStraffic_chart.add_data(CStraffic3GErlU900, titles_from_data = True)
CStraffic_chart.set_categories(x_values)
CStraffic_chart.legend.position = 'b'
daily_sheet.add_chart(CStraffic_chart, "A18")

PStraffic_chart = LineChart()
PStraffic_chart.width = 40
PStraffic_chart.height = 10
PStraffic_chart.add_data(PStraffic3GULDLGB, titles_from_data = True)  #
PStraffic_chart.add_data(PStraffic3GULDLGBU2100, titles_from_data = True)
PStraffic_chart.add_data(PStraffic3GULDLGBU900, titles_from_data = True)
PStraffic_chart.set_categories(x_values)
PStraffic_chart.legend.position = 'b'
daily_sheet.add_chart(PStraffic_chart, "A38")

CSdrop_chart = LineChart()
CSdrop_chart.width = 40
CSdrop_chart.height = 10
CSdrop_chart.add_data(CSRABDropRate, titles_from_data = True)  #
CSdrop_chart.add_data(CSRABDropRateU2100, titles_from_data = True)
CSdrop_chart.add_data(CSRABDropRateU900, titles_from_data = True)
CSdrop_chart.set_categories(x_values)
CSdrop_chart.legend.position = 'b'
daily_sheet.add_chart(CSdrop_chart, "A58")

PSdrop_chart = LineChart()
PSdrop_chart.width = 40
PSdrop_chart.height = 10
PSdrop_chart.add_data(PSRABDropRate, titles_from_data = True)  #
PSdrop_chart.add_data(PSRABDropRateU2100, titles_from_data = True)
PSdrop_chart.add_data(PSRABDropRateU900, titles_from_data = True)
PSdrop_chart.set_categories(x_values)
PSdrop_chart.legend.position = 'b'
daily_sheet.add_chart(PSdrop_chart, "A78")

RRCdrop_chart = LineChart()
RRCdrop_chart.width = 40
RRCdrop_chart.height = 10
RRCdrop_chart.add_data(RRCDropRateCSBH, titles_from_data = True)  #
RRCdrop_chart.add_data(RRCDropRateCSBHU2100, titles_from_data = True)
RRCdrop_chart.add_data(RRCDropRateCSBHU900, titles_from_data = True)
RRCdrop_chart.set_categories(x_values)
RRCdrop_chart.legend.position = 'b'
daily_sheet.add_chart(RRCdrop_chart, "A98")

HSDPAThroughput_chart = LineChart()
HSDPAThroughput_chart.width = 40
HSDPAThroughput_chart.height = 10
HSDPAThroughput_chart.add_data(HSDPAThroughputkbps, titles_from_data = True)  #
HSDPAThroughput_chart.add_data(HSDPAThroughputkbpsU2100, titles_from_data = True)
HSDPAThroughput_chart.add_data(HSDPAThroughputkbpsU900, titles_from_data = True)
HSDPAThroughput_chart.set_categories(x_values)
HSDPAThroughput_chart.legend.position = 'b'
daily_sheet.add_chart(HSDPAThroughput_chart, "A118")

HSUPAThroughput_chart = LineChart()
HSUPAThroughput_chart.width = 40
HSUPAThroughput_chart.height = 10
HSUPAThroughput_chart.add_data(HSUPAThroughputkbps, titles_from_data = True)  #
HSUPAThroughput_chart.add_data(HSUPAThroughputkbpsU2100, titles_from_data = True)
HSUPAThroughput_chart.add_data(HSUPAThroughputkbpsU900, titles_from_data = True)
HSUPAThroughput_chart.set_categories(x_values)
HSUPAThroughput_chart.legend.position = 'b'
daily_sheet.add_chart(HSUPAThroughput_chart, "A138")

Handover_chart = LineChart()
Handover_chart.width = 40
Handover_chart.height = 10
Handover_chart.add_data(SoftHandoverSuccessrate, titles_from_data = True)  #
Handover_chart.add_data(SoftHandoverSuccessrateU2100, titles_from_data = True)
Handover_chart.add_data(SoftHandoverSuccessrateU900, titles_from_data = True)
Handover_chart.add_data(HardHandoverSuccessrate, titles_from_data = True)  #
Handover_chart.add_data(HardHandoverSuccessrateU2100, titles_from_data = True)
Handover_chart.add_data(HardHandoverSuccessrateU900, titles_from_data = True)
Handover_chart.add_data(CSW2GInterRATHandoverOutSR, titles_from_data = True)  #
Handover_chart.add_data(CSW2GInterRATHandoverOutSRU2100, titles_from_data = True)
Handover_chart.add_data(CSW2GInterRATHandoverOutSRU900, titles_from_data = True)
Handover_chart.set_categories(x_values)
Handover_chart.legend.position = 'b'
daily_sheet.add_chart(Handover_chart, "A158")

RRCassign_chart = LineChart()
RRCassign_chart.width = 40
RRCassign_chart.height = 10
RRCassign_chart.add_data(RRCAssignmentSucessRateCSBH, titles_from_data = True)  #
RRCassign_chart.add_data(RRCAssignmentSucessRateCSBHU2100, titles_from_data = True)
RRCassign_chart.add_data(RRCAssignmentSucessRateCSBHU900, titles_from_data = True)
RRCassign_chart.set_categories(x_values)
RRCassign_chart.legend.position = 'b'
daily_sheet.add_chart(RRCassign_chart, "A178")

RABCSassign_chart = LineChart()
RABCSassign_chart.width = 40
RABCSassign_chart.height = 10
RABCSassign_chart.add_data(RABAssignmentSuccessRateCS, titles_from_data = True)  #
RABCSassign_chart.add_data(RABAssignmentSuccessRateCSU2100, titles_from_data = True)
RABCSassign_chart.add_data(RABAssignmentSuccessRateCSU900, titles_from_data = True)
RABCSassign_chart.set_categories(x_values)
RABCSassign_chart.legend.position = 'b'
daily_sheet.add_chart(RABCSassign_chart, "A198")

RABPSassign_chart = LineChart()
RABPSassign_chart.width = 40
RABPSassign_chart.height = 10
RABPSassign_chart.add_data(RABAssignmentSuccessRatePS, titles_from_data = True)  #
RABPSassign_chart.add_data(RABAssignmentSuccessRatePSU2100, titles_from_data = True)
RABPSassign_chart.add_data(RABAssignmentSuccessRatePSU900, titles_from_data = True)
RABPSassign_chart.set_categories(x_values)
RABPSassign_chart.legend.position = 'b'
daily_sheet.add_chart(RABPSassign_chart, "A218")

CCSR3G_chart = LineChart()
CCSR3G_chart.width = 40
CCSR3G_chart.height = 10
CCSR3G_chart.add_data(CCSR3G, titles_from_data = True)  #
CCSR3G_chart.add_data(CCSR3GU2100, titles_from_data = True)
CCSR3G_chart.add_data(CCSR3GU900, titles_from_data = True)
CCSR3G_chart.set_categories(x_values)
CCSR3G_chart.legend.position = 'b'
daily_sheet.add_chart(CCSR3G_chart, "A238")

DCSR3G_chart = LineChart()
DCSR3G_chart.width = 40
DCSR3G_chart.height = 10
DCSR3G_chart.add_data(DCSR3G, titles_from_data = True)  #
DCSR3G_chart.add_data(DCSR3GU2100, titles_from_data = True)
DCSR3G_chart.add_data(DCSR3GU900, titles_from_data = True)
DCSR3G_chart.set_categories(x_values)
DCSR3G_chart.legend.position = 'b'
daily_sheet.add_chart(DCSR3G_chart, "A258")
#daily NodeB
x_valuesN = Reference(dailyN_sheet, range_string=(f"NodeBdaily!$B$2:$B${last_row_dailyN}"))
MeanThrHSDPAkbps= Reference(dailyN_sheet, min_col=3, min_row=1, max_row=last_row_dailyN)
MeanThrHSDPADCkbps= Reference(dailyN_sheet, min_col=4, min_row=1, max_row=last_row_dailyN)
MeanThrHSUPAkbps= Reference(dailyN_sheet, min_col=5, min_row=1, max_row=last_row_dailyN)
MeanThrHSDPAU2100kbps= Reference(dailyN_sheet, min_col=6, min_row=1, max_row=last_row_dailyN)
MeanThrHSDPAU2100DCkbps= Reference(dailyN_sheet, min_col=7, min_row=1, max_row=last_row_dailyN)
MeanThrHSUPAU2100kbps= Reference(dailyN_sheet, min_col=8, min_row=1, max_row=last_row_dailyN)
MeanThrHSDPAU900kbps= Reference(dailyN_sheet, min_col=9, min_row=1, max_row=last_row_dailyN)
MeanThrHSDPAU900DCkbps= Reference(dailyN_sheet, min_col=10, min_row=1, max_row=last_row_dailyN)
MeanThrHSUPAU900kbps= Reference(dailyN_sheet, min_col=11, min_row=1, max_row=last_row_dailyN)

MeanThrHSDPA_chart = LineChart()
MeanThrHSDPA_chart.width = 40
MeanThrHSDPA_chart.height = 10
MeanThrHSDPA_chart.add_data(MeanThrHSDPAkbps, titles_from_data = True)  #
MeanThrHSDPA_chart.add_data(MeanThrHSDPAU2100kbps, titles_from_data = True)  #
MeanThrHSDPA_chart.add_data(MeanThrHSDPAU900kbps, titles_from_data = True)  #
MeanThrHSDPA_chart.set_categories(x_valuesN)
MeanThrHSDPA_chart.legend.position = 'b'
dailyN_sheet.add_chart(MeanThrHSDPA_chart, "A18")

MeanThrHSUPAkbps_chart = LineChart()
MeanThrHSUPAkbps_chart.width = 40
MeanThrHSUPAkbps_chart.height = 10
MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAkbps, titles_from_data = True)  #
MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAU2100kbps, titles_from_data = True)  #
MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAU900kbps, titles_from_data = True)  #
MeanThrHSUPAkbps_chart.set_categories(x_valuesN)
MeanThrHSUPAkbps_chart.legend.position = 'b'
dailyN_sheet.add_chart(MeanThrHSUPAkbps_chart, "A38")

MeanThrHSDPADC_chart = LineChart()
MeanThrHSDPADC_chart.width = 40
MeanThrHSDPADC_chart.height = 10
MeanThrHSDPADC_chart.add_data(MeanThrHSDPADCkbps, titles_from_data = True)  #
MeanThrHSDPADC_chart.add_data(MeanThrHSDPAU2100DCkbps, titles_from_data = True)  #
MeanThrHSDPADC_chart.add_data(MeanThrHSDPAU900DCkbps, titles_from_data = True)  #
MeanThrHSDPADC_chart.set_categories(x_valuesN)
MeanThrHSDPADC_chart.legend.position = 'b'
dailyN_sheet.add_chart(MeanThrHSDPADC_chart, "A58")


# графики почасовые hourly_sheet  last_row_hourly
x_values = Reference(hourly_sheet, range_string=(f"hourly!$A$2:$B${last_row_hourly}"))

CStraffic3GErl= Reference(hourly_sheet, min_col=3, min_row=1, max_row=last_row_hourly)
PStraffic3GULDLGB= Reference(hourly_sheet, min_col=4, min_row=1, max_row=last_row_hourly)
CSRABDropRate= Reference(hourly_sheet, min_col=5, min_row=1, max_row=last_row_hourly)
PSBlockingRate= Reference(hourly_sheet, min_col=6, min_row=1, max_row=last_row_hourly)
PSRABDropRate= Reference(hourly_sheet, min_col=7, min_row=1, max_row=last_row_hourly)
PSHSDropRate= Reference(hourly_sheet, min_col=8, min_row=1, max_row=last_row_hourly)
HSDPAThroughputkbps= Reference(hourly_sheet, min_col=9, min_row=1, max_row=last_row_hourly)
HSUPAThroughputkbps= Reference(hourly_sheet, min_col=10, min_row=1, max_row=last_row_hourly)
SoftHandoverSuccessrate= Reference(hourly_sheet, min_col=11, min_row=1, max_row=last_row_hourly)
HardHandoverSuccessrate= Reference(hourly_sheet, min_col=12, min_row=1, max_row=last_row_hourly)
CSW2GInterRATHandoverOutSR= Reference(hourly_sheet, min_col=13, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRateCSBH= Reference(hourly_sheet, min_col=14, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRatePSBH= Reference(hourly_sheet, min_col=15, min_row=1, max_row=last_row_hourly)
RRCDropRateCSBH= Reference(hourly_sheet, min_col=16, min_row=1, max_row=last_row_hourly)
RRCDropRatePSBH= Reference(hourly_sheet, min_col=17, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRateCS= Reference(hourly_sheet, min_col=18, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRatePS= Reference(hourly_sheet, min_col=19, min_row=1, max_row=last_row_hourly)
CCSR3G= Reference(hourly_sheet, min_col=20, min_row=1, max_row=last_row_hourly)
DCSR3G= Reference(hourly_sheet, min_col=21, min_row=1, max_row=last_row_hourly)
CStraffic3GErlU2100= Reference(hourly_sheet, min_col=22, min_row=1, max_row=last_row_hourly)
PStraffic3GULDLGBU2100= Reference(hourly_sheet, min_col=23, min_row=1, max_row=last_row_hourly)
CSRABDropRateU2100= Reference(hourly_sheet, min_col=24, min_row=1, max_row=last_row_hourly)
PSBlockingRateU2100= Reference(hourly_sheet, min_col=25, min_row=1, max_row=last_row_hourly)
PSRABDropRateU2100= Reference(hourly_sheet, min_col=26, min_row=1, max_row=last_row_hourly)
PSHSDropRateU2100= Reference(hourly_sheet, min_col=27, min_row=1, max_row=last_row_hourly)
HSDPAThroughputkbpsU2100= Reference(hourly_sheet, min_col=28, min_row=1, max_row=last_row_hourly)
HSUPAThroughputkbpsU2100= Reference(hourly_sheet, min_col=29, min_row=1, max_row=last_row_hourly)
SoftHandoverSuccessrateU2100= Reference(hourly_sheet, min_col=30, min_row=1, max_row=last_row_hourly)
HardHandoverSuccessrateU2100= Reference(hourly_sheet, min_col=31, min_row=1, max_row=last_row_hourly)
CSW2GInterRATHandoverOutSRU2100= Reference(hourly_sheet, min_col=32, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRateCSBHU2100= Reference(hourly_sheet, min_col=33, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRatePSBHU2100= Reference(hourly_sheet, min_col=34, min_row=1, max_row=last_row_hourly)
RRCDropRateCSBHU2100= Reference(hourly_sheet, min_col=35, min_row=1, max_row=last_row_hourly)
RRCDropRatePSBHU2100= Reference(hourly_sheet, min_col=36, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRateCSU2100= Reference(hourly_sheet, min_col=37, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRatePSU2100= Reference(hourly_sheet, min_col=38, min_row=1, max_row=last_row_hourly)
CCSR3GU2100= Reference(hourly_sheet, min_col=39, min_row=1, max_row=last_row_hourly)
DCSR3GU2100= Reference(hourly_sheet, min_col=40, min_row=1, max_row=last_row_hourly)
CStraffic3GErlU900= Reference(hourly_sheet, min_col=41, min_row=1, max_row=last_row_hourly)
PStraffic3GULDLGBU900= Reference(hourly_sheet, min_col=42, min_row=1, max_row=last_row_hourly)
CSRABDropRateU900= Reference(hourly_sheet, min_col=43, min_row=1, max_row=last_row_hourly)
PSBlockingRateU900= Reference(hourly_sheet, min_col=44, min_row=1, max_row=last_row_hourly)
PSRABDropRateU900= Reference(hourly_sheet, min_col=45, min_row=1, max_row=last_row_hourly)
PSHSDropRateU900= Reference(hourly_sheet, min_col=46, min_row=1, max_row=last_row_hourly)
HSDPAThroughputkbpsU900= Reference(hourly_sheet, min_col=47, min_row=1, max_row=last_row_hourly)
HSUPAThroughputkbpsU900= Reference(hourly_sheet, min_col=48, min_row=1, max_row=last_row_hourly)
SoftHandoverSuccessrateU900= Reference(hourly_sheet, min_col=49, min_row=1, max_row=last_row_hourly)
HardHandoverSuccessrateU900= Reference(hourly_sheet, min_col=50, min_row=1, max_row=last_row_hourly)
CSW2GInterRATHandoverOutSRU900= Reference(hourly_sheet, min_col=51, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRateCSBHU900= Reference(hourly_sheet, min_col=52, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRatePSBHU900= Reference(hourly_sheet, min_col=53, min_row=1, max_row=last_row_hourly)
RRCDropRateCSBHU900= Reference(hourly_sheet, min_col=54, min_row=1, max_row=last_row_hourly)
RRCDropRatePSBHU900= Reference(hourly_sheet, min_col=55, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRateCSU900= Reference(hourly_sheet, min_col=56, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRatePSU900= Reference(hourly_sheet, min_col=57, min_row=1, max_row=last_row_hourly)
CCSR3GU900= Reference(hourly_sheet, min_col=58, min_row=1, max_row=last_row_hourly)
DCSR3GU900= Reference(hourly_sheet, min_col=59, min_row=1, max_row=last_row_hourly)
CStraffic3GErl10612= Reference(hourly_sheet, min_col=60, min_row=1, max_row=last_row_hourly)
PStraffic3GULDLGB10612= Reference(hourly_sheet, min_col=61, min_row=1, max_row=last_row_hourly)
CSRABDropRate10612= Reference(hourly_sheet, min_col=62, min_row=1, max_row=last_row_hourly)
PSBlockingRate10612= Reference(hourly_sheet, min_col=63, min_row=1, max_row=last_row_hourly)
PSRABDropRate10612= Reference(hourly_sheet, min_col=64, min_row=1, max_row=last_row_hourly)
PSHSDropRate10612= Reference(hourly_sheet, min_col=65, min_row=1, max_row=last_row_hourly)
HSDPAThroughputkbps10612= Reference(hourly_sheet, min_col=66, min_row=1, max_row=last_row_hourly)
HSUPAThroughputkbps10612= Reference(hourly_sheet, min_col=67, min_row=1, max_row=last_row_hourly)
SoftHandoverSuccessrate10612= Reference(hourly_sheet, min_col=68, min_row=1, max_row=last_row_hourly)
HardHandoverSuccessrate10612= Reference(hourly_sheet, min_col=69, min_row=1, max_row=last_row_hourly)
CSW2GInterRATHandoverOutSR10612= Reference(hourly_sheet, min_col=70, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRateCSBH10612= Reference(hourly_sheet, min_col=71, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRatePSBH10612= Reference(hourly_sheet, min_col=72, min_row=1, max_row=last_row_hourly)
RRCDropRateCSBH10612= Reference(hourly_sheet, min_col=73, min_row=1, max_row=last_row_hourly)
RRCDropRatePSBH10612= Reference(hourly_sheet, min_col=74, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRateCS10612= Reference(hourly_sheet, min_col=75, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRatePS10612= Reference(hourly_sheet, min_col=76, min_row=1, max_row=last_row_hourly)
CCSR3G10612= Reference(hourly_sheet, min_col=77, min_row=1, max_row=last_row_hourly)
DCSR3G10612= Reference(hourly_sheet, min_col=78, min_row=1, max_row=last_row_hourly)
CStraffic3GErl10637= Reference(hourly_sheet, min_col=79, min_row=1, max_row=last_row_hourly)
PStraffic3GULDLGB10637= Reference(hourly_sheet, min_col=80, min_row=1, max_row=last_row_hourly)
CSRABDropRate10637= Reference(hourly_sheet, min_col=81, min_row=1, max_row=last_row_hourly)
PSBlockingRate10637= Reference(hourly_sheet, min_col=82, min_row=1, max_row=last_row_hourly)
PSRABDropRate10637= Reference(hourly_sheet, min_col=83, min_row=1, max_row=last_row_hourly)
PSHSDropRate10637= Reference(hourly_sheet, min_col=84, min_row=1, max_row=last_row_hourly)
HSDPAThroughputkbps10637= Reference(hourly_sheet, min_col=85, min_row=1, max_row=last_row_hourly)
HSUPAThroughputkbps10637= Reference(hourly_sheet, min_col=86, min_row=1, max_row=last_row_hourly)
SoftHandoverSuccessrate10637= Reference(hourly_sheet, min_col=87, min_row=1, max_row=last_row_hourly)
HardHandoverSuccessrate10637= Reference(hourly_sheet, min_col=88, min_row=1, max_row=last_row_hourly)
CSW2GInterRATHandoverOutSR10637= Reference(hourly_sheet, min_col=89, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRateCSBH10637= Reference(hourly_sheet, min_col=90, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRatePSBH10637= Reference(hourly_sheet, min_col=91, min_row=1, max_row=last_row_hourly)
RRCDropRateCSBH10637= Reference(hourly_sheet, min_col=92, min_row=1, max_row=last_row_hourly)
RRCDropRatePSBH10637= Reference(hourly_sheet, min_col=93, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRateCS10637= Reference(hourly_sheet, min_col=94, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRatePS10637= Reference(hourly_sheet, min_col=95, min_row=1, max_row=last_row_hourly)
CCSR3G10637= Reference(hourly_sheet, min_col=96, min_row=1, max_row=last_row_hourly)
DCSR3G10637= Reference(hourly_sheet, min_col=97, min_row=1, max_row=last_row_hourly)
CStraffic3GErl2937= Reference(hourly_sheet, min_col=98, min_row=1, max_row=last_row_hourly)
PStraffic3GULDLGB2937= Reference(hourly_sheet, min_col=99, min_row=1, max_row=last_row_hourly)
CSRABDropRate2937= Reference(hourly_sheet, min_col=100, min_row=1, max_row=last_row_hourly)
PSBlockingRate2937= Reference(hourly_sheet, min_col=101, min_row=1, max_row=last_row_hourly)
PSRABDropRate2937= Reference(hourly_sheet, min_col=102, min_row=1, max_row=last_row_hourly)
PSHSDropRate2937= Reference(hourly_sheet, min_col=103, min_row=1, max_row=last_row_hourly)
HSDPAThroughputkbps2937= Reference(hourly_sheet, min_col=104, min_row=1, max_row=last_row_hourly)
HSUPAThroughputkbps2937= Reference(hourly_sheet, min_col=105, min_row=1, max_row=last_row_hourly)
SoftHandoverSuccessrate2937= Reference(hourly_sheet, min_col=106, min_row=1, max_row=last_row_hourly)
HardHandoverSuccessrate2937= Reference(hourly_sheet, min_col=107, min_row=1, max_row=last_row_hourly)
CSW2GInterRATHandoverOutSR2937= Reference(hourly_sheet, min_col=108, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRateCSBH2937= Reference(hourly_sheet, min_col=109, min_row=1, max_row=last_row_hourly)
RRCAssignmentSucessRatePSBH2937= Reference(hourly_sheet, min_col=110, min_row=1, max_row=last_row_hourly)
RRCDropRateCSBH2937= Reference(hourly_sheet, min_col=111, min_row=1, max_row=last_row_hourly)
RRCDropRatePSBH2937= Reference(hourly_sheet, min_col=112, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRateCS2937= Reference(hourly_sheet, min_col=113, min_row=1, max_row=last_row_hourly)
RABAssignmentSuccessRatePS2937= Reference(hourly_sheet, min_col=114, min_row=1, max_row=last_row_hourly)
CCSR3G2937= Reference(hourly_sheet, min_col=115, min_row=1, max_row=last_row_hourly)
DCSR3G2937= Reference(hourly_sheet, min_col=116, min_row=1, max_row=last_row_hourly)

CStraffic_chart = LineChart()
CStraffic_chart.width = 40
CStraffic_chart.height = 10
CStraffic_chart.add_data(CStraffic3GErl, titles_from_data = True)  #
CStraffic_chart.add_data(CStraffic3GErlU2100, titles_from_data = True)
CStraffic_chart.add_data(CStraffic3GErlU900, titles_from_data = True)
CStraffic_chart.set_categories(x_values)
CStraffic_chart.legend.position = 'b'
hourly_sheet.add_chart(CStraffic_chart, "A18")

PStraffic_chart = LineChart()
PStraffic_chart.width = 40
PStraffic_chart.height = 10
PStraffic_chart.add_data(PStraffic3GULDLGB, titles_from_data = True)  #
PStraffic_chart.add_data(PStraffic3GULDLGBU2100, titles_from_data = True)
PStraffic_chart.add_data(PStraffic3GULDLGBU900, titles_from_data = True)
PStraffic_chart.set_categories(x_values)
PStraffic_chart.legend.position = 'b'
hourly_sheet.add_chart(PStraffic_chart, "A38")

CSdrop_chart = LineChart()
CSdrop_chart.width = 40
CSdrop_chart.height = 10
CSdrop_chart.add_data(CSRABDropRate, titles_from_data = True)  #
CSdrop_chart.add_data(CSRABDropRateU2100, titles_from_data = True)
CSdrop_chart.add_data(CSRABDropRateU900, titles_from_data = True)
CSdrop_chart.set_categories(x_values)
CSdrop_chart.legend.position = 'b'
hourly_sheet.add_chart(CSdrop_chart, "A58")

PSdrop_chart = LineChart()
PSdrop_chart.width = 40
PSdrop_chart.height = 10
PSdrop_chart.add_data(PSRABDropRate, titles_from_data = True)  #
PSdrop_chart.add_data(PSRABDropRateU2100, titles_from_data = True)
PSdrop_chart.add_data(PSRABDropRateU900, titles_from_data = True)
PSdrop_chart.set_categories(x_values)
PSdrop_chart.legend.position = 'b'
hourly_sheet.add_chart(PSdrop_chart, "A78")

RRCdrop_chart = LineChart()
RRCdrop_chart.width = 40
RRCdrop_chart.height = 10
RRCdrop_chart.add_data(RRCDropRateCSBH, titles_from_data = True)  #
RRCdrop_chart.add_data(RRCDropRateCSBHU2100, titles_from_data = True)
RRCdrop_chart.add_data(RRCDropRateCSBHU900, titles_from_data = True)
RRCdrop_chart.set_categories(x_values)
RRCdrop_chart.legend.position = 'b'
hourly_sheet.add_chart(RRCdrop_chart, "A98")

HSDPAThroughput_chart = LineChart()
HSDPAThroughput_chart.width = 40
HSDPAThroughput_chart.height = 10
HSDPAThroughput_chart.add_data(HSDPAThroughputkbps, titles_from_data = True)  #
HSDPAThroughput_chart.add_data(HSDPAThroughputkbpsU2100, titles_from_data = True)
HSDPAThroughput_chart.add_data(HSDPAThroughputkbpsU900, titles_from_data = True)
HSDPAThroughput_chart.set_categories(x_values)
HSDPAThroughput_chart.legend.position = 'b'
hourly_sheet.add_chart(HSDPAThroughput_chart, "A118")

HSUPAThroughput_chart = LineChart()
HSUPAThroughput_chart.width = 40
HSUPAThroughput_chart.height = 10
HSUPAThroughput_chart.add_data(HSUPAThroughputkbps, titles_from_data = True)  #
HSUPAThroughput_chart.add_data(HSUPAThroughputkbpsU2100, titles_from_data = True)
HSUPAThroughput_chart.add_data(HSUPAThroughputkbpsU900, titles_from_data = True)
HSUPAThroughput_chart.set_categories(x_values)
HSUPAThroughput_chart.legend.position = 'b'
hourly_sheet.add_chart(HSUPAThroughput_chart, "A138")

Handover_chart = LineChart()
Handover_chart.width = 40
Handover_chart.height = 10
Handover_chart.add_data(SoftHandoverSuccessrate, titles_from_data = True)  #
Handover_chart.add_data(SoftHandoverSuccessrateU2100, titles_from_data = True)
Handover_chart.add_data(SoftHandoverSuccessrateU900, titles_from_data = True)
Handover_chart.add_data(HardHandoverSuccessrate, titles_from_data = True)  #
Handover_chart.add_data(HardHandoverSuccessrateU2100, titles_from_data = True)
Handover_chart.add_data(HardHandoverSuccessrateU900, titles_from_data = True)
Handover_chart.add_data(CSW2GInterRATHandoverOutSR, titles_from_data = True)  #
Handover_chart.add_data(CSW2GInterRATHandoverOutSRU2100, titles_from_data = True)
Handover_chart.add_data(CSW2GInterRATHandoverOutSRU900, titles_from_data = True)
Handover_chart.set_categories(x_values)
Handover_chart.legend.position = 'b'
hourly_sheet.add_chart(Handover_chart, "A158")

RRCassign_chart = LineChart()
RRCassign_chart.width = 40
RRCassign_chart.height = 10
RRCassign_chart.add_data(RRCAssignmentSucessRateCSBH, titles_from_data = True)  #
RRCassign_chart.add_data(RRCAssignmentSucessRateCSBHU2100, titles_from_data = True)
RRCassign_chart.add_data(RRCAssignmentSucessRateCSBHU900, titles_from_data = True)
RRCassign_chart.set_categories(x_values)
RRCassign_chart.legend.position = 'b'
hourly_sheet.add_chart(RRCassign_chart, "A178")

RABCSassign_chart = LineChart()
RABCSassign_chart.width = 40
RABCSassign_chart.height = 10
RABCSassign_chart.add_data(RABAssignmentSuccessRateCS, titles_from_data = True)  #
RABCSassign_chart.add_data(RABAssignmentSuccessRateCSU2100, titles_from_data = True)
RABCSassign_chart.add_data(RABAssignmentSuccessRateCSU900, titles_from_data = True)
RABCSassign_chart.set_categories(x_values)
RABCSassign_chart.legend.position = 'b'
hourly_sheet.add_chart(RABCSassign_chart, "A198")

RABPSassign_chart = LineChart()
RABPSassign_chart.width = 40
RABPSassign_chart.height = 10
RABPSassign_chart.add_data(RABAssignmentSuccessRatePS, titles_from_data = True)  #
RABPSassign_chart.add_data(RABAssignmentSuccessRatePSU2100, titles_from_data = True)
RABPSassign_chart.add_data(RABAssignmentSuccessRatePSU900, titles_from_data = True)
RABPSassign_chart.set_categories(x_values)
RABPSassign_chart.legend.position = 'b'
hourly_sheet.add_chart(RABPSassign_chart, "A218")

CCSR3G_chart = LineChart()
CCSR3G_chart.width = 40
CCSR3G_chart.height = 10
CCSR3G_chart.add_data(CCSR3G, titles_from_data = True)  #
CCSR3G_chart.add_data(CCSR3GU2100, titles_from_data = True)
CCSR3G_chart.add_data(CCSR3GU900, titles_from_data = True)
CCSR3G_chart.set_categories(x_values)
CCSR3G_chart.legend.position = 'b'
hourly_sheet.add_chart(CCSR3G_chart, "A238")

DCSR3G_chart = LineChart()
DCSR3G_chart.width = 40
DCSR3G_chart.height = 10
DCSR3G_chart.add_data(DCSR3G, titles_from_data = True)  #
DCSR3G_chart.add_data(DCSR3GU2100, titles_from_data = True)
DCSR3G_chart.add_data(DCSR3GU900, titles_from_data = True)
DCSR3G_chart.set_categories(x_values)
DCSR3G_chart.legend.position = 'b'
hourly_sheet.add_chart(DCSR3G_chart, "A258")

#hourly NodeB
x_valuesN = Reference(hourlyN_sheet, range_string=(f"NodeBhourly!$A$2:$B${last_row_hourlyN}"))
MeanThrHSDPAkbps= Reference(hourlyN_sheet, min_col=3, min_row=1, max_row=last_row_hourlyN)
MeanThrHSDPADCkbps= Reference(hourlyN_sheet, min_col=4, min_row=1, max_row=last_row_hourlyN)
MeanThrHSUPAkbps= Reference(hourlyN_sheet, min_col=5, min_row=1, max_row=last_row_hourlyN)
MeanThrHSDPAU2100kbps= Reference(hourlyN_sheet, min_col=6, min_row=1, max_row=last_row_hourlyN)
MeanThrHSDPAU2100DCkbps= Reference(hourlyN_sheet, min_col=7, min_row=1, max_row=last_row_hourlyN)
MeanThrHSUPAU2100kbps= Reference(hourlyN_sheet, min_col=8, min_row=1, max_row=last_row_hourlyN)
MeanThrHSDPAU900kbps= Reference(hourlyN_sheet, min_col=9, min_row=1, max_row=last_row_hourlyN)
MeanThrHSDPAU900DCkbps= Reference(hourlyN_sheet, min_col=10, min_row=1, max_row=last_row_hourlyN)
MeanThrHSUPAU900kbps= Reference(hourlyN_sheet, min_col=11, min_row=1, max_row=last_row_hourlyN)

MeanThrHSDPA_chart = LineChart()
MeanThrHSDPA_chart.width = 40
MeanThrHSDPA_chart.height = 10
MeanThrHSDPA_chart.add_data(MeanThrHSDPAkbps, titles_from_data = True)  #
MeanThrHSDPA_chart.add_data(MeanThrHSDPAU2100kbps, titles_from_data = True)  #
MeanThrHSDPA_chart.add_data(MeanThrHSDPAU900kbps, titles_from_data = True)  #
MeanThrHSDPA_chart.set_categories(x_valuesN)
MeanThrHSDPA_chart.legend.position = 'b'
hourlyN_sheet.add_chart(MeanThrHSDPA_chart, "A18")

MeanThrHSUPAkbps_chart = LineChart()
MeanThrHSUPAkbps_chart.width = 40
MeanThrHSUPAkbps_chart.height = 10
MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAkbps, titles_from_data = True)  #
MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAU2100kbps, titles_from_data = True)  #
MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAU900kbps, titles_from_data = True)  #
MeanThrHSUPAkbps_chart.set_categories(x_valuesN)
MeanThrHSUPAkbps_chart.legend.position = 'b'
hourlyN_sheet.add_chart(MeanThrHSUPAkbps_chart, "A38")

MeanThrHSDPADC_chart = LineChart()
MeanThrHSDPADC_chart.width = 40
MeanThrHSDPADC_chart.height = 10
MeanThrHSDPADC_chart.add_data(MeanThrHSDPADCkbps, titles_from_data = True)  #
MeanThrHSDPADC_chart.add_data(MeanThrHSDPAU2100DCkbps, titles_from_data = True)  #
MeanThrHSDPADC_chart.add_data(MeanThrHSDPAU900DCkbps, titles_from_data = True)  #
MeanThrHSDPADC_chart.set_categories(x_valuesN)
MeanThrHSDPADC_chart.legend.position = 'b'
hourlyN_sheet.add_chart(MeanThrHSDPADC_chart, "A58")

#######!!!!!!!!!!!
# графики ЧНН-ые busy_hourCS_sheet  last_row_BHCS
x_values = Reference(busy_hourCS_sheet, range_string=(f"busy_hourCS!$A$2:$B${last_row_BHCS}"))
x_valuesP = Reference(busy_hourPS_sheet, range_string=(f"busy_hourPS!$A$2:$B${last_row_BHPS}"))

CStraffic3GErl= Reference(busy_hourCS_sheet, min_col=3, min_row=1, max_row=last_row_BHCS)
PStraffic3GULDLGB= Reference(busy_hourPS_sheet, min_col=3, min_row=1, max_row=last_row_BHPS)
CSRABDropRate= Reference(busy_hourCS_sheet, min_col=4, min_row=1, max_row=last_row_BHCS)
PSBlockingRate= Reference(busy_hourPS_sheet, min_col=4, min_row=1, max_row=last_row_BHPS)
PSRABDropRate= Reference(busy_hourPS_sheet, min_col=5, min_row=1, max_row=last_row_BHCS)
PSHSDropRate= Reference(busy_hourPS_sheet, min_col=6, min_row=1, max_row=last_row_BHPS)
HSDPAThroughputkbps= Reference(busy_hourPS_sheet, min_col=7, min_row=1, max_row=last_row_BHPS)
HSUPAThroughputkbps= Reference(busy_hourPS_sheet, min_col=8, min_row=1, max_row=last_row_BHPS)
SoftHandoverSuccessrate= Reference(busy_hourCS_sheet, min_col=5, min_row=1, max_row=last_row_BHCS)
HardHandoverSuccessrate= Reference(busy_hourCS_sheet, min_col=6, min_row=1, max_row=last_row_BHCS)
CSW2GInterRATHandoverOutSR= Reference(busy_hourCS_sheet, min_col=7, min_row=1, max_row=last_row_BHCS)
RRCAssignmentSucessRateCSBH= Reference(busy_hourCS_sheet, min_col=8, min_row=1, max_row=last_row_BHCS)
RRCAssignmentSucessRatePSBH= Reference(busy_hourPS_sheet, min_col=9, min_row=1, max_row=last_row_BHPS)
RRCDropRateCSBH= Reference(busy_hourCS_sheet, min_col=9, min_row=1, max_row=last_row_BHCS)
RRCDropRatePSBH= Reference(busy_hourPS_sheet, min_col=10, min_row=10, max_row=last_row_BHPS)
RABAssignmentSuccessRateCS= Reference(busy_hourCS_sheet, min_col=10, min_row=1, max_row=last_row_BHCS)
RABAssignmentSuccessRatePS= Reference(busy_hourPS_sheet, min_col=11, min_row=1, max_row=last_row_BHPS)
CCSR3G= Reference(busy_hourCS_sheet, min_col=11, min_row=1, max_row=last_row_BHCS)
DCSR3G= Reference(busy_hourPS_sheet, min_col=12, min_row=1, max_row=last_row_BHPS)


CStraffic_chart = LineChart()
CStraffic_chart.width = 40
CStraffic_chart.height = 10
CStraffic_chart.add_data(CStraffic3GErl, titles_from_data = True)  #
#CStraffic_chart.add_data(CStraffic3GErlU2100, titles_from_data = True)
#CStraffic_chart.add_data(CStraffic3GErlU900, titles_from_data = True)
CStraffic_chart.set_categories(x_values)
CStraffic_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(CStraffic_chart, "A18")

PStraffic_chart = LineChart()
PStraffic_chart.width = 40
PStraffic_chart.height = 10
PStraffic_chart.add_data(PStraffic3GULDLGB, titles_from_data = True)  #
#PStraffic_chart.add_data(PStraffic3GULDLGBU2100, titles_from_data = True)
#PStraffic_chart.add_data(PStraffic3GULDLGBU900, titles_from_data = True)
PStraffic_chart.set_categories(x_valuesP)
PStraffic_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(PStraffic_chart, "A38")

CSdrop_chart = LineChart()
CSdrop_chart.width = 40
CSdrop_chart.height = 10
CSdrop_chart.add_data(CSRABDropRate, titles_from_data = True)  #
#CSdrop_chart.add_data(CSRABDropRateU2100, titles_from_data = True)
#CSdrop_chart.add_data(CSRABDropRateU900, titles_from_data = True)
CSdrop_chart.set_categories(x_values)
CSdrop_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(CSdrop_chart, "A58")

PSdrop_chart = LineChart()
PSdrop_chart.width = 40
PSdrop_chart.height = 10
PSdrop_chart.add_data(PSRABDropRate, titles_from_data = True)  #
#PSdrop_chart.add_data(PSRABDropRateU2100, titles_from_data = True)
#PSdrop_chart.add_data(PSRABDropRateU900, titles_from_data = True)
PSdrop_chart.set_categories(x_valuesP)
PSdrop_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(PSdrop_chart, "A78")

RRCdrop_chart = LineChart()
RRCdrop_chart.width = 40
RRCdrop_chart.height = 10
RRCdrop_chart.add_data(RRCDropRateCSBH, titles_from_data = True)  #
#RRCdrop_chart.add_data(RRCDropRateCSBHU2100, titles_from_data = True)
#RRCdrop_chart.add_data(RRCDropRateCSBHU900, titles_from_data = True)
RRCdrop_chart.set_categories(x_values)
RRCdrop_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(RRCdrop_chart, "A98")

HSDPAThroughput_chart = LineChart()
HSDPAThroughput_chart.width = 40
HSDPAThroughput_chart.height = 10
HSDPAThroughput_chart.add_data(HSDPAThroughputkbps, titles_from_data = True)  #
#HSDPAThroughput_chart.add_data(HSDPAThroughputkbpsU2100, titles_from_data = True)
#HSDPAThroughput_chart.add_data(HSDPAThroughputkbpsU900, titles_from_data = True)
HSDPAThroughput_chart.set_categories(x_valuesP)
HSDPAThroughput_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(HSDPAThroughput_chart, "A118")

HSUPAThroughput_chart = LineChart()
HSUPAThroughput_chart.width = 40
HSUPAThroughput_chart.height = 10
HSUPAThroughput_chart.add_data(HSUPAThroughputkbps, titles_from_data = True)  #
#HSUPAThroughput_chart.add_data(HSUPAThroughputkbpsU2100, titles_from_data = True)
#HSUPAThroughput_chart.add_data(HSUPAThroughputkbpsU900, titles_from_data = True)
HSUPAThroughput_chart.set_categories(x_valuesP)
HSUPAThroughput_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(HSUPAThroughput_chart, "A138")

Handover_chart = LineChart()
Handover_chart.width = 40
Handover_chart.height = 10
Handover_chart.add_data(SoftHandoverSuccessrate, titles_from_data = True)  #
#Handover_chart.add_data(SoftHandoverSuccessrateU2100, titles_from_data = True)
#Handover_chart.add_data(SoftHandoverSuccessrateU900, titles_from_data = True)
Handover_chart.add_data(HardHandoverSuccessrate, titles_from_data = True)  #
#Handover_chart.add_data(HardHandoverSuccessrateU2100, titles_from_data = True)
#Handover_chart.add_data(HardHandoverSuccessrateU900, titles_from_data = True)
Handover_chart.add_data(CSW2GInterRATHandoverOutSR, titles_from_data = True)  #
#Handover_chart.add_data(CSW2GInterRATHandoverOutSRU2100, titles_from_data = True)
#Handover_chart.add_data(CSW2GInterRATHandoverOutSRU900, titles_from_data = True)
Handover_chart.set_categories(x_values)
Handover_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(Handover_chart, "A158")

RRCassign_chart = LineChart()
RRCassign_chart.width = 40
RRCassign_chart.height = 10
RRCassign_chart.add_data(RRCAssignmentSucessRateCSBH, titles_from_data = True)  #
#RRCassign_chart.add_data(RRCAssignmentSucessRateCSBHU2100, titles_from_data = True)
#RRCassign_chart.add_data(RRCAssignmentSucessRateCSBHU900, titles_from_data = True)
RRCassign_chart.set_categories(x_values)
RRCassign_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(RRCassign_chart, "A178")

RABCSassign_chart = LineChart()
RABCSassign_chart.width = 40
RABCSassign_chart.height = 10
RABCSassign_chart.add_data(RABAssignmentSuccessRateCS, titles_from_data = True)  #
#RABCSassign_chart.add_data(RABAssignmentSuccessRateCSU2100, titles_from_data = True)
#RABCSassign_chart.add_data(RABAssignmentSuccessRateCSU900, titles_from_data = True)
RABCSassign_chart.set_categories(x_values)
RABCSassign_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(RABCSassign_chart, "A198")

RABPSassign_chart = LineChart()
RABPSassign_chart.width = 40
RABPSassign_chart.height = 10
RABPSassign_chart.add_data(RABAssignmentSuccessRatePS, titles_from_data = True)  #
#RABPSassign_chart.add_data(RABAssignmentSuccessRatePSU2100, titles_from_data = True)
#RABPSassign_chart.add_data(RABAssignmentSuccessRatePSU900, titles_from_data = True)
RABPSassign_chart.set_categories(x_valuesP)
RABPSassign_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(RABPSassign_chart, "A218")

CCSR3G_chart = LineChart()
CCSR3G_chart.width = 40
CCSR3G_chart.height = 10
CCSR3G_chart.add_data(CCSR3G, titles_from_data = True)  #
#CCSR3G_chart.add_data(CCSR3GU2100, titles_from_data = True)
#CCSR3G_chart.add_data(CCSR3GU900, titles_from_data = True)
CCSR3G_chart.set_categories(x_values)
CCSR3G_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(CCSR3G_chart, "A238")

DCSR3G_chart = LineChart()
DCSR3G_chart.width = 40
DCSR3G_chart.height = 10
DCSR3G_chart.add_data(DCSR3G, titles_from_data = True)  #
#DCSR3G_chart.add_data(DCSR3GU2100, titles_from_data = True)
#DCSR3G_chart.add_data(DCSR3GU900, titles_from_data = True)
DCSR3G_chart.set_categories(x_valuesP)
DCSR3G_chart.legend.position = 'b'
busy_hourCS_sheet.add_chart(DCSR3G_chart, "A258")

# BHhourly NodeB   hourlyPSN   hourlyPSN_sheet
x_valuesN = Reference(hourlyPSN_sheet, range_string=(f"NodeB_BH!$A$2:$B${last_row_BHPSN}"))
MeanThrHSDPAkbps= Reference(hourlyPSN_sheet, min_col=3, min_row=1, max_row=last_row_BHPSN)
MeanThrHSDPADCkbps= Reference(hourlyPSN_sheet, min_col=4, min_row=1, max_row=last_row_BHPSN)
MeanThrHSUPAkbps= Reference(hourlyPSN_sheet, min_col=5, min_row=1, max_row=last_row_BHPSN)
#MeanThrHSDPAU2100kbps= Reference(hourlyPSN_sheet, min_col=6, min_row=1, max_row=last_row_BHPSN)
#MeanThrHSDPAU2100DCkbps= Reference(hourlyPSN_sheet, min_col=7, min_row=1, max_row=last_row_BHPSN)
#MeanThrHSUPAU2100kbps= Reference(hourlyPSN_sheet, min_col=8, min_row=1, max_row=last_row_BHPSN)
#MeanThrHSDPAU900kbps= Reference(hourlyPSN_sheet, min_col=9, min_row=1, max_row=last_row_BHPSN)
#MeanThrHSDPAU900DCkbps= Reference(hourlyPSN_sheet, min_col=10, min_row=1, max_row=last_row_BHPSN)
#MeanThrHSUPAU900kbps= Reference(hourlyPSN_sheet, min_col=11, min_row=1, max_row=last_row_BHPSN)

MeanThrHSDPA_chart = LineChart()
MeanThrHSDPA_chart.width = 40
MeanThrHSDPA_chart.height = 10
MeanThrHSDPA_chart.add_data(MeanThrHSDPAkbps, titles_from_data = True)  #
#MeanThrHSDPA_chart.add_data(MeanThrHSDPAU2100kbps, titles_from_data = True)  #
#MeanThrHSDPA_chart.add_data(MeanThrHSDPAU900kbps, titles_from_data = True)  #
MeanThrHSDPA_chart.set_categories(x_valuesN)
MeanThrHSDPA_chart.legend.position = 'b'
hourlyPSN_sheet.add_chart(MeanThrHSDPA_chart, "A18")

MeanThrHSUPAkbps_chart = LineChart()
MeanThrHSUPAkbps_chart.width = 40
MeanThrHSUPAkbps_chart.height = 10
MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAkbps, titles_from_data = True)  #
#MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAU2100kbps, titles_from_data = True)  #
#MeanThrHSUPAkbps_chart.add_data(MeanThrHSUPAU900kbps, titles_from_data = True)  #
MeanThrHSUPAkbps_chart.set_categories(x_valuesN)
MeanThrHSUPAkbps_chart.legend.position = 'b'
hourlyPSN_sheet.add_chart(MeanThrHSUPAkbps_chart, "A38")

MeanThrHSDPADC_chart = LineChart()
MeanThrHSDPADC_chart.width = 40
MeanThrHSDPADC_chart.height = 10
MeanThrHSDPADC_chart.add_data(MeanThrHSDPADCkbps, titles_from_data = True)  #
#MeanThrHSDPADC_chart.add_data(MeanThrHSDPAU2100DCkbps, titles_from_data = True)  #
#MeanThrHSDPADC_chart.add_data(MeanThrHSDPAU900DCkbps, titles_from_data = True)  #
MeanThrHSDPADC_chart.set_categories(x_valuesN)
MeanThrHSDPADC_chart.legend.position = 'b'
hourlyPSN_sheet.add_chart(MeanThrHSDPADC_chart, "A58")


my_file.save(f"{directory}{csv_name1}{output_comment}.xlsx")

print('готово')
winsound.Beep(2500, 1000)
winsound.Beep(3000, 1000)