import csv
import sqlite3
import pandas as pd
import winsound
import openpyxl
from openpyxl.chart import (LineChart, Reference)
import openpyxl.styles
from openpyxl.styles import PatternFill
#from openpyxl.utils import cell_range
''' '''

# подключаемся к базе данных
conn = sqlite3.connect('C:/SQLite/firstDB/stsDB.db')
# создаем курсор для выполнения запросов
cursor = conn.cursor()
start_date = '2023-03-25'
end_date = '2023-04-03' # надо брать на день позже

# используйте операторы сравнения для выборки строк в заданном диапазоне
# columns = ['K3014:Traffic Volume on TCH (Erl)', 'Start Time', 'GCELL']
query1 = f'''SELECT "Start Time", "GCELL", "K3014:Traffic Volume on TCH (Erl)" FROM GSMsts WHERE `Start Time` >= '{start_date}' AND `Start Time` <= '{end_date}' '''

cursor.execute(query1)
data = cursor.fetchall()
data = [[None if col == 'NIL' else col for col in row] for row in data]
sts_df_GSM = pd.DataFrame(data, columns=[i[0] for i in cursor.description])
#print(sts_df_GSM)

sts_df_GSM = pd.pivot_table(sts_df_GSM, index = 'GCELL', columns = 'Start Time', values = 'K3014:Traffic Volume on TCH (Erl)')

## UMTS
query2 = f'''SELECT "Start Time", "BSC6910UCell", "CS Voice Traffic Volume (Erl)" FROM UMTS_1v2 WHERE `Start Time` >= '{start_date}' AND `Start Time` <= '{end_date}' '''

cursor.execute(query2)
data = cursor.fetchall()
data = [[None if col == 'NIL' else col for col in row] for row in data]
sts_df_UMTS = pd.DataFrame(data, columns=[i[0] for i in cursor.description])
#print(sts_df_GSM)

sts_df_UMTS = pd.pivot_table(sts_df_UMTS, index = 'BSC6910UCell', columns = 'Start Time', values = 'CS Voice Traffic Volume (Erl)')

##LTE
query3 = f'''SELECT "Start Time", "Cell", "L.ChMeas.PRB.DL.Used.Avg (None)" FROM LTEsts WHERE `Start Time` >= '{start_date}' AND `Start Time` <= '{end_date}' '''

cursor.execute(query3)
data = cursor.fetchall()
data = [[None if col == 'NIL' else col for col in row] for row in data]
sts_df_LTE = pd.DataFrame(data, columns=[i[0] for i in cursor.description])
#print(sts_df_GSM)
conn.close()
sts_df_LTE = pd.pivot_table(sts_df_LTE, index = 'Cell', columns = 'Start Time', values = 'L.ChMeas.PRB.DL.Used.Avg (None)')

conn.close()


with pd.ExcelWriter('C:/test2/traffic_alarm.xlsx', engine='openpyxl') as writer:
    sts_df_GSM.to_excel(writer, sheet_name='GSM')
    sts_df_UMTS.to_excel(writer, sheet_name='UMTS')
    sts_df_LTE.to_excel(writer, sheet_name='LTE')


my_file = openpyxl.load_workbook(f"C:/test2/traffic_alarm.xlsx")
GSM_sheet = my_file["GSM"]
UMTS_sheet = my_file["UMTS"]
LTE_sheet = my_file["LTE"]

# определение количества строк и столбцов в таблицах
last_row_GSM = GSM_sheet.max_row
last_column_GSM = GSM_sheet.max_column

last_row_UMTS = UMTS_sheet.max_row
last_column_UMTS = UMTS_sheet.max_column

last_row_LTE = LTE_sheet.max_row
last_column_LTE = LTE_sheet.max_column


# выставление переноса строк для даты/времени
for cell in GSM_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in UMTS_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
for cell in LTE_sheet[1]:
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)


fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# выставление правильного формата для столбцов с датами
# for r in range(2,(last_row_GSM+1)):
#     GSM_sheet[f'B{r}'].number_format ='DD.MM.YYYY'

for row in range(1, GSM_sheet.max_row):
    for col in range(2, last_column_GSM):
        cell = GSM_sheet.cell(row=row, column=col)
        if cell.value == 0:
            cell.fill = fill

for row in range(1, UMTS_sheet.max_row):
    for col in range(2, last_column_UMTS):
        cell = UMTS_sheet.cell(row=row, column=col)
        if cell.value == 0:
            cell.fill = fill
for row in range(1, LTE_sheet.max_row):
    for col in range(2, last_column_LTE):
        cell = LTE_sheet.cell(row=row, column=col)
        if cell.value == 0:
            cell.fill = fill

my_file.save(f"C:/test2/traffic_alarm.xlsx")

